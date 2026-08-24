# -*- coding: utf-8 -*-
"""LG 라인스케줄 '잔업' 시트 파서 (백엔드/적재 공용). parse_line_schedule(path_or_bytes, anchor_date)->dict.
   2026-08 신형식: 주 생산라인(No.=CAC##)만 추출(추가라인 1층/2층 조립·공청기 제외).
   셀값은 보여지는대로 저장(B/A/E·잔업시간숫자·SKD/rac이동/CC지원 등). 구 '특수일 행(row15)'은 폐지(이제 데이터라인이라 오독)."""
import re, datetime, io
import openpyxl

LINES_DEFAULT = ['C1', 'CA', 'CE', 'CG', 'CH', 'CJ', 'CK', 'CM']

def parse_line_schedule(src, anchor_date):
    """src=파일경로 또는 bytes. anchor_date=datetime.date(파일명 기준일). '잔업' 시트 파싱."""
    if isinstance(src, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(src), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if '잔업' not in wb.sheetnames:
        raise ValueError("'잔업' 시트가 없습니다.")
    ws = wb['잔업']
    rows = list(ws.iter_rows(min_row=1, max_row=40, max_col=131, values_only=True))
    def cell(r, c): return rows[r-1][c-1] if r-1 < len(rows) and c-1 < len(rows[r-1]) else None
    # 날짜열 (행5 = 일자: '29일' 등)
    day_of = {}
    for c in range(6, 132):
        v = cell(5, c)
        if v is not None:
            mm = re.search(r'(\d+)', str(v))
            if mm: day_of[c] = int(mm.group(1))
    cols = sorted(day_of)
    if not cols: raise ValueError("날짜열(행5)을 찾을 수 없습니다.")
    def build(ac): return {c: anchor_date + datetime.timedelta(days=(c - ac)) for c in cols}
    best = None
    for ac in [c for c in cols if day_of[c] == anchor_date.day]:
        dd = build(ac)
        if all(dd[c].day == day_of[c] for c in cols): best = ac; break
    if best is None: best = cols[0]  # 폴백
    dates = build(best)
    # 데이터행: 주 생산라인만(No.=CAC## 패턴). 추가라인(1층/2층 조립·공청기 등)은 제외 → 위치 이동에도 견고.
    recs = []; meta = []; ord_ = 0
    for r in range(7, 41):
        line = cell(r, 3); no = str(cell(r, 4) or '').strip()
        if not line: continue
        if not re.match(r'^CAC\d+', no, re.I): continue  # 주 생산라인만
        line = str(line).strip(); ord_ += 1
        meta.append({"line_no": line, "sort_ord": ord_, "gubun": str(cell(r, 2) or '').strip(),
                     "model_no": no, "jindo": str(cell(r, 5) or '').strip()})
        for c in cols:
            v = cell(r, c)
            if v is not None and str(v).strip():
                recs.append({"line_no": line, "ymd": dates[c], "code": str(v).strip().replace('\n', '/')[:20]})
    # 특수일 행 폐지(신형식엔 전용 행 없음)
    events = []
    return {"recs": recs, "meta": meta, "events": events,
            "date_from": dates[cols[0]], "date_to": dates[cols[-1]], "anchor": anchor_date}


if __name__ == '__main__':
    import sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.path.insert(0, r'd:\피앤씨인더스트리\100_AI_AGENT\Projects\New_ERP')
    import db_client, pyodbc
    fp = r'D:\피앤씨인더스트리\100_AI_AGENT\Projects\NEW_ERP_1\RAC_SAC_Changed Plan_20260723.xlsx'
    m = re.search(r'(\d{8})', fp)
    anchor = datetime.date(int(m.group(1)[:4]), int(m.group(1)[4:6]), int(m.group(1)[6:8]))
    res = parse_line_schedule(fp, anchor)
    print(f"파싱: recs {len(res['recs'])} · meta {len(res['meta'])} · events {len(res['events'])} · {res['date_from']}~{res['date_to']}")
    cn = pyodbc.connect(f'DRIVER={{SQL Server}};SERVER={db_client.DB_SERVER},{db_client.DB_PORT};DATABASE=PARTNER_ERP_TEST3;UID={db_client.DB_USER};PWD={db_client.DB_PASSWORD}', autocommit=False)
    c = cn.cursor()
    c.execute("IF OBJECT_ID('nx.line_calendar') IS NULL CREATE TABLE nx.line_calendar (line_no NVARCHAR(20), cal_ymd DATE, work_code NVARCHAR(20), note NVARCHAR(50), upd_dt DATETIME DEFAULT GETDATE(), CONSTRAINT pk_lcal PRIMARY KEY(line_no,cal_ymd))")
    c.execute("IF OBJECT_ID('nx.line_cal_event') IS NULL CREATE TABLE nx.line_cal_event (cal_ymd DATE, event NVARCHAR(50), CONSTRAINT pk_lcev PRIMARY KEY(cal_ymd,event))")
    c.execute("IF OBJECT_ID('nx.line_cal_meta') IS NULL CREATE TABLE nx.line_cal_meta (line_no NVARCHAR(20) PRIMARY KEY, sort_ord INT, gubun NVARCHAR(20), model_no NVARCHAR(20), jindo NVARCHAR(30), upd_dt DATETIME DEFAULT GETDATE())")
    cn.commit()
    df, dt = res['date_from'].isoformat(), res['date_to'].isoformat()
    c.execute("DELETE FROM nx.line_calendar WHERE cal_ymd BETWEEN ? AND ?", df, dt)
    c.execute("DELETE FROM nx.line_cal_event WHERE cal_ymd BETWEEN ? AND ?", df, dt)
    for x in res['recs']:
        c.execute("INSERT INTO nx.line_calendar(line_no,cal_ymd,work_code) VALUES(?,?,?)", x['line_no'], x['ymd'].isoformat(), x['code'])
    for x in res['events']:
        c.execute("INSERT INTO nx.line_cal_event(cal_ymd,event) VALUES(?,?)", x['ymd'].isoformat(), x['event'])
    for x in res['meta']:
        c.execute("DELETE FROM nx.line_cal_meta WHERE line_no=?", x['line_no'])
        c.execute("INSERT INTO nx.line_cal_meta(line_no,sort_ord,gubun,model_no,jindo) VALUES(?,?,?,?,?)", x['line_no'], x['sort_ord'], x['gubun'], x['model_no'], x['jindo'])
    cn.commit()
    c.execute("SELECT COUNT(*) FROM nx.line_calendar"); print("nx.line_calendar:", c.fetchone()[0])
    c.execute("SELECT COUNT(*) FROM nx.line_cal_event"); print("nx.line_cal_event:", c.fetchone()[0])
