# -*- coding: utf-8 -*-
"""원소재(동) 정산 엔진 v2 — 담당 수불 자료 정본 앵커.
   설계(사용자 확정 방향):
     - 부품 단위 동중량(geom) = 담당 파이프 수불 자료의 Φ/T/L에서 산출 (검증완료 ×0.02809, 파일중량 99.9%일치)
       → BOM 재귀전개 폐기(과다전개 오류), 품목마스터 치수는 fallback
     - 입고수량 = ERP 확정입고(진짜매입: COST_TAG='1' 매입단가 거래처=입고거래처)  ← 자동화
     - 동 소비 = 부품 geom × ERP 입고수량, 협력사별 집계
   수불 자료 위치: <프로젝트>\{4,5,6}월 파이프 수불 자료_대표님 전달 건.xlsx (매월 갱신)
"""
import os, re, sys
sys.path.insert(0, r"d:\피앤씨인더스트리\100_AI_AGENT\Projects\NEW_ERP_1\_harness")
from nx_cost_engine import _nx
import openpyxl

BASE = r"d:\피앤씨인더스트리\100_AI_AGENT\Projects\NEW_ERP_1"
CU_K = 0.02809 * 0.001
COOP_CUST = {'2148':'대원산업','2096':'미래정밀','2306':'명진산업','2048':'중앙정밀','2068':'이젠터',
             '2266':'케이비','2142':'세광산업','233':'썬텍코리아','2028':'썬텍','2067':'MTS','2250':'수테크'}
_CODE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-\.]{4,}$")
_SUF = re.compile(r"-\d+(-\d+)?$")

def _norm(c):
    c = re.sub(r"\s+", "", str(c).strip().upper())
    c = re.sub(r"[\(\+].*$", "", c)            # (링/(마·+용접링 꼬리
    c = re.sub(r"번$|품\d+$", "", c)            # 5998A20002A-2번 / MAA..품7
    return _SUF.sub("", c).strip()

def geom(d, t, l):
    try: d, t, l = float(d), float(t), float(l)
    except (TypeError, ValueError): return 0.0
    return (d - t) * t * l * CU_K if (d and t and l and t > 0 and d > t and l > 0) else 0.0


class CopperSettlement:
    def __init__(self, pipe_files=None):
        self.cn = _nx(); self.cur = self.cn.cursor(); self.L = "PARTNER_ERP.dbo."
        self._pur = None; self._dim = {}
        self.ref = {}       # 부품 geom 정본: norm(code) -> geom(kg)
        self.ref_raw = {}   # 원본코드 -> geom (표시용)
        self.assy_ref = {}  # Assy 단위동중량 정본: norm(assy) -> Σ(부품 geom×소요)
        files = pipe_files or [f"{m}월 파이프 수불 자료_대표님 전달 건.xlsx" for m in ("4", "5", "6")]
        for fn in files:
            p = os.path.join(BASE, fn)
            if os.path.exists(p): self._load_pipe(p)

    def _load_pipe(self, path):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            if sn == "MASTER": continue
            rows = list(wb[sn].iter_rows(min_row=1, max_row=5000, values_only=True))
            hdr = None; cm = {}
            for i in range(min(8, len(rows))):
                for j, v in enumerate(rows[i] or []):
                    s = str(v or '')
                    if '소요량' in s: cm['soyo'] = j; hdr = i
                    if '외경' in s or s.strip() == 'Φ': cm['phi'] = j
                    if s.strip() == 'T': cm['t'] = j
                    if '길이' in s: cm['l'] = j
                    if '하위' in s: cm['part'] = j
                if hdr is not None and 'phi' in cm: break
            if hdr is None or 'phi' not in cm: continue
            pc = cm.get('part', cm['phi'] - 1); ac = cm.get('assy', max(0, cm['phi'] - 2)); sc = cm.get('soyo')
            assy_acc = {}   # 이 시트 Assy별 Σ(부품 geom×소요)
            for r in rows[hdr + 1:]:
                if not r or pc >= len(r): continue
                p = str(r[pc] or '').strip()
                if not p or not _CODE.match(p): continue
                g = geom(r[cm['phi']], r[cm.get('t', cm['phi']+1)], r[cm.get('l', cm['phi']+2)])
                if g > 0:
                    self.ref.setdefault(_norm(p), g)     # 부품 단위 geom (부품간 일관 검증완료)
                    self.ref_raw.setdefault(p.upper(), g)
                    a = str(r[ac] or '').strip() if ac < len(r) else ''
                    if a and _CODE.match(a):
                        try: so = float(r[sc]) if (sc is not None and sc < len(r) and r[sc] is not None) else 1.0
                        except (TypeError, ValueError): so = 1.0
                        assy_acc[_norm(a)] = assy_acc.get(_norm(a), 0.0) + g * (so or 1.0)
            for a, tot in assy_acc.items():
                self.assy_ref.setdefault(a, tot)          # 최초 파일/시트 값(전체 부품구성)
        wb.close()

    def _purmap(self):
        if self._pur is None:
            self.cur.execute(f"SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))), LTRIM(RTRIM(ISNULL(CUST_CODE,''))) "
                             f"FROM {self.L}PR_M_ITEM_COST WHERE COST_TAG='1'")
            m = {}
            for ic, cc in self.cur.fetchall(): m.setdefault(str(ic).strip(), set()).add(str(cc).strip())
            self._pur = m
        return self._pur

    def _master_geom(self, code):
        if code not in self._dim:
            self.cur.execute(f"SELECT ITEM_DIAM, ITEM_THICK, ITEM_LENGTH FROM {self.L}PR_M_ITEM "
                             f"WHERE LTRIM(RTRIM(ITEM_CODE))=?", code)
            r = self.cur.fetchone(); self._dim[code] = geom(*r) if r else 0.0
        return self._dim[code]

    def unit_copper(self, code):
        """1개당 동중량(kg). 1순위=수불 부품, 2순위=수불 Assy(Σ부품), 3순위=품목마스터. (source, geom)."""
        n = _norm(code)
        if code.upper() in self.ref_raw: return ("수불부품", self.ref_raw[code.upper()])
        if n in self.ref: return ("수불부품~", self.ref[n])
        if n in self.assy_ref: return ("수불Assy", self.assy_ref[n])
        g = self._master_geom(code)
        if g > 0: return ("마스터", g)
        return ("치수없음", 0.0)

    def receipts(self, yymm, custs=None):
        custs = list(custs or COOP_CUST.keys()); ph = ",".join("?" * len(custs))
        f, t = yymm + "01", yymm + "31"
        self.cur.execute(f"""SELECT UPPER(LTRIM(RTRIM(A.MAT_CODE))) ic, LTRIM(RTRIM(A.CUST_CODE)) cc, SUM(A.MAINT_QTY) q
            FROM {self.L}PU_T_STOCK_MAINT A
            WHERE A.MAINT_TAG IN ('9','S','C','G','H') AND A.MAINT_QTY>0 AND A.MAINT_YMD BETWEEN ? AND ?
              AND LTRIM(RTRIM(A.CUST_CODE)) IN ({ph})
              AND ((ISNULL(A.INSP_FLAG,'N') IN ('','N')) OR (ISNULL(A.INSP_FLAG,'N') IN ('S','F')))
            GROUP BY UPPER(LTRIM(RTRIM(A.MAT_CODE))), LTRIM(RTRIM(A.CUST_CODE))""", f, t, *custs)
        return [(str(r[0]).strip(), str(r[1]).strip(), float(r[2] or 0)) for r in self.cur.fetchall()]

    def by_vendor(self, yymm, custs=None, real_only=True):
        pur = self._purmap(); agg = {}; detail = []
        for ic, cc, q in self.receipts(yymm, custs):
            if real_only and cc not in pur.get(ic, set()):
                agg.setdefault(cc, {"copper":0.0,"items":0,"nonpur":0,"nodim":0})["nonpur"] += 1; continue
            src, uc = self.unit_copper(ic); cop = uc * q
            a = agg.setdefault(cc, {"copper":0.0,"items":0,"nonpur":0,"nodim":0})
            a["copper"] += cop; a["items"] += 1
            if uc == 0: a["nodim"] += 1
            detail.append((cc, ic, src, round(uc, 5), q, round(cop, 3)))
        return agg, detail

    def close(self):
        try: self.cn.close()
        except Exception: pass


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    eng = CopperSettlement()
    print(f"수불 정본: 부품 geom {len(eng.ref_raw)}개 · Assy {len(eng.assy_ref)}개 로드")
    agg, detail = eng.by_vendor("2606")
    print("\n=== 26년 6월 협력사별 동 소비(kg) — 수불정본 앵커 ===")
    for cc, a in sorted(agg.items(), key=lambda x: -x[1]["copper"]):
        print(f"  {COOP_CUST.get(cc,cc):<10}({cc:<5}) 동소비 {a['copper']:>11,.1f}kg · 매입 {a['items']:>4} · 매입아님 {a['nonpur']:>4} · 치수없음 {a['nodim']:>3}")
    # source 분포
    from collections import Counter
    sc = Counter(d[2] for d in detail)
    print("\n단위동 출처:", dict(sc))
    eng.close()
