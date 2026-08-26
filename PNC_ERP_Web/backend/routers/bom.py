# -*- coding: utf-8 -*-
"""bom 도메인 라우터 — app.py에서 분리. 공유헬퍼는 common.py."""
import os, math, json, base64, time, hashlib, mimetypes
from datetime import datetime, timedelta
from urllib.parse import quote as _urlquote
from fastapi import APIRouter, Query, Body, HTTPException, Response, UploadFile, File, Form
from common import (_conn, _num, _run_sp, _shape, _nx, _nx_tx, _b, _d6, _ym, _ITEM_WORK, _get_cost_engine, _reset_cost_engine, _COST_LOCK, SP_SIL, SP_NAE, NxCostEngine, _HERE, _geom_weight)

from common import _ITEM_MAKE
router = APIRouter()

@router.get("/api/bom/search")
def bom_search(q: str = Query('', description="품번/품명 부분검색"),
               include_past: int = Query(0, description="0=현행(status='사용')만[기본], 1=과거(휴면)포함"),
               all_active: int = Query(0, description="1=BOM연결 무관 status='사용' 전체(재고 입출고 화면용: orphan 원소재도 노출). include_past 우선."),
               use: str = Query("", description="사용여부(nx.item.use_flag): ''=전체, '1'=사용중, '0'=사용중지")):
    q = q.strip()
    usef = "AND ISNULL(i.use_flag,1)=1" if use.strip() == "1" else ("AND ISNULL(i.use_flag,1)=0" if use.strip() == "0" else "")
    cn = _nx(); cur = cn.cursor()
    try:
        like = f'%{q}%'
        # ★현행/과거 토글 + ★미사용 orphan 숨김(2026-08-13): 기본 = status='사용' AND (BOM보유 OR 다른BOM의 자식으로 사용).
        #   → 빈 SUB shell(BOM없고 아무 BOM에도 자식으로 안 쓰이는 _S01·_S07·-은납-S7 등)은 기본검색에서 숨김. include_past=1이면 전체.
        # ★현행 판정: (현행자식 cs_except=0으로 사용) OR (BOM보유 AND 아무데도 자식아님=최상위제품). → 비현행 변형(예 태국 -F&T, 현행자식 0)·빈 shell 숨김. include_past=1이면 전체.
        # ★성능: 파생 DISTINCT 조인(전체 bom_line 스캔 2회) → 상관 EXISTS로 교체(578ms→56ms). 매칭 item만 검사.
        # ★all_active(2026-08-15): 재고 입출고 화면은 BOM연결 무관하게 입고 가능한 전 활성자재를 검색해야 함(orphan 원소재 7072AR9374M/X/L 등 노출). orphan 필터만 해제·status는 유지.
        if int(include_past or 0):
            past = ""
        elif int(all_active or 0):
            past = "AND ISNULL(i.status,N'사용')=N'사용'"
        else:
            past = """AND ISNULL(i.status,N'사용')=N'사용'
              AND ( EXISTS(SELECT 1 FROM nx.bom_line uc WHERE uc.child_item=i.item_code AND ISNULL(uc.cs_calc_except,0)=0)
                 OR ( EXISTS(SELECT 1 FROM nx.bom_header h2 WHERE h2.item_code=i.item_code)
                      AND NOT EXISTS(SELECT 1 FROM nx.bom_line u WHERE u.child_item=i.item_code) ) )"""
        cur.execute(f"""
            SELECT TOP 60 i.item_code, i.item_name, i.item_type, ISNULL(i.status,'') st,
              CASE WHEN EXISTS(SELECT 1 FROM nx.bom_header h WHERE h.item_code=i.item_code) THEN 1 ELSE 0 END AS has_bom,
              ISNULL(i.use_flag,1) uf
            FROM nx.item i
            WHERE (i.item_code LIKE ? OR i.item_name LIKE ?) {past} {usef}
            ORDER BY has_bom DESC, i.item_code""", like, like)
        rows = [{"item": r[0], "name": r[1], "type": r[2], "status": str(r[3]).strip(), "has_bom": bool(r[4]), "use_flag": int(r[5])} for r in cur.fetchall()]
        return {"rows": rows, "include_past": int(include_past or 0)}
    finally:
        cn.close()

@router.get("/api/bom/iteminfo")
def bom_iteminfo(item: str = Query(...)):
    """신규 BOM 등록 그리드: 기존 품번 입력 시 마스터(품명·규격·치수·재질·중량·분류·매입처) 자동조회. 3방식(LG불러오기·복사·새로) 공통."""
    code = item.strip().upper()
    if not code:
        return {"found": False}
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""SELECT item_name, item_spec, diam, thick, length, metal_gubun, net_weight, unit,
              lgroup, sgroup, make_type, cost_gubun, in_cust FROM nx.item WHERE item_code=?""", code)
        r = cur.fetchone()
        if not r:
            return {"found": False}
        keys = ["item_name", "item_spec", "diam", "thick", "length", "metal_gubun", "net_weight",
                "unit", "lgroup", "sgroup", "make_type", "cost_gubun", "in_cust"]
        num = {"diam", "thick", "length", "net_weight"}
        d = {}
        for k, v in zip(keys, r):
            if k in num:
                try: d[k] = (float(v) if v is not None else None)
                except (TypeError, ValueError): d[k] = None
            else:
                d[k] = ("" if v is None else str(v).strip())
        cust = ""
        if d.get("in_cust"):
            cur.execute("SELECT CUST_DESC FROM PARTNER_ERP_TEST3.nx.CM_M_CUST WHERE CUST_CODE=?", d["in_cust"])
            rc = cur.fetchone()
            if rc and rc[0]:
                cust = str(rc[0]).strip()
        d["cust_name"] = cust
        d["found"] = True
        return d
    finally:
        cn.close()

@router.get("/api/bom/get")
def bom_get(item: str = Query(..., description="품번")):
    item = item.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT item_name, item_type, ISNULL(cut_gubun,'') FROM nx.item WHERE item_code=?", item)
        pi = cur.fetchone()
        if not pi:
            raise HTTPException(404, f"품목 {item} 없음")
        cur.execute("""SELECT GAGONG_PROC_CODE, GAGONG_PROC_DESC FROM PARTNER_ERP_TEST3.nx.PR_M_PROC_GAGONG
                       WHERE ISNULL(GAGONG_PROC_CODE,'')<>'' ORDER BY SORT_KEY, GAGONG_PROC_CODE""")
        procs = [{"code": str(r[0]).strip(), "name": (str(r[1]).strip() if r[1] else str(r[0]).strip())} for r in cur.fetchall()]
        cur.execute("SELECT bom_id, version, status FROM nx.bom_header WHERE item_code=?", item)
        h = cur.fetchone()
        if not h:
            return {"item": item, "name": pi[0], "cut_gubun": pi[2], "header": None, "lines": [], "procs": procs}
        bom_id = h[0]
        cur.execute("""
            SELECT l.seq, l.child_item, ci.item_name, l.qty, l.node_type,
                   l.cs_calc_except, l.lme_except, l.sagub_default, l.is_optional,
                   l.from_ymd, l.to_ymd, l.except_flag, l.set_except, l.kitting, l.vir_item,
                   l.proc_gubun, l.gagong_proc, l.s_work, l.wh_gagong, l.in_gagong, l.cust_code, l.remarks,
                   ci.item_spec, ci.metal_gubun, ci.diam, ci.thick, ci.length, ci.item_type AS child_type,
                   ci.net_weight, ci.unit, ci.sgroup, ci.lgroup, ci.make_type, ci.cost_gubun, ci.status,
                   ISNULL(ci.in_cust,'') AS in_cust, ISNULL(pv.partner_name, pc.CUST_DESC) AS cust_name
            FROM nx.bom_line l
            LEFT JOIN nx.item ci ON ci.item_code = l.child_item
            LEFT JOIN nx.partner pv ON pv.partner_code = ci.in_cust
            LEFT JOIN PARTNER_ERP_TEST3.nx.CM_M_CUST pc ON pc.CUST_CODE = ci.in_cust
            WHERE l.bom_id = ? ORDER BY l.seq""", bom_id)
        cols = [d[0] for d in cur.description]
        lines = []
        for r in cur.fetchall():
            d = {}
            for k, v in zip(cols, r):
                d[k] = bool(v) if isinstance(v, bool) else v
            lines.append(d)
        return {"item": item, "name": pi[0], "type": pi[1], "cut_gubun": pi[2],
                "header": {"bom_id": bom_id, "version": h[1], "status": h[2]}, "lines": lines, "procs": procs}
    finally:
        cn.close()

# 품목 마스터 편집용 드롭다운 코드(CM_M_MASTER_DETAIL) + 마스터 저장(nx.item)
_CODE_GROUPS = {"lgroup": "PR005", "sgroup": "PR006", "metal": "PR019",
                "material": "PR020", "unit": "CM002", "obtain": "PR007", "itemclass": "PR008"}

@router.get("/api/codes")
def codes():
    cn = _conn(); cur = cn.cursor()
    try:
        out = {}
        for key, grp in _CODE_GROUPS.items():
            cur.execute("""SELECT LTRIM(RTRIM(DETAIL_CODE)), LTRIM(RTRIM(DETAIL_DESC))
                FROM PARTNER_ERP_TEST3.nx.CM_M_MASTER_DETAIL WHERE KIND_CODE=? AND ISNULL(USE_FLAG,'1')='1'
                ORDER BY SORT_SEQ, DETAIL_CODE""", grp)
            out[key] = [{"code": r[0], "name": r[1]} for r in cur.fetchall()]
        out["make_type"] = [{"code": k, "name": v} for k, v in _ITEM_MAKE.items() if k]   # 통일: _ITEM_MAKE(품목조회 일치)
        out["cost_gubun"] = [{"code": "2", "name": "구매단가"}, {"code": "3", "name": "소재단가"},
                             {"code": "1", "name": "내부단가"}, {"code": "5", "name": "기타"}]
        out["status"] = [{"code": "사용", "name": "사용"}, {"code": "중지", "name": "중지"}]
        # 거래처(협력사) 검색용 상위 — 프론트는 item/vendorsearch로 별도 검색
        return out
    finally:
        cn.close()

@router.get("/api/item/vendorsearch")
def item_vendorsearch(q: str = Query("")):
    """매입처(거래처) 검색 — nx.partner + 라이브 CM_M_CUST."""
    cn = _nx(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT TOP 30 c.CUST_CODE, c.CUST_DESC FROM PARTNER_ERP_TEST3.nx.CM_M_CUST c
            WHERE c.CUST_CODE LIKE ? OR c.CUST_DESC LIKE ? ORDER BY c.CUST_CODE""", like, like)
        return {"rows": [{"code": r[0], "name": r[1]} for r in cur.fetchall()]}
    finally:
        cn.close()

@router.post("/api/item/save")
def item_save(payload: dict = Body(...)):
    """품목 마스터 일괄 저장(nx.item) — 원소재 스펙~거래처~분류~생산구분. 상향식 등록."""
    rows = payload.get("rows", []) or []
    nx = _nx(); cur = nx.cursor()
    try:
        saved = 0; errs = []
        for r in rows:
            code = str(r.get("item_code", "")).strip()
            if not code:
                continue
            def s(k): v = r.get(k); return None if v in (None, "") else str(v).strip()
            def n(k):
                v = r.get(k)
                try: return float(v) if v not in (None, "") else None
                except: return None
            _mg, _d, _t, _l, _nw = s("metal_gubun"), n("diam"), n("thick"), n("length"), n("net_weight")
            if str(s("cost_gubun") or "").strip() == "3":   # ★원소재=기하중량 재계산(레거시 f_get_weight3·저장값무시)
                _gw = _geom_weight(_mg, _d, _t, _l)
                if _gw is not None: _nw = _gw
            vals = (s("item_name"), s("item_spec"), _mg, _d, _t, _l,
                    _nw, s("unit"), s("in_cust"), s("sgroup"), s("lgroup"), s("make_type"),
                    s("cost_gubun"), s("status"))
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", code)
            if cur.fetchone():
                cur.execute("""UPDATE nx.item SET item_name=ISNULL(?,item_name), item_spec=?, metal_gubun=?,
                    diam=?, thick=?, length=?, net_weight=?, unit=ISNULL(?,unit), in_cust=?, sgroup=?, lgroup=?,
                    make_type=?, cost_gubun=?, status=ISNULL(?,status) WHERE item_code=?""", *vals, code)
            else:
                cur.execute("""INSERT INTO nx.item(item_code,item_name,item_spec,metal_gubun,diam,thick,length,
                    net_weight,unit,in_cust,sgroup,lgroup,make_type,cost_gubun,status,item_type)
                    VALUES(?,?,?,?,?,?,?,?,ISNULL(?,'EA'),?,?,?,?,?,ISNULL(?,'사용'),'부품')""", code, *vals)
            saved += 1
        _reset_cost_engine()   # 스펙(치수·재질·중량·조달) 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "count": saved, "errors": errs}
    finally:
        nx.close()

def _bom_tree_route(item, route_id):
    """★조달후보(nx.sourcing_route_line 계층)를 마스터 bom/tree와 **동일한 트리 JSON 스키마**로 반환.
       레벨0=제품 / 레벨1=최상위 라인(parent_line NULL) / 레벨2+=SUB 하위(parent_line 기반). node_kind='SUB'=haskids.
       ★route_id>0 전용. route_id=0/미지정 경로는 절대 여기 오지 않음(마스터 완전 불변)."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT item_code, ISNULL(route_no,0), ISNULL(route_name,''), ISNULL(approve_flag,0) FROM nx.sourcing_route WHERE route_id=?", route_id)
        h = cur.fetchone()
        if not h:
            raise HTTPException(404, f"조달후보 route_id={route_id} 없음")
        ritem = str(h[0]).strip()
        cur.execute("""SELECT line_id, ISNULL(sort_seq,0), ISNULL(child_item,''), ISNULL(child_name,''), qty,
              ISNULL(gubun,''), ISNULL(vendor_code,''), ISNULL(is_rawmat,0), diam, thick, len_val, ISNULL(material,''),
              ISNULL(spec,''), ISNULL(node_kind,'PART'), parent_line, ISNULL(sub_item,'')
            FROM nx.sourcing_route_line WHERE route_id=? ORDER BY sort_seq, line_id""", route_id)
        lines = []
        for r in cur.fetchall():
            lines.append({"line_id": int(r[0]), "seq": int(r[1] or 0), "code": str(r[2]).strip(), "nm": (r[3] or ""),
                "qty": float(r[4] or 0), "gubun": str(r[5] or ""), "cust": str(r[6]).strip(), "is_rawmat": int(r[7] or 0),
                "diam": float(r[8] or 0), "thick": float(r[9] or 0), "length": float(r[10] or 0),
                "metal": str(r[11] or ""), "spec": str(r[12] or ""), "node_kind": str(r[13] or 'PART'),
                "parent_line": (int(r[14]) if r[14] is not None else None), "sub_item": str(r[15] or '')})
        # 벤더 코드→이름(라이브 CM_M_CUST, 크로스DB) + 제품명
        vcodes = sorted({l["cust"] for l in lines if l["cust"]})
        vmap = {}
        for i in range(0, len(vcodes), 900):
            ch = vcodes[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"SELECT CUST_CODE, ISNULL(CUST_DESC,'') FROM PARTNER_ERP_TEST3.nx.CM_M_CUST WHERE CUST_CODE IN ({ph})", *ch)
            for rr in cur.fetchall(): vmap[str(rr[0]).strip()] = rr[1]
        cur.execute("SELECT ISNULL(item_name,'') FROM nx.item WHERE item_code=?", ritem)
        rr = cur.fetchone(); rootnm = rr[0] if rr else ""
    finally:
        nx.close()
    kids = {}
    for l in lines:
        kids.setdefault(l["parent_line"], []).append(l)
    for k in kids: kids[k].sort(key=lambda x: x["seq"])
    out = [{"level": 0, "code": ritem, "nm": rootnm, "spec": "", "qty": 1, "cust": "", "custnm": "", "gubun": "제품",
            "sag": "", "se": "", "kt": "", "vir": "", "ce": "", "le": "", "gp": "", "sw": "",
            "metal": "", "diam": 0, "thick": 0, "length": 0, "haskids": bool(kids.get(None))}]
    def emit(parent_line, lvl):
        for l in kids.get(parent_line, []):
            haskids = (l["node_kind"] == 'SUB') or bool(kids.get(l["line_id"]))
            code = l["sub_item"] or l["code"]
            gub = l["gubun"]
            out.append({"level": lvl, "code": code, "nm": l["nm"], "spec": l["spec"], "qty": l["qty"],
                "cust": l["cust"], "custnm": vmap.get(l["cust"], l["cust"]), "gubun": gub,
                "sag": ("1" if ("사급" in gub) else ""), "se": "", "kt": "", "vir": "", "ce": "", "le": "",
                "gp": "", "sw": "", "metal": l["metal"], "diam": l["diam"], "thick": l["thick"], "length": l["length"],
                "haskids": haskids})
            if haskids:
                emit(l["line_id"], lvl + 1)
    emit(None, 1)
    maxlvl = max((r["level"] for r in out), default=0)
    return {"item": item, "name": rootnm, "rows": out, "count": len(out) - 1, "maxlevel": maxlvl,
            "route_id": route_id, "route_no": int(h[1]), "route_name": h[2], "is_route": True}

def _bom_tree_nx(item, real, expandbuy=0):
    """★단일 정규화 BOM(nx.bom_line) 전개 — #1 이관본. 용접봉(RAC) 제외(설계=공정종속 nx.proc_weld), 자도번→품번_S{nn}(nx.sub_alias) 표시 정규화.
       real=1: cs_calc_except=0(현행) + 라이브 PR_M_ITEM.MAKE_TYPE='1' 하위전개(현행 CS bom/tree와 동일 grain → 리프·수량 diff0).
       ★expandbuy=1(real=1과 함께): 현행필터(cs_calc_except=0)는 유지하되 MAKE_TYPE 게이트 제거 → 매입 SUB(예 명진 -19-1) 하위도 전개(비현행 변형 태국 -F&T=cs_except1은 여전히 제외). 라우팅 화면 '현행 전체전개'용.
       구조/수량=nx.bom_line, 상세(품명·매입처·치수)=라이브 PR_M_ITEM(자도번 코드 기준, 표시코드만 정규화)."""
    exc = "AND ISNULL(bl.cs_calc_except,0)=0" if real else ""
    mk = "" if (not real or expandbuy) else "JOIN PARTNER_ERP_TEST3.nx.PR_M_ITEM pt ON pt.ITEM_CODE=t.c AND ISNULL(pt.MAKE_TYPE,'')='1'"
    cn = _nx(); cur = cn.cursor()
    try:
        # ★성능(2026-08-13): 13컬럼 CAST를 재귀 안에서 계산하면 525ms. 재귀는 (부모,자식)만 슬림 전개(~10ms) →
        #   전개된 부모집합 대상으로 엣지 상세를 비재귀 1쿼리로 가져옴. 결과 동일, 원격DB 지연 대폭 감소.
        cur.execute(f"""WITH tree AS (
            SELECT h.item_code p, bl.child_item c, 1 lvl
            FROM nx.bom_header h JOIN nx.bom_line bl ON bl.bom_id=h.bom_id
            WHERE h.item_code=? {exc}
            UNION ALL
            SELECT h2.item_code, bl.child_item, t.lvl+1
            FROM tree t
            {mk}
            JOIN nx.bom_header h2 ON h2.item_code=t.c
            JOIN nx.bom_line bl ON bl.bom_id=h2.bom_id {exc}
            WHERE t.lvl < 8)
            SELECT DISTINCT p FROM tree OPTION(MAXRECURSION 50)""", item)
        parents = [(r[0] or '').strip() for r in cur.fetchall() if (r[0] or '').strip()]
        edges = {}
        if parents:
            pin = ",".join("N'" + p.replace("'", "''") + "'" for p in parents)
            cur.execute(f"""SELECT h.item_code p, bl.child_item c, CAST(bl.qty AS decimal(18,6)) q,
                   CAST(ISNULL(bl.sagub_default,0) AS int) sag, CAST(ISNULL(bl.set_except,0) AS int) se,
                   CAST(ISNULL(bl.kitting,0) AS int) kt, CAST(ISNULL(bl.vir_item,0) AS int) vir,
                   CAST(ISNULL(bl.cs_calc_except,0) AS int) ce, CAST(ISNULL(bl.lme_except,0) AS int) le,
                   ISNULL(bl.gagong_proc,'') gp, ISNULL(bl.s_work,'') sw, ISNULL(bl.seq,0) sq
                FROM nx.bom_header h JOIN nx.bom_line bl ON bl.bom_id=h.bom_id
                WHERE h.item_code IN ({pin}) {exc}""")
            for r in cur.fetchall():
                edges.setdefault((r[0] or '').strip(), []).append({
                    "child": (r[1] or '').strip(), "q": float(r[2] or 0), "sag": ('1' if r[3] else '0'),
                    "se": ('1' if r[4] else ''), "kt": ('1' if r[5] else ''), "vir": ('1' if r[6] else ''),
                    "ce": ('1' if r[7] else ''), "le": ('1' if r[8] else ''), "gp": str(r[9]).strip(),
                    "sw": str(r[10]).strip(), "sq": int(r[11] or 0)})
        for p in edges: edges[p].sort(key=lambda x: x["sq"])
        nl = list({item} | {e["child"] for lst in edges.values() for e in lst} | set(edges.keys()))
        info = {}; alias = {}
        for i in range(0, len(nl), 900):
            chunk = nl[i:i+900]
            inl = ",".join("N'" + str(x).replace("'", "''") + "'" for x in chunk)   # ★성능: 파라미터 IN(524ms)→N인라인(11ms). pyodbc param 오버헤드 회피
            cur.execute(f"""SELECT m.ITEM_CODE, ISNULL(m.ITEM_DESC,''), ISNULL(m.ITEM_SPEC,''),
                  ISNULL(m.IN_CUST_CODE,''), ISNULL(c.CUST_DESC,''), ISNULL(m.METAL_GUBUN,''),
                  ISNULL(m.ITEM_DIAM,0), ISNULL(m.ITEM_THICK,0), ISNULL(m.ITEM_LENGTH,0)
                FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM m LEFT JOIN PARTNER_ERP_TEST3.nx.CM_M_CUST c ON c.CUST_CODE=m.IN_CUST_CODE
                WHERE m.ITEM_CODE IN ({inl})""")
            for r in cur.fetchall():
                info[(r[0] or '').strip()] = {"nm": r[1], "spec": r[2], "cust": str(r[3]).strip(), "custnm": r[4],
                      "metal": r[5], "diam": float(r[6] or 0), "thick": float(r[7] or 0), "length": float(r[8] or 0)}
            # ★sub_alias 쿼리 제거(성능): 정리 후 canonical 전부 NULL=매핑 없음. 필요시 재도입. 원격DB 왕복 1회 절감.
        def disp(code): return alias.get(code, code)   # 리프변형 표시(자재 canonical). SUB는 아래 tree-order.
        # ★SUB 표시 = {ASSY품번}_S{순번} 트리순서(현재 임시=위치기반). raw=원본코드(navi/edit),
        #   정본식별 S#####(nx.sub_registry/sub_code_map)은 dedup·후보채번용 내부 identity(표시 아님).
        # ★★정본설계(2026-08-15, _schema/SUB_CODE_MASKS_REAL_ASSY.md §7-1): 코드=출생라벨 {첫작업ASSY}_R{첫route}_S{nn}
        #   (태어난 자리 박제·영구재사용·영속번호). 실제 제품(깨끗한코드)은 개명 금지=아래 subdisp에서 실제코드 유지(적용됨).
        #   출생라벨·영속번호·공용flag = 마이그레이션 정본화 과제(현재는 위치기반 표시 유지).
        sub_seq = [0]; sub_map = {}
        def subdisp(child):
            if child in edges:   # 하위 보유 = SUB
                # ★2026-08-15: 실제 제품코드(접미사 -N-N·_S 없는 깨끗한 코드)는 개명 금지 → 도면 품번 그대로 표시.
                #   합성/변형 SUB(base-N-N·_S)만 {부모}_S{nn} 채번. (실제 ASSY가 SUB코드로 가려지던 문제 수정)
                if ('-' not in child) and ('_' not in child):
                    return disp(child)
                if child not in sub_map:
                    sub_seq[0] += 1; sub_map[child] = f"{item}_S{sub_seq[0]:02d}"
                return sub_map[child]
            return disp(child)
        rootnm = info.get(item, {}).get("nm", "")
        out = [{"level": 0, "code": disp(item), "raw": item, "nm": rootnm, "spec": info.get(item, {}).get("spec", ""),
                "qty": 1, "cust": "", "custnm": "", "sag": "", "se": "", "kt": "", "vir": "", "ce": "", "le": "",
                "gp": "", "sw": "", "metal": info.get(item, {}).get("metal", ""),
                "diam": info.get(item, {}).get("diam", 0), "thick": info.get(item, {}).get("thick", 0), "length": info.get(item, {}).get("length", 0),
                "haskids": item in edges}]
        seen = set()
        def walk(code, lvl):
            if code in seen: return
            seen.add(code)
            for e in edges.get(code, []):
                ci = info.get(e["child"], {})
                out.append({"level": lvl, "code": subdisp(e["child"]), "raw": e["child"], "parent": code, "pseq": e.get("sq", 0), "nm": ci.get("nm", ""), "spec": ci.get("spec", ""),
                    "qty": e["q"], "cust": ci.get("cust", ""), "custnm": ci.get("custnm", ""),
                    "sag": e["sag"], "se": e["se"], "kt": e["kt"], "vir": e["vir"], "ce": e["ce"], "le": e["le"],
                    "gp": e["gp"], "sw": e["sw"], "metal": ci.get("metal", ""),
                    "diam": ci.get("diam", 0), "thick": ci.get("thick", 0), "length": ci.get("length", 0),
                    "haskids": e["child"] in edges})
                walk(e["child"], lvl + 1)
            seen.discard(code)
        walk(item, 1)
        maxlvl = max((r["level"] for r in out), default=0)
        return {"item": item, "name": rootnm, "rows": out, "count": len(out) - 1, "maxlevel": maxlvl, "src": "nx"}
    finally:
        cn.close()

@router.get("/api/bom/tree")
def bom_tree(item: str = Query(..., description="품번"), real: int = Query(1, description="1=실사용전개(원가제외 스킵+제작품만 전개,매입중단)=실원가용 일치, 0=전체전개"),
             route_id: int = Query(0, description="0/미지정=마스터 실사용 BOM(완전불변), >0=조달후보(nx.sourcing_route_line) 구조를 동일스키마로 반환"),
             src: str = Query('nx', description="nx=단일BOM nx.bom_line [기본, 컷오버후 라이브](용접봉제외+자도번→품번_S{nn}정규화, 레거시SP구조 39/40검증). cs=현행 라이브 CS_M_ITEM_BOM(대조·롤백용)"),
             expandbuy: int = Query(0, description="1(real=1과 함께): 현행필터(cs_calc_except=0) 유지하되 매입SUB 하위도 전개(비현행 변형은 여전히 제외). 라우팅 '현행 전체전개'용")):
    """다단계 BOM 트리(레벨별) — CS_M_ITEM_BOM 재귀전개. 매입처=컴포넌트 IN_CUST_CODE(현행 벤더).
    real=1(기본): 견적원가조회(실원가용, SP_CS_견적서(BOM)) 전개와 일치 — CS_CALC_EXCEPT_FLAG='1'(원가제외=현행아닌 조달경로) 제외 + MAKE_TYPE='1'(제작/자체)만 하위전개, 매입/구매품(구매완제)은 전개중단.
    route_id>0: 조달후보 계층(nx.sourcing_route_line)을 동일 트리 스키마로 반환(마스터 미조회)."""
    item = item.strip()
    if int(route_id or 0) > 0:
        return _bom_tree_route(item, int(route_id))
    real = 1 if real is None else int(real)
    if str(src or 'nx').lower() != 'cs':
        return _bom_tree_nx(item, real, int(expandbuy or 0))   # ★#1 이관: 단일 정규화 BOM(nx.bom_line). src=cs면 아래 현행 CS 전개(대조·롤백)
    exc_a = "AND ISNULL(CS_CALC_EXCEPT_FLAG,'0')<>'1'" if real else ""      # 원가제외 라인 스킵
    exc_r = "AND ISNULL(b.CS_CALC_EXCEPT_FLAG,'0')<>'1'" if real else ""
    mk_gate = "JOIN PARTNER_ERP_TEST3.nx.PR_M_ITEM pt ON pt.ITEM_CODE=t.c AND ISNULL(pt.MAKE_TYPE,'')='1'" if real else ""  # 제작품만 하위전개
    cn = _conn(); cur = cn.cursor()  # live PARTNER_ERP
    try:
        cur.execute(f"""WITH tree AS (
            SELECT ITEM_CODE p, MAT_CODE c, CAST(USE_QTY AS decimal(18,6)) q, ISNULL(SAGUB_FLAG,'0') sag,
                   ISNULL(SET_EXCEPT_FLAG,'') se, ISNULL(KITTING_FLAG,'') kt, ISNULL(VIR_ITEM_FLAG,'') vir,
                   ISNULL(CS_CALC_EXCEPT_FLAG,'') ce, ISNULL(LME_EXCEPT_FLAG,'') le,
                   ISNULL(GAGONG_PROC_CODE,'') gp, ISNULL(S_WORK_CODE,'') sw, ISNULL(BOM_SEQ,0) sq, 1 lvl
            FROM PARTNER_ERP_TEST3.nx.CS_M_ITEM_BOM WHERE ITEM_CODE=? AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101' {exc_a}
            UNION ALL
            SELECT b.ITEM_CODE, b.MAT_CODE, CAST(b.USE_QTY AS decimal(18,6)), ISNULL(b.SAGUB_FLAG,'0'),
                   ISNULL(b.SET_EXCEPT_FLAG,''), ISNULL(b.KITTING_FLAG,''), ISNULL(b.VIR_ITEM_FLAG,''),
                   ISNULL(b.CS_CALC_EXCEPT_FLAG,''), ISNULL(b.LME_EXCEPT_FLAG,''),
                   ISNULL(b.GAGONG_PROC_CODE,''), ISNULL(b.S_WORK_CODE,''), ISNULL(b.BOM_SEQ,0), t.lvl+1
            FROM tree t JOIN PARTNER_ERP_TEST3.nx.CS_M_ITEM_BOM b ON b.ITEM_CODE=t.c AND b.FROM_APPLY_YMD<='991231' AND b.TO_APPLY_YMD>='260101' {exc_r}
            {mk_gate}
            WHERE t.lvl < 8)
            SELECT p,c,q,sag,se,kt,vir,ce,le,gp,sw,sq,lvl FROM tree OPTION(MAXRECURSION 50)""", item)
        edges = {}
        for r in cur.fetchall():
            edges.setdefault(r[0], []).append({
                "child": r[1], "q": float(r[2] or 0), "sag": str(r[3]), "se": str(r[4]), "kt": str(r[5]),
                "vir": str(r[6]), "ce": str(r[7]), "le": str(r[8]), "gp": str(r[9]).strip(),
                "sw": str(r[10]).strip(), "sq": int(r[11] or 0)})
        for p in edges:
            edges[p].sort(key=lambda x: x["sq"])
        # 노드 상세(품명·매입처=IN_CUST·치수) 일괄
        nodes = {item} | {e["child"] for lst in edges.values() for e in lst} | set(edges.keys())
        info = {}
        if nodes:
            nl = list(nodes)
            for i in range(0, len(nl), 900):
                chunk = nl[i:i+900]; ph = ",".join("?" * len(chunk))
                cur.execute(f"""SELECT m.ITEM_CODE, ISNULL(m.ITEM_DESC,''), ISNULL(m.ITEM_SPEC,''),
                      ISNULL(m.IN_CUST_CODE,''), ISNULL(c.CUST_DESC,''), ISNULL(m.METAL_GUBUN,''),
                      ISNULL(m.ITEM_DIAM,0), ISNULL(m.ITEM_THICK,0), ISNULL(m.ITEM_LENGTH,0)
                    FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM m LEFT JOIN PARTNER_ERP_TEST3.nx.CM_M_CUST c ON c.CUST_CODE=m.IN_CUST_CODE
                    WHERE m.ITEM_CODE IN ({ph})""", *chunk)
                for r in cur.fetchall():
                    info[r[0]] = {"nm": r[1], "spec": r[2], "cust": str(r[3]).strip(), "custnm": r[4],
                                  "metal": r[5], "diam": float(r[6] or 0), "thick": float(r[7] or 0), "length": float(r[8] or 0)}
        rootnm = info.get(item, {}).get("nm", "")
        out = [{"level": 0, "code": item, "nm": rootnm, "spec": info.get(item, {}).get("spec", ""),
                "qty": 1, "cust": "", "custnm": "", "sag": "", "se": "", "kt": "", "vir": "", "ce": "", "le": "",
                "gp": "", "sw": "", "metal": info.get(item, {}).get("metal", ""),
                "diam": info.get(item, {}).get("diam", 0), "thick": info.get(item, {}).get("thick", 0), "length": info.get(item, {}).get("length", 0),
                "haskids": item in edges}]
        seen = set()
        def walk(code, lvl):
            if code in seen: return   # 순환 방지
            seen.add(code)
            for e in edges.get(code, []):
                ci = info.get(e["child"], {})
                out.append({"level": lvl, "code": e["child"], "nm": ci.get("nm", ""), "spec": ci.get("spec", ""),
                    "qty": e["q"], "cust": ci.get("cust", ""), "custnm": ci.get("custnm", ""),
                    "sag": e["sag"], "se": e["se"], "kt": e["kt"], "vir": e["vir"], "ce": e["ce"], "le": e["le"],
                    "gp": e["gp"], "sw": e["sw"], "metal": ci.get("metal", ""),
                    "diam": ci.get("diam", 0), "thick": ci.get("thick", 0), "length": ci.get("length", 0),
                    "haskids": e["child"] in edges})
                walk(e["child"], lvl + 1)
            seen.discard(code)
        walk(item, 1)
        maxlvl = max((r["level"] for r in out), default=0)
        return {"item": item, "name": rootnm, "rows": out, "count": len(out) - 1, "maxlevel": maxlvl}
    finally:
        cn.close()

@router.get("/api/bom/whereused")
def bom_whereused(item: str = Query(..., description="품번 — 이 품번을 하위자재로 쓰는 상위 품번들을 역전개(where-used)")):
    """역전개(where-used): 이 품번을 하위구성으로 쓰는 상위 품번들을 다단계 상향전개.
    ★우리 재설계 단일BOM(nx.bom_header/bom_line) 미러 — forward(_bom_tree_nx)의 상하반전
      (bl.child_item=대상 → h.item_code=부모, 상향 재귀). 자도번→품번_S{nn} 표시정규화(nx.sub_alias) 동일 적용.
    where-used는 원가제외/사급 여부와 무관하게 '어디에 쓰이나'를 모두 보여주되 플래그(제외·사급·세트제외)는 표시."""
    item = item.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""WITH up AS (
            SELECT h.item_code p, bl.child_item c, CAST(bl.qty AS decimal(18,6)) q,
                   CAST(ISNULL(bl.sagub_default,0) AS int) sag, CAST(ISNULL(bl.set_except,0) AS int) se,
                   CAST(ISNULL(bl.cs_calc_except,0) AS int) ce, ISNULL(bl.seq,0) sq, 1 lvl
            FROM nx.bom_line bl JOIN nx.bom_header h ON h.bom_id=bl.bom_id
            WHERE bl.child_item=?
            UNION ALL
            SELECT h.item_code, bl.child_item, CAST(bl.qty AS decimal(18,6)),
                   CAST(ISNULL(bl.sagub_default,0) AS int), CAST(ISNULL(bl.set_except,0) AS int),
                   CAST(ISNULL(bl.cs_calc_except,0) AS int), ISNULL(bl.seq,0), u.lvl+1
            FROM up u
            JOIN nx.bom_line bl ON bl.child_item=u.p
            JOIN nx.bom_header h ON h.bom_id=bl.bom_id
            WHERE u.lvl < 8)
            SELECT p,c,q,sag,se,ce,sq,lvl FROM up OPTION(MAXRECURSION 50)""", item)
        parents = {}   # child c -> [ {부모 p, 소요량 q, 플래그...} ]  (c를 쓰는 상위 품번들)
        for r in cur.fetchall():
            parents.setdefault((r[1] or '').strip(), []).append({
                "parent": (r[0] or '').strip(), "q": float(r[2] or 0),
                "sag": ('1' if r[3] else '0'), "se": ('1' if r[4] else ''), "ce": ('1' if r[5] else ''),
                "sq": int(r[6] or 0)})
        for c in parents: parents[c].sort(key=lambda x: x["parent"])
        nodes = {item} | {p["parent"] for lst in parents.values() for p in lst} | set(parents.keys())
        nl = list(nodes); info = {}; alias = {}
        for i in range(0, len(nl), 900):
            chunk = nl[i:i+900]
            inl = ",".join("N'" + str(x).replace("'", "''") + "'" for x in chunk)   # ★성능: 파라미터 IN→N인라인(524ms→11ms)
            cur.execute(f"""SELECT m.ITEM_CODE, ISNULL(m.ITEM_DESC,''), ISNULL(m.ITEM_SPEC,''),
                  ISNULL(m.IN_CUST_CODE,''), ISNULL(c.CUST_DESC,''), ISNULL(m.METAL_GUBUN,''),
                  ISNULL(m.ITEM_DIAM,0), ISNULL(m.ITEM_THICK,0), ISNULL(m.ITEM_LENGTH,0)
                FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM m LEFT JOIN PARTNER_ERP_TEST3.nx.CM_M_CUST c ON c.CUST_CODE=m.IN_CUST_CODE
                WHERE m.ITEM_CODE IN ({inl})""")
            for r in cur.fetchall():
                info[(r[0] or '').strip()] = {"nm": r[1], "spec": r[2], "cust": str(r[3]).strip(), "custnm": r[4],
                      "metal": r[5], "diam": float(r[6] or 0), "thick": float(r[7] or 0), "length": float(r[8] or 0)}
            # ★sub_alias 쿼리 제거(성능): 정리 후 canonical 전부 NULL. 원격DB 왕복 1회 절감.
        def disp(code): return alias.get(code, code)   # 역전개=원본코드 표시(글로벌 S코드 미표시, 정본식별은 내부 dedup용)
        rootnm = info.get(item, {}).get("nm", "")
        out = [{"level": 0, "code": disp(item), "raw": item, "nm": rootnm, "spec": info.get(item, {}).get("spec", ""),
                "qty": None, "custnm": "", "sag": "", "se": "", "ce": "",
                "metal": info.get(item, {}).get("metal", ""), "diam": info.get(item, {}).get("diam", 0),
                "thick": info.get(item, {}).get("thick", 0), "hasparents": item in parents}]
        seen = set()
        def walk(code, lvl):
            if code in seen: return   # 순환 방지
            seen.add(code)
            for p in parents.get(code, []):
                pi = info.get(p["parent"], {})
                out.append({"level": lvl, "code": disp(p["parent"]), "raw": p["parent"], "nm": pi.get("nm", ""),
                    "spec": pi.get("spec", ""), "qty": p["q"], "custnm": pi.get("custnm", ""),
                    "sag": p["sag"], "se": p["se"], "ce": p["ce"], "metal": pi.get("metal", ""),
                    "diam": pi.get("diam", 0), "thick": pi.get("thick", 0),
                    "hasparents": p["parent"] in parents})
                walk(p["parent"], lvl + 1)
            seen.discard(code)
        walk(item, 1)
        maxlvl = max((r["level"] for r in out), default=0)
        return {"item": item, "name": rootnm, "rows": out, "count": len(out) - 1, "maxlevel": maxlvl, "src": "nx"}
    finally:
        cn.close()


def _sub_signature(cur, children, weld):
    """SUB 시그니처(레지스트리와 동일 규칙): children[RAC%제외, 자식SUB는 자기 sig로]+weld[weld_item·st·use_qty] → 'S:'+md5[:12].
       children=[{item,qty}], weld=[{weld_item,weld_st,use_qty}]. cur=nx 커서(자식 sig 조회용)."""
    codes = [str(c.get("item", "")).strip() for c in children
             if str(c.get("item", "")).strip() and not str(c.get("item", "")).upper().startswith('RAC')]
    keymap = {}   # 자식코드 -> 'S:xxxx'(SUB) ; 없으면 리프
    if codes:
        uc = list(dict.fromkeys(codes))
        for i in range(0, len(uc), 900):
            chunk = uc[i:i+900]; ph = ",".join("?" * len(chunk))
            cur.execute(f"""SELECT m.raw_item, r.sig FROM nx.sub_code_map m
                            JOIN nx.sub_registry r ON r.sub_code=m.sub_code WHERE m.raw_item IN ({ph})""", *chunk)
            for raw, sig in cur.fetchall(): keymap[(raw or '').strip()] = sig
            cur.execute(f"SELECT sub_code, sig FROM nx.sub_registry WHERE sub_code IN ({ph})", *chunk)
            for sc, sig in cur.fetchall(): keymap[(sc or '').strip()] = sig
    items = []
    for c in children:
        it = str(c.get("item", "")).strip()
        if not it or it.upper().startswith('RAC'): continue
        q = round(float(c.get("qty", 1) or 1), 6)
        items.append((it, q, keymap.get(it, 'L:' + it)))
    items.sort(key=lambda x: (x[0], x[1]))
    parts = [f"{key}#{q}" for it, q, key in items]
    wl = sorted((str(w.get("weld_item", "")).strip(), round(float(w.get("weld_st", 0) or 0), 4),
                 round(float(w.get("use_qty", 0) or 0), 6)) for w in (weld or []))
    weldstr = ';'.join(f"{wi}|{st}|{uq}" for wi, st, uq in wl)
    raw = f"C[{','.join(parts)}]W[{weldstr}]"
    return 'S:' + hashlib.md5(raw.encode('utf-8')).hexdigest()[:12]


def _mint_sub(cur, sig, rep_item, nm=''):
    """레지스트리 dedup-safe 등록(append-only): sig 존재 시 기존 S반환(신규 생성 안 함=완전차단),
       없으면 다음 S##### 발급+nx.sub_registry/sub_code_map 등록. 반환 (sub_code, is_new).
       ★주의: 최초 시드(sub_registry_build DROP+재빌드) 이후 레지스트리는 append-only. mint된 후보SUB는 bom_line에 없으므로 DROP+재빌드 재실행 금지(코드 안정성)."""
    rep_item = (rep_item or '').strip()
    cur.execute("SELECT sub_code FROM nx.sub_registry WHERE sig=?", sig)
    r = cur.fetchone()
    if r:
        return ((r[0] or '').strip(), False)
    cur.execute("SELECT ISNULL(MAX(CAST(SUBSTRING(sub_code,2,10) AS INT)),0) FROM nx.sub_registry WHERE sub_code LIKE 'S[0-9][0-9][0-9][0-9][0-9]'")
    code = f"S{int(cur.fetchone()[0]) + 1:05d}"
    cur.execute("INSERT INTO nx.sub_registry(sub_code,sig,rep_item,nm,members) VALUES(?,?,?,?,1)", code, sig, rep_item[:50], (nm or '')[:200])
    cur.execute("IF NOT EXISTS(SELECT 1 FROM nx.sub_code_map WHERE raw_item=?) INSERT INTO nx.sub_code_map(raw_item,sub_code) VALUES(?,?)", rep_item[:50], rep_item[:50], code)
    return (code, True)


@router.post("/api/bom/sub/dedup")
def sub_dedup(payload: dict = Body(...)):
    """★SUB 재사용 판정(조달후보 모달 '묶기' 시): children+weld의 시그니처로 nx.sub_registry 조회.
       동일 SUB 있으면 기존 S코드 반환(강제 재사용) → 중복 생성 차단. 읽기전용(채번은 후보 저장시)."""
    children = payload.get("children", []) or []
    weld = payload.get("weld", []) or []
    if not children:
        raise HTTPException(400, "children(구성) 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        sig = _sub_signature(cur, children, weld)
        cur.execute("SELECT sub_code, rep_item, nm, members FROM nx.sub_registry WHERE sig=?", sig)
        r = cur.fetchone()
        if r:
            return {"exists": True, "sub_code": (r[0] or '').strip(), "rep_item": (r[1] or '').strip(),
                    "nm": r[2] or '', "members": int(r[3] or 0), "sig": sig}
        return {"exists": False, "sub_code": None, "sig": sig}
    finally:
        cn.close()


@router.get("/api/bom/sub/impact")
def sub_impact(item: str = Query(..., description="SUB의 raw코드 또는 정본 S코드")):
    """★공유 영향 — 이 SUB(=1개 bom_header) 편집이 몇 개 제품에 자동 전파되나.
       direct_parents = 이 raw코드를 직속 참조하는 부모(같은 bom_header 공유 → 편집 시 **자동 전파** 대상).
       ※ 구조만 같은 변형(다른 raw코드)은 별도 bom_header라 자동전파 아님 — 여기서 병합/처리 안 함(사용자 결정)."""
    item = item.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        if item[:1] == 'S' and item[1:].isdigit():
            scode = item
            cur.execute("SELECT rep_item FROM nx.sub_registry WHERE sub_code=?", item)
            rr = cur.fetchone(); target_raw = ((rr[0] or '').strip() if rr and rr[0] else item)
        else:
            cur.execute("SELECT sub_code FROM nx.sub_code_map WHERE raw_item=?", item)
            r = cur.fetchone(); scode = ((r[0] or '').strip() if r else None); target_raw = item
        # direct = 이 raw를 직속 자식으로 갖는 부모(자동전파 대상) + 이름
        cur.execute("""SELECT DISTINCT h.item_code, ISNULL(m.ITEM_DESC,'')
                       FROM nx.bom_line bl JOIN nx.bom_header h ON h.bom_id=bl.bom_id
                       LEFT JOIN PARTNER_ERP_TEST3.nx.PR_M_ITEM m ON m.ITEM_CODE=h.item_code
                       WHERE bl.child_item=?""", target_raw)
        direct = [{"code": (x[0] or '').strip(), "name": x[1] or ''} for x in cur.fetchall() if (x[0] or '').strip()]
        return {"item": item, "sub_code": scode, "target_raw": target_raw,
                "direct_parents": direct, "direct_count": len(direct), "shared": len(direct) > 1}
    finally:
        cn.close()


@router.post("/api/bom/save")
def bom_save(payload: dict = Body(...)):
    """BOM 구성 전체 교체 저장. 마스터 가드: 참조무결성·중복·순환·필수값. (마감/재고 가드는 재고·실적 프로그램용, BOM 미적용)"""
    item = str(payload.get("item", "")).strip()
    lines = payload.get("lines", []) or []
    if not item:
        raise HTTPException(400, "item(품번) 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", item)
        if not cur.fetchone():
            # 신규 BOM: 대상 도번 미등록 시 최소등록(프론트에서 "신규 생성?" 확인 후 target_name 전달)
            tname = str(payload.get("target_name", "") or "").strip() or item
            cur.execute("INSERT INTO nx.item(item_code,item_name,item_type) VALUES(?,?,N'제품')", item, tname)
        errs, seen = [], set()
        for i, ln in enumerate(lines, 1):
            ch = str(ln.get("child_item", "")).strip()
            if not ch:
                errs.append(f"{i}행: 자품목코드 필요"); continue
            if ch == item:
                errs.append(f"{i}행: 자기참조 불가 ({ch})")
            if ch in seen:
                errs.append(f"{i}행: 중복 자품목 ({ch})")
            seen.add(ch)
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", ch)
            if not cur.fetchone():
                errs.append(f"{i}행: 미등록 품목 ({ch})")
            # 순환: 자품목의 BOM이 이 품목을 포함하면 순환
            cur.execute("""SELECT 1 FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id
                           WHERE h.item_code=? AND l.child_item=?""", ch, item)
            if cur.fetchone():
                errs.append(f"{i}행: 순환참조 ({ch} 가 {item} 를 이미 포함)")
        if errs:
            return {"ok": False, "errors": errs}
        # 헤더 확보
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", item)
        h = cur.fetchone()
        if h:
            bom_id = h[0]
        else:
            cur.execute("INSERT INTO nx.bom_header(item_code,version,apply_from,status) VALUES(?,1,'2000-01-01',N'확정')", item)
            cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", item)
            bom_id = cur.fetchone()[0]
        # 전체 교체 — ★용접봉(RAC)은 BOM구성행 아님·공정종속 자재 → nx.proc_weld로 라우팅(bom_line 제외)
        cur.execute("DELETE FROM nx.bom_line WHERE bom_id=?", bom_id)
        cur.execute("IF OBJECT_ID('nx.proc_weld','U') IS NOT NULL DELETE FROM nx.proc_weld WHERE parent_item=?", item)
        seq = 0; nweld = 0
        for ln in lines:
            ch = str(ln.get("child_item", "")).strip()
            if ch.upper().startswith("RAC"):   # 용접봉 → proc_weld(공정종속 자재)
                cur.execute("""INSERT INTO nx.proc_weld(parent_item,weld_item,weld_base,use_qty,cs_calc_except,lme_except,from_ymd,to_ymd,tag,src)
                    VALUES(?,?,?,?,?,?,?,?,'W','bom_save')""",
                    item, ch, ch.split('-')[0], float(ln.get("qty") or 0),
                    _b(ln.get("cs_calc_except")), _b(ln.get("lme_except")),
                    (ln.get("from_ymd") or None), (ln.get("to_ymd") or None))
                nweld += 1; continue
            seq += 1
            cur.execute("""INSERT INTO nx.bom_line
                (bom_id,seq,child_item,qty,node_type,cs_calc_except,lme_except,sagub_default,is_optional,
                 from_ymd,to_ymd,except_flag,set_except,kitting,vir_item,proc_gubun,gagong_proc,s_work,wh_gagong,in_gagong,cust_code,remarks)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                bom_id, seq, ch, float(ln.get("qty") or 0),
                str(ln.get("node_type") or "부품"),
                _b(ln.get("cs_calc_except")), _b(ln.get("lme_except")), _b(ln.get("sagub_default")), _b(ln.get("is_optional")),
                (ln.get("from_ymd") or None), (ln.get("to_ymd") or None),
                _b(ln.get("except_flag")), _b(ln.get("set_except")), _b(ln.get("kitting")), _b(ln.get("vir_item")),
                (ln.get("proc_gubun") or None), (ln.get("gagong_proc") or None), (ln.get("s_work") or None),
                (ln.get("wh_gagong") or None), (ln.get("in_gagong") or None), (ln.get("cust_code") or None),
                (ln.get("remarks") or None))
        _reset_cost_engine()   # BOM 구성/소요량 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "count": seq, "weld": nweld, "bom_id": bom_id}
    finally:
        cn.close()

@router.post("/api/bom/delete")
def bom_delete(payload: dict = Body(...)):
    """품번 삭제 — 레거시 방식(하위 구성 제거 후 품번 삭제). BOM구성(자식관계)·용접공정·서브테이블 제거 후 nx.item에서 품번 삭제.
       ★가드: 이 품번을 자식으로 쓰는 다른 BOM이 하나라도 있으면 차단(사용중=삭제불가)."""
    item = str(payload.get("item", "")).strip()
    if not item:
        raise HTTPException(400, "item(품번) 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", item)
        if not cur.fetchone():
            return {"ok": False, "errors": [f"미등록 품번 ({item})"]}
        # ★가드: 다른 BOM이 이 품번을 자식으로 사용 → 삭제불가(상위에서 먼저 제거해야 함)
        cur.execute("""SELECT DISTINCT TOP 20 h.item_code FROM nx.bom_line l
                       JOIN nx.bom_header h ON h.bom_id=l.bom_id WHERE l.child_item=?""", item)
        used = [r[0] for r in cur.fetchall()]
        if used:
            more = "..." if len(used) >= 20 else ""
            return {"ok": False, "errors": [
                f"삭제 불가 — 이 품번을 자식(구성)으로 사용하는 BOM {len(used)}건이 있습니다. 상위 BOM에서 먼저 제거하세요:",
                ", ".join(used[:15]) + more]}
        # 삭제: 하위 구성(bom_line) → 헤더 → 용접공정 → 서브테이블 → 품번(마스터)
        nline = 0
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", item)
        h = cur.fetchone()
        if h:
            cur.execute("DELETE FROM nx.bom_line WHERE bom_id=?", h[0]); nline = cur.rowcount
            cur.execute("DELETE FROM nx.bom_header WHERE bom_id=?", h[0])
        cur.execute("IF OBJECT_ID('nx.proc_weld','U') IS NOT NULL DELETE FROM nx.proc_weld WHERE parent_item=?", item)
        for t in ("item_sub", "item_valve", "routing", "item_weld"):
            cur.execute(f"IF OBJECT_ID('nx.{t}','U') IS NOT NULL DELETE FROM nx.{t} WHERE item_code=?", item)
        cur.execute("DELETE FROM nx.item WHERE item_code=?", item)
        nitem = cur.rowcount
        cn.commit()
        _reset_cost_engine()   # 구성·품목 삭제 → 원가엔진 캐시 무효화
        return {"ok": True, "item": item, "lines_removed": nline, "item_removed": nitem}
    finally:
        cn.close()

@router.post("/api/bom/deleteline")
def bom_deleteline(payload: dict = Body(...)):
    """다단계 편집: 특정 부모 BOM에서 자식 1행(구성) 삭제. body {parent, child, seq?}.
       seq 주면 정확히 그 라인만, 없으면 (parent,child) 매칭 전부. 용접봉(RAC)은 proc_weld에서 제거."""
    parent = str(payload.get("parent", "")).strip()
    child = str(payload.get("child", "")).strip()
    seq = payload.get("seq")
    if not parent or not child:
        raise HTTPException(400, "parent·child 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        if child.upper().startswith("RAC"):   # 용접봉 = 공정종속(proc_weld)
            cur.execute("IF OBJECT_ID('nx.proc_weld','U') IS NOT NULL DELETE FROM nx.proc_weld WHERE parent_item=? AND weld_item=?", parent, child)
            n = cur.rowcount; cn.commit(); _reset_cost_engine()
            return {"ok": True, "deleted": n, "parent": parent, "child": child, "weld": True}
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", parent)
        h = cur.fetchone()
        if not h:
            return {"ok": False, "errors": [f"부모 BOM 없음 ({parent})"]}
        bom_id = h[0]
        if seq is not None and str(seq) != "":
            cur.execute("DELETE FROM nx.bom_line WHERE bom_id=? AND child_item=? AND seq=?", bom_id, child, int(seq))
        else:
            cur.execute("DELETE FROM nx.bom_line WHERE bom_id=? AND child_item=?", bom_id, child)
        n = cur.rowcount
        cn.commit()
        _reset_cost_engine()   # 구성 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "deleted": n, "parent": parent, "child": child}
    finally:
        cn.close()

@router.post("/api/bom/addline")
def bom_addline(payload: dict = Body(...)):
    """다단계 편집: 부모 BOM에 자식 1행 추가(삭제 되돌리기·구성추가). body {parent, child, qty}.
       가드: 부모·자식 등록확인·자기참조·순환·(용접봉 RAC은 proc_weld). 중복은 허용(dup 카운트 반환)."""
    parent = str(payload.get("parent", "")).strip()
    child = str(payload.get("child", "")).strip().upper()
    try:
        qty = float(payload.get("qty") or 1)
    except (TypeError, ValueError):
        qty = 1.0
    if not parent or not child:
        raise HTTPException(400, "parent·child 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", parent)
        if not cur.fetchone():
            return {"ok": False, "errors": [f"부모 미등록 ({parent})"]}
        if child == parent:
            return {"ok": False, "errors": ["자기참조 불가"]}
        # 헤더 확보
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", parent)
        h = cur.fetchone()
        if h:
            bom_id = h[0]
        else:
            cur.execute("INSERT INTO nx.bom_header(item_code,version,apply_from,status) VALUES(?,1,'2000-01-01',N'확정')", parent)
            cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", parent); bom_id = cur.fetchone()[0]
        if child.upper().startswith("RAC"):   # 용접봉 = 공정종속(proc_weld)
            cur.execute("""INSERT INTO nx.proc_weld(parent_item,weld_item,weld_base,use_qty,tag,src)
                VALUES(?,?,?,?,'W','addline')""", parent, child, child.split('-')[0], qty)
            cn.commit(); _reset_cost_engine()
            return {"ok": True, "parent": parent, "child": child, "weld": True, "qty": qty}
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", child)
        if not cur.fetchone():
            return {"ok": False, "errors": [f"자식 미등록 품목 ({child}) — 품목마스터에 먼저 등록하세요"]}
        # 순환: 자식의 BOM이 부모를 포함하면 순환
        cur.execute("""SELECT 1 FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id
                       WHERE h.item_code=? AND l.child_item=?""", child, parent)
        if cur.fetchone():
            return {"ok": False, "errors": [f"순환참조 ({child} 가 {parent} 를 이미 포함)"]}
        cur.execute("SELECT COUNT(*) FROM nx.bom_line WHERE bom_id=? AND child_item=?", bom_id, child)
        dup = cur.fetchone()[0]
        cur.execute("SELECT ISNULL(MAX(seq),0)+1 FROM nx.bom_line WHERE bom_id=?", bom_id)
        seq = cur.fetchone()[0]
        cur.execute("INSERT INTO nx.bom_line(bom_id,seq,child_item,qty,node_type) VALUES(?,?,?,?,N'부품')", bom_id, seq, child, qty)
        cn.commit(); _reset_cost_engine()
        return {"ok": True, "parent": parent, "child": child, "seq": seq, "qty": qty, "dup": dup}
    finally:
        cn.close()

@router.post("/api/bom/qty")
def bom_qty(payload: dict = Body(...)):
    """BOM 단건 소요량 수정(부모+자식) — 다른 라인 속성 보존(전체교체 아님). 내부원가 수정에서 안전 갱신."""
    parent = str(payload.get("parent", "")).strip()
    child = str(payload.get("child", "")).strip()
    if not parent or not child:
        raise HTTPException(400, "parent·child 필요")
    try:
        q = float(payload.get("qty"))
    except Exception:
        raise HTTPException(400, "qty(숫자) 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""UPDATE l SET l.qty=? FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id
            WHERE h.item_code=? AND l.child_item=?""", q, parent, child)
        n = cur.rowcount
        _reset_cost_engine()
        return {"ok": n > 0, "updated": int(n)}
    finally:
        cn.close()

@router.post("/api/item/spec")
def item_spec(payload: dict = Body(...)):
    """품목 원소재 스펙 단건 수정(외경·두께·재질·중량) — ★지정한 필드만 갱신(item/save와 달리 미지정 필드 NULL 덮어쓰기 없음)."""
    code = str(payload.get("item_code", "")).strip()
    if not code:
        raise HTTPException(400, "item_code 필요")
    sets, vals = [], []
    for key, col in (("diam", "diam"), ("thick", "thick"), ("metal_gubun", "metal_gubun"), ("net_weight", "net_weight")):
        if key in payload and payload[key] not in (None, ""):
            if col == "metal_gubun":
                v = str(payload[key]).strip()
            else:
                try: v = float(payload[key])
                except Exception: raise HTTPException(400, f"{key}(숫자) 필요")
            sets.append(f"{col}=?"); vals.append(v)
    if not sets:
        return {"ok": True, "updated": 0}
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute(f"UPDATE nx.item SET {','.join(sets)} WHERE item_code=?", *vals, code)
        n = cur.rowcount
        _reset_cost_engine()
        return {"ok": n > 0, "updated": int(n)}
    finally:
        cn.close()

@router.post("/api/item/cutgubun")
def item_cutgubun(payload: dict = Body(...)):
    """품목 절삭/설치 구분(cut_gubun) 단건 수정 — nx.item.cut_gubun. 값: 절삭/설치/분지관/이지링크/빈값(미분류).
       경영 대시보드·영업예상매출 절삭/설치 분류의 정본(품목마스터 속성). 원가 무영향."""
    code = str(payload.get("item_code", "")).strip()
    if not code:
        raise HTTPException(400, "item_code 필요")
    g = str(payload.get("cut_gubun", "")).strip()
    if g not in ("", "절삭", "설치", "분지관", "이지링크"):
        raise HTTPException(400, "cut_gubun 값 오류(절삭/설치/분지관/이지링크/빈값)")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("IF COL_LENGTH('nx.item','cut_gubun') IS NULL ALTER TABLE nx.item ADD cut_gubun nvarchar(20)")
        cur.execute("UPDATE nx.item SET cut_gubun=NULLIF(?,'') WHERE item_code=?", g, code)
        return {"ok": cur.rowcount > 0, "updated": int(cur.rowcount)}
    finally:
        cn.close()


# ===================== LG BOM 조회 (nx.lg_bom, LG 원본 BOM Explosion 56,522행) =====================
@router.get("/api/lgbom/search")
def lgbom_search(q: str = Query(""), werks: str = Query(""), limit: int = Query(200)):
    """LG BOM 상위모델(완제품) 검색. model=상위품번. werks: DMZ(SAC)/DGZ(RAC). 유효기간 현행."""
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["b.stufe=1"]; p = []
        if q: w.append("(b.model LIKE ? OR b.parent_code LIKE ?)"); p += [f"%{q}%", f"%{q}%"]
        if werks: w.append("b.werks=?"); p.append(werks)
        cur.execute(f"""SELECT TOP {int(limit)} b.model, b.werks, MAX(b.parent_code) parent_code,
              MAX(ISNULL(pi.ITEM_DESC,'')) modelnm, COUNT(*) child_cnt, MAX(b.valid_from) valid_from
            FROM nx.lg_bom b LEFT JOIN PARTNER_ERP_TEST3.nx.PR_M_ITEM pi ON pi.ITEM_CODE=b.model
            WHERE {' AND '.join(w)} GROUP BY b.model, b.werks ORDER BY b.model""", *p)
        cols = [d[0] for d in cur.description]
        return {"rows": [dict(zip(cols, r)) for r in cur.fetchall()]}
    finally:
        cn.close()

@router.get("/api/lgbom/tree")
def lgbom_tree(model: str = Query(...), werks: str = Query("")):
    """선택 모델의 LG BOM 전개(전 레벨). stufe/posnr 순. 프론트에서 parent_code→child_code 트리 조립."""
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["b.model=?"]; p = [model]
        if werks: w.append("b.werks=?"); p.append(werks)
        cur.execute(f"""SELECT b.id, b.werks, b.stufe, b.posnr, b.parent_code, b.child_code,
              b.child_desc, b.child_spec, b.qty, b.unit, b.supply_type, b.mmsta, b.matty, b.lowest_flg,
              b.main_mat, b.matkl, b.valid_from, b.valid_to, ISNULL(i.ITEM_DESC,'') nx_desc
            FROM nx.lg_bom b LEFT JOIN PARTNER_ERP_TEST3.nx.PR_M_ITEM i ON i.ITEM_CODE=b.child_code
            WHERE {' AND '.join(w)} ORDER BY b.stufe, b.posnr, b.id""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        # 최상위 parent(model) 정보
        cur.execute("SELECT ISNULL(ITEM_DESC,'') FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM WHERE ITEM_CODE=?", model)
        mn = cur.fetchone()
        return {"model": model, "modelnm": (mn[0] if mn else ""), "rows": rows}
    finally:
        cn.close()

@router.post("/api/lgbom/upload")
async def lgbom_upload(file: UploadFile = File(...)):
    """LG BOM Explosion 엑셀 업로드 → nx.lg_bom 적재(모델·werks별 교체). 신규 BOM 등록 전 사전업로드용.
       헤더 1행: MODEL/WERKS/STUFE/POSNR/MATNR(부모)/IDNRK(자식)/OJTXP(품명)/CHI_SPECI(규격)/MENGE(수량)/MEINS(단위)/ETEXT(supply_type)/MATTY/LOWEST_FLG/MATKL/DATAB/DATVT 등."""
    import io as _io
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb.active
    itr = ws.iter_rows(values_only=True)
    try:
        hdr = next(itr)
    except StopIteration:
        raise HTTPException(400, "빈 파일")
    ix = {str(h or '').strip().upper(): i for i, h in enumerate(hdr)}
    if not all(k in ix for k in ('MODEL', 'MATNR', 'IDNRK')):
        raise HTTPException(400, "LG BOM 형식 아님 — 헤더 1행에 MODEL·MATNR·IDNRK 필요")
    def gv(r, n):
        i = ix.get(n); v = r[i] if (i is not None and i < len(r)) else None
        return None if v in (None, '') else v
    def gs(r, n, ln=None):
        v = gv(r, n); s = '' if v is None else str(v).strip(); return s[:ln] if ln else s
    def gi(r, n):
        try: return int(float(gv(r, n) or 0))
        except Exception: return 0
    def gf(r, n):
        try: return float(gv(r, n) or 0)
        except Exception: return 0.0
    recs = []; models = set()
    for r in itr:
        if not r or not any(x not in (None, '') for x in r): continue
        model = gs(r, 'MODEL')
        if not model: continue
        werks = gs(r, 'WERKS', 4)
        recs.append(('C', werks, model, gi(r, 'STUFE'), gs(r, 'POSNR', 10),
            gs(r, 'MATNR', 30), gs(r, 'IDNRK', 30), gs(r, 'OJTXP', 150), gs(r, 'CHI_SPECI', 200),
            gf(r, 'MENGE'), gs(r, 'MEINS', 6), gs(r, 'PAR_UIT', 4), gs(r, 'ETEXT', 30),
            gs(r, 'MMSTA', 6), gs(r, 'MTSTB', 80), gs(r, 'MATTY', 6), (gs(r, 'LOWEST_FLG', 1) or 'N'),
            gs(r, 'ALT_ITEM', 30), gs(r, 'MAIN_MAT', 6), gs(r, 'MATKL', 20), gs(r, 'DATAB', 10), gs(r, 'DATVT', 10)))
        models.add((werks, model))
    wb.close()
    if not recs:
        raise HTTPException(400, "데이터 행 없음")
    cn = _nx(); cur = cn.cursor()
    try:
        for (werks, model) in models:
            if werks:
                cur.execute("DELETE FROM nx.lg_bom WHERE model=? AND werks=?", model, werks)
            else:
                cur.execute("DELETE FROM nx.lg_bom WHERE model=? AND ISNULL(werks,'')=''", model)
        cur.executemany("""INSERT INTO nx.lg_bom
            (cr,werks,model,stufe,posnr,parent_code,child_code,child_desc,child_spec,qty,unit,uit,supply_type,
             mmsta,mtstb,matty,lowest_flg,alt_item,main_mat,matkl,valid_from,valid_to,src_valid,load_dt)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CONVERT(varchar(10),GETDATE(),120),GETDATE())""", recs)
        return {"ok": True, "rows": len(recs), "models": sorted({m for (w, m) in models}),
                "werks": sorted({w for (w, m) in models if w}), "file": file.filename}
    finally:
        cn.close()

@router.post("/api/lgbom/parse")
async def lgbom_parse(file: UploadFile = File(...)):
    """방식① 신규BOM용 — LG BOM Explosion 엑셀 업로드→파싱→**구성 초안 반환(nx 미저장)**.
       엑셀 헤더(실측): WERKS·MODEL·MATNR(부모)·IDNRK(자식)·OJTXP(품명)·CHI_SPECI(규격)·MENGE(수량)·MEINS(단위)·ETEXT(supply_type)·STUFE·POSNR·LOWEST_FLG·MATTY.
       반환: top_item=MODEL, lines=MODEL 직속자식(STUFE=1), 용접봉(RAC*)은 is_weld=1 플래그(구성서 분리 유도). 저장은 프론트에서 /api/bom/save 별도."""
    import io as _io
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb.active
    itr = ws.iter_rows(values_only=True)
    try:
        hdr = next(itr)
    except StopIteration:
        raise HTTPException(400, "빈 파일")
    ix = {str(h or '').strip().upper(): i for i, h in enumerate(hdr)}
    if not all(k in ix for k in ('MODEL', 'MATNR', 'IDNRK')):
        raise HTTPException(400, "LG BOM 형식 아님 — 헤더에 MODEL·MATNR·IDNRK 필요")
    def gv(r, n):
        i = ix.get(n); v = r[i] if (i is not None and i < len(r)) else None
        return None if v in (None, '') else v
    def gs(r, n, ln=None):
        v = gv(r, n); s = '' if v is None else str(v).strip(); return s[:ln] if ln else s
    def gf(r, n):
        try: return float(gv(r, n) or 0)
        except Exception: return 0.0
    def gi(r, n):
        try: return int(float(gv(r, n) or 0))
        except Exception: return 0
    model = ''; werks = ''; rows = []
    for r in itr:
        if not r or not any(x not in (None, '') for x in r): continue
        m = gs(r, 'MODEL')
        if not m: continue
        if not model: model = m; werks = gs(r, 'WERKS', 4)
        rows.append({"stufe": gi(r, 'STUFE'), "parent": gs(r, 'MATNR', 30), "child": gs(r, 'IDNRK', 30),
                     "name": gs(r, 'OJTXP', 150), "spec": gs(r, 'CHI_SPECI', 200), "qty": gf(r, 'MENGE'),
                     "unit": gs(r, 'MEINS', 6), "supply_type": gs(r, 'ETEXT', 30), "lowest": gs(r, 'LOWEST_FLG', 1)})
    wb.close()
    if not model:
        raise HTTPException(400, "데이터 행 없음")
    # 직속자식(STUFE=1 또는 parent==model). 용접봉(RAC*)은 is_weld 분리
    direct = [x for x in rows if x["stufe"] == 1 or x["parent"] == model]
    if not direct:
        direct = rows
    lines, weld = [], []
    seen = set()
    for x in direct:
        ch = x["child"]
        if not ch or ch in seen: continue
        seen.add(ch)
        rec = {"child_item": ch, "item_name": x["name"], "item_spec": x["spec"], "qty": x["qty"],
               "unit": x["unit"], "supply_type": x["supply_type"], "is_weld": 1 if ch.upper().startswith("RAC") else 0}
        (weld if rec["is_weld"] else lines).append(rec)
    return {"ok": True, "top_item": model, "werks": werks, "total_rows": len(rows),
            "lines": lines, "weld_children": weld, "file": file.filename}

@router.post("/api/bom/copy")
def bom_copy(payload: dict = Body(...)):
    """BOM 복사(nx 전용): source의 직접자식 구성을 target으로 복제. source nx.bom_line 우선, 없으면 라이브 CS_M_ITEM_BOM 직접자식.
       target 미등록시 nx.item 최소등록. 라이브는 미변경. 유사공정 협력사 변형용."""
    source = str(payload.get("source", "")).strip()
    target = str(payload.get("target", "")).strip().upper()
    if not source or not target:
        raise HTTPException(400, "source·target 필요")
    if source == target:
        return {"ok": False, "error": "원본과 대상이 같습니다."}
    warn = ""
    cn = _nx(); cur = cn.cursor()
    try:
        # 대상 품목 nx.item 확보(미등록시 최소등록: source 이름 기반)
        cur.execute("SELECT item_name FROM nx.item WHERE item_code=?", target)
        if not cur.fetchone():
            cur.execute("SELECT ISNULL(item_name,'') FROM nx.item WHERE item_code=?", source)
            sr = cur.fetchone(); snm = (sr[0] if sr else "") or source
            cur.execute("INSERT INTO nx.item(item_code,item_name,item_type) VALUES(?,?,N'제품')", target, f"{snm} (복사:{source})")
            warn = f"대상 품번 {target} 신규 등록(품목마스터에서 속성 보완 필요)"
        # source 직접자식: nx.bom_line 우선
        cur.execute("""SELECT l.child_item,l.qty,l.node_type,l.cs_calc_except,l.lme_except,l.sagub_default,l.is_optional,
                 l.from_ymd,l.to_ymd,l.except_flag,l.set_except,l.kitting,l.vir_item,l.proc_gubun,l.gagong_proc,l.s_work,
                 l.wh_gagong,l.in_gagong,l.cust_code,l.remarks
              FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id WHERE h.item_code=? ORDER BY l.seq""", source)
        rows = cur.fetchall()
        src = "nx"
        if not rows:
            # 라이브 CS_M_ITEM_BOM 직접자식(현행 유효일자)로 대체
            lc = _conn(); lcur = lc.cursor()
            try:
                lcur.execute("""SELECT LTRIM(RTRIM(MAT_CODE)), CAST(USE_QTY AS decimal(18,6)), ISNULL(CS_CALC_EXCEPT_FLAG,''),
                       ISNULL(LME_EXCEPT_FLAG,''), ISNULL(SAGUB_FLAG,'0'), ISNULL(SET_EXCEPT_FLAG,''), ISNULL(KITTING_FLAG,''),
                       ISNULL(VIR_ITEM_FLAG,''), ISNULL(GAGONG_PROC_CODE,''), ISNULL(S_WORK_CODE,''), ISNULL(BOM_SEQ,0)
                    FROM PARTNER_ERP_TEST3.nx.CS_M_ITEM_BOM WHERE ITEM_CODE=? AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101' ORDER BY BOM_SEQ""", source)
                lrows = lcur.fetchall()
            finally:
                lc.close()
            rows = [(r[0], r[1], "부품", r[2], r[3], r[4], 0, None, None, "", r[5], r[6], r[7], None, r[8], r[9], None, None, None, None) for r in lrows]
            src = "live"
        if not rows:
            return {"ok": False, "error": f"원본 {source} 의 BOM 구성을 찾을 수 없습니다."}
        # target 헤더 확보 + 전체교체
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", target)
        h = cur.fetchone()
        if h:
            bom_id = h[0]
        else:
            cur.execute("INSERT INTO nx.bom_header(item_code,version,apply_from,status) VALUES(?,1,'2000-01-01',N'확정')", target)
            cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", target); bom_id = cur.fetchone()[0]
        cur.execute("DELETE FROM nx.bom_line WHERE bom_id=?", bom_id)
        seq = 0; nweld = 0
        for r in rows:
            ch = str(r[0]).strip()
            if ch.upper().startswith("RAC"):   # 용접봉 → proc_weld(공정종속)
                cur.execute("""INSERT INTO nx.proc_weld(parent_item,weld_item,weld_base,use_qty,cs_calc_except,lme_except,from_ymd,to_ymd,tag,src)
                    VALUES(?,?,?,?,?,?,?,?,'W','bom_copy')""", target, ch, ch.split('-')[0], float(r[1] or 0), _b(r[3]), _b(r[4]), r[7], r[8])
                nweld += 1; continue
            seq += 1
            cur.execute("""INSERT INTO nx.bom_line
                (bom_id,seq,child_item,qty,node_type,cs_calc_except,lme_except,sagub_default,is_optional,
                 from_ymd,to_ymd,except_flag,set_except,kitting,vir_item,proc_gubun,gagong_proc,s_work,wh_gagong,in_gagong,cust_code,remarks)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                bom_id, seq, ch, float(r[1] or 0), r[2], _b(r[3]), _b(r[4]), _b(r[5]), _b(r[6]),
                r[7], r[8], _b(r[9]), _b(r[10]), _b(r[11]), _b(r[12]), r[13], r[14], r[15], r[16], r[17], r[18], r[19])
        # 원본이 nx이면 source의 proc_weld(용접봉)도 target으로 복사
        cur.execute("IF OBJECT_ID('nx.proc_weld','U') IS NOT NULL DELETE FROM nx.proc_weld WHERE parent_item=? AND src<>'bom_copy'", target)
        cur.execute("""IF OBJECT_ID('nx.proc_weld','U') IS NOT NULL
            INSERT INTO nx.proc_weld(parent_item,weld_item,weld_base,pipe_diam,weld_st,unit_qty,use_qty,cs_calc_except,lme_except,from_ymd,to_ymd,tag,src)
            SELECT ?,weld_item,weld_base,pipe_diam,weld_st,unit_qty,use_qty,cs_calc_except,lme_except,from_ymd,to_ymd,'W','bom_copy' FROM nx.proc_weld WHERE parent_item=?""", target, source)
        # ★item_weld(용접 관경 detail) 복사 — 팝업 프리로드/재계산 정합
        cur.execute("IF OBJECT_ID('nx.item_weld','U') IS NOT NULL DELETE FROM nx.item_weld WHERE item_code=?", target)
        cur.execute("""IF OBJECT_ID('nx.item_weld','U') IS NOT NULL
            INSERT INTO nx.item_weld(item_code,weld_item,pipe_diam,weld_qty,use_qty)
            SELECT ?,weld_item,pipe_diam,weld_qty,use_qty FROM nx.item_weld WHERE item_code=?""", target, source)
        # ★routing(공정=가공비) 복사 — 미복사 시 가공비 누락. ★제품 조립공정은 p_item=제품·item_code=RAC(용접봉carrier) 구조 →
        #   item_code=source(제품자체 공정) OR p_item=source(용접/은납/체결 carrier공정) 둘 다 복사. source→target 치환(carrier코드는 유지).
        cur.execute("IF OBJECT_ID('nx.routing','U') IS NOT NULL DELETE FROM nx.routing WHERE item_code=? OR p_item=?", target, target)
        cur.execute("""IF OBJECT_ID('nx.routing','U') IS NOT NULL
            INSERT INTO nx.routing(p_item,item_code,proc_code,work_qty,prod_uph,calc_gubun,sort_seq)
            SELECT CASE WHEN p_item=? THEN ? ELSE p_item END, CASE WHEN item_code=? THEN ? ELSE item_code END,
                   proc_code,work_qty,prod_uph,calc_gubun,sort_seq
            FROM nx.routing WHERE item_code=? OR p_item=?""", source, target, source, target, source, source)
        _reset_cost_engine()   # ★신규 품번 → 엔진 캐시(_hasbom 등) 무효화(미호출 시 재료=0 발생)
        return {"ok": True, "count": seq, "weld": nweld, "source_from": src, "warn": warn}
    finally:
        cn.close()


# ===================== 공정(routing) = 개발 정본(품목별 공정관리, nx.routing=CS_T_ITEM_PROC) CRUD + 복사 =====================
#  개발이 공정을 지정. 없으면 내부공정 복사 → 수정 → 등록. 협력사견적은 공정 기준 아님(단가).
@router.get("/api/routing/get")
def routing_get(item: str = Query(...)):
    """품목 공정 목록(개발 정본 nx.routing). 전체 공정마스터(CS_M_PROC<90) + 해당 품목 ST(work_qty) 병합 → 편집용."""
    item = item.strip()
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT proc_code, ISNULL(work_qty,0), ISNULL(prod_uph,0), ISNULL(sort_seq,0), ISNULL(calc_gubun,''), ISNULL(p_item,'') FROM nx.routing WHERE item_code=?", item)
        cur_rows = {str(r[0]).strip(): {"work_qty": float(r[1] or 0), "prod_uph": float(r[2] or 0), "sort_seq": int(r[3] or 0),
                                        "calc_gubun": str(r[4] or '').strip(), "p_item": str(r[5] or '').strip()} for r in cur.fetchall()}
    finally:
        nx.close()
    # 공정마스터(이름·정렬), proc_code<90(관리/운반/이윤 제외)
    cn = _conn(); c2 = cn.cursor()
    try:
        c2.execute("SELECT PROC_CODE, PROC_DESC, ISNULL(SORT_SEQ,0), ISNULL(PROD_UPH,0) FROM PARTNER_ERP_TEST3.nx.CS_M_PROC WHERE ISNULL(TRY_CONVERT(int,PROC_CODE),99)<90 AND ISNULL(USE_FLAG,'1')<>'0' ORDER BY ISNULL(SORT_SEQ,0), PROC_CODE")
        procs = []
        for r in c2.fetchall():
            pc = str(r[0]).strip(); ex = cur_rows.get(pc, {})
            procs.append({"proc_code": pc, "name": str(r[1]).strip(), "work_qty": ex.get("work_qty", 0),
                          "prod_uph": ex.get("prod_uph") or float(r[3] or 0),   # 품목ST 없으면 공정마스터 기본 UPH
                          "calc_gubun": ex.get("calc_gubun", "") or "3", "p_item": ex.get("p_item", "")})  # calc_gubun 기본 3(임율기반)
    finally:
        cn.close()
    return {"item": item, "procs": procs, "has_routing": any(p["work_qty"] > 0 for p in procs)}

@router.post("/api/routing/save")
def routing_save(payload: dict = Body(...)):
    """품목 공정 등록/수정(개발 정본 nx.routing). rows=[{proc_code, work_qty, prod_uph, calc_gubun, p_item}]. work_qty>0만 저장. 전체교체.
       ★calc_gubun(계산구분 3임율/8중량/9적용율) + p_item(공정귀속, 일반=''·은납품=부모) 저장 필수 — 미저장시 엔진 가공비=0."""
    item = str(payload.get("item", "")).strip()
    rows = payload.get("rows", []) or []
    if not item:
        raise HTTPException(400, "item 필요")
    keep = [(str(r.get("proc_code", "")).strip(), float(r.get("work_qty") or 0), float(r.get("prod_uph") or 0),
             (str(r.get("calc_gubun", "")).strip() or "3"), str(r.get("p_item", "")).strip())
            for r in rows if str(r.get("proc_code", "")).strip() and float(r.get("work_qty") or 0) > 0]
    nx = _nx(); cur = nx.cursor()
    try:
        # 가공공정(proc<90)만 교체 — 91/92/93/98/99(일반율·운반·이윤율 overhead) 행은 보존(공정그리드 미노출·율마스터 성격)
        cur.execute("DELETE FROM nx.routing WHERE item_code=? AND ISNULL(TRY_CONVERT(int,proc_code),99)<90", item)
        for seq, (pc, wq, uph, cg, pit) in enumerate(keep, 1):
            cur.execute("INSERT INTO nx.routing(p_item, item_code, proc_code, work_qty, prod_uph, calc_gubun, sort_seq) VALUES(?,?,?,?,?,?,?)", pit, item, pc, wq, uph, cg, seq * 10)
        _reset_cost_engine()   # 공정 변경 → 원가엔진 캐시 무효화(다음 계산 최신 반영)
        return {"ok": True, "count": len(keep)}
    finally:
        nx.close()

@router.post("/api/routing/copy")
def routing_copy(payload: dict = Body(...)):
    """내부공정 복사: source 품목의 공정(nx.routing)을 target으로 복사(전체교체). 개발이 유사품 공정을 빠르게 지정."""
    source = str(payload.get("source", "")).strip()
    target = str(payload.get("target", "")).strip()
    if not source or not target or source == target:
        raise HTTPException(400, "source·target(서로 다른) 필요")
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT proc_code, ISNULL(work_qty,0), ISNULL(prod_uph,0), ISNULL(sort_seq,0) FROM nx.routing WHERE item_code=? AND ISNULL(work_qty,0)>0 ORDER BY sort_seq", source)
        src = cur.fetchall()
        if not src:
            return {"ok": False, "error": f"원본 {source} 에 공정(ST>0)이 없습니다."}
        cur.execute("DELETE FROM nx.routing WHERE item_code=?", target)
        for r in src:
            cur.execute("INSERT INTO nx.routing(p_item, item_code, proc_code, work_qty, prod_uph, sort_seq) VALUES(?,?,?,?,?,?)", target, target, str(r[0]).strip(), float(r[1] or 0), float(r[2] or 0), int(r[3] or 0))
        _reset_cost_engine()   # 공정 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "count": len(src)}
    finally:
        nx.close()
