# -*- coding: utf-8 -*-
"""price 도메인 라우터 — app.py에서 분리. 공유헬퍼는 common.py."""
import os, math, json, base64, time, hashlib, mimetypes
from datetime import datetime, timedelta
from urllib.parse import quote as _urlquote
from fastapi import APIRouter, Query, Body, HTTPException, Response, UploadFile, File, Form
from common import (_conn, _num, _run_sp, _shape, _nx, _nx_tx, _b, _d6, _ym, _ITEM_WORK, _get_cost_engine, _reset_cost_engine, _COST_LOCK, SP_SIL, SP_NAE, NxCostEngine, _HERE, _closed, _validate_alloc, _ensure_modelbom, _pur_src, _custnm_map, _kindmap, _dig4, _cur_ym, _sale_win, _SALE_MAGAM, DOC_STORAGE_PATH, _hashlib, _mimetypes)

from common import _NATURE_ALL
router = APIRouter()

# ============ 기준정보: 단가변동내역(전사 라이브 피드) — 품목단가조회에 통합 ============
_COST_TAG = {"1": "매입", "E": "판매(수출)", "S": "판매(내수)"}
@router.get("/api/price/history")
def price_history(from_ymd: str = Query(""), to_ymd: str = Query(""), item: str = Query(""),
                  tag: str = Query(""), changed: str = Query(""),
                  lgroup: str = Query(""), sgroup: str = Query(""), cust: str = Query("")):
    """전사 단가변동 피드: PR_M_ITEM_COST 적용일 내림차순 + 직전단가 대비 Δ(LAG). 소스=nx(우리 DB)·읽기전용."""
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("H.apply_ymd>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("H.apply_ymd<=?"); p.append(_d6(to_ymd))
        if item.strip(): w.append("H.item LIKE ?"); p.append(f"%{item.strip()}%")
        if tag.strip():  w.append("H.tag=?"); p.append(tag.strip())
        if lgroup.strip(): w.append("i.ITEM_LGROUP=?"); p.append(lgroup.strip())
        if sgroup.strip(): w.append("i.ITEM_SGROUP=?"); p.append(sgroup.strip())
        if cust.strip(): w.append("(H.cust LIKE ? OR c.CUST_DESC LIKE ?)"); p += [f"%{cust.strip()}%"] * 2
        if changed == "1": w.append("H.prev IS NOT NULL AND H.cost<>H.prev")
        cur.execute(f"""WITH H AS (
            SELECT ITEM_CODE item, COST_TAG tag, ISNULL(CUST_CODE,'') cust, ISNULL(MKT,'') mkt,
                   ISNULL(CURRENCY,'') curr, COST_APPLY_YMD apply_ymd, ITEM_COST cost, MAT_COST mat,
                   PROC_COST procc, OTHER_COST oth, PUR_RATE rate, ISNULL(INSERT_USER_ID,'') usr,
                   INSERT_DATETIME idt, ISNULL(REMARKS,'') remarks,
                   LAG(ITEM_COST) OVER (PARTITION BY ITEM_CODE,COST_TAG,ISNULL(CUST_CODE,''),ISNULL(MKT,'')
                                        ORDER BY COST_APPLY_YMD, INSERT_DATETIME) prev
            FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM_COST)
          SELECT TOP 3000 H.item, ISNULL(i.ITEM_DESC,'') nm, H.tag, H.cust, ISNULL(c.CUST_DESC,'') cust_nm,
                 H.mkt, H.curr, H.apply_ymd, H.cost, H.mat, H.procc, H.oth, H.rate, H.prev, H.usr,
                 H.idt, H.remarks
          FROM H LEFT JOIN PARTNER_ERP_TEST3.nx.PR_M_ITEM i ON i.ITEM_CODE=H.item LEFT JOIN PARTNER_ERP_TEST3.nx.CM_M_CUST c ON c.CUST_CODE=H.cust
          WHERE {' AND '.join(w)} ORDER BY H.apply_ymd DESC, H.idt DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            for k in ("cost", "mat", "procc", "oth", "rate", "prev"):
                r[k] = float(r[k]) if r[k] is not None else None
            r["delta"] = round(r["cost"] - r["prev"], 2) if (r["prev"] is not None and r["cost"] is not None) else None
            r["tag_nm"] = _COST_TAG.get(str(r["tag"]).strip(), str(r["tag"]))
            r["idt"] = str(r["idt"] or "")[:19]
        dLG = _kindmap(cur, "PR005"); dSG = _kindmap(cur, "PR006")
        return {"rows": rows, "cnt": len(rows),
                "changed": sum(1 for r in rows if r["delta"] not in (None, 0)),
                "lgroups": [{"code": k, "nm": v} for k, v in sorted(dLG.items())],
                "sgroups": [{"code": k, "nm": v} for k, v in sorted(dSG.items())]}
    finally:
        cn.close()

# ---------- 품목별 단가조회(라이브) — 레거시 w_pr_master_150: 거래처별·적용월 시계열 ----------
_CURR_NM = {"KRW": "원화", "USD": "달러", "RMB": "위안", "EUR": "유로", "JPY": "엔"}
@router.get("/api/price/search")
def price_search(q: str = Query(""), lgroup: str = Query(""), sgroup: str = Query(""), cust: str = Query(""), limit: int = Query(1000)):
    """단가보유 품목 검색(품번/품명 + 대/소분류 + 거래처 필터, AND). 좌측 리스트. 분류·거래처 코드→이름."""
    cn = _conn(); cur = cn.cursor()
    try:
        dLG = _kindmap(cur, "PR005"); dSG = _kindmap(cur, "PR006")
        w = ""; p = []
        if q.strip(): w += " AND (i.ITEM_CODE LIKE ? OR i.ITEM_DESC LIKE ?)"; p += [f"%{q.strip()}%"] * 2
        if lgroup.strip(): w += " AND i.ITEM_LGROUP=?"; p.append(lgroup.strip())
        if sgroup.strip(): w += " AND i.ITEM_SGROUP=?"; p.append(sgroup.strip())
        # 거래처 필터: 해당 거래처(코드=오토컴플리트값 / 명칭 LIKE) 단가가 있는 품목만 (AND)
        cust_cond = "EXISTS(SELECT 1 FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM_COST x WHERE x.ITEM_CODE=i.ITEM_CODE"
        if cust.strip():
            cust_cond += " AND (x.CUST_CODE=? OR EXISTS(SELECT 1 FROM PARTNER_ERP_TEST3.nx.CM_M_CUST c WHERE c.CUST_CODE=x.CUST_CODE AND c.CUST_DESC LIKE ?))"
            p2 = [cust.strip(), f"%{cust.strip()}%"]
        else:
            p2 = []
        cust_cond += ")"
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),1000))} i.ITEM_CODE, ISNULL(i.ITEM_DESC,'') nm, ISNULL(i.ITEM_SPEC,'') spec,
              ISNULL(i.ITEM_LGROUP,'') lg, ISNULL(i.ITEM_SGROUP,'') sg, (SELECT COUNT(*) FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM_COST x WHERE x.ITEM_CODE=i.ITEM_CODE) cnt
            FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM i
            WHERE {cust_cond}{w}
            ORDER BY i.ITEM_CODE""", *(p2 + p))
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            d["lg_nm"] = dLG.get(str(d.get("lg", "")).strip(), str(d.get("lg", "")).strip())
            d["sg_nm"] = dSG.get(str(d.get("sg", "")).strip(), str(d.get("sg", "")).strip())
            rows.append(d)
        cur.execute("""SELECT DISTINCT ISNULL(i.ITEM_LGROUP,'') lg, ISNULL(i.ITEM_SGROUP,'') sg FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM i
            WHERE EXISTS(SELECT 1 FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM_COST x WHERE x.ITEM_CODE=i.ITEM_CODE)""")
        lgs, sgs = set(), set()
        for r in cur.fetchall():
            if str(r[0]).strip(): lgs.add(str(r[0]).strip())
            if str(r[1]).strip(): sgs.add(str(r[1]).strip())
        return {"rows": rows, "cnt": len(rows),
                "lgroups": [{"code": s, "nm": dLG.get(s, s)} for s in sorted(lgs)],
                "sgroups": [{"code": s, "nm": dSG.get(s, s)} for s in sorted(sgs)]}
    finally:
        cn.close()

@router.get("/api/price/item")
def price_item(item: str = Query("")):
    """품번별 단가 이력(거래처별·단가구분·적용월 시계열, 코드→이름). 소급조회의 원장."""
    item = item.strip()
    if not item: return {"item": item, "nm": "", "spec": "", "rows": []}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT ISNULL(ITEM_DESC,''), ISNULL(ITEM_SPEC,'') FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM WHERE ITEM_CODE=?", item)
        r0 = cur.fetchone(); nm, spec = (r0[0], r0[1]) if r0 else ("", "")
        cur.execute("""SELECT h.COST_TAG, ISNULL(h.CUST_CODE,'') cust, ISNULL(c.CUST_DESC,'') cust_nm,
              h.COST_APPLY_YMD, ISNULL(h.CURRENCY,'') curr, ISNULL(h.MAIN_FLAG,'') main_flag, ISNULL(h.MKT,'') mkt,
              h.ITEM_COST, h.MAT_COST, h.PROC_COST, h.OTHER_COST, ISNULL(h.REMARKS,'') remarks
            FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM_COST h LEFT JOIN PARTNER_ERP_TEST3.nx.CM_M_CUST c ON c.CUST_CODE=h.CUST_CODE
            WHERE h.ITEM_CODE=? ORDER BY h.COST_APPLY_YMD DESC, h.COST_TAG, h.CUST_CODE""", item)
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            for k in ("ITEM_COST", "MAT_COST", "PROC_COST", "OTHER_COST"):
                d[k.lower()] = float(d.pop(k)) if d.get(k) is not None else 0.0
            d["tag_nm"] = _COST_TAG.get(str(d["COST_TAG"]).strip(), str(d["COST_TAG"]))
            d["curr_nm"] = _CURR_NM.get(str(d["curr"]).strip(), str(d["curr"]).strip())
            d["main"] = 1 if str(d["main_flag"]).strip() == "1" else 0
            d["apply_ymd"] = str(d.pop("COST_APPLY_YMD") or "")
            d["tag"] = str(d.pop("COST_TAG")).strip()
            rows.append(d)
        # ★업로드된 사급가(nx.price_item vendor='LG', COSP)도 함께 표시 — 크로스DB
        try:
            cur.execute("""SELECT apply_ymd, price, ISNULL(currency,'KRW') FROM PARTNER_ERP_TEST3.nx.price_item
                WHERE item_code=? AND price_type='매입' AND vendor_code='LG' ORDER BY apply_ymd DESC""", item)
            for u in cur.fetchall():
                cc = str(u[2] or "KRW").strip()
                rows.append({"tag": "매입", "tag_nm": "사급가(업로드)", "cust": "LG", "cust_nm": "LG(COSP 업로드)",
                             "apply_ymd": str(u[0] or ""), "curr": cc, "curr_nm": _CURR_NM.get(cc, cc),
                             "main_flag": "", "main": 0, "mkt": "",
                             "item_cost": float(u[1]) if u[1] is not None else 0.0, "mat_cost": 0.0,
                             "proc_cost": 0.0, "other_cost": 0.0, "remarks": "COSP 업로드"})
        except Exception:
            pass
        rows.sort(key=lambda d: (d.get("apply_ymd") or ""), reverse=True)   # 최신일자 우선(업로드분 병합)
        return {"item": item, "nm": nm, "spec": spec, "rows": rows, "cnt": len(rows)}
    finally:
        cn.close()


# ---------- 품목조회(라이브 PR_M_ITEM 전 컬럼, 코드→이름) — 레거시 w_pr_master_010 ----------
_ITEM_MAKE = {"": "", "1": "자체생산", "2": "외주가공", "3": "매입", "4": "사급가공", "5": "외주완성"}
_ITEM_WORK = {"": "", "P1": "용접", "P2": "가공", "D1": "직납"}
def _kindmap(cur, kind):
    cur.execute("SELECT DETAIL_CODE, DETAIL_DESC FROM PARTNER_ERP_TEST3.nx.CM_M_MASTER_DETAIL WHERE KIND_CODE=?", kind)
    return {str(r[0]).strip(): str(r[1] or "").strip() for r in cur.fetchall()}

_MAT_SGROUP = ('210', '220', '230', '310', '910', '991', '992', '993')  # 자재(원소재/원자재/부자재/사급/잡자재/소모품)
@router.get("/api/item/list")
def item_list(q: str = Query(""), lgroup: str = Query(""), sgroup: str = Query(""), mat: str = Query(""),
              nature: str = Query(""), limit: int = Query(500)):
    """품목조회(라이브). 대/소분류·품목형태·단위·재질·협력사·작업처·제작유형 코드→이름. mat=1→자재만. nature=성격6그룹(nx.item 조인)."""
    cn = _conn(); cur = cn.cursor()
    try:
        dLG = _kindmap(cur, "PR005"); dSG = _kindmap(cur, "PR006"); dPK = _kindmap(cur, "PR021")
        dUN = _kindmap(cur, "CM002"); dMT = _kindmap(cur, "PR019")
        cur.execute("SELECT CUST_CODE, ISNULL(CUST_DESC,'') FROM PARTNER_ERP_TEST3.nx.CM_M_CUST")
        dCust = {str(r[0]).strip(): r[1] for r in cur.fetchall()}
        w = ["1=1"]; p = []
        if q.strip(): w.append("(i.ITEM_CODE LIKE ? OR i.ITEM_DESC LIKE ?)"); p += [f"%{q.strip()}%"] * 2
        if lgroup.strip(): w.append("i.ITEM_LGROUP=?"); p.append(lgroup.strip())
        if sgroup.strip(): w.append("i.ITEM_SGROUP=?"); p.append(sgroup.strip())
        if mat == "1": w.append(f"i.ITEM_SGROUP IN ({','.join('?'*len(_MAT_SGROUP))})"); p += list(_MAT_SGROUP)
        if nature.strip(): w.append("nx.nature=?"); p.append(nature.strip())
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),3000))} i.ITEM_CODE, ISNULL(i.ITEM_DESC,'') nm, ISNULL(i.ITEM_SPEC,'') spec,
              ISNULL(i.ITEM_LGROUP,'') lg, ISNULL(i.ITEM_SGROUP,'') sg, ISNULL(i.PIPE_KIND,'') pk, ISNULL(i.UNIT,'') un,
              i.ITEM_DIAM, i.ITEM_THICK, i.ITEM_LENGTH, i.ITEM_WEIGHT, ISNULL(i.METAL_GUBUN,'') metal, ISNULL(i.IN_CUST_CODE,'') incust,
              ISNULL(i.WORK_CODE,'') work, ISNULL(i.MAKE_TYPE,'') mk, ISNULL(i.COST_GUBUN,'') cg, ISNULL(i.ITEM_STATUS,'') status,
              ISNULL(i.SAFE_STOCK_MIN,0), ISNULL(i.SAFE_STOCK_MAX,0), ISNULL(i.KITTING_MIN,0),
              ISNULL(i.WELD_POINT_IN,0), ISNULL(i.WELD_POINT_OUT,0), ISNULL(i.TARIFF_RATE,0), ISNULL(i.REMARKS,'') remarks, ISNULL(i.ITEM_COST,0),
              ISNULL(nx.nature,'') nature, nx.active
            FROM PARTNER_ERP_TEST3.nx.PR_M_ITEM i LEFT JOIN PARTNER_ERP_TEST3.nx.item nx ON nx.item_code = i.ITEM_CODE COLLATE DATABASE_DEFAULT
            WHERE {' AND '.join(w)} ORDER BY i.ITEM_CODE""", *p)
        rows = []
        for r in cur.fetchall():
            g = lambda i: str(r[i] or "").strip()
            num = lambda i: (float(r[i]) if r[i] is not None else 0)
            rows.append({
                "item_code": g(0), "nm": g(1), "spec": g(2),
                "lgroup": dLG.get(g(3), g(3)), "sgroup": dSG.get(g(4), g(4)), "pipe_kind": dPK.get(g(5), g(5)),
                "unit": dUN.get(g(6), g(6)), "diam": num(7), "thick": num(8), "length": num(9), "weight": num(10),
                "metal": dMT.get(g(11), g(11)), "in_cust": dCust.get(g(12), g(12)), "work": _ITEM_WORK.get(g(13), g(13)),
                "make_type": _ITEM_MAKE.get(g(14), g(14)), "cost_gubun": g(15), "status": ("사용" if g(16) in ("", "1", "사용") else g(16)),
                "safe_min": num(17), "safe_max": num(18), "kitting_min": num(19),
                "weld_in": num(20), "weld_out": num(21), "tariff": num(22), "remarks": g(23), "item_cost": num(24),
                "nature": g(25), "active": (0 if (r[26] is not None and not r[26]) else 1)})
        # 분류 드롭다운
        return {"rows": rows, "cnt": len(rows),
                "lgroups": [{"code": k, "nm": v} for k, v in sorted(dLG.items())],
                "sgroups": [{"code": k, "nm": v} for k, v in sorted(dSG.items())],
                "natures": [{"code": n, "nm": n} for n in _NATURE_ALL]}
    finally:
        cn.close()


# ============ 사급가(COSP Sales Price) 업로드 — LG 사급 부품가를 nx.price_item에 Start Date 반영 upsert ============
@router.post("/api/price/sagub_upload")
async def price_sagub_upload(file: UploadFile = File(...)):
    """COSP Sales Price 엑셀(LG 사급 부품가) 업로드. 각 행의 Start Date를 적용일(apply_ymd)로 nx.price_item에 upsert.
       price_type='매입', vendor_code='LG'. 최신가는 as-of(적용일 최신)로 자동 반영. nx.item 미등록 품번은 스킵."""
    import io as _io
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb.active; itr = ws.iter_rows(values_only=True)
    try:
        hdr = next(itr)
    except StopIteration:
        raise HTTPException(400, "빈 파일")
    ix = {str(h or '').strip().lower(): i for i, h in enumerate(hdr)}
    if not all(k in ix for k in ('material', 'sales price', 'start date')):
        raise HTTPException(400, "COSP 형식 아님 — 헤더에 Material·Sales Price·Start Date 필요")
    def gv(r, n):
        i = ix.get(n); v = r[i] if (i is not None and i < len(r)) else None
        return None if v in (None, '') else v
    def _ymd6(dt):
        if hasattr(dt, 'year'): return f"{dt.year % 100:02d}{dt.month:02d}{dt.day:02d}"
        s = ''.join(ch for ch in str(dt) if ch.isdigit())
        return s[2:8] if len(s) >= 8 else (s[-6:] if len(s) >= 6 else '')
    seen = {}
    for r in itr:
        if not r or not any(x not in (None, '') for x in r): continue
        mat = str(gv(r, 'material') or '').strip()
        if not mat: continue
        try: price = float(gv(r, 'sales price') or 0)
        except Exception: continue
        ay = _ymd6(gv(r, 'start date'))
        if not ay or price <= 0: continue
        curc = (str(gv(r, 'curr') or 'KRW').strip()[:3] or 'KRW')
        seen[(mat, ay)] = (price, curc)          # (품번,적용일)별 dedup(엑셀 중복행)
    wb.close()
    if not seen:
        raise HTTPException(400, "데이터 행 없음")
    nx = _nx(); cur = nx.cursor()
    try:
        up = 0; skip = 0; items = set()
        for (mat, ay), (price, curc) in seen.items():
            try:
                cur.execute("""MERGE nx.price_item AS t
                  USING (SELECT ? item_code,'매입' price_type,'LG' vendor_code,? apply_ymd) s
                  ON t.item_code=s.item_code AND t.price_type=s.price_type AND t.vendor_code=s.vendor_code AND t.apply_ymd=s.apply_ymd
                  WHEN MATCHED THEN UPDATE SET price=?, currency=?
                  WHEN NOT MATCHED THEN INSERT(item_code,price_type,vendor_code,currency,apply_ymd,price)
                    VALUES(?,'매입','LG',?,?,?);""",
                    mat, ay, price, curc, mat, curc, ay, price)
                up += 1; items.add(mat)
            except Exception:
                skip += 1                        # nx.item 미등록(FK) 등 스킵
        return {"ok": True, "rows": up, "items": len(items), "skipped": skip, "file": file.filename}
    finally:
        nx.close()


@router.get("/api/price/sagub_list")
def price_sagub_list(q: str = Query("")):
    """업로드된 사급가(nx.price_item vendor='LG') 품번별 최신값 목록."""
    nx = _nx(); cur = nx.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT p.item_code, ISNULL(i.item_name,''), p.apply_ymd, p.price, p.currency
            FROM nx.price_item p LEFT JOIN nx.item i ON i.item_code=p.item_code
            WHERE p.price_type='매입' AND p.vendor_code='LG'
              AND (?='' OR p.item_code LIKE ? OR i.item_name LIKE ?)
              AND p.apply_ymd=(SELECT MAX(x.apply_ymd) FROM nx.price_item x
                    WHERE x.item_code=p.item_code AND x.price_type='매입' AND x.vendor_code='LG')
            ORDER BY p.item_code""", q.strip(), like, like)
        rows = [{"item": r[0], "name": r[1], "apply_ymd": r[2], "price": float(r[3]), "currency": r[4]} for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()


# ============ LG 판가(PO Price) 업로드 — nx.price_item vendor 1010(SAC)/1020(RAC), TAGE(수출)/TAGS(내수) upsert ============
_LGP_VEND = {'SAC': '1010', 'RAC': '1020'}   # 사업부→LG 사업장 vendor (실측 검증: SAC∩1010=1876, RAC∩1020=687)
@router.post("/api/price/lgprice_upload")
async def price_lgprice_upload(file: UploadFile = File(...), biz: str = Query("")):
    """LG PO Price 엑셀(LG 판가) 업로드. biz=SAC(vendor 1010)/RAC(1020). 헤더=Material·MKT·Unit Price·Start Date·Curr.
       MKT 1→TAGS(내수)·2→TAGE(수출). Start Date=적용일(apply_ymd). nx.price_item upsert(item·type·vendor·apply_ymd).
       원가/손익 엔진의 LG판가 소스(vendor 1010/1020, TAGE/TAGS as-of)로 바로 사용됨. nx.item 미등록 품번은 스킵."""
    biz = biz.strip().upper(); vendor = _LGP_VEND.get(biz)
    if not vendor:
        raise HTTPException(400, "사업부(biz)=SAC 또는 RAC 필요")
    import io as _io
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb.active; itr = ws.iter_rows(values_only=True)
    try:
        hdr = next(itr)
    except StopIteration:
        raise HTTPException(400, "빈 파일")
    ix = {str(h or '').strip().lower().replace('\n', ' '): i for i, h in enumerate(hdr)}
    def col(*names):
        for n in names:
            if n in ix: return ix[n]
        return None
    ci_mat = col('material'); ci_price = col('unit price', 'price'); ci_sd = col('start date'); ci_mkt = col('mkt'); ci_cur = col('curr', 'currency')
    if ci_mat is None or ci_price is None or ci_sd is None:
        raise HTTPException(400, "PO Price 형식 아님 — 헤더에 Material·Unit Price·Start Date 필요")
    def _ymd6(dt):
        if hasattr(dt, 'year'): return f"{dt.year % 100:02d}{dt.month:02d}{dt.day:02d}"
        s = ''.join(ch for ch in str(dt) if ch.isdigit()); return s[2:8] if len(s) >= 8 else (s[-6:] if len(s) >= 6 else '')
    def gv(r, i):
        return r[i] if (i is not None and i < len(r)) else None
    seen = {}
    for r in itr:
        if not r: continue
        mat = str(gv(r, ci_mat) or '').strip()
        if not mat: continue
        try:
            price = float(gv(r, ci_price) or 0)
        except Exception:
            continue
        ay = _ymd6(gv(r, ci_sd))
        if not ay or price <= 0: continue
        mkt = str(gv(r, ci_mkt) or '').strip()
        ptype = 'TAGS' if mkt in ('1', '1.0') else ('TAGE' if mkt in ('2', '2.0') else 'TAGS')
        cv = gv(r, ci_cur); curc = (str(cv).strip()[:3] if cv else 'KRW') or 'KRW'
        seen[(mat.upper(), ptype, ay)] = (price, curc)   # (품번,구분,적용일)별 dedup
    wb.close()
    if not seen:
        raise HTTPException(400, "데이터 행 없음")
    nx = _nx(); cur = nx.cursor()
    try:
        up = 0; skip = 0; items = set()
        for (mat, ptype, ay), (price, curc) in seen.items():
            try:
                cur.execute("""MERGE nx.price_item AS t
                  USING (SELECT ? item_code,? price_type,? vendor_code,? apply_ymd) s
                  ON t.item_code=s.item_code AND t.price_type=s.price_type AND t.vendor_code=s.vendor_code AND t.apply_ymd=s.apply_ymd
                  WHEN MATCHED THEN UPDATE SET price=?, currency=?
                  WHEN NOT MATCHED THEN INSERT(item_code,price_type,vendor_code,currency,apply_ymd,price)
                    VALUES(?,?,?,?,?,?);""",
                    mat, ptype, vendor, ay, price, curc, mat, ptype, vendor, curc, ay, price)
                up += 1; items.add(mat)
            except Exception:
                skip += 1
        return {"ok": True, "rows": up, "items": len(items), "skipped": skip, "biz": biz, "vendor": vendor, "file": file.filename}
    finally:
        nx.close()

@router.get("/api/price/lgprice_list")
def price_lgprice_list(q: str = Query(""), biz: str = Query("")):
    """업로드된 LG판가(price_item vendor 1010/1020, TAGE/TAGS) 품번×구분별 최신값. biz=SAC/RAC 필터(빈값=전체)."""
    b = biz.strip().upper()
    vendors = [_LGP_VEND[b]] if b in _LGP_VEND else ['1010', '1020']
    vin = ",".join("'" + v + "'" for v in vendors)
    nx = _nx(); cur = nx.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute(f"""SELECT p.item_code, ISNULL(i.item_name,''), p.vendor_code, p.price_type, p.apply_ymd, p.price, p.currency
            FROM nx.price_item p LEFT JOIN nx.item i ON i.item_code=p.item_code
            WHERE p.vendor_code IN ({vin}) AND p.price_type IN ('TAGE','TAGS')
              AND (?='' OR p.item_code LIKE ? OR i.item_name LIKE ?)
              AND p.apply_ymd=(SELECT MAX(x.apply_ymd) FROM nx.price_item x
                    WHERE x.item_code=p.item_code AND x.vendor_code=p.vendor_code AND x.price_type=p.price_type)
            ORDER BY p.item_code, p.vendor_code, p.price_type""", q.strip(), like, like)
        rows = [{"item": r[0], "name": r[1], "vendor": r[2], "type": r[3], "apply_ymd": r[4], "price": float(r[5]), "currency": r[6]} for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()


# ============ 특이 단가목록: 실 입고가 > 실 유상사급 출고가(판가 역전) ============
# 해당월에 실제 입고(자재입고명세서)되고 유상사급 출고(자재불출명세서)된 품목 중, 실 입고가 > 실 출고가(비싸게 사서 싸게 사급) = 손해.
# 입고=PU_T_STOCK_MAINT(9,S,C,G,H) · 출고=PU_T_STOCK_MAINT(tag'5', 출고가>0=유상). 원소재·용접봉·소모품 제외(용접링 유지). 상위 Assy=BOM 역전개.
_INV_EXCL_SG = ('210', '220', '910', '991', '992', '993')  # 원소재/원자재/잡자재(용접봉·나이프)/생산소모품/일반소모품/수불예외
@router.get("/api/price/inversion")
def price_inversion(ym: str = Query(""), q: str = Query(""), limit: int = Query(3000)):
    """특이 단가목록: 해당월(ym=YYMM, 미지정=당월) 실 입고가 > 실 유상사급 출고가 역전 부품. 입고/출고 둘 다 있는 품목만.
       입고가=PU_T_STOCK_MAINT(9,S,C,G,H) 가중평균 · 출고가=tag'5' 가중평균(>0). 원소재·용접봉·소모품 제외. 상위 Assy(BOM역전개)."""
    nx = _nx(); cur = nx.cursor()
    try:
        ymv = "".join(ch for ch in (ym or "") if ch.isdigit())[:4]
        if len(ymv) != 4:
            cur.execute("SELECT FORMAT(GETDATE(),'yyMM')"); ymv = str(cur.fetchone()[0])
        cur.execute("SELECT CUST_CODE, ISNULL(CUST_DESC,'') FROM nx.CM_M_CUST")
        dCust = {str(r[0]).strip(): r[1] for r in cur.fetchall()}
        dSG = _kindmap(cur, "PR006")
        lim = max(1, min(int(limit), 8000))
        exsg = ",".join("?" * len(_INV_EXCL_SG))
        w = []; p = []
        if q.strip(): w.append("(inb.mat LIKE ? OR m.ITEM_DESC LIKE ?)"); p += [f"%{q.strip()}%"] * 2
        extra = (" AND " + " AND ".join(w)) if w else ""
        cur.execute(f"""
        WITH inb AS (
          SELECT MAT_CODE mat, SUM(CAST(MAINT_QTY AS FLOAT)) q, SUM(CAST(MAINT_AMT AS FLOAT)) amt
          FROM nx.PU_T_STOCK_MAINT WHERE LEFT(MAINT_YMD,4)=? AND MAINT_TAG IN ('9','S','C','G','H') AND MAINT_QTY>0 GROUP BY MAT_CODE),
        inc AS (SELECT mat, CUST_CODE FROM (
            SELECT MAT_CODE mat, CUST_CODE, ROW_NUMBER() OVER(PARTITION BY MAT_CODE ORDER BY SUM(CAST(MAINT_QTY AS FLOAT)) DESC) rn
            FROM nx.PU_T_STOCK_MAINT WHERE LEFT(MAINT_YMD,4)=? AND MAINT_TAG IN ('9','S','C','G','H') AND MAINT_QTY>0 GROUP BY MAT_CODE, CUST_CODE) x WHERE rn=1),
        outb AS (
          SELECT MAT_CODE mat, SUM(CAST(MAINT_QTY AS FLOAT)) q, SUM(CAST(MAINT_AMT AS FLOAT)) amt
          FROM nx.PU_T_STOCK_MAINT WHERE LEFT(MAINT_YMD,4)=? AND MAINT_TAG='5' GROUP BY MAT_CODE),
        outc AS (SELECT mat, CUST_CODE FROM (
            SELECT MAT_CODE mat, CUST_CODE, ROW_NUMBER() OVER(PARTITION BY MAT_CODE ORDER BY SUM(CAST(MAINT_QTY AS FLOAT)) DESC) rn
            FROM nx.PU_T_STOCK_MAINT WHERE LEFT(MAINT_YMD,4)=? AND MAINT_TAG='5' GROUP BY MAT_CODE, CUST_CODE) x WHERE rn=1)
        SELECT TOP {lim} inb.mat item, ISNULL(m.ITEM_DESC,'') nm, ISNULL(m.ITEM_SGROUP,'') sg,
          ic.CUST_CODE pur_cust, CAST(inb.amt/NULLIF(inb.q,0) AS DECIMAL(18,2)) pur, CAST(inb.q AS DECIMAL(18,2)) inq,
          oc.CUST_CODE sag_cust, CAST(outb.amt/NULLIF(outb.q,0) AS DECIMAL(18,2)) sag, CAST(-outb.q AS DECIMAL(18,2)) outq,
          CAST((inb.amt/NULLIF(inb.q,0)) - (outb.amt/NULLIF(outb.q,0)) AS DECIMAL(18,2)) diff
        FROM inb JOIN outb ON inb.mat=outb.mat
          JOIN nx.PR_M_ITEM m ON m.ITEM_CODE=inb.mat
          LEFT JOIN inc ic ON ic.mat=inb.mat LEFT JOIN outc oc ON oc.mat=outb.mat
        WHERE (outb.amt/NULLIF(outb.q,0)) > 0
          AND (inb.amt/NULLIF(inb.q,0)) > (outb.amt/NULLIF(outb.q,0))
          AND ( m.ITEM_SGROUP NOT IN ({exsg}) OR m.ITEM_DESC LIKE N'%용접링%' )
          AND ( m.ITEM_CODE NOT LIKE 'RAC%' OR m.ITEM_DESC LIKE N'%용접링%' )
          {extra}
        ORDER BY ((inb.amt/NULLIF(inb.q,0)) - (outb.amt/NULLIF(outb.q,0))) DESC""",
        ymv, ymv, ymv, ymv, *(list(_INV_EXCL_SG) + p))
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        # 상위 Assy(완제품 조상) — reverse-BOM 인메모리 워크(1쿼리). nx 출고 ITEM_CODE(ASY) 미적재라 BOM으로 파생.
        cur.execute("SELECT bl.child_item, h.item_code FROM nx.bom_line bl JOIN nx.bom_header h ON bl.bom_id=h.bom_id")
        par = {}
        for ch, pa in cur.fetchall():
            par.setdefault(str(ch).strip(), []).append(str(pa).strip())
        def tops(it):
            seen = {it}; out = set(); stk = list(par.get(it, []))
            while stk:
                nn = stk.pop()
                if nn in seen: continue
                seen.add(nn)
                ps = par.get(nn)
                if ps: stk.extend(ps)
                else: out.add(nn)   # 부모 없음 = 최상위 완제품
            return sorted(out)
        for r in rows:
            for k in ("pur", "sag", "diff", "inq", "outq"):
                r[k] = float(r[k]) if r[k] is not None else 0.0
            r["pur_cust_nm"] = dCust.get(str(r["pur_cust"] or "").strip(), str(r["pur_cust"] or "").strip())
            r["sag_cust_nm"] = dCust.get(str(r["sag_cust"] or "").strip(), str(r["sag_cust"] or "").strip())
            r["sg_nm"] = dSG.get(str(r["sg"]).strip(), str(r["sg"]).strip())
            r["assy"] = tops(str(r["item"]).strip())[:10]
        return {"rows": rows, "cnt": len(rows), "ym": ymv}
    finally:
        nx.close()
