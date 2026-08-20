# -*- coding: utf-8 -*-
"""rawmat 도메인 라우터 — app.py에서 분리. 공유헬퍼는 common.py."""
import os, math, json, base64, time, hashlib, mimetypes
from datetime import datetime, timedelta
from urllib.parse import quote as _urlquote
from fastapi import APIRouter, Query, Body, HTTPException, Response, UploadFile, File, Form
from common import (_conn, _num, _run_sp, _shape, _nx, _nx_tx, _b, _d6, _ym, _ITEM_WORK, _get_cost_engine, _reset_cost_engine, _COST_LOCK, SP_SIL, SP_NAE, NxCostEngine, _HERE, _closed, _validate_alloc, _ensure_modelbom, _pur_src, _custnm_map, _kindmap, _dig4, _cur_ym, _sale_win, _SALE_MAGAM, DOC_STORAGE_PATH, _hashlib, _mimetypes)

router = APIRouter()

# ============ 원소재 마스터 + 5가격 통합 뷰 (nx.raw_material ↔ price_metal/price_item/lg_lme_costtable) ============
_MATCODE = {"구리": "CU", "고강도관": "고강도", "알루미늄": "AL", "STS": "STS"}
# LG LME 인정가 = lg_lme_costtable 재료비(재질 대표구분). CU←L/W·고강도←고강도관. (구 price_lme_lg 빈테이블 통합제거)
_LG_GUBUN = {"CU": "L/W", "고강도": "고강도관"}
@router.get("/api/rawmat/list")
def rawmat_list(q: str = Query(""), material: str = Query("")):
    """원소재 마스터(외경×두께×재질×조질) + 5가격: 시세·파트너(사급)·매입·매출·LG인증. 규격/재질/품번 조인."""
    nx = _nx(); cur = nx.cursor()
    try:
        # 원소재 마스터
        w = ["1=1"]; p = []
        if material.strip(): w.append("material=?"); p.append(material.strip())
        if q.strip(): w.append("(CAST(outer_diam AS varchar)+'x'+CAST(thickness AS varchar) LIKE ? OR part_no LIKE ? OR src_codes LIKE ?)"); p += [f"%{q.strip()}%"] * 3
        cur.execute(f"""SELECT raw_id, material, outer_diam, thickness, temper, part_no, unit, std_length,
              y2026_kg, src_codes, vendors FROM nx.raw_material WHERE {' AND '.join(w)}
              ORDER BY material, outer_diam, thickness""", *p)
        rms = cur.fetchall()
        # price_metal 규격별 최신 시세/파트너
        cur.execute("""SELECT metal_gubun, diam, thick, std_price, partner_price FROM nx.price_metal pm
              WHERE apply_ym = (SELECT MAX(apply_ym) FROM nx.price_metal x WHERE x.metal_gubun=pm.metal_gubun AND x.diam=pm.diam AND x.thick=pm.thick)""")
        pmet = {}
        for r in cur.fetchall():
            pmet[(str(r[0]).strip(), float(r[1]) if r[1] is not None else None, float(r[2]) if r[2] is not None else None)] = (r[3], r[4])
        # LG 인증가 = lg_lme_costtable 재료비(최신월, 재질 대표구분 L/W·고강도관) — 단일정본, 매월 자동반영
        cur.execute("""SELECT gubun, MAX(jaeryo) FROM nx.lg_lme_costtable
              WHERE apply_ym=(SELECT MAX(apply_ym) FROM nx.lg_lme_costtable) AND gubun IN ('L/W','고강도관') GROUP BY gubun""")
        _g2m = {"L/W": "CU", "고강도관": "고강도"}
        plme = {_g2m[str(r[0]).strip()]: r[1] for r in cur.fetchall() if str(r[0]).strip() in _g2m}
        # 품번별 최신 매입/사급/매출 (price_item)
        cur.execute("""SELECT item_code, price_type, price FROM nx.price_item pi
              WHERE apply_ymd=(SELECT MAX(apply_ymd) FROM nx.price_item x WHERE x.item_code=pi.item_code AND x.price_type=pi.price_type)""")
        pit = {}
        for r in cur.fetchall():
            pit.setdefault(str(r[0]).strip(), {})[str(r[1]).strip()] = r[2]
        rows = []
        for r in rms:
            mat = str(r[1] or "").strip(); od = float(r[2]) if r[2] is not None else None; th = float(r[3]) if r[3] is not None else None
            sise, partner = pmet.get((_MATCODE.get(mat, mat), od, th), (None, None))
            # 매입/사급/매출 = 원천품번들 중 대표(최대) 값
            codes = [c.strip() for c in str(r[9] or "").split(",") if c.strip()]
            buy = sale = sagub = None
            for c in codes:
                d = pit.get(c, {})
                for k, dst in (("매입", "buy"), ("TAGS", "sagub"), ("TAGE", "sale")):
                    v = d.get(k)
                    if v and float(v) > 0:
                        if dst == "buy": buy = max(buy or 0, float(v))
                        elif dst == "sagub": sagub = max(sagub or 0, float(v))
                        else: sale = max(sale or 0, float(v))
            rows.append({
                "raw_id": r[0], "material": mat, "outer_diam": od, "thickness": th,
                "temper": r[4] or "", "part_no": r[5] or "", "unit": r[6] or "KG", "std_length": r[7],
                "y2026_kg": float(r[8]) if r[8] is not None else 0, "codes_cnt": len(codes), "vendors": r[10] or "",
                "price_sise": float(sise) if sise is not None else None,     # 직거래 시세
                "price_partner": float(partner) if partner is not None else None,  # 파트너(사급)가
                "price_buy": buy, "price_sale": sale, "price_sagub_item": sagub,
                "price_lg_recog": float(plme[_MATCODE.get(mat, mat)]) if _MATCODE.get(mat, mat) in plme else None})
        return {"rows": rows, "cnt": len(rows),
                "materials": [{"code": m, "nm": m} for m in ["구리", "고강도관", "알루미늄", "STS"]]}
    finally:
        nx.close()


@router.get("/api/rawmat/prices")
def rawmat_prices(raw_id: int = Query(...)):
    """선택 원소재의 월별 단가 시계열: LG인증가(lg_lme_costtable 재료비)·LG사급가(rawmat_lg_sagub, 입력값 미입력=0)·현물가/협력사사급가(price_metal). 최신월 상단."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT material, outer_diam, thickness, temper, part_no FROM nx.raw_material WHERE raw_id=?", raw_id)
        r = cur.fetchone()
        if not r:
            return {"raw_id": raw_id, "material": "", "spec": "", "rows": []}
        mat = str(r[0] or "").strip(); od = r[1]; th = r[2]
        mg = _MATCODE.get(mat, mat)
        series = {}
        cur.execute("SELECT apply_ym, std_price, partner_price FROM nx.price_metal WHERE metal_gubun=? AND diam=? AND thick=?", mg, od, th)
        for ym, sp, pp in cur.fetchall():
            d = series.setdefault(str(ym).strip(), {})
            d["sise"] = float(sp) if sp is not None else None
            d["partner"] = float(pp) if pp is not None else None
        _lgg = _LG_GUBUN.get(mg)
        if _lgg:
            cur.execute("SELECT apply_ym, MAX(jaeryo) FROM nx.lg_lme_costtable WHERE gubun=? GROUP BY apply_ym", _lgg)
            for ym, v in cur.fetchall():
                series.setdefault(str(ym).strip(), {})["lg_recog"] = float(v) if v is not None else None
        cur.execute("SELECT apply_ym, price FROM nx.rawmat_lg_sagub WHERE raw_id=?", raw_id)
        for ym, v in cur.fetchall():
            series.setdefault(str(ym).strip(), {})["lg_sagub"] = float(v) if v is not None else 0
        rows = []
        for ym in sorted(series.keys(), reverse=True):
            d = series[ym]
            rows.append({"ym": ym, "lg_recog": d.get("lg_recog"), "lg_sagub": d.get("lg_sagub", 0),
                         "sise": d.get("sise"), "partner": d.get("partner")})
        spec = f"⌀{od}×{th}" + (f" {r[3]}" if r[3] else "")
        return {"raw_id": raw_id, "material": mat, "spec": spec, "part_no": r[4] or "", "rows": rows}
    finally:
        nx.close()


@router.post("/api/rawmat/lg_sagub/save")
def rawmat_lg_sagub_save(payload: dict = Body(...)):
    """원소재별 월별 LG사급가 저장(upsert). rows=[{ym,price}]. 미입력월은 0."""
    raw_id = int(payload.get("raw_id") or 0)
    rows = payload.get("rows", []) or []
    if not raw_id:
        return {"ok": False, "error": "raw_id 필요"}
    nx = _nx(); cur = nx.cursor()
    try:
        n = 0
        for r in rows:
            ym = "".join(ch for ch in str(r.get("ym", "")) if ch.isdigit())[:6]
            if len(ym) != 6:
                continue
            price = float(r.get("price") or 0)
            cur.execute("""MERGE nx.rawmat_lg_sagub AS t USING (SELECT ? raw_id, ? apply_ym) s
                ON t.raw_id=s.raw_id AND t.apply_ym=s.apply_ym
                WHEN MATCHED THEN UPDATE SET price=?, upd_user='web', upd_dt=GETDATE()
                WHEN NOT MATCHED THEN INSERT(raw_id,apply_ym,price,upd_user,upd_dt) VALUES(?,?,?,'web',GETDATE());""",
                raw_id, ym, price, raw_id, ym, price)
            n += 1
        return {"ok": True, "count": n}
    finally:
        nx.close()


# ============ 절삭재료비 — 원소재 마스터 탭 ============
# 조회 = 라이브 dbo.CS_M_METERIAL_COST(레거시 w_cs_master_050) ∪ 웹입력 nx.cut_matcost_web. 쓰기 = 웹전용 nx.
# 웹 필드 ↔ CS_M_METERIAL_COST 컬럼. market=MARKET_COST(수치)·remarks=REMARKS(텍스트 L/W·직관 등).
_MC_MAP = [("diam", "ITEM_DIAM", 1), ("thick", "ITEM_THICK", 1), ("matcost", "MAT_COST", 1),
           ("proccost", "PROC_COST", 1), ("exrate", "EXCHANGE_RATE", 1), ("totcost", "TOT_COST", 1),
           ("totcust", "TOT_COST_CUST", 1), ("totsub", "TOT_COST_SUB", 1), ("market", "MARKET_COST", 1),
           ("remarks", "REMARKS", 0)]
_MC_COLS = ",".join(c for _, c, _ in _MC_MAP)

def _mc_f(v):
    if v in (None, ""): return None
    try: return float(v)
    except Exception: return None

def _mc_obj(r, user, dt):
    o = {}
    for i, (f, c, n) in enumerate(_MC_MAP):
        o[f] = _num(r[i]) if n else str(r[i] or "").strip()
    o["user"] = str(user or "").strip(); o["dt"] = (str(dt)[:19] if dt else "")
    return o

def _mc_live_rows(cur, metal, ym):
    cur.execute(f"""SELECT {_MC_COLS}, ISNULL(UPDATE_USER_ID,INSERT_USER_ID), ISNULL(UPDATE_DATETIME,INSERT_DATETIME)
        FROM dbo.CS_M_METERIAL_COST WHERE LTRIM(RTRIM(METAL_GUBUN))=? AND LTRIM(RTRIM(APPLY_YYYYMM))=?
        ORDER BY ITEM_DIAM, ITEM_THICK""", metal, ym)
    n = len(_MC_MAP)
    return [_mc_obj(r[:n], r[n], r[n + 1]) for r in cur.fetchall()]

def _mc_web_rows(curnx, metal, ym):
    curnx.execute(f"""SELECT {_MC_COLS}, upd_user, upd_at FROM nx.cut_matcost_web
        WHERE METAL_GUBUN=? AND APPLY_YYYYMM=? ORDER BY ITEM_DIAM, ITEM_THICK""", metal, ym)
    n = len(_MC_MAP)
    return [_mc_obj(r[:n], r[n], r[n + 1]) for r in curnx.fetchall()]

def _mc_web_months(curnx, metal):
    curnx.execute("SELECT DISTINCT LTRIM(RTRIM(APPLY_YYYYMM)) FROM nx.cut_matcost_web WHERE METAL_GUBUN=?", metal)
    return {str(r[0]).strip() for r in curnx.fetchall()}

def _mc_write_web(curnx, metal, ym, rows, user):
    """소재+년월 통째 교체(DELETE 후 INSERT) → nx.cut_matcost_web."""
    curnx.execute("DELETE FROM nx.cut_matcost_web WHERE METAL_GUBUN=? AND APPLY_YYYYMM=?", metal, ym)
    ph = ",".join(["?"] * len(_MC_MAP))
    ins = f"INSERT INTO nx.cut_matcost_web(METAL_GUBUN,APPLY_YYYYMM,{_MC_COLS},upd_user) VALUES(?,?,{ph},?)"
    n = 0
    for r in rows:
        vals = [metal, ym]
        for f, c, num in _MC_MAP:
            v = r.get(f)
            vals.append(_mc_f(v) if num else (None if v is None else str(v)[:200]))
        vals.append(user)
        curnx.execute(ins, *vals); n += 1
    return n

@router.get("/api/rawmat/matcost")
def rawmat_matcost(metal: str = Query("CU"), ym: str = Query("")):
    """절삭재료비 조회 — 조회=라이브 CS_M_METERIAL_COST ∪ 웹입력(cut_matcost_web).
       그 소재·월에 웹 입력분이 있으면 웹 우선(src=web), 없으면 라이브(src=live). 요청월 없으면 최신월 폴백."""
    metal = (metal or "CU").strip()
    ymq = "".join(ch for ch in str(ym or "") if ch.isdigit())[:6]
    cn = _conn(); cur = cn.cursor(); nx = _nx(); curnx = nx.cursor()
    try:
        cur.execute("SELECT DISTINCT LTRIM(RTRIM(METAL_GUBUN)) FROM dbo.CS_M_METERIAL_COST WHERE ISNULL(METAL_GUBUN,'')<>'' ORDER BY 1")
        metals = [str(r[0]).strip() for r in cur.fetchall()]
        if metal not in metals and metals:
            metal = "CU" if "CU" in metals else metals[0]
        cur.execute("SELECT DISTINCT LTRIM(RTRIM(APPLY_YYYYMM)) FROM dbo.CS_M_METERIAL_COST WHERE LTRIM(RTRIM(METAL_GUBUN))=?", metal)
        live_months = {str(r[0]).strip() for r in cur.fetchall()}
        web_months = _mc_web_months(curnx, metal)
        months = sorted(live_months | web_months, reverse=True)
        ym_used = ymq if (ymq and ymq in months) else (months[0] if months else "")
        rows = []; src = "live"
        if ym_used:
            if ym_used in web_months:
                rows = _mc_web_rows(curnx, metal, ym_used); src = "web"
            else:
                rows = _mc_live_rows(cur, metal, ym_used)
        return {"metal": metal, "ym_req": ymq, "ym_used": ym_used, "is_fallback": bool(ymq and ymq != ym_used),
                "metals": metals, "months": months, "web_months": sorted(web_months, reverse=True),
                "rows": rows, "src": src}
    finally:
        cn.close(); nx.close()

@router.post("/api/rawmat/matcost/save")
def rawmat_matcost_save(payload: dict = Body(...)):
    """절삭재료비 저장 — 웹전용 nx.cut_matcost_web에 소재+년월 통째 교체. rows=[{diam,thick,matcost,...}]."""
    metal = str(payload.get("metal") or "").strip()
    ym = "".join(ch for ch in str(payload.get("ym") or "") if ch.isdigit())[:6]
    rows = payload.get("rows", []) or []
    user = str(payload.get("user") or "web")[:50]
    if not metal or len(ym) != 6:
        return {"ok": False, "error": "metal·ym(YYYYMM) 필요"}
    nx = _nx_tx(); cur = nx.cursor()
    try:
        n = _mc_write_web(cur, metal, ym, rows, user)
        nx.commit()
        return {"ok": True, "count": n, "metal": metal, "ym": ym}
    except Exception as e:
        nx.rollback(); return {"ok": False, "error": str(e)}
    finally:
        nx.close()

@router.post("/api/rawmat/matcost/copy")
def rawmat_matcost_copy(payload: dict = Body(...)):
    """방식1 — 다른 월 통째 복사: src_ym(라이브∪웹) 데이터를 dst_ym으로 복사(웹테이블에 통째 저장)."""
    metal = str(payload.get("metal") or "").strip()
    src = "".join(ch for ch in str(payload.get("src_ym") or "") if ch.isdigit())[:6]
    dst = "".join(ch for ch in str(payload.get("dst_ym") or "") if ch.isdigit())[:6]
    user = str(payload.get("user") or "web")[:50]
    if not metal or len(src) != 6 or len(dst) != 6:
        return {"ok": False, "error": "metal·src_ym·dst_ym(YYYYMM) 필요"}
    if src == dst:
        return {"ok": False, "error": "복사 원본월과 대상월이 같습니다"}
    cn = _conn(); curL = cn.cursor(); nx = _nx_tx(); cur = nx.cursor()
    try:
        srcrows = _mc_web_rows(cur, metal, src) if src in _mc_web_months(cur, metal) else _mc_live_rows(curL, metal, src)
        if not srcrows:
            return {"ok": False, "error": f"{src} 원본 데이터가 없습니다"}
        n = _mc_write_web(cur, metal, dst, srcrows, user)
        nx.commit()
        return {"ok": True, "count": n, "metal": metal, "src_ym": src, "dst_ym": dst}
    except Exception as e:
        nx.rollback(); return {"ok": False, "error": str(e)}
    finally:
        cn.close(); nx.close()

@router.post("/api/rawmat/matcost/upload")
def rawmat_matcost_upload(file: UploadFile = File(...), metal: str = Form("CU"),
                          ym: str = Form(...), user: str = Form("web")):
    """방식2 — LG '원소재 사급가 변경' 엑셀 드래그&드롭 등록(웹테이블 저장).
       파싱: 구분(→remarks)·외경(diam)·t(thick)·사급가 변경후(→ 원재료비 totcost).
       최신월(라이브∪웹)을 템플릿으로 복사 후 (diam,thick) 매칭 행의 totcost=변경후 덮어쓰기 → 대상 ym 통째 저장."""
    import openpyxl, io as _io
    metal = (metal or "CU").strip()
    ymd = "".join(ch for ch in str(ym or "") if ch.isdigit())[:6]
    if len(ymd) != 6:
        return {"ok": False, "error": "등록 년월(YYYYMM) 필요"}
    try:
        raw = file.file.read()
        wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return {"ok": False, "error": f"엑셀 읽기 실패: {e}"}
    hr = None
    for r in range(1, min(ws.max_row, 12) + 1):
        for cc in range(1, ws.max_column + 1):
            if str(ws.cell(r, cc).value or "").strip() == "외경":
                hr = r; break
        if hr: break
    if not hr:
        return {"ok": False, "error": "헤더('외경')를 찾지 못했습니다"}
    col = {}
    for cc in range(1, ws.max_column + 1):
        h = str(ws.cell(hr, cc).value or "").strip()
        if h == "구분": col["gubun"] = cc
        elif h == "외경": col["diam"] = cc
        elif h in ("t", "두께"): col["thick"] = cc
    after_col = None
    for cc in range(1, ws.max_column + 1):
        if str(ws.cell(hr + 1, cc).value or "").strip() == "변경후":
            after_col = cc; break
    if not after_col or "diam" not in col or "thick" not in col:
        return {"ok": False, "error": "외경/두께/변경후 컬럼을 찾지 못했습니다"}
    parsed = []
    for r in range(hr + 2, ws.max_row + 1):
        d = ws.cell(r, col["diam"]).value; t = ws.cell(r, col["thick"]).value
        pa = ws.cell(r, after_col).value
        if d in (None, "") or pa in (None, ""):
            continue
        try:
            parsed.append({"diam": float(d), "thick": (float(t) if t not in (None, "") else None),
                           "after": float(pa),
                           "gubun": (str(ws.cell(r, col["gubun"]).value).strip() if col.get("gubun") and ws.cell(r, col["gubun"]).value else None)})
        except Exception:
            continue
    if not parsed:
        return {"ok": False, "error": "파싱된 데이터가 없습니다"}
    cn = _conn(); curL = cn.cursor(); nx = _nx_tx(); cur = nx.cursor()
    try:
        # 템플릿 = 해당 소재 최신월(대상월 제외, 라이브∪웹)
        wm = sorted([m for m in _mc_web_months(cur, metal) if m != ymd], reverse=True)
        curL.execute("SELECT TOP 1 LTRIM(RTRIM(APPLY_YYYYMM)) FROM dbo.CS_M_METERIAL_COST WHERE LTRIM(RTRIM(METAL_GUBUN))=? AND LTRIM(RTRIM(APPLY_YYYYMM))<>? ORDER BY 1 DESC", metal, ymd)
        lr = curL.fetchone(); lm = str(lr[0]).strip() if lr else None
        tmpl_ym = max([x for x in (wm[0] if wm else None, lm) if x], default=None)
        tmpl = {}
        if tmpl_ym:
            trows = _mc_web_rows(cur, metal, tmpl_ym) if tmpl_ym in _mc_web_months(cur, metal) else _mc_live_rows(curL, metal, tmpl_ym)
            for tr in trows:
                tmpl[(round(float(tr.get("diam") or 0), 3), round(float(tr.get("thick") or 0), 3))] = tr
        matched = 0; added = 0; out = []
        for p in parsed:
            key = (round(p["diam"], 3), round(p["thick"] or 0, 3))
            base = dict(tmpl.get(key, {}))
            if key in tmpl: matched += 1
            else: added += 1
            base.pop("user", None); base.pop("dt", None)
            base["diam"] = p["diam"]; base["thick"] = p["thick"]
            base["totcost"] = p["after"]                    # 사급가 변경후 → 원재료비
            if p["gubun"]: base["remarks"] = p["gubun"]     # 구분(L/W·고강도관·직관&P/C) → 비고(REMARKS)
            out.append(base)
        n = _mc_write_web(cur, metal, ymd, out, user)
        nx.commit()
        return {"ok": True, "count": n, "matched": matched, "added": added,
                "tmpl_ym": tmpl_ym, "metal": metal, "ym": ymd}
    except Exception as e:
        nx.rollback(); return {"ok": False, "error": str(e)}
    finally:
        cn.close(); nx.close()
