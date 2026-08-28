# -*- coding: utf-8 -*-
"""LG사급현황 — LG 사급 실적(월별 유상사급 입고) 엑셀 업로드 + 조회. 사급가 업로드와 유사 패턴.
   저장=nx.lg_sagub_actual. 실제 사급 입고(월 18~22억)를 넣어 우리 계산(예상/원가분석)과 대사용."""
import io as _io
import re as _re
from fastapi import APIRouter, Query, UploadFile, File, HTTPException
from common import _nx, _conn

router = APIRouter()

# ── BOM기준 동 소요(규격별) + 절삭재료비 신규 사급가 as-of (LG사급현황 대사) ──
try:
    import nx_soyo_engine as _soyo            # common.py가 _harness를 sys.path에 추가
    from nx_cost_engine import NxCostEngine
except Exception:
    _soyo = None; NxCostEngine = None
_WENG = None


def _weng():
    global _WENG
    if _WENG is None:
        _WENG = NxCostEngine()
    return _WENG


def _bom_sig(cur):
    cur.execute("SELECT COUNT(*), ISNULL(CHECKSUM_AGG(BINARY_CHECKSUM(bom_id,child_item,qty,ISNULL(qty_pr,qty),ISNULL(except_flag,0))),0) FROM nx.bom_line")
    r = cur.fetchone()
    return "%s:%s" % (r[0], r[1])


def _ensure_dong_cache(cur):
    """완제품 규격별 동중량 캐시 + BOM 서명가드(변경시 무효)."""
    global _WENG
    cur.execute("""IF OBJECT_ID('nx.item_dong_spec') IS NULL CREATE TABLE nx.item_dong_spec(
        item_code varchar(30) NOT NULL, metal varchar(20) NOT NULL, diam float NOT NULL, thick float NOT NULL, per_unit float,
        CONSTRAINT pk_item_dong_spec PRIMARY KEY(item_code,metal,diam,thick))""")
    cur.execute("IF OBJECT_ID('nx.item_dong_spec_meta') IS NULL CREATE TABLE nx.item_dong_spec_meta(id int PRIMARY KEY, bom_sig varchar(80), built_dt datetime)")
    sig = _bom_sig(cur)
    cur.execute("SELECT bom_sig FROM nx.item_dong_spec_meta WHERE id=1")
    r = cur.fetchone()
    if (not r) or (r[0] != sig):
        _WENG = None
        cur.execute("TRUNCATE TABLE nx.item_dong_spec")
        cur.execute("DELETE FROM nx.item_dong_spec_meta WHERE id=1")
        cur.execute("INSERT INTO nx.item_dong_spec_meta(id,bom_sig,built_dt) VALUES(1,?,getdate())", sig)


def _dong_of(cur, item):
    """완제품 규격별 동중량 {(metal,diam,thick): per_unit} — ★nx.bom_flat(검증정본·변형SUB dedup) 기반.
       ★copper_by_spec(nx_soyo_engine, 소스 nx.bom_line)는 변형 SUB 두 경로(-3-1/-20-1 등)로 같은 동을 2중계상(정확히 2배 과다)해서 폐기.
         (AJR30004702: copper_by_spec 0.6986 = bom_flat 0.3493×2. LG BOM·bom_flat은 1회.) 기록 §7·LME과다·subvariant 계열.
       중량=bom_flat.weight_actual(우리실측)×qty, 규격=(metal_gubun[nx.item], fin_diam, fin_thick).
       ★동 재질만 = metal_gubun IN ('CU','고강도')(_WT_COPPER). role LIKE '%동%'는 STS 제작동관까지 잡아 오염(STS 22.2×1 등 절삭재료비 미매칭)→ 재질필터로 교체."""
    cur.execute("""SELECT LTRIM(RTRIM(i.metal_gubun)) mg, ISNULL(bf.fin_diam,0) d, ISNULL(bf.fin_thick,0) t,
                     SUM(ISNULL(bf.weight_actual,0)*ISNULL(bf.qty,0)) w
                   FROM nx.bom_flat bf
                   JOIN nx.item i ON UPPER(LTRIM(RTRIM(i.item_code)))=UPPER(LTRIM(RTRIM(bf.leaf_code)))
                   WHERE UPPER(LTRIM(RTRIM(bf.item_code)))=? AND ISNULL(bf.weight_actual,0)>0
                     AND LTRIM(RTRIM(i.metal_gubun)) IN (N'CU', N'고강도')
                   GROUP BY LTRIM(RTRIM(i.metal_gubun)), ISNULL(bf.fin_diam,0), ISNULL(bf.fin_thick,0)""",
                item.strip().upper())
    out = {}
    for mg, d, t, w in cur.fetchall():
        k = ((mg or '').strip(), float(d or 0), float(t or 0))
        out[k] = out.get(k, 0.0) + float(w or 0)
    return out


def _dong_of_batch(cur, items):
    """★성능: 여러 완제품의 우리 실측 동중량(bom_flat 제작동관·CU/고강도)을 1쿼리로 = {item: {(metal,diam,thick): kg}}.
       _dong_of 개별 N회(느림) 대체."""
    out = {}
    safe = [str(m).strip().upper() for m in (items or []) if m and all(c.isalnum() or c in '-_' for c in str(m).strip())]
    if not safe:
        return out
    inl = ",".join("'" + m + "'" for m in safe)
    cur.execute(f"""SELECT UPPER(LTRIM(RTRIM(bf.item_code))), LTRIM(RTRIM(i.metal_gubun)),
                     ISNULL(bf.fin_diam,0), ISNULL(bf.fin_thick,0), SUM(ISNULL(bf.weight_actual,0)*ISNULL(bf.qty,0))
                   FROM nx.bom_flat bf JOIN nx.item i ON UPPER(LTRIM(RTRIM(i.item_code)))=UPPER(LTRIM(RTRIM(bf.leaf_code)))
                   WHERE UPPER(LTRIM(RTRIM(bf.item_code))) IN ({inl}) AND ISNULL(bf.weight_actual,0)>0
                     AND LTRIM(RTRIM(i.metal_gubun)) IN (N'CU', N'고강도')
                   GROUP BY UPPER(LTRIM(RTRIM(bf.item_code))), LTRIM(RTRIM(i.metal_gubun)), ISNULL(bf.fin_diam,0), ISNULL(bf.fin_thick,0)""")
    for it, mg, d, t, w in cur.fetchall():
        out.setdefault(it, {})[((mg or '').strip(), float(d or 0), float(t or 0))] = float(w or 0)
    return out


def _model_in_sql(models):
    """models(set/iterable) → ' AND UPPER(LTRIM(RTRIM(model))) IN (...)' 조각. 성능: 필요 모델만 전개.
       코드 안전문자(영숫자·-_)만 통과(인젝션 방지). 비면 빈문자."""
    if not models:
        return ""
    safe = [str(m).strip().upper() for m in models if m and all(c.isalnum() or c in '-_' for c in str(m).strip())]
    if not safe:
        return ""
    return " AND UPPER(LTRIM(RTRIM(model))) IN (" + ",".join("'" + m + "'" for m in safe) + ")"


def _lg_ap_all(cur, ver_date, models=None):
    """LG BOM 버전(point-in-time) 사급(Assembly Pull) 동 원소재 소요 = {model: {(metal,diam,thick): per_unit_kg}}.
       ★LG전자 사급 소요 산출방식 = LG BOM 다단계 트리전개(롤업): 동이 L2(서브 밑)면 L1 서브 수량을 곱해 누적.
         (구 flat합은 L1 수량 미곱 → 과소. 예 AJR30004702 P7.0 L1 ×7 반영.)
       동 원소재 = matkl='MJU0631'·supply_type='Assembly Pull'·ALUMINUM 제외. Supplier·사급부품(matkl≠)은 전개 관통만 하고 미계상.
       point-in-time = model·werks별 ver_from<=ver_date 최신 버전. werks 다중이면 전개합 MAX(양공장 중복방지).
       규격/재질 = nx.item 우선, 없으면 child_spec 파싱. root = model(STUFE1 부모=model).
       ★models(대문자 set) 지정시 그 모델만 전개(성능: 전 모델 전개 회피)."""
    from collections import defaultdict as _dd
    minl = _model_in_sql(models)
    cur.execute(f"""
      WITH latest AS (
        SELECT model, ISNULL(werks,'') w, MAX(ver_from) mv
        FROM nx.lg_bom_ver WHERE ver_from<=? {minl} GROUP BY model, ISNULL(werks,''))
      SELECT UPPER(LTRIM(RTRIM(r.model))), ISNULL(r.werks,''),
             UPPER(LTRIM(RTRIM(r.parent_code))), UPPER(LTRIM(RTRIM(r.child_code))),
             r.matkl, LTRIM(RTRIM(ISNULL(r.supply_type,''))), ISNULL(r.child_spec,''), CONVERT(float,ISNULL(r.qty,0)),
             ISNULL(ic.metal_gubun,''), ISNULL(ic.diam,0), ISNULL(ic.thick,0)
      FROM nx.lg_bom_ver r
      JOIN latest l ON l.model=r.model AND l.w=ISNULL(r.werks,'') AND r.ver_from=l.mv
      LEFT JOIN nx.item ic ON UPPER(LTRIM(RTRIM(ic.item_code)))=UPPER(LTRIM(RTRIM(r.child_code)))
    """, ver_date)
    MW = _dd(lambda: _dd(list))    # model -> werks -> [(parent,child,matkl,sup,spec,qty,mg,diam,thick)]
    for md, w, p, c, mk, sup, spec, q, mg, idiam, ithick in cur.fetchall():
        MW[md][w].append((p, c, mk, sup, spec, float(q or 0), (mg or '').strip(), float(idiam or 0), float(ithick or 0)))

    def _key(spec, mg, idiam, ithick):
        od = idiam if idiam > 0 else None; thk = ithick if ithick > 0 else None
        if od is None:
            m = _re.search(r'P(\d+(?:\.\d+)?)', spec); od = float(m.group(1)) if m else 0.0
        if thk is None:
            m = _re.search(r'T(\d+(?:\.\d+)?)', spec); thk = float(m.group(1)) if m else 0.0
        metal = mg if mg else ('고강도' if '고강도' in spec else 'CU')
        return (metal, float(od), float(thk))

    out = {}
    for md, wmap in MW.items():
        best = None; best_tot = -1.0
        for w, edges in wmap.items():
            ch = _dd(list)
            for e in edges:
                ch[e[0]].append(e)
            acc = _dd(float); tot = [0.0]

            def dfs(node, mult, depth, path):
                if depth > 25:
                    return
                for (p, c, mk, sup, spec, q, mg, idiam, ithick) in ch.get(node, ()):
                    # ★q=1.0(정확값) 동 = LG 데이터 플레이스홀더(검증 6모델: 정상 0.008~0.5인데 1.0). 제외.
                    if mk == 'MJU0631' and sup == 'Assembly Pull' and 'ALUMINUM' not in spec.upper() and abs(q - 1.0) > 1e-9:
                        cv = q * mult
                        acc[_key(spec, mg, idiam, ithick)] += cv; tot[0] += cv
                    if c != node and c not in path:       # EA 중간노드=수량 곱해 관통, 동 leaf=자식없어 종료, cycle 방지
                        dfs(c, mult * q, depth + 1, path | {c})
            dfs(md, 1.0, 0, {md})
            if tot[0] > best_tot:
                best_tot = tot[0]; best = dict(acc)
        if best:
            out[md] = best
    return out


def _lg_ap_split(cur, ver_date, models=None, jjset=None):
    """★B: LG BOM AP 동 소요(=전체 사급 동, 우리가 협력사에 사급 주는 소재 포함)를 **분할**(2중계상 없음, 전체=우리절삭+협력사사급).
       분할 기준 = 각 동(Tube,Raw)의 부모(절단관)가 **우리 제작동관(bom_flat role='제작동관', 우리가 깎음)** 이면 '우리절삭', 아니면(사급 SUB=협력사가 깎음, 우리가 소재 사급) '협력사사급'.
       반환 {model: {'our':{spec:kg}, 'coop':{spec:kg}}}. 전개·롤업·플레이스홀더제외는 _lg_ap_all과 동일. models 지정시 그 모델만(성능)."""
    from collections import defaultdict as _dd
    # 우리 제작동관 코드집합(우리가 직접 깎는 절단관). jjset 주면 재사용(월별 반복 성능).
    if jjset is None:
        cur.execute("SELECT DISTINCT UPPER(LTRIM(RTRIM(leaf_code))) FROM nx.bom_flat WHERE role=N'제작동관'")
        jjset = set(r[0] for r in cur.fetchall())
    minl = _model_in_sql(models)
    cur.execute(f"""
      WITH latest AS (
        SELECT model, ISNULL(werks,'') w, MAX(ver_from) mv
        FROM nx.lg_bom_ver WHERE ver_from<=? {minl} GROUP BY model, ISNULL(werks,''))
      SELECT UPPER(LTRIM(RTRIM(r.model))), ISNULL(r.werks,''),
             UPPER(LTRIM(RTRIM(r.parent_code))), UPPER(LTRIM(RTRIM(r.child_code))),
             r.matkl, LTRIM(RTRIM(ISNULL(r.supply_type,''))), ISNULL(r.child_spec,''), CONVERT(float,ISNULL(r.qty,0)),
             ISNULL(ic.metal_gubun,''), ISNULL(ic.diam,0), ISNULL(ic.thick,0)
      FROM nx.lg_bom_ver r
      JOIN latest l ON l.model=r.model AND l.w=ISNULL(r.werks,'') AND r.ver_from=l.mv
      LEFT JOIN nx.item ic ON UPPER(LTRIM(RTRIM(ic.item_code)))=UPPER(LTRIM(RTRIM(r.child_code)))
    """, ver_date)
    MW = _dd(lambda: _dd(list))
    for md, w, p, c, mk, sup, spec, q, mg, idiam, ithick in cur.fetchall():
        MW[md][w].append((p, c, mk, sup, spec, float(q or 0), (mg or '').strip(), float(idiam or 0), float(ithick or 0)))

    def _key(spec, mg, idiam, ithick):
        od = idiam if idiam > 0 else None; thk = ithick if ithick > 0 else None
        if od is None:
            m = _re.search(r'P(\d+(?:\.\d+)?)', spec); od = float(m.group(1)) if m else 0.0
        if thk is None:
            m = _re.search(r'T(\d+(?:\.\d+)?)', spec); thk = float(m.group(1)) if m else 0.0
        metal = mg if mg else ('고강도' if '고강도' in spec else 'CU')
        return (metal, float(od), float(thk))

    out = {}
    for md, wmap in MW.items():
        best = None; best_tot = -1.0
        for w, edges in wmap.items():
            ch = _dd(list)
            for e in edges:
                ch[e[0]].append(e)
            our = _dd(float); coop = _dd(float); tot = [0.0]

            def dfs(node, mult, depth, path):
                if depth > 25:
                    return
                for (p, c, mk, sup, spec, q, mg, idiam, ithick) in ch.get(node, ()):
                    if mk == 'MJU0631' and sup == 'Assembly Pull' and 'ALUMINUM' not in spec.upper() and abs(q - 1.0) > 1e-9:
                        cv = q * mult; k = _key(spec, mg, idiam, ithick)
                        (our if node in jjset else coop)[k] += cv    # node=이 동의 부모(절단관). 우리 제작동관이면 우리절삭
                        tot[0] += cv
                    if c != node and c not in path:
                        dfs(c, mult * q, depth + 1, path | {c})
            dfs(md, 1.0, 0, {md})
            if tot[0] > best_tot:
                best_tot = tot[0]; best = {'our': dict(our), 'coop': dict(coop)}
        if best:
            out[md] = best
    return out


def _matcost_asof(ym6):
    """절삭재료비 신규 사급가 as-of: {(metal,diam,thick): TOT_COST}, APPLY_YYYYMM ≤ ym6(YYYYMM) 최신."""
    c2 = _conn(); cu = c2.cursor()
    try:
        cu.execute("""SELECT METAL_GUBUN,ITEM_DIAM,ITEM_THICK,TOT_COST FROM PARTNER_ERP.dbo.CS_M_METERIAL_COST
            WHERE APPLY_YYYYMM=(SELECT MAX(APPLY_YYYYMM) FROM PARTNER_ERP.dbo.CS_M_METERIAL_COST WHERE APPLY_YYYYMM<=?)""", ym6)
        return {(str(r[0]).strip(), float(r[1] or 0), float(r[2] or 0)): float(r[3] or 0) for r in cu.fetchall()}
    finally:
        c2.close()

_DDL = """IF OBJECT_ID('nx.lg_sagub_actual') IS NULL
CREATE TABLE nx.lg_sagub_actual(
  id int IDENTITY(1,1) PRIMARY KEY, ym varchar(4), ymd varchar(8), biz varchar(4), item_code varchar(50), item_name nvarchar(200),
  qty float, amt float, price float, cust_code varchar(20), remarks nvarchar(300),
  src_file nvarchar(200), upload_dt datetime DEFAULT getdate())"""
# 기존 테이블 마이그레이션: ymd(일자)·biz(사업부 RAC/SAC) 컬럼 없으면 추가
_MIGRATE = """IF COL_LENGTH('nx.lg_sagub_actual','ymd') IS NULL ALTER TABLE nx.lg_sagub_actual ADD ymd varchar(8);
IF COL_LENGTH('nx.lg_sagub_actual','biz') IS NULL ALTER TABLE nx.lg_sagub_actual ADD biz varchar(4);"""

def _prep(cur):
    cur.execute(_DDL); cur.execute(_MIGRATE)

def _biz_norm(v):
    """사업부 정규화: rac/dgz→RAC, sac/dmz→SAC. 그 외는 대문자 그대로."""
    s = str(v or "").strip().lower()
    if not s: return ""
    if "rac" in s or "dgz" in s: return "RAC"
    if "sac" in s or "dmz" in s: return "SAC"
    return str(v).strip().upper()[:4]

def _ymd6(v):
    """값에서 YYMMDD(6) 추출. 날짜/문자 모두. 실패시 ''."""
    if v is None or v == "": return ""
    if hasattr(v, "year"):
        return f"{v.year % 100:02d}{v.month:02d}{v.day:02d}"
    s = "".join(ch for ch in str(v) if ch.isdigit())
    if len(s) >= 8: return s[2:8]   # YYYYMMDD(+시간) → YYMMDD
    if len(s) == 6: return s        # YYMMDD
    return ""

# 엑셀 헤더 → 컬럼 매핑(유연 감지). 소문자/공백제거 비교.
_ALIAS = {
    "item": ["품번", "품목", "품목코드", "자재", "자재코드", "material", "item", "itemcode", "code", "partno", "part", "품번호"],
    "name": ["품명", "품목명", "품명규격", "name", "itemname", "desc", "description", "material desc"],
    "qty": ["수량", "입고수량", "사급수량", "qty", "quantity", "inputqty", "recvqty", "입고량"],
    "amt": ["금액", "입고금액", "사급금액", "가액", "공급가액", "합계금액", "amount", "amt", "value", "총액"],
    "price": ["단가", "사급가", "사급단가", "price", "unitprice", "unit"],
    "ym": ["월", "기준월", "년월", "ym", "yearmonth"],
    "ymd": ["일자", "입고일", "입고일자", "적용일", "date", "ymd", "startdate", "start date", "transaction_ymd", "in_ymd"],
    "cust": ["거래처", "업체", "사업장", "매입처", "cust", "vendor", "site", "plant"],
}


def _norm(h):
    return "".join(str(h or "").strip().lower().split())


def _ym_of(v):
    """값에서 YYMM 추출(날짜/문자/월)."""
    if v is None or v == "":
        return ""
    if hasattr(v, "year"):
        return f"{v.year % 100:02d}{v.month:02d}"
    s = "".join(ch for ch in str(v) if ch.isdigit())
    if len(s) >= 6:      # YYYYMMDD.. or YYYYMM
        return s[2:6] if len(s) >= 8 else s[2:6]
    if len(s) == 4:      # YYMM
        return s
    if len(s) <= 2:      # 월만(1~12)
        try:
            m = int(s)
            if 1 <= m <= 12:
                return None  # 월만으론 연도 모름 → 호출측 기준연도 사용
        except Exception:
            pass
    return ""


@router.post("/api/lgsagub/upload")
async def lgsagub_upload(file: UploadFile = File(...), ym: str = Query(""), base_year: str = Query("26"), biz: str = Query("")):
    """LG 사급 실적 엑셀 업로드 → nx.lg_sagub_actual. 컬럼 유연감지(품번/수량/금액/단가/일자). 감지결과 반환.
       ★biz(사업부 RAC/SAC)=업로드 시 사용자 선택(파일명에 없어서). 일자(Transaction Date)→ymd(YYMMDD), ym=ymd 앞4."""
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb.active
    rows_all = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_all:
        raise HTTPException(400, "빈 파일")
    # 헤더행 자동탐지: 앞 10행 중 alias 매칭 최다 행
    best_hi, best_map, best_score = 0, {}, -1
    for hi in range(min(10, len(rows_all))):
        hdr = rows_all[hi]
        nm = {i: _norm(h) for i, h in enumerate(hdr)}
        mp = {}
        for key, al in _ALIAS.items():
            als = ["".join(a.split()) for a in al]
            for i, h in nm.items():
                if h and (h in als or any(h == a or a in h for a in als)):
                    mp.setdefault(key, i)
        score = len(mp) + (2 if "item" in mp else 0) + (1 if ("qty" in mp or "amt" in mp) else 0)
        if score > best_score:
            best_hi, best_map, best_score = hi, mp, score
    ixmap = best_map
    detected = {k: str(rows_all[best_hi][i]) for k, i in ixmap.items()}
    if "item" not in ixmap:
        return {"ok": False, "error": "품번 컬럼 감지 실패", "header_row": [str(x) for x in rows_all[best_hi]], "detected": detected}

    def gv(r, key):
        i = ixmap.get(key)
        if i is None or i >= len(r):
            return None
        v = r[i]
        return None if v in (None, "") else v

    forced_ym = ym.strip()
    bizv = _biz_norm(biz) or _biz_norm(file.filename)   # 사용자선택 우선, 없으면 파일명(lg_rac/lg_sac)에서 추론
    _SUMMARY = ("total", "합계", "subtotal", "소계", "grand total", "총계")
    recs = []
    for r in rows_all[best_hi + 1:]:
        if not r or not any(x not in (None, "") for x in r):
            continue
        # ★요약/Total 행 스킵(OSP 파일 끝 'Total' 행이 데이터에 섞여 이중계상 방지)
        if any(isinstance(x, str) and x.strip().lower() in _SUMMARY for x in r):
            continue
        it = str(gv(r, "item") or "").strip()
        if not it or it.lower() in ("품번", "material", "품목", "합계", "total", "소계"):
            continue
        def _f(key):
            v = gv(r, key)
            try:
                return float(str(v).replace(",", "")) if v is not None else None
            except Exception:
                return None
        q = _f("qty"); a = _f("amt"); p = _f("price")
        # 일자(Transaction Date)→ymd(YYMMDD), 월=ymd 앞4. forced_ym 있으면 우선.
        rymd = _ymd6(gv(r, "ymd"))
        rym = forced_ym or (rymd[:4] if rymd else "")
        if not rym:
            rym = _ym_of(gv(r, "ym"))
            if rym is None:  # 월만 있음
                mm = "".join(ch for ch in str(gv(r, "ym") or gv(r, "ymd") or "") if ch.isdigit())
                rym = f"{base_year}{int(mm):02d}" if mm.isdigit() and 1 <= int(mm) <= 12 else ""
        recs.append((rym or "", rymd, bizv, it, str(gv(r, "name") or "")[:200], q, a, p,
                     str(gv(r, "cust") or "")[:20], ""))
    if not recs:
        return {"ok": False, "error": "데이터 행 없음", "detected": detected, "header_row": [str(x) for x in rows_all[best_hi]]}
    nx = _nx(); cur = nx.cursor()
    try:
        _prep(cur)
        # ★재업로드=최신 업데이트(#9): 같은 사업부의 "그 파일에 담긴 일자(ymd)" 또는 같은 파일명 기존행 삭제 후 삽입.
        #   OSP 파일명은 다운로드마다 번호가 바뀌므로 파일명만으론 중복제거 안 됨 → biz+ymd 기준으로 같은 사업부·같은 날짜는 최신이 덮어씀.
        fn = (file.filename or "")[:200]
        ymds = sorted({x[1] for x in recs if x[1]})
        if ymds:
            inl = ",".join("'" + y + "'" for y in ymds)
            cur.execute(f"DELETE FROM nx.lg_sagub_actual WHERE ISNULL(biz,'')=? AND (ISNULL(ymd,'') IN ({inl}) OR src_file=?)", bizv, fn)
        else:
            cur.execute("DELETE FROM nx.lg_sagub_actual WHERE ISNULL(biz,'')=? AND src_file=?", bizv, fn)
        # ★배치 다중행 INSERT — 파라미터 2100 한도: 11컬럼 × 150 = 1650 안전
        try: cur.fast_executemany = True
        except Exception: pass
        cols = "(ym,ymd,biz,item_code,item_name,qty,amt,price,cust_code,remarks,src_file)"
        n = 0; BATCH = 150
        for i in range(0, len(recs), BATCH):
            chunk = recs[i:i + BATCH]
            vals = ",".join("(?,?,?,?,?,?,?,?,?,?,?)" for _ in chunk)
            flat = []
            for ymv, ymdv, bzv, it, nmv, q, a, p, cst, rm in chunk:
                flat += [ymv, ymdv, bzv, it, nmv, q, a, p, cst, rm, fn]
            cur.execute(f"INSERT INTO nx.lg_sagub_actual{cols} VALUES {vals}", *flat)
            n += len(chunk)
        nx.commit()
        # 업로드 요약
        cur.execute("SELECT ym, COUNT(*), SUM(ISNULL(qty,0)), SUM(ISNULL(amt,0)) FROM nx.lg_sagub_actual WHERE src_file=? GROUP BY ym ORDER BY ym", fn)
        by_ym = [{"ym": r[0], "rows": r[1], "qty": float(r[2] or 0), "amt": float(r[3] or 0)} for r in cur.fetchall()]
        return {"ok": True, "file": fn, "rows": n, "detected": detected, "by_ym": by_ym}
    finally:
        nx.close()


@router.get("/api/lgsagub/summary")
def lgsagub_summary():
    """업로드분 요약: 월범위·사업부별·파일목록(기간필터 기본값·필터콤보용)."""
    nx = _nx(); cur = nx.cursor()
    try:
        _prep(cur)
        cur.execute("""SELECT ym, COUNT(DISTINCT item_code) items, SUM(ISNULL(qty,0)) qty, SUM(ISNULL(amt,0)) amt
            FROM nx.lg_sagub_actual GROUP BY ym ORDER BY ym""")
        by_ym = [{"ym": r[0], "items": r[1], "qty": float(r[2] or 0), "amt": float(r[3] or 0)} for r in cur.fetchall()]
        cur.execute("SELECT ISNULL(biz,'') b, COUNT(*) c, SUM(ISNULL(amt,0)) amt FROM nx.lg_sagub_actual GROUP BY ISNULL(biz,'') ORDER BY b")
        by_biz = [{"biz": r[0], "rows": r[1], "amt": float(r[2] or 0)} for r in cur.fetchall()]
        cur.execute("SELECT MIN(NULLIF(ymd,'')), MAX(NULLIF(ymd,'')) FROM nx.lg_sagub_actual")
        _rr = cur.fetchone(); ymd_min = _rr[0] or ""; ymd_max = _rr[1] or ""
        cur.execute("SELECT src_file, ISNULL(biz,''), MIN(ym), MAX(ym), COUNT(*), MAX(upload_dt) FROM nx.lg_sagub_actual GROUP BY src_file, ISNULL(biz,'') ORDER BY MAX(upload_dt) DESC")
        files = [{"file": r[0], "biz": r[1], "ym_from": r[2], "ym_to": r[3], "rows": r[4], "dt": str(r[5])[:19]} for r in cur.fetchall()]
        return {"by_ym": by_ym, "by_biz": by_biz, "files": files, "ymd_min": ymd_min, "ymd_max": ymd_max}
    finally:
        nx.close()


def _cls_of(name):
    """분류: 품명에 TUBE 있으면 원소재(동파이프 등), 나머지는 사급부품(사용자 규칙)."""
    return "원소재" if "TUBE" in str(name or "").upper() else "사급부품"


def _range_wh(ym, ym_from, ym_to, ymd_from, ymd_to):
    """기간 필터 조건 생성: ymd(일자 YYMMDD) 범위 우선, 없으면 ym(월). 반환 (조건리스트, 파라미터)."""
    wh = []; p = []
    if ymd_from.strip() or ymd_to.strip():
        if ymd_from.strip(): wh.append("ISNULL(s.ymd,'')>=?"); p.append(ymd_from.strip())
        if ymd_to.strip():   wh.append("ISNULL(s.ymd,'')<=?"); p.append(ymd_to.strip())
    elif ym.strip():
        wh.append("s.ym=?"); p.append(ym.strip())
    else:
        if ym_from.strip(): wh.append("s.ym>=?"); p.append(ym_from.strip())
        if ym_to.strip():   wh.append("s.ym<=?"); p.append(ym_to.strip())
    return wh, p


@router.get("/api/lgsagub/list")
def lgsagub_list(ym: str = Query(""), ym_from: str = Query(""), ym_to: str = Query(""), ymd_from: str = Query(""), ymd_to: str = Query(""),
                 biz: str = Query(""), cls: str = Query(""), q: str = Query(""), limit: int = Query(3000)):
    """LG사급 실적 품목별 요약목록(기간·사업부·분류 필터). 기간=일자범위 ymd_from~ymd_to(YYMMDD) 우선, 없으면 월(ym).
       biz=품목별 사업부(콤마구분), pmin/pmax=단가(다르면 pchg=1), ndays=일자수. cls=원소재(품명 TUBE)/사급부품."""
    nx = _nx(); cur = nx.cursor()
    try:
        _prep(cur)
        wh = ["1=1"]; p = []
        rwh, rp = _range_wh(ym, ym_from, ym_to, ymd_from, ymd_to); wh += rwh; p += rp
        if biz.strip(): wh.append("ISNULL(s.biz,'')=?"); p.append(biz.strip())
        if q.strip():
            wh.append("(s.item_code LIKE ? OR s.item_name LIKE ?)"); p += [f"%{q.strip()}%", f"%{q.strip()}%"]
        whs = ' AND '.join(wh)
        # 품목별 사업부(#8): 같은 필터로 (item, biz) 집계 → item별 사업부 콤마목록
        cur.execute(f"SELECT s.item_code, ISNULL(s.biz,'') b FROM nx.lg_sagub_actual s WHERE {whs} GROUP BY s.item_code, ISNULL(s.biz,'')", *p)
        bizmap = {}
        for ic, b in cur.fetchall(): bizmap.setdefault(ic, set()).add(b or '')
        cur.execute(f"""SELECT TOP {int(limit)} s.item_code,
              MAX(ISNULL(NULLIF(s.item_name,''), ISNULL(i.item_name,''))) nm,
              SUM(ISNULL(s.qty,0)) qty, SUM(ISNULL(s.amt,0)) amt,
              MIN(NULLIF(s.price,0)) pmin, MAX(s.price) pmax, COUNT(*) cnt, COUNT(DISTINCT ISNULL(s.ymd,'')) ndays
            FROM nx.lg_sagub_actual s LEFT JOIN nx.item i ON i.item_code=s.item_code COLLATE DATABASE_DEFAULT
            WHERE {whs}
            GROUP BY s.item_code ORDER BY SUM(ISNULL(s.amt,0)) DESC, s.item_code""", *p)
        rows = []
        clsf = cls.strip()
        for r in cur.fetchall():
            pmin = float(r[4] or 0); pmax = float(r[5] or 0)
            cl = _cls_of(r[1])
            if clsf and cl != clsf:
                continue
            rows.append({"item": r[0], "name": r[1], "cls": cl,
                         "biz": ",".join(sorted(x for x in bizmap.get(r[0], set()) if x)),
                         "qty": float(r[2] or 0), "amt": float(r[3] or 0),
                         "pmin": pmin, "pmax": pmax, "cnt": r[6], "ndays": r[7],
                         "pchg": 1 if (pmin and pmax and abs(pmax - pmin) > 1e-6) else 0})
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()


@router.get("/api/lgsagub/detail")
def lgsagub_detail(item: str = Query(""), ym_from: str = Query(""), ym_to: str = Query(""),
                   ymd_from: str = Query(""), ymd_to: str = Query(""), biz: str = Query(""), limit: int = Query(3000)):
    """품번별 개별 사급 기록(일자·사업부·단가별). 가격 변동일자 확인용. 같은 일자·사업부·단가는 합산."""
    nx = _nx(); cur = nx.cursor()
    try:
        _prep(cur)
        if not item.strip():
            return {"rows": [], "cnt": 0}
        wh = ["s.item_code=?"]; p = [item.strip()]
        rwh, rp = _range_wh("", "", "", ymd_from, ymd_to); wh += rwh; p += rp
        if not (ymd_from.strip() or ymd_to.strip()):
            if ym_from.strip(): wh.append("s.ym>=?"); p.append(ym_from.strip())
            if ym_to.strip():   wh.append("s.ym<=?"); p.append(ym_to.strip())
        if biz.strip():     wh.append("ISNULL(s.biz,'')=?"); p.append(biz.strip())
        cur.execute(f"""SELECT TOP {int(limit)} ISNULL(s.ymd,'') ymd, ISNULL(s.biz,'') biz, s.price,
              SUM(ISNULL(s.qty,0)) qty, SUM(ISNULL(s.amt,0)) amt, COUNT(*) cnt
            FROM nx.lg_sagub_actual s
            WHERE {' AND '.join(wh)}
            GROUP BY ISNULL(s.ymd,''), ISNULL(s.biz,''), s.price
            ORDER BY ISNULL(s.ymd,''), ISNULL(s.biz,''), s.price""", *p)
        rows = [{"ymd": r[0], "biz": r[1], "price": float(r[2] or 0), "qty": float(r[3] or 0),
                 "amt": float(r[4] or 0), "cnt": r[5]} for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()


# ============================================================================
#  LG 정산 원단위(동정산) — 리시빙 비교용. Assy별 동부자재 소요(사급/직거래 구분).
#  ★중량 = (외경−T)·T·길이·π·8.94/1e6 × 수량  → 수량이 이미 반영됨(합산만, ×수량 금지).
#  ★사급=LG유상사급 인정동 / 직거래=우리 직매입(미인정, 설치동류).
# ============================================================================
_SETTLE_DDL = """IF OBJECT_ID('nx.lg_settle_unit') IS NULL
CREATE TABLE nx.lg_settle_unit(
  id int IDENTITY(1,1) PRIMARY KEY, ym varchar(4), coop nvarchar(60),
  assy_pn varchar(50), assy_desc nvarchar(200), sub_pn varchar(50), sub_desc nvarchar(200),
  qty float, gubun1 nvarchar(10), gubun2 nvarchar(30),
  od float, thk float, leng float, weight float, mat_cost float,
  eff_ym varchar(4), src_file nvarchar(200), upload_dt datetime DEFAULT getdate())"""
# eff_ym = Update 일정(품목 추가/변경 시점, YYMM). 유효일자 필터(리시빙월 ≤ eff_ym 제외)용.
_SETTLE_MIGRATE = "IF COL_LENGTH('nx.lg_settle_unit','eff_ym') IS NULL ALTER TABLE nx.lg_settle_unit ADD eff_ym varchar(4);"

def _settle_prep(cur):
    cur.execute(_SETTLE_DDL); cur.execute(_SETTLE_MIGRATE)


def _upd_ym(v):
    """Update 일정 문자열 → YYMM(효력월). '19.05월'→'1905', '16.9월->17.6월'→'1706'(최신),
       '17.11월(황동물 추가)'→'1711'. 빈값/'수정'→'' (기초행=항상 유효). 여러 날짜면 마지막 채택."""
    import re
    s = str(v or "")
    ms = re.findall(r"(\d{1,2})[.\-/](\d{1,2})", s)
    if not ms:
        return ""
    y, m = ms[-1]                      # 마지막 날짜(→ 이동 후 최신)
    y = int(y); m = int(m)
    if y >= 100: y = y % 100
    return f"{y:02d}{m:02d}"


def _pick_sheet(wb):
    """원단위 파일에서 피앤씨 탭 선택(이름에 '피앤씨' 포함, 공백무시). 없으면 첫 시트."""
    for sn in wb.sheetnames:
        if "피앤씨" in sn.replace(" ", ""):
            return wb[sn]
    return wb[wb.sheetnames[0]]


@router.post("/api/lgsagub/settle_upload")
async def settle_upload(file: UploadFile = File(...), ym: str = Query(""), sheet: str = Query("")):
    """LG 동정산 원단위 엑셀 업로드 → nx.lg_settle_unit. 피앤씨 탭의 (Assy×하위동부자재) 소요.
       컬럼: Assy P/N·Desc·협력사·P/N(하위1)·Desc·수량·구분1·구분2·외경·T·길이·중량(+소재비).
       ★중량은 수량 반영값이라 그대로 저장. ym=기준월(미지정시 파일에서 추론 안됨→필수 권장)."""
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else _pick_sheet(wb)
    rows_all = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_all:
        raise HTTPException(400, "빈 시트")

    def nrm(v):
        return "".join(str(v or "").strip().lower().split()).replace("\n", "")

    # 헤더행 탐지: 'assyp/n' 포함 행
    hi = -1
    for i in range(min(8, len(rows_all))):
        joined = [nrm(x) for x in rows_all[i]]
        if any("assyp/n" == c or "assyp/n" in c for c in joined):
            hi = i; break
    if hi < 0:
        raise HTTPException(400, "헤더행(Assy P/N) 감지 실패")
    hdr = [nrm(x) for x in rows_all[hi]]

    def find(*keys, after=None):
        start = (after + 1) if after is not None else 0
        for j in range(start, len(hdr)):
            if hdr[j] in keys:
                return j
        return None
    ci_assy = find("assyp/n")
    ci_assydesc = ci_assy + 1 if ci_assy is not None else None  # Assy 바로 뒤 Desc
    ci_coop = find("협력사")
    ci_sub = find("p/n(하위1)", "p/n하위1", "p/n(하위)", after=ci_coop)
    if ci_sub is None:
        ci_sub = find("p/n(하위1)", "p/n하위1")
    ci_subdesc = ci_sub + 1 if ci_sub is not None else None
    ci_qty = find("수량")
    ci_g1 = find("구분1")
    ci_g2 = find("구분2")
    ci_od = find("외경")
    ci_thk = find("t", "두께")
    ci_len = find("길이")
    ci_wt = find("중량")
    ci_mc = find("소재비변경후", "소재비")
    ci_upd = find("update일정", "update", "업데이트일정")
    if ci_assy is None or ci_g1 is None or ci_wt is None:
        return {"ok": False, "error": "필수컬럼(Assy P/N·구분1·중량) 감지 실패",
                "header": [str(x) for x in rows_all[hi]]}

    def gv(r, i):
        return r[i] if (i is not None and i < len(r)) else None
    def _f(r, i):
        v = gv(r, i)
        try:
            return float(str(v).replace(",", "")) if v not in (None, "") else 0.0
        except Exception:
            return 0.0

    recs = []
    for r in rows_all[hi + 1:]:
        assy = str(gv(r, ci_assy) or "").strip()
        if not assy or nrm(assy) in ("assyp/n", "합계", "total"):
            continue
        g1 = str(gv(r, ci_g1) or "").strip()
        if not g1:
            continue
        recs.append((ym.strip(), str(gv(r, ci_coop) or "")[:60], assy[:50],
                     str(gv(r, ci_assydesc) or "")[:200], str(gv(r, ci_sub) or "")[:50],
                     str(gv(r, ci_subdesc) or "")[:200], _f(r, ci_qty), g1[:10],
                     str(gv(r, ci_g2) or "")[:30], _f(r, ci_od), _f(r, ci_thk),
                     _f(r, ci_len), _f(r, ci_wt), _f(r, ci_mc), _upd_ym(gv(r, ci_upd))))
    if not recs:
        return {"ok": False, "error": "데이터 행 없음", "header": [str(x) for x in rows_all[hi]]}

    nx = _nx(); cur = nx.cursor()
    try:
        _settle_prep(cur)
        fn = (file.filename or "")[:200]
        # 재업로드=덮어쓰기: 같은 ym(+협력사)의 기존행 삭제. ym 없으면 파일명 기준.
        if ym.strip():
            coops = sorted({x[1] for x in recs if x[1]})
            cur.execute("DELETE FROM nx.lg_settle_unit WHERE ISNULL(ym,'')=?", ym.strip())
        else:
            cur.execute("DELETE FROM nx.lg_settle_unit WHERE src_file=?", fn)
        try: cur.fast_executemany = True
        except Exception: pass
        cols = "(ym,coop,assy_pn,assy_desc,sub_pn,sub_desc,qty,gubun1,gubun2,od,thk,leng,weight,mat_cost,eff_ym,src_file)"
        n = 0; BATCH = 120
        for i in range(0, len(recs), BATCH):
            chunk = recs[i:i + BATCH]
            vals = ",".join("(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)" for _ in chunk)
            flat = []
            for t in chunk:
                flat += list(t) + [fn]
            cur.execute(f"INSERT INTO nx.lg_settle_unit{cols} VALUES {vals}", *flat)
            n += len(chunk)
        nx.commit()
        cur.execute("""SELECT gubun1, COUNT(*), SUM(ISNULL(weight,0)) FROM nx.lg_settle_unit
                       WHERE src_file=? GROUP BY gubun1""", fn)
        by_g = [{"gubun1": r[0], "rows": r[1], "weight": float(r[2] or 0)} for r in cur.fetchall()]
        return {"ok": True, "file": fn, "sheet": ws.title, "ym": ym.strip(), "rows": n, "by_gubun1": by_g}
    finally:
        nx.close()


@router.get("/api/lgsagub/settle_summary")
def settle_summary():
    """업로드된 원단위 요약(기준월·협력사·구분1별)."""
    nx = _nx(); cur = nx.cursor()
    try:
        _settle_prep(cur)
        cur.execute("""SELECT ym, COUNT(DISTINCT assy_pn) assys, COUNT(*) rows,
              SUM(CASE WHEN gubun1=N'사급' THEN weight ELSE 0 END) sg_wt,
              SUM(CASE WHEN gubun1=N'직거래' THEN weight ELSE 0 END) jk_wt, MAX(upload_dt) dt
            FROM nx.lg_settle_unit GROUP BY ym ORDER BY ym""")
        by_ym = [{"ym": r[0], "assys": r[1], "rows": r[2], "sagub_wt": float(r[3] or 0),
                  "jikgae_wt": float(r[4] or 0), "dt": str(r[5])[:19]} for r in cur.fetchall()]
        return {"by_ym": by_ym}
    finally:
        nx.close()


@router.get("/api/lgsagub/recvcompare")
def recvcompare(ym: str = Query(""), ymd_from: str = Query(""), ymd_to: str = Query(""), limit: int = Query(2000)):
    """리시빙 비교(원소재 동): 기간 LG리시빙 품목별 동소요를 두 축으로 비교 + IN OSP 대사.
       (1)우리 BOM 기준 = nx BOM 전개(_dong_of, LG 미인정분 포함)  (2)LG BOM 기준 = 사급(Assembly Pull)만.
       우리<LG 품목 = 우리 BOM이 동을 덜 잡은 정교화 대상. 기간(ymd_from~ymd_to) 우선, 없으면 월(ym)."""
    f = lambda v: float(v or 0)
    nx = _nx(); cur = nx.cursor()
    try:
        rwh, rp, yms = _recv_where(ym, ymd_from, ymd_to)
        # 유효일자 컷오프 = 리시빙 최종월(as-of 단가/버전 선택용).
        eff_cut = (max(yms) if yms else (ym.strip() or (ymd_to.strip()[:4] if len(ymd_to.strip()) >= 4 else "2608")))
        # ★신규 사급가(절삭재료비 as-of): 2월~8/6 인상전(202605) / 8/7~ 인상후(202608). 리시빙 일자로 단가 갈림.
        pm_pre = _matcost_asof('202606')      # 인상전(≤8/6) = 202605
        pm_post = _matcost_asof('202612')     # 인상후(8/7~) = 202608
        _CUT = '260807'                        # 인상후 적용 시작일(YYMMDD)
        # ★BOM기준 동소요 = LG BOM 사급(Assembly Pull) 동 point-in-time(nx.lg_bom_ver). eff_cut(YYMM) 시점 버전.
        ap_ver = (f"20{ymd_to.strip()[:2]}-{ymd_to.strip()[2:4]}-{ymd_to.strip()[4:6]}" if len(ymd_to.strip()) >= 6
                  else f"20{eff_cut[:2]}-{eff_cut[2:4]}-28")
        # ★B: 전체 사급 동소요 = LG BOM AP(우리가 협력사에 사급 주는 소재 포함). 분할 = 우리 직접절삭 + 협력사 사급분(2중계상 없음).
        #   split은 리시빙 품목만 전개(성능) — recvrows 확정 후 아래에서 계산.
        split = {}; our_map = {}     # recvrows 확정 후 배치로딩(성능)
        #   규격키 (metal,diam,thick) 동일, 금액 = 절삭재료비 as-of 단가(pm, 없으면 인상후 fallback).
        def _kv(spec, pm):                     # {(metal,diam,thick): 중량} → (중량합, 금액합)
            kg = amt = 0.0
            for sp, w in spec.items():
                kg += w
                amt += w * (pm.get(sp) or pm_post.get(sp) or 0.0)
            return kg, amt
        def _actual_kv(it, pm):                # (참고) 우리 실측 중량(bom_flat 제작동관·배치) — 정산차액용
            return _kv(our_map.get(it, {}), pm)
        def _split_kv(it, pm):                 # LG BOM 분할: (우리절삭kg/amt, 협력사사급kg/amt)
            d = split.get(it, {"our": {}, "coop": {}})
            ok, oa = _kv(d.get("our", {}), pm); ck, ca = _kv(d.get("coop", {}), pm)
            return ok, oa, ck, ca

        cur.execute(f"""SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))) it,
              CASE WHEN RECEIVING_YMD >= '{_CUT}' THEN 1 ELSE 0 END post,
              SUM(CASE WHEN GUBUN='C' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END) qc,
              SUM(CASE WHEN GUBUN='R' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END) qr,
              SUM(CASE WHEN GUBUN='C' THEN ISNULL(RECV_AMT,0) ELSE 0 END) ac
            FROM nx.SA_T_LG_RECEIVING_DTL WHERE {rwh}
            GROUP BY UPPER(LTRIM(RTRIM(ITEM_CODE))), CASE WHEN RECEIVING_YMD >= '{_CUT}' THEN 1 ELSE 0 END""", *rp)
        recvrows = [(r[0], int(r[1] or 0), f(r[2]), f(r[3]), f(r[4])) for r in cur.fetchall()]
        # ★절삭만 = LG 사급(동 정합 대상). 설치·이지링크·미분류 = 직거래(PNC 자체구매) → 제외 (nx.item.cut_gubun)
        cur.execute("SELECT UPPER(LTRIM(RTRIM(item_code))) FROM nx.item WHERE cut_gubun=N'절삭'")
        cutset = set(r[0] for r in cur.fetchall())
        recvrows = [row for row in recvrows if row[0] in cutset]
        # ★성능: 리시빙에 실제 등장한 완제품만 LG BOM 전개 + 우리 실측 배치로딩(개별 N쿼리 회피)
        _rec_models = set(row[0] for row in recvrows)
        split = _lg_ap_split(cur, ap_ver, models=_rec_models)
        our_map = _dong_of_batch(cur, _rec_models)
        # 품명 매핑
        nm = {}
        cur.execute("SELECT UPPER(LTRIM(RTRIM(item_code))), MAX(item_name) FROM nx.item GROUP BY UPPER(LTRIM(RTRIM(item_code)))")
        for a, b in cur.fetchall():
            nm[a] = b
        agg = {}; total_qc = 0.0
        for it, post, qc, qr, ac in recvrows:
            total_qc += qc
            pm = pm_post if post else pm_pre
            net = qc - qr
            ok, oa, ck, ca = _split_kv(it, pm)              # LG BOM 분할: 우리절삭 / 협력사사급 /개
            au_kg, au_amt = _actual_kv(it, pm)              # (참고) 우리 실측 중량 /개
            has = (ok + ck) > 0                              # LG인정 = LG BOM 사급(AP) 동 보유
            d = agg.get(it)
            if d is None:
                d = agg[it] = {"item": it, "name": nm.get(it, ""), "recv_c": 0.0, "recv_r": 0.0, "recv_amt": 0.0,
                               "matched": 1 if has else 0,
                               "ourcut_kg": 0.0, "ourcut_amt": 0.0, "coop_kg": 0.0, "coop_amt": 0.0, "actual_kg": 0.0}
            d["recv_c"] += qc; d["recv_r"] += qr; d["recv_amt"] += ac
            d["ourcut_kg"] += net * ok; d["ourcut_amt"] += net * oa
            d["coop_kg"] += net * ck; d["coop_amt"] += net * ca
            d["actual_kg"] += net * au_kg
        items = list(agg.values())
        for d in items:                                     # 전체 = 우리절삭 + 협력사사급 (2중계상 없음)
            d["total_kg"] = d["ourcut_kg"] + d["coop_kg"]
            d["total_amt"] = d["ourcut_amt"] + d["coop_amt"]
        matched_qc = sum(d["recv_c"] for d in items if d["matched"])
        unmatched = sum(1 for d in items if not d["matched"])
        TOT_KG = sum(d["total_kg"] for d in items); TOT_AMT = sum(d["total_amt"] for d in items)
        OURCUT_KG = sum(d["ourcut_kg"] for d in items); COOP_KG = sum(d["coop_kg"] for d in items)
        ACTUAL_KG = sum(d["actual_kg"] for d in items)
        coop_items = sum(1 for d in items if d["coop_kg"] > 1e-6)  # 협력사 사급분 있는 품목수
        # IN OSP(사급입고) — 원소재/사급부품
        inl = ",".join("'" + y + "'" for y in yms) if yms else "''"
        cur.execute(f"""SELECT CASE WHEN UPPER(item_name) LIKE '%TUBE%' THEN N'원소재' ELSE N'사급부품' END cl,
              SUM(ISNULL(qty,0)) q, SUM(ISNULL(amt,0)) a FROM nx.lg_sagub_actual WHERE ym IN ({inl}) GROUP BY
              CASE WHEN UPPER(item_name) LIKE '%TUBE%' THEN N'원소재' ELSE N'사급부품' END""")
        osp = {r[0]: {"qty": f(r[1]), "amt": f(r[2])} for r in cur.fetchall()}
        in_raw = osp.get("원소재", {"qty": 0, "amt": 0})
        price = (in_raw["amt"] / in_raw["qty"]) if in_raw["qty"] else 0.0   # OSP 원소재 평균단가(참고)
        items.sort(key=lambda x: -x["total_kg"])
        return {
            "ym": ym.strip(),
            "copper": {
                # 전체 사급 동소요(LG BOM AP·우리가 협력사 사급 포함) = 우리절삭 + 협력사사급
                "total_net": TOT_KG, "total_net_amt": TOT_AMT,
                "ourcut_net": OURCUT_KG, "coop_net": COOP_KG,
                "actual_net": ACTUAL_KG,     # (참고) 우리 실측 중량(bom_flat)
                "in_osp_kg": in_raw["qty"], "in_osp_amt": in_raw["amt"], "osp_price": price,
            },
            "parts_in": osp.get("사급부품", {"qty": 0, "amt": 0}),
            "coverage": {"matched_qty": matched_qc, "total_qty": total_qc, "unmatched_items": unmatched,
                         "coop_items": coop_items,
                         "rate": (matched_qc / total_qc * 100) if total_qc else 0},
            "items": items[:int(limit)],
        }
    finally:
        nx.close()


@router.get("/api/lgsagub/recvcompare_ledger")
def recvcompare_ledger(from_ym: str = Query(""), to_ym: str = Query("")):
    """동 원소재 수불(월별 누적): 기초 + 입고(OSP TUBE) − 소요 = 기말. 소요 2축(우리 BOM / LG BOM 사급 AP).
       from_ym(기초0 시작월, 미지정=OSP 첫 입고월)~to_ym(미지정=OSP 최신월). LG BOM은 월별 point-in-time 버전.
       ★기초=0 가정: 시작월 이전 동재고 없음. OSP 데이터 없는 달로 시작하면 입고0→마이너스 되므로 첫 OSP월 권장."""
    f = lambda v: float(v or 0)
    nx = _nx(); cur = nx.cursor()
    try:
        # ★사급 원소재 수불 개시월 = 2607(2026.07) 사용자 확정(LG BOM 버전 baseline=2026-07-01과 정합). 기초0. to_ym 기본 = OSP 최신월.
        cur.execute("SELECT MIN(ym), MAX(ym) FROM nx.lg_sagub_actual WHERE UPPER(item_name) LIKE '%TUBE%'")
        r0 = cur.fetchone(); osp_min = (r0[0] if r0 and r0[0] else "") or ""; osp_max = (r0[1] if r0 and r0[1] else "") or ""
        LEDGER_START = "2607"
        frm = from_ym.strip() or LEDGER_START
        if osp_min and frm < osp_min:    # OSP 데이터 없는 이전달로 시작하면 입고0→마이너스 → 첫 OSP월로 클램프
            frm = osp_min
        to = to_ym.strip() or osp_max or frm

        def ym_next(y):
            yy = int(y[:2]); mm = int(y[2:]) + 1
            if mm > 12: mm = 1; yy += 1
            return f"{yy:02d}{mm:02d}"
        months = []; m = frm; guard = 0
        while m <= to and guard < 120:
            months.append(m); m = ym_next(m); guard += 1

        # ★두 소요 기준(수불): (LG BOM기준)=LG BOM AP 전체(우리절삭LG+협력사)  (우리 BOM기준)=우리 실측절삭(bom_flat)+협력사사급. 협력사분 공통.
        pm_pre = _matcost_asof('202606'); pm_post = _matcost_asof('202612')
        def _mdate(M):    # YYMM → 그 달 말(28일) date: point-in-time 버전 선택(ver_from<=)
            return f"20{M[:2]}-{M[2:4]}-28"
        cur.execute("SELECT UPPER(LTRIM(RTRIM(item_code))) FROM nx.item WHERE cut_gubun=N'절삭'")
        cutset = set(r[0] for r in cur.fetchall())   # ★절삭만 = LG 사급, 설치/이지링크=직거래 제외
        # ★성능: 우리 실측(bom_flat) + 제작동관셋 1회 배치로딩(월별 재쿼리 회피)
        cur.execute("SELECT DISTINCT UPPER(LTRIM(RTRIM(ITEM_CODE))) FROM nx.SA_T_LG_RECEIVING_DTL WHERE LEFT(RECEIVING_YMD,4)>=? AND LEFT(RECEIVING_YMD,4)<=?", frm, to)
        allitems = set(r[0] for r in cur.fetchall()) & cutset
        our_map = _dong_of_batch(cur, allitems)
        cur.execute("SELECT DISTINCT UPPER(LTRIM(RTRIM(leaf_code))) FROM nx.bom_flat WHERE role=N'제작동관'")
        jjset = set(r[0] for r in cur.fetchall())
        rows = []; bal_our_kg = 0.0; bal_our_amt = 0.0; bal_bom_kg = 0.0; bal_bom_amt = 0.0
        for M in months:
            cur.execute("""SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))) it,
                  SUM(CASE WHEN GUBUN='C' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END)
                 -SUM(CASE WHEN GUBUN='R' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END) net
                FROM nx.SA_T_LG_RECEIVING_DTL WHERE LEFT(RECEIVING_YMD,4)=?
                GROUP BY UPPER(LTRIM(RTRIM(ITEM_CODE)))""", M)
            recvlist = [(r[0], f(r[1])) for r in cur.fetchall()]
            recvlist = [(it, net) for it, net in recvlist if it in cutset]   # ★절삭만(LG사급)
            pmM = pm_post if M >= '2608' else pm_pre        # 인상후(8월~) / 인상전
            spM = _lg_ap_split(cur, _mdate(M), models=set(it for it, _ in recvlist), jjset=jjset)  # 우리절삭/협력사 분할
            soyo_our_kg = soyo_our_amt = 0.0; soyo_bom_kg = soyo_bom_amt = 0.0
            for it, net in recvlist:
                sd = spM.get(it, {"our": {}, "coop": {}})
                for sp, w in sd.get("our", {}).items():         # LG기준 우리절삭(LG인증값)
                    cva = w * net * (pmM.get(sp) or pm_post.get(sp) or 0.0)
                    soyo_bom_kg += w * net; soyo_bom_amt += cva
                for sp, w in sd.get("coop", {}).items():        # 협력사 사급분(양 기준 공통)
                    cvv = w * net; cva = w * net * (pmM.get(sp) or pm_post.get(sp) or 0.0)
                    soyo_bom_kg += cvv; soyo_bom_amt += cva
                    soyo_our_kg += cvv; soyo_our_amt += cva
                for sp, w in our_map.get(it, {}).items():       # 우리기준 우리절삭(우리 실측값)
                    soyo_our_kg += w * net; soyo_our_amt += w * net * (pmM.get(sp) or pm_post.get(sp) or 0.0)
            cur.execute("""SELECT SUM(ISNULL(qty,0)), SUM(ISNULL(amt,0)) FROM nx.lg_sagub_actual
                           WHERE ym=? AND UPPER(item_name) LIKE '%TUBE%'""", M)
            r = cur.fetchone(); in_kg = f(r[0]); in_amt = f(r[1])
            open_our_kg = bal_our_kg; open_our_amt = bal_our_amt
            bal_our_kg = open_our_kg + in_kg - soyo_our_kg
            bal_our_amt = open_our_amt + in_amt - soyo_our_amt
            open_bom_kg = bal_bom_kg; open_bom_amt = bal_bom_amt
            bal_bom_kg = open_bom_kg + in_kg - soyo_bom_kg
            bal_bom_amt = open_bom_amt + in_amt - soyo_bom_amt
            rows.append({"ym": M, "in_kg": in_kg, "in_amt": in_amt,
                         "open_our_kg": open_our_kg, "soyo_our_kg": soyo_our_kg, "close_our_kg": bal_our_kg,
                         "soyo_our_amt": soyo_our_amt, "close_our_amt": bal_our_amt,
                         "open_bom_kg": open_bom_kg, "soyo_bom_kg": soyo_bom_kg, "close_bom_kg": bal_bom_kg,
                         "soyo_bom_amt": soyo_bom_amt, "close_bom_amt": bal_bom_amt})
        return {"from_ym": frm, "to_ym": to, "osp_min": osp_min, "rows": rows}
    finally:
        nx.close()


# ── 사급부품(소분류310) BOM 전개 캐시 ──
_PARTS_MAPS = None
def _parts_maps(cur):
    """CS_M_ITEM_BOM 부모→자식 맵 + 사급부품(SGROUP 310) 집합. 모듈캐시(1회 로드).
       ★CS_CALC_EXCEPT_FLAG<>'1' 필터 유지: 이 플래그는 변형SUB 이중계상 방지용.
         예) AJR30077403은 MJX65072203을 (a)직접행(except='1') + (b)-F&T 변형서브 안 = 2경로로 가짐.
         except 필터가 (a)를 걸러 1회만 계상. 제거하면 2배 이중계상됨(실측 확인)."""
    global _PARTS_MAPS
    if _PARTS_MAPS is not None:
        return _PARTS_MAPS
    cur.execute("""SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))), UPPER(LTRIM(RTRIM(MAT_CODE))), ISNULL(USE_QTY,0)
                   FROM nx.CS_M_ITEM_BOM WHERE ISNULL(CS_CALC_EXCEPT_FLAG,'0')<>'1'""")
    ch = {}
    for p, c2, q in cur.fetchall():
        ch.setdefault(p, []).append((c2, float(q or 0)))
    cur.execute("SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))) FROM nx.PR_M_ITEM WHERE LTRIM(RTRIM(ITEM_SGROUP))='310'")
    sg310 = set(r[0] for r in cur.fetchall())
    _PARTS_MAPS = (ch, sg310)
    return _PARTS_MAPS

def _explode_parts(item, ch, sg310, memo):
    """1개 완제품 → {사급부품(310): 소요개수}. 310 도달시 계상 후 정지(LG가 완성제공)."""
    if item in memo:
        return memo[item]
    memo[item] = {}
    acc = {}
    for c2, q in ch.get(item, []):
        if q <= 0:
            continue
        if c2 in sg310:
            acc[c2] = acc.get(c2, 0.0) + q
        else:
            for k, v in _explode_parts(c2, ch, sg310, memo).items():
                acc[k] = acc.get(k, 0.0) + v * q
    memo[item] = acc
    return acc


@router.get("/api/lgsagub/settle_list")
def settle_list(ym: str = Query(""), q: str = Query(""), gubun1: str = Query(""), limit: int = Query(20000)):
    """원단위 관리 목록: 적용월(ym)의 (Assy×하위동부자재) 행 조회 + 월목록(드롭다운용)."""
    f = lambda v: float(v or 0)
    nx = _nx(); cur = nx.cursor()
    try:
        _settle_prep(cur)
        cur.execute("SELECT ym, COUNT(*) FROM nx.lg_settle_unit GROUP BY ym ORDER BY ym DESC")
        yms = [{"ym": r[0], "rows": r[1]} for r in cur.fetchall()]
        if not ym.strip() and yms:
            ym = yms[0]["ym"]
        wh = ["ym=?"]; p = [ym.strip()]
        if gubun1.strip():
            wh.append("gubun1=?"); p.append(gubun1.strip())
        if q.strip():
            wh.append("(assy_pn LIKE ? OR sub_pn LIKE ? OR assy_desc LIKE ? OR sub_desc LIKE ?)")
            p += [f"%{q.strip()}%"] * 4
        cur.execute(f"""SELECT TOP {int(limit)} assy_pn, assy_desc, coop, sub_pn, sub_desc, qty,
              gubun1, gubun2, od, thk, leng, weight
            FROM nx.lg_settle_unit WHERE {' AND '.join(wh)}
            ORDER BY assy_pn, sub_pn""", *p)
        rows = [{"assy_pn": r[0], "assy_desc": r[1], "coop": r[2], "sub_pn": r[3], "sub_desc": r[4],
                 "qty": f(r[5]), "gubun1": r[6], "gubun2": r[7], "od": f(r[8]), "thk": f(r[9]),
                 "leng": f(r[10]), "weight": f(r[11])} for r in cur.fetchall()]
        return {"ym": ym.strip(), "yms": yms, "rows": rows, "cnt": len(rows)}
    finally:
        nx.close()


@router.post("/api/lgsagub/settle_copy")
def settle_copy(from_ym: str = Query(...), to_ym: str = Query(...)):
    """전월 복사로 신규월: from_ym 원단위를 to_ym로 복제(기존 to_ym 삭제 후). 상대적 불변이라 복사후 부분수정용."""
    nx = _nx(); cur = nx.cursor()
    try:
        _settle_prep(cur)
        cur.execute("DELETE FROM nx.lg_settle_unit WHERE ym=?", to_ym.strip())
        cur.execute("""INSERT INTO nx.lg_settle_unit
              (ym,coop,assy_pn,assy_desc,sub_pn,sub_desc,qty,gubun1,gubun2,od,thk,leng,weight,mat_cost,src_file)
            SELECT ?, coop,assy_pn,assy_desc,sub_pn,sub_desc,qty,gubun1,gubun2,od,thk,leng,weight,mat_cost,
              N'copy<'+ym+N'>' FROM nx.lg_settle_unit WHERE ym=?""", to_ym.strip(), from_ym.strip())
        n = cur.rowcount
        nx.commit()
        return {"ok": True, "from_ym": from_ym.strip(), "to_ym": to_ym.strip(), "rows": n}
    finally:
        nx.close()


def _ym_list(a, b):
    """YYMMDD 범위 → 걸치는 YYMM 리스트. 둘 다 없으면 []."""
    if not a and not b:
        return []
    a = a or b; b = b or a
    try:
        ya, ma = int(a[:2]), int(a[2:4]); yb, mb = int(b[:2]), int(b[2:4])
    except Exception:
        return []
    out = []; y, m = ya, ma
    while (y, m) <= (yb, mb) and len(out) < 60:
        out.append(f"{y:02d}{m:02d}")
        m += 1
        if m > 12: m = 1; y += 1
    return out

def _recv_where(ym, ymd_from, ymd_to):
    """리시빙 필터: 기간(ymd_from~ymd_to YYMMDD) 우선, 없으면 월(ym=YYMM=LEFT4). (where, params, yms)."""
    if ymd_from.strip() or ymd_to.strip():
        return ("RECEIVING_YMD>=? AND RECEIVING_YMD<=?",
                [ymd_from.strip() or "000000", ymd_to.strip() or "999999"],
                _ym_list(ymd_from.strip(), ymd_to.strip()))
    return ("LEFT(RECEIVING_YMD,4)=?", [ym.strip()], [ym.strip()] if ym.strip() else [])


@router.get("/api/lgsagub/recvcompare_parts")
def recvcompare_parts(ym: str = Query(""), ymd_from: str = Query(""), ymd_to: str = Query(""), limit: int = Query(3000)):
    """리시빙 비교(부품): 기간 리시빙 × BOM 사급부품(310) 소요개수 = OUT vs OSP 사급부품입고 IN.
       ★부품=개수 단위. 소비=C(정상)만(R은 반품 아님→차감안함). BOM=CS_M_ITEM_BOM 전체(원가제외필터 미적용).
       기간(ymd_from~ymd_to) 우선, 없으면 월(ym)."""
    f = lambda v: float(v or 0)
    nx = _nx(); cur = nx.cursor()
    try:
        _prep(cur)
        ch, sg310 = _parts_maps(cur)
        rwh, rp, yms = _recv_where(ym, ymd_from, ymd_to)
        # ② IN OSP 사급부품 (lg_sagub, 품명 TUBE 아님) — 기간내 월합. ★이 OSP 목록 자체가 '사급부품 정의' = 전개 정지점.
        inl = ",".join("'" + y + "'" for y in yms) if yms else "''"
        cur.execute(f"""SELECT UPPER(LTRIM(RTRIM(item_code))) it, MAX(item_name) nm,
              SUM(ISNULL(qty,0)) q, SUM(ISNULL(amt,0)) a,
              SUM(ISNULL(amt,0))/NULLIF(SUM(ISNULL(qty,0)),0) p
            FROM nx.lg_sagub_actual WHERE ym IN ({inl}) AND UPPER(item_name) NOT LIKE '%TUBE%'
            GROUP BY UPPER(LTRIM(RTRIM(item_code)))""")
        in_map = {r[0]: {"nm": r[1], "q": f(r[2]), "a": f(r[3]), "p": f(r[4])} for r in cur.fetchall()}
        # ★전개 정지=OSP 목록(SGROUP=310 아님). 310으로만 멈추면 MAZ30083301(sg230, 구매단가)처럼
        #   310 아닌 사급부품을 통째로 놓침. OSP에 있는 부품이면 어디서든 정지·계상.
        osp_set = set(in_map)
        # ③ 리시빙 → 사급부품 소요개수 (C+R 전부. R은 반품 아니라 정상 리시빙 다른 구분)
        cur.execute(f"""SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))) it,
              SUM(CASE WHEN GUBUN='C' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END) qc,
              SUM(CASE WHEN GUBUN='R' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END) qr
            FROM nx.SA_T_LG_RECEIVING_DTL WHERE {rwh} GROUP BY UPPER(LTRIM(RTRIM(ITEM_CODE)))""", *rp)
        recv = [(r[0], f(r[1]), f(r[2])) for r in cur.fetchall()]
        memo = {}
        out_c = {}; out_r = {}   # 사급부품별 OUT 소요개수
        for it, qc, qr in recv:
            pmap = _explode_parts(it, ch, osp_set, memo)   # ★정지=OSP
            for part, per in pmap.items():
                out_c[part] = out_c.get(part, 0.0) + qc * per
                out_r[part] = out_r.get(part, 0.0) + qr * per
        # ★① 우리 ERP 확정입고(입고기준): 확정입고집계표와 동일 원천 = PU_T_STOCK_MAINT(9/S/C/G/H 검사통과)+_C(수입 DIVISION=P). MAT_CODE별 기간합.
        #   ★2026-08-21 수정: 라이브 dbo 직독. 웹 자재입고관리(stock.py)가 nx.PU_T_STOCK_MAINT(미러 테이블)에 직접 등록 →
        #   레거시 미러분과 이중계상(EBD64385805 300→600). 라이브 원본이 확정입고 정본이므로 nx 대신 라이브 읽음.
        erp_map = {}
        if yms:
            cn2 = _conn(); cur2 = cn2.cursor()
            try:
                cur2.execute(f"""SELECT UPPER(LTRIM(RTRIM(mat))) it, SUM(qty) q FROM (
                      SELECT MAT_CODE mat, CONVERT(float,ISNULL(MAINT_QTY,0)) qty FROM dbo.PU_T_STOCK_MAINT
                        WHERE LEFT(MAINT_YMD,4) IN ({inl}) AND MAINT_TAG IN ('9','S','C','G','H')
                          AND ((ISNULL(INSP_FLAG,'N') IN ('','N')) OR (ISNULL(INSP_FLAG,'N') IN ('S','F') AND INSP_PROC_YMD >= ''))
                      UNION ALL
                      SELECT MAT_CODE, CONVERT(float,ISNULL(MAINT_QTY,0)) FROM dbo.PU_T_STOCK_MAINT_C
                        WHERE LEFT(MAINT_YMD,4) IN ({inl}) AND DIVISION='P'
                    ) t GROUP BY UPPER(LTRIM(RTRIM(mat)))""")
                for r in cur2.fetchall():
                    erp_map[r[0]] = f(r[1])
            finally:
                cn2.close()
        # ★OSP에 나오는 부품(=LG 사급 목록)만 대상. 품명은 OSP(in_map)에 이미 있음(nx.item 전체조회 제거=속도).
        parts = set(in_map)
        items = []
        tot_out = tot_in_q = tot_in_a = tot_erp = 0.0
        for p in parts:
            # ★소비 = C+R 전부. GUBUN='R'은 반품이 아니라 정상 리시빙의 다른 구분(R전용 제품 다수)이라
            #   C만 세면 그 제품들의 부품이 통째로 0이 됨(EAP65270720 등). C+R로 ③/②=0.72→0.97 검증.
            oc = out_c.get(p, 0.0) + out_r.get(p, 0.0)
            ind = in_map.get(p, {})
            iq = ind.get("q", 0.0); ia = ind.get("a", 0.0); price = ind.get("p", 0.0)
            ei = erp_map.get(p, 0.0)   # ① 우리 ERP 확정입고
            tot_out += oc; tot_in_q += iq; tot_in_a += ia; tot_erp += ei
            # ★3-way: ①우리ERP입고(erp_in) vs ②LG OSP(in_qty) vs ③LG리시빙소비(out_net=C+R).
            #   ①≈② 정상(둘 다 공급)·③≈② 정상(넓은 기간)·diff_erp=②−①(기록불일치)·diff=②−③(선입고).
            items.append({"item": p, "name": ind.get("nm") or "",
                          "erp_in": ei, "out_net": oc,
                          "in_qty": iq, "in_amt": ia, "price": price,
                          "diff": iq - oc, "diff_erp": iq - ei})
        items.sort(key=lambda x: -(x["in_qty"] + x["erp_in"] + abs(x["out_net"])))
        return {
            "ym": ym.strip(),
            "summary": {"erp_in": tot_erp, "out_net": tot_out,
                        "in_qty": tot_in_q, "in_amt": tot_in_a, "parts": len(parts)},
            "items": items[:int(limit)],
        }
    finally:
        nx.close()


@router.get("/api/lgsagub/recvcompare_parts_ledger")
def recvcompare_parts_ledger(from_ym: str = Query(""), to_ym: str = Query("")):
    """사급부품 월별 수불(원소재 수불과 동일 형태): 기초 + 입고(OSP 사급부품) − 소요(리시빙×BOM 부품) = 기말. 개수 단위·1월(2601)부터.
       입고=nx.lg_sagub_actual(NOT TUBE) 월합. 소요=리시빙(C+R)×_explode_parts(정지=OSP 사급부품). 금액=개수×전기간 평균단가."""
    f = lambda v: float(v or 0)
    nx = _nx(); cur = nx.cursor()
    try:
        _prep(cur)
        ch, sg310 = _parts_maps(cur)
        # 사급부품 정의(전개 정지점) + 평균단가 = 전기간 OSP(NOT TUBE)
        cur.execute("""SELECT UPPER(LTRIM(RTRIM(item_code))), SUM(ISNULL(qty,0)), SUM(ISNULL(amt,0))
                       FROM nx.lg_sagub_actual WHERE UPPER(item_name) NOT LIKE '%TUBE%'
                       GROUP BY UPPER(LTRIM(RTRIM(item_code)))""")
        osp_all = {r[0]: (f(r[1]), f(r[2])) for r in cur.fetchall()}
        osp_set = set(osp_all)
        price = {k: (a / q if q else 0.0) for k, (q, a) in osp_all.items()}
        cur.execute("SELECT MIN(ym), MAX(ym) FROM nx.lg_sagub_actual WHERE UPPER(item_name) NOT LIKE '%TUBE%'")
        r0 = cur.fetchone(); osp_min = (r0[0] if r0 and r0[0] else "") or ""; osp_max = (r0[1] if r0 and r0[1] else "") or ""
        LEDGER_START = "2601"                 # ★사용자: 1월부터
        frm = from_ym.strip() or LEDGER_START
        to = to_ym.strip() or osp_max or frm

        def ym_next(y):
            yy = int(y[:2]); mm = int(y[2:]) + 1
            if mm > 12: mm = 1; yy += 1
            return f"{yy:02d}{mm:02d}"
        months = []; m = frm; guard = 0
        while m <= to and guard < 120:
            months.append(m); m = ym_next(m); guard += 1

        memo = {}
        rows = []; bal_q = 0.0; bal_a = 0.0
        for M in months:
            cur.execute("""SELECT SUM(ISNULL(qty,0)), SUM(ISNULL(amt,0)) FROM nx.lg_sagub_actual
                           WHERE ym=? AND UPPER(item_name) NOT LIKE '%TUBE%'""", M)
            r = cur.fetchone(); in_q = f(r[0]); in_a = f(r[1])
            cur.execute("""SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))) it,
                  SUM(CASE WHEN GUBUN='C' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END)
                 +SUM(CASE WHEN GUBUN='R' THEN CONVERT(float,ISNULL(RECV_QTY,0)) ELSE 0 END) qty
                FROM nx.SA_T_LG_RECEIVING_DTL WHERE LEFT(RECEIVING_YMD,4)=?
                GROUP BY UPPER(LTRIM(RTRIM(ITEM_CODE)))""", M)
            out_q = out_a = 0.0
            for it, qty in [(r[0], f(r[1])) for r in cur.fetchall()]:
                for part, per in _explode_parts(it, ch, osp_set, memo).items():
                    out_q += qty * per
                    out_a += qty * per * price.get(part, 0.0)
            open_q = bal_q; open_a = bal_a
            bal_q = open_q + in_q - out_q
            bal_a = open_a + in_a - out_a
            rows.append({"ym": M, "in_kg": in_q, "in_amt": in_a,
                         "open_bom_kg": open_q, "soyo_bom_kg": out_q, "close_bom_kg": bal_q,
                         "soyo_bom_amt": out_a, "close_bom_amt": bal_a})
        return {"from_ym": frm, "to_ym": to, "osp_min": osp_min, "rows": rows}
    finally:
        nx.close()


# ===================== 원소재 사급전환율 (LG BOM Assembly Pull 대조) =====================
def _parse_cu_spec(spec):
    """동 원소재 규격 파싱: 'CUTTING CU P9.52 T0.7 L/W C1220T-O ALL' → 외경/두께/재질/형태."""
    s = str(spec or "")
    od = thk = metal = form = ""
    m = _re.search(r'P(\d+(?:\.\d+)?)', s)
    if m: od = m.group(1)
    m = _re.search(r'T(\d+(?:\.\d+)?)', s)
    if m: thk = m.group(1)
    m = _re.search(r'C\d{3,4}[A-Z]?(?:-[A-Z])?', s)      # C1220T-O, C1220C-O
    if m: metal = m.group(0)
    elif "COPPER" in s.upper(): metal = "Copper Alloy"
    if "L/W" in s: form = "L/W"           # Level Wound(코일)
    elif "S/L" in s: form = "S/L"         # Straight Length(직관)
    return od, thk, metal, form


def _num(v):
    try:
        f = float(v)
        return f if f > 0 else None
    except Exception:
        return None


@router.get("/api/lgsagub/sagub_convert")
def lgsagub_sagub_convert(werks: str = Query(""), status: str = Query("supplier"),
                          mt: str = Query("1,2,5"), scope: str = Query("all"),
                          cutg: str = Query("절삭"), q: str = Query(""), limit: int = Query(6000)):
    """원소재 사급전환율: LG BOM의 동 원소재(child_desc='Tube,Raw')가 사급(Assembly Pull)으로
       전환됐는지 대조. Supplier=미전환(우리가 구매)·Assembly Pull=전환(LG 사급).
       기본=미전환(Supplier)·제작품(parent) 제작유형 1/2/5.
       ★치수·재질은 우리 정본 nx.item(diam/thick/length/net_weight/metal_gubun) 우선,
         없으면 LG child_spec 파싱(BOM 정본: LG 치수·bom_dim.fin_weight는 부정확)."""
    nx = _nx(); cur = nx.cursor()
    try:
        # 전환율 요약 (전체 Tube,Raw edge 기준)
        cur.execute("SELECT supply_type, COUNT(*) FROM nx.lg_bom WHERE child_desc='Tube,Raw' GROUP BY supply_type")
        sm = {r[0]: r[1] for r in cur.fetchall()}
        pull = sm.get("Assembly Pull", 0); sup = sm.get("Supplier", 0)
        rate = round(pull * 100.0 / (pull + sup), 1) if (pull + sup) else 0.0

        wh = ["r.child_desc='Tube,Raw'"]; p = []
        if status == "supplier":  wh.append("r.supply_type='Supplier'")
        elif status == "pull":    wh.append("r.supply_type='Assembly Pull'")
        # status=all → 전체
        if werks.strip(): wh.append("r.werks=?"); p.append(werks.strip())
        if scope == "active":   # 사용중 = LG 리시빙(완제품 출하) 2025.01~ 실적 있는 ASSY만
            wh.append("r.model IN (SELECT DISTINCT ITEM_CODE FROM nx.SA_T_LG_RECEIVING_DTL WHERE RECEIVING_YMD>='250101')")
        if cutg.strip() and cutg != "all":   # 제품군 = 완제품(model) cut_gubun (절삭/설치/분지관/이지링크/(없음))
            if cutg == "(없음)":
                wh.append("ISNULL(im.cut_gubun,'')=''")
            else:
                wh.append("im.cut_gubun=?"); p.append(cutg.strip())
        mtl = [x.strip() for x in mt.split(",") if x.strip()]
        if mtl:
            wh.append("ISNULL(ip.make_type,'') IN (%s)" % ",".join("?" * len(mtl))); p += mtl
        if q.strip():
            wh.append("(r.model LIKE ? OR r.parent_code LIKE ? OR r.child_code LIKE ? OR im.item_name LIKE ? OR ip.item_name LIKE ?)")
            qq = "%" + q.strip() + "%"; p += [qq] * 5
        sql = ("""SELECT r.werks, r.model, im.item_name,
                         r.parent_code, ip.item_name, ISNULL(ip.make_type,''), ISNULL(ip.lgroup,''), ISNULL(ip.cut_gubun,''),
                         r.child_code, ic.item_name, r.supply_type,
                         ic.diam, ic.thick, ic.length, ic.net_weight, ic.metal_gubun,
                         MAX(r.child_spec), SUM(CONVERT(float, ISNULL(r.qty,0))), MAX(r.unit), COUNT(*)
                  FROM nx.lg_bom r
                  LEFT JOIN nx.item im ON UPPER(LTRIM(RTRIM(im.item_code)))=UPPER(LTRIM(RTRIM(r.model)))
                  LEFT JOIN nx.item ip ON UPPER(LTRIM(RTRIM(ip.item_code)))=UPPER(LTRIM(RTRIM(r.parent_code)))
                  LEFT JOIN nx.item ic ON UPPER(LTRIM(RTRIM(ic.item_code)))=UPPER(LTRIM(RTRIM(r.child_code)))
                  WHERE """ + " AND ".join(wh) + """
                  GROUP BY r.werks, r.model, im.item_name, r.parent_code, ip.item_name, ip.make_type, ip.lgroup, ip.cut_gubun,
                           r.child_code, ic.item_name, r.supply_type, ic.diam, ic.thick, ic.length, ic.net_weight, ic.metal_gubun
                  ORDER BY r.model, r.parent_code, r.child_code""")
        cur.execute(sql, *p)
        rows = []
        for r in cur.fetchall():
            spec = r[16] or ""
            od, thk, metal, form = _parse_cu_spec(spec)
            # ★치수: 우리 정본(nx.item) 우선, 없으면 LG spec 파싱
            i_od, i_thk, i_len, i_wt, i_metal = _num(r[11]), _num(r[12]), _num(r[13]), _num(r[14]), (r[15] or "")
            dim_src = "우리" if (i_od or i_thk) else ("LG" if (od or thk) else "")
            rows.append({
                "werks": r[0] or "", "model": r[1] or "", "model_name": r[2] or "",
                "parent": r[3] or "", "parent_name": r[4] or "",
                "make_type": r[5] or "", "lgroup": r[6] or "", "cut_gubun": r[7] or "",
                "child": r[8] or "", "child_name": r[9] or "",
                "od": i_od if i_od else (float(od) if od else None),
                "thk": i_thk if i_thk else (float(thk) if thk else None),
                "length": i_len,
                "weight": i_wt,
                "metal": i_metal or metal, "form": form, "dim_src": dim_src,
                "qty": float(r[17] or 0), "unit": r[18] or "",
                "status": "미전환" if r[10] == "Supplier" else ("전환" if r[10] == "Assembly Pull" else (r[10] or "")),
                "spec": spec, "cnt": r[19],
            })
        # 전월(직전 월) LG 리시빙 수량 per model(완제품) — 우선순위/정렬용(GUBUN C+R = 소비 전량)
        import datetime as _dt
        _td = _dt.date.today()
        _py = _td.year if _td.month > 1 else _td.year - 1
        _pm = _td.month - 1 or 12
        prev_ym = "%02d%02d" % (_py % 100, _pm)   # 전월 YYMM (예 2607)
        recv = {}
        mset = list({(x["model"] or "").upper().strip() for x in rows if x["model"]})
        for i in range(0, len(mset), 800):
            ck = mset[i:i + 800]
            ph = ",".join("?" * len(ck))
            cur.execute("SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))), SUM(CONVERT(float,ISNULL(RECV_QTY,0))) "
                        "FROM nx.SA_T_LG_RECEIVING_DTL WHERE RECEIVING_YMD LIKE ? AND UPPER(LTRIM(RTRIM(ITEM_CODE))) IN (%s) "
                        "GROUP BY UPPER(LTRIM(RTRIM(ITEM_CODE)))" % ph, prev_ym + "%", *ck)
            for r in cur.fetchall():
                recv[r[0]] = float(r[1] or 0)
        for x in rows:
            x["recv_prev"] = recv.get((x["model"] or "").upper().strip(), 0)
        rows.sort(key=lambda x: -x["recv_prev"])   # 기본 = 전월 리시빙 많은 순
        total = len(rows)
        models = len(set(x["model"] for x in rows))
        parents = len(set((x["model"], x["parent"]) for x in rows))
        return {"rate": rate, "pull": pull, "supplier": sup, "prev_ym": prev_ym,
                "rows": rows[:int(limit)], "total": total, "shown": min(total, int(limit)),
                "models": models, "parents": parents}
    finally:
        nx.close()
