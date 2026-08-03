# -*- coding: utf-8 -*-
"""
PNC ERP 원가/손익 백엔드 API (FastAPI, 읽기전용 PARTNER_ERP)
- 견적원가조회(w_cs_esti_010) + 손익분석(w_cs_esti_020) 데이터를 라이브 SP로 산출
- 실원가용 SP_CS_견적서(실원가용)_250910, 내부용 SP_CS_견적서(내부용)_250704
- 엔진(engine_full.py)이 이 SP와 일치함을 검증완료 → 향후 신DB 전환 시 엔진으로 교체
실행(내부망):  uvicorn app:app --host 0.0.0.0 --port 8010   → 직원 PC 브라우저: http://<서버IP>:8010/
실행(단독):    uvicorn app:app --host 127.0.0.1 --port 8010
"""
import os, sys, warnings
from datetime import datetime, timedelta
warnings.filterwarnings('ignore')
# ★경로는 이 파일(backend) 기준 상대경로 — 서버 어느 드라이브/폴더에 복사해도 동작(트리만 유지: Projects\{New_ERP, NEW_ERP_1})
_HERE = os.path.dirname(os.path.abspath(__file__))                       # ...\NEW_ERP_1\PNC_ERP_Web\backend
sys.path.insert(0, os.path.join(_HERE, '..', '..', '..', 'New_ERP'))    # Projects\New_ERP (db_client)
sys.path.insert(0, os.path.join(_HERE, '..', '..', '_harness'))         # NEW_ERP_1\_harness (nx_cost_engine)
import db_client, pyodbc
from fastapi import FastAPI, Query, HTTPException, Body, Response, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
try:
    from nx_cost_engine import NxCostEngine   # 검증완료 nx 원가엔진(재료·가공·일반/운반/이윤·손익)
except Exception as _e:
    NxCostEngine = None

# ── 지속 원가엔진(캐시 유지) ── 매 요청 새 엔진=콜드캐시(느림, 3~6초). 엔진 재사용→웜캐시(수십ms).
#    pyodbc 커서 스레드 비안전 → _COST_LOCK 안에서만 사용. 공정/BOM/단가 변경 시 _reset_cost_engine()으로 무효화.
import threading as _threading
_COST_ENG = None
_COST_LOCK = _threading.RLock()
def _get_cost_engine(fresh=False):
    global _COST_ENG
    if fresh and _COST_ENG is not None:
        try: _COST_ENG.close()
        except Exception: pass
        _COST_ENG = None
    if _COST_ENG is None:
        _COST_ENG = NxCostEngine()
    return _COST_ENG
def _reset_cost_engine():
    """원가 입력(공정/BOM/단가) 변경 후 캐시 무효화 → 다음 계산은 최신 DB 반영."""
    global _COST_ENG
    with _COST_LOCK:
        if _COST_ENG is not None:
            try: _COST_ENG.close()
            except Exception: pass
            _COST_ENG = None

SP_SIL = 'SP_CS_견적서(실원가용)_250910'
SP_NAE = 'SP_CS_견적서(내부용)_250704'

app = FastAPI(title="PNC ERP 원가/손익 API", version="1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
from fastapi.middleware.gzip import GZipMiddleware   # data.js(25MB)·대용량 응답 gzip 압축 → 초기로딩 가속
app.add_middleware(GZipMiddleware, minimum_size=1024)

# 라이브 조회 API(읽기전용 PARTNER_ERP) — 스냅샷 화면 승격용
from live_api import live_router
app.include_router(live_router)
from routers import prod as _r_prod   # 생산실적 도메인 라우터(app.py 분리)
app.include_router(_r_prod.router)
from routers import cost as _r_cost   # 원가 도메인 라우터
app.include_router(_r_cost.router)
from routers import qc as _r_qc   # 품질 도메인 라우터
app.include_router(_r_qc.router)
from routers import esticost as _r_esti   # 견적원가+납품 도메인 라우터
app.include_router(_r_esti.router)
from routers import planinput as _r_pi   # 생산계획입력 도메인 라우터
app.include_router(_r_pi.router)
from routers import sales as _r_sales   # 사급+매출+권한 도메인 라우터
app.include_router(_r_sales.router)
from routers import setin as _r_setin   # 협력사 세트입고 도메인 라우터
app.include_router(_r_setin.router)
import weight_calc  # 무게정산(중량조정) 계산

import re as _re_guard
# ★라이브 PARTNER_ERP 쓰기 가드: DML(INSERT/UPDATE/DELETE/MERGE/TRUNCATE/DROP/ALTER/CREATE/GRANT/REVOKE) 첫키워드 거부.
#   허용: SELECT·WITH·SET·DECLARE·EXEC/EXECUTE(조회SP). 쓰기는 nx(_nx, PARTNER_ERP_TEST3)에서만. _nx엔 가드 미적용.
_LIVE_DML = _re_guard.compile(r'^(INSERT|UPDATE|DELETE|MERGE|TRUNCATE|DROP|ALTER|CREATE|GRANT|REVOKE)\b', _re_guard.IGNORECASE)

def _sql_lead(sql):
    """주석(-- , /* */)·공백·선행 세미콜론 제거 후 남은 문장 앞부분 반환(첫 키워드 판정용)."""
    s = str(sql or '')
    while True:
        s = s.lstrip(' \t\r\n;')
        if s[:2] == '--':
            i = s.find('\n'); s = '' if i < 0 else s[i + 1:]
        elif s[:2] == '/*':
            i = s.find('*/'); s = '' if i < 0 else s[i + 2:]
        else:
            return s

class _ROCursor:
    """라이브(PARTNER_ERP) 읽기전용 가드 커서 — DML 첫키워드 거부, 조회/EXEC/SET/DECLARE 허용."""
    def __init__(self, cur): object.__setattr__(self, '_cur', cur)
    def execute(self, sql, *a, **k):
        if _LIVE_DML.match(_sql_lead(sql)):
            raise PermissionError("라이브 PARTNER_ERP는 읽기전용입니다 — INSERT/UPDATE/DELETE/DDL 차단(쓰기는 nx=PARTNER_ERP_TEST3에서만).")
        return self._cur.execute(sql, *a, **k)
    def executemany(self, *a, **k):
        raise PermissionError("라이브 PARTNER_ERP는 읽기전용입니다 — executemany(쓰기) 차단.")
    def __iter__(self): return iter(self._cur)
    def __enter__(self): return self
    def __exit__(self, *a):
        c = getattr(self._cur, 'close', None);
        if c: c()
        return False
    def __getattr__(self, n): return getattr(self._cur, n)

class _ROConn:
    """라이브 읽기전용 가드 커넥션 — .cursor()가 가드 커서 반환. 그 외 위임."""
    def __init__(self, cn): object.__setattr__(self, '_cn', cn)
    def cursor(self): return _ROCursor(self._cn.cursor())
    def __enter__(self): return self
    def __exit__(self, *a):
        c = getattr(self._cn, 'close', None)
        if c: c()
        return False
    def __getattr__(self, n): return getattr(self._cn, n)

def _conn():
    # ApplicationIntent=ReadOnly: 테스트 연결·SELECT 1 성공 확인됨(2026-07-29). AG 미사용 시 무해(라우팅 힌트).
    #   실질 쓰기차단은 _ROConn/_ROCursor 코드가드. 쓰기는 _nx(PARTNER_ERP_TEST3)에서만 허용.
    cs = (f'DRIVER={{SQL Server}};SERVER={db_client.DB_SERVER},{db_client.DB_PORT};'
          f'DATABASE=PARTNER_ERP;UID={db_client.DB_USER};PWD={db_client.DB_PASSWORD};ApplicationIntent=ReadOnly')
    return _ROConn(pyodbc.connect(cs, autocommit=True))

def _num(x):
    try: return round(float(x), 2)
    except Exception: return 0.0

def _run_sp(sp, item, ymd):
    """SP를 실행하고 마지막 결과셋(컬럼, 행)을 반환."""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SET NOCOUNT ON; EXEC [dbo].[" + sp + "] ?, ?", item, ymd)
        cols, rows = None, None
        while True:
            if cur.description:
                cols = [d[0] for d in cur.description]
                try: rows = cur.fetchall()
                except Exception: rows = []
            if not cur.nextset(): break
        return cols, (rows or [])
    finally:
        cn.close()

def _shape(cols, rows):
    if not cols:
        return {"rows": [], "agg": {}}
    ix = {c: i for i, c in enumerate(cols)}
    def g(r, c): return r[ix[c]] if c in ix else None
    grid, top = [], None
    for r in rows:
        lvl = int(g(r, 'C_ITEM_LEVEL') or 0)
        grid.append({
            "level": lvl,
            "part": str(g(r, 'C_ITEM_CODE') or '').strip(),
            "desc": str(g(r, 'C_ITEM_DESC') or '').strip(),
            "sgroup": str(g(r, 'ITEM_SGROUP') or '').strip(),
            "diam": _num(g(r, 'C_ITEM_DIAM')), "thick": _num(g(r, 'C_ITEM_THICK')), "length": _num(g(r, 'C_ITEM_LENGTH')),
            "metal": str(g(r, 'C_METAL_GUBUN') or '').strip(),
            "cust": str(g(r, 'CUST_DESC') or '').strip(),
            "unit_gubun": str(g(r, 'COST_GUBUN') or '').strip(),
            "use_qty": _num(g(r, 'USE_QTY')),
            "won_mat": _num(g(r, 'WON_MAT_COST')),
            "jai": _num(g(r, 'JAI_COST')),
        })
        if lvl == 0: top = r
    agg = {}
    if top is not None:
        for f in ['JAI_COST', 'GAGONG_AMT', 'ILBAN_AMT', 'UNBAN_AMT', 'PROFIT_AMT', 'TOT_AMT',
                  'LME_CHA_AMT', 'LG_COST', 'WON_JAI_AMT', 'BU_JAI_AMT', 'SA_JAI_AMT']:
            agg[f] = _num(g(top, f)) if f in ix else None
        agg['SONIK'] = (agg.get('LG_COST') or 0) - (agg.get('TOT_AMT') or 0)
    return {"rows": grid, "agg": agg}

@app.get("/api/health")
def health():
    return {"ok": True, "sp_sil": SP_SIL, "sp_nae": SP_NAE}

# ===================== 품목 BOM 관리 (nx.bom, 쓰기=TEST3) =====================
def _nx():
    cs = (f'DRIVER={{SQL Server}};SERVER={db_client.DB_SERVER},{db_client.DB_PORT};'
          f'DATABASE=PARTNER_ERP_TEST3;UID={db_client.DB_USER};PWD={db_client.DB_PASSWORD}')
    return pyodbc.connect(cs, autocommit=True)

def _nx_tx():
    """nx 쓰기 트랜잭션 커넥션(autocommit=False). ★그룹 단위 원자성 전용:
    호출측은 반드시 try:...commit() / except: rollback();raise / finally: close() 로 감쌀 것.
    멀티행 그룹(이동 2행·백플러시 다건+log·재키 delete+insert)이 부분실패로 net-0/멱등 불변식을 깨지 않게 함."""
    cs = (f'DRIVER={{SQL Server}};SERVER={db_client.DB_SERVER},{db_client.DB_PORT};'
          f'DATABASE=PARTNER_ERP_TEST3;UID={db_client.DB_USER};PWD={db_client.DB_PASSWORD}')
    return pyodbc.connect(cs, autocommit=False)

def _b(x):  # 파이썬 truthy → BIT
    return 1 if x in (1, True, '1', 'Y', 'y', 'true', 'True') else 0

@app.get("/api/bom/search")
def bom_search(q: str = Query('', description="품번/품명 부분검색")):
    q = q.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        like = f'%{q}%'
        # BOM 보유 품목 우선. 인덱스(item_code PK) 활용, TOP 60 제한.
        cur.execute("""
            SELECT TOP 60 i.item_code, i.item_name, i.item_type,
              CASE WHEN h.item_code IS NOT NULL THEN 1 ELSE 0 END AS has_bom
            FROM nx.item i
            LEFT JOIN (SELECT DISTINCT item_code FROM nx.bom_header) h ON h.item_code = i.item_code
            WHERE i.item_code LIKE ? OR i.item_name LIKE ?
            ORDER BY has_bom DESC, i.item_code""", like, like)
        rows = [{"item": r[0], "name": r[1], "type": r[2], "has_bom": bool(r[3])} for r in cur.fetchall()]
        return {"rows": rows}
    finally:
        cn.close()

@app.get("/api/bom/get")
def bom_get(item: str = Query(..., description="품번")):
    item = item.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT item_name, item_type FROM nx.item WHERE item_code=?", item)
        pi = cur.fetchone()
        if not pi:
            raise HTTPException(404, f"품목 {item} 없음")
        cur.execute("""SELECT GAGONG_PROC_CODE, GAGONG_PROC_DESC FROM PARTNER_ERP.dbo.PR_M_PROC_GAGONG
                       WHERE ISNULL(GAGONG_PROC_CODE,'')<>'' ORDER BY SORT_KEY, GAGONG_PROC_CODE""")
        procs = [{"code": str(r[0]).strip(), "name": (str(r[1]).strip() if r[1] else str(r[0]).strip())} for r in cur.fetchall()]
        cur.execute("SELECT bom_id, version, status FROM nx.bom_header WHERE item_code=?", item)
        h = cur.fetchone()
        if not h:
            return {"item": item, "name": pi[0], "header": None, "lines": [], "procs": procs}
        bom_id = h[0]
        cur.execute("""
            SELECT l.seq, l.child_item, ci.item_name, l.qty, l.node_type,
                   l.cs_calc_except, l.lme_except, l.sagub_default, l.is_optional,
                   l.from_ymd, l.to_ymd, l.except_flag, l.set_except, l.kitting, l.vir_item,
                   l.proc_gubun, l.gagong_proc, l.s_work, l.wh_gagong, l.in_gagong, l.cust_code, l.remarks,
                   ci.item_spec, ci.metal_gubun, ci.diam, ci.thick, ci.length, ci.item_type AS child_type,
                   ci.net_weight, ci.unit, ci.sgroup, ci.lgroup, ci.make_type, ci.cost_gubun, ci.status,
                   ISNULL(ci.in_cust,'') AS in_cust, ISNULL(pv.partner_name, pc.CUST_DESC) AS cust_name
            FROM nx.bom_line l
            LEFT JOIN nx.item ci ON ci.item_code = l.child_item
            LEFT JOIN nx.partner pv ON pv.partner_code = ci.in_cust
            LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST pc ON pc.CUST_CODE = ci.in_cust
            WHERE l.bom_id = ? ORDER BY l.seq""", bom_id)
        cols = [d[0] for d in cur.description]
        lines = []
        for r in cur.fetchall():
            d = {}
            for k, v in zip(cols, r):
                d[k] = bool(v) if isinstance(v, bool) else v
            lines.append(d)
        return {"item": item, "name": pi[0], "type": pi[1],
                "header": {"bom_id": bom_id, "version": h[1], "status": h[2]}, "lines": lines, "procs": procs}
    finally:
        cn.close()

# 품목 마스터 편집용 드롭다운 코드(CM_M_MASTER_DETAIL) + 마스터 저장(nx.item)
_CODE_GROUPS = {"lgroup": "PR005", "sgroup": "PR006", "metal": "PR019",
                "material": "PR020", "unit": "CM002", "obtain": "PR007", "itemclass": "PR008"}

@app.get("/api/codes")
def codes():
    cn = _conn(); cur = cn.cursor()
    try:
        out = {}
        for key, grp in _CODE_GROUPS.items():
            cur.execute("""SELECT LTRIM(RTRIM(DETAIL_CODE)), LTRIM(RTRIM(DETAIL_DESC))
                FROM CM_M_MASTER_DETAIL WHERE KIND_CODE=? AND ISNULL(USE_FLAG,'1')='1'
                ORDER BY SORT_SEQ, DETAIL_CODE""", grp)
            out[key] = [{"code": r[0], "name": r[1]} for r in cur.fetchall()]
        out["make_type"] = [{"code": k, "name": v} for k, v in _ITEM_MAKE.items() if k]   # 통일: _ITEM_MAKE(품목조회 일치)
        out["cost_gubun"] = [{"code": "2", "name": "구매단가"}, {"code": "3", "name": "소재단가"},
                             {"code": "1", "name": "내부단가"}, {"code": "5", "name": "기타"}]
        out["status"] = [{"code": "사용", "name": "사용"}, {"code": "중지", "name": "중지"}]
        # 거래처(협력사) 검색용 상위 — 프론트는 item/vendorsearch로 별도 검색
        return out
    finally:
        cn.close()

@app.get("/api/item/vendorsearch")
def item_vendorsearch(q: str = Query("")):
    """매입처(거래처) 검색 — nx.partner + 라이브 CM_M_CUST."""
    cn = _nx(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT TOP 30 c.CUST_CODE, c.CUST_DESC FROM PARTNER_ERP.dbo.CM_M_CUST c
            WHERE c.CUST_CODE LIKE ? OR c.CUST_DESC LIKE ? ORDER BY c.CUST_CODE""", like, like)
        return {"rows": [{"code": r[0], "name": r[1]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.post("/api/item/save")
def item_save(payload: dict = Body(...)):
    """품목 마스터 일괄 저장(nx.item) — 원소재 스펙~거래처~분류~생산구분. 상향식 등록."""
    rows = payload.get("rows", []) or []
    nx = _nx(); cur = nx.cursor()
    try:
        saved = 0; errs = []
        for r in rows:
            code = str(r.get("item_code", "")).strip()
            if not code:
                continue
            def s(k): v = r.get(k); return None if v in (None, "") else str(v).strip()
            def n(k):
                v = r.get(k)
                try: return float(v) if v not in (None, "") else None
                except: return None
            vals = (s("item_name"), s("item_spec"), s("metal_gubun"), n("diam"), n("thick"), n("length"),
                    n("net_weight"), s("unit"), s("in_cust"), s("sgroup"), s("lgroup"), s("make_type"),
                    s("cost_gubun"), s("status"))
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", code)
            if cur.fetchone():
                cur.execute("""UPDATE nx.item SET item_name=ISNULL(?,item_name), item_spec=?, metal_gubun=?,
                    diam=?, thick=?, length=?, net_weight=?, unit=ISNULL(?,unit), in_cust=?, sgroup=?, lgroup=?,
                    make_type=?, cost_gubun=?, status=ISNULL(?,status) WHERE item_code=?""", *vals, code)
            else:
                cur.execute("""INSERT INTO nx.item(item_code,item_name,item_spec,metal_gubun,diam,thick,length,
                    net_weight,unit,in_cust,sgroup,lgroup,make_type,cost_gubun,status,item_type)
                    VALUES(?,?,?,?,?,?,?,?,ISNULL(?,'EA'),?,?,?,?,?,ISNULL(?,'사용'),'부품')""", code, *vals)
            saved += 1
        _reset_cost_engine()   # 스펙(치수·재질·중량·조달) 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "count": saved, "errors": errs}
    finally:
        nx.close()

@app.get("/api/bom/tree")
def bom_tree(item: str = Query(..., description="품번"), real: int = Query(1, description="1=실사용전개(원가제외 스킵+제작품만 전개,매입중단)=실원가용 일치, 0=전체전개")):
    """다단계 BOM 트리(레벨별) — CS_M_ITEM_BOM 재귀전개. 매입처=컴포넌트 IN_CUST_CODE(현행 벤더).
    real=1(기본): 견적원가조회(실원가용, SP_CS_견적서(BOM)) 전개와 일치 — CS_CALC_EXCEPT_FLAG='1'(원가제외=현행아닌 조달경로) 제외 + MAKE_TYPE='1'(제작/자체)만 하위전개, 매입/구매품(구매완제)은 전개중단."""
    item = item.strip()
    real = 1 if real is None else int(real)
    exc_a = "AND ISNULL(CS_CALC_EXCEPT_FLAG,'0')<>'1'" if real else ""      # 원가제외 라인 스킵
    exc_r = "AND ISNULL(b.CS_CALC_EXCEPT_FLAG,'0')<>'1'" if real else ""
    mk_gate = "JOIN PR_M_ITEM pt ON pt.ITEM_CODE=t.c AND ISNULL(pt.MAKE_TYPE,'')='1'" if real else ""  # 제작품만 하위전개
    cn = _conn(); cur = cn.cursor()  # live PARTNER_ERP
    try:
        cur.execute(f"""WITH tree AS (
            SELECT ITEM_CODE p, MAT_CODE c, CAST(USE_QTY AS decimal(18,6)) q, ISNULL(SAGUB_FLAG,'0') sag,
                   ISNULL(SET_EXCEPT_FLAG,'') se, ISNULL(KITTING_FLAG,'') kt, ISNULL(VIR_ITEM_FLAG,'') vir,
                   ISNULL(CS_CALC_EXCEPT_FLAG,'') ce, ISNULL(LME_EXCEPT_FLAG,'') le,
                   ISNULL(GAGONG_PROC_CODE,'') gp, ISNULL(S_WORK_CODE,'') sw, ISNULL(BOM_SEQ,0) sq, 1 lvl
            FROM CS_M_ITEM_BOM WHERE ITEM_CODE=? AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101' {exc_a}
            UNION ALL
            SELECT b.ITEM_CODE, b.MAT_CODE, CAST(b.USE_QTY AS decimal(18,6)), ISNULL(b.SAGUB_FLAG,'0'),
                   ISNULL(b.SET_EXCEPT_FLAG,''), ISNULL(b.KITTING_FLAG,''), ISNULL(b.VIR_ITEM_FLAG,''),
                   ISNULL(b.CS_CALC_EXCEPT_FLAG,''), ISNULL(b.LME_EXCEPT_FLAG,''),
                   ISNULL(b.GAGONG_PROC_CODE,''), ISNULL(b.S_WORK_CODE,''), ISNULL(b.BOM_SEQ,0), t.lvl+1
            FROM tree t JOIN CS_M_ITEM_BOM b ON b.ITEM_CODE=t.c AND b.FROM_APPLY_YMD<='991231' AND b.TO_APPLY_YMD>='260101' {exc_r}
            {mk_gate}
            WHERE t.lvl < 8)
            SELECT p,c,q,sag,se,kt,vir,ce,le,gp,sw,sq,lvl FROM tree OPTION(MAXRECURSION 50)""", item)
        edges = {}
        for r in cur.fetchall():
            edges.setdefault(r[0], []).append({
                "child": r[1], "q": float(r[2] or 0), "sag": str(r[3]), "se": str(r[4]), "kt": str(r[5]),
                "vir": str(r[6]), "ce": str(r[7]), "le": str(r[8]), "gp": str(r[9]).strip(),
                "sw": str(r[10]).strip(), "sq": int(r[11] or 0)})
        for p in edges:
            edges[p].sort(key=lambda x: x["sq"])
        # 노드 상세(품명·매입처=IN_CUST·치수) 일괄
        nodes = {item} | {e["child"] for lst in edges.values() for e in lst} | set(edges.keys())
        info = {}
        if nodes:
            nl = list(nodes)
            for i in range(0, len(nl), 900):
                chunk = nl[i:i+900]; ph = ",".join("?" * len(chunk))
                cur.execute(f"""SELECT m.ITEM_CODE, ISNULL(m.ITEM_DESC,''), ISNULL(m.ITEM_SPEC,''),
                      ISNULL(m.IN_CUST_CODE,''), ISNULL(c.CUST_DESC,''), ISNULL(m.METAL_GUBUN,''),
                      ISNULL(m.ITEM_DIAM,0), ISNULL(m.ITEM_THICK,0), ISNULL(m.ITEM_LENGTH,0)
                    FROM PR_M_ITEM m LEFT JOIN CM_M_CUST c ON c.CUST_CODE=m.IN_CUST_CODE
                    WHERE m.ITEM_CODE IN ({ph})""", *chunk)
                for r in cur.fetchall():
                    info[r[0]] = {"nm": r[1], "spec": r[2], "cust": str(r[3]).strip(), "custnm": r[4],
                                  "metal": r[5], "diam": float(r[6] or 0), "thick": float(r[7] or 0), "length": float(r[8] or 0)}
        rootnm = info.get(item, {}).get("nm", "")
        out = [{"level": 0, "code": item, "nm": rootnm, "spec": info.get(item, {}).get("spec", ""),
                "qty": 1, "cust": "", "custnm": "", "sag": "", "se": "", "kt": "", "vir": "", "ce": "", "le": "",
                "gp": "", "sw": "", "metal": info.get(item, {}).get("metal", ""),
                "diam": info.get(item, {}).get("diam", 0), "thick": info.get(item, {}).get("thick", 0), "length": info.get(item, {}).get("length", 0),
                "haskids": item in edges}]
        seen = set()
        def walk(code, lvl):
            if code in seen: return   # 순환 방지
            seen.add(code)
            for e in edges.get(code, []):
                ci = info.get(e["child"], {})
                out.append({"level": lvl, "code": e["child"], "nm": ci.get("nm", ""), "spec": ci.get("spec", ""),
                    "qty": e["q"], "cust": ci.get("cust", ""), "custnm": ci.get("custnm", ""),
                    "sag": e["sag"], "se": e["se"], "kt": e["kt"], "vir": e["vir"], "ce": e["ce"], "le": e["le"],
                    "gp": e["gp"], "sw": e["sw"], "metal": ci.get("metal", ""),
                    "diam": ci.get("diam", 0), "thick": ci.get("thick", 0), "length": ci.get("length", 0),
                    "haskids": e["child"] in edges})
                walk(e["child"], lvl + 1)
            seen.discard(code)
        walk(item, 1)
        maxlvl = max((r["level"] for r in out), default=0)
        return {"item": item, "name": rootnm, "rows": out, "count": len(out) - 1, "maxlevel": maxlvl}
    finally:
        cn.close()

@app.post("/api/bom/save")
def bom_save(payload: dict = Body(...)):
    """BOM 구성 전체 교체 저장. 마스터 가드: 참조무결성·중복·순환·필수값. (마감/재고 가드는 재고·실적 프로그램용, BOM 미적용)"""
    item = str(payload.get("item", "")).strip()
    lines = payload.get("lines", []) or []
    if not item:
        raise HTTPException(400, "item(품번) 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", item)
        if not cur.fetchone():
            # 신규 BOM: 대상 도번 미등록 시 최소등록(프론트에서 "신규 생성?" 확인 후 target_name 전달)
            tname = str(payload.get("target_name", "") or "").strip() or item
            cur.execute("INSERT INTO nx.item(item_code,item_name,item_type) VALUES(?,?,N'제품')", item, tname)
        errs, seen = [], set()
        for i, ln in enumerate(lines, 1):
            ch = str(ln.get("child_item", "")).strip()
            if not ch:
                errs.append(f"{i}행: 자품목코드 필요"); continue
            if ch == item:
                errs.append(f"{i}행: 자기참조 불가 ({ch})")
            if ch in seen:
                errs.append(f"{i}행: 중복 자품목 ({ch})")
            seen.add(ch)
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", ch)
            if not cur.fetchone():
                errs.append(f"{i}행: 미등록 품목 ({ch})")
            # 순환: 자품목의 BOM이 이 품목을 포함하면 순환
            cur.execute("""SELECT 1 FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id
                           WHERE h.item_code=? AND l.child_item=?""", ch, item)
            if cur.fetchone():
                errs.append(f"{i}행: 순환참조 ({ch} 가 {item} 를 이미 포함)")
        if errs:
            return {"ok": False, "errors": errs}
        # 헤더 확보
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", item)
        h = cur.fetchone()
        if h:
            bom_id = h[0]
        else:
            cur.execute("INSERT INTO nx.bom_header(item_code,version,apply_from,status) VALUES(?,1,'2000-01-01',N'확정')", item)
            cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", item)
            bom_id = cur.fetchone()[0]
        # 전체 교체
        cur.execute("DELETE FROM nx.bom_line WHERE bom_id=?", bom_id)
        for seq, ln in enumerate(lines, 1):
            cur.execute("""INSERT INTO nx.bom_line
                (bom_id,seq,child_item,qty,node_type,cs_calc_except,lme_except,sagub_default,is_optional,
                 from_ymd,to_ymd,except_flag,set_except,kitting,vir_item,proc_gubun,gagong_proc,s_work,wh_gagong,in_gagong,cust_code,remarks)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                bom_id, seq, str(ln.get("child_item", "")).strip(), float(ln.get("qty") or 0),
                str(ln.get("node_type") or "부품"),
                _b(ln.get("cs_calc_except")), _b(ln.get("lme_except")), _b(ln.get("sagub_default")), _b(ln.get("is_optional")),
                (ln.get("from_ymd") or None), (ln.get("to_ymd") or None),
                _b(ln.get("except_flag")), _b(ln.get("set_except")), _b(ln.get("kitting")), _b(ln.get("vir_item")),
                (ln.get("proc_gubun") or None), (ln.get("gagong_proc") or None), (ln.get("s_work") or None),
                (ln.get("wh_gagong") or None), (ln.get("in_gagong") or None), (ln.get("cust_code") or None),
                (ln.get("remarks") or None))
        _reset_cost_engine()   # BOM 구성/소요량 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "count": len(lines), "bom_id": bom_id}
    finally:
        cn.close()

@app.post("/api/bom/qty")
def bom_qty(payload: dict = Body(...)):
    """BOM 단건 소요량 수정(부모+자식) — 다른 라인 속성 보존(전체교체 아님). 내부원가 수정에서 안전 갱신."""
    parent = str(payload.get("parent", "")).strip()
    child = str(payload.get("child", "")).strip()
    if not parent or not child:
        raise HTTPException(400, "parent·child 필요")
    try:
        q = float(payload.get("qty"))
    except Exception:
        raise HTTPException(400, "qty(숫자) 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""UPDATE l SET l.qty=? FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id
            WHERE h.item_code=? AND l.child_item=?""", q, parent, child)
        n = cur.rowcount
        _reset_cost_engine()
        return {"ok": n > 0, "updated": int(n)}
    finally:
        cn.close()

@app.post("/api/item/spec")
def item_spec(payload: dict = Body(...)):
    """품목 원소재 스펙 단건 수정(외경·두께·재질·중량) — ★지정한 필드만 갱신(item/save와 달리 미지정 필드 NULL 덮어쓰기 없음)."""
    code = str(payload.get("item_code", "")).strip()
    if not code:
        raise HTTPException(400, "item_code 필요")
    sets, vals = [], []
    for key, col in (("diam", "diam"), ("thick", "thick"), ("metal_gubun", "metal_gubun"), ("net_weight", "net_weight")):
        if key in payload and payload[key] not in (None, ""):
            if col == "metal_gubun":
                v = str(payload[key]).strip()
            else:
                try: v = float(payload[key])
                except Exception: raise HTTPException(400, f"{key}(숫자) 필요")
            sets.append(f"{col}=?"); vals.append(v)
    if not sets:
        return {"ok": True, "updated": 0}
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute(f"UPDATE nx.item SET {','.join(sets)} WHERE item_code=?", *vals, code)
        n = cur.rowcount
        _reset_cost_engine()
        return {"ok": n > 0, "updated": int(n)}
    finally:
        cn.close()


# ===================== LG BOM 조회 (nx.lg_bom, LG 원본 BOM Explosion 56,522행) =====================
@app.get("/api/lgbom/search")
def lgbom_search(q: str = Query(""), werks: str = Query(""), limit: int = Query(200)):
    """LG BOM 상위모델(완제품) 검색. model=상위품번. werks: DMZ(SAC)/DGZ(RAC). 유효기간 현행."""
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["b.stufe=1"]; p = []
        if q: w.append("(b.model LIKE ? OR b.parent_code LIKE ?)"); p += [f"%{q}%", f"%{q}%"]
        if werks: w.append("b.werks=?"); p.append(werks)
        cur.execute(f"""SELECT TOP {int(limit)} b.model, b.werks, MAX(b.parent_code) parent_code,
              MAX(ISNULL(pi.ITEM_DESC,'')) modelnm, COUNT(*) child_cnt, MAX(b.valid_from) valid_from
            FROM nx.lg_bom b LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM pi ON pi.ITEM_CODE=b.model
            WHERE {' AND '.join(w)} GROUP BY b.model, b.werks ORDER BY b.model""", *p)
        cols = [d[0] for d in cur.description]
        return {"rows": [dict(zip(cols, r)) for r in cur.fetchall()]}
    finally:
        cn.close()

@app.get("/api/lgbom/tree")
def lgbom_tree(model: str = Query(...), werks: str = Query("")):
    """선택 모델의 LG BOM 전개(전 레벨). stufe/posnr 순. 프론트에서 parent_code→child_code 트리 조립."""
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["b.model=?"]; p = [model]
        if werks: w.append("b.werks=?"); p.append(werks)
        cur.execute(f"""SELECT b.id, b.werks, b.stufe, b.posnr, b.parent_code, b.child_code,
              b.child_desc, b.child_spec, b.qty, b.unit, b.supply_type, b.mmsta, b.matty, b.lowest_flg,
              b.main_mat, b.matkl, b.valid_from, b.valid_to, ISNULL(i.ITEM_DESC,'') nx_desc
            FROM nx.lg_bom b LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM i ON i.ITEM_CODE=b.child_code
            WHERE {' AND '.join(w)} ORDER BY b.stufe, b.posnr, b.id""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        # 최상위 parent(model) 정보
        cur.execute("SELECT ISNULL(ITEM_DESC,'') FROM PARTNER_ERP.dbo.PR_M_ITEM WHERE ITEM_CODE=?", model)
        mn = cur.fetchone()
        return {"model": model, "modelnm": (mn[0] if mn else ""), "rows": rows}
    finally:
        cn.close()

@app.post("/api/lgbom/upload")
async def lgbom_upload(file: UploadFile = File(...)):
    """LG BOM Explosion 엑셀 업로드 → nx.lg_bom 적재(모델·werks별 교체). 신규 BOM 등록 전 사전업로드용.
       헤더 1행: MODEL/WERKS/STUFE/POSNR/MATNR(부모)/IDNRK(자식)/OJTXP(품명)/CHI_SPECI(규격)/MENGE(수량)/MEINS(단위)/ETEXT(supply_type)/MATTY/LOWEST_FLG/MATKL/DATAB/DATVT 등."""
    import io as _io
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl 미설치(서버)")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 열기 실패: {str(e)[:120]}")
    ws = wb.active
    itr = ws.iter_rows(values_only=True)
    try:
        hdr = next(itr)
    except StopIteration:
        raise HTTPException(400, "빈 파일")
    ix = {str(h or '').strip().upper(): i for i, h in enumerate(hdr)}
    if not all(k in ix for k in ('MODEL', 'MATNR', 'IDNRK')):
        raise HTTPException(400, "LG BOM 형식 아님 — 헤더 1행에 MODEL·MATNR·IDNRK 필요")
    def gv(r, n):
        i = ix.get(n); v = r[i] if (i is not None and i < len(r)) else None
        return None if v in (None, '') else v
    def gs(r, n, ln=None):
        v = gv(r, n); s = '' if v is None else str(v).strip(); return s[:ln] if ln else s
    def gi(r, n):
        try: return int(float(gv(r, n) or 0))
        except Exception: return 0
    def gf(r, n):
        try: return float(gv(r, n) or 0)
        except Exception: return 0.0
    recs = []; models = set()
    for r in itr:
        if not r or not any(x not in (None, '') for x in r): continue
        model = gs(r, 'MODEL')
        if not model: continue
        werks = gs(r, 'WERKS', 4)
        recs.append(('C', werks, model, gi(r, 'STUFE'), gs(r, 'POSNR', 10),
            gs(r, 'MATNR', 30), gs(r, 'IDNRK', 30), gs(r, 'OJTXP', 150), gs(r, 'CHI_SPECI', 200),
            gf(r, 'MENGE'), gs(r, 'MEINS', 6), gs(r, 'PAR_UIT', 4), gs(r, 'ETEXT', 30),
            gs(r, 'MMSTA', 6), gs(r, 'MTSTB', 80), gs(r, 'MATTY', 6), (gs(r, 'LOWEST_FLG', 1) or 'N'),
            gs(r, 'ALT_ITEM', 30), gs(r, 'MAIN_MAT', 6), gs(r, 'MATKL', 20), gs(r, 'DATAB', 10), gs(r, 'DATVT', 10)))
        models.add((werks, model))
    wb.close()
    if not recs:
        raise HTTPException(400, "데이터 행 없음")
    cn = _nx(); cur = cn.cursor()
    try:
        for (werks, model) in models:
            if werks:
                cur.execute("DELETE FROM nx.lg_bom WHERE model=? AND werks=?", model, werks)
            else:
                cur.execute("DELETE FROM nx.lg_bom WHERE model=? AND ISNULL(werks,'')=''", model)
        cur.executemany("""INSERT INTO nx.lg_bom
            (cr,werks,model,stufe,posnr,parent_code,child_code,child_desc,child_spec,qty,unit,uit,supply_type,
             mmsta,mtstb,matty,lowest_flg,alt_item,main_mat,matkl,valid_from,valid_to,src_valid,load_dt)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CONVERT(varchar(10),GETDATE(),120),GETDATE())""", recs)
        return {"ok": True, "rows": len(recs), "models": sorted({m for (w, m) in models}),
                "werks": sorted({w for (w, m) in models if w}), "file": file.filename}
    finally:
        cn.close()

@app.post("/api/bom/copy")
def bom_copy(payload: dict = Body(...)):
    """BOM 복사(nx 전용): source의 직접자식 구성을 target으로 복제. source nx.bom_line 우선, 없으면 라이브 CS_M_ITEM_BOM 직접자식.
       target 미등록시 nx.item 최소등록. 라이브는 미변경. 유사공정 협력사 변형용."""
    source = str(payload.get("source", "")).strip()
    target = str(payload.get("target", "")).strip().upper()
    if not source or not target:
        raise HTTPException(400, "source·target 필요")
    if source == target:
        return {"ok": False, "error": "원본과 대상이 같습니다."}
    warn = ""
    cn = _nx(); cur = cn.cursor()
    try:
        # 대상 품목 nx.item 확보(미등록시 최소등록: source 이름 기반)
        cur.execute("SELECT item_name FROM nx.item WHERE item_code=?", target)
        if not cur.fetchone():
            cur.execute("SELECT ISNULL(item_name,'') FROM nx.item WHERE item_code=?", source)
            sr = cur.fetchone(); snm = (sr[0] if sr else "") or source
            cur.execute("INSERT INTO nx.item(item_code,item_name) VALUES(?,?)", target, f"{snm} (복사:{source})")
            warn = f"대상 품번 {target} 신규 등록(품목마스터에서 속성 보완 필요)"
        # source 직접자식: nx.bom_line 우선
        cur.execute("""SELECT l.child_item,l.qty,l.node_type,l.cs_calc_except,l.lme_except,l.sagub_default,l.is_optional,
                 l.from_ymd,l.to_ymd,l.except_flag,l.set_except,l.kitting,l.vir_item,l.proc_gubun,l.gagong_proc,l.s_work,
                 l.wh_gagong,l.in_gagong,l.cust_code,l.remarks
              FROM nx.bom_line l JOIN nx.bom_header h ON h.bom_id=l.bom_id WHERE h.item_code=? ORDER BY l.seq""", source)
        rows = cur.fetchall()
        src = "nx"
        if not rows:
            # 라이브 CS_M_ITEM_BOM 직접자식(현행 유효일자)로 대체
            lc = _conn(); lcur = lc.cursor()
            try:
                lcur.execute("""SELECT LTRIM(RTRIM(MAT_CODE)), CAST(USE_QTY AS decimal(18,6)), ISNULL(CS_CALC_EXCEPT_FLAG,''),
                       ISNULL(LME_EXCEPT_FLAG,''), ISNULL(SAGUB_FLAG,'0'), ISNULL(SET_EXCEPT_FLAG,''), ISNULL(KITTING_FLAG,''),
                       ISNULL(VIR_ITEM_FLAG,''), ISNULL(GAGONG_PROC_CODE,''), ISNULL(S_WORK_CODE,''), ISNULL(BOM_SEQ,0)
                    FROM CS_M_ITEM_BOM WHERE ITEM_CODE=? AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101' ORDER BY BOM_SEQ""", source)
                lrows = lcur.fetchall()
            finally:
                lc.close()
            rows = [(r[0], r[1], "부품", r[2], r[3], r[4], 0, None, None, "", r[5], r[6], r[7], None, r[8], r[9], None, None, None, None) for r in lrows]
            src = "live"
        if not rows:
            return {"ok": False, "error": f"원본 {source} 의 BOM 구성을 찾을 수 없습니다."}
        # target 헤더 확보 + 전체교체
        cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", target)
        h = cur.fetchone()
        if h:
            bom_id = h[0]
        else:
            cur.execute("INSERT INTO nx.bom_header(item_code,version,apply_from,status) VALUES(?,1,'2000-01-01',N'확정')", target)
            cur.execute("SELECT bom_id FROM nx.bom_header WHERE item_code=?", target); bom_id = cur.fetchone()[0]
        cur.execute("DELETE FROM nx.bom_line WHERE bom_id=?", bom_id)
        for seq, r in enumerate(rows, 1):
            cur.execute("""INSERT INTO nx.bom_line
                (bom_id,seq,child_item,qty,node_type,cs_calc_except,lme_except,sagub_default,is_optional,
                 from_ymd,to_ymd,except_flag,set_except,kitting,vir_item,proc_gubun,gagong_proc,s_work,wh_gagong,in_gagong,cust_code,remarks)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                bom_id, seq, r[0], float(r[1] or 0), r[2], _b(r[3]), _b(r[4]), _b(r[5]), _b(r[6]),
                r[7], r[8], _b(r[9]), _b(r[10]), _b(r[11]), _b(r[12]), r[13], r[14], r[15], r[16], r[17], r[18], r[19])
        return {"ok": True, "count": len(rows), "source_from": src, "warn": warn}
    finally:
        cn.close()


# ===================== 공정(routing) = 개발 정본(품목별 공정관리, nx.routing=CS_T_ITEM_PROC) CRUD + 복사 =====================
#  개발이 공정을 지정. 없으면 내부공정 복사 → 수정 → 등록. 협력사견적은 공정 기준 아님(단가).
@app.get("/api/routing/get")
def routing_get(item: str = Query(...)):
    """품목 공정 목록(개발 정본 nx.routing). 전체 공정마스터(CS_M_PROC<90) + 해당 품목 ST(work_qty) 병합 → 편집용."""
    item = item.strip()
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT proc_code, ISNULL(work_qty,0), ISNULL(prod_uph,0), ISNULL(sort_seq,0), ISNULL(calc_gubun,''), ISNULL(p_item,'') FROM nx.routing WHERE item_code=?", item)
        cur_rows = {str(r[0]).strip(): {"work_qty": float(r[1] or 0), "prod_uph": float(r[2] or 0), "sort_seq": int(r[3] or 0),
                                        "calc_gubun": str(r[4] or '').strip(), "p_item": str(r[5] or '').strip()} for r in cur.fetchall()}
    finally:
        nx.close()
    # 공정마스터(이름·정렬), proc_code<90(관리/운반/이윤 제외)
    cn = _conn(); c2 = cn.cursor()
    try:
        c2.execute("SELECT PROC_CODE, PROC_DESC, ISNULL(SORT_SEQ,0), ISNULL(PROD_UPH,0) FROM CS_M_PROC WHERE ISNULL(TRY_CONVERT(int,PROC_CODE),99)<90 AND ISNULL(USE_FLAG,'1')<>'0' ORDER BY ISNULL(SORT_SEQ,0), PROC_CODE")
        procs = []
        for r in c2.fetchall():
            pc = str(r[0]).strip(); ex = cur_rows.get(pc, {})
            procs.append({"proc_code": pc, "name": str(r[1]).strip(), "work_qty": ex.get("work_qty", 0),
                          "prod_uph": ex.get("prod_uph") or float(r[3] or 0),   # 품목ST 없으면 공정마스터 기본 UPH
                          "calc_gubun": ex.get("calc_gubun", "") or "3", "p_item": ex.get("p_item", "")})  # calc_gubun 기본 3(임율기반)
    finally:
        cn.close()
    return {"item": item, "procs": procs, "has_routing": any(p["work_qty"] > 0 for p in procs)}

@app.post("/api/routing/save")
def routing_save(payload: dict = Body(...)):
    """품목 공정 등록/수정(개발 정본 nx.routing). rows=[{proc_code, work_qty, prod_uph, calc_gubun, p_item}]. work_qty>0만 저장. 전체교체.
       ★calc_gubun(계산구분 3임율/8중량/9적용율) + p_item(공정귀속, 일반=''·은납품=부모) 저장 필수 — 미저장시 엔진 가공비=0."""
    item = str(payload.get("item", "")).strip()
    rows = payload.get("rows", []) or []
    if not item:
        raise HTTPException(400, "item 필요")
    keep = [(str(r.get("proc_code", "")).strip(), float(r.get("work_qty") or 0), float(r.get("prod_uph") or 0),
             (str(r.get("calc_gubun", "")).strip() or "3"), str(r.get("p_item", "")).strip())
            for r in rows if str(r.get("proc_code", "")).strip() and float(r.get("work_qty") or 0) > 0]
    nx = _nx(); cur = nx.cursor()
    try:
        # 가공공정(proc<90)만 교체 — 91/92/93/98/99(일반율·운반·이윤율 overhead) 행은 보존(공정그리드 미노출·율마스터 성격)
        cur.execute("DELETE FROM nx.routing WHERE item_code=? AND ISNULL(TRY_CONVERT(int,proc_code),99)<90", item)
        for seq, (pc, wq, uph, cg, pit) in enumerate(keep, 1):
            cur.execute("INSERT INTO nx.routing(p_item, item_code, proc_code, work_qty, prod_uph, calc_gubun, sort_seq) VALUES(?,?,?,?,?,?,?)", pit, item, pc, wq, uph, cg, seq * 10)
        _reset_cost_engine()   # 공정 변경 → 원가엔진 캐시 무효화(다음 계산 최신 반영)
        return {"ok": True, "count": len(keep)}
    finally:
        nx.close()

@app.post("/api/routing/copy")
def routing_copy(payload: dict = Body(...)):
    """내부공정 복사: source 품목의 공정(nx.routing)을 target으로 복사(전체교체). 개발이 유사품 공정을 빠르게 지정."""
    source = str(payload.get("source", "")).strip()
    target = str(payload.get("target", "")).strip()
    if not source or not target or source == target:
        raise HTTPException(400, "source·target(서로 다른) 필요")
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT proc_code, ISNULL(work_qty,0), ISNULL(prod_uph,0), ISNULL(sort_seq,0) FROM nx.routing WHERE item_code=? AND ISNULL(work_qty,0)>0 ORDER BY sort_seq", source)
        src = cur.fetchall()
        if not src:
            return {"ok": False, "error": f"원본 {source} 에 공정(ST>0)이 없습니다."}
        cur.execute("DELETE FROM nx.routing WHERE item_code=?", target)
        for r in src:
            cur.execute("INSERT INTO nx.routing(p_item, item_code, proc_code, work_qty, prod_uph, sort_seq) VALUES(?,?,?,?,?,?)", target, target, str(r[0]).strip(), float(r[1] or 0), float(r[2] or 0), int(r[3] or 0))
        _reset_cost_engine()   # 공정 변경 → 원가엔진 캐시 무효화
        return {"ok": True, "count": len(src)}
    finally:
        nx.close()


# ===================== 자재 재고 (nx.stock_ledger 통합원장) =====================
# 조정/입고/출고 3화면 = 동일 원장, MAINT_TAG 프리셋으로 구분. 출고는 음수저장(양수표시).
STOCK_SCREENS = {
    "adjust":  {"name": "자재개별재고조정", "tags": ["1", "2", "3", "A"], "sign": 0},   # ± 조정
    "receipt": {"name": "자재입고관리",     "tags": ["9", "S", "C", "G", "H"], "sign": 1},  # + 입고
    "issue":   {"name": "자재출고관리",     "tags": ["4"], "sign": -1},                 # - 출고(양수표시)
    "return":  {"name": "자재반품",         "tags": ["RT"], "sign": -1},               # - 반품(≤현재고 가드, 다음공정 이동분은 이미 재고감소=반품불가)
}

def _ym(ymd):  # MAINT_YMD(YYMMDD/YYYYMMDD) → 마감월 YYMM
    y = str(ymd or "").strip()
    return y[:4] if len(y) >= 6 else ""

@app.get("/api/stock/list")
def stock_list(screen: str = Query("adjust"), ymd_from: str = Query(...), ymd_to: str = Query(...),
               q: str = Query("")):
    sc = STOCK_SCREENS.get(screen)
    if not sc:
        raise HTTPException(400, "screen 오류")
    cn = _nx(); cur = cn.cursor()
    try:
        tags = "','".join(sc["tags"])
        like = f"%{q.strip()}%"
        sign = "-1" if sc["sign"] == -1 else "1"
        cur.execute(f"""
            SELECT TOP 500 l.MAINT_YMD, l.MAINT_SEQ, l.MAINT_TAG, tg.name AS tag_name,
                   l.CUST_CODE, pc.partner_name AS cust_name, l.GAGONG_PROC_CODE,
                   l.MAT_CODE, i.item_name, i.item_spec, l.ITEM_CODE,
                   (l.MAINT_QTY * {sign}) AS qty, l.MAINT_COST, l.MAINT_AMT, l.REMARKS,
                   l.SHEET_NO, l.INSP_FLAG, l.WORK_CODE, l.TO_GAGONG_PROC_CODE, l.OUT_WH_GUBUN,
                   l.INSERT_USER_ID, l.INSERT_DATETIME
            FROM nx.stock_ledger l
            LEFT JOIN nx.item i ON i.item_code = l.MAT_CODE
            LEFT JOIN nx.partner pc ON pc.partner_code = l.CUST_CODE
            LEFT JOIN nx.stock_tag tg ON tg.tag = l.MAINT_TAG
            WHERE l.STOCK_POINT='MAT' AND l.MAINT_YMD BETWEEN ? AND ? AND l.MAINT_TAG IN ('{tags}')
              AND (? = '%%' OR l.MAT_CODE LIKE ? OR l.CUST_CODE LIKE ?)
            ORDER BY l.MAINT_YMD DESC, l.MAINT_SEQ DESC""",
            ymd_from.strip(), ymd_to.strip(), like, like, like)
        cols = [d[0] for d in cur.description]
        rows = [{c: (v.isoformat() if hasattr(v, "isoformat") else v) for c, v in zip(cols, r)} for r in cur.fetchall()]
        return {"screen": screen, "name": sc["name"], "sign": sc["sign"], "rows": rows}
    finally:
        cn.close()

@app.post("/api/stock/save")
def stock_save(payload: dict = Body(...)):
    """재고원장 저장(신규행 insert). 가드: 마감월 잠금·FK·출고 재고부족(음수방지)."""
    screen = str(payload.get("screen", "")).strip()
    rows = payload.get("rows", []) or []
    sc = STOCK_SCREENS.get(screen)
    if not sc:
        raise HTTPException(400, "screen 오류")
    cn = _nx(); cur = cn.cursor()
    try:
        # 마감월 집합
        cur.execute("SELECT ym FROM nx.stock_close WHERE close_flag=1")
        closed = {str(r[0]).strip() for r in cur.fetchall()}   # ym=char(6) 패딩 제거(_ym은 4자 → 집합비교 일치)
        errs = []
        for idx, r in enumerate(rows, 1):
            ymd = str(r.get("MAINT_YMD", "")).strip()
            mat = str(r.get("MAT_CODE", "")).strip()
            qty = float(r.get("qty") or 0)
            if not ymd or len(ymd) < 6:
                errs.append(f"{idx}행: 일자 필요"); continue
            if _ym(ymd) in closed:
                errs.append(f"{idx}행: 마감월({_ym(ymd)}) 편집 불가")
            if not mat:
                errs.append(f"{idx}행: 자도번 필요"); continue
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", mat)
            if not cur.fetchone():
                errs.append(f"{idx}행: 미등록 품목({mat})")
            # 조정=부호입력 허용(불량·개발불출 −, 장부수정 ±), 그 외=양수만
            if screen == "adjust":
                if qty == 0:
                    errs.append(f"{idx}행: 조정수량은 0일 수 없습니다(증가 +, 감소 −)")
            elif qty <= 0:
                errs.append(f"{idx}행: 수량은 0보다 커야 함")
            # 재고 음수방지: 출고·반품(가용 이내) / 조정 감소(결과재고 ≥ 0). 현재고=원장 SUM.
            if mat and screen in ("issue", "return"):
                cur.execute("SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE MAT_CODE=?", mat)
                avail = float(cur.fetchone()[0] or 0)
                if qty > avail:
                    lbl = "반품" if screen == "return" else "출고"
                    errs.append(f"{idx}행: 재고부족 ({mat} 가용 {avail:g} < {lbl} {qty:g}) — 다음공정 이동분은 반품 불가")
            elif mat and screen == "adjust" and qty < 0:
                cur.execute("SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE MAT_CODE=?", mat)
                avail = float(cur.fetchone()[0] or 0)
                if avail + qty < 0:
                    errs.append(f"{idx}행: 음수재고 유발 ({mat} 결과재고 {avail+qty:g} < 0)")
        if errs:
            return {"ok": False, "errors": errs}
        # insert (일자별 SEQ 채번, 출고 음수 저장)
        saved = 0
        for r in rows:
            ymd = str(r.get("MAINT_YMD", "")).strip()
            cur.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd)
            seq = cur.fetchone()[0]
            tag = str(r.get("MAINT_TAG") or sc["tags"][0]).strip()
            qty = float(r.get("qty") or 0)
            store_qty = -abs(qty) if sc["sign"] == -1 else qty
            cur.execute("""INSERT INTO nx.stock_ledger
                (STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,CUST_CODE,GAGONG_PROC_CODE,TO_GAGONG_PROC_CODE,OUT_WH_GUBUN,
                 MAT_CODE,ITEM_CODE,WORK_CODE,MAINT_QTY,MAINT_COST,MAINT_AMT,REMARKS,SHEET_NO,INSERT_USER_ID,INSERT_DATETIME)
                VALUES('MAT',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,GETDATE())""",
                ymd, seq, tag, (r.get("CUST_CODE") or None), (r.get("GAGONG_PROC_CODE") or None),
                (r.get("TO_GAGONG_PROC_CODE") or None), (r.get("OUT_WH_GUBUN") or None),
                str(r.get("MAT_CODE", "")).strip(), (r.get("ITEM_CODE") or None), (r.get("WORK_CODE") or None),
                store_qty, float(r.get("MAINT_COST") or 0), float(r.get("MAINT_AMT") or 0),
                (r.get("REMARKS") or None), (r.get("SHEET_NO") or None), "web")
            saved += 1
        return {"ok": True, "count": saved}
    finally:
        cn.close()

@app.get("/api/stock/kanban")
def stock_kanban(q: str = Query("")):
    """자재입고진행현황(읽기전용 집계): 품목별 현재고=원장 SUM. 계획대비는 추후 확장."""
    cn = _nx(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""
            SELECT TOP 300 l.MAT_CODE, i.item_name, i.item_spec, MAX(l.GAGONG_PROC_CODE) AS part,
                   SUM(l.MAINT_QTY) AS stock_qty,
                   SUM(CASE WHEN l.MAINT_QTY>0 THEN l.MAINT_QTY ELSE 0 END) AS in_qty,
                   SUM(CASE WHEN l.MAINT_QTY<0 THEN -l.MAINT_QTY ELSE 0 END) AS out_qty
            FROM nx.stock_ledger l LEFT JOIN nx.item i ON i.item_code=l.MAT_CODE
            WHERE l.STOCK_POINT='MAT' AND (? = '%%' OR l.MAT_CODE LIKE ?)
            GROUP BY l.MAT_CODE, i.item_name, i.item_spec
            HAVING SUM(l.MAINT_QTY) <> 0
            ORDER BY SUM(l.MAINT_QTY) DESC""", like, like)
        cols = [d[0] for d in cur.description]
        rows = [{c: v for c, v in zip(cols, r)} for r in cur.fetchall()]
        return {"rows": rows}
    finally:
        cn.close()

# ============ 자재입고: 발주분 입고(057 개별일괄 / 057_1 PO바코드) — 발주잔량 차감·nx.stock_ledger 기록 ============
# 발주잔량 = 발주(PU_T_PURCHASE_DTL.PUR_QTY) − 레거시기입고(IN_QTY) − 취소(CANCEL_QTY) − nx웹입고(발주링크 SUM). [[nextgen-erp-ledger-consistency]] 원장파생.
@app.get("/api/matrecv/po_pending")
def matrecv_po_pending(cust: str = Query(""), item: str = Query(""), sheet: str = Query(""),
                       from_ymd: str = Query(""), to_ymd: str = Query("")):
    """발주분 입고대기(발주잔량>0). 개별일괄/PO바코드 공용. sheet=발주번호(PUR_SEQ, 바코드 PO뒤 숫자)."""
    C = " COLLATE DATABASE_DEFAULT"
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["(p.PUR_QTY - ISNULL(p.IN_QTY,0) - ISNULL(p.CANCEL_QTY,0)) > 0"]; pr = []
        if cust.strip(): w.append("p.CUST_CODE LIKE ?"); pr.append(f"%{cust.strip()}%")
        if item.strip(): w.append("p.ITEM_CODE LIKE ?"); pr.append(f"%{item.strip()}%")
        if sheet.strip(): w.append("p.PUR_SEQ=?"); pr.append(sheet.strip())
        if from_ymd.strip(): w.append("p.PUR_YMD>=?"); pr.append(_d6(from_ymd))
        if to_ymd.strip(): w.append("p.PUR_YMD<=?"); pr.append(_d6(to_ymd))
        cur.execute(f"""SELECT TOP 800 p.PUR_YMD, p.PUR_SEQ, p.PUR_SEQ_ROW, p.CUST_CODE, ISNULL(cu.CUST_DESC,'') cust_nm,
              p.ITEM_CODE, ISNULL(it.ITEM_DESC,'') nm, ISNULL(it.ITEM_SPEC,'') spec, ISNULL(it.UNIT,'') unit, p.DLVY_YMD,
              p.PUR_QTY, ISNULL(p.IN_QTY,0) in_qty, ISNULL(p.CANCEL_QTY,0) cancel_qty, ISNULL(nx.q,0) nx_in,
              ISNULL(p.PUR_COST,0) pur_cost, ISNULL(p.MAT_INSPECTION,'') insp,
              (p.PUR_QTY - ISNULL(p.IN_QTY,0) - ISNULL(p.CANCEL_QTY,0) - ISNULL(nx.q,0)) remain
            FROM PARTNER_ERP.dbo.PU_T_PURCHASE_DTL p
            LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST cu ON cu.CUST_CODE=p.CUST_CODE
            LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM it ON it.ITEM_CODE=p.ITEM_CODE
            LEFT JOIN (SELECT PUR_YMD,PUR_SEQ,PUR_SEQ_ROW,SUM(MAINT_QTY) q FROM nx.stock_ledger
                       WHERE MAINT_TAG='9' AND ISNULL(PUR_YMD,'')<>'' GROUP BY PUR_YMD,PUR_SEQ,PUR_SEQ_ROW) nx
              ON nx.PUR_YMD{C}=p.PUR_YMD{C} AND nx.PUR_SEQ{C}=p.PUR_SEQ{C} AND nx.PUR_SEQ_ROW=p.PUR_SEQ_ROW
            WHERE {' AND '.join(w)}
            ORDER BY p.PUR_YMD DESC, p.PUR_SEQ, p.PUR_SEQ_ROW""", *pr)
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            d["remain"] = float(d["remain"] or 0)
            if d["remain"] <= 0:  # nx입고로 이미 충족분 제외
                continue
            for k in ("PUR_QTY", "in_qty", "cancel_qty", "nx_in", "pur_cost"):
                d[k] = float(d[k] or 0)
            d["DLVY_YMD"] = str(d["DLVY_YMD"] or "")
            rows.append(d)
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

@app.post("/api/matrecv/receive")
def matrecv_receive(payload: dict = Body(...)):
    """발주분 입고 확정 → nx.stock_ledger(MAINT_TAG='9', 발주링크). 마감월/미등록품목/발주잔량초과 가드.
    입고수량이 발주잔량 초과시 차단. 삭제는 /api/stock/delete(원장행 제거=역진행)."""
    ymd = str(payload.get("ymd") or "").strip()      # 입고일자 YYMMDD
    wh = str(payload.get("wh") or "IS0001").strip()   # 입고창고(gagong_proc_code)
    rows = payload.get("rows", []) or []
    if not ymd or len(ymd) < 6:
        raise HTTPException(400, "입고일자 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        if _closed(cur, ymd):
            return {"ok": False, "errors": [f"마감월({_ym(ymd)}) 입고 불가"]}
        errs = []
        # 검증: 품목등록·발주잔량 초과
        for idx, r in enumerate(rows, 1):
            item = str(r.get("item", "")).strip(); qty = float(r.get("qty") or 0)
            py = str(r.get("pur_ymd", "")).strip(); ps = str(r.get("pur_seq", "")).strip(); prw = r.get("pur_seq_row")
            if qty <= 0: errs.append(f"{idx}행: 입고수량>0 필요"); continue
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", item)
            if not cur.fetchone(): errs.append(f"{idx}행: 미등록품목({item})")
            if py and ps and prw is not None:  # 발주잔량 초과 가드
                cur.execute("""SELECT (p.PUR_QTY-ISNULL(p.IN_QTY,0)-ISNULL(p.CANCEL_QTY,0)
                    -ISNULL((SELECT SUM(MAINT_QTY) FROM nx.stock_ledger WHERE MAINT_TAG='9' AND PUR_YMD=? AND PUR_SEQ=? AND PUR_SEQ_ROW=?),0))
                    FROM PARTNER_ERP.dbo.PU_T_PURCHASE_DTL p WHERE p.PUR_YMD=? AND p.PUR_SEQ=? AND p.PUR_SEQ_ROW=?""",
                    py, ps, int(prw), py, ps, int(prw))
                rem = cur.fetchone()
                remain = float(rem[0]) if rem and rem[0] is not None else None
                if remain is not None and qty > remain + 0.001:
                    errs.append(f"{idx}행: 발주잔량 초과({item} 잔량 {remain:g} < 입고 {qty:g})")
        if errs:
            return {"ok": False, "errors": errs}
        saved = 0
        for r in rows:
            item = str(r.get("item", "")).strip(); qty = float(r.get("qty") or 0)
            if qty <= 0: continue
            cost = float(r.get("cost") or 0); vat = float(r.get("vat") or round(qty * cost * 0.1))
            py = (str(r.get("pur_ymd", "")).strip() or None); ps = (str(r.get("pur_seq", "")).strip() or None)
            prw = r.get("pur_seq_row")
            cur.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd)
            seq = cur.fetchone()[0]
            cur.execute("""INSERT INTO nx.stock_ledger
                (STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,CUST_CODE,GAGONG_PROC_CODE,MAT_CODE,MAINT_QTY,MAINT_COST,MAINT_AMT,MAINT_VAT,
                 PUR_YMD,PUR_SEQ,PUR_SEQ_ROW,INSP_FLAG,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
                VALUES('MAT',?,?, '9',?,?,?,?,?,?,?, ?,?,?,?,?, ?, GETDATE())""",
                ymd, seq, (r.get("cust") or None), wh, item, qty, cost, round(qty * cost), vat,
                py, ps, (int(prw) if prw is not None else None), (r.get("insp") or None),
                (r.get("remarks") or "발주입고"), "web")
            saved += 1
        return {"ok": True, "count": saved}
    finally:
        cn.close()

# ============ 자재입고: 가공이동전표 입고(057_2 바코드) — PU_T_STOCK_MAINT_GAGONG_MOVE → nx.stock_ledger(tag C) ============
@app.get("/api/matrecv/gagong_pending")
def matrecv_gagong_pending(sheet: str = Query(""), item: str = Query("")):
    """가공이동전표 미입고분. sheet=바코드(MV+MAINT_GROUP_SEQ). 미입고=IN_CONFIRM_FLAG≠1 & nx웹입고 미충족."""
    C = " COLLATE DATABASE_DEFAULT"
    cn = _nx(); cur = cn.cursor()
    try:
        w = ["ISNULL(g.IN_CONFIRM_FLAG,'0')<>'1'"]; pr = []
        if sheet.strip():
            gs = "".join(ch for ch in sheet.strip() if ch.isdigit())
            w.append("g.MAINT_GROUP_SEQ=?"); pr.append(int(gs) if gs else -1)
        if item.strip(): w.append("g.MAT_CODE LIKE ?"); pr.append(f"%{item.strip()}%")
        cur.execute(f"""SELECT TOP 500 g.MAINT_GROUP_SEQ, g.MAT_CODE, ISNULL(it.ITEM_DESC,'') nm, ISNULL(it.ITEM_SPEC,'') spec,
              ISNULL(it.UNIT,'') unit, g.ITEM_CODE upper_code, g.MAINT_QTY, g.GAGONG_PROC_CODE, g.TO_GAGONG_PROC_CODE,
              g.MAINT_YMD, ISNULL(nx.q,0) nx_in
            FROM PARTNER_ERP.dbo.PU_T_STOCK_MAINT_GAGONG_MOVE g
            LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM it ON it.ITEM_CODE{C}=g.MAT_CODE{C}
            LEFT JOIN (SELECT MAINT_GROUP_SEQ, SUM(MAINT_QTY) q FROM nx.stock_ledger WHERE MAINT_TAG='C' AND MAINT_GROUP_SEQ IS NOT NULL GROUP BY MAINT_GROUP_SEQ) nx
              ON nx.MAINT_GROUP_SEQ=g.MAINT_GROUP_SEQ
            WHERE {' AND '.join(w)}
            ORDER BY g.MAINT_YMD DESC, g.MAINT_GROUP_SEQ DESC""", *pr)
        cols = [d[0] for d in cur.description]; rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            d["MAINT_QTY"] = float(d["MAINT_QTY"] or 0); d["nx_in"] = float(d["nx_in"] or 0)
            d["remain"] = d["MAINT_QTY"] - d["nx_in"]
            d["MAINT_YMD"] = str(d["MAINT_YMD"] or "")
            if d["remain"] <= 0: continue
            rows.append(d)
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

@app.post("/api/matrecv/gagong_receive")
def matrecv_gagong_receive(payload: dict = Body(...)):
    """가공이동전표 입고 확정 → nx.stock_ledger(MAINT_TAG='C', MAINT_GROUP_SEQ 링크, 입고창고=TO_GAGONG_PROC_CODE). 마감 가드."""
    ymd = str(payload.get("ymd") or "").strip()
    rows = payload.get("rows", []) or []
    if not ymd or len(ymd) < 6:
        raise HTTPException(400, "입고일자 필요")
    cn = _nx(); cur = cn.cursor()
    try:
        if _closed(cur, ymd):
            return {"ok": False, "errors": [f"마감월({_ym(ymd)}) 입고 불가"]}
        errs = []; saved = 0
        for idx, r in enumerate(rows, 1):
            item = str(r.get("item", "")).strip(); qty = float(r.get("qty") or 0)
            if qty <= 0: errs.append(f"{idx}행: 입고수량>0 필요"); continue
            cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", item)
            if not cur.fetchone(): errs.append(f"{idx}행: 미등록품목({item})")
        if errs:
            return {"ok": False, "errors": errs}
        for r in rows:
            item = str(r.get("item", "")).strip(); qty = float(r.get("qty") or 0)
            if qty <= 0: continue
            gseq = r.get("group_seq"); wh = str(r.get("to_gagong") or r.get("wh") or "IS0001").strip()
            cur.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd)
            seq = cur.fetchone()[0]
            cur.execute("""INSERT INTO nx.stock_ledger
                (STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_GROUP_SEQ,MAINT_TAG,GAGONG_PROC_CODE,TO_GAGONG_PROC_CODE,MAT_CODE,ITEM_CODE,MAINT_QTY,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
                VALUES('MAT',?,?,?, 'C', ?,?,?,?,?, ?, ?, GETDATE())""",
                ymd, seq, (int(gseq) if gseq is not None else None), wh, (r.get("gagong") or None),
                item, (r.get("upper") or None), qty, (r.get("remarks") or "가공이동입고"), "web")
            saved += 1
        return {"ok": True, "count": saved}
    finally:
        cn.close()

def _closed(cur, ymd):
    cur.execute("SELECT close_flag FROM nx.stock_close WHERE ym=?", _ym(ymd))
    r = cur.fetchone()
    return bool(r and r[0])

# ===================== ★Phase5: nx 재고 월마감 스냅샷 (STOCK_POINT별 기초→기말=기초+ΣMAINT) =====================
# 기말 스냅샷=다음달 기초 연속성·마감후 파생 고정. 잠금=기존 nx.stock_close(ym) 플래그 재사용(옵션).
# ★사고 재발방지: stock_ledger 무삭제. 재계산은 자기생성 근거키(ym+point)의 stock_close_snap만 갱신.
@app.post("/api/stockclose/run")
def stockclose_run(payload: dict = Body(...)):
    """월마감 실행: (ym, point) 기말 스냅샷 산출·저장(set-based, 고속). 기초=단일원장 누적(<ym01, =직전월기말 동치). 기말=기초+입−출.
    lock=true면 nx.stock_close(ym) 잠금 플래그 set(기존 가드 발동=이전 원장 쓰기잠금). 멱등(같은 ym+point 재실행=재계산).
    ★사고 재발방지: stock_ledger 무삭제. 정리는 자기생성 근거키(ym+point)의 stock_close_snap만."""
    ym = str(payload.get("ym", "")).strip()
    point = str(payload.get("point", "")).strip().upper()
    lock = bool(payload.get("lock", False))
    user = (str(payload.get("user", "")).strip() or "web")[:30]
    if len(ym) != 4 or point not in ("MAT", "PRD", "ASY", "RDY", "SAG"):
        raise HTTPException(400, "ym=YYMM(4자)·point=MAT/PRD/ASY/RDY/SAG 필수")
    y01, y99 = ym + "01", ym + "99"
    cn = _nx(); cur = cn.cursor()
    try:
        # 자기생성 근거키(ym+point) 재계산분만 제거(멱등 — 이 마감의 스냅샷만)
        cur.execute("DELETE FROM nx.stock_close_snap WHERE ym=? AND stock_point=?", ym, point)
        # set-based: RTRIM 정규화 GROUP BY(후행공백 PK중복 방지), 기초=Σ(<y01)·입출=당월. 기말=기초+입−출.
        cur.execute("""INSERT INTO nx.stock_close_snap(ym,stock_point,item_key,gpc,cust,base_qty,in_qty,out_qty,end_qty,close_user,close_dt)
            SELECT ?, ?, LEFT(t.k,40), LEFT(t.g,20), LEFT(t.c,10), t.base, t.inq, t.outq, t.base+t.inq-t.outq, ?, GETDATE()
            FROM (
              SELECT COALESCE(NULLIF(RTRIM(L.MAT_CODE),''),RTRIM(L.ITEM_CODE)) k,
                     RTRIM(ISNULL(L.GAGONG_PROC_CODE,'')) g, RTRIM(ISNULL(L.CUST_CODE,'')) c,
                     SUM(CASE WHEN L.MAINT_YMD<? THEN L.MAINT_QTY ELSE 0 END) base,
                     SUM(CASE WHEN L.MAINT_YMD BETWEEN ? AND ? AND L.MAINT_QTY>0 THEN L.MAINT_QTY ELSE 0 END) inq,
                     SUM(CASE WHEN L.MAINT_YMD BETWEEN ? AND ? AND L.MAINT_QTY<0 THEN -L.MAINT_QTY ELSE 0 END) outq
              FROM nx.stock_ledger L WHERE L.STOCK_POINT=?
              GROUP BY COALESCE(NULLIF(RTRIM(L.MAT_CODE),''),RTRIM(L.ITEM_CODE)), RTRIM(ISNULL(L.GAGONG_PROC_CODE,'')), RTRIM(ISNULL(L.CUST_CODE,''))
            ) t
            WHERE (t.base<>0 OR t.inq<>0 OR t.outq<>0) AND t.k IS NOT NULL AND t.k<>''""",
            ym, point, user, y01, y01, y99, y01, y99, point)
        n = cur.rowcount
        if lock:  # 기존 nx.stock_close(ym) 플래그 재사용(신설 아님) — 이전 원장 쓰기잠금 발동
            cur.execute("IF EXISTS(SELECT 1 FROM nx.stock_close WHERE ym=?) UPDATE nx.stock_close SET close_flag=1,close_user=?,close_dt=GETDATE() WHERE ym=? ELSE INSERT INTO nx.stock_close(ym,close_flag,close_user,close_dt) VALUES(?,1,?,GETDATE())", ym, user, ym, ym, user)
        cur.execute("SELECT ISNULL(SUM(end_qty),0) FROM nx.stock_close_snap WHERE ym=? AND stock_point=?", ym, point)
        endsum = float(cur.fetchone()[0] or 0)
        return {"ok": True, "ym": ym, "point": point, "rows": n, "end_total": round(endsum, 3),
                "base_from": "단일원장 누적(<ym01 = 직전월기말 동치)", "locked": lock}
    finally:
        cn.close()

@app.get("/api/stockclose/status")
def stockclose_status(ym: str = Query(""), point: str = Query("")):
    """마감 현황: 스냅샷 요약(ym·point별 행수·기말합) + 잠금 플래그."""
    cn = _nx(); cur = cn.cursor()
    try:
        w = []; p = []
        if ym.strip(): w.append("ym=?"); p.append(ym.strip())
        if point.strip(): w.append("stock_point=?"); p.append(point.strip().upper())
        wh = ("WHERE " + " AND ".join(w)) if w else ""
        cur.execute(f"""SELECT ym, stock_point, COUNT(*) rows, ISNULL(SUM(end_qty),0) end_total, MAX(close_dt) close_dt
            FROM nx.stock_close_snap {wh} GROUP BY ym, stock_point ORDER BY ym DESC, stock_point""", *p)
        snaps = [{"ym": r[0], "point": r[1], "rows": int(r[2]), "end_total": round(float(r[3] or 0), 3),
                  "close_dt": str(r[4] or "")[:19]} for r in cur.fetchall()]
        cur.execute("SELECT ym, close_flag, close_user, close_dt FROM nx.stock_close ORDER BY ym DESC")
        locks = [{"ym": str(r[0]).strip(), "locked": bool(r[1]), "user": r[2], "dt": str(r[3] or "")[:19]} for r in cur.fetchall()]
        return {"snapshots": snaps, "locks": locks}
    finally:
        cn.close()

@app.post("/api/stock/update")
def stock_update(payload: dict = Body(...)):
    """기존 원장행 수정(값 필드만). 키(MAINT_YMD,MAINT_SEQ)·자도번 불변, 저장부호 보존.
    가드: 대상존재·마감월 잠금·수량>0·음수재고 유발 차단."""
    screen = str(payload.get("screen", "")).strip()
    sc = STOCK_SCREENS.get(screen)
    if not sc:
        raise HTTPException(400, "screen 오류")
    ymd = str(payload.get("MAINT_YMD", "")).strip()
    try:
        seq = int(payload.get("MAINT_SEQ"))
    except (TypeError, ValueError):
        raise HTTPException(400, "MAINT_SEQ 오류")
    qty = float(payload.get("qty") or 0)
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT MAT_CODE, MAINT_QTY FROM nx.stock_ledger WHERE MAINT_YMD=? AND MAINT_SEQ=?", ymd, seq)
        row = cur.fetchone()
        if not row:
            return {"ok": False, "errors": [f"대상 없음 ({ymd}/{seq})"]}
        mat = str(row[0] or "").strip()
        old_stored = float(row[1] or 0)
        errs = []
        if _closed(cur, ymd):
            errs.append(f"마감월({_ym(ymd)}) 편집 불가")
        if screen == "adjust":
            # 조정=부호 그대로(불량·개발불출 −, 장부수정 ±). 음수재고는 아래 new_sum 검증에서 차단.
            if qty == 0:
                errs.append("조정수량은 0일 수 없습니다(증가 +, 감소 −)")
            new_stored = qty
        else:
            if qty <= 0:
                errs.append("수량은 0보다 커야 함")
            # 저장부호 보존(기존 음수→음수, 0이면 화면부호)
            neg = old_stored < 0 or (old_stored == 0 and sc["sign"] == -1)
            new_stored = -abs(qty) if neg else abs(qty)
        # 음수재고 유발 차단(악화 시에만)
        cur.execute("SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE MAT_CODE=?", mat)
        cur_sum = float(cur.fetchone()[0] or 0)
        new_sum = cur_sum - old_stored + new_stored
        if new_sum < 0 and new_sum < cur_sum:
            errs.append(f"음수재고 유발 ({mat} 결과재고 {new_sum:g} < 0)")
        if errs:
            return {"ok": False, "errors": errs}
        cur.execute("""UPDATE nx.stock_ledger
            SET MAINT_QTY=?, MAINT_TAG=?, CUST_CODE=?, GAGONG_PROC_CODE=?, REMARKS=?,
                UPDATE_USER_ID=?, UPDATE_DATETIME=GETDATE()
            WHERE MAINT_YMD=? AND MAINT_SEQ=?""",
            new_stored, (str(payload.get("MAINT_TAG") or sc["tags"][0]).strip()),
            (str(payload.get("CUST_CODE") or "").strip() or None),
            (str(payload.get("GAGONG_PROC_CODE") or "").strip() or None),
            (str(payload.get("REMARKS") or "").strip() or None),
            "web", ymd, seq)
        return {"ok": True, "stored_qty": new_stored, "stock": new_sum}
    finally:
        cn.close()

@app.post("/api/stock/delete")
def stock_delete(payload: dict = Body(...)):
    """기존 원장행 삭제. 가드: 대상존재·마감월 잠금·삭제 시 음수재고 유발 차단(입고행 삭제로 재고<0 방지)."""
    ymd = str(payload.get("MAINT_YMD", "")).strip()
    try:
        seq = int(payload.get("MAINT_SEQ"))
    except (TypeError, ValueError):
        raise HTTPException(400, "MAINT_SEQ 오류")
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT MAT_CODE, MAINT_QTY FROM nx.stock_ledger WHERE MAINT_YMD=? AND MAINT_SEQ=?", ymd, seq)
        row = cur.fetchone()
        if not row:
            return {"ok": False, "errors": [f"대상 없음 ({ymd}/{seq})"]}
        mat = str(row[0] or "").strip()
        old_stored = float(row[1] or 0)
        errs = []
        if _closed(cur, ymd):
            errs.append(f"마감월({_ym(ymd)}) 삭제 불가")
        cur.execute("SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE MAT_CODE=?", mat)
        cur_sum = float(cur.fetchone()[0] or 0)
        new_sum = cur_sum - old_stored
        if new_sum < 0 and new_sum < cur_sum:
            errs.append(f"음수재고 유발 ({mat} 삭제 후 재고 {new_sum:g} < 0)")
        if errs:
            return {"ok": False, "errors": errs}
        cur.execute("DELETE FROM nx.stock_ledger WHERE MAINT_YMD=? AND MAINT_SEQ=?", ymd, seq)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        cn.close()

# ================= 매출마감처리 (w_pu_sale_020 재설계) — 협력사 매출(tag5) 업체별 마감·조정·사유 =================
def _dig4(s):
    d = "".join(ch for ch in str(s or "") if ch.isdigit())
    return d[2:6] if len(d) == 6 else d[:4]   # YYYYMM→YYMM, YYMM→그대로(방어적)

def _cur_ym():
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT FORMAT(GETDATE(),'yyMM')"); return cur.fetchone()[0]
    finally:
        cn.close()

_SALE_MAGAM = """WITH MAGAM(CUST_CODE,JUN_YYMM,JUN_MAGAM_DAY,MAGAM_DAY) AS (
  SELECT CUST_CODE, format(dateadd(MONTH,-1,convert(date,'{ym}'+'01',12)),'yyMM') JUN_YYMM,
    ISNULL((SELECT TOP 1 MAGAM_DAY FROM CM_M_CUST_MAGAM WHERE CUST_CODE=A.CUST_CODE AND APPLY_YYMM<=format(dateadd(MONTH,-1,convert(date,'{ym}'+'01',12)),'yyMM') ORDER BY APPLY_YYMM DESC),'31') JUN_MAGAM_DAY,
    ISNULL((SELECT TOP 1 MAGAM_DAY FROM CM_M_CUST_MAGAM WHERE CUST_CODE=A.CUST_CODE AND APPLY_YYMM<='{ym}' ORDER BY APPLY_YYMM DESC),'31') MAGAM_DAY
  FROM CM_M_CUST A)"""

def _sale_win():
    return "A.MAINT_YMD > mg.JUN_YYMM+mg.JUN_MAGAM_DAY AND A.MAINT_YMD <= '{ym}'+mg.MAGAM_DAY"

@app.get("/api/salemagam/list")
def salemagam_list(ym: str = Query("")):
    """매출마감 업체별 집계(협력사판매 tag5, 마감기준) + nx 마감상태·조정합."""
    y = _dig4(ym) or _cur_ym()
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute(f"""{_SALE_MAGAM.format(ym=y)}
          SELECT A.CUST_CODE cc, MAX(C.CUST_DESC) nm, MAX(C.CUST_TYPE) ct,
            MAX(LTRIM(RTRIM(ISNULL(NULLIF(C.CHARGE_USER_ID,''),ISNULL(C.CHARGE_NAME,''))))) chg,
            SUM(-A.MAINT_QTY) qty, SUM(-A.MAINT_AMT) amt, SUM(-A.MAINT_VAT) vat, COUNT(DISTINCT A.MAT_CODE) items
          FROM PU_T_STOCK_MAINT A JOIN CM_M_CUST C ON A.CUST_CODE=C.CUST_CODE JOIN MAGAM mg ON A.CUST_CODE=mg.CUST_CODE
          WHERE A.MAINT_TAG='5' AND {_sale_win().format(ym=y)}
          GROUP BY A.CUST_CODE HAVING SUM(-A.MAINT_AMT)<>0 ORDER BY SUM(-A.MAINT_AMT) DESC""")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        cn.close()
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("SELECT cust_code,close_flag,bill_flag FROM nx.sale_close WHERE ym=?", y)
        st = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in nc.fetchall()}
        nc.execute("SELECT cust_code, SUM(delta_amt) FROM nx.sale_adjust WHERE ym=? GROUP BY cust_code", y)
        adj = {r[0]: float(r[1] or 0) for r in nc.fetchall()}
    finally:
        nx.close()
    for r in rows:
        cc = r["cc"]; s = st.get(cc, (0, 0))
        r["qty"] = float(r["qty"] or 0); r["amt"] = float(r["amt"] or 0); r["vat"] = float(r["vat"] or 0); r["items"] = int(r["items"] or 0)
        r["close_flag"] = s[0]; r["bill_flag"] = s[1]
        r["adj_amt"] = adj.get(cc, 0.0); r["final_amt"] = round(r["amt"] + adj.get(cc, 0.0), 2)
    return {"ym": y, "rows": rows}

@app.get("/api/salemagam/detail")
def salemagam_detail(ym: str = Query(""), cc: str = Query(...)):
    """업체 마감상세: 품목×일자 피벗 + 저장된 조정내역."""
    y = _dig4(ym) or _cur_ym()
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute(f"""{_SALE_MAGAM.format(ym=y)}
          SELECT A.MAT_CODE mat, MAX(M.ITEM_DESC) nm, MAX(M.ITEM_SPEC) spec, MAX(M.UNIT) unit, A.MAINT_COST cost,
            CAST(RIGHT(A.MAINT_YMD,2) AS INT) d, SUM(-A.MAINT_QTY) q, SUM(-A.MAINT_AMT) amt
          FROM PU_T_STOCK_MAINT A JOIN PR_M_ITEM M ON A.MAT_CODE=M.ITEM_CODE JOIN MAGAM mg ON A.CUST_CODE=mg.CUST_CODE
          WHERE A.MAINT_TAG='5' AND A.CUST_CODE=? AND {_sale_win().format(ym=y)}
          GROUP BY A.MAT_CODE, A.MAINT_COST, CAST(RIGHT(A.MAINT_YMD,2) AS INT)""", cc)
        raw = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    finally:
        cn.close()
    items = {}; days = set()
    for r in raw:
        mat = str(r["mat"]).strip(); d = int(r["d"] or 0); days.add(d)
        it = items.setdefault(mat, {"mat": mat, "nm": r["nm"], "spec": r["spec"], "unit": r["unit"],
                                    "cost": float(r["cost"] or 0), "qty": 0.0, "amt": 0.0, "_bd": {}})
        qv = float(r["q"] or 0); av = float(r["amt"] or 0); cv = float(r["cost"] or 0)
        it["qty"] += qv; it["amt"] += av
        bd = it["_bd"].setdefault(d, {"d": d, "qty": 0.0, "amt": 0.0, "cost": cv})
        bd["qty"] += qv; bd["amt"] += av; bd["cost"] = cv
    for it in items.values():
        it["byday"] = sorted(it.pop("_bd").values(), key=lambda x: x["d"])
    items_list = sorted(items.values(), key=lambda x: -abs(x["amt"]))
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("""SELECT adj_seq,adj_type,scope,mat_code,target_ymd,old_cost,new_cost,old_qty,new_qty,delta_amt,reason_code,reason_detail
                      FROM nx.sale_adjust WHERE ym=? AND cust_code=? ORDER BY adj_seq""", y, cc)
        adjs = [{"adj_type": r[1], "scope": r[2], "mat_code": r[3], "target_ymd": r[4],
                 "old_cost": (float(r[5]) if r[5] is not None else None), "new_cost": (float(r[6]) if r[6] is not None else None),
                 "old_qty": (float(r[7]) if r[7] is not None else None), "new_qty": (float(r[8]) if r[8] is not None else None),
                 "delta_amt": float(r[9] or 0), "reason_code": r[10], "reason_detail": r[11]} for r in nc.fetchall()]
        nc.execute("SELECT close_flag FROM nx.sale_close WHERE ym=? AND cust_code=?", y, cc)
        cr = nc.fetchone(); closed = int(cr[0]) if cr else 0
    finally:
        nx.close()
    return {"ym": y, "cc": cc, "days": sorted(days), "items": items_list, "adjustments": adjs, "close_flag": closed}

@app.get("/api/salemagam/reasons")
def salemagam_reasons():
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("SELECT reason_code,reason_name,category FROM nx.close_reason WHERE use_flag=1 ORDER BY sort_no")
        return {"rows": [{"code": r[0], "name": r[1], "cat": r[2]} for r in nc.fetchall()]}
    finally:
        nx.close()

@app.get("/api/salemagam/custsearch")
def salemagam_custsearch(q: str = Query("")):
    """거래처 단일선택 검색(코드로 구분 — 동명이인 방지). 삼화코리아 2건도 코드로 구분."""
    cn = _conn(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT TOP 30 CUST_CODE, CUST_DESC, CUST_TYPE FROM CM_M_CUST
                       WHERE CUST_CODE LIKE ? OR CUST_DESC LIKE ? ORDER BY CUST_DESC, CUST_CODE""", like, like)
        return {"rows": [{"cc": r[0], "nm": r[1], "ct": r[2]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.post("/api/salemagam/save")
def salemagam_save(payload: dict = Body(...)):
    """조정내역(단가변경/총액증감/품목무관) 저장 + 선택 시 마감. 이미 마감이면 거부."""
    y = _dig4(payload.get("ym")); cc = str(payload.get("cust_code", "")).strip()
    if not y or not cc:
        raise HTTPException(400, "ym/cust_code 필요")
    adjs = payload.get("adjustments", []) or []
    do_close = bool(payload.get("close"))
    base_amt = float(payload.get("base_amt") or 0)
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("SELECT close_flag FROM nx.sale_close WHERE ym=? AND cust_code=?", y, cc)
        r = nc.fetchone()
        if r and r[0]:
            return {"ok": False, "errors": ["이미 마감된 업체 — 마감취소 후 수정하세요"]}
        # 사유 필수(조정이 있으면)
        errs = []
        for i, a in enumerate(adjs, 1):
            if not (a.get("reason_code") or (a.get("reason_detail") or "").strip()):
                errs.append(f"{i}행: 변경사유 필요")
        if errs:
            return {"ok": False, "errors": errs}
        nc.execute("DELETE FROM nx.sale_adjust WHERE ym=? AND cust_code=?", y, cc)
        tot = 0.0
        for i, a in enumerate(adjs, 1):
            d = float(a.get("delta_amt") or 0); tot += d
            nc.execute("""INSERT INTO nx.sale_adjust(ym,cust_code,adj_seq,adj_type,scope,mat_code,target_ymd,old_cost,new_cost,old_qty,new_qty,delta_amt,reason_code,reason_detail,ins_user)
                          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                       y, cc, i, str(a.get("adj_type") or "ITEM_ADJ"), (a.get("scope") or None), (a.get("mat_code") or None),
                       (str(a.get("target_ymd")) if a.get("target_ymd") else None),
                       (a.get("old_cost") if a.get("old_cost") is not None else None),
                       (a.get("new_cost") if a.get("new_cost") is not None else None),
                       (a.get("old_qty") if a.get("old_qty") is not None else None),
                       (a.get("new_qty") if a.get("new_qty") is not None else None),
                       d, (a.get("reason_code") or None), ((a.get("reason_detail") or "").strip() or None), "web")
        final = round(base_amt + tot, 2)
        cf = 1 if do_close else 0
        nc.execute("""MERGE nx.sale_close AS t USING (SELECT ? ym, ? cust_code) s ON t.ym=s.ym AND t.cust_code=s.cust_code
          WHEN MATCHED THEN UPDATE SET close_flag=?, base_amt=?, adj_amt=?, final_amt=?,
             close_user=CASE WHEN ?=1 THEN 'web' ELSE close_user END, close_dt=CASE WHEN ?=1 THEN GETDATE() ELSE close_dt END
          WHEN NOT MATCHED THEN INSERT(ym,cust_code,close_flag,base_amt,adj_amt,final_amt,close_user,close_dt)
             VALUES(?,?,?,?,?,?,CASE WHEN ?=1 THEN 'web' ELSE NULL END, CASE WHEN ?=1 THEN GETDATE() ELSE NULL END);""",
          y, cc, cf, base_amt, tot, final, cf, cf, y, cc, cf, base_amt, tot, final, cf, cf)
        return {"ok": True, "final_amt": final, "adj_amt": tot, "closed": do_close}
    finally:
        nx.close()

@app.post("/api/salemagam/reopen")
def salemagam_reopen(payload: dict = Body(...)):
    """마감 취소(재수정 허용)."""
    y = _dig4(payload.get("ym")); cc = str(payload.get("cust_code", "")).strip()
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("UPDATE nx.sale_close SET close_flag=0, close_dt=NULL WHERE ym=? AND cust_code=?", y, cc)
        return {"ok": True, "reopened": nc.rowcount}
    finally:
        nx.close()

@app.get("/api/salemagam/weight")
def salemagam_weight(ym: str = Query("")):
    """무게정산(중량조정): 업체별 원소재/용접봉 출고−업체가공입고=차액, ×(시세−사급가).
       기초 불필요·매월 증/차감. 출고=확정(tag5), 입고=마스터(PR_M_ITEM.ITEM_WEIGHT+CS_M_ITEM_BOM 잠정)."""
    y = _dig4(ym) or _cur_ym()
    # 시세·사급가(원소재/용접봉) — nx.mat_price_month
    px = {"원소재": (25000.0, 20000.0), "용접봉": (None, 21100.0)}
    try:
        nx = _nx(); nc = nx.cursor()
        nc.execute("SELECT category, real_price, sagub_price FROM nx.mat_price_month WHERE apply_ym=?", y)
        for cat, rp, sp in nc.fetchall():
            px[cat] = ((float(rp) if rp is not None else None), (float(sp) if sp is not None else px.get(cat, (None, None))[1]))
        nx.close()
    except Exception:
        pass
    rr, sr = px.get("원소재", (25000.0, 20000.0))
    rw, sw = px.get("용접봉", (None, 21100.0))
    try:
        data = weight_calc.compute(y, real_raw=(rr if rr is not None else 25000.0), sagub_raw=(sr if sr is not None else 20000.0),
                                   real_weld=rw, sagub_weld=(sw if sw is not None else 21100.0))
    except Exception as e:
        raise HTTPException(500, f"무게정산 계산 오류: {e}")
    return {"ym": y, "real_raw": (rr if rr is not None else 25000.0), "sagub_raw": (sr if sr is not None else 20000.0),
            "real_weld": rw, "sagub_weld": (sw if sw is not None else 21100.0), "rows": data}

@app.get("/api/matprice/list")
def matprice_list(ym: str = Query("")):
    """월별 원소재/용접봉 시세·사급가 조회."""
    y = _dig4(ym) or _cur_ym()
    defaults = {"원소재": 20000.0, "용접봉": 21100.0}
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("SELECT category, real_price, sagub_price, note FROM nx.mat_price_month WHERE apply_ym=?", y)
        rows = {r[0]: {"category": r[0], "real_price": (float(r[1]) if r[1] is not None else None),
                       "sagub_price": (float(r[2]) if r[2] is not None else defaults.get(r[0])), "note": r[3] or ""} for r in nc.fetchall()}
        for cat, sg in defaults.items():
            if cat not in rows:
                rows[cat] = {"category": cat, "real_price": None, "sagub_price": sg, "note": ""}
        return {"ym": y, "rows": [rows["원소재"], rows["용접봉"]]}
    finally:
        nx.close()

@app.post("/api/matprice/save")
def matprice_save(payload: dict = Body(...)):
    y = _dig4(payload.get("ym"))
    if not y:
        raise HTTPException(400, "ym 필요")
    nx = _nx(); nc = nx.cursor()
    try:
        for it in (payload.get("rows") or []):
            cat = str(it.get("category", "")).strip()
            if cat not in ("원소재", "용접봉"):
                continue
            rp = it.get("real_price"); sp = it.get("sagub_price")
            rp = float(rp) if rp not in (None, "") else None
            sp = float(sp) if sp not in (None, "") else None
            nc.execute("""MERGE nx.mat_price_month AS T USING (SELECT ? ym, ? cat) S ON T.apply_ym=S.ym AND T.category=S.cat
                WHEN MATCHED THEN UPDATE SET real_price=?, sagub_price=?, upd_user='web', upd_dt=GETDATE()
                WHEN NOT MATCHED THEN INSERT(apply_ym,category,real_price,sagub_price,upd_user,upd_dt) VALUES(?,?,?,?,'web',GETDATE());""",
                y, cat, rp, sp, y, cat, rp, sp)
        return {"ok": True}
    finally:
        nx.close()

# ===================== 품질 반성회의록 CRUD (nx.meeting ← 레거시 cm_user_meeting_1) =====================
# 근거: w_cm_user_meeting_200/205. 코드마스터 없음(순수 텍스트). 비용=(인원+1)×시간×358.3.
_MEETING_COLS = ["meeting_type", "meeting_ymd", "subject", "member", "member_count", "duration_min", "pay_amount",
                 "note", "note2", "organizer",
                 "action1_desc", "action1_person", "action1_due", "action2_desc", "action2_person", "action2_due",
                 "action3_desc", "action3_person", "action3_due", "action4_desc", "action4_person", "action4_due",
                 "action5_desc", "action5_person", "action5_due"]
_MEETING_INT = {"member_count", "duration_min", "pay_amount"}

@app.get("/api/meeting/list")
def meeting_list(q: str = Query(""), from_ymd: str = Query(""), to_ymd: str = Query(""), limit: int = Query(300)):
    """반성회의록 목록(nx.meeting). 제목/작성자/참석자 검색 + 회의일자 범위."""
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["1=1"]; p = []
        if q.strip(): w.append("(subject LIKE ? OR organizer LIKE ? OR member LIKE ?)"); p += [f"%{q.strip()}%"] * 3
        if from_ymd.strip(): w.append("meeting_ymd >= ?"); p.append(from_ymd.strip())
        if to_ymd.strip(): w.append("meeting_ymd <= ?"); p.append(to_ymd.strip())
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),2000))} meeting_id,meeting_type,meeting_ymd,subject,member,
            member_count,duration_min,pay_amount,note,note2,organizer,
            action1_desc,action1_person,action1_due,action2_desc,action2_person,action2_due,
            action3_desc,action3_person,action3_due,action4_desc,action4_person,action4_due,
            action5_desc,action5_person,action5_due
            FROM nx.meeting WHERE {' AND '.join(w)} ORDER BY meeting_ymd DESC, meeting_id DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [{c: ("" if v is None else v) for c, v in zip(cols, r)} for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()

@app.post("/api/meeting/save")
def meeting_save(payload: dict = Body(...)):
    """반성회의록 등록/수정. 제목 필수. 비용 서버 자동계산(방어)."""
    p = payload
    if not str(p.get("subject", "") or "").strip():
        raise HTTPException(400, "회의 제목은 필수입니다.")
    def s(k):
        v = p.get(k); return None if v in (None, "") else str(v).strip()
    def i(k):
        v = p.get(k)
        try: return int(float(v)) if v not in (None, "") else None
        except Exception: return None
    mc, du = i("member_count"), i("duration_min")
    pay = int(round((mc + 1) * du * 358.3)) if (mc is not None and du is not None) else i("pay_amount")
    vals = [pay if k == "pay_amount" else (i(k) if k in _MEETING_INT else s(k)) for k in _MEETING_COLS]
    mid = p.get("meeting_id")
    nx = _nx(); cur = nx.cursor()
    try:
        if mid:
            sets = ",".join(f"{k}=?" for k in _MEETING_COLS)
            cur.execute(f"UPDATE nx.meeting SET {sets},upd_user='web',upd_dt=GETDATE() WHERE meeting_id=?", *vals, int(mid))
            return {"ok": True, "mode": "update", "meeting_id": int(mid), "pay_amount": pay}
        cur.execute(f"INSERT INTO nx.meeting({','.join(_MEETING_COLS)},upd_user,upd_dt) OUTPUT INSERTED.meeting_id "
                    f"VALUES({','.join(['?']*len(_MEETING_COLS))},'web',GETDATE())", *vals)
        newid = cur.fetchone()[0]
        return {"ok": True, "mode": "insert", "meeting_id": int(newid), "pay_amount": pay}
    finally:
        nx.close()

@app.post("/api/meeting/delete")
def meeting_delete(payload: dict = Body(...)):
    ids = [int(x) for x in (payload.get("ids", []) or []) if str(x).strip().lstrip('-').isdigit()]
    if not ids: return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute(f"DELETE FROM nx.meeting WHERE meeting_id IN ({','.join('?'*len(ids))})", *ids)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()

# ===================== 설계도면조회 + 도면/시방 파일첨부 (nx.doc + 레거시 blob) =====================
# 근거: w_pr_master_200. 일반도면(개발)=DRAWING.PR_M_DWG · 시방도면(품질)=QA_T_SPEC_REV_BLOB(FILE_TAG='2').
# 신규 업로드=NAS경로(DOC_STORAGE_PATH)+nx.doc 메타. 기존 15.9GB=레거시 blob 읽기 폴백.
import os as _os, hashlib as _hashlib, mimetypes as _mimetypes
from urllib.parse import quote as _urlquote
DOC_STORAGE_PATH = _os.getenv("DOC_STORAGE_PATH", r"F:\NEW_ERP_FILES")   # 배포시 NAS 마운트(\\200.200.200.15\...)로 교체
_DOC_KIND = {"GENERAL_DWG": "일반도면", "SPEC_DWG": "시방도면", "SPEC_SHEET": "시방서", "ITEM_ATTACH": "품목첨부"}

@app.get("/api/doc/list")
def doc_list(item_code: str = Query("")):
    """설계도면조회: 품번/파일명 검색 = nx.doc(신규) ∪ 일반도면(PR_M_DWG) ∪ 시방도면(QA_T_SPEC_REV).
       ★빈 검색이면 각 소스 최근 전체(TOP N)를 반환해 조회 즉시 파일이 보이게(브라우즈)."""
    item = item_code.strip()
    like = f"%{item}%"
    rows = []
    # nx.doc(신규) — 있으면 최상단
    nx = _nx(); ncur = nx.cursor()
    try:
        if item:
            ncur.execute("""SELECT doc_id,doc_kind,orig_filename,ext,byte_size,insert_user,insert_dt,rev_yymd,rev_no
                FROM nx.doc WHERE del_flag=0 AND (item_code=? OR orig_filename LIKE ?) ORDER BY insert_dt DESC""", item, like)
        else:
            ncur.execute("""SELECT doc_id,doc_kind,orig_filename,ext,byte_size,insert_user,insert_dt,rev_yymd,rev_no
                FROM nx.doc WHERE del_flag=0 ORDER BY insert_dt DESC""")
        for r in ncur.fetchall():
            rows.append({"src": "doc", "key": str(r[0]), "kind": r[1], "kind_nm": _DOC_KIND.get(r[1], r[1]),
                         "filename": r[2], "rev": (f"{r[7]}/{r[8]}" if r[7] else ""), "spec_no": "",
                         "dt": (r[6].isoformat() if hasattr(r[6], "isoformat") else str(r[6] or "")).replace("T", " ")[:19],
                         "user": r[5] or "", "size": int(r[4] or 0), "editable": (r[1] == "GENERAL_DWG"), "gubun": "1"})
    finally:
        nx.close()
    # ★레거시 w_pr_master_200 동일: 일반도면(PR_M_DWG 도면구분1) + 시방도면(QA_T_SPEC_REV DRAWING_FILE<>'' 도면구분2). 캡·blob필터 제거(전건).
    cn = _conn(); cur = cn.cursor()
    try:
        # ① 일반도면 (도면구분 1)
        if item:
            cur.execute("""SELECT FILE_NAME, FILE_DATETIME, ISNULL(UPDATE_USER_ID,ISNULL(INSERT_USER_ID,'')), ISNULL(UPDATE_DATETIME,FILE_DATETIME)
                FROM DRAWING.DBO.PR_M_DWG WHERE FILE_NAME LIKE ? ORDER BY FILE_DATETIME DESC""", like)
        else:
            cur.execute("""SELECT FILE_NAME, FILE_DATETIME, ISNULL(UPDATE_USER_ID,ISNULL(INSERT_USER_ID,'')), ISNULL(UPDATE_DATETIME,FILE_DATETIME)
                FROM DRAWING.DBO.PR_M_DWG ORDER BY FILE_DATETIME DESC""")
        for r in cur.fetchall():
            rows.append({"src": "dwg", "key": f"{r[0]}|{r[1]}", "kind": "GENERAL_DWG", "kind_nm": "일반도면",
                         "filename": r[0], "rev": "", "spec_no": "",
                         "dt": (r[1].isoformat() if hasattr(r[1], "isoformat") else str(r[1] or "")).replace("T", " ")[:19],
                         "user": r[2] or "", "size": 0, "editable": False, "gubun": "1"})
        # ② 시방도면 (도면구분 2) — DRAWING_FILE<>'' 헤더 기준(blob 존재 강제 제거)
        if item:
            cur.execute("""SELECT h.REV_YYMD, h.REV_NO, ISNULL(h.DRAWING_FILE,''), ISNULL(h.ISSUE_YYMD,''),
                  ISNULL(h.UPDATE_USER_ID,ISNULL(h.INSERT_USER_ID,'RPA')), ISNULL(h.UPDATE_DATETIME,h.INSERT_DATETIME)
                FROM QA_T_SPEC_REV h WHERE (h.DRAWING_FILE LIKE ? OR h.ITEM_CODE LIKE ?) AND ISNULL(h.DRAWING_FILE,'')<>''
                ORDER BY h.REV_YYMD DESC, h.REV_NO DESC""", like, like)
        else:
            cur.execute("""SELECT h.REV_YYMD, h.REV_NO, ISNULL(h.DRAWING_FILE,''), ISNULL(h.ISSUE_YYMD,''),
                  ISNULL(h.UPDATE_USER_ID,ISNULL(h.INSERT_USER_ID,'RPA')), ISNULL(h.UPDATE_DATETIME,h.INSERT_DATETIME)
                FROM QA_T_SPEC_REV h WHERE ISNULL(h.DRAWING_FILE,'')<>'' ORDER BY h.REV_YYMD DESC, h.REV_NO DESC""")
        for r in cur.fetchall():
            rows.append({"src": "spec", "key": f"{r[0]}|{r[1]}|2", "kind": "SPEC_DWG", "kind_nm": "시방도면",
                         "filename": r[2], "rev": f"{r[0]}/{r[1]}", "spec_no": f"{r[0]}-{r[1]}",
                         "dt": (r[5].isoformat() if hasattr(r[5], "isoformat") else str(r[3] or "")).replace("T", " ")[:19],
                         "user": r[4] or "RPA", "size": 0, "editable": False, "gubun": "2"})
    finally:
        cn.close()
    # ★레거시 동일: 일반+시방 통합 후 파일일시 내림차순(최신 시방이 최상단)
    rows.sort(key=lambda x: x.get("dt") or "", reverse=True)
    return {"rows": rows, "cnt": len(rows)}

@app.get("/api/doc/download")
def doc_download(src: str = Query(...), key: str = Query(...), disp: str = Query("attach")):
    """다운로드/열기: doc=nx파일 / dwg=PR_M_DWG단일 / spec=QA blob 분할조립. disp=inline이면 브라우저에서 바로 열기(뷰)."""
    data = b""; fname = "file"
    if src == "doc":
        nx = _nx(); cur = nx.cursor()
        try:
            cur.execute("SELECT orig_filename, storage_uri FROM nx.doc WHERE doc_id=? AND del_flag=0", int(key))
            r = cur.fetchone()
            if not r: raise HTTPException(404, "문서 없음")
            path = _os.path.join(DOC_STORAGE_PATH, r[1]); fname = r[0]
            if not _os.path.exists(path): raise HTTPException(404, f"파일 없음: {r[1]}")
            with open(path, "rb") as fp: data = fp.read()
        finally: nx.close()
    elif src == "dwg":
        fn, fdt = key.split("|", 1)
        cn = _conn(); cur = cn.cursor()
        try:
            cur.execute("SELECT FILE_BLOB FROM DRAWING.DBO.PR_M_DWG WHERE FILE_NAME=? AND FILE_DATETIME=?", fn, fdt)
            r = cur.fetchone()
            if not r: raise HTTPException(404, "도면 없음")
            data = bytes(r[0]) if r[0] is not None else b""; fname = fn
        finally: cn.close()
    elif src == "spec":
        ry, rn, tag = key.split("|")
        cn = _conn(); cur = cn.cursor()
        try:
            cur.execute("SELECT ISNULL(DRAWING_FILE,''), ISNULL(SPECS_FILE,'') FROM QA_T_SPEC_REV WHERE REV_YYMD=? AND REV_NO=?", ry, int(rn))
            h = cur.fetchone()
            fname = ((h[0] if tag == '2' else h[1]) or f"{ry}_{rn}.pdf") if h else f"{ry}_{rn}.pdf"
            cur.execute("SELECT FILE_BLOB FROM QA_T_SPEC_REV_BLOB WHERE REV_YYMD=? AND REV_NO=? AND FILE_TAG=? ORDER BY FILE_SEQ", ry, int(rn), tag)
            data = b"".join(bytes(x[0]) for x in cur.fetchall() if x[0] is not None)
        finally: cn.close()
    elif src == "sibang":   # 품목시방 PPT (DRAWING.PR_M_SIBANG, PR_M_DWG 쌍둥이 = 단일 blob)
        fn, fdt = key.split("|", 1)
        cn = _conn(); cur = cn.cursor()
        try:
            cur.execute("SELECT FILE_BLOB FROM DRAWING.DBO.PR_M_SIBANG WHERE FILE_NAME=? AND FILE_DATETIME=?", fn, fdt)
            r = cur.fetchone()
            if not r: raise HTTPException(404, "시방파일 없음")
            data = bytes(r[0]) if r[0] is not None else b""; fname = fn
        finally: cn.close()
    elif src == "itemblob":   # 품목 첨부 (PR_M_ITEM_BLOB, 청크 조립, 파일명 합성)
        ic, ft = key.split("|", 1)
        cn = _conn(); cur = cn.cursor()
        try:
            cur.execute("SELECT TOP 1 ISNULL(FILE_EXT,'') FROM PR_M_ITEM_BLOB WHERE ITEM_CODE=? AND FILE_TYPE=?", ic, ft)
            e = cur.fetchone(); ext = (e[0].strip() if e and e[0] else "dat")
            fname = f"{ic}_{ft}.{ext}"
            cur.execute("SELECT MODULE_BLOB FROM PR_M_ITEM_BLOB WHERE ITEM_CODE=? AND FILE_TYPE=? ORDER BY MODULE_SEQ", ic, ft)
            data = b"".join(bytes(x[0]) for x in cur.fetchall() if x[0] is not None)
        finally: cn.close()
    else:
        raise HTTPException(400, "src 오류")
    mime = _mimetypes.guess_type(fname)[0] or "application/octet-stream"
    cd = "inline" if str(disp).lower() == "inline" else "attachment"
    return Response(content=data, media_type=mime,
                    headers={"Content-Disposition": f"{cd}; filename*=UTF-8''{_urlquote(fname)}"})

@app.post("/api/doc/upload")
async def doc_upload(file: UploadFile = File(...), doc_kind: str = Form("GENERAL_DWG"),
                     item_code: str = Form(""), rev_yymd: str = Form(""), rev_no: str = Form(""),
                     user: str = Form("웹사용자")):
    """업로드: NAS경로(DOC_STORAGE_PATH) 저장 + nx.doc 메타. sha256 중복검사."""
    raw = await file.read()
    if not raw: raise HTTPException(400, "빈 파일입니다.")
    fname = file.filename or "file"
    ext = ((fname.rsplit(".", 1)[-1] if "." in fname else "") or "").lower()[:10]
    sha = _hashlib.sha256(raw).hexdigest()
    sub = _os.path.join(doc_kind, (item_code.strip() or "_misc"))
    d = _os.path.join(DOC_STORAGE_PATH, sub)
    try:
        _os.makedirs(d, exist_ok=True)
    except Exception as e:
        raise HTTPException(500, f"저장경로 생성 실패({DOC_STORAGE_PATH}): {e}")
    safe = f"{sha[:12]}_{fname}"
    with open(_os.path.join(d, safe), "wb") as fp: fp.write(raw)
    rel = _os.path.join(sub, safe)
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""INSERT INTO nx.doc(doc_kind,item_code,rev_yymd,rev_no,orig_filename,storage_uri,ext,byte_size,sha256,insert_user,insert_dt)
            OUTPUT INSERTED.doc_id VALUES(?,?,?,?,?,?,?,?,?,?,GETDATE())""",
            doc_kind, (item_code.strip() or None), (rev_yymd.strip() or None),
            (int(rev_no) if str(rev_no).strip().isdigit() else None),
            fname, rel, ext, len(raw), sha, (user or "웹사용자")[:20])
        did = cur.fetchone()[0]
        return {"ok": True, "doc_id": int(did), "size": len(raw), "path": rel}
    finally:
        nx.close()

@app.post("/api/doc/delete")
def doc_delete(payload: dict = Body(...)):
    """삭제 — nx.doc GENERAL_DWG(일반도면)만. 시방도면은 시방변경에서."""
    did = payload.get("doc_id")
    if not did: return {"ok": False, "errors": ["doc_id 필요"]}
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT doc_kind, storage_uri FROM nx.doc WHERE doc_id=? AND del_flag=0", int(did))
        r = cur.fetchone()
        if not r: return {"ok": False, "errors": ["문서 없음"]}
        if r[0] != "GENERAL_DWG":
            return {"ok": False, "errors": ["일반도면만 삭제 가능합니다. 시방도면은 시방변경관리에서 삭제하세요."]}
        cur.execute("UPDATE nx.doc SET del_flag=1 WHERE doc_id=?", int(did))
        try:
            fp = _os.path.join(DOC_STORAGE_PATH, r[1])
            if _os.path.exists(fp): _os.remove(fp)
        except Exception: pass
        return {"ok": True}
    finally:
        nx.close()

@app.get("/api/itemspec/list")
def itemspec_list(item_code: str = Query("")):
    """품목시방관리(w_pr_master_210): 품번별 = nx.doc(ITEM_ATTACH) ∪ 시방PPT(DRAWING.PR_M_SIBANG) ∪ 품목첨부14종(PR_M_ITEM_BLOB, PR010)."""
    item = item_code.strip()   # ★레거시 w_pr_master_210: 빈 품번=전건 브라우즈(PR_M_SIBANG 정본)
    like = f"%{item}%"
    rows = []
    nx = _nx(); ncur = nx.cursor()
    try:
        if item:
            ncur.execute("""SELECT doc_id,orig_filename,ext,byte_size,insert_user,insert_dt,ISNULL(file_tag,'')
                FROM nx.doc WHERE del_flag=0 AND doc_kind='ITEM_ATTACH' AND item_code=? ORDER BY insert_dt DESC""", item)
        else:
            ncur.execute("""SELECT doc_id,orig_filename,ext,byte_size,insert_user,insert_dt,ISNULL(file_tag,'')
                FROM nx.doc WHERE del_flag=0 AND doc_kind='ITEM_ATTACH' ORDER BY insert_dt DESC""")
        for r in ncur.fetchall():
            rows.append({"src": "doc", "key": str(r[0]), "atype": r[6], "atype_nm": "신규첨부",
                         "filename": r[1], "user": r[4] or "", "size": int(r[3] or 0),
                         "dt": (r[5].isoformat() if hasattr(r[5], "isoformat") else ""), "editable": True})
    finally:
        nx.close()
    cn = _conn(); cur = cn.cursor()
    try:
        pr010 = _kindmap(cur, "PR010")
        # 시방 PPT(정본) — 빈 품번=전건. blob 크기 스캔 제거(전건 성능).
        if item:
            cur.execute("""SELECT FILE_NAME, FILE_DATETIME, ISNULL(INSERT_USER_ID,'')
                FROM DRAWING.DBO.PR_M_SIBANG WHERE FILE_NAME LIKE ? ORDER BY FILE_DATETIME DESC""", like)
        else:
            cur.execute("""SELECT FILE_NAME, FILE_DATETIME, ISNULL(INSERT_USER_ID,'')
                FROM DRAWING.DBO.PR_M_SIBANG ORDER BY FILE_DATETIME DESC""")
        for r in cur.fetchall():
            rows.append({"src": "sibang", "key": f"{r[0]}|{r[1]}", "atype": "SIBANG_PPT", "atype_nm": "시방(PPT)",
                         "filename": r[0], "user": r[2], "size": 0, "dt": str(r[1]), "editable": False})
        # 품목첨부 14종 — 품번 지정 시에만(품목별 상세)
        if item:
            cur.execute("""SELECT FILE_TYPE, MAX(ISNULL(FILE_EXT,'')), SUM(DATALENGTH(MODULE_BLOB)),
                  MAX(ISNULL(INSERT_USER_ID,'')), MAX(INSERT_DATETIME)
                FROM PR_M_ITEM_BLOB WHERE ITEM_CODE=? GROUP BY FILE_TYPE ORDER BY FILE_TYPE""", item)
            for r in cur.fetchall():
                ft = str(r[0]).strip(); ext = (r[1].strip() if r[1] else "dat"); nm = pr010.get(ft, ft)
                rows.append({"src": "itemblob", "key": f"{item}|{ft}", "atype": ft, "atype_nm": nm,
                             "filename": f"{item}_{nm}.{ext}", "user": r[3] or "", "size": int(r[2] or 0),
                             "dt": (r[4].isoformat() if hasattr(r[4], "isoformat") else ""), "editable": False})
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

@app.get("/api/qc/spec/files")
def qc_spec_files(rev_ymd: str = Query(""), rev_no: str = Query("")):
    """시방 첨부파일 목록: 레거시 QA blob(도면 tag2/시방서 tag1) + nx.doc(SPEC_DWG/SPEC_SHEET)."""
    ry = rev_ymd.strip(); rn = rev_no.strip()
    out = []
    if ry and rn.isdigit():
        cn = _conn(); cur = cn.cursor()
        try:
            cur.execute("SELECT ISNULL(DRAWING_FILE,''), ISNULL(SPECS_FILE,'') FROM QA_T_SPEC_REV WHERE REV_YYMD=? AND REV_NO=?", ry, int(rn))
            h = cur.fetchone()
            for tag, kind, fn in [('2', '도면', (h[0] if h else '')), ('1', '시방서', (h[1] if h else ''))]:
                cur.execute("SELECT SUM(DATALENGTH(FILE_BLOB)) FROM QA_T_SPEC_REV_BLOB WHERE REV_YYMD=? AND REV_NO=? AND FILE_TAG=?", ry, int(rn), tag)
                sz = cur.fetchone()[0]
                if sz:
                    out.append({"kind": kind, "src": "spec", "key": f"{ry}|{rn}|{tag}",
                                "filename": (fn or f"{ry}_{rn}_{kind}.pdf"), "size": int(sz), "editable": False})
        finally: cn.close()
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""SELECT doc_id, doc_kind, orig_filename, byte_size FROM nx.doc
            WHERE del_flag=0 AND doc_kind IN ('SPEC_DWG','SPEC_SHEET') AND rev_yymd=? AND rev_no=?""",
            ry, (int(rn) if rn.isdigit() else None))
        for r in cur.fetchall():
            out.append({"kind": ("도면" if r[1] == 'SPEC_DWG' else "시방서"), "src": "doc", "key": str(r[0]),
                        "filename": r[2], "size": int(r[3] or 0), "editable": True})
    finally: nx.close()
    return {"rows": out, "cnt": len(out)}

# ===================== 생산전표출력관리 (w_pr_input_468 등) — 전표/간판/라벨 조회·발행(nx)·인쇄 =====================
# 근거: jp_proc_method(J전표/G간판). 계획=nx.plan_part(도번). 발행=nx.sheet_issue(컷오버=nx 신규원장).
@app.get("/api/prodsheet/list")
def prodsheet_list(from_ymd: str = Query(""), to_ymd: str = Query(""), line: str = Query(""),
                   item: str = Query(""), limit: int = Query(1000)):
    """도번별 생산계획 + jp_proc_method(J전표/G간판) + 발행현황(nx.sheet_issue)."""
    def d6(s):
        d = ''.join(c for c in str(s or '') if c.isdigit())
        return d[2:8] if len(d) >= 8 else d
    cn = _conn(); cur = cn.cursor()
    nx = _nx(); ncur = nx.cursor()
    try:
        jp = {}
        cur.execute("SELECT ITEM_CODE, MAX(JP_PROC_METHOD) FROM PR_M_ITEM_PROC_GAGONG WHERE JP_PROC_METHOD IN ('J','G') GROUP BY ITEM_CODE")
        for r in cur.fetchall(): jp[str(r[0]).strip()] = str(r[1]).strip()
        w = ["1=1"]; p = []
        if from_ymd.strip(): w.append("pp.PLAN_YMD>=?"); p.append(d6(from_ymd))
        if to_ymd.strip(): w.append("pp.PLAN_YMD<=?"); p.append(d6(to_ymd))
        if line.strip(): w.append("pp.WORK_CENTER=?"); p.append(line.strip())
        if item.strip(): w.append("pp.PART_CODE LIKE ?"); p.append(f"%{item.strip()}%")
        ncur.execute(f"""SELECT TOP {max(1,min(int(limit),3000))} pp.PLAN_YMD, pp.WORK_ORDER, pp.ASSY_ITEM_CODE, pp.PART_CODE,
              pp.WORK_CENTER, pp.PLAN_QTY,
              (SELECT COUNT(*) FROM nx.sheet_issue si WHERE si.item_code=pp.PART_CODE AND si.plan_ymd=pp.PLAN_YMD AND si.kind='G'),
              (SELECT COUNT(*) FROM nx.sheet_issue si WHERE si.item_code=pp.PART_CODE AND si.plan_ymd=pp.PLAN_YMD AND si.kind='L'),
              (SELECT COUNT(*) FROM nx.sheet_issue si WHERE si.item_code=pp.PART_CODE AND si.plan_ymd=pp.PLAN_YMD AND si.kind='J')
            FROM nx.plan_part pp WHERE {' AND '.join(w)} ORDER BY pp.PLAN_YMD DESC, pp.PART_CODE""", *p)
        rows = []; parts = set()
        for r in ncur.fetchall():
            g = lambda i: str(r[i] if r[i] is not None else "").strip()
            pc = g(3)
            rows.append({"plan_ymd": g(0), "work_order": g(1), "assy": g(2), "item_code": pc, "work_center": g(4),
                         "plan_qty": float(r[5] or 0), "method": jp.get(pc, ""),
                         "gcnt": int(r[6] or 0), "lcnt": int(r[7] or 0), "jcnt": int(r[8] or 0)})
            parts.add(pc)
        nm = {}; pl = [x for x in parts if x]
        for i in range(0, len(pl), 900):
            ch = pl[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
            for a, b in cur.fetchall(): nm[str(a).strip()] = b
        for x in rows:
            x["nm"] = nm.get(x["item_code"], "")
            x["method_nm"] = {"J": "전표", "G": "간판"}.get(x["method"], "미지정")
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close(); nx.close()

@app.post("/api/prodsheet/issue")
def prodsheet_issue(payload: dict = Body(...)):
    """발행 채번 → nx.sheet_issue. kind J전표/G간판/L라벨. box_no/print_seq/sheet_no=nx max+1."""
    kind = str(payload.get("kind", "")).strip()
    rows = payload.get("rows", []) or []
    if kind not in ("J", "G", "L"):
        raise HTTPException(400, "kind 오류(J전표/G간판/L라벨)")
    user = (str(payload.get("user", "") or "").strip() or "웹사용자")[:20]
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT ISNULL(MAX(box_no),0), ISNULL(MAX(print_seq),0), ISNULL(MAX(sheet_no),0) FROM nx.sheet_issue")
        mx = cur.fetchone(); box = int(mx[0] or 0); seq = int(mx[1] or 0); sheet = int(mx[2] or 0)
        issued = 0
        for r in rows:
            ic = str(r.get("item_code", "") or "").strip()
            if not ic:
                continue
            qty = float(r.get("plan_qty") or 0)
            sheet += 1; bx = ps = qf = qt = None
            if kind == "G":
                box += 1; bx = box
            elif kind == "L":
                seq += 1; ps = seq
                qf = f"{ic}{sheet:08d}0001"; qt = f"{ic}{sheet:08d}{int(max(qty,1)):04d}"
            cur.execute("""INSERT INTO nx.sheet_issue(kind,item_code,assy_code,work_order,line_no,plan_ymd,plan_qty,
                box_no,print_seq,qr_from,qr_to,sheet_no,issue_user,issue_dt)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,GETDATE())""",
                kind, ic, (r.get("assy") or None), (r.get("work_order") or None), (r.get("work_center") or None),
                (r.get("plan_ymd") or None), qty, bx, ps, qf, qt, sheet, user)
            issued += 1
        return {"ok": True, "issued": issued, "kind": kind}
    finally:
        nx.close()

# ===================== 공정별 바코드생산실적 (w_pr_input_520) — 스캔→자동채움→등록/취소(nx.proc_barcode) =====================
# 근거: 220/527 근사(520 소스 부재). 간판(GP+box_no)/라벨(QR) 스캔→마스터조회→토글. ★커밋 상세는 원본 확보 후 대조.
@app.get("/api/procbc/lookup")
def procbc_lookup(barcode: str = Query(...), proc_code: str = Query("")):
    """바코드 파싱→간판/라벨 마스터(nx.sheet_issue+레거시) 조회→품번·수량·등록/취소 판정."""
    bc = barcode.strip()
    if not bc:
        raise HTTPException(400, "바코드가 필요합니다.")
    item = None; qty = 0.0; sheet = None; kind = None; src = None
    nx = _nx(); cur = nx.cursor()
    try:
        if bc.upper().startswith("GP") and bc[2:].isdigit():
            cur.execute("SELECT TOP 1 item_code, plan_qty, sheet_no FROM nx.sheet_issue WHERE kind='G' AND box_no=?", int(bc[2:]))
            r = cur.fetchone()
            if r: item, qty, sheet, kind, src = str(r[0]).strip(), float(r[1] or 0), int(r[2] or 0), "간판", "nx"
        if not item:
            cur.execute("SELECT TOP 1 item_code, plan_qty, sheet_no FROM nx.sheet_issue WHERE qr_from=? OR qr_to=?", bc, bc)
            r = cur.fetchone()
            if r: item, qty, sheet, kind, src = str(r[0]).strip(), float(r[1] or 0), int(r[2] or 0), "라벨", "nx"
        if not item:
            cn = _conn(); c2 = cn.cursor()
            try:
                if bc.upper().startswith("GP") and bc[2:].isdigit():
                    c2.execute("SELECT TOP 1 ITEM_CODE, PLAN_QTY, SHEET_NO FROM PR_T_INDI_SHEET2 WHERE BOX_NO=?", int(bc[2:]))
                    r = c2.fetchone()
                    if r: item, qty, sheet, kind, src = str(r[0]).strip(), float(r[1] or 0), int(r[2] or 0), "간판", "legacy"
                if not item:
                    c2.execute("SELECT TOP 1 ITEM_CODE, PRINT_QTY, SHEET_NO FROM PR_T_PRINT_STICKER WHERE QR_BARCODE_FROM=? OR QR_BARCODE_TO=?", bc, bc)
                    r = c2.fetchone()
                    if r: item, qty, sheet, kind, src = str(r[0]).strip(), float(r[1] or 0), int(r[2] or 0), "라벨", "legacy"
                nm = ""
                if item:
                    c2.execute("SELECT ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE=?", item)
                    rr = c2.fetchone(); nm = rr[0] if rr else ""
            finally: cn.close()
        else:
            cn = _conn(); c2 = cn.cursor()
            try:
                c2.execute("SELECT ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE=?", item)
                rr = c2.fetchone(); nm = rr[0] if rr else ""
            finally: cn.close()
        if not item:
            return {"found": False, "msg": "바코드를 찾을 수 없습니다 (간판 GP…/라벨 QR)"}
        cur.execute("SELECT ISNULL(SUM(prod_qty),0) FROM nx.proc_barcode WHERE barcode=? AND ISNULL(proc_code,'')=?", bc, proc_code.strip())
        prev = float(cur.fetchone()[0] or 0)
        return {"found": True, "barcode": bc, "kind": kind, "src": src, "item_code": item, "item_name": nm,
                "qty": qty, "sheet_no": sheet, "prev_qty": prev, "action": ("취소" if prev > 0 else "등록")}
    finally:
        nx.close()

@app.post("/api/procbc/save")
def procbc_save(payload: dict = Body(...)):
    """스캔 실적 저장 — 등록(+수량)/취소(−기실적) 토글. nx.proc_barcode.
    ★자동 백플러시: 완성공정(proc==품목 MAX PROC_SEQ 또는 finish_flag='1') AND INNER_PROD=1 이면
      등록→backflush post(전체BOM×수량, +ASY/+PRD), 취소→reverse. bc_id+ref_key(BC:{bc}:{proc}) 멱등.
      바코드 INSERT + 백플러시(소비/생산/log)를 _nx_tx 동일 트랜잭션(부분실패 전체 롤백)."""
    p = payload
    bc = str(p.get("barcode", "") or "").strip()
    proc = str(p.get("proc_code", "") or "").strip()
    item = str(p.get("item_code", "") or "").strip()
    qty = float(p.get("qty") or 0)
    user = (str(p.get("user") or "웹사용자")[:20])
    if not bc or not item:
        raise HTTPException(400, "바코드·품번이 필요합니다.")
    if not proc:
        raise HTTPException(400, "공정을 선택하세요.")
    cn = _conn(); nx = _nx_tx(); cur = nx.cursor()   # cn=RO(완성공정·INNER_PROD 판정), nx=쓰기 tx
    try:
        cur.execute("SELECT ISNULL(SUM(prod_qty),0) FROM nx.proc_barcode WHERE barcode=? AND ISNULL(proc_code,'')=?", bc, proc)
        prev = float(cur.fetchone()[0] or 0)
        store = -prev if prev > 0 else qty
        if store == 0:
            nx.rollback(); return {"ok": False, "errors": ["처리할 수량이 없습니다."]}
        cur.execute("""INSERT INTO nx.proc_barcode(barcode,proc_code,item_code,work_code,prod_tag,mach_code,worker_code,sheet_no,prod_qty,prod_datetime,insert_user)
            OUTPUT INSERTED.bc_id VALUES(?,?,?,?,?,?,?,?,?,GETDATE(),?)""",
            bc, proc, item, (str(p.get("work_code") or "").strip() or None), (str(p.get("prod_tag") or "").strip() or None),
            (str(p.get("mach_code") or "").strip() or None), (str(p.get("worker_code") or "").strip() or None),
            (int(p.get("sheet_no")) if str(p.get("sheet_no") or "").strip().isdigit() else None),
            store, user)
        new_bc = int(cur.fetchone()[0])
        # ★완성공정 자동 백플러시 훅
        bf = None
        final = _final_proc_code(cn, item)
        is_final = (bool(final) and proc == final) or str(p.get("finish_flag", "")).strip() == "1"
        if is_final and _is_inner_prod(cn, item):
            ref_key = f"BC:{bc}:{proc}"
            gpc = proc
            wo = (str(p.get("sheet_no") or "").strip() or f"BC{new_bc}")
            if store > 0:   # 등록 → 백플러시 post
                bf = _backflush_core(cn, nx, item, store, wo, gpc, "post", user, ref_key, ref_bc=new_bc)
            else:           # 취소 → 백플러시 reverse
                bf = _backflush_core(cn, nx, item, prev, wo, gpc, "reverse", user, ref_key, ref_bc=new_bc)
        nx.commit()
        return {"ok": True, "action": ("취소" if prev > 0 else "등록"), "qty": store, "bc_id": new_bc,
                "final_proc": (proc if is_final else ""), "backflush": bf}
    except Exception as e:
        try: nx.rollback()
        except Exception: pass
        return {"ok": False, "errors": [str(e)[:200]]}
    finally:
        cn.close(); nx.close()

@app.get("/api/procbc/list")
def procbc_list(from_ymd: str = Query(""), to_ymd: str = Query(""), proc: str = Query(""), limit: int = Query(300)):
    """바코드 실적 이력(nx.proc_barcode). 품명 디코드."""
    nx = _nx(); cur = nx.cursor()
    cn = _conn(); c2 = cn.cursor()
    try:
        w = ["1=1"]; pr = []
        if from_ymd.strip(): w.append("b.prod_datetime>=?"); pr.append(from_ymd.strip() + " 00:00:00")
        if to_ymd.strip(): w.append("b.prod_datetime<=?"); pr.append(to_ymd.strip() + " 23:59:59")
        if proc.strip(): w.append("b.proc_code=?"); pr.append(proc.strip())
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),2000))} b.bc_id,b.barcode,b.proc_code,b.item_code,b.prod_qty,
              b.worker_code,b.mach_code,b.sheet_no,b.prod_datetime,b.insert_user
            FROM nx.proc_barcode b WHERE {' AND '.join(w)} ORDER BY b.prod_datetime DESC, b.bc_id DESC""", *pr)
        rows = []; items = set()
        for r in cur.fetchall():
            g = lambda i: str(r[i] if r[i] is not None else "").strip()
            rows.append({"bc_id": r[0], "barcode": g(1), "proc": g(2), "item_code": g(3), "qty": float(r[4] or 0),
                         "worker": g(5), "mach": g(6), "sheet_no": r[7],
                         "dt": (r[8].isoformat() if hasattr(r[8], "isoformat") else ""), "user": g(9)})
            items.add(g(3))
        nm = {}; il = [x for x in items if x]
        for i in range(0, len(il), 900):
            ch = il[i:i+900]; ph = ",".join("?" * len(ch))
            c2.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
            for a, b in c2.fetchall(): nm[str(a).strip()] = b
        for x in rows: x["nm"] = nm.get(x["item_code"], "")
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close(); cn.close()

# ===================== 준비실적처리(키팅) (w_pr_input_250/460_new) — 준비등록/취소(nx.ready_ledger) =====================
# 근거: 준비필요=계획−준비완료. 본체키팅=자재무차감(준비마킹). 잔량=SUM 파생. write=nx 신규원장(확정).
@app.get("/api/ready/plan")
def ready_plan(from_ymd: str = Query(""), to_ymd: str = Query(""), line: str = Query(""),
               item: str = Query(""), limit: int = Query(1000)):
    """자재별 정본 자재소요(nx.plan_part_mat) + 준비완료(nx.ready_ledger SUM) + 준비필요=소요−준비.
       ★정본 파이프라인 전환: nx.plan_part(구98%) → nx.plan_part_mat(레거시 STEP5→6→7 100%검증)."""
    def d6(s):
        d = ''.join(c for c in str(s or '') if c.isdigit())
        return d[2:8] if len(d) >= 8 else d
    cn = _conn(); cur = cn.cursor(); nx = _nx(); ncur = nx.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd.strip(): w.append("pp.PLAN_YMD>=?"); p.append(d6(from_ymd))
        if to_ymd.strip(): w.append("pp.PLAN_YMD<=?"); p.append(d6(to_ymd))
        if line.strip(): w.append("pp.MAT_WORK_CENTER_CODE=?"); p.append(line.strip())
        if item.strip(): w.append("pp.MAT_CODE LIKE ?"); p.append(f"%{item.strip()}%")
        ncur.execute(f"""SELECT TOP {max(1,min(int(limit),3000))} pp.PLAN_YMD, pp.WORK_ORDER, pp.MAT_CODE, pp.MAT_WORK_CENTER_CODE,
              SUM(CAST(pp.PART_PLAN_QTY AS float)) plan_qty,
              ISNULL((SELECT SUM(rl.MAINT_QTY) FROM nx.stock_ledger rl WHERE rl.STOCK_POINT='RDY' AND rl.ITEM_CODE=pp.MAT_CODE AND ISNULL(rl.WORK_ORDER,'')=ISNULL(pp.WORK_ORDER,'') AND ISNULL(rl.INPUT_YMD,'')=ISNULL(pp.PLAN_YMD,'')),0) ready_qty
            FROM nx.plan_part_mat pp WHERE {' AND '.join(w)}
            GROUP BY pp.PLAN_YMD, pp.WORK_ORDER, pp.MAT_CODE, pp.MAT_WORK_CENTER_CODE
            ORDER BY pp.PLAN_YMD DESC, pp.MAT_CODE""", *p)
        rows = []; parts = set()
        for r in ncur.fetchall():
            g = lambda i: str(r[i] if r[i] is not None else "").strip()
            pq = float(r[4] or 0); rq = float(r[5] or 0)
            rows.append({"plan_ymd": g(0), "work_order": g(1), "item_code": g(2), "work_center": g(3),
                         "plan_qty": pq, "ready_qty": rq, "need_qty": round(max(pq - rq, 0), 2)})
            parts.add(g(2))
        nm = {}; pl = [x for x in parts if x]
        for i in range(0, len(pl), 900):
            ch = pl[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
            for a, b in cur.fetchall(): nm[str(a).strip()] = b
        for x in rows: x["nm"] = nm.get(x["item_code"], "")
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close(); nx.close()

@app.post("/api/ready/register")
def ready_register(payload: dict = Body(...)):
    """준비등록(+)/취소(−) 다건. ★Phase1: 단일원장 nx.stock_ledger(STOCK_POINT='RDY', K1/K2). 셀키=item·wo·gpc(파트)·plan_ymd(INPUT_YMD).
       취소는 준비잔량 이내. flag-only(자재무차감). 쓰기 nx만."""
    mode = str(payload.get("mode", "register")).strip()
    rows = payload.get("rows", []) or []
    user = (str(payload.get("user", "") or "").strip() or "웹사용자")[:20]
    tag = "K2" if mode == "cancel" else "K1"; remk = "키팅취소" if mode == "cancel" else "키팅확인"
    nx = _nx(); cur = nx.cursor()
    try:
        n = 0; skipped = 0
        for r in rows:
            ic = str(r.get("item_code", "") or "").strip()
            qty = float(r.get("qty") or 0)
            if not ic or qty <= 0:
                skipped += 1; continue
            wo = str(r.get("work_order", "") or "").strip(); py = str(r.get("plan_ymd", "") or "").strip()
            gpc = (str(r.get("gpc") or r.get("work_center") or "").strip() or None)   # 파트(gpc) 우선, 없으면 작업처
            if mode == "cancel":
                cur.execute("""SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE STOCK_POINT='RDY'
                      AND ITEM_CODE=? AND ISNULL(GAGONG_PROC_CODE,'')=? AND ISNULL(WORK_ORDER,'')=? AND ISNULL(INPUT_YMD,'')=?""",
                      ic, (gpc or ''), wo, py)
                if qty > float(cur.fetchone()[0] or 0):
                    skipped += 1; continue
            cur.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=RIGHT(CONVERT(varchar(8),GETDATE(),112),6)")
            seq = int(cur.fetchone()[0] or 1)
            cur.execute("""INSERT INTO nx.stock_ledger(STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,CUST_CODE,ITEM_CODE,GAGONG_PROC_CODE,
                  WORK_ORDER,INPUT_YMD,MAINT_QTY,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
                VALUES('RDY',RIGHT(CONVERT(varchar(8),GETDATE(),112),6),?,?,'Z99990',?,?,?,?,?,?,?,GETDATE())""",
                seq, tag, ic, gpc, (wo or None), (py or None), (-qty if mode == "cancel" else qty), remk, user)
            n += 1
        return {"ok": True, "count": n, "skipped": skipped, "mode": mode}
    finally:
        nx.close()

# ================= 파트MASTER (기준정보, w_pr_master_280) — PR_M_PROC_GAGONG 라이브 CRUD =================
# 파트(가공공정)마스터. PROD_RATE=생산효율(=키팅 회수율). 공유마스터라 라이브 직접편집(원가·계획·키팅 즉시 일관). 권한게이트=프론트.
_GC_GUBUN = {'W': '자재창고', 'P': '생산파트', 'V': '생산창고', 'Q': '가공파트'}

@app.get("/api/partmaster/list")
def partmaster_list(q: str = Query(""), grp: str = Query("")):
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["1=1"]; p = []
        if q.strip():   w.append("(g.GAGONG_PROC_CODE LIKE ? OR g.GAGONG_PROC_DESC LIKE ?)"); p += [f"%{q.strip()}%", f"%{q.strip()}%"]
        if grp.strip(): w.append("ISNULL(g.PART_GROUP_CODE,'')=?"); p.append(grp.strip())
        cur.execute(f"""SELECT g.GAGONG_PROC_CODE code, g.GAGONG_PROC_DESC nm, ISNULL(g.GC_GUBUN,'') gubun,
              ISNULL(g.WORK_CODE,'') wc, ISNULL(w.WORK_DESC,'') wcnm, ISNULL(g.IN_CUST_CODE,'') wh, ISNULL(c.CUST_DESC,'') whnm,
              ISNULL(g.SORT_KEY,0) sortkey, ISNULL(g.PROD_RATE,0) rate, ISNULL(g.PART_GROUP_CODE,'') grp,
              ISNULL(g.WH_IP_ADDRESS,'') ip, ISNULL(g.RACK_NUMBER,0) rack, ISNULL(g.UPDATE_USER_ID,'') uid, g.UPDATE_DATETIME udt
            FROM PR_M_PROC_GAGONG g
            LEFT JOIN PR_M_WORK w ON w.WORK_CODE=g.WORK_CODE
            LEFT JOIN CM_M_CUST c ON c.CUST_CODE=g.IN_CUST_CODE
            WHERE {' AND '.join(w)} ORDER BY g.WORK_CODE, g.SORT_KEY, g.GAGONG_PROC_CODE""", *p)
        cols = [d[0] for d in cur.description]; rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r['gubunnm'] = _GC_GUBUN.get(r['gubun'], r['gubun'])
            r['rate'] = float(r['rate'] or 0); r['sortkey'] = int(r['sortkey'] or 0); r['rack'] = int(r['rack'] or 0)
            r['udt'] = str(r['udt'])[:19] if r['udt'] else ''
        return {"rows": rows, "cnt": len(rows), "gubuns": _GC_GUBUN}
    finally:
        cn.close()

@app.post("/api/partmaster/save")
def partmaster_save(payload: dict = Body(...)):
    r = payload.get('row', {}); user = (payload.get('user') or '웹')[:20]
    code = (r.get('code') or '').strip()
    if not code: return {"ok": False, "detail": "파트코드 필수"}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM PR_M_PROC_GAGONG WHERE GAGONG_PROC_CODE=?", code)
        exists = cur.fetchone()[0] > 0
        args = (r.get('nm', '') or '', (r.get('gubun', '') or '')[:1], (r.get('grp', '') or '')[:2], (r.get('wc', '') or '')[:4],
                (r.get('wh', '') or '')[:10], int(r.get('sortkey') or 0), float(r.get('rate') or 0),
                (r.get('ip', '') or '')[:30], int(r.get('rack') or 0), user)
        if exists:
            cur.execute("""UPDATE PR_M_PROC_GAGONG SET GAGONG_PROC_DESC=?, GC_GUBUN=?, PART_GROUP_CODE=?, WORK_CODE=?,
                  IN_CUST_CODE=?, SORT_KEY=?, PROD_RATE=?, WH_IP_ADDRESS=?, RACK_NUMBER=?,
                  UPDATE_USER_ID=?, UPDATE_DATETIME=getdate(), UPDATE_WINDOW='web_partmaster'
                WHERE GAGONG_PROC_CODE=?""", *args, code)
        else:
            cur.execute("""INSERT INTO PR_M_PROC_GAGONG(GAGONG_PROC_CODE, GAGONG_PROC_DESC, GC_GUBUN, PART_GROUP_CODE, WORK_CODE,
                  IN_CUST_CODE, SORT_KEY, PROD_RATE, WH_IP_ADDRESS, RACK_NUMBER, UPDATE_USER_ID, UPDATE_DATETIME, UPDATE_WINDOW)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,getdate(),'web_partmaster')""", code, *args)
        cn.commit()
        return {"ok": True, "mode": "update" if exists else "insert"}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()

@app.post("/api/partmaster/delete")
def partmaster_delete(payload: dict = Body(...)):
    code = (payload.get('code') or '').strip()
    if not code: return {"ok": False, "detail": "코드 필수"}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("DELETE FROM PR_M_PROC_GAGONG WHERE GAGONG_PROC_CODE=?", code); cn.commit()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()

@app.get("/api/partmaster/workers")
def partmaster_workers(part: str = Query(..., description="파트코드(GAGONG_PROC_CODE)")):
    """파트별 작업자 목록 (레거시 w_pr_master_350 하단그리드). 원천 PR_M_PROC_GAGONG_WORKER.
       WORKER_CODE=작업자명, WORK_FLAG='1'=실작업자. 실작업자 우선·이름순."""
    part = (part or '').strip()
    if not part:
        return {"part": part, "rows": [], "cnt": 0}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("""SELECT ISNULL(WORKER_CODE,''), ISNULL(WORK_FLAG,''),
              ISNULL(INSERT_USER_ID,''), CONVERT(varchar(19),INSERT_DATETIME,120),
              ISNULL(UPDATE_USER_ID,''), CONVERT(varchar(19),UPDATE_DATETIME,120)
            FROM PR_M_PROC_GAGONG_WORKER WHERE GAGONG_PROC_CODE=?
            ORDER BY WORK_FLAG DESC, WORKER_CODE""", part)
        rows = [{"worker": str(r[0]).strip(), "real": str(r[1]).strip() == '1',
                 "ins_user": str(r[2] or '').strip(), "ins_dt": str(r[3] or '').strip(),
                 "upd_user": str(r[4] or '').strip(), "upd_dt": str(r[5] or '').strip()} for r in cur.fetchall()]
        return {"part": part, "rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

# ================= 준비실적처리(키팅) 그리드 (w_pr_input_460_new ≈ dw_pr_input_410_t1_new2) =================
@app.get("/api/kitting/grid")
def kitting_grid(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
                 part: str = Query(""), pgroup: str = Query(""), line: str = Query(""),
                 assy: str = Query(""), jado: str = Query(""), gigan: int = Query(2),
                 wh_part: str = Query("IS0001"),
                 view: str = Query("전체"), unfin: str = Query("전체"), limit: int = Query(20000)):
    """준비실적처리(키팅) 그리드 — ★레거시 정본 SP `SP_PR_CREATE_PLAN_파트별_생산계획계산_생산준비등록_NEW` 로직 복제(실행X, .sql 이식).
       source=PR_T_PLAN_PART_COPY, 필터 GC_GUBUN='P'(생산파트)·GAGONG_PROC_SEQ=1·투입파트(WH_GAGONG_PROC_CODE=@wh_part, 기본 IS0001, BOM CTE).
       ★본행 grain = (GAGONG_PROC_CODE, WORK_ORDER, SPLIT_WORK_ORDER, ASSY_ITEM_CODE, UPPER_ITEM_CODE, ITEM_CODE).
       당일이전(plan_qty_00)=part_plan_ymd<from · 일자셀=from+N(달력일). 재고충당 순서=출하(sa_t_sale_dtl,tag90)→ASSY재고(sa_t_item_stock×use,tag70)
       →준비재고(pu_t_ready_stock cust=Z99990·proc=파트,tag50). finish_qty_NN=finish+ready. finish_tag→color: 90주황/70노랑/50·10녹/30회색/else백.
       입력값=라이브 직독, 색=SP로직 복제. ★라이브 PARTNER_ERP 읽기전용(SP 미실행)."""
    from datetime import datetime as _dt, timedelta as _td
    def _yadd(y6, n):
        try: return (_dt.strptime('20' + y6, '%Y%m%d') + _td(days=n)).strftime('%y%m%d')
        except Exception: return y6
    cn = _conn(); cur = cn.cursor()
    try:
        d6a = _d6(from_ymd) or _dt.now().strftime('%y%m%d')
        d6b = _d6(to_ymd) or _yadd(d6a, max(0, int(gigan) - 1))   # to = from + (기간-1) 달력일
        dates = [_yadd(d6a, i) for i in range(max(1, int(gigan)))]  # 지평 일자셀(달력일)
        whp = (wh_part.strip() or 'IS0001')
        # ★성능: SP #TEMP_CTE(투입파트 재귀BOM)를 메인쿼리 JOIN에서 분리 → KEYS set 선(先)조회 후 파이썬 필터
        #   (재귀CTE를 메인 GROUP 조인에 인라인하면 재구체화로 ~5초. 분리 시 ~1.5초. 값·색 로직 불변)
        keys = set()
        try:
            cur.execute("""
                ;WITH CTE (ITEM_CODE, MAT_CODE, GAGONG_PROC_CODE, WH_GAGONG_PROC_CODE, VIR_ITEM_FLAG) AS (
                     SELECT a.ITEM_CODE, B.MAT_CODE, B.GAGONG_PROC_CODE, B.WH_GAGONG_PROC_CODE, B.VIR_ITEM_FLAG
                       FROM PR_T_PLAN_PART_COPY a WITH(NOLOCK) JOIN pr_m_item_bom B WITH(NOLOCK) ON A.ITEM_CODE=B.ITEM_CODE
                      WHERE a.part_plan_ymd BETWEEN '' AND ? AND a.GC_GUBUN='P'
                     UNION ALL
                     SELECT a.ITEM_CODE, B.MAT_CODE, B.GAGONG_PROC_CODE, B.WH_GAGONG_PROC_CODE, B.VIR_ITEM_FLAG
                       FROM CTE a JOIN pr_m_item_bom B WITH(NOLOCK) ON A.MAT_CODE=B.ITEM_CODE WHERE A.VIR_ITEM_FLAG='1'
                )
                SELECT DISTINCT ITEM_CODE, GAGONG_PROC_CODE FROM CTE WHERE WH_GAGONG_PROC_CODE=? OPTION(MAXRECURSION 0)""", d6b, whp)
            for rr in cur.fetchall(): keys.add((rr[0], rr[1]))
        except Exception: pass
        w = ["a.part_plan_ymd<=?", "a.GC_GUBUN='P'", "a.GAGONG_PROC_SEQ=1"]; p = [d6b]
        if wc.strip():     w.append("a.WORK_CODE=?"); p.append(wc.strip())
        if part.strip():   w.append("a.GAGONG_PROC_CODE=?"); p.append(part.strip())
        if line.strip():   w.append("a.LINE_NO=?"); p.append(line.strip())
        if assy.strip():   w.append("a.ASSY_ITEM_CODE LIKE ?"); p.append(f"%{assy.strip()}%")   # 도번 필터=ASSY
        if jado.strip():   w.append("a.ITEM_CODE LIKE ?"); p.append(f"%{jado.strip()}%")          # 자도번 필터=도번(item)
        if pgroup.strip(): w.append("pg.PART_GROUP_CODE=?"); p.append(pgroup.strip())
        cur.execute(f"""SELECT TOP {int(limit) * 40}
              a.ASSY_ITEM_CODE assy, a.UPPER_ITEM_CODE upper, a.ITEM_CODE item,
              a.GAGONG_PROC_CODE gpc, COALESCE(pg.GAGONG_PROC_DESC, a.GAGONG_PROC_CODE) gpcnm,
              ISNULL(pg.PART_GROUP_CODE,'') pgc, a.WORK_CODE wc,
              COALESCE(wk.WORK_DESC, cu.CUST_DESC, a.WORK_CODE) wcnm, MAX(ISNULL(a.LINE_NO,'')) line,
              a.WORK_ORDER wo, a.SPLIT_WORK_ORDER swo, a.PART_PLAN_YMD ymd,
              MAX(ISNULL(a.PART_OUTPUT_HM,'')) inhm, ISNULL(ib.ITEM_DESC,'') nm,
              ISNULL(pg.PROD_RATE,100) rate, ISNULL(st.st,0) st, MAX(CAST(ISNULL(a.USE_QTY,1) AS float)) useq,
              MIN(ISNULL(a.PLAN_YMD,'')) plan_ymd, SUM(CAST(a.PART_PLAN_QTY AS float)) pl
            FROM PR_T_PLAN_PART_COPY a WITH(NOLOCK)
            JOIN pr_m_item b WITH(NOLOCK) ON a.ASSY_ITEM_CODE=b.ITEM_CODE
            JOIN pr_m_item ib WITH(NOLOCK) ON a.ITEM_CODE=ib.ITEM_CODE
            JOIN PR_M_PROC_GAGONG pg WITH(NOLOCK) ON a.GAGONG_PROC_CODE=pg.GAGONG_PROC_CODE
            LEFT JOIN PR_M_WORK wk WITH(NOLOCK) ON wk.WORK_CODE=a.WORK_CODE
            LEFT JOIN CM_M_CUST cu WITH(NOLOCK) ON cu.CUST_CODE=pg.IN_CUST_CODE
            LEFT JOIN (SELECT ITEM_CODE, SUM(CAST(ISNULL(TOT_ST,0) AS float)) st FROM PR_M_ITEM_PROC_GAGONG GROUP BY ITEM_CODE) st ON st.ITEM_CODE=a.ITEM_CODE
            WHERE {' AND '.join(w)}
            GROUP BY a.GAGONG_PROC_CODE, COALESCE(pg.GAGONG_PROC_DESC, a.GAGONG_PROC_CODE), ISNULL(pg.PART_GROUP_CODE,''),
              a.WORK_CODE, COALESCE(wk.WORK_DESC, cu.CUST_DESC, a.WORK_CODE), a.WORK_ORDER, a.SPLIT_WORK_ORDER,
              a.ASSY_ITEM_CODE, a.UPPER_ITEM_CODE, a.ITEM_CODE, a.PART_PLAN_YMD, ISNULL(ib.ITEM_DESC,''),
              ISNULL(pg.PROD_RATE,100), ISNULL(st.st,0)""", *p)
        cols = [d[0] for d in cur.description]
        raw = [d for d in (dict(zip(cols, r)) for r in cur.fetchall()) if (d["item"], d["gpc"]) in keys]   # ★투입파트 KEYS 필터
        # ── 본행 grain = (gpc,wo,swo,assy,upper,item), 일자셀 = 달력일 피벗 ──
        keyed = {}
        for r in raw:
            q = float(r["pl"] or 0); ymd = r["ymd"]
            bucket = 'P' if ymd < d6a else (ymd if ymd in dates else None)
            if bucket is None: continue   # to 초과분 방어(WHERE로 이미 <=to)
            k = (r["gpc"], r["wo"], r["swo"] or '', r["assy"], r["upper"] or '', r["item"])
            g = keyed.get(k)
            if not g:
                g = {"assy": r["assy"], "upper": r["upper"] or '', "item": r["item"], "nm": r["nm"],
                     "gpc": r["gpc"], "gpcnm": r["gpcnm"], "pgc": r["pgc"], "wc": r["wc"], "wcnm": r["wcnm"],
                     "line": r["line"], "inhm": r["inhm"], "rate": float(r["rate"] or 100),
                     "item_st": float(r["st"] or 0), "use_qty": float(r["useq"] or 1),
                     "wo": r["wo"], "swo": r["swo"] or '', "plan_ymd": (r["plan_ymd"] or ''),
                     "days": {}, "prior_plan": 0.0, "plan_qty": 0.0, "_cells": {}}
                keyed[k] = g
            if (r["plan_ymd"] or '') and (not g["plan_ymd"] or (r["plan_ymd"] or '') < g["plan_ymd"]): g["plan_ymd"] = r["plan_ymd"]
            cell = g["_cells"].get(bucket)
            if not cell:
                cell = {"bucket": bucket, "ymd": ymd, "plan": 0.0, "finish": 0.0, "ready": 0.0, "tag": 0}
                g["_cells"][bucket] = cell
            cell["plan"] += q
            if bucket == 'P': g["prior_plan"] += q
            else: g["days"][bucket] = g["days"].get(bucket, 0.0) + q
            g["plan_qty"] += q
        rows = list(keyed.values())
        capped = len(rows) >= int(limit); rows = rows[:int(limit)]
        # ── 충당 소스 조회(라이브 직독, SP 소스와 동일) ──
        rstock = {}; assystk = {}; saled = {}; nxcell = {}; midstk = {}; fixstk = {}
        try:  # 준비재고: pu_t_ready_stock cust='Z99990', (proc_gubun=파트, item)
            cur.execute("SELECT ITEM_CODE, PROC_GUBUN, SUM(STOCK_QTY) FROM PU_T_READY_STOCK WHERE CUST_CODE='Z99990' GROUP BY ITEM_CODE, PROC_GUBUN")
            for rr in cur.fetchall(): rstock[(rr[0], rr[1] or '')] = float(rr[2] or 0)
        except Exception: pass
        try:  # ASSY 현재고: sa_t_item_stock (item)
            cur.execute("SELECT ITEM_CODE, SUM(STOCK_QTY) FROM SA_T_ITEM_STOCK GROUP BY ITEM_CODE")
            for rr in cur.fetchall(): assystk[rr[0]] = float(rr[1] or 0)
        except Exception: pass
        # ★중간공정 파트재고 롤업(SP #TEMP_MAT_STOCK T_SUB_CTE): 자재/생산/사급/스태커 재고 + 재귀BOM 도번고정 → tag70.
        #   ★필터 무관 전역 재고롤업이라 색(tag70)에만 영향(값/개수/계획합계는 매요청 라이브 재조회) → 90초 TTL 캐시로 재귀비용 회피(~2초 유지).
        _cache = getattr(kitting_grid, "_rollup_cache", None)
        _now = _dt.now().timestamp()
        if _cache and (_now - _cache["ts"] < 90) and _cache["mid"]:
            midstk = _cache["mid"]; fixstk = _cache["fix"]
        else:
            try:
                cur.execute("IF OBJECT_ID('tempdb..#tms') IS NOT NULL DROP TABLE #tms")   # 풀링 재사용 대비 선정리(#temp만, 가드 통과=IF 시작)
                cur.execute("""
                    ;WITH T_SUB_CTE (item_code, upper_item_code, mat_code, stock_qty, pr_stock_qty, fix_pr_stock_qty) AS (
                        SELECT s.mat_code, s.mat_code, s.mat_code,
                               CONVERT(int, ISNULL(SUM(s.stock_qty),0)), CONVERT(int, ISNULL(SUM(s.pr_stock_qty),0)), 0
                          FROM ( SELECT mat_code, 0 stock_qty, STOCK_QTY pr_stock_qty FROM pr_t_mat_stock_wh WITH(NOLOCK)
                                 UNION ALL SELECT a.mat_code,0,a.STOCK_QTY FROM PU_T_SAGUB_STOCK a WITH(NOLOCK) JOIN pr_m_item m WITH(NOLOCK) ON a.MAT_CODE=m.ITEM_CODE WHERE m.SAGUB_STOCK_FLAG='1'
                                 UNION ALL SELECT mat_code, stock_qty, 0 FROM pu_t_mat_stock_wh WITH(NOLOCK) WHERE cust_code='Z99990' AND gagong_proc_code NOT IN ('SA1','SA2','SB1','SB2')
                                 UNION ALL SELECT mat_code, stock_qty, 0 FROM PU_T_STACKER_STOCK WITH(NOLOCK) ) s
                         GROUP BY s.mat_code HAVING SUM(s.stock_qty)<>0 OR SUM(s.pr_stock_qty)<>0
                        UNION ALL
                        SELECT cb.item_code, b.item_code, b.mat_code, 0, 0,
                               CONVERT(int, (CASE WHEN cb.fix_pr_stock_qty<>0 THEN cb.fix_pr_stock_qty ELSE (cb.pr_stock_qty+cb.stock_qty) END) * b.use_qty)
                          FROM T_SUB_CTE cb JOIN pr_m_item_bom b WITH(NOLOCK) ON cb.mat_code=b.item_code WHERE ISNULL(b.except_flag,'0')<>'1'
                    )
                    SELECT item_code, upper_item_code, mat_code, stock_qty, pr_stock_qty, fix_pr_stock_qty INTO #tms FROM T_SUB_CTE OPTION(MAXRECURSION 0)""")
                cur.execute("SELECT mat_code, SUM(stock_qty), SUM(pr_stock_qty) FROM #tms GROUP BY mat_code")   # 자재+생산재고(item)
                for rr in cur.fetchall(): midstk[rr[0]] = float(rr[1] or 0) + float(rr[2] or 0)
                cur.execute("SELECT upper_item_code, mat_code, SUM(fix_pr_stock_qty) FROM #tms GROUP BY upper_item_code, mat_code")  # 도번고정(upper,item)
                for rr in cur.fetchall(): fixstk[(rr[0], rr[1])] = float(rr[2] or 0)
                kitting_grid._rollup_cache = {"ts": _now, "mid": midstk, "fix": fixstk}
            except Exception: pass
        try:  # 출하: sa_t_sale_dtl (wo, split, item=assy, finish_flag='0') — ★결과 WORK_ORDER로 제한(전체 GROUP BY 3.7s→회피)
            wos = list({g["wo"] for g in rows if g["wo"]})
            for i in range(0, len(wos), 900):
                ck = wos[i:i + 900]; ph = ",".join("?" * len(ck))
                cur.execute(f"SELECT WORK_ORDER, ISNULL(SPLIT_WORK_ORDER,''), ITEM_CODE, SUM(SALE_QTY) FROM SA_T_SALE_DTL WHERE FINISH_FLAG='0' AND WORK_ORDER IN ({ph}) GROUP BY WORK_ORDER, ISNULL(SPLIT_WORK_ORDER,''), ITEM_CODE", *ck)
                for rr in cur.fetchall(): saled[(rr[0], rr[1] or '', rr[2])] = float(rr[3] or 0)
        except Exception: pass
        try:  # ★Phase1: nx 셀단위 준비 flag = 단일원장 nx.stock_ledger(STOCK_POINT='RDY') SUM. (item×wo×파트gpc×일자INPUT_YMD)
            nxc = _nx(); nc = nxc.cursor()
            nc.execute("""SELECT ITEM_CODE, ISNULL(WORK_ORDER,''), ISNULL(GAGONG_PROC_CODE,''), ISNULL(INPUT_YMD,''), ISNULL(SUM(MAINT_QTY),0)
                FROM nx.stock_ledger WHERE STOCK_POINT='RDY'
                GROUP BY ITEM_CODE, ISNULL(WORK_ORDER,''), ISNULL(GAGONG_PROC_CODE,''), ISNULL(INPUT_YMD,'')""")
            for rr in nc.fetchall(): nxcell[(rr[0], rr[1], rr[2], rr[3])] = float(rr[4] or 0)
            nxc.close()
        except Exception: pass
        # finish_tag → color(fin) 매핑: 90출하→'6' / 70생산→'4' / 50·10준비→'3' / 30자재→'2' / else '0'
        _TAG2FIN = {90: '6', 70: '4', 50: '3', 10: '3', 30: '2'}
        def _alloc(cellseq, pool, tag, key):
            """SP 커서 재고충당: 계획순 셀에 pool 충당. 완전충당 셀=tag, 부분=태그유지. key='finish' or 'ready'."""
            pool = max(float(pool or 0), 0.0)
            for c in cellseq:
                if pool <= 0: break
                jan = c["plan"] - c["finish"] - (c["ready"] if key == 'ready' else 0.0)
                if jan <= 0: continue
                if jan > pool:
                    c[key] += pool; pool = 0.0                       # 부분충당 → tag 미변경(NULL)
                else:
                    c[key] += jan; pool -= jan
                    if tag > c["tag"] or c["tag"] == 0: c["tag"] = tag  # 완전충당 → tag(최고단계 유지)
        for g in rows:
            it = g["item"]
            g["part_ymd"] = min([c["ymd"] for c in g["_cells"].values()] or [''])   # 당일이전 셀 키(=최소 계획일)
            seq = ([g["_cells"]['P']] if 'P' in g["_cells"] else []) + [g["_cells"][y] for y in dates if y in g["_cells"]]
            # 1) 출하(sale, tag90) — pool=sa_t_sale_dtl[(wo,swo,assy)]
            _alloc(seq, saled.get((g["wo"], g["swo"], g["assy"]), 0.0), 90, 'finish')
            # 2) ASSY 현재고(tag70) — pool=sa_t_item_stock[assy] × use_qty
            _alloc(seq, assystk.get(g["assy"], 0.0) * g["use_qty"], 70, 'finish')
            # 2-1) 도번고정재고(tag70) — pool=재귀BOM 롤업 fixstk[(upper,item)] (SP 도번고정 감안)
            _alloc(seq, max(fixstk.get((g["upper"], it), 0.0), 0.0), 70, 'finish')
            # 2-2) 중간공정 파트재고(tag70) — pool=자재재고+생산재고 midstk[item] (SP 파트재고 감안)
            _alloc(seq, max(midstk.get(it, 0.0), 0.0), 70, 'finish')
            # 3) 준비재고(tag50) — pool=pu_t_ready_stock[(item,파트)]  → ready_qty
            _alloc(seq, max(rstock.get((it, g["gpc"]), 0.0), 0.0), 50, 'ready')
            # 4) ★nx 셀단위 준비 flag 오버레이(우리 확인분, 셀별) — 라이브 PU와 별도 합산(이중가산X), 커버 시 tag50 녹
            for c in seq:
                ck = g["part_ymd"] if c["bucket"] == 'P' else c["ymd"]
                nq = nxcell.get((it, g["wo"], g["gpc"], ck), 0.0)
                if nq > 0:
                    rem = max(c["plan"] - c["finish"] - c["ready"], 0.0)
                    if rem > 0: c["ready"] += min(nq, rem)
                    if c["plan"] > 0 and (c["finish"] + c["ready"]) >= c["plan"] and c["tag"] < 50: c["tag"] = 50
            # 셀 표시: finish_qty_NN = finish + ready; fin = tag매핑
            g["dcov"] = {}; g["dfin"] = {}
            pc = g["_cells"].get('P')
            g["prior_cover"] = round((pc["finish"] + pc["ready"]), 2) if pc else 0.0
            g["prior_fin"] = _TAG2FIN.get(pc["tag"], '0') if pc else '0'
            for y in g["days"]:
                c = g["_cells"].get(y)
                g["dcov"][y] = round((c["finish"] + c["ready"]), 2) if c else 0.0
                g["dfin"][y] = _TAG2FIN.get(c["tag"], '0') if c else '0'
            g["finish"] = round(sum(c["finish"] for c in g["_cells"].values()), 2)         # 완료수량=충당 finish합(SP finish_qty)
            g["ready_stock"] = round(max(rstock.get((it, g["gpc"]), 0.0), 0.0), 2)          # 준비재고(파트버킷)
            g["ready_qty"] = round(sum(c["ready"] for c in g["_cells"].values()), 2)        # 준비수량=충당 ready합(SP ready_qty)
            g["prod_stock"] = round(assystk.get(it, 0.0), 2)
            g["assy_stock"] = round(assystk.get(g["assy"], 0.0), 2)
            g["sale"] = round(saled.get((g["wo"], g["swo"], g["assy"]), 0.0), 2)
            g["need_qty"] = round(max(g["plan_qty"] - g["ready_qty"], 0.0), 2)
            fins = [_TAG2FIN.get(c["tag"], '0') for c in g["_cells"].values() if c["plan"] > 0]
            g["fin"] = (g["prior_fin"] if g["prior_plan"] > 0 else (g["dfin"][min(g["days"])] if g["days"] else '0'))
            g["_done_all"] = bool(fins) and all(f in ('4', '6') for f in fins)
            g["_has_unkit"] = any(f == '0' for f in fins)
            g["splits"] = [{"gpc": g["gpc"], "gpcnm": g["gpcnm"], "prior_plan": g["prior_plan"], "days": dict(g["days"])}]
            del g["_cells"]
        uf = unfin.strip()
        if uf == '미생산':   rows = [r for r in rows if not r["_done_all"]]
        elif uf == '미키팅': rows = [r for r in rows if r["_has_unkit"]]
        for r in rows: r.pop("_done_all", None); r.pop("_has_unkit", None)
        # 정렬 = 레거시 DW sort=: part_plan_ymd_output_hm → plan_ymd → gagong_proc_code → output_hm → work_order → split_work_order
        rows.sort(key=lambda x: ((x["part_ymd"] or "") + (x["inhm"] or ""), x["plan_ymd"] or "",
                                 x["gpc"] or "", x["inhm"] or "", x["wo"] or "", x["swo"] or ""))
        note = f"⚠ 상위 {limit}건 초과 — 투입파트·작업처·도번으로 필터하세요." if capped else ""
        return {"dates": dates, "rows": rows, "cnt": len(rows),
                "plan_sum": sum(r["plan_qty"] for r in rows), "ready_sum": sum(r["ready_qty"] for r in rows), "note": note}
    finally:
        cn.close()

def _kit_cell_guard(item, wo, swo, gpc, ymd, qty, assy):
    """셀 확인/취소 서버 가드(라이브 RO): 월마감(PU_T_MONTH_READY_STOCK) 이후·출하완료분 금지. (ok, detail)."""
    cn = _conn(); cur = cn.cursor()
    try:
        cellm = _d6(ymd) if ymd else ''
        try:
            cur.execute("SELECT ISNULL(MAX(stock_yymm),'0000') FROM pu_t_month_ready_stock")
            mclose = (cur.fetchone()[0] or '0000')
        except Exception:
            mclose = '0000'
        if cellm and len(cellm) == 6 and cellm <= (mclose + '99'):   # 셀 일자가 마감월 이내면 금지
            return (False, f"월마감({mclose}) 완료 일자 — 확인/취소 불가")
        if assy and qty > 0:   # 출하완료분 금지(sa_t_sale_dtl 미마감 출하 ≥ 셀잔량)
            try:
                cur.execute("SELECT ISNULL(SUM(SALE_QTY),0) FROM SA_T_SALE_DTL WHERE FINISH_FLAG='0' AND WORK_ORDER=? AND ISNULL(SPLIT_WORK_ORDER,'')=? AND ITEM_CODE=?",
                            wo, (swo or ''), assy)
                if float(cur.fetchone()[0] or 0) >= qty:
                    return (False, "출하완료분 — 키팅 확인 불가")
            except Exception: pass
        return (True, "")
    finally:
        cn.close()

@app.post("/api/kitting/cell-confirm")
def kitting_cell_confirm(payload: dict = Body(...)):
    """준비실적처리 셀단위 준비완료 등록(레거시 250창 우클릭 '확인'). flag-only(자재무차감) — nx.ready_ledger INSERT(tag '1').
       키=item_code·work_order·proc_code(파트gpc)·plan_ymd(셀 일자, 당일이전=행 part_ymd). qty=셀 잔량(계획−완료). ★쓰기 nx만."""
    item = (payload.get("item") or "").strip(); wo = (payload.get("wo") or "").strip()
    swo = (payload.get("swo") or "").strip(); gpc = (payload.get("gpc") or "").strip()
    ymd = _d6(payload.get("ymd") or ""); qty = float(payload.get("qty") or 0)
    assy = (payload.get("assy") or "").strip()
    user = (str(payload.get("user", "") or "").strip() or "웹사용자")[:20]
    if not item or not gpc or qty <= 0: return {"ok": False, "detail": "item·파트·수량(>0) 필수"}
    ok, detail = _kit_cell_guard(item, wo, swo, gpc, ymd, qty, assy)
    if not ok: return {"ok": False, "detail": detail}
    nx = _nx(); nc = nx.cursor()
    try:  # ★Phase1: 단일원장 nx.stock_ledger(STOCK_POINT='RDY', tag K1=+확인). 셀키=item·wo·gpc·plan_ymd(INPUT_YMD). flag-only(자재무차감)
        nc.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=RIGHT(CONVERT(varchar(8),GETDATE(),112),6)")
        seq = int(nc.fetchone()[0] or 1)
        nc.execute("""INSERT INTO nx.stock_ledger(STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,CUST_CODE,ITEM_CODE,GAGONG_PROC_CODE,
              WORK_ORDER,INPUT_YMD,MAINT_QTY,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
            VALUES('RDY',RIGHT(CONVERT(varchar(8),GETDATE(),112),6),?,'K1','Z99990',?,?,?,?,?,'키팅확인',?,GETDATE())""",
            seq, item, gpc, (wo or None), (ymd or None), qty, user)
        return {"ok": True, "qty": qty}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        nx.close()

@app.post("/api/kitting/cell-cancel")
def kitting_cell_cancel(payload: dict = Body(...)):
    """준비실적처리 셀단위 준비취소(레거시 250창 우클릭 '취소'). ★Phase1: nx.stock_ledger(RDY) 상쇄 INSERT(tag K2, −qty). 잔량 이내. 쓰기 nx만."""
    item = (payload.get("item") or "").strip(); wo = (payload.get("wo") or "").strip()
    swo = (payload.get("swo") or "").strip(); gpc = (payload.get("gpc") or "").strip()
    ymd = _d6(payload.get("ymd") or ""); assy = (payload.get("assy") or "").strip()
    user = (str(payload.get("user", "") or "").strip() or "웹사용자")[:20]
    if not item or not gpc: return {"ok": False, "detail": "item·파트 필수"}
    nx = _nx(); nc = nx.cursor()
    try:  # 셀 현재 net(RDY 원장 우리 flag) 계산 → 그 이내로 취소
        nc.execute("""SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE STOCK_POINT='RDY'
              AND ITEM_CODE=? AND ISNULL(GAGONG_PROC_CODE,'')=? AND ISNULL(WORK_ORDER,'')=? AND ISNULL(INPUT_YMD,'')=?""",
                   item, gpc, (wo or ''), (ymd or ''))
        cur_net = float(nc.fetchone()[0] or 0)
        if cur_net <= 0: return {"ok": False, "detail": "취소할 준비완료(우리 확인분) 없음"}
        req = float(payload.get("qty") or 0)
        cancel = min(req, cur_net) if req > 0 else cur_net
        ok, detail = _kit_cell_guard(item, wo, swo, gpc, ymd, cancel, assy)
        if not ok: return {"ok": False, "detail": detail}
        nc.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=RIGHT(CONVERT(varchar(8),GETDATE(),112),6)")
        seq = int(nc.fetchone()[0] or 1)
        nc.execute("""INSERT INTO nx.stock_ledger(STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,CUST_CODE,ITEM_CODE,GAGONG_PROC_CODE,
              WORK_ORDER,INPUT_YMD,MAINT_QTY,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
            VALUES('RDY',RIGHT(CONVERT(varchar(8),GETDATE(),112),6),?,'K2','Z99990',?,?,?,?,?,'키팅취소',?,GETDATE())""",
            seq, item, gpc, (wo or None), (ymd or None), -cancel, user)
        return {"ok": True, "qty": cancel}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        nx.close()

# ================= ★Phase2: 생산실적 백플러시 엔진 (실사용BOM×생산량 소비, 회수율 제외) =================
def _is_inner_prod(cro, item):
    """사내생산(INNER_PROD=1) 판정: MAKE_TYPE='1' 또는 가공공정(PR_M_ITEM_PROC_GAGONG) 보유. 라이브 RO."""
    c = cro.cursor()
    try:
        c.execute("SELECT ISNULL(MAKE_TYPE,'') FROM PR_M_ITEM WHERE ITEM_CODE=?", item)
        r = c.fetchone()
        if r and str(r[0]).strip() == '1': return True
        c.execute("SELECT COUNT(*) FROM PR_M_ITEM_PROC_GAGONG WHERE ITEM_CODE=?", item)
        return (c.fetchone()[0] or 0) > 0
    except Exception:
        return False

def _backflush_bom(nxc, root, cro=None):
    """실사용BOM 전개(nx.bom): 제작서브(children보유·is_lowest≠Y) 전개, 최말단 자재/구매품 소비.
       용접봉(role='용접봉')=공정종속 → ★별도수집(완성공정 1회 함께 소비, base RAC 코드별 종류별. 정본 qty=nx.bom 재빌드된 CS_M_ITEM_BOM.USE_QTY=ITEM_USE_QTY×1.5).
       ★사내한정 가드: 용접봉 −W는 사내 용접(부모노드 root=INNER_PROD 또는 MAKE_TYPE='1' 제작)만. 외주 용접봉은 사급출고(tag5)로 이미 −재고 → 이중차감 방지(결정 I). cro=라이브RO(사내판정), None=전량(하위호환).
       반환 (comps[(child,cum_qty)], weld{base_rac:cum_qty}). 회수율 미개입."""
    c = nxc.cursor()
    c.execute("SELECT parent_code, child_code, CAST(qty AS float), ISNULL(role,''), ISNULL(is_lowest,'') FROM nx.bom")
    kids = {}
    for p, ch, q, role, low in c.fetchall():
        kids.setdefault(p, []).append((ch, q or 0.0, role, low))
    _mkc = {}
    def _sanae(node):   # 사내 용접 판정: root(INNER_PROD 게이트) 또는 부모 MAKE_TYPE='1'(제작)
        if node == root: return True
        if cro is None: return True
        n = str(node).strip()
        if n not in _mkc:
            cc = cro.cursor(); cc.execute("SELECT ISNULL(MAKE_TYPE,'') FROM PARTNER_ERP.dbo.PR_M_ITEM WHERE ITEM_CODE=?", n)
            r = cc.fetchone(); _mkc[n] = bool(r and str(r[0]).strip() == '1')
        return _mkc[n]
    out = {}; weld = {}
    def walk(node, mult, depth):
        if depth > 15: return
        for ch, q, role, low in kids.get(node, []):
            cq = mult * q
            if '용접봉' in (role or ''):                    # ★용접봉=공정종속
                if str(ch).upper().startswith('RAC') and _sanae(node):   # RAC + 사내용접만 −W(외주=사급출고 이미 −재고)
                    weld[str(ch).split('-')[0]] = weld.get(str(ch).split('-')[0], 0.0) + cq
                continue                                    # 그 외 role=용접봉(3H·용접SUB)·외주용접봉 = 스킵
            if ch in kids and str(low) != 'Y':             # 제작 서브 → 전개
                walk(ch, cq, depth + 1)
            else:                                          # 소비 leaf(자재/구매품)
                out[ch] = out.get(ch, 0.0) + cq
    walk(root, 1.0, 0)
    return list(out.items()), weld

def _weld_proc_code(nxc, base_rac):
    """용접봉 투입공정(GAGONG_PROC_CODE) — nx.bom_line 대표값(Q1000/Q2000 용접봉창고), 없으면 'Q1000' 기본."""
    c = nxc.cursor()
    c.execute("SELECT TOP 1 ISNULL(gagong_proc,'') FROM nx.bom_line WHERE child_item LIKE ? AND ISNULL(gagong_proc,'')<>'' ORDER BY seq", base_rac + '%')
    r = c.fetchone()
    return (str(r[0]).strip() if r and r[0] else 'Q1000')

def _final_proc_code(cro, item):
    """완성공정(최종) gagong_proc_code = MAX(PROC_SEQ). method 무관·PROC_SEQ 최댓값. 라이브 RO."""
    c = cro.cursor()
    try:
        c.execute("SELECT TOP 1 ISNULL(GAGONG_PROC_CODE,'') FROM PR_M_ITEM_PROC_GAGONG WHERE ITEM_CODE=? ORDER BY PROC_SEQ DESC", item)
        r = c.fetchone()
        return str(r[0]).strip() if r and r[0] else ""
    except Exception:
        return ""

def _is_final_product(nxc, item):
    """최종제품(ASY) 판정: nx.bom에 child로 없으면 최상위=제품(ASY), child면 반제품(PRD)."""
    c = nxc.cursor()
    c.execute("SELECT COUNT(*) FROM nx.bom WHERE child_code=?", item)
    return (c.fetchone()[0] or 0) == 0

def _backflush_core(cro, nx, item, prod_qty, wo, gpc, mode, user, ref_key, ref_bc=None):
    """★백플러시 코어(트랜잭션 미관리 — 호출측 commit/rollback). cro=RO conn, nx=쓰기 tx conn.
       완성공정 1회 전체BOM×생산량 소비(−P4: RDY 우선 없으면 MAT) + 생산품 +ASY(최종제품)/+PRD(반제품, tag P7).
       회수율 제외. INNER_PROD=1만. 멱등=ref_key(바코드=BC:{barcode}:{proc} / 수기=wo|item|ymd)."""
    nc = nx.cursor()
    if not item or prod_qty <= 0: return {"ok": False, "detail": "item·생산량(>0) 필수"}
    if not _is_inner_prod(cro, item): return {"ok": False, "detail": "사내생산(INNER_PROD=1) 아님 — 백플러시 제외(사급회수·매입·직납)"}
    import datetime as _d
    ymd6 = _d.datetime.now().strftime('%y%m%d')
    nc.execute("SELECT bf_id FROM nx.backflush_log WHERE ref_key=? AND state='posted'", ref_key)
    ex = nc.fetchone()
    if mode == "post" and ex: return {"ok": False, "detail": f"이미 백플러시됨(중복방지) — ref {ref_key}"}
    if mode == "reverse" and not ex: return {"ok": False, "detail": "되돌릴 백플러시 없음"}
    f = -1.0 if mode == "reverse" else 1.0
    comps, weld = _backflush_bom(nx, item, cro)   # ★cro=라이브RO(용접봉 사내한정 판정)
    if not comps and not weld: return {"ok": False, "detail": "nx.bom 전개결과 없음(소비 BOM 없음)"}
    out_sp = 'ASY' if _is_final_product(nx, item) else 'PRD'   # ★완성=최종제품 ASY / 반제품 PRD
    def _seq():
        nc.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd6)
        return int(nc.fetchone()[0] or 1)
    def _post(sp, child, qty, tag, remk, gpc_over=None):
        if abs(qty) < 1e-9: return
        nc.execute("""INSERT INTO nx.stock_ledger(STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,CUST_CODE,ITEM_CODE,MAT_CODE,
              GAGONG_PROC_CODE,WORK_ORDER,MAINT_QTY,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
            VALUES(?,?,?,?,'Z99990',?,?,?,?,?,?,?,GETDATE())""",
            sp, ymd6, _seq(), tag, (child if sp in ('PRD','ASY','RDY') else None),
            (child if sp == 'MAT' else None), (gpc_over or gpc or None), (wo or None), qty, remk, user)
            # ★RDY도 ITEM_CODE축(키팅 예약과 정합) / MAT만 MAT_CODE축 — −RDY가 키팅 +RDY를 정확히 상쇄
    seq_from = _seq(); consumed = 0.0
    for child, cq in comps:                       # 소비(−P4): RDY 우선 없으면 MAT
        need = cq * prod_qty
        nc.execute("SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger WHERE STOCK_POINT='RDY' AND ITEM_CODE=?", child)
        rdy = max(float(nc.fetchone()[0] or 0), 0.0)
        from_rdy = min(need, rdy); from_mat = need - from_rdy
        _post('RDY', child, -from_rdy * f, 'P4', '백플러시소비(준비)')
        _post('MAT', child, -from_mat * f, 'P4', '백플러시소비(자재)')
        consumed += need
    weld_consumed = 0.0                            # ★용접봉 소비(−MAT, tag 'W', base RAC, 투입공정): 완성공정 1회 자재와 함께
    for base_rac, wq in weld.items():
        wneed = wq * prod_qty
        if abs(wneed) < 1e-9: continue
        _post('MAT', base_rac, -wneed * f, 'W', '백플러시 용접봉소비', gpc_over=_weld_proc_code(nx, base_rac))
        weld_consumed += wneed
    _post(out_sp, item, prod_qty * f, 'P7', f'백플러시 생산입고({out_sp})')   # 생산품 +ASY/+PRD
    nc.execute("SELECT ISNULL(MAX(MAINT_SEQ),0) FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd6)
    seq_to = int(nc.fetchone()[0] or 0)
    if mode == "post":
        nc.execute("""INSERT INTO nx.backflush_log(prod_ymd,work_order,item_code,gpc,prod_qty,ref_key,ref_bc,state,maint_ymd,seq_from,seq_to,ins_user)
            VALUES(?,?,?,?,?,?,?, 'posted', ?,?,?,?)""",
            ymd6, (wo or None), item, (gpc or None), prod_qty, ref_key, ref_bc, ymd6, seq_from, seq_to, user)
    else:
        nc.execute("UPDATE nx.backflush_log SET state='reversed' WHERE bf_id=?", ex[0])
    # 협력사 용접봉 무게정산(weight_calc) 연계는 후속(TODO) — 여기선 물리적 재고소비만.
    return {"ok": True, "mode": mode, "item": item, "prod_qty": prod_qty, "out_point": out_sp,
            "components": len(comps), "consumed_qty": round(consumed, 3),
            "weld_kinds": len(weld), "weld_consumed": round(weld_consumed, 4), "ref_key": ref_key}

@app.post("/api/backflush/post")
def backflush_post(payload: dict = Body(...)):
    """수기 백플러시(테스트/보정). 실운영 자동트리거=바코드생산실적(procbc_save 완성공정). mode=post/reverse. INNER_PROD=1만. 쓰기 nx만."""
    item = (payload.get("item") or "").strip(); wo = (payload.get("work_order") or payload.get("wo") or "").strip()
    gpc = (payload.get("gpc") or "").strip(); prod_qty = float(payload.get("prod_qty") or 0)
    mode = str(payload.get("mode", "post")).strip()
    user = (str(payload.get("user", "") or "").strip() or "웹사용자")[:20]
    import datetime as _d
    ref_key = f"{wo}|{item}|{_d.datetime.now().strftime('%y%m%d')}"   # 수기 멱등키(WO·품목·일자)
    cn = _conn(); nx = _nx_tx()   # ★원자성: 소비(−P4)+생산입고(+P7/ASY)+backflush_log 동일 트랜잭션
    try:
        r = _backflush_core(cn, nx, item, prod_qty, wo, gpc, mode, user, ref_key)
        nx.commit() if r.get("ok") else nx.rollback()
        return r
    except Exception as e:
        try: nx.rollback()
        except Exception: pass
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close(); nx.close()

# ================= 가공생산진척관리(전표발행) (w_pr_input_420_new) — PR_T_PLAN_PART_DTL 스냅샷 직독 =================
@app.get("/api/gagong/prog420")
def gagong_prog420(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query("P2"),
                   part: str = Query(""), item: str = Query(""), jado: str = Query(""),
                   unfin: str = Query("전체"), limit: int = Query(3000)):
    """가공생산진척관리. ★레거시 암호화SP `SP_PR_가공생산진척관리_260602`(from,to,mat_work_code) 직접 실행 → 100% 동일.
       당일이전=index00, 기준일~기간=index01+. 완료=finish_qty_NN·셀색=color_NN(BGR long). 자도번작업처=assy_work_center(이름)."""
    cn = _conn(); cur = cn.cursor()
    try:
        import datetime as _dt
        d6f = _d6(from_ymd) if from_ymd else '260729'
        d6t = _d6(to_ymd) if to_ymd else d6f
        try:
            da = _dt.date(2000+int(d6f[:2]), int(d6f[2:4]), int(d6f[4:6])); db = _dt.date(2000+int(d6t[:2]), int(d6t[2:4]), int(d6t[4:6]))
        except Exception:
            da = db = _dt.date(2026, 7, 29)
        ndays = max(1, min(8, (db - da).days + 1))
        dates = [(da + _dt.timedelta(days=i)).strftime('%y%m%d') for i in range(ndays)]   # 기준일~기간
        cur.execute("EXEC [dbo].[SP_PR_가공생산진척관리_260602] ?, ?, ?", d6f, d6t, (wc.strip() or 'P2'))
        cols = [d[0].lower() for d in cur.description]; sp = [dict(zip(cols, r)) for r in cur.fetchall()]
        def _css(v):
            v = int(v or 0)
            if v in (16777215, 553648127) or v <= 0: return ''
            return 'background:#%02x%02x%02x' % (v % 256, (v // 256) % 256, (v // 65536) % 256)
        def _f(x):
            try: return float(x or 0)
            except Exception: return 0.0
        allitems = [x for x in ({str(r.get('c_item_code') or '') for r in sp} | {str(r.get('mat_code') or '') for r in sp}) if x]; nm = {}
        gpcs = [x for x in {str(r.get('gagong_proc_code') or '') for r in sp} if x]; gpn = {}
        for i in range(0, len(allitems), 1000):
            ck = allitems[i:i + 1000]; ph = ",".join("?" * len(ck))
            cur.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ck)
            for a, b in cur.fetchall(): nm[a] = b
        if gpcs:
            ph = ",".join("?" * len(gpcs))
            cur.execute(f"SELECT GAGONG_PROC_CODE, ISNULL(GAGONG_PROC_DESC,'') FROM PR_M_PROC_GAGONG WHERE GAGONG_PROC_CODE IN ({ph})", *gpcs)
            for a, b in cur.fetchall(): gpn[a] = b
        rows = []
        for r in sp:
            assy = str(r.get('assy_item_code') or ''); mat = str(r.get('mat_code') or '')
            if item.strip() and item.strip().lower() not in assy.lower(): continue
            if jado.strip() and jado.strip().lower() not in mat.lower(): continue
            gpc = str(r.get('gagong_proc_code') or '')
            _wcc = str(r.get('mat_work_code') or '').strip() or (wc.strip())   # 자도번작업처 코드 정본=mat_work_code(P2 등)
            _wcd = _ITEM_WORK.get(_wcc, _wcc)                                    # 코드→이름(P2→가공)
            g = {"assy": assy, "jado": mat, "jnm": nm.get(mat, ''), "gpcnm": gpn.get(gpc, gpc),
                 "wcc": _wcc, "wcd": _wcd, "st": round(_f(r.get('item_st')) * _f(r.get('plan_qty')) / 3600.0, 2),
                 "plan_qty": _f(r.get('plan_qty')), "finish": _f(r.get('finish_qty')),
                 "sale": _f(r.get('sale_qty')), "prs": _f(r.get('pr_stock_qty')),
                 "assyst": _f(r.get('assy_stock_qty')), "fixst": _f(r.get('fix_pr_stock_qty')),
                 "wo": r.get('work_order') or '',
                 "prior_pl": _f(r.get('plan_qty_00')), "prior_fn": _f(r.get('finish_qty_00')), "prior_bg": _css(r.get('color_00')),
                 "days": {}, "done": {}, "colors": {}}
            for i in range(1, ndays + 1):
                ii = '%02d' % i; ymd = dates[i - 1]
                g["days"][ymd] = _f(r.get('plan_qty_' + ii)); g["done"][ymd] = _f(r.get('finish_qty_' + ii)); g["colors"][ymd] = _css(r.get('color_' + ii))
            rows.append(g)
        uf = unfin.strip()
        if uf == "미생산": rows = [r for r in rows if r["finish"] < r["plan_qty"]]
        elif uf == "미키팅": rows = [r for r in rows if r["prior_fn"] <= 0]   # 근사(SP에 키팅상태 미분리)
        rows = rows[:int(limit)]
        return {"dates": dates, "rows": rows, "cnt": len(rows),
                "plan_sum": sum(r["plan_qty"] for r in rows), "done_sum": sum(r["finish"] for r in rows), "note": ""}
    finally:
        cn.close()

# ================= 4주간 가공계획현황 (w_pr_outside_410_work) — 도번×라인×작업처, 자도번LIST 묶기 =================
@app.get("/api/gagong/plan4w")
def gagong_plan4w(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
                  item: str = Query(""), part: str = Query(""), mat_flag: str = Query("1"), limit: int = Query(2500)):
    """레거시 정본 SP_PR_4주간_가공계획현황_250703 본문 인라인 재현(EXEC 권한없어 SELECT로).
       SP는 WO별 행 생성 → ★표시 grain=도번(c_item_code)로 묶음. 자도번LIST=f_find_cust_mat_list2,
       자도번작업처=mat_work_code(P2=가공, 필터), 작업처=work_center. dates[0]=plan_qty_01(당일이전누적 plan_ymd<=기준일),
       dates[k]=plan_qty_(k+1). 수량=ceiling(plan_qty×use_qty×prod_rate/100). 참조 _legacy_analysis/SP_DUMP."""
    import datetime as _dt, os, sys as _sys
    _bd = os.path.dirname(os.path.abspath(__file__))
    if _bd not in _sys.path: _sys.path.insert(0, _bd)
    from _sp_4wk import SQL_4WK
    cn = _conn(); cur = cn.cursor()
    try:
        d6f = _d6(from_ymd) if from_ymd else '260729'
        d6t = _d6(to_ymd) if to_ymd else None
        if not d6t:
            _y = _dt.date(2000+int(d6f[:2]), int(d6f[2:4]), int(d6f[4:6])) + _dt.timedelta(days=30)
            d6t = _y.strftime('%y%m%d')
        wcp = (wc.strip() or 'P2'); mf = (mat_flag.strip() or '1')
        sql = SQL_4WK
        for k, v in [('@@WC@@', wcp), ('@@MAT@@', '%'), ('@@FLAG@@', mf),
                     ('@@FROM@@', d6f), ('@@TO@@', d6t), ('@@ITEM@@', '%')]:
            sql = sql.replace(k, "'%s'" % v)
        cur.execute(sql)
        cols = [d[0].lower() for d in cur.description]; ix = {c: i for i, c in enumerate(cols)}
        def gv(r, n, d=None):
            i = ix.get(n); return r[i] if i is not None else d
        raw = cur.fetchall()
        # 날짜 캘린더: dates[0]=기준일(=plan_qty_01 당일이전누적), 이후 plan_qty_02..31
        da = _dt.date(2000+int(d6f[:2]), int(d6f[2:4]), int(d6f[4:6]))
        db = _dt.date(2000+int(d6t[:2]), int(d6t[2:4]), int(d6t[4:6]))
        dates = []; cu = da
        while cu <= db and len(dates) < 31: dates.append(cu.strftime('%y%m%d')); cu += _dt.timedelta(days=1)
        keyed = {}
        for r in raw:
            doban = str(gv(r, 'c_item_code', '') or '').strip()
            if not doban: continue
            g = keyed.get(doban)
            if not g:
                g = {"assy": doban, "nm": "", "awcnm": str(gv(r, 'mat_work_desc', '') or '').strip(),
                     "mwcnm": str(gv(r, 'work_center', '') or '').strip(), "jado": str(gv(r, 'mat_list', '') or '').strip(),
                     "line": str(gv(r, 'line_no', '') or '').strip(), "lot": 0.0, "matq": 0.0,
                     "days": {}, "done": {}, "colors": {}, "_wos": set()}
                keyed[doban] = g
            g["lot"] += float(gv(r, 'lot_qty', 0) or 0)
            g["matq"] += float(gv(r, 'plan_qty', 0) or 0)
            _wo = str(gv(r, 'work_order', '') or '').strip()
            if _wo: g["_wos"].add((_wo, str(gv(r, 'split_work_order', '') or '').strip()))
            for k, ymd in enumerate(dates):
                g["days"][ymd] = g["days"].get(ymd, 0.0) + float(gv(r, 'plan_qty_%02d' % (k+1), 0) or 0)
        rows = list(keyed.values())
        if item.strip(): rows = [g for g in rows if item.strip() in g["assy"]]
        if part.strip(): rows = [g for g in rows if part.strip() in g["jado"]]
        rows.sort(key=lambda x: x["assy"])
        capped = len(rows) > int(limit); rows = rows[:int(limit)]
        # 품명 채우기(도번=PR_M_ITEM)
        codes = [g["assy"] for g in rows]; nm = {}
        for i in range(0, len(codes), 900):
            ch = codes[i:i+900]; qm = ",".join("?" * len(ch))
            cur.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({qm})", *ch)
            for a, b in cur.fetchall(): nm[str(a).strip()] = b
        # 자도번LIST = f_find_cust_mat_list2 재현(함수 EXECUTE 거부) — SP CTE_BOM 로직으로 도번의 mat_work_code(P2)자재 BOM전개
        jadomap = {}
        if codes:
            from collections import defaultdict as _dd
            jm = _dd(list)
            for i in range(0, len(codes), 500):
                ch = codes[i:i+500]; vals = ",".join("(?)" for _ in ch)
                bomsql = f"""
                WITH SEED(item_code) AS (SELECT item_code FROM (VALUES {vals}) v(item_code)),
                CTE_BOM AS (
                  SELECT CONVERT(int,1) level_no, CONVERT(varchar(50),s.item_code) item_code, CONVERT(varchar(50),s.item_code) mat_code,
                     CONVERT(decimal(18,5),1) cum_use_qty, CONVERT(varchar(20),CASE WHEN c.work_code>'' THEN c.work_code ELSE c.in_cust_code END) mwc
                  FROM SEED s JOIN pr_m_item c ON c.item_code=s.item_code
                  UNION ALL
                  SELECT cb.level_no+1, cb.item_code, CONVERT(varchar(50),b.mat_code),
                     CONVERT(decimal(18,5), cb.cum_use_qty*b.use_qty),
                     CONVERT(varchar(20),CASE WHEN m.work_code>'' THEN m.work_code ELSE m.in_cust_code END)
                  FROM CTE_BOM cb JOIN pr_m_item_bom b ON cb.mat_code=b.item_code JOIN pr_m_item m ON b.mat_code=m.item_code
                  WHERE ISNULL(b.EXCEPT_FLAG,'0')='0' AND cb.level_no<10 )
                SELECT item_code, mat_code, SUM(CONVERT(float,cum_use_qty)) q
                FROM CTE_BOM WHERE mwc=? AND level_no>1 GROUP BY item_code, mat_code
                ORDER BY item_code, mat_code OPTION(MAXRECURSION 0)"""
                cur.execute(bomsql, *ch, wcp)
                for it, mc, q in cur.fetchall():
                    jm[str(it).strip()].append("%s{%d}" % (str(mc).strip(), int(q or 0)))
            jadomap = {k: ",".join(v) for k, v in jm.items()}
        # ★완료/색 = 준비실적처리(키팅, kitting_grid)와 동일 워터폴 이식: 출하(주황)→ASSY재고(노랑)→도번고정(노랑)→중간재고(노랑)→준비재고(녹) 순 계획일 충당.
        # 소스: 라이브 직독(SA_T_ITEM_STOCK·PU_T_READY_STOCK·SA_T_SALE_DTL + 중간재고롤업 kitting캐시). 도번(=ITEM_CODE) 단위 합산.
        assystk = {}; rstock = {}; saled = {}; midstk = {}; fixstk = {}
        try:
            cur.execute("SELECT ITEM_CODE, SUM(STOCK_QTY) FROM SA_T_ITEM_STOCK GROUP BY ITEM_CODE")
            for rr in cur.fetchall(): assystk[str(rr[0]).strip()] = float(rr[1] or 0)
        except Exception: pass
        try:
            cur.execute("SELECT ITEM_CODE, SUM(STOCK_QTY) FROM PU_T_READY_STOCK WHERE CUST_CODE='Z99990' GROUP BY ITEM_CODE")
            for rr in cur.fetchall(): rstock[str(rr[0]).strip()] = float(rr[1] or 0)
        except Exception: pass
        try:  # 출하는 ★계획 WO로 제한(키팅과 동일, 무관 WO 출하 과다합산 방지). 키=(wo,swo,item)
            _pwos = list({wo for g in rows for (wo, sw) in g["_wos"]})
            for i in range(0, len(_pwos), 900):
                ck = _pwos[i:i+900]; ph = ",".join("?" * len(ck))
                cur.execute(f"SELECT WORK_ORDER, ISNULL(SPLIT_WORK_ORDER,''), ITEM_CODE, SUM(SALE_QTY) FROM SA_T_SALE_DTL WHERE FINISH_FLAG='0' AND WORK_ORDER IN ({ph}) GROUP BY WORK_ORDER, ISNULL(SPLIT_WORK_ORDER,''), ITEM_CODE", *ck)
                for rr in cur.fetchall(): saled[(str(rr[0]).strip(), str(rr[1] or '').strip(), str(rr[2]).strip())] = float(rr[3] or 0)
        except Exception: pass
        try:  # 중간공정 자재/생산재고 롤업 = kitting_grid 캐시 재사용, 없으면 자체계산(전역·필터무관, 색tag70용)
            _rc = getattr(kitting_grid, "_rollup_cache", None)
            if not (_rc and _rc.get("mid")):
                cur.execute("IF OBJECT_ID('tempdb..#tms4') IS NOT NULL DROP TABLE #tms4")
                cur.execute("""
                    ;WITH T_SUB_CTE (item_code, upper_item_code, mat_code, stock_qty, pr_stock_qty, fix_pr_stock_qty) AS (
                        SELECT s.mat_code, s.mat_code, s.mat_code, CONVERT(int, ISNULL(SUM(s.stock_qty),0)), CONVERT(int, ISNULL(SUM(s.pr_stock_qty),0)), 0
                          FROM ( SELECT mat_code, 0 stock_qty, STOCK_QTY pr_stock_qty FROM pr_t_mat_stock_wh WITH(NOLOCK)
                                 UNION ALL SELECT a.mat_code,0,a.STOCK_QTY FROM PU_T_SAGUB_STOCK a WITH(NOLOCK) JOIN pr_m_item m WITH(NOLOCK) ON a.MAT_CODE=m.ITEM_CODE WHERE m.SAGUB_STOCK_FLAG='1'
                                 UNION ALL SELECT mat_code, stock_qty, 0 FROM pu_t_mat_stock_wh WITH(NOLOCK) WHERE cust_code='Z99990' AND gagong_proc_code NOT IN ('SA1','SA2','SB1','SB2')
                                 UNION ALL SELECT mat_code, stock_qty, 0 FROM PU_T_STACKER_STOCK WITH(NOLOCK) ) s
                         GROUP BY s.mat_code HAVING SUM(s.stock_qty)<>0 OR SUM(s.pr_stock_qty)<>0
                        UNION ALL
                        SELECT cb.item_code, b.item_code, b.mat_code, 0, 0, CONVERT(int, (CASE WHEN cb.fix_pr_stock_qty<>0 THEN cb.fix_pr_stock_qty ELSE (cb.pr_stock_qty+cb.stock_qty) END) * b.use_qty)
                          FROM T_SUB_CTE cb JOIN pr_m_item_bom b WITH(NOLOCK) ON cb.mat_code=b.item_code WHERE ISNULL(b.except_flag,'0')<>'1'
                    )
                    SELECT item_code, upper_item_code, mat_code, stock_qty, pr_stock_qty, fix_pr_stock_qty INTO #tms4 FROM T_SUB_CTE OPTION(MAXRECURSION 0)""")
                _mid = {}; _fix = {}
                cur.execute("SELECT mat_code, SUM(stock_qty)+SUM(pr_stock_qty) FROM #tms4 GROUP BY mat_code")
                for rr in cur.fetchall(): _mid[str(rr[0]).strip()] = float(rr[1] or 0)
                cur.execute("SELECT upper_item_code, mat_code, SUM(fix_pr_stock_qty) FROM #tms4 GROUP BY upper_item_code, mat_code")
                for rr in cur.fetchall(): _fix[(str(rr[0]).strip(), str(rr[1]).strip())] = float(rr[2] or 0)
                import time as _tm
                kitting_grid._rollup_cache = {"ts": _tm.time(), "mid": _mid, "fix": _fix}
                _rc = kitting_grid._rollup_cache
            midstk = _rc["mid"]
            for (up, it), v in _rc.get("fix", {}).items(): fixstk[it] = fixstk.get(it, 0.0) + float(v or 0)
        except Exception: pass
        def _alloc4(cells, pool, tag, key):
            pool = max(float(pool or 0), 0.0)
            for c in cells:
                if pool <= 0: break
                jan = c["plan"] - c["finish"] - (c["ready"] if key == 'ready' else 0.0)
                if jan <= 0: continue
                if jan > pool: c[key] += pool; pool = 0.0
                else:
                    c[key] += jan; pool -= jan
                    if tag > c["tag"] or c["tag"] == 0: c["tag"] = tag
        _TAGCOLOR = {90: '#fac090', 70: '#ffff00', 50: '#669900'}   # 출하주황·생산노랑·키팅녹
        for g in rows:
            g["nm"] = nm.get(g["assy"], "")
            if jadomap.get(g["assy"]): g["jado"] = jadomap[g["assy"]]
            g["matcnt"] = (g["jado"].count(",") + 1) if g["jado"] else 0
            it = g["assy"]
            cells = [{"ymd": d, "plan": g["days"].get(d, 0.0), "finish": 0.0, "ready": 0.0, "tag": 0} for d in dates]
            _sale = sum(saled.get((wo, sw, it), 0.0) for (wo, sw) in g["_wos"])   # 계획 WO 출하만
            _alloc4(cells, _sale, 90, 'finish')       # 출하 → 주황
            _alloc4(cells, assystk.get(it, 0.0), 70, 'finish')     # ASSY재고(생산완료) → 노랑
            # ★도번고정재고·중간공정재고는 완료에서 제외 — 레거시(w_pr_outside_410_work) 대조결과 in-progress라 완료 아님.
            #   (이 둘을 포함하면 ACJ75119301 완료 500(과다) vs 레거시 200. 제외 시 ASSY재고 200 = 레거시 일치)
            _alloc4(cells, rstock.get(it, 0.0), 50, 'ready')       # 준비재고(키팅) → 녹
            fintot = 0.0
            for c in cells:
                cov = c["finish"] + c["ready"]; g["done"][c["ymd"]] = round(cov, 2); fintot += cov
                g["colors"][c["ymd"]] = _TAGCOLOR.get(c["tag"], '')
            g["finish"] = round(fintot, 2)
            g["plan_qty"] = max(0.0, round(g["matq"] - fintot, 2))   # ★요청수량=자재−완료(미완료 잔량, 레거시 공식)
            g["fin"] = '0'
            g.pop("_wos", None)
        note = (f"⚠ 상위 {limit}건만 표시 — 도번/자도번으로 필터하세요." if capped else "")
        return {"dates": dates, "rows": rows, "cnt": len(rows),
                "plan_sum": sum(r["matq"] for r in rows), "done_sum": sum(r["finish"] for r in rows), "note": note}
    finally:
        cn.close()

# ================= 가공전표이력현황 (w_pr_processing_010) — BOX_NO 마스터-디테일 =================
@app.get("/api/gagong/jeohist")
def gagong_jeohist(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
                   item: str = Query(""), jado: str = Query(""), box_no: str = Query(""), limit: int = Query(500)):
    """가공전표이력. box_no 없으면 좌측 마스터(전표=바코드 목록), 있으면 우측 디테일(공정순서별).
       원천: 최신 PR_T_INDI_CUTTING(전표) + PR_T_INDI_CUTTING_PROC_GAGONG(공정 S_WORK_CODE=가공공정, STD_SIZE=작업표준)."""
    cn = _conn(); cur = cn.cursor()
    try:
        if box_no.strip():   # ---- 우측 디테일: BOX_NO 공정 실적(★레거시 정본 PR_T_PROD_DTL_GAGONG) ----
            # 가공공정=S_WORK_CODE(PR_M_WORK_SINGLE 컷팅/축관…), 파트=GAGONG_PROC_CODE, 설비=MACH_CODE(QA_M_MACHINE), 생산완료=PROD_QTY.
            # 공정횟수(WORK_QTY)·작업표준(STD_SIZE)은 PROD_DTL_GAGONG 부재 → INDI_CUTTING_PROC_GAGONG 보충(있을 때만, 없으면 NULL=담당확인).
            cur.execute("""SELECT p.PROC_SEQ, p.GAGONG_PROC_CODE gpc,
                  COALESCE(NULLIF(pg.GAGONG_PROC_DESC,''), p.GAGONG_PROC_CODE) partnm,
                  ISNULL(p.S_WORK_CODE,'') swork, ISNULL(ws.WORK_DESC,'') sworknm,
                  ISNULL(p.MACH_CODE,'') mach, ISNULL(mm.MACH_DESC,'') machnm,
                  ISNULL(p.PROD_QTY,0) doneq, ic2.WORK_QTY proc_cnt, ic2.STD_SIZE std
                FROM PR_T_PROD_DTL_GAGONG p
                LEFT JOIN PR_M_PROC_GAGONG pg ON pg.GAGONG_PROC_CODE=p.GAGONG_PROC_CODE
                LEFT JOIN PR_M_WORK_SINGLE ws ON ws.S_WORK_CODE=p.S_WORK_CODE
                LEFT JOIN QA_M_MACHINE mm ON mm.MACH_CODE=p.MACH_CODE
                LEFT JOIN PR_T_INDI_CUTTING_PROC_GAGONG ic2 ON ic2.BOX_NO=p.BOX_NO
                     AND ISNULL(ic2.S_WORK_CODE,'')=ISNULL(p.S_WORK_CODE,'') AND ic2.PROC_SEQ=p.PROC_SEQ
                WHERE p.BOX_NO=?
                ORDER BY p.PROC_SEQ, p.PROD_SEQ""", box_no.strip())
            cols = [d[0] for d in cur.description]
            det = []
            for r in cur.fetchall():
                d = dict(zip(cols, r))
                d["doneq"] = float(d["doneq"] or 0)
                d["proc_cnt"] = (float(d["proc_cnt"]) if d["proc_cnt"] is not None else None)   # None=담당확인(원천부재)
                d["std"] = (str(d["std"]).strip() if d["std"] is not None else None)
                det.append(d)
            return {"box_no": box_no.strip(), "detail": det, "cnt": len(det)}
        # ---- 좌측 마스터: 전표(바코드) 목록 (레거시 w_pr_processing_010 전 컬럼) ----
        # ★필터=전표출력기간=PRINT_DATETIME(PLAN_YMD 아님), 자도번=MAT_CODE. 작업처명=work_code→이름 or in_cust→벤더명.
        w = ["CONVERT(date, ic.PRINT_DATETIME) BETWEEN ? AND ?"]
        p = [(from_ymd or "2000-01-01"), (to_ymd or "2099-12-31")]
        if item.strip(): w.append("ic.ITEM_CODE LIKE ?"); p.append(f"%{item.strip()}%")   # 도번=상위도번=ITEM_CODE
        if jado.strip(): w.append("ic.MAT_CODE LIKE ?"); p.append(f"%{jado.strip()}%")     # 자도번=MAT_CODE
        if wc.strip():   # ★작업처(레거시 cust_code 필터) = 자도번 매입처/작업처 코드·명
            w.append("(ma.IN_CUST_CODE LIKE ? OR mac.CUST_DESC LIKE ? OR maw.WORK_DESC LIKE ?)"); p += [f"%{wc.strip()}%"] * 3
        cur.execute(f"""SELECT TOP {int(limit)} ic.BOX_NO,
              ISNULL(ic.ITEM_CODE,'') doban, ISNULL(ic.MAT_CODE,'') jado,
              ISNULL(ma.IN_CUST_CODE,'') wcen, COALESCE(NULLIF(mac.CUST_DESC,''), maw.WORK_DESC, '') wcennm,
              ISNULL(CONVERT(varchar(20), ic.ITEM_DIAM), '') diam, ISNULL(CONVERT(varchar(20), ic.ITEM_THICK), '') thick,
              '' inspdt, ISNULL(ic.CUT_FLAG,'') cutflag, ISNULL(ic.CUT_USER_ID,'') cutuser,
              ISNULL(CONVERT(varchar(19), ic.CUT_DATETIME, 120), '') cutdt,
              ISNULL(ic.ASSY_ITEM_CODE,'') assy,
              COALESCE(NULLIF(aac.CUST_DESC,''), aaw.WORK_DESC, '') assywc,
              COALESCE(NULLIF(iac.CUST_DESC,''), iaw.WORK_DESC, '') dobanwc,
              COALESCE(wh.GAGONG_PROC_DESC, ic.WH_GAGONG_PROC_CODE, '') inwh,
              CONVERT(varchar(19), ic.PRINT_DATETIME, 120) prt, ISNULL(pn.proc_n,0) proc_n
            FROM PR_T_INDI_CUTTING ic
            LEFT JOIN PR_M_ITEM ma ON ma.ITEM_CODE=ic.MAT_CODE
            LEFT JOIN CM_M_CUST mac ON mac.CUST_CODE=ma.IN_CUST_CODE
            LEFT JOIN PR_M_WORK maw ON maw.WORK_CODE=ma.WORK_CODE
            LEFT JOIN PR_M_ITEM ia ON ia.ITEM_CODE=ic.ITEM_CODE
            LEFT JOIN CM_M_CUST iac ON iac.CUST_CODE=ia.IN_CUST_CODE
            LEFT JOIN PR_M_WORK iaw ON iaw.WORK_CODE=ia.WORK_CODE
            LEFT JOIN PR_M_ITEM aa ON aa.ITEM_CODE=ic.ASSY_ITEM_CODE
            LEFT JOIN CM_M_CUST aac ON aac.CUST_CODE=aa.IN_CUST_CODE
            LEFT JOIN PR_M_WORK aaw ON aaw.WORK_CODE=aa.WORK_CODE
            LEFT JOIN PR_M_PROC_GAGONG wh ON wh.GAGONG_PROC_CODE=ic.WH_GAGONG_PROC_CODE
            LEFT JOIN (SELECT BOX_NO, COUNT(*) proc_n FROM PR_T_INDI_CUTTING_PROC_GAGONG GROUP BY BOX_NO) pn ON pn.BOX_NO=ic.BOX_NO
            WHERE {' AND '.join(w)}
            ORDER BY ic.BOX_NO DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

# ================= 가공스크랩관리 (w_qa_raw_input_100/120/125) — QA_T_RAW_ERROR 조회 ∪ nx.scrap_raw(쓰기) =================
# 조회=라이브 QA_T_RAW_ERROR(RO)+nx.scrap_raw 합집합 / 추가·수정·삭제·복사=nx만. 컬럼=SEQ·구분·불량일자·P/No·품명·작업처·작업자·소재항목·불량내용·발생공정·스크랩중량(kg).
_SCRAP_TAG = {"1": "재료", "2": "가공스크랩"}     # 구분(error_tag) — 코드마스터 부재, 데이터 추론(정확명 담당확인)
_SCRAP_SOJE = ["Scrap 일반", "동칩", "동가루", "작업 불량", "고강도", "세팅 불량"]   # 소재항목(실측 고정값)

@app.get("/api/scrap/list")
def scrap_list(from_ymd: str = Query(""), to_ymd: str = Query(""), tag: str = Query(""),
               item: str = Query(""), src: str = Query("")):
    """가공스크랩 목록. 불량기간(error_ymd)·구분(error_tag)·품번 필터. src: L(라이브)/N(nx)/공백=합집합. 하단합계=건수·총중량(kg)."""
    f6 = _d6(from_ymd) if from_ymd else ""; t6 = _d6(to_ymd) if to_ymd else ""
    rows = []
    if src != "N":   # 라이브 QA_T_RAW_ERROR (읽기전용)
        cn = _conn(); cur = cn.cursor()
        try:
            w = ["1=1"]; p = []
            if f6: w.append("e.ERROR_YMD>=?"); p.append(f6)
            if t6: w.append("e.ERROR_YMD<=?"); p.append(t6)
            if tag.strip(): w.append("e.ERROR_TAG=?"); p.append(tag.strip())
            if item.strip(): w.append("e.ITEM_CODE LIKE ?"); p.append(f"%{item.strip()}%")
            cur.execute(f"""SELECT TOP 3000 e.SEQ seq, ISNULL(e.ERROR_TAG,'') tag, e.ERROR_YMD ymd, ISNULL(e.ITEM_CODE,'') item,
                  ISNULL(e.ITEM_DESC,'') item_desc, ISNULL(e.WORK_CODE,'') work, ISNULL(w.WORK_DESC,'') work_desc,
                  ISNULL(e.ERROR_MEMBER_NAME,'') worker, ISNULL(e.ERROR_ITEM,'') soje, ISNULL(e.ERROR_DESC,'') err_desc,
                  ISNULL(e.PROC_CODE,'') pcode, ISNULL(g.GAGONG_PROC_DESC,'') proc_desc, ISNULL(e.LOT_QTY,0) wt, ISNULL(e.INSERT_USER_ID,'') usr
                FROM QA_T_RAW_ERROR e
                LEFT JOIN PR_M_WORK w ON w.WORK_CODE=e.WORK_CODE
                LEFT JOIN PR_M_PROC_GAGONG g ON g.GAGONG_PROC_CODE=e.PROC_CODE
                WHERE {' AND '.join(w)} ORDER BY e.ERROR_YMD DESC, e.SEQ DESC""", *p)
            cols = [c[0] for c in cur.description]
            for r in cur.fetchall():
                d = dict(zip(cols, r)); d["wt"] = float(d["wt"] or 0)
                d["id"] = f"L{int(d['seq'])}"; d["src"] = "L"; rows.append(d)
        finally: cn.close()
    if src != "L":   # nx.scrap_raw (신규 쓰기분)
        nx = _nx(); cur = nx.cursor()
        try:
            w = ["1=1"]; p = []
            if f6: w.append("s.error_ymd>=?"); p.append(f6)
            if t6: w.append("s.error_ymd<=?"); p.append(t6)
            if tag.strip(): w.append("s.error_tag=?"); p.append(tag.strip())
            if item.strip(): w.append("s.item_code LIKE ?"); p.append(f"%{item.strip()}%")
            cur.execute(f"""SELECT s.id seq, ISNULL(s.error_tag,'') tag, s.error_ymd ymd, ISNULL(s.item_code,'') item,
                  ISNULL(s.item_desc,'') item_desc, ISNULL(s.work_code,'') work, ISNULL(w.WORK_DESC,'') work_desc,
                  ISNULL(s.error_member_name,'') worker, ISNULL(s.error_item,'') soje, ISNULL(s.error_desc,'') err_desc,
                  ISNULL(s.proc_code,'') pcode, ISNULL(g.GAGONG_PROC_DESC,'') proc_desc, ISNULL(s.lot_qty,0) wt, ISNULL(s.insert_user_id,'') usr
                FROM nx.scrap_raw s
                LEFT JOIN PARTNER_ERP.dbo.PR_M_WORK w ON w.WORK_CODE=s.work_code
                LEFT JOIN PARTNER_ERP.dbo.PR_M_PROC_GAGONG g ON g.GAGONG_PROC_CODE=s.proc_code
                WHERE {' AND '.join(w)} ORDER BY s.error_ymd DESC, s.id DESC""", *p)
            cols = [c[0] for c in cur.description]
            for r in cur.fetchall():
                d = dict(zip(cols, r)); d["wt"] = float(d["wt"] or 0)
                d["id"] = f"N{int(d['seq'])}"; d["src"] = "N"; rows.append(d)
        finally: nx.close()
    rows.sort(key=lambda r: (str(r["ymd"]), str(r["src"]), r["seq"]), reverse=True)
    tags = [{"code": k, "name": f"{k} {v}"} for k, v in _SCRAP_TAG.items()]
    procs = sorted({(r["pcode"], r["proc_desc"]) for r in rows if r["pcode"]})
    works = sorted({(r["work"], r["work_desc"]) for r in rows if r["work"]})
    workers = sorted({r["worker"] for r in rows if r["worker"]})
    return {"rows": rows, "cnt": len(rows), "total_wt": round(sum(r["wt"] for r in rows), 2),
            "tags": tags, "sojes": _SCRAP_SOJE, "procs": [{"code": c, "name": (n or c)} for c, n in procs],
            "works": [{"code": c, "name": (n or c)} for c, n in works], "workers": workers}

@app.post("/api/scrap/save")
def scrap_save(payload: dict = Body(...)):
    """가공스크랩 추가/수정 → nx.scrap_raw. ★라이브(L*) 수정불가(읽기전용, 신규만 nx). 필수=불량일자·스크랩중량(>0)."""
    p = payload
    ymd = _d6(str(p.get("error_ymd", "")))
    lot = float(p.get("lot_qty") or 0)
    if not ymd:
        raise HTTPException(400, "불량일자는 필수입니다.")
    if lot <= 0:
        raise HTTPException(400, "스크랩중량(kg)은 0보다 커야 합니다.")
    tag = str(p.get("error_tag", "")).strip()[:2]
    item = str(p.get("item_code", "")).strip()[:20]
    idesc = str(p.get("item_desc", "")).strip()[:60]
    work = str(p.get("work_code", "") or "P2").strip()[:4]
    wcust = str(p.get("work_cust_code", "") or "Z99990").strip()[:10]
    proc = str(p.get("proc_code", "") or "P0001").strip()[:10]
    mach = str(p.get("mach_code", "")).strip()[:10]
    worker = str(p.get("error_member_name", "")).strip()[:30]
    soje = str(p.get("error_item", "")).strip()[:100]
    edesc = str(p.get("error_desc", "")).strip()[:300]
    eqty = float(p.get("error_qty") or 0)
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    rid = str(p.get("id", "") or "").strip()
    nx = _nx(); cur = nx.cursor()
    try:
        if rid:
            if not rid.startswith("N"):
                raise HTTPException(409, "라이브(레거시) 자료는 수정할 수 없습니다 — 신규 등록분만 수정 가능(nx).")
            sid = int(rid[1:])
            cur.execute("""UPDATE nx.scrap_raw SET error_ymd=?, error_tag=?, item_code=?, item_desc=?, work_code=?, work_cust_code=?,
                  proc_code=?, mach_code=?, error_member_name=?, error_item=?, error_desc=?, lot_qty=?, error_qty=?,
                  update_user_id=?, update_datetime=getdate() WHERE id=?""",
                  ymd, tag, item, idesc, work, wcust, proc, mach, worker, soje, edesc, lot, eqty, usr, sid)
            if cur.rowcount == 0:
                raise HTTPException(404, f"대상 없음(N{sid})")
            return {"ok": True, "id": f"N{sid}", "mode": "update"}
        cur.execute("""INSERT INTO nx.scrap_raw(error_ymd,error_tag,item_code,item_desc,work_code,work_cust_code,proc_code,mach_code,
              error_member_name,error_item,error_desc,lot_qty,error_qty,insert_user_id)
            OUTPUT INSERTED.id VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            ymd, tag, item, idesc, work, wcust, proc, mach, worker, soje, edesc, lot, eqty, usr)
        nid = int(cur.fetchone()[0])
        return {"ok": True, "id": f"N{nid}", "mode": "insert"}
    finally:
        nx.close()

@app.post("/api/scrap/delete")
def scrap_delete(payload: dict = Body(...)):
    """가공스크랩 삭제(nx만). 라이브(L*) 삭제 불가."""
    ids = [str(x) for x in (payload.get("ids", []) or []) if str(x).strip()]
    live = [x for x in ids if not x.startswith("N")]
    if live:
        raise HTTPException(409, f"라이브(레거시) 자료는 삭제할 수 없습니다(id: {','.join(live)}).")
    nids = [int(x[1:]) for x in ids if x.startswith("N")]
    if not nids:
        return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        ph = ",".join("?" * len(nids))
        cur.execute(f"DELETE FROM nx.scrap_raw WHERE id IN ({ph})", *nids)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()

@app.post("/api/scrap/copy")
def scrap_copy(payload: dict = Body(...)):
    """가공스크랩 복사(원본 L/N → nx 신규 1행). 원본 필드 복제, 신규 채번."""
    rid = str(payload.get("id", "") or "").strip()
    usr = (str(payload.get("user", "")).strip() or "웹사용자")[:30]
    if not rid:
        raise HTTPException(400, "복사할 자료 id 필요")
    src = None
    if rid.startswith("L"):   # 라이브 원본 읽기(RO)
        cn = _conn(); c = cn.cursor()
        try:
            c.execute("""SELECT ERROR_YMD,ISNULL(ERROR_TAG,''),ISNULL(ITEM_CODE,''),ISNULL(ITEM_DESC,''),ISNULL(WORK_CODE,''),
                  ISNULL(WORK_CUST_CODE,''),ISNULL(PROC_CODE,''),ISNULL(MACH_CODE,''),ISNULL(ERROR_MEMBER_NAME,''),
                  ISNULL(ERROR_ITEM,''),ISNULL(ERROR_DESC,''),ISNULL(LOT_QTY,0),ISNULL(ERROR_QTY,0)
                FROM QA_T_RAW_ERROR WHERE SEQ=?""", int(rid[1:]))
            src = c.fetchone()
        finally: cn.close()
    elif rid.startswith("N"):
        nx0 = _nx(); c = nx0.cursor()
        try:
            c.execute("""SELECT error_ymd,ISNULL(error_tag,''),ISNULL(item_code,''),ISNULL(item_desc,''),ISNULL(work_code,''),
                  ISNULL(work_cust_code,''),ISNULL(proc_code,''),ISNULL(mach_code,''),ISNULL(error_member_name,''),
                  ISNULL(error_item,''),ISNULL(error_desc,''),ISNULL(lot_qty,0),ISNULL(error_qty,0)
                FROM nx.scrap_raw WHERE id=?""", int(rid[1:]))
            src = c.fetchone()
        finally: nx0.close()
    if not src:
        raise HTTPException(404, "복사 원본 없음")
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""INSERT INTO nx.scrap_raw(error_ymd,error_tag,item_code,item_desc,work_code,work_cust_code,proc_code,mach_code,
              error_member_name,error_item,error_desc,lot_qty,error_qty,insert_user_id)
            OUTPUT INSERTED.id VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            str(src[0]), str(src[1]), str(src[2]), str(src[3]), str(src[4]), str(src[5]), str(src[6]), str(src[7]),
            str(src[8]), str(src[9]), str(src[10]), float(src[11] or 0), float(src[12] or 0), usr)
        nid = int(cur.fetchone()[0])
        return {"ok": True, "id": f"N{nid}"}
    finally:
        nx.close()

# ================= 가공창고 이동계획 (w_pr_input_580) — 도번×라인, 자도번LIST + 이동필요/완료 =================
@app.get("/api/gagong/move580")
def gagong_move580(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
                   item: str = Query(""), part: str = Query(""), mv: str = Query("전체"), limit: int = Query(2500)):
    """가공창고 이동계획. 계획=PR_T_PLAN_PART_MAT, 이동완료=PU_T_STOCK_MAINT_GAGONG_MOVE(IN_CONFIRM_FLAG='1').
       이동필요수=계획−이동완료. 도번(ASSY)×라인 그룹, 자도번LIST 묶기. (레거시 SP 암호화→라이브 역설계)"""
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["pp.PART_PLAN_QTY>0"]; p = []
        if from_ymd: w.append("pp.PART_PLAN_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("pp.PART_PLAN_YMD<=?"); p.append(_d6(to_ymd))
        if wc.strip():   w.append("ia.WORK_CODE=?"); p.append(wc.strip())
        if item.strip(): w.append("pp.ASSY_ITEM_CODE LIKE ?"); p.append(f"%{item.strip()}%")
        if part.strip(): w.append("pp.MAT_CODE LIKE ?"); p.append(f"%{part.strip()}%")
        cur.execute(f"""SELECT TOP {int(limit) * 60} pp.ASSY_ITEM_CODE assy, ISNULL(ia.ITEM_DESC,'') nm,
              ISNULL(pp.LINE_NO,'') line, COALESCE(cw.WORK_DESC, cc.CUST_DESC, pp.MAT_WORK_CENTER_CODE, '') dest,
              pp.MAT_CODE mat, pp.PART_PLAN_YMD ymd, MIN(ISNULL(pp.PART_OUTPUT_HM,'')) hm,
              SUM(CAST(pp.PART_PLAN_QTY AS float)) q
            FROM PR_T_PLAN_PART_MAT pp
            JOIN PR_M_ITEM ia ON ia.ITEM_CODE=pp.ASSY_ITEM_CODE
            LEFT JOIN PR_M_WORK cw ON cw.WORK_CODE=pp.MAT_WORK_CENTER_CODE
            LEFT JOIN CM_M_CUST cc ON cc.CUST_CODE=pp.MAT_WORK_CENTER_CODE
            WHERE {' AND '.join(w)}
            GROUP BY pp.ASSY_ITEM_CODE, ISNULL(ia.ITEM_DESC,''), pp.LINE_NO,
              COALESCE(cw.WORK_DESC, cc.CUST_DESC, pp.MAT_WORK_CENTER_CODE, ''), pp.MAT_CODE, pp.PART_PLAN_YMD
            ORDER BY assy, line""", *p)
        cols = [d[0] for d in cur.description]; raw = [dict(zip(cols, r)) for r in cur.fetchall()]
        keyed = {}
        for r in raw:
            k = (r["assy"], r["line"], r["dest"])
            g = keyed.get(k)
            if not g:
                g = {"assy": r["assy"], "nm": r["nm"], "line": r["line"], "dest": r["dest"],
                     "days": {}, "mats": {}, "plan_qty": 0.0, "part_ymd": r["ymd"], "hm": r["hm"] or ""}
                keyed[k] = g
            q = float(r["q"] or 0)
            g["days"][r["ymd"]] = g["days"].get(r["ymd"], 0) + q
            g["mats"][r["mat"]] = g["mats"].get(r["mat"], 0) + q
            g["plan_qty"] += q
            if r["ymd"] < g["part_ymd"]: g["part_ymd"] = r["ymd"]; g["hm"] = r["hm"] or ""
        rows = list(keyed.values()); capped = len(keyed) > int(limit); rows = rows[:int(limit)]
        for g in rows:
            g["jado"] = ",".join(f"{m}{{{int(v)}}}" for m, v in sorted(g["mats"].items()))
            g["matcnt"] = len(g["mats"]); g["matlist"] = list(g["mats"].keys()); del g["mats"]
        # 이동완료 = 이동원장(IN_CONFIRM_FLAG='1') by MAT_CODE, 조회범위 date-scope
        matset = list({m for g in rows for m in g["matlist"]}); moved = {}
        d6a = _d6(from_ymd) if from_ymd else None; d6b = _d6(to_ymd) if to_ymd else None
        CH = 1000
        for i in range(0, len(matset), CH):
            ck = matset[i:i + CH]; ph = ",".join("?" * len(ck)); pr = list(ck)
            q = f"SELECT MAT_CODE, SUM(CAST(MAINT_QTY AS float)) FROM PU_T_STOCK_MAINT_GAGONG_MOVE WHERE IN_CONFIRM_FLAG='1' AND MAT_CODE IN ({ph})"
            if d6a: q += " AND MAINT_YMD>=?"; pr.append(d6a)
            if d6b: q += " AND MAINT_YMD<=?"; pr.append(d6b)
            q += " GROUP BY MAT_CODE"
            try:
                cur.execute(q, *pr)
                for rr in cur.fetchall(): moved[rr[0]] = moved.get(rr[0], 0.0) + float(rr[1] or 0)
            except Exception:
                pass
        for g in rows:
            g["moved"] = sum(moved.get(m, 0.0) for m in g["matlist"])
            g["need"] = max(0.0, g["plan_qty"] - g["moved"])
            del g["matlist"]
        m = mv.strip()
        if m == "이동필요": rows = [r for r in rows if r["need"] > 0]
        elif m == "이동완료": rows = [r for r in rows if r["need"] <= 0]
        dates = sorted({ymd for g in rows for ymd in g["days"]})
        rows.sort(key=lambda x: (x["part_ymd"], x["assy"]))
        note = f"⚠ 상위 {limit}건만 표시 — 작업처·도번으로 필터하세요." if capped else ""
        return {"dates": dates, "rows": rows, "cnt": len(rows),
                "plan_sum": sum(r["plan_qty"] for r in rows), "need_sum": sum(r["need"] for r in rows),
                "moved_sum": sum(r["moved"] for r in rows), "note": note}
    finally:
        cn.close()

# ================= 가공바코드실적처리 (w_pr_input_018) — 스캔조회 + nx미러 실적등록/취소 =================
# ★018 원본 미발견: 데이터모델·필드매핑은 라이브확정, 쓰기는 nx미러(라이브 직접커밋 금지). durable=_legacy_analysis/BARCODE_RESULT_018_ANALYSIS.md
def _bc_box(barcode):
    """바코드 → box_no. 접두어(CT/GP/가공) 무시, 뒤 숫자 추출."""
    import re
    m = re.search(r'(\d+)\s*$', str(barcode if barcode is not None else '').strip())
    return int(m.group(1)) if m else None

@app.get("/api/gagong/barcode/scan")
def gagong_bc_scan(barcode: str = Query("")):
    """바코드 스캔 → 전표(BOX_NO) 정보 조회(읽기전용 라이브 + nx미러 기등록분)."""
    box = _bc_box(barcode)
    if not box: return {"ok": False, "msg": "바코드 형식 오류(숫자 없음)"}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("""SELECT ISNULL(ic.ASSY_ITEM_CODE,''), ISNULL(ic.MAT_CODE,''), ISNULL(ic.ITEM_CODE,''),
              ISNULL(ic.PLAN_QTY,0), ISNULL(ic.PROD_QTY,0), ISNULL(ic.PROD_FLAG,'0'),
              ISNULL(ic.WH_GAGONG_PROC_CODE,''), ISNULL(pg.GAGONG_PROC_DESC, ic.WH_GAGONG_PROC_CODE), ISNULL(im.ITEM_DESC,'')
            FROM PR_T_INDI_CUTTING ic
            LEFT JOIN PR_M_PROC_GAGONG pg ON pg.GAGONG_PROC_CODE=ic.WH_GAGONG_PROC_CODE
            LEFT JOIN PR_M_ITEM im ON im.ITEM_CODE=ic.MAT_CODE
            WHERE ic.BOX_NO=?""", box)
        r = cur.fetchone()
        if not r: return {"ok": False, "msg": f"전표(바코드 {box})가 존재하지 않습니다."}
        cur.execute("SELECT ISNULL(SUM(CAST(ISNULL(ERROR_QTY,0) AS int)),0), COUNT(*) FROM QA_T_ERROR WHERE BOX_NO=?", box)
        er = cur.fetchone(); err_qty = int(er[0] or 0); err_cnt = int(er[1] or 0)
        cur.execute("SELECT GOOD_QTY, BAD_QTY, ISNULL(REG_USER,''), CONVERT(varchar(19),REG_DATETIME,120) FROM PARTNER_ERP_TEST3.nx.gagong_barcode_result WHERE BOX_NO=?", box)
        nxr = cur.fetchone()
        return {"ok": True, "box_no": box, "assy": r[0], "mat": r[1], "matnm": r[8], "plan_qty": int(r[3] or 0),
                "prod_qty": int(r[4] or 0) + (int(nxr[0]) if nxr else 0), "prod_flag": r[5], "wh": r[7],
                "err_qty": err_qty, "err_cnt": err_cnt, "already": bool(nxr),
                "reg_good": (int(nxr[0]) if nxr else 0), "reg_bad": (int(nxr[1]) if nxr else 0),
                "reg_user": (nxr[2] if nxr else ""), "reg_dt": (nxr[3] if nxr else "")}
    finally:
        cn.close()

@app.post("/api/gagong/barcode/register")
def gagong_bc_register(payload: dict = Body(...)):
    """처리바코드 2회 일치 검증 → nx미러 실적등록(양품/불량). 라이브 직접커밋 안 함."""
    box = _bc_box(payload.get("box_no") or payload.get("scan1"))
    scan2 = _bc_box(payload.get("scan2"))
    good = int(float(payload.get("good_qty") or 0)); bad = int(float(payload.get("bad_qty") or 0))
    user = str(payload.get("user") or "웹")[:30]
    if not box: return {"ok": False, "msg": "바코드 오류"}
    if scan2 != box: return {"ok": False, "msg": "처리바코드가 스캔바코드와 일치하지 않습니다."}
    if good <= 0 and bad <= 0: return {"ok": False, "msg": "양품/불량 수량을 입력하세요."}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT ISNULL(ASSY_ITEM_CODE,''), ISNULL(MAT_CODE,''), ISNULL(WH_GAGONG_PROC_CODE,''), ISNULL(PROD_FLAG,'0'), ISNULL(PLAN_QTY,0) FROM PR_T_INDI_CUTTING WHERE BOX_NO=?", box)
        lr = cur.fetchone()
        if not lr: return {"ok": False, "msg": f"전표(바코드 {box}) 없음"}
        ymd = _d6(str(payload.get("ymd") or "")) or None
        cur.execute("DELETE FROM PARTNER_ERP_TEST3.nx.gagong_barcode_result WHERE BOX_NO=?", box)
        cur.execute("""INSERT INTO PARTNER_ERP_TEST3.nx.gagong_barcode_result(BOX_NO,ASSY_ITEM_CODE,MAT_CODE,GOOD_QTY,BAD_QTY,REG_YMD,REG_USER,REG_DATETIME,WH_GAGONG_PROC_CODE)
            VALUES(?,?,?,?,?,?,?,getdate(),?)""", box, lr[0], lr[1], good, bad, ymd, user, lr[2])
        cn.commit()
        return {"ok": True, "box_no": box, "good_qty": good, "bad_qty": bad, "msg": "실적 등록 완료(nx 미러)"}
    finally:
        cn.close()

@app.post("/api/gagong/barcode/cancel")
def gagong_bc_cancel(payload: dict = Body(...)):
    """실적 취소 — nx미러 등록분 삭제."""
    box = _bc_box(payload.get("box_no"))
    if not box: return {"ok": False, "msg": "바코드 오류"}
    cn = _conn(); cur = cn.cursor()
    try:
        n = cur.execute("DELETE FROM PARTNER_ERP_TEST3.nx.gagong_barcode_result WHERE BOX_NO=?", box).rowcount
        cn.commit()
        return {"ok": True, "box_no": box, "deleted": n, "msg": ("실적 취소 완료" if n else "등록분 없음")}
    finally:
        cn.close()

# ================= 매입마감처리 (구매/자재, w_pu_sale_010) — 거래처별, 확정입고(매입) =================
def _pur_src(win):
    """확정입고(매입) 원천: 9/S/C/G/H(검사통과) + 수입(_C DIVISION=P). 금액 양수. win=마감기준 조건(mg 참조)."""
    return f"""
    SELECT A.CUST_CODE cc, A.MAT_CODE mat, A.MAINT_COST cost, A.MAINT_YMD ymd, A.MAINT_QTY qty, A.MAINT_AMT amt, A.MAINT_VAT vat
     FROM PU_T_STOCK_MAINT A JOIN MAGAM mg ON A.CUST_CODE=mg.CUST_CODE
     WHERE {win} AND A.MAINT_TAG IN ('9','S','C','G','H')
       AND ((ISNULL(A.INSP_FLAG,'N') IN ('','N')) OR (ISNULL(A.INSP_FLAG,'N') IN ('S','F') AND A.INSP_PROC_YMD >= ''))
    UNION ALL
    SELECT A.CUST_CODE, A.MAT_CODE, A.MAINT_COST, A.MAINT_YMD, A.MAINT_QTY, A.MAINT_AMT, ISNULL(A.TAXPAYERS,0)
     FROM PU_T_STOCK_MAINT_C A JOIN MAGAM mg ON A.CUST_CODE=mg.CUST_CODE
     WHERE {win} AND A.DIVISION='P'"""

@app.get("/api/purmagam/list")
def purmagam_list(ym: str = Query("")):
    """매입마감 거래처별 집계(확정입고, 마감기준) + nx 마감상태·조정합."""
    y = _dig4(ym) or _cur_ym()
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute(f"""{_SALE_MAGAM.format(ym=y)}
          SELECT S.cc cc, MAX(C.CUST_DESC) nm, MAX(C.CUST_TYPE) ct,
            MAX(LTRIM(RTRIM(ISNULL(NULLIF(C.CHARGE_USER_ID,''),ISNULL(C.CHARGE_NAME,''))))) chg,
            SUM(S.qty) qty, SUM(S.amt) amt, SUM(S.vat) vat, COUNT(DISTINCT S.mat) items
          FROM ({_pur_src(_sale_win().format(ym=y))}) S JOIN CM_M_CUST C ON S.cc=C.CUST_CODE
          GROUP BY S.cc HAVING SUM(S.amt)<>0 ORDER BY SUM(S.amt) DESC""")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        cn.close()
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("SELECT cust_code,close_flag,bill_flag FROM nx.pur_close WHERE ym=?", y)
        st = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in nc.fetchall()}
        nc.execute("SELECT cust_code, SUM(delta_amt) FROM nx.pur_adjust WHERE ym=? GROUP BY cust_code", y)
        adj = {r[0]: float(r[1] or 0) for r in nc.fetchall()}
    finally:
        nx.close()
    for r in rows:
        cc = r["cc"]; s = st.get(cc, (0, 0))
        r["qty"] = float(r["qty"] or 0); r["amt"] = float(r["amt"] or 0); r["vat"] = float(r["vat"] or 0); r["items"] = int(r["items"] or 0)
        r["close_flag"] = s[0]; r["bill_flag"] = s[1]
        r["adj_amt"] = adj.get(cc, 0.0); r["final_amt"] = round(r["amt"] + adj.get(cc, 0.0), 2)
    return {"ym": y, "rows": rows}

@app.get("/api/purmagam/detail")
def purmagam_detail(ym: str = Query(""), cc: str = Query(...)):
    """매입 거래처 마감상세: 품목×일자 + 저장된 조정."""
    y = _dig4(ym) or _cur_ym()
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute(f"""{_SALE_MAGAM.format(ym=y)}
          SELECT S.mat mat, MAX(M.ITEM_DESC) nm, MAX(M.ITEM_SPEC) spec, MAX(M.UNIT) unit, S.cost cost,
            CAST(RIGHT(S.ymd,2) AS INT) d, SUM(S.qty) q, SUM(S.amt) amt
          FROM ({_pur_src(_sale_win().format(ym=y))}) S JOIN PR_M_ITEM M ON S.mat=M.ITEM_CODE
          WHERE S.cc=? GROUP BY S.mat, S.cost, CAST(RIGHT(S.ymd,2) AS INT)""", cc)
        raw = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    finally:
        cn.close()
    items = {}; days = set()
    for r in raw:
        mat = str(r["mat"]).strip(); d = int(r["d"] or 0); days.add(d)
        it = items.setdefault(mat, {"mat": mat, "nm": r["nm"], "spec": r["spec"], "unit": r["unit"],
                                    "cost": float(r["cost"] or 0), "qty": 0.0, "amt": 0.0, "_bd": {}})
        qv = float(r["q"] or 0); av = float(r["amt"] or 0); cv = float(r["cost"] or 0)
        it["qty"] += qv; it["amt"] += av
        bd = it["_bd"].setdefault(d, {"d": d, "qty": 0.0, "amt": 0.0, "cost": cv})
        bd["qty"] += qv; bd["amt"] += av; bd["cost"] = cv
    for it in items.values():
        it["byday"] = sorted(it.pop("_bd").values(), key=lambda x: x["d"])
    items_list = sorted(items.values(), key=lambda x: -abs(x["amt"]))
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("""SELECT adj_seq,adj_type,scope,mat_code,target_ymd,old_cost,new_cost,old_qty,new_qty,delta_amt,reason_code,reason_detail
                      FROM nx.pur_adjust WHERE ym=? AND cust_code=? ORDER BY adj_seq""", y, cc)
        adjs = [{"adj_type": r[1], "scope": r[2], "mat_code": r[3], "target_ymd": r[4],
                 "old_cost": (float(r[5]) if r[5] is not None else None), "new_cost": (float(r[6]) if r[6] is not None else None),
                 "old_qty": (float(r[7]) if r[7] is not None else None), "new_qty": (float(r[8]) if r[8] is not None else None),
                 "delta_amt": float(r[9] or 0), "reason_code": r[10], "reason_detail": r[11]} for r in nc.fetchall()]
        nc.execute("SELECT close_flag FROM nx.pur_close WHERE ym=? AND cust_code=?", y, cc)
        cr = nc.fetchone(); closed = int(cr[0]) if cr else 0
    finally:
        nx.close()
    return {"ym": y, "cc": cc, "days": sorted(days), "items": items_list, "adjustments": adjs, "close_flag": closed}

@app.post("/api/purmagam/save")
def purmagam_save(payload: dict = Body(...)):
    """매입 조정 replace-all + 선택시 마감. 가드: 사유필수·이미마감 거부."""
    y = _dig4(payload.get("ym")); cc = str(payload.get("cust_code", "")).strip()
    adjs = payload.get("adjustments", []) or []; do_close = bool(payload.get("close"))
    base = float(payload.get("base_amt", 0) or 0)
    if not y or not cc:
        raise HTTPException(400, "ym·cust_code 필요")
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("SELECT close_flag FROM nx.pur_close WHERE ym=? AND cust_code=?", y, cc)
        cr = nc.fetchone()
        if cr and int(cr[0]) == 1:
            raise HTTPException(409, "이미 마감된 거래처입니다. 마감취소 후 수정하세요.")
        for a in adjs:
            if float(a.get("delta_amt", 0) or 0) != 0 and not (a.get("reason_code") or (a.get("reason_detail") or "").strip()):
                raise HTTPException(400, "사유(코드 또는 세부내역)가 필요한 조정이 있습니다.")
        nc.execute("DELETE FROM nx.pur_adjust WHERE ym=? AND cust_code=?", y, cc)
        for a in adjs:
            nc.execute("""INSERT INTO nx.pur_adjust(ym,cust_code,adj_type,scope,mat_code,target_ymd,old_cost,new_cost,old_qty,new_qty,delta_amt,reason_code,reason_detail,ins_user)
                          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", y, cc, a.get("adj_type"), a.get("scope"), a.get("mat_code"),
                       a.get("target_ymd"), a.get("old_cost"), a.get("new_cost"), a.get("old_qty"), a.get("new_qty"),
                       float(a.get("delta_amt", 0) or 0), a.get("reason_code"), a.get("reason_detail"), "web")
        adj_sum = sum(float(a.get("delta_amt", 0) or 0) for a in adjs)
        nc.execute("""MERGE nx.pur_close AS T USING (SELECT ? ym, ? cc) AS S ON T.ym=S.ym AND T.cust_code=S.cc
                      WHEN MATCHED THEN UPDATE SET base_amt=?, adj_amt=?, final_amt=?, close_flag=?, close_user=?, close_dt=?
                      WHEN NOT MATCHED THEN INSERT(ym,cust_code,base_amt,adj_amt,final_amt,close_flag,close_user,close_dt)
                        VALUES(?,?,?,?,?,?,?,?);""",
                   y, cc, base, adj_sum, base+adj_sum, (1 if do_close else 0), ("web" if do_close else None), (None),
                   y, cc, base, adj_sum, base+adj_sum, (1 if do_close else 0), ("web" if do_close else None), None)
        if do_close:
            nc.execute("UPDATE nx.pur_close SET close_flag=1, close_user='web', close_dt=GETDATE() WHERE ym=? AND cust_code=?", y, cc)
        return {"ok": True, "closed": do_close, "adj_sum": adj_sum}
    finally:
        nx.close()

@app.post("/api/purmagam/reopen")
def purmagam_reopen(payload: dict = Body(...)):
    y = _dig4(payload.get("ym")); cc = str(payload.get("cust_code", "")).strip()
    nx = _nx(); nc = nx.cursor()
    try:
        nc.execute("UPDATE nx.pur_close SET close_flag=0, close_dt=NULL WHERE ym=? AND cust_code=?", y, cc)
        return {"ok": True, "reopened": nc.rowcount}
    finally:
        nx.close()

# ================= 수동발주 (구매/자재, w_pr_input_410 시나리오) =================
@app.get("/api/manorder/vendors")
def manorder_vendors(q: str = Query("")):
    """매입처 검색(그 업체가 납품하는 품목 보유=IN_CUST_CODE). 단일선택 코드 구분."""
    cn = _conn(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT TOP 30 C.CUST_CODE, MAX(C.CUST_DESC) nm, MAX(C.CUST_TYPE) ct, COUNT(M.ITEM_CODE) items
          FROM CM_M_CUST C JOIN PR_M_ITEM M ON M.IN_CUST_CODE=C.CUST_CODE AND ISNULL(M.ITEM_STATUS,'1') IN ('1','2')
          WHERE (C.CUST_CODE LIKE ? OR C.CUST_DESC LIKE ?)
          GROUP BY C.CUST_CODE HAVING COUNT(M.ITEM_CODE)>0
          ORDER BY COUNT(M.ITEM_CODE) DESC""", like, like)
        return {"rows": [{"cc": r[0], "nm": r[1], "ct": r[2], "items": r[3]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.get("/api/manorder/items")
def manorder_items(cc: str = Query(...), ym: str = Query("")):
    """선택 업체 품목별 계획수량·현재고. ★계획 윈도우 = 오늘~+1개월(from6~to6). 좌측 계획수량·우측 일자별 동일 윈도우 → 계=계획수량. ym(YYMM) 지정 시 해당 월 전체."""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT FORMAT(GETDATE(),'yyMMdd'), FORMAT(DATEADD(MONTH,1,GETDATE()),'yyMMdd')")
        from6, to6 = cur.fetchone()
        y = _dig4(ym)
        if y:                       # 특정 월 지정 시 그 달 전체
            from6, to6 = y + "01", y + "99"
        cur.execute("SELECT MAX(STOCK_YYMM) FROM PU_T_MONTH_STOCK_WH")
        smax = cur.fetchone()[0]
        # 계획수량: 부품 접미사 제거한 부모 도번 기준(부모별 1회 집계 후 조인=고속). 기발주=PU_T_PURCHASE_DTL 미입고잔량.
        cur.execute("""
          WITH PLANP AS (
            SELECT LEFT(C_ITEM_CODE, CASE WHEN CHARINDEX('-',C_ITEM_CODE)>0 THEN CHARINDEX('-',C_ITEM_CODE)-1 ELSE LEN(C_ITEM_CODE) END) parent, SUM(PLAN_QTY) pq
            FROM PR_T_PLAN_ITEM_DTL WHERE PLAN_YMD BETWEEN ? AND ?
            GROUP BY LEFT(C_ITEM_CODE, CASE WHEN CHARINDEX('-',C_ITEM_CODE)>0 THEN CHARINDEX('-',C_ITEM_CODE)-1 ELSE LEN(C_ITEM_CODE) END))
          SELECT M.ITEM_CODE ic, M.ITEM_DESC nm, ISNULL(M.ITEM_SPEC,'') spec, ISNULL(M.UNIT,'EA') unit,
            ISNULL(PP.pq,0) plan_qty, ISNULL(S.sq,0) stock_qty, ISNULL(PO.remain,0) po_qty
          FROM PR_M_ITEM M
          LEFT JOIN PLANP PP ON PP.parent = LEFT(M.ITEM_CODE, CASE WHEN CHARINDEX('-',M.ITEM_CODE)>0 THEN CHARINDEX('-',M.ITEM_CODE)-1 ELSE LEN(M.ITEM_CODE) END)
          LEFT JOIN (SELECT MAT_CODE, SUM(STOCK_QTY) sq FROM PU_T_MONTH_STOCK_WH WHERE STOCK_YYMM=? GROUP BY MAT_CODE) S ON S.MAT_CODE=M.ITEM_CODE
          LEFT JOIN (SELECT ITEM_CODE, SUM(PUR_QTY-ISNULL(IN_QTY,0)-ISNULL(CANCEL_QTY,0)) remain
             FROM PU_T_PURCHASE_DTL WHERE CUST_CODE=? AND ISNULL(IN_FINISH_FLAG,'N')<>'Y'
             GROUP BY ITEM_CODE HAVING SUM(PUR_QTY-ISNULL(IN_QTY,0)-ISNULL(CANCEL_QTY,0))>0) PO ON PO.ITEM_CODE=M.ITEM_CODE
          WHERE M.IN_CUST_CODE=? AND ISNULL(M.ITEM_STATUS,'1') IN ('1','2')
          ORDER BY ISNULL(PP.pq,0) DESC, M.ITEM_CODE""", from6, to6, smax, cc, cc)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r["plan_qty"] = float(r["plan_qty"] or 0); r["stock_qty"] = float(r["stock_qty"] or 0)
            r["po_qty"] = float(r["po_qty"] or 0)  # 기발주 = PU_T_PURCHASE_DTL 미입고 발주잔량
        # ★우측 협력사 일자별 계획 = 좌측과 동일 소스(PR_T_PLAN_ITEM_DTL). 부모 도번별 PLAN_YMD 분포 → 일자별 합 = 좌측 계획수량.
        cur.execute("""
          WITH ITM AS (SELECT DISTINCT LEFT(ITEM_CODE, CASE WHEN CHARINDEX('-',ITEM_CODE)>0 THEN CHARINDEX('-',ITEM_CODE)-1 ELSE LEN(ITEM_CODE) END) parent
                       FROM PR_M_ITEM WHERE IN_CUST_CODE=? AND ISNULL(ITEM_STATUS,'1') IN ('1','2'))
          SELECT LEFT(D.C_ITEM_CODE, CASE WHEN CHARINDEX('-',D.C_ITEM_CODE)>0 THEN CHARINDEX('-',D.C_ITEM_CODE)-1 ELSE LEN(D.C_ITEM_CODE) END) parent,
                 D.PLAN_YMD ymd, SUM(D.PLAN_QTY) pq
          FROM PR_T_PLAN_ITEM_DTL D
          JOIN ITM ON ITM.parent = LEFT(D.C_ITEM_CODE, CASE WHEN CHARINDEX('-',D.C_ITEM_CODE)>0 THEN CHARINDEX('-',D.C_ITEM_CODE)-1 ELSE LEN(D.C_ITEM_CODE) END)
          WHERE D.PLAN_YMD BETWEEN ? AND ?
          GROUP BY LEFT(D.C_ITEM_CODE, CASE WHEN CHARINDEX('-',D.C_ITEM_CODE)>0 THEN CHARINDEX('-',D.C_ITEM_CODE)-1 ELSE LEN(D.C_ITEM_CODE) END), D.PLAN_YMD""", cc, from6, to6)
        daily = {}; dset = set()
        for pr, ymd, pq in cur.fetchall():
            ymd = str(ymd).strip(); dset.add(ymd)
            daily.setdefault(str(pr).strip(), {})[ymd] = float(pq or 0)
        dates = sorted(dset)
        def _par(ic):
            ic = str(ic or ""); i = ic.find("-"); return ic[:i] if i > 0 else ic
        for r in rows:
            r["days"] = daily.get(_par(r["ic"]), {})
        cn2 = _conn(); c2 = cn2.cursor()
        try:
            c2.execute("SELECT CUST_DESC FROM CM_M_CUST WHERE CUST_CODE=?", cc)
            rr = c2.fetchone(); nm = rr[0] if rr else cc
        finally:
            cn2.close()
        ymlbl = f"{from6[0:2]}/{from6[2:4]}/{from6[4:6]}~{to6[0:2]}/{to6[2:4]}/{to6[4:6]}"
        return {"cc": cc, "cust_name": nm, "ym": ymlbl, "from_ymd": from6, "to_ymd": to6, "stock_ym": smax, "rows": rows, "dates": dates}
    finally:
        cn.close()

# ================= 생산: 주문UPLOAD(w_pr_plan_010) · 생산계획UPLOAD(w_pr_plan_020) =================
# 소스=LG PU-SCS 2.0 엑셀(Purchase Order / Production Plan Status). 레거시 SP_LGE_RECV_ORDER 매핑을
# 실측검증(품번·단가·워크오더·납기·CR_FLAG 0불일치, 생산계획 WO총량 100%)한 규칙 그대로 적재. 저장=nx(TEST3).
def _d6(s):
    d = ''.join(ch for ch in str(s or '') if ch.isdigit())
    if len(d) == 8: return d[2:8]      # yyyymmdd → yymmdd
    return d[-6:] if len(d) >= 6 else d

def _po_rows(ws, cr):
    """Purchase Order 시트 → recv 튜플. WORK_ORDER=P/S Order '-' 앞부분, ITEM_COST=Unit Price."""
    def y6(v):
        s = str(v); return (s[2:4]+s[5:7]+s[8:10]) if len(s) >= 10 and s[4] == '-' else ''
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[3] is None or str(r[0]) == 'Total Sum': continue
        on = f"{str(r[3]).strip()}-{str(r[4]).strip() if r[4] is not None else ''}"
        ps = str(r[17]).strip() if r[17] else ''
        out.append((on, str(r[1]).strip() if r[1] else '', y6(r[5]), '0000', int(r[9] or 0), int(r[8] or 0),
                    ps[:8], ps, y6(r[33]), cr, str(r[2] or '')[:40], round(float(r[24] or 0), 2)))  # WORK_ORDER=LEFT(P/S,8) 레거시 정본
    return out

def _plan_rows(rows, cr):
    """Production Plan Status 행들 → plan 튜플(WO,일자별). 일별 컬럼(MM/DD) 전개."""
    import re as _re
    def cymd(h):
        m = _re.match(r'(\d\d)/(\d\d)', str(h)); return ('26'+m.group(1)+m.group(2)) if m else None
    hdr = rows[0]; dcol = {i: cymd(hdr[i]) for i in range(len(hdr)) if cymd(hdr[i])}
    agg = {}
    for r in rows[1:]:
        if not r or r[3] is None: continue
        wo = str(r[3]).strip(); line = str(r[1]).strip() if r[1] else ''; sg = str(r[2]).strip() if r[2] else ''
        model = str(r[4]).strip() if r[4] else ''; buyer = str(r[5]).strip() if r[5] else ''
        tot = int(float(r[6] or 0)); rem = int(float(r[7] or 0))
        st = str(r[46]) if len(r) > 46 and r[46] else ''
        sh = (st[11:13]+st[14:16]) if len(st) >= 16 else ''
        if not ('0000' <= sh <= '2359'): sh = '0800'   # 레거시: 무효 Start Time → 0800
        fs = str(r[47]).strip() if len(r) > 47 and r[47] is not None else ''
        ts = str(r[48]).strip() if len(r) > 48 and r[48] is not None else ''
        tool = str(r[49]).strip() if len(r) > 49 and r[49] else ''
        for ci, ymd in dcol.items():
            if ci < len(r) and r[ci] and float(r[ci]) > 0:
                q = int(float(r[ci])); k = (wo, ymd)
                if k in agg:
                    p = list(agg[k]); p[6] += q; agg[k] = tuple(p)
                else:
                    agg[k] = (ymd, wo, model, buyer, line, sg, q, tot, rem, sh, tool[:40], fs[:20], ts[:20], cr)
    return list(agg.values())

def _load_xlsx(b64):
    import base64, io as _io, openpyxl
    raw = base64.b64decode(str(b64).split(',')[-1])
    return openpyxl.load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)

@app.get("/api/order/list")
def order_list(from_ymd: str = Query(""), to_ymd: str = Query(""), need_from: str = Query(""),
               need_to: str = Query(""), done: str = Query("all"), item: str = Query(""),
               wo: str = Query(""), cr: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("r.ORDER_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("r.ORDER_YMD<=?"); p.append(_d6(to_ymd))
        if need_from: w.append("r.NEED_BY_YMD>=?"); p.append(_d6(need_from))
        if need_to:   w.append("r.NEED_BY_YMD<=?"); p.append(_d6(need_to))
        if item.strip(): w.append("r.ITEM_CODE LIKE ?"); p.append(f"%{item.strip()}%")
        if wo.strip():   w.append("(r.WORK_ORDER LIKE ? OR r.PS_ORDER LIKE ?)"); p += [f"%{wo.strip()}%"]*2
        if cr in ('C', 'R'): w.append("r.CR_FLAG=?"); p.append(cr)
        if done == 'done':   w.append("r.REMAIN_QTY<=0")
        elif done == 'undone': w.append("r.REMAIN_QTY>0")
        cur.execute(f"""SELECT TOP 5000 r.ORDER_NO,r.ORDER_YMD,r.ITEM_CODE,ISNULL(i.ITEM_DESC,'') nm,
            r.ORDER_QTY,r.REMAIN_QTY,r.NEED_BY_YMD,r.NEED_BY_HM,r.WORK_ORDER,r.PS_ORDER,r.ITEM_COST,r.CR_FLAG,r.PO_TYPE
          FROM nx.recv_dtl r LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM i ON i.ITEM_CODE=r.ITEM_CODE
          WHERE {' AND '.join(w)} ORDER BY r.ORDER_YMD DESC, r.ORDER_NO""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, row)) for row in cur.fetchall()]
        for r in rows:
            r["ITEM_COST"] = float(r["ITEM_COST"] or 0)
            r["AMT"] = round(r["ITEM_COST"] * (r["ORDER_QTY"] or 0))
        return {"rows": rows, "count": len(rows),
                "sum_qty": sum(r["ORDER_QTY"] or 0 for r in rows),
                "sum_amt": sum(r["AMT"] for r in rows)}
    finally:
        nx.close()

@app.post("/api/order/upload")
def order_upload(payload: dict = Body(...)):
    cr = (str(payload.get("cr", "C")).strip() or "C")[:1]
    try:
        wb = _load_xlsx(payload.get("b64", ""))
    except Exception as e:
        raise HTTPException(400, f"엑셀 파싱 실패: {e}")
    recs = _po_rows(wb[wb.sheetnames[0]], cr); wb.close()
    if not recs:
        return {"ok": True, "inserted": 0, "updated": 0, "total": 0, "cr": cr}
    nx = _nx(); cur = nx.cursor()
    try:
        # 레거시 방식(temp→일괄): 세트기반 upsert로 고속 처리
        cur.execute("IF OBJECT_ID('tempdb..#s') IS NOT NULL DROP TABLE #s")
        cur.execute("""CREATE TABLE #s(ORDER_NO varchar(24),ITEM_CODE varchar(20),NEED_BY_YMD varchar(6),NEED_BY_HM varchar(4),
            ORDER_QTY int,REMAIN_QTY int,WORK_ORDER varchar(20),PS_ORDER varchar(30),ORDER_YMD varchar(6),
            CR_FLAG varchar(1),PO_TYPE varchar(40),ITEM_COST decimal(18,2))""")
        cur.fast_executemany = True
        cur.executemany("INSERT INTO #s VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", recs)
        upd = cur.execute("SELECT COUNT(*) FROM nx.recv_dtl r JOIN #s s ON r.ORDER_NO=s.ORDER_NO").fetchone()[0]
        cur.execute("DELETE r FROM nx.recv_dtl r JOIN #s s ON r.ORDER_NO=s.ORDER_NO")
        cur.execute("""INSERT INTO nx.recv_dtl(ORDER_NO,ITEM_CODE,NEED_BY_YMD,NEED_BY_HM,ORDER_QTY,REMAIN_QTY,
            WORK_ORDER,PS_ORDER,ORDER_YMD,CR_FLAG,PO_TYPE,ITEM_COST,UPLOAD_DT)
            SELECT ORDER_NO,ITEM_CODE,NEED_BY_YMD,NEED_BY_HM,ORDER_QTY,REMAIN_QTY,WORK_ORDER,PS_ORDER,ORDER_YMD,
                   CR_FLAG,PO_TYPE,ITEM_COST,getdate() FROM #s""")
        return {"ok": True, "inserted": len(recs) - upd, "updated": upd, "total": len(recs), "cr": cr}
    finally:
        nx.close()

@app.get("/api/plan/list")
def plan_list(from_ymd: str = Query(""), to_ymd: str = Query(""), line: str = Query(""),
              sched: str = Query(""), wo: str = Query(""), model: str = Query(""), cr: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("PLAN_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("PLAN_YMD<=?"); p.append(_d6(to_ymd))
        if line.strip():  w.append("LINE_NO=?"); p.append(line.strip())
        if sched.strip(): w.append("SCHED_GROUP=?"); p.append(sched.strip())
        if wo.strip():    w.append("WORK_ORDER LIKE ?"); p.append(f"%{wo.strip()}%")
        if model.strip(): w.append("MODEL_NO LIKE ?"); p.append(f"%{model.strip()}%")
        if cr in ('C', 'R'): w.append("CR_FLAG=?"); p.append(cr)
        cur.execute(f"""SELECT PLAN_YMD,WORK_ORDER,MODEL_NO,BUYER_MODEL,LINE_NO,SCHED_GROUP,PLAN_QTY,
            TOTAL_QTY,REMAIN_QTY,START_HM,TOOL,FROM_SEQ,TO_SEQ,CR_FLAG
          FROM nx.plan_dtl WHERE {' AND '.join(w)}""", *p)
        cols = [d[0] for d in cur.description]
        raw = [dict(zip(cols, row)) for row in cur.fetchall()]
        dates = sorted({r["PLAN_YMD"] for r in raw})
        wos = {}
        for r in raw:
            k = r["WORK_ORDER"]
            g = wos.get(k)
            if not g:
                g = {"wo": k, "model": r["MODEL_NO"], "buyer": r["BUYER_MODEL"], "line": r["LINE_NO"],
                     "sched": r["SCHED_GROUP"], "total": r["TOTAL_QTY"], "remain": r["REMAIN_QTY"],
                     "tool": r["TOOL"], "cr": r["CR_FLAG"], "days": {}}
                wos[k] = g
            g["days"][r["PLAN_YMD"]] = (g["days"].get(r["PLAN_YMD"], 0) + (r["PLAN_QTY"] or 0))
        rows = sorted(wos.values(), key=lambda x: (x["line"] or "", x["wo"]))
        return {"dates": dates, "rows": rows, "wo_count": len(rows),
                "sum_qty": sum(r["PLAN_QTY"] or 0 for r in raw)}
    finally:
        nx.close()

@app.post("/api/plan/upload")
def plan_upload(payload: dict = Body(...)):
    cr = (str(payload.get("cr", "C")).strip() or "C")[:1]
    try:
        wb = _load_xlsx(payload.get("b64", ""))
    except Exception as e:
        raise HTTPException(400, f"엑셀 파싱 실패: {e}")
    ws = wb[wb.sheetnames[0]]; rows = list(ws.iter_rows(values_only=True)); wb.close()
    recs = _plan_rows(rows, cr)
    if not recs:
        return {"ok": True, "inserted": 0, "updated": 0, "total": 0, "cr": cr}
    nx = _nx(); cur = nx.cursor()
    try:  # recs t=(PLAN_YMD,WORK_ORDER,MODEL_NO,BUYER_MODEL,LINE_NO,SCHED_GROUP,PLAN_QTY,TOTAL_QTY,REMAIN_QTY,START_HM,TOOL,FROM_SEQ,TO_SEQ,CR_FLAG)
        cur.execute("IF OBJECT_ID('tempdb..#p') IS NOT NULL DROP TABLE #p")
        cur.execute("""CREATE TABLE #p(PLAN_YMD varchar(6),WORK_ORDER varchar(20),MODEL_NO varchar(30),BUYER_MODEL varchar(30),
            LINE_NO varchar(10),SCHED_GROUP varchar(6),PLAN_QTY int,TOTAL_QTY int,REMAIN_QTY int,START_HM varchar(4),
            TOOL varchar(40),FROM_SEQ varchar(20),TO_SEQ varchar(20),CR_FLAG varchar(1))""")
        cur.fast_executemany = True
        cur.executemany("INSERT INTO #p VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)", recs)
        # 레거시 STEP0: "cr별 삭제 후 재적재"(full replace). 업로드 파일 = 해당 CR의 완전한 현재 계획.
        # ★해당 CR 전체 삭제(과거일자 포함) → 재적재. 계획일자 이동/재업로드 시 stale행 누적(2배)·
        #   과거일자 잔재(compose_mat 부풀림) 방지. 과거 이력은 별도 _daily 백업 대상(현 미구현).
        fmin = min(r[0] for r in recs)
        upd = cur.execute("SELECT COUNT(*) FROM nx.plan_dtl WHERE CR_FLAG=?", cr).fetchone()[0]
        cur.execute("DELETE FROM nx.plan_dtl WHERE CR_FLAG=?", cr)
        cur.execute("""INSERT INTO nx.plan_dtl(PLAN_YMD,WORK_ORDER,MODEL_NO,BUYER_MODEL,LINE_NO,SCHED_GROUP,PLAN_QTY,
            TOTAL_QTY,REMAIN_QTY,START_HM,TOOL,FROM_SEQ,TO_SEQ,CR_FLAG,UPLOAD_DT)
            SELECT PLAN_YMD,WORK_ORDER,MODEL_NO,BUYER_MODEL,LINE_NO,SCHED_GROUP,PLAN_QTY,TOTAL_QTY,REMAIN_QTY,START_HM,
                   TOOL,FROM_SEQ,TO_SEQ,CR_FLAG,getdate() FROM #p""")
        # full-replace(cr별): 기존 upd행 삭제 후 recs행 재적재
        return {"ok": True, "inserted": len(recs), "replaced": upd, "total": len(recs), "cr": cr, "from_ymd": fmin}
    finally:
        nx.close()

# ================= 협력사 계획 편성 엔진 (생산계획업로드 → 자도번 라우팅) =================
# 레거시 SP_PR_CREATE_PLAN_협력사계획_생성 정렬(98% 재현): PR_M_ITEM_BOM(except≠1) + 가공처(work_code‖in_cust)
#  + charindex 중복제거(조상에 같은 가공처면 컷) + 조달프로파일 오버레이(유효기간·배분).
def _compose_maps():
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT ITEM_CODE, LTRIM(RTRIM(ISNULL(WORK_CODE,''))), ISNULL(IN_CUST_CODE,'') FROM PR_M_ITEM")
        WCEN = {}
        for ic, wc, inc in cur.fetchall():
            WCEN[ic] = wc if wc > '' else str(inc).strip()
        cur.execute("""SELECT ITEM_CODE, MAT_CODE, USE_QTY FROM PR_M_ITEM_BOM
            WHERE ISNULL(EXCEPT_FLAG,'0')<>'1' AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101'""")
        CH = {}
        for p, c, q in cur.fetchall():
            CH.setdefault(p, []).append((c, float(q or 0)))
        return WCEN, CH
    finally:
        cn.close()

def _compose_assy(assy, WCEN, CH, memo):
    """assy → {(part, work_center): cum_qty}. 레거시 앵커(ASSY 자신=level0 파트) + charindex 중복제거."""
    if assy in memo:
        return memo[assy]
    out = {}
    root_wc = WCEN.get(assy, '')
    out[(assy, root_wc)] = 1.0   # ★앵커멤버: ASSY 자신을 파트로(레거시 bom_level 0)
    def rec(item, cq, path):
        for c, q in CH.get(item, []):
            wc = WCEN.get(c, ''); nq = cq * q
            if wc not in path:
                k = (c, wc); out[k] = out.get(k, 0.0) + nq
            rec(c, nq, path | {wc})
    rec(assy, 1.0, {root_wc})
    memo[assy] = out
    return out

@app.post("/api/plan/compose")
def plan_compose(payload: dict = Body(...)):
    """nx.plan_dtl 전량 편성 → nx.plan_part. 조달프로파일(현행,유효기간,배분) 오버레이."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("IF OBJECT_ID('nx.plan_part') IS NULL CREATE TABLE nx.plan_part(PLAN_YMD varchar(6),WORK_ORDER varchar(20),ASSY_ITEM_CODE varchar(20),PART_CODE varchar(20),WORK_CENTER varchar(20),SUPPLY_GUBUN varchar(10),PROFILE_ID int,USE_QTY decimal(18,6),PLAN_QTY decimal(18,3),COMPOSE_DT datetime DEFAULT getdate())")
        # ★STEP M 신규모델생성 (레거시 ue_make_model 재현): 주문(nx.recv_dtl) ⋈ 계획(nx.plan_dtl)을 제번(WORK_ORDER)으로 조인
        #   → model→AJR(ASSY) 매핑을 nx.model_bom에 동적생성. use_qty=CEILING(발주/LOT), 3중제외(모델BOM·중복·EXCEPT).
        cur.execute("DELETE FROM nx.model_bom WHERE REMARKS='신규모델자동'")
        cur.execute("""INSERT INTO nx.model_bom(MODEL_NO,C_ITEM_CODE,USE_QTY,APPLY_FROM,APPLY_TO,REMARKS,INS_DT)
            SELECT p.model_no, r.item_code,
               MAX(CASE WHEN r.order_qty<p.lot THEN 1 ELSE CEILING(CAST(r.order_qty AS float)/NULLIF(p.lot,0)) END),
               MIN(p.plan_ymd),'999999','신규모델자동',getdate()
            FROM (SELECT RTRIM(MODEL_NO) model_no,WORK_ORDER,MAX(TOTAL_QTY) lot,MIN(PLAN_YMD) plan_ymd
                  FROM nx.plan_dtl WHERE ISNULL(MODEL_NO,'')>'' GROUP BY RTRIM(MODEL_NO),WORK_ORDER) p
            JOIN (SELECT RTRIM(ITEM_CODE) item_code,WORK_ORDER,SUM(ORDER_QTY) order_qty
                  FROM nx.recv_dtl WHERE ISNULL(ITEM_CODE,'')>'' GROUP BY RTRIM(ITEM_CODE),WORK_ORDER) r
              ON p.WORK_ORDER=r.WORK_ORDER
            WHERE NOT EXISTS(SELECT 1 FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM b WHERE b.MODEL_NO=p.model_no AND b.C_ITEM_CODE=r.item_code)
              AND NOT EXISTS(SELECT 1 FROM nx.model_bom m WHERE m.MODEL_NO=p.model_no AND m.C_ITEM_CODE=r.item_code)
              AND NOT EXISTS(SELECT 1 FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM_EXCEPT e WHERE e.MODEL_NO=p.model_no AND e.C_ITEM_CODE=r.item_code)
            GROUP BY p.model_no,r.item_code""")
        # ASSY 매핑: ①모델BOM(MODEL→[ASSY×use_qty], 유효일자 버전) ②주문 fallback. + 제외조건.
        mbom = {}
        cur.execute("SELECT MODEL_NO, C_ITEM_CODE, USE_QTY, MAKE_YMD, TO_APPLY_YMD FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM")
        for m, ci, uq, my, ty in cur.fetchall():
            mbom.setdefault(str(m).strip(), []).append((str(ci).strip(), float(uq or 1), str(my or '').strip(), str(ty or '').strip()))
        try:  # 우리 신규등록 모델BOM(nx) union
            cur.execute("SELECT MODEL_NO, C_ITEM_CODE, USE_QTY, APPLY_FROM, APPLY_TO FROM nx.model_bom")
            for m, ci, uq, my, ty in cur.fetchall():
                mbom.setdefault(str(m).strip(), []).append((str(ci).strip(), float(uq or 1), str(my or '').strip(), str(ty or '').strip()))
        except Exception:
            pass
        mbexcept = set()  # 모델BOM제외조건
        cur.execute("SELECT MODEL_NO, C_ITEM_CODE FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM_EXCEPT")
        for m, ci in cur.fetchall(): mbexcept.add((str(m).strip(), str(ci).strip()))
        recvmap = {}
        cur.execute("SELECT DISTINCT WORK_ORDER, ITEM_CODE FROM PARTNER_ERP.dbo.sa_t_recv_dtl WHERE WORK_ORDER>''")
        for wo, ic in cur.fetchall(): recvmap.setdefault(str(wo).strip(), set()).add(str(ic).strip())
        # 조달프로파일(현행 활성): item→[(sg,vc,alloc,af,at)]
        cur.execute("""SELECT item_code, supply_gubun, ISNULL(vendor_code,''), ISNULL(alloc_ratio,100),
            CONVERT(varchar(10),apply_from,23), CONVERT(varchar(10),apply_to,23), profile_id
            FROM nx.sourcing_profile WHERE is_active=1 AND is_internal=0""")
        prof = {}
        for ic, sg, vc, al, af, at, pid in cur.fetchall():
            prof.setdefault(ic, []).append((sg, str(vc).strip(), float(al or 100), af, at, pid))
        WCEN, CH = _compose_maps()
        cur.execute("SELECT WORK_ORDER, PLAN_YMD, PLAN_QTY, MODEL_NO FROM nx.plan_dtl WHERE PLAN_QTY>0")
        plan = cur.fetchall()
        memo = {}; rows = []; mapped = unmapped = 0
        for wo, ymd, q, model in plan:
            wos = str(wo).strip(); mk = str(model).strip()
            cand = mbom.get(mk)                            # [(assy, use_qty, make_ymd, to_apply_ymd)]
            assys = None
            if cand:                                       # 유효일자(plan날짜)만. ★MODEL_BOM_EXCEPT는 STEP M(신규모델생성)전용 —
                assys = [(a, mq) for (a, mq, my, ty) in cand   #   전개에 적용 금지(레거시 STEP5 품목별생성은 EXCEPT 미적용).
                         if (not my or my <= ymd) and (not ty or ty >= ymd)]  #   대원산업 외주완성 서포터(EXCEPT=1) 드롭 방지.
            if not assys:
                rc = recvmap.get(wos)
                assys = [(a, 1.0) for a in rc] if rc else None
            if not assys: unmapped += 1; continue
            mapped += 1
            d = f"20{ymd[:2]}-{ymd[2:4]}-{ymd[4:6]}"
            for assy, mq in assys:
                base = float(q) * mq
                for (part, wc), soyo in _compose_assy(assy, WCEN, CH, memo).items():
                    ps = [p for p in prof.get(part, []) if (not p[3] or p[3] <= d) and (not p[4] or p[4] >= d)]
                    if ps:
                        for sg, vc, al, af, at, pid in ps:
                            who = vc if (sg in ('유상사급', '매입') and vc) else wc
                            rows.append((ymd, wos, assy, part, who, sg, pid, soyo, base * soyo * al / 100.0))
                    else:
                        rows.append((ymd, wos, assy, part, wc, None, None, soyo, base * soyo))
        cur.execute("DELETE FROM nx.plan_part")
        cur.fast_executemany = True
        cur.executemany("INSERT INTO nx.plan_part(PLAN_YMD,WORK_ORDER,ASSY_ITEM_CODE,PART_CODE,WORK_CENTER,SUPPLY_GUBUN,PROFILE_ID,USE_QTY,PLAN_QTY) VALUES(?,?,?,?,?,?,?,?,?)", rows)
        return {"ok": True, "mapped": mapped, "unmapped": unmapped, "part_lines": len(rows)}
    finally:
        nx.close()

# ================= ★자재소요 정본 엔진 (레거시 STEP5→6→7 충실이식, 수량100%·총량1.00000x 검증) =================
# STEP5 nx.plan_item_dtl(LOT합산·모델→ASSY 유효일자, EXCEPT미적용) → STEP6 nx.plan_part_dtl(10레벨BOM+가공공정 공정전이)
#  → STEP7 nx.plan_part_mat(사급중단 NOT EXISTS PART_DTL + 최하위집계 + charindex중복 + 용접봉sgroup910 제외=공정처리).
# 검증: 설계2건(용접봉·체결SUB이중계상)제외시 웹 vs 레거시 PR_T_PLAN_PART_MAT 수량완전일치100%. [[newerp-plan-soyo-verify]]
_P = "PARTNER_ERP.dbo."

@app.post("/api/plan/compose_mat")
def plan_compose_mat(payload: dict = Body(...)):
    nx = _nx(); cur = nx.cursor()
    try:
        # ── STEP M 신규모델생성(주문⋈계획 제번조인, use=CEILING(order/lot), 3중제외) ──
        cur.execute("DELETE FROM nx.model_bom WHERE REMARKS='신규모델자동'")
        cur.execute("""INSERT INTO nx.model_bom(MODEL_NO,C_ITEM_CODE,USE_QTY,APPLY_FROM,APPLY_TO,REMARKS,INS_DT)
            SELECT p.model_no, r.item_code,
               MAX(CASE WHEN r.order_qty<p.lot THEN 1 ELSE CEILING(CAST(r.order_qty AS float)/NULLIF(p.lot,0)) END),
               MIN(p.plan_ymd),'999999','신규모델자동',getdate()
            FROM (SELECT RTRIM(MODEL_NO) model_no,WORK_ORDER,MAX(TOTAL_QTY) lot,MIN(PLAN_YMD) plan_ymd
                  FROM nx.plan_dtl WHERE ISNULL(MODEL_NO,'')>'' GROUP BY RTRIM(MODEL_NO),WORK_ORDER) p
            JOIN (SELECT RTRIM(ITEM_CODE) item_code,WORK_ORDER,SUM(ORDER_QTY) order_qty
                  FROM nx.recv_dtl WHERE ISNULL(ITEM_CODE,'')>'' GROUP BY RTRIM(ITEM_CODE),WORK_ORDER) r ON p.WORK_ORDER=r.WORK_ORDER
            WHERE NOT EXISTS(SELECT 1 FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM b WHERE b.MODEL_NO=p.model_no AND b.C_ITEM_CODE=r.item_code)
              AND NOT EXISTS(SELECT 1 FROM nx.model_bom m WHERE m.MODEL_NO=p.model_no AND m.C_ITEM_CODE=r.item_code)
              AND NOT EXISTS(SELECT 1 FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM_EXCEPT e WHERE e.MODEL_NO=p.model_no AND e.C_ITEM_CODE=r.item_code)
            GROUP BY p.model_no,r.item_code""")
        # ── STEP5 nx.plan_item_dtl: (제번,모델) LOT합산 → 모델→ASSY 전개(유효일자, ★EXCEPT미적용) ──
        from collections import defaultdict as _dd
        mbom = _dd(list)
        cur.execute("SELECT MODEL_NO,C_ITEM_CODE,USE_QTY,MAKE_YMD,TO_APPLY_YMD FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM")
        for m, ci, uq, my, ty in cur.fetchall(): mbom[str(m).strip()].append((str(ci).strip(), float(uq or 1), str(my or '').strip(), str(ty or '').strip()))
        cur.execute("SELECT MODEL_NO,C_ITEM_CODE,USE_QTY,APPLY_FROM,APPLY_TO FROM nx.model_bom")
        for m, ci, uq, my, ty in cur.fetchall(): mbom[str(m).strip()].append((str(ci).strip(), float(uq or 1), str(my or '').strip(), str(ty or '').strip()))
        recvmap = _dd(set)
        cur.execute("SELECT DISTINCT WORK_ORDER,ITEM_CODE FROM PARTNER_ERP.dbo.sa_t_recv_dtl WHERE WORK_ORDER>''")
        for wo, ic in cur.fetchall(): recvmap[str(wo).strip()].add(str(ic).strip())
        prate = {}
        cur.execute("SELECT ITEM_CODE, ISNULL(PROD_RATE,100) FROM PARTNER_ERP.dbo.PR_M_ITEM")
        for ic, pr in cur.fetchall(): prate[str(ic).strip()] = float(pr or 100)
        cur.execute("""IF OBJECT_ID('nx.plan_item_dtl') IS NULL CREATE TABLE nx.plan_item_dtl(
            PLAN_YMD varchar(6),WORK_ORDER varchar(20),SPLIT_WORK_ORDER varchar(30),C_ITEM_CODE varchar(20),
            USE_QTY decimal(18,5),LOT_QTY int,PLAN_QTY int,ORG_PLAN_YMD varchar(6),LINE_NO varchar(6),OUTPUT_HM varchar(4),PROD_RATE numeric(9,2))""")
        cur.execute("DELETE FROM nx.plan_item_dtl")
        cur.execute("SELECT WORK_ORDER,MODEL_NO,SUM(CAST(PLAN_QTY AS int)),MIN(PLAN_YMD) FROM nx.plan_dtl WHERE PLAN_QTY>0 GROUP BY WORK_ORDER,MODEL_NO")
        irows = []; lot = _dd(int)
        for wo, model, pq, ymd in cur.fetchall():
            wos = str(wo).strip(); mk = str(model).strip(); pq = int(pq or 0); ymd = str(ymd).strip()
            cand = mbom.get(mk); assys = None
            if cand:
                best = {}
                for a, mq, my, ty in cand:
                    if (not my or my <= ymd) and (not ty or ty >= ymd):
                        if a not in best or my > best[a][1]: best[a] = (mq, my)
                assys = [(a, best[a][0]) for a in best]
            if not assys:
                rc = recvmap.get(wos); assys = [(a, 1.0) for a in rc] if rc else None
            if not assys: continue
            for a, mq in assys:
                irows.append([ymd, wos, wos, a, mq, 0, pq, ymd, '', '0800', prate.get(a, 100)]); lot[wos] = max(lot[wos], pq)
        for rr in irows: rr[5] = lot[rr[1]]
        cur.fast_executemany = True
        cur.executemany("INSERT INTO nx.plan_item_dtl(PLAN_YMD,WORK_ORDER,SPLIT_WORK_ORDER,C_ITEM_CODE,USE_QTY,LOT_QTY,PLAN_QTY,ORG_PLAN_YMD,LINE_NO,OUTPUT_HM,PROD_RATE) VALUES(?,?,?,?,?,?,?,?,?,?,?)", irows)
        # ── STEP6 nx.plan_part_dtl: 10레벨 BOM전개 → 가공공정 → 공정전이지점 ──
        _step6_sql(cur)
        # ── STEP7 nx.plan_part_mat: 사급중단+최하위집계+charindex+용접봉(sgroup910)제외 ──
        _step7_sql(cur)
        # ── 조달 프로파일 오버레이 → nx.plan_mat_source (공급방식·공급처·수량) ──
        #   ①활성 프로파일 있으면 supply_gubun·vendor·배분(alloc) ②없으면 BOM기본(MAKE_TYPE→매입/사급/외주/자체 + IN_CUST vendor).
        cur.execute("""IF OBJECT_ID('nx.plan_mat_source') IS NULL CREATE TABLE nx.plan_mat_source(
            WORK_ORDER varchar(20),MAT_CODE varchar(20),SUPPLY_GUBUN varchar(20),VENDOR_CODE varchar(20),
            QTY decimal(18,3),SOURCE varchar(10),COMPOSE_DT datetime DEFAULT getdate())""")
        cur.execute("DELETE FROM nx.plan_mat_source")
        MKF = {}; INCF = {}
        cur.execute("SELECT ITEM_CODE, ISNULL(MAKE_TYPE,''), ISNULL(IN_CUST_CODE,'') FROM PARTNER_ERP.dbo.PR_M_ITEM")
        for ic, mkt, inc in cur.fetchall(): ic = str(ic).strip(); MKF[ic] = str(mkt).strip(); INCF[ic] = str(inc).strip()
        PRF = {}
        cur.execute("SELECT item_code, supply_gubun, ISNULL(vendor_code,''), ISNULL(alloc_ratio,100) FROM nx.sourcing_profile WHERE is_active=1 AND is_internal=0")
        for ic, sg, v, al in cur.fetchall(): PRF.setdefault(str(ic).strip(), []).append((str(sg).strip(), str(v).strip(), float(al or 100)))
        _MKMAP = {'1': '자체', '2': '외주가공', '3': '매입', '4': '유상사급', '5': '외주완성'}  # '자체'=프로파일 라벨과 통일
        cur.execute("SELECT work_order, mat_code, SUM(CAST(part_plan_qty AS float)) FROM nx.plan_part_mat GROUP BY work_order, mat_code")
        srows = []
        for wo, mat, qty in cur.fetchall():
            wo = str(wo).strip(); mat = str(mat).strip(); qty = float(qty or 0)
            ps = PRF.get(mat)
            if ps:
                for sg, v, al in ps: srows.append((wo, mat, sg, v, qty * al / 100.0, '프로파일'))
            else:
                srows.append((wo, mat, _MKMAP.get(MKF.get(mat, ''), '미지정'), INCF.get(mat, ''), qty, 'BOM기본'))
        cur.fast_executemany = True
        cur.executemany("INSERT INTO nx.plan_mat_source(WORK_ORDER,MAT_CODE,SUPPLY_GUBUN,VENDOR_CODE,QTY,SOURCE) VALUES(?,?,?,?,?,?)", srows)
        cur.execute("SELECT COUNT(*), COUNT(DISTINCT work_order) FROM nx.plan_part_mat")
        n, woc = cur.fetchone()
        return {"ok": True, "item_lines": len(irows), "mat_lines": int(n), "mat_work_orders": int(woc), "sourcing_lines": len(srows)}
    finally:
        nx.close()

@app.get("/api/plan/sourcing")
def plan_sourcing(mode: str = Query("gubun"), gubun: str = Query(""), vendor: str = Query(""),
                  mat: str = Query(""), wo: str = Query("")):
    """조달 소요 조회. mode=gubun(공급방식별 집계)·vendor(공급처별)·detail(제번×자재 명세)."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("IF OBJECT_ID('nx.plan_mat_source') IS NULL SELECT 1 WHERE 1=0")
        w = ["1=1"]; p = []
        if gubun.strip(): w.append("s.SUPPLY_GUBUN=?"); p.append(gubun.strip())
        if vendor.strip(): w.append("s.VENDOR_CODE=?"); p.append(vendor.strip())
        if mat.strip(): w.append("s.MAT_CODE LIKE ?"); p.append(f"%{mat.strip()}%")
        if wo.strip(): w.append("s.WORK_ORDER LIKE ?"); p.append(f"%{wo.strip()}%")
        wh = " AND ".join(w)
        try:
            if mode == "vendor":
                cur.execute(f"""SELECT s.SUPPLY_GUBUN, s.VENDOR_CODE, ISNULL(cu.CUST_DESC,'') vname,
                    COUNT(DISTINCT s.MAT_CODE) mats, SUM(s.QTY) qty FROM nx.plan_mat_source s
                    LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST cu ON s.VENDOR_CODE COLLATE DATABASE_DEFAULT=cu.CUST_CODE COLLATE DATABASE_DEFAULT
                    WHERE {wh} GROUP BY s.SUPPLY_GUBUN, s.VENDOR_CODE, cu.CUST_DESC ORDER BY SUM(s.QTY) DESC""", p)
            elif mode == "detail":
                cur.execute(f"""SELECT TOP 2000 s.WORK_ORDER, s.MAT_CODE, ISNULL(it.ITEM_DESC,'') mname, s.SUPPLY_GUBUN,
                    s.VENDOR_CODE, ISNULL(cu.CUST_DESC,'') vname, s.QTY, s.SOURCE FROM nx.plan_mat_source s
                    LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM it ON s.MAT_CODE COLLATE DATABASE_DEFAULT=it.ITEM_CODE COLLATE DATABASE_DEFAULT
                    LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST cu ON s.VENDOR_CODE COLLATE DATABASE_DEFAULT=cu.CUST_CODE COLLATE DATABASE_DEFAULT
                    WHERE {wh} ORDER BY s.QTY DESC""", p)
            else:  # gubun
                cur.execute(f"""SELECT s.SUPPLY_GUBUN, COUNT(DISTINCT s.MAT_CODE) mats, SUM(s.QTY) qty,
                    SUM(CASE WHEN s.SOURCE='프로파일' THEN s.QTY ELSE 0 END) prof_qty FROM nx.plan_mat_source s
                    WHERE {wh} GROUP BY s.SUPPLY_GUBUN ORDER BY SUM(s.QTY) DESC""", p)
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            return {"ok": True, "mode": mode, "rows": rows}
        except Exception as e:
            return {"ok": False, "error": str(e)[:200], "rows": []}
    finally:
        nx.close()

@app.get("/api/sales/forecast")
def sales_forecast(base: str = Query("")):
    """★영업예상매출현황 라이브 API (레거시 dw_pr_plan_190 재현, 정적스냅샷 대체).
       소스=sa_t_plan_item_dtl(union1)+pr_t_plan_input(union4). 단가=pr_m_item_cost(COST_TAG in S/E=LG판매가, 품목단위 최신, cust무관) KRW.
       gross=차감전(=라이브190). net=차감후=gross − union4(pr_t_plan_input)의 첫계획일 과대분 제거. [[nextgen-erp-sales-forecast-190]]"""
    cn = _conn(); cur = cn.cursor()
    try:
        b = _d6(base) if base.strip() else None
        cur.execute("SELECT FORMAT(GETDATE(),'yyMMdd')")
        today = str(cur.fetchone()[0])
        b = b or today
        # union1(sa_t_plan_item_dtl) + union4(pr_t_plan_input), item×ymd×src
        cur.execute(f"""
          SELECT C_ITEM_CODE item, PLAN_YMD ymd, 'u1' src, SUM(CAST(PLAN_QTY AS float)) q
            FROM sa_t_plan_item_dtl WHERE PLAN_YMD>=? GROUP BY C_ITEM_CODE, PLAN_YMD
          UNION ALL
          SELECT ITEM_CODE item, PLAN_YMD ymd, 'u4' src, SUM(CAST(PLAN_QTY AS float)) q
            FROM pr_t_plan_input WHERE PLAN_YMD>=? GROUP BY ITEM_CODE, PLAN_YMD""", b, b)
        src = [(str(a).strip(), str(y).strip(), str(s).strip(), float(qq or 0)) for a, y, s, qq in cur.fetchall()]
        if not src:
            return {"base": b, "days": [], "rows": []}
        base_ymd = min(y for _, y, _, _ in src)  # 첫 계획일(차감 기준)
        # 단가: COST_TAG in (S,E) 최신 COST_APPLY_YMD, 품목단위(cust무관)
        cur.execute("""SELECT c.ITEM_CODE, c.ITEM_COST FROM pr_m_item_cost c
            JOIN (SELECT ITEM_CODE, MAX(COST_APPLY_YMD) mx FROM pr_m_item_cost WHERE COST_TAG IN('S','E') GROUP BY ITEM_CODE) m
              ON c.ITEM_CODE=m.ITEM_CODE AND c.COST_APPLY_YMD=m.mx WHERE c.COST_TAG IN('S','E')""")
        cost = {}
        for ic, ct in cur.fetchall():
            k = str(ic).strip()
            if k not in cost: cost[k] = float(ct or 0)
        cur.execute("SELECT ITEM_CODE, ISNULL(ITEM_DESC,''), ISNULL(WORK_CODE,'') FROM PR_M_ITEM")
        nmm = {}; wcm = {}
        for ic, d, wc in cur.fetchall(): k = str(ic).strip(); nmm[k] = d; wcm[k] = str(wc).strip()
        agg = {}; days = set()
        for item, ymd, s, qty in src:
            days.add(ymd)
            g = agg.get(item)
            if not g:
                g = {"item": item, "nm": nmm.get(item, ""), "wc": wcm.get(item, ""), "cost": cost.get(item, 0), "gdays": {}, "ndays": {}}
                agg[item] = g
            g["gdays"][ymd] = g["gdays"].get(ymd, 0) + qty
            if not (s == 'u4' and ymd == base_ymd):   # ★차감: pr_t_plan_input(u4) 첫날분 제외
                g["ndays"][ymd] = g["ndays"].get(ymd, 0) + qty
        rows = []
        for g in agg.values():
            gq = sum(g["gdays"].values()); nq = sum(g["ndays"].values()); c = g["cost"]
            g["gq"] = gq; g["nq"] = nq; g["gamt"] = round(gq * c); g["namt"] = round(nq * c)
            rows.append(g)
        return {"base": base_ymd, "days": sorted(days), "rows": rows,
                "gross_amt": round(sum(r["gamt"] for r in rows)), "net_amt": round(sum(r["namt"] for r in rows))}
    finally:
        cn.close()

def _step6_sql(cur):
    P = _P
    cur.execute("IF OBJECT_ID('nx.plan_part_temp') IS NOT NULL DROP TABLE nx.plan_part_temp")
    cur.execute(("""
    WITH CTE_BOM(assy_item_code, level_no, item_code, p_item_code, mat_code, cum_use_qty, in_cust_code, vir_item_flag, cum_item_code) AS (
      SELECT DISTINCT a.c_item_code,0,a.c_item_code,a.c_item_code,a.c_item_code,CONVERT(decimal(18,5),1),ISNULL(c.in_cust_code,''),'0',CONVERT(varchar(500),'{'+a.c_item_code+'}')
      FROM nx.plan_item_dtl a JOIN {P}PR_M_ITEM c ON a.c_item_code=c.item_code
      WHERE NOT EXISTS(SELECT 1 FROM {P}PR_M_MAT WHERE mat_code=a.c_item_code)
      UNION ALL
      SELECT cb.assy_item_code,cb.level_no+1,b.item_code,CASE cb.vir_item_flag WHEN '1' THEN cb.p_item_code ELSE b.item_code END,
             b.mat_code,CONVERT(decimal(18,5),cb.cum_use_qty*b.use_qty),ISNULL(c.in_cust_code,''),
             CASE b.vir_item_flag WHEN '1' THEN '1' ELSE '0' END,CONVERT(varchar(500),cb.cum_item_code+'{'+b.mat_code+'}')
      FROM CTE_BOM cb JOIN {P}PR_M_ITEM_BOM b ON cb.mat_code=b.item_code JOIN {P}PR_M_ITEM c ON b.mat_code=c.item_code
      WHERE ISNULL(b.except_flag,'0')<>'1' AND cb.level_no<10 AND NOT EXISTS(SELECT 1 FROM {P}PR_M_MAT WHERE mat_code=b.mat_code))
    SELECT assy_item_code,level_no,item_code,MAX(p_item_code) p_item_code,mat_code,SUM(cum_use_qty) cum_use_qty,MAX(in_cust_code) in_cust_code,MAX(vir_item_flag) vir_item_flag
    INTO nx.plan_part_temp FROM CTE_BOM GROUP BY assy_item_code,level_no,item_code,mat_code OPTION(MAXRECURSION 0)""").replace("{P}", P))
    cur.execute("IF OBJECT_ID('nx.plan_part_gagong') IS NOT NULL DROP TABLE nx.plan_part_gagong")
    cur.execute(("""SELECT a.assy_item_code,a.level_no,a.item_code,a.mat_code,a.p_item_code,a.vir_item_flag,b.proc_seq,g.gc_gubun,a.cum_use_qty,s.gagong_proc_code,b.gagong_proc_seq,b.s_work_code,ISNULL(b.lt_hr,0) lt_hr
    INTO nx.plan_part_gagong FROM nx.plan_part_temp a
    JOIN {P}PR_M_ITEM_PROC_GAGONG b ON a.mat_code=b.item_code JOIN {P}PR_M_WORK_SINGLE s ON b.s_work_code=s.s_work_code JOIN {P}PR_M_PROC_GAGONG g ON s.gagong_proc_code=g.gagong_proc_code
    WHERE a.vir_item_flag='0' AND ISNULL(a.in_cust_code,'') IN ('','2228')""").replace("{P}", P))
    cur.execute("IF OBJECT_ID('nx.plan_part_swork') IS NOT NULL DROP TABLE nx.plan_part_swork")
    cur.execute(("""SELECT b.plan_ymd,b.work_order,b.split_work_order,a.assy_item_code,a.level_no AS bom_level,a.item_code AS upper_item_code,a.mat_code AS item_code,a.p_item_code,a.proc_seq,a.gc_gubun,
      b.line_no,a.cum_use_qty AS use_qty,b.lot_qty,CEILING(CONVERT(float,b.plan_qty)*ISNULL(b.use_qty,1)*ISNULL(c.prod_rate,100)/100) AS plan_qty,
      a.gagong_proc_code,a.gagong_proc_seq,a.s_work_code,a.lt_hr,CEILING(CONVERT(float,b.plan_qty)*ISNULL(b.use_qty,1)*ISNULL(c.prod_rate,100)/100)*a.cum_use_qty AS part_plan_qty
    INTO nx.plan_part_swork FROM nx.plan_part_gagong a JOIN nx.plan_item_dtl b ON a.assy_item_code=b.c_item_code JOIN {P}PR_M_ITEM c ON a.assy_item_code=c.item_code""").replace("{P}", P))
    cur.execute("IF OBJECT_ID('nx.plan_part_dtl') IS NOT NULL DROP TABLE nx.plan_part_dtl")
    cur.execute("""SELECT a.* INTO nx.plan_part_dtl FROM nx.plan_part_swork a
      WHERE a.gagong_proc_code <> ISNULL((SELECT TOP 1 b.gagong_proc_code FROM nx.plan_part_swork b
        WHERE b.plan_ymd=a.plan_ymd AND b.work_order=a.work_order AND b.split_work_order=a.split_work_order AND b.assy_item_code=a.assy_item_code
          AND b.bom_level=a.bom_level AND b.upper_item_code=a.upper_item_code AND b.item_code=a.item_code AND b.proc_seq<a.proc_seq ORDER BY b.proc_seq DESC),'')""")

def _step7_sql(cur):
    P = _P
    cur.execute("IF OBJECT_ID('nx.plan_part_mat_tmp') IS NOT NULL DROP TABLE nx.plan_part_mat_tmp")
    cur.execute(("""
    WITH CTE_BOM(plan_ymd,work_order,split_work_order,assy_item_code,bom_level,upper_item_code,item_code,proc_seq,bom_mat_code,mat_work_center_code,cum_use_qty,cum_in_cust_code,mat_flag,use_qty,part_plan_qty,gc_gubun,cust_flag) AS (
      SELECT a.plan_ymd,a.work_order,a.split_work_order,a.assy_item_code,a.bom_level,a.upper_item_code,a.item_code,a.proc_seq,a.item_code,
         CASE WHEN c.work_code>'' THEN c.work_code ELSE ISNULL(c.in_cust_code,'') END,CONVERT(decimal(18,5),a.use_qty),
         CONVERT(varchar(500),'||'+CASE WHEN c.work_code>'' THEN c.work_code ELSE ISNULL(c.in_cust_code,'') END+'|'),'1',a.use_qty,CONVERT(float,a.part_plan_qty)/NULLIF(a.use_qty,0),a.gc_gubun,'0'
      FROM nx.plan_part_dtl a JOIN {P}PR_M_ITEM c ON a.item_code=c.item_code WHERE a.proc_seq=1
      UNION ALL
      SELECT a.plan_ymd,a.work_order,a.split_work_order,a.c_item_code,0,a.c_item_code,a.c_item_code,1,a.c_item_code,
         CASE WHEN c.work_code>'' THEN c.work_code ELSE ISNULL(c.in_cust_code,'') END,CONVERT(decimal(18,5),a.use_qty),
         CONVERT(varchar(500),'||'+CASE WHEN c.work_code>'' THEN c.work_code ELSE ISNULL(c.in_cust_code,'') END+'|'),'1',a.use_qty,CEILING(CONVERT(float,a.plan_qty)*ISNULL(a.use_qty,1)*ISNULL(c.prod_rate,100)/100),'','1'
      FROM nx.plan_item_dtl a JOIN {P}PR_M_ITEM c ON a.c_item_code=c.item_code
      WHERE NOT EXISTS(SELECT 1 FROM nx.plan_part_dtl d WHERE d.work_order=a.work_order AND d.split_work_order=a.split_work_order AND d.item_code=a.c_item_code)
      UNION ALL
      SELECT cb.plan_ymd,cb.work_order,cb.split_work_order,cb.assy_item_code,cb.bom_level,cb.upper_item_code,cb.item_code,cb.proc_seq,b.mat_code,
         CASE WHEN m.work_code>'' THEN m.work_code ELSE ISNULL(m.in_cust_code,'') END,CONVERT(decimal(18,5),CASE WHEN cb.cum_use_qty=0 THEN 0 ELSE cb.cum_use_qty*b.use_qty END),
         CONVERT(varchar(500),cb.cum_in_cust_code+'|'+CASE WHEN m.work_code>'' THEN m.work_code ELSE ISNULL(m.in_cust_code,'') END+'|'),
         ISNULL((SELECT '2' FROM {P}PR_M_MAT WHERE mat_code=b.mat_code),'1'),cb.use_qty,cb.part_plan_qty,'','1'
      FROM CTE_BOM cb JOIN {P}PR_M_ITEM_BOM b ON cb.bom_mat_code=b.item_code JOIN {P}PR_M_ITEM m ON b.mat_code=m.item_code
      WHERE ISNULL(b.except_flag,'0')<>'1'
        AND NOT EXISTS(SELECT 1 FROM nx.plan_part_dtl d WHERE d.plan_ymd=cb.plan_ymd AND d.work_order=cb.work_order AND d.split_work_order=cb.split_work_order
            AND d.assy_item_code=cb.assy_item_code AND d.bom_level=cb.bom_level+1 AND d.upper_item_code=b.item_code AND d.item_code=b.mat_code))
    SELECT * INTO nx.plan_part_mat_tmp FROM CTE_BOM
    WHERE CHARINDEX('||'+mat_work_center_code+'||',cum_in_cust_code)=0 AND NOT (cust_flag='0' AND gc_gubun='P') OPTION(MAXRECURSION 0)""").replace("{P}", P))
    cur.execute("IF OBJECT_ID('nx.plan_part_mat') IS NOT NULL DROP TABLE nx.plan_part_mat")
    # 최하위집계 + ★용접봉(sgroup910)=공정처리 제외
    cur.execute(("""SELECT a.plan_ymd,a.work_order,a.split_work_order,a.assy_item_code,a.bom_level,a.upper_item_code,a.item_code,a.proc_seq,a.bom_mat_code AS mat_code,
        SUM(a.part_plan_qty*a.cum_use_qty) AS part_plan_qty,MAX(a.mat_flag) mat_flag,MAX(a.mat_work_center_code) mat_work_center_code
    INTO nx.plan_part_mat FROM nx.plan_part_mat_tmp a
    WHERE NOT EXISTS(SELECT 1 FROM nx.plan_part_mat_tmp d WHERE d.work_order=a.work_order AND d.split_work_order=a.split_work_order AND d.assy_item_code=a.assy_item_code AND d.bom_level>a.bom_level AND d.bom_mat_code=a.bom_mat_code)
      AND NOT EXISTS(SELECT 1 FROM {P}PR_M_ITEM wj WHERE wj.item_code=a.bom_mat_code AND wj.item_sgroup='910')
    GROUP BY a.plan_ymd,a.work_order,a.split_work_order,a.assy_item_code,a.bom_level,a.upper_item_code,a.item_code,a.proc_seq,a.bom_mat_code""").replace("{P}", P))

@app.get("/api/plan/part")
def plan_part(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
              part: str = Query(""), assy: str = Query("")):
    # ★정본 파이프라인 전환: nx.plan_part(구 단일패스 98%) → nx.plan_part_mat(레거시 STEP5→6→7 100%검증).
    #   자도번(파트)=mat_code, 작업처=mat_work_center_code, 소요=part_plan_qty.
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("IF OBJECT_ID('nx.plan_part_mat') IS NULL SELECT 1 WHERE 1=0")
        w = ["1=1"]; p = []
        if from_ymd: w.append("pp.PLAN_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("pp.PLAN_YMD<=?"); p.append(_d6(to_ymd))
        if wc.strip():   w.append("pp.MAT_WORK_CENTER_CODE=?"); p.append(wc.strip())
        if part.strip(): w.append("pp.MAT_CODE LIKE ?"); p.append(f"%{part.strip()}%")
        if assy.strip(): w.append("pp.ASSY_ITEM_CODE LIKE ?"); p.append(f"%{assy.strip()}%")
        try:
            cur.execute(f"""SELECT pp.PLAN_YMD, pp.ASSY_ITEM_CODE, pp.MAT_CODE, pp.MAT_WORK_CENTER_CODE,
                  COALESCE(w.WORK_DESC, cu.CUST_DESC, pp.MAT_WORK_CENTER_CODE) wcnm, ISNULL(i.ITEM_DESC,'') nm,
                  SUM(CAST(pp.PART_PLAN_QTY AS float)) q
                FROM nx.plan_part_mat pp
                LEFT JOIN PARTNER_ERP.dbo.PR_M_WORK w ON w.WORK_CODE COLLATE DATABASE_DEFAULT=pp.MAT_WORK_CENTER_CODE COLLATE DATABASE_DEFAULT
                LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST cu ON cu.CUST_CODE COLLATE DATABASE_DEFAULT=pp.MAT_WORK_CENTER_CODE COLLATE DATABASE_DEFAULT
                LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM i ON i.ITEM_CODE COLLATE DATABASE_DEFAULT=pp.MAT_CODE COLLATE DATABASE_DEFAULT
                WHERE {' AND '.join(w)}
                GROUP BY pp.PLAN_YMD, pp.ASSY_ITEM_CODE, pp.MAT_CODE, pp.MAT_WORK_CENTER_CODE,
                  COALESCE(w.WORK_DESC, cu.CUST_DESC, pp.MAT_WORK_CENTER_CODE), i.ITEM_DESC""", *p)
        except Exception:
            return {"dates": [], "rows": [], "part_count": 0, "sum_qty": 0, "note": "편성 먼저 실행(/compose_mat)"}
        cols = [d[0] for d in cur.description]; raw = [dict(zip(cols, r)) for r in cur.fetchall()]
        dates = sorted({r["PLAN_YMD"] for r in raw})
        keyed = {}
        for r in raw:
            k = (r["ASSY_ITEM_CODE"], r["MAT_CODE"], r["MAT_WORK_CENTER_CODE"])
            g = keyed.get(k)
            if not g:
                g = {"assy": r["ASSY_ITEM_CODE"], "part": r["MAT_CODE"], "nm": r["nm"], "wc": r["MAT_WORK_CENTER_CODE"],
                     "wcnm": r["wcnm"], "sg": "", "days": {}, "tot": 0}
                keyed[k] = g
            q = float(r["q"] or 0); g["days"][r["PLAN_YMD"]] = g["days"].get(r["PLAN_YMD"], 0) + q; g["tot"] += q
        rows = sorted(keyed.values(), key=lambda x: (x["wcnm"] or "", x["part"]))
        return {"dates": dates, "rows": rows, "part_count": len(rows), "sum_qty": sum(float(r["q"] or 0) for r in raw)}
    finally:
        nx.close()

# ================= 모델BOM 관리 (w_pr_master_060/020) — 모델→도번 매핑(신규모델 등록) =================
# 조회=PR_M_MODEL_BOM(라이브 62762) ∪ nx.model_bom(우리 신규등록). 편성이 둘 다 사용.
def _ensure_modelbom(cur):
    cur.execute("""IF OBJECT_ID('nx.model_bom') IS NULL CREATE TABLE nx.model_bom(
        MODEL_NO varchar(30) NOT NULL, C_ITEM_CODE varchar(20) NOT NULL, USE_QTY decimal(18,4) DEFAULT 1,
        APPLY_FROM varchar(6) DEFAULT '000000', APPLY_TO varchar(6) DEFAULT '999999',
        REMARKS varchar(100), INS_USER varchar(20), INS_DT datetime DEFAULT getdate(),
        CONSTRAINT PK_nx_model_bom PRIMARY KEY(MODEL_NO,C_ITEM_CODE))""")

@app.get("/api/modelbom/search")
def modelbom_search(q: str = Query(""), by: str = Query("model")):
    """by=model: 모델검색 / by=item: 도번(역방향) 검색."""
    cn = _conn(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        if by == "item":
            cur.execute("""SELECT TOP 100 C_ITEM_CODE cd, COUNT(DISTINCT MODEL_NO) n FROM PR_M_MODEL_BOM
                WHERE C_ITEM_CODE LIKE ? GROUP BY C_ITEM_CODE ORDER BY C_ITEM_CODE""", like)
            return {"by": "item", "rows": [{"code": r[0], "n": r[1]} for r in cur.fetchall()]}
        cur.execute("""SELECT TOP 100 MODEL_NO cd, COUNT(*) n FROM PR_M_MODEL_BOM
            WHERE MODEL_NO LIKE ? GROUP BY MODEL_NO ORDER BY MODEL_NO""", like)
        return {"by": "model", "rows": [{"code": r[0], "n": r[1]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.get("/api/modelbom/get")
def modelbom_get(model: str = Query(""), item: str = Query("")):
    """모델→도번(정방향) 또는 도번→모델(역방향). 라이브 ∪ nx.model_bom."""
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_modelbom(cur)
        if item.strip():  # 역방향
            cur.execute("""SELECT MODEL_NO, C_ITEM_CODE, USE_QTY, CONVERT(varchar,MAKE_YMD), CONVERT(varchar,TO_APPLY_YMD), 'live'
                  FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM WHERE C_ITEM_CODE=?
                UNION ALL SELECT MODEL_NO, C_ITEM_CODE, USE_QTY, APPLY_FROM, APPLY_TO, 'nx' FROM nx.model_bom WHERE C_ITEM_CODE=?
                ORDER BY 1""", item.strip(), item.strip())
        else:
            cur.execute("""SELECT MODEL_NO, C_ITEM_CODE, USE_QTY, CONVERT(varchar,MAKE_YMD), CONVERT(varchar,TO_APPLY_YMD), 'live'
                  FROM PARTNER_ERP.dbo.PR_M_MODEL_BOM WHERE MODEL_NO=?
                UNION ALL SELECT MODEL_NO, C_ITEM_CODE, USE_QTY, APPLY_FROM, APPLY_TO, 'nx' FROM nx.model_bom WHERE MODEL_NO=?
                ORDER BY 2""", model.strip(), model.strip())
        rows = []
        for r in cur.fetchall():
            rows.append({"model": r[0], "item": r[1], "use_qty": float(r[2] or 1),
                         "from": str(r[3] or ''), "to": str(r[4] or ''), "src": r[5]})
        # 도번 품명
        codes = list({r["item"] for r in rows})
        nm = {}
        if codes:
            for i in range(0, len(codes), 900):
                ch = codes[i:i+900]; ph = ",".join("?" * len(ch))
                cur.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,''), ISNULL(IN_CUST_CODE,''), LTRIM(RTRIM(ISNULL(WORK_CODE,''))) FROM PARTNER_ERP.dbo.PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
                for x in cur.fetchall(): nm[x[0]] = {"nm": x[1], "wc": (x[3] if x[3] else x[2])}
        for r in rows:
            info = nm.get(r["item"], {}); r["nm"] = info.get("nm", ""); r["wc"] = info.get("wc", "")
        return {"model": model, "item": item, "rows": rows}
    finally:
        nx.close()

@app.post("/api/modelbom/save")
def modelbom_save(payload: dict = Body(...)):
    """신규 모델→도번 등록/수정(nx.model_bom). 라이브 PR_M_MODEL_BOM은 읽기전용."""
    model = str(payload.get("model", "")).strip()
    rows = payload.get("rows", []) or []
    if not model:
        raise HTTPException(400, "model 필요")
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_modelbom(cur)
        cur.execute("DELETE FROM nx.model_bom WHERE MODEL_NO=?", model)
        saved = 0
        for r in rows:
            it = str(r.get("item", "")).strip()
            if not it: continue
            cur.execute("""INSERT INTO nx.model_bom(MODEL_NO,C_ITEM_CODE,USE_QTY,APPLY_FROM,APPLY_TO,REMARKS,INS_USER)
                VALUES(?,?,?,?,?,?,'web')""", model, it, float(r.get("use_qty") or 1),
                _d6(r.get("from")) or "000000", _d6(r.get("to")) or "999999", (r.get("remarks") or None))
            saved += 1
        return {"ok": True, "count": saved}
    finally:
        nx.close()

# ================= 협력사 ①: 협력사계획현황 (w_pr_outside_040) — nx.plan_part 편성결과 =================
@app.get("/api/partner/workcenters")
def partner_workcenters():
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("IF OBJECT_ID('nx.plan_part_mat') IS NULL SELECT 1 WHERE 1=0")
        C = " COLLATE DATABASE_DEFAULT"
        try:
            cur.execute(f"""SELECT pp.MAT_WORK_CENTER_CODE, COALESCE(w.WORK_DESC, cu.CUST_DESC, pp.MAT_WORK_CENTER_CODE) nm, COUNT(*) n
                FROM nx.plan_part_mat pp
                LEFT JOIN PARTNER_ERP.dbo.PR_M_WORK w ON w.WORK_CODE{C}=pp.MAT_WORK_CENTER_CODE{C}
                LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST cu ON cu.CUST_CODE{C}=pp.MAT_WORK_CENTER_CODE{C}
                WHERE pp.MAT_WORK_CENTER_CODE>'' GROUP BY pp.MAT_WORK_CENTER_CODE, COALESCE(w.WORK_DESC, cu.CUST_DESC, pp.MAT_WORK_CENTER_CODE)
                ORDER BY COUNT(*) DESC""")
            return {"rows": [{"cc": r[0], "nm": r[1], "n": r[2]} for r in cur.fetchall()]}
        except Exception:
            return {"rows": []}
    finally:
        nx.close()

@app.get("/api/partner/planstatus")
def partner_planstatus(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
                       part: str = Query(""), assy: str = Query(""), line: str = Query(""),
                       gubun: str = Query("외주")):
    """협력사(납품업체)별 자도번 일자계획. gubun: 외주(협력사=CUST, 기본)/자체(내부공정=WORK)/전체.
       ★정본 파이프라인 전환: nx.plan_part(구98%) → nx.plan_part_mat(레거시 STEP5→6→7 100%검증). 가공처=mat_work_center_code, 자도번=mat_code."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("IF OBJECT_ID('nx.plan_part_mat') IS NULL SELECT 1 WHERE 1=0")
        C = " COLLATE DATABASE_DEFAULT"
        w = ["1=1"]; p = []
        if from_ymd: w.append("pp.PLAN_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("pp.PLAN_YMD<=?"); p.append(_d6(to_ymd))
        if wc.strip():   w.append("pp.MAT_WORK_CENTER_CODE=?"); p.append(wc.strip())
        if part.strip(): w.append("pp.MAT_CODE LIKE ?"); p.append(f"%{part.strip()}%")
        if assy.strip(): w.append("pp.ASSY_ITEM_CODE LIKE ?"); p.append(f"%{assy.strip()}%")
        if line.strip(): w.append("pd.LINE_NO=?"); p.append(line.strip())
        if gubun == "외주":   w.append("w.WORK_CODE IS NULL AND cu.CUST_CODE IS NOT NULL")  # 거래처(협력사)만
        elif gubun == "자체": w.append("w.WORK_CODE IS NOT NULL")                            # 내부공정(P1/P2)
        # ★정본 nx.plan_part_mat은 자재단위라 행수가 큼(외주 5만+) → 브라우저 과부하 방지: 자도번(part)×가공처 단위로 먼저 집계(일자는 유지)
        #   후 상한(CAP). 필터(가공처/제번/자도번) 걸면 좁혀짐.
        CAP = 4000
        try:
            cur.execute(f"""SELECT TOP {CAP} pp.PLAN_YMD, pp.MAT_WORK_CENTER_CODE, COALESCE(w.WORK_DESC, cu.CUST_DESC, pp.MAT_WORK_CENTER_CODE) wcnm,
                  pp.WORK_ORDER, pp.ASSY_ITEM_CODE, pp.MAT_CODE, ISNULL(i.ITEM_DESC,'') nm,
                  ISNULL(pd.LINE_NO,'') line, ISNULL(pd.MODEL_NO,'') model, SUM(CAST(pp.PART_PLAN_QTY AS float)) q
                FROM nx.plan_part_mat pp
                LEFT JOIN (SELECT WORK_ORDER, MAX(LINE_NO) LINE_NO, MAX(MODEL_NO) MODEL_NO FROM nx.plan_dtl GROUP BY WORK_ORDER) pd ON pd.WORK_ORDER=pp.WORK_ORDER
                LEFT JOIN PARTNER_ERP.dbo.PR_M_WORK w ON w.WORK_CODE{C}=pp.MAT_WORK_CENTER_CODE{C}
                LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST cu ON cu.CUST_CODE{C}=pp.MAT_WORK_CENTER_CODE{C}
                LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM i ON i.ITEM_CODE{C}=pp.MAT_CODE{C}
                WHERE {' AND '.join(w)}
                GROUP BY pp.PLAN_YMD, pp.MAT_WORK_CENTER_CODE, COALESCE(w.WORK_DESC, cu.CUST_DESC, pp.MAT_WORK_CENTER_CODE),
                  pp.WORK_ORDER, pp.ASSY_ITEM_CODE, pp.MAT_CODE, i.ITEM_DESC, pd.LINE_NO, pd.MODEL_NO
                ORDER BY wcnm, pp.WORK_ORDER, pp.MAT_CODE""", *p)
        except Exception as e:
            return {"dates": [], "rows": [], "cnt": 0, "sum_qty": 0, "note": "편성 먼저 실행(생산계획업로드 → 🧾자재소요·조달 편성). 오류: " + str(e)[:120]}
        cols = [d[0] for d in cur.description]; raw = [dict(zip(cols, r)) for r in cur.fetchall()]
        capped = len(raw) >= CAP
        dates = sorted({r["PLAN_YMD"] for r in raw})
        keyed = {}
        for r in raw:
            k = (r["MAT_WORK_CENTER_CODE"], r["WORK_ORDER"], r["ASSY_ITEM_CODE"], r["MAT_CODE"])
            g = keyed.get(k)
            if not g:
                g = {"wc": r["MAT_WORK_CENTER_CODE"], "wcnm": r["wcnm"], "wo": r["WORK_ORDER"], "assy": r["ASSY_ITEM_CODE"],
                     "part": r["MAT_CODE"], "nm": r["nm"], "line": r["line"], "model": r["model"], "days": {}, "tot": 0}
                keyed[k] = g
            q = float(r["q"] or 0); g["days"][r["PLAN_YMD"]] = g["days"].get(r["PLAN_YMD"], 0) + q; g["tot"] += q
        rows = sorted(keyed.values(), key=lambda x: (x["wcnm"] or "", x["line"], x["wo"], x["part"]))
        note = f"⚠ 결과가 많아 상위 {CAP}건만 표시했습니다. 협력사(가공처)·제번·자도번으로 필터하세요." if capped else ""
        return {"dates": dates, "rows": rows, "cnt": len(rows), "sum_qty": sum(float(r["q"] or 0) for r in raw), "note": note}
    finally:
        nx.close()

# ================= 생산 ②: 파트별 생산계획 (w_pr_input_410) — PR_T_PLAN_PART_MAT 라이브 =================
# 협력사계획 생성결과(SP_PR_CREATE_PLAN_협력사계획_생성) = 도번→자도번 전개 + 작업처 라우팅 + 일자별 계획.
@app.get("/api/partplan/list")
def partplan_list(from_ymd: str = Query(""), to_ymd: str = Query(""), wc: str = Query(""),
                  part: str = Query(""), assy: str = Query(""), line: str = Query(""),
                  diam: str = Query(""), thick: str = Query(""), pipe: str = Query("")):
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["p.PART_PLAN_QTY>0"]; pr = []
        if from_ymd: w.append("p.PART_PLAN_YMD>=?"); pr.append(_d6(from_ymd))
        if to_ymd:   w.append("p.PART_PLAN_YMD<=?"); pr.append(_d6(to_ymd))
        if wc.strip():   w.append("p.MAT_WORK_CENTER_CODE=?"); pr.append(wc.strip())
        if part.strip(): w.append("p.MAT_CODE LIKE ?"); pr.append(f"%{part.strip()}%")
        if assy.strip(): w.append("p.ASSY_ITEM_CODE LIKE ?"); pr.append(f"%{assy.strip()}%")
        if line.strip(): w.append("p.LINE_NO=?"); pr.append(line.strip())
        if diam.strip():  w.append("i.ITEM_DIAM=?"); pr.append(float(diam))
        if thick.strip(): w.append("i.ITEM_THICK=?"); pr.append(float(thick))
        if pipe == '1':   w.append("i.METAL_GUBUN IN ('CU','고강도') AND ISNULL(i.ITEM_DIAM,0)>0")  # 동파이프만
        cur.execute(f"""SELECT p.PART_PLAN_YMD, p.ASSY_ITEM_CODE, p.MAT_CODE, MAX(p.LINE_NO) line,
              p.MAT_WORK_CENTER_CODE wc,
              MAX(COALESCE(w.WORK_DESC, cu.CUST_DESC, '')) wcnm, MAX(ISNULL(i.ITEM_DESC,'')) nm,
              MAX(ISNULL(i.ITEM_DIAM,0)) diam, MAX(ISNULL(i.ITEM_THICK,0)) thick, MAX(ISNULL(i.ITEM_LENGTH,0)) length,
              MAX(p.CUM_USE_QTY) useq, SUM(p.PART_PLAN_QTY) pq
            FROM PR_T_PLAN_PART_MAT p
            LEFT JOIN PR_M_WORK w ON w.WORK_CODE=p.MAT_WORK_CENTER_CODE
            LEFT JOIN CM_M_CUST cu ON cu.CUST_CODE=p.MAT_WORK_CENTER_CODE
            LEFT JOIN PR_M_ITEM i ON i.ITEM_CODE=p.MAT_CODE
            WHERE {' AND '.join(w)}
            GROUP BY p.PART_PLAN_YMD, p.ASSY_ITEM_CODE, p.MAT_CODE, p.MAT_WORK_CENTER_CODE""", *pr)
        cols = [d[0] for d in cur.description]
        raw = [dict(zip(cols, r)) for r in cur.fetchall()]
        dates = sorted({r["PART_PLAN_YMD"] for r in raw})
        keyed = {}
        for r in raw:
            k = (r["ASSY_ITEM_CODE"], r["MAT_CODE"], r["wc"])
            g = keyed.get(k)
            if not g:
                g = {"assy": r["ASSY_ITEM_CODE"], "part": r["MAT_CODE"], "nm": r["nm"], "line": r["line"],
                     "wc": r["wc"], "wcnm": r["wcnm"], "use": float(r["useq"] or 0),
                     "diam": float(r["diam"] or 0), "thick": float(r["thick"] or 0), "length": float(r["length"] or 0),
                     "days": {}, "tot": 0}
                keyed[k] = g
            q = float(r["pq"] or 0); g["days"][r["PART_PLAN_YMD"]] = g["days"].get(r["PART_PLAN_YMD"], 0) + q; g["tot"] += q
        rows = sorted(keyed.values(), key=lambda x: (x["wcnm"] or "", x["part"]))
        return {"dates": dates, "rows": rows, "part_count": len(rows),
                "sum_qty": sum(float(r["pq"] or 0) for r in raw)}
    finally:
        cn.close()

@app.get("/api/partplan/workcenters")
def partplan_workcenters():
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("""SELECT p.MAT_WORK_CENTER_CODE cc, COALESCE(w.WORK_DESC, cu.CUST_DESC, '') nm, COUNT(*) n
            FROM PR_T_PLAN_PART_MAT p
            LEFT JOIN PR_M_WORK w ON w.WORK_CODE=p.MAT_WORK_CENTER_CODE
            LEFT JOIN CM_M_CUST cu ON cu.CUST_CODE=p.MAT_WORK_CENTER_CODE
            WHERE p.PART_PLAN_QTY>0 AND p.MAT_WORK_CENTER_CODE>''
            GROUP BY p.MAT_WORK_CENTER_CODE, COALESCE(w.WORK_DESC, cu.CUST_DESC, '')
            ORDER BY COUNT(*) DESC""")
        return {"rows": [{"cc": r[0], "nm": r[1], "n": r[2]} for r in cur.fetchall()]}
    finally:
        cn.close()

# ============ 생산 ③④: 생산실적현황(w_pr_list_010) / 파트별생산실적현황(w_pr_list_090) — PR_T_PROD_DTL ============
def _pivot_prod(raw, keyname, extra):
    dates = sorted({r["PROD_YMD"] for r in raw})
    keyed = {}
    for r in raw:
        k = r[keyname]; g = keyed.get(k)
        if not g:
            g = dict(extra(r)); g["days"] = {}; g["tot"] = 0; keyed[k] = g
        q = float(r["q"] or 0); g["days"][r["PROD_YMD"]] = g["days"].get(r["PROD_YMD"], 0) + q; g["tot"] += q
    return dates, list(keyed.values())

# ================= 조달 프로파일 관리 (발주규칙: 유효기간 1순위 + 다중시 배분) — nx.sourcing_profile =================
def _d(s):  # 'yyyy-mm-dd' → date str or None
    s = str(s or "").strip()
    return s[:10] if len(s) >= 10 and s[4] == '-' else None

@app.get("/api/profile/search")
def profile_search(q: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT TOP 80 sp.item_code, MAX(ISNULL(i.item_name,'')) nm, COUNT(*) profiles,
              SUM(CASE WHEN sp.is_active=1 THEN 1 ELSE 0 END) actives,
              MAX(CASE WHEN sp.supply_gubun='유상사급' THEN 1 ELSE 0 END) has_sagub
            FROM nx.sourcing_profile sp LEFT JOIN nx.item i ON i.item_code=sp.item_code
            WHERE sp.item_code LIKE ? OR i.item_name LIKE ?
            GROUP BY sp.item_code ORDER BY sp.item_code""", like, like)
        cols = [d[0] for d in cur.description]
        return {"rows": [dict(zip(cols, r)) for r in cur.fetchall()]}
    finally:
        nx.close()

@app.get("/api/profile/get")
def profile_get(item: str = Query(...)):
    item = item.strip()
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT item_name FROM nx.item WHERE item_code=?", item)
        r = cur.fetchone(); nm = r[0] if r else ""
        cur.execute("""SELECT sp.profile_id, sp.profile_name, sp.supply_gubun, ISNULL(sp.vendor_code,'') vendor_code,
              ISNULL(c.CUST_DESC,'') vendor_name, sp.lme_flag, CONVERT(varchar(10),sp.apply_from,23) apply_from,
              CONVERT(varchar(10),sp.apply_to,23) apply_to, sp.is_active, sp.is_internal,
              sp.alloc_ratio, sp.priority
            FROM nx.sourcing_profile sp
            LEFT JOIN PARTNER_ERP.dbo.CM_M_CUST c ON c.CUST_CODE=sp.vendor_code
            WHERE sp.item_code=? ORDER BY sp.is_internal DESC, sp.supply_gubun, sp.profile_id""", item)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r["lme_flag"] = bool(r["lme_flag"]); r["is_active"] = bool(r["is_active"]); r["is_internal"] = bool(r["is_internal"])
            r["alloc_ratio"] = float(r["alloc_ratio"]) if r["alloc_ratio"] is not None else None
        return {"item": item, "item_name": nm, "rows": rows}
    finally:
        nx.close()

def _validate_alloc(profs):
    """profs=[(af,at,alloc)] 활성 비내부 프로파일. 유효기간 겹치는 모든 구간에서 배분합=100% 강제.
       단독(그 구간에 1개, alloc=None)은 암묵 100%. 반환: 위반 메시지 리스트(빈=정상)."""
    FAR = "2099-12-31"
    norm = [(af, at or FAR, alloc) for (af, at, alloc) in profs]
    pts = sorted({p[0] for p in norm} | {p[1] for p in norm})
    errs = []
    for seg in pts:
        cover = [al for (af, at, al) in norm if af <= seg <= at]
        if not cover:
            continue
        if len(cover) == 1 and cover[0] is None:
            continue  # 단독=암묵 100%
        total = sum((al or 0) for al in cover)
        if abs(total - 100.0) > 0.01:
            errs.append(f"{seg} 시점 배분합 {total:g}% — 정확히 100%가 되어야 합니다 (활성 프로파일 {len(cover)}개)")
    return list(dict.fromkeys(errs))

@app.post("/api/profile/save")
def profile_save(payload: dict = Body(...)):
    """발주규칙 저장: 유효기간(1순위)·활성·배분(2개이상). 검증: 겹치는 구간마다 배분합=100% 강제(위반시 저장차단·미기록)."""
    rows = payload.get("rows", []) or []
    item = str(payload.get("item", "")).strip()
    nx = _nx(); cur = nx.cursor()
    try:
        # is_internal 맵(내부용은 계획배분에서 제외 = 참조전용)
        internal = {}
        if item:
            cur.execute("SELECT profile_id, is_internal FROM nx.sourcing_profile WHERE item_code=?", item)
            internal = {r[0]: bool(r[1]) for r in cur.fetchall()}
        norm = []; active_non_internal = []
        for r in rows:
            pid = int(r.get("profile_id"))
            af = _d(r.get("apply_from")) or "2000-01-01"
            at = _d(r.get("apply_to"))
            act = bool(r.get("is_active"))
            ratio = r.get("alloc_ratio"); ratio = float(ratio) if (ratio not in (None, "", "null")) else None
            prio = r.get("priority"); prio = int(prio) if (prio not in (None, "", "null")) else None
            norm.append((pid, af, at, 1 if act else 0, ratio, prio))
            if act and not internal.get(pid, False):
                active_non_internal.append((af, at, ratio))
        # ★검증 먼저(쓰기 전) — 위반시 미기록
        errs = _validate_alloc(active_non_internal)
        if errs:
            return {"ok": False, "errors": errs}
        for (pid, af, at, act, ratio, prio) in norm:
            cur.execute("""UPDATE nx.sourcing_profile SET apply_from=?, apply_to=?, is_active=?, alloc_ratio=?, priority=?
                WHERE profile_id=?""", af, at, act, ratio, prio, pid)
        return {"ok": True, "count": len(norm)}
    finally:
        nx.close()


# ================= 생산 쓰기화면 공용 룩업 =================
@app.get("/api/wr/itemsearch")
def wr_itemsearch(q: str = Query("")):
    """품번/품명 부분검색 (자재·도번 입력 도우미)"""
    q = q.strip()
    cn = _conn(); cur = cn.cursor()
    try:
        like = f"%{q}%"
        cur.execute("""SELECT TOP 40 ITEM_CODE, ISNULL(ITEM_DESC,'') nm, ISNULL(ITEM_SGROUP,'') sg
            FROM PR_M_ITEM WHERE ITEM_CODE LIKE ? OR ITEM_DESC LIKE ? ORDER BY ITEM_CODE""", like, like)
        return {"rows": [{"item": r[0], "nm": r[1], "sg": r[2]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.get("/api/wr/works")
def wr_works():
    """작업장 목록 (PR_M_WORK)"""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT WORK_CODE, ISNULL(WORK_DESC,'') nm FROM PR_M_WORK ORDER BY WORK_CODE")
        return {"rows": [{"code": r[0], "nm": r[1]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.get("/api/wr/sworks")
def wr_sworks():
    """공정(S_WORK_CODE) 목록 — 실적 상위 사용코드"""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("""SELECT TOP 40 S_WORK_CODE, COUNT(*) c FROM PR_T_PROD_DTL
            WHERE S_WORK_CODE IS NOT NULL AND S_WORK_CODE>0 GROUP BY S_WORK_CODE ORDER BY c DESC""")
        return {"rows": [{"code": int(r[0])} for r in cur.fetchall()]}
    finally:
        cn.close()


# ================= 470 자재개별재고조정 (w_pr_stock_470) — ★Phase3 단일원장 fold: nx.stock_ledger(STOCK_POINT='PRD') =================
# 생산파트재고조정 = PRD 조정(±). 태그: 불량→'1', 재고조정→'2', 기타→'PE'(레거시 '4', STOCK_POINT로 격리).
# ID = "YMD-SEQ" (원장 복합키). 수정=기존행 삭제 후 신규(일자·부호 변경 안전). 마감월 잠금 가드.
STOCKMAINT_TAGS = {"1": "불량", "2": "재고조정", "4": "기타"}
_SM_UI2LED = {"4": "PE"}   # UI 태그 → 원장 태그
_SM_LED2UI = {"PE": "4"}   # 원장 태그 → UI 태그

@app.get("/api/stockmaint/list")
def stockmaint_list(from_ymd: str = Query(""), to_ymd: str = Query(""), tag: str = Query(""),
                    mat: str = Query(""), wc: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["l.STOCK_POINT='PRD'", "l.MAINT_TAG IN ('1','2','PE')"]; p = []
        if from_ymd: w.append("l.MAINT_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("l.MAINT_YMD<=?"); p.append(_d6(to_ymd))
        if tag.strip():
            ui = tag.strip()[:1]; w.append("l.MAINT_TAG=?"); p.append(_SM_UI2LED.get(ui, ui))
        if mat.strip():  w.append("(l.MAT_CODE LIKE ? OR l.ITEM_CODE LIKE ?)"); p += [f"%{mat.strip()}%"]*2
        if wc.strip():   w.append("(l.WORK_CODE=? OR l.TO_GAGONG_PROC_CODE=?)"); p += [wc.strip(), wc.strip()]
        cur.execute(f"""SELECT TOP 3000 l.MAINT_YMD, l.MAINT_SEQ, ISNULL(l.MAINT_TAG,'') tag,
              ISNULL(l.WORK_CODE,'') work_code, ISNULL(l.GAGONG_PROC_CODE,'') part_code,
              ISNULL(l.MAT_CODE,'') mat_code, ISNULL(im.ITEM_DESC,'') mat_nm,
              ISNULL(l.ITEM_CODE,'') item_code, ISNULL(ii.ITEM_DESC,'') item_nm,
              l.MAINT_QTY, l.MAINT_COST, l.MAINT_AMT, ISNULL(l.REMARKS,'') remarks,
              ISNULL(l.TO_GAGONG_PROC_CODE,'') prod_work_code, ISNULL(l.INSERT_USER_ID,'') usr, l.INSERT_DATETIME
            FROM nx.stock_ledger l
            LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM im ON im.ITEM_CODE=l.MAT_CODE
            LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM ii ON ii.ITEM_CODE=l.ITEM_CODE
            WHERE {' AND '.join(w)} ORDER BY l.MAINT_YMD DESC, l.MAINT_SEQ DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            ui = _SM_LED2UI.get(r["tag"], r["tag"])
            r["ID"] = f'{r["MAINT_YMD"]}-{r["MAINT_SEQ"]}'; r["tag"] = ui
            r["MAINT_QTY"] = float(r["MAINT_QTY"] or 0); r["MAINT_COST"] = float(r["MAINT_COST"] or 0)
            r["MAINT_AMT"] = float(r["MAINT_AMT"] or 0)
            r["tag_nm"] = STOCKMAINT_TAGS.get(ui, ui)
            r["INSERT_DATETIME"] = str(r["INSERT_DATETIME"] or "")[:19]
        return {"rows": rows, "cnt": len(rows), "sum_qty": sum(r["MAINT_QTY"] for r in rows),
                "sum_amt": sum(r["MAINT_AMT"] for r in rows)}
    finally:
        nx.close()

@app.post("/api/stockmaint/save")
def stockmaint_save(payload: dict = Body(...)):
    p = payload
    ymd = _d6(str(p.get("maint_ymd", "")))
    mat = str(p.get("mat_code", "")).strip()[:20]
    if not ymd or not mat:
        raise HTTPException(400, "조정일자·자재코드는 필수입니다.")
    ui_tag = (str(p.get("maint_tag", "2")).strip() or "2")[:1]
    led_tag = _SM_UI2LED.get(ui_tag, ui_tag)
    work = str(p.get("work_code", "")).strip()[:10]
    part = str(p.get("part_code", "")).strip()[:10]
    item = str(p.get("item_code", "")).strip()[:20]
    qty = float(p.get("maint_qty") or 0)
    cost = float(p.get("maint_cost") or 0)
    amt = round(qty * cost, 2)
    rem = str(p.get("remarks", "")).strip()[:255]
    pwc = str(p.get("prod_work_code", "")).strip()[:10]
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    if qty == 0:
        raise HTTPException(400, "조정수량은 0일 수 없습니다(증가 +, 감소 −).")
    mid = p.get("id")
    nx = _nx(); cur = nx.cursor()
    try:
        if _closed(cur, ymd):
            raise HTTPException(400, f"마감월({_ym(ymd)}) 편집 불가")
        cur.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd)
        seq = cur.fetchone()[0]   # 삭제 전 채번 → 수정 시 신규 SEQ(기존과 상이)
        if mid:  # 수정 = 기존행 삭제 후 신규(재키)
            try:
                oy, osq = str(mid).split("-"); osq = int(osq)
                if _closed(cur, oy):
                    raise HTTPException(400, f"마감월({_ym(oy)}) 편집 불가")
                cur.execute("DELETE FROM nx.stock_ledger WHERE STOCK_POINT='PRD' AND MAINT_YMD=? AND MAINT_SEQ=?", oy, osq)
            except (ValueError, AttributeError):
                pass
        cur.execute("""INSERT INTO nx.stock_ledger
            (STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_TAG,GAGONG_PROC_CODE,WORK_CODE,TO_GAGONG_PROC_CODE,
             MAT_CODE,ITEM_CODE,MAINT_QTY,MAINT_COST,MAINT_AMT,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
            VALUES('PRD',?,?,?,?,?,?,?,?,?,?,?,?,?,GETDATE())""",
            ymd, seq, led_tag, (part or None), (work or None), (pwc or None),
            mat, (item or None), qty, cost, amt, (rem or None), usr)
        return {"ok": True, "id": f"{ymd}-{seq}", "mode": ("update" if mid else "insert")}
    finally:
        nx.close()

@app.post("/api/stockmaint/delete")
def stockmaint_delete(payload: dict = Body(...)):
    ids = [str(x) for x in (payload.get("ids", []) or []) if str(x).strip()]
    if not ids:
        return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        dl = 0
        for x in ids:
            try:
                y, sq = x.split("-"); sq = int(sq)
            except ValueError:
                continue
            if _closed(cur, y):
                raise HTTPException(400, f"마감월({_ym(y)}) 삭제 불가")
            cur.execute("DELETE FROM nx.stock_ledger WHERE STOCK_POINT='PRD' AND MAINT_YMD=? AND MAINT_SEQ=?", y, sq)
            dl += cur.rowcount
        return {"ok": True, "deleted": dl}
    finally:
        nx.close()


# ================= 260 공정별 생산실적등록 (w_pr_input_260, PR_T_PROD_DTL) — nx.proc_result =================
@app.get("/api/procreg/list")
def procreg_list(from_ymd: str = Query(""), to_ymd: str = Query(""), swork: str = Query(""),
                 line: str = Query(""), item: str = Query(""), wo: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("d.PROD_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("d.PROD_YMD<=?"); p.append(_d6(to_ymd))
        if swork.strip(): w.append("d.S_WORK_CODE=?"); p.append(int(swork.strip()))
        if line.strip():  w.append("d.LINE_NO=?"); p.append(line.strip())
        if item.strip():  w.append("d.ITEM_CODE LIKE ?"); p.append(f"%{item.strip()}%")
        if wo.strip():    w.append("d.WORK_ORDER LIKE ?"); p.append(f"%{wo.strip()}%")
        cur.execute(f"""SELECT TOP 3000 d.ID, d.PROD_YMD, d.PROD_HMS, ISNULL(d.WORK_ORDER,'') wo,
              ISNULL(d.SPLIT_WORK_ORDER,'') swo, ISNULL(d.ITEM_CODE,'') item, ISNULL(ii.ITEM_DESC,'') nm,
              ISNULL(d.LINE_NO,'') line, ISNULL(d.PART_CODE,'') part, d.S_WORK_CODE sw, d.PROD_QTY,
              ISNULL(d.WORK_CODE,'') work_code, ISNULL(d.FINISH_FLAG,'') fin, ISNULL(d.PROD_USER_ID,'') usr,
              d.INSERT_DATETIME
            FROM nx.proc_result d LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM ii ON ii.ITEM_CODE=d.ITEM_CODE
            WHERE {' AND '.join(w)} ORDER BY d.PROD_YMD DESC, d.PROD_HMS DESC, d.ID DESC""", *p)
        cols = [dd[0] for dd in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r["PROD_QTY"] = float(r["PROD_QTY"] or 0); r["sw"] = str(r["sw"] if r["sw"] is not None else "")
            r["INSERT_DATETIME"] = str(r["INSERT_DATETIME"] or "")[:19]
        return {"rows": rows, "cnt": len(rows), "sum_qty": sum(r["PROD_QTY"] for r in rows)}
    finally:
        nx.close()

@app.post("/api/procreg/save")
def procreg_save(payload: dict = Body(...)):
    from datetime import datetime as _dt
    p = payload
    ymd = _d6(str(p.get("prod_ymd", "")))
    item = str(p.get("item_code", "")).strip()[:20]
    if not ymd or not item:
        raise HTTPException(400, "실적일자·품번은 필수입니다.")
    hms = str(p.get("prod_hms", "")).strip()[:6] or _dt.now().strftime("%H%M%S")
    wo = str(p.get("work_order", "")).strip()[:20]
    swo = str(p.get("split_work_order", "")).strip()[:30]
    line = str(p.get("line_no", "")).strip()[:10]
    part = str(p.get("part_code", "")).strip()[:10]
    sw = p.get("s_work_code"); sw = int(sw) if str(sw).strip() not in ("", "None", "null") else 0
    qty = int(float(p.get("prod_qty") or 0))
    work = str(p.get("work_code", "")).strip()[:10]
    fin = (str(p.get("finish_flag", "0")).strip() or "0")[:1]
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    mid = p.get("id")
    nx = _nx(); cur = nx.cursor()
    try:
        if mid:
            cur.execute("""UPDATE nx.proc_result SET PROD_YMD=?, PROD_HMS=?, WORK_ORDER=?, SPLIT_WORK_ORDER=?,
                ITEM_CODE=?, LINE_NO=?, PART_CODE=?, S_WORK_CODE=?, PROD_QTY=?, WORK_CODE=?, FINISH_FLAG=?,
                PROD_USER_ID=?, UPDATE_USER_ID=?, UPDATE_DATETIME=getdate() WHERE ID=?""",
                ymd, hms, wo, swo, item, line, part, sw, qty, work, fin, usr, usr, int(mid))
            return {"ok": True, "id": int(mid), "mode": "update"}
        cur.execute("""INSERT INTO nx.proc_result(PROD_YMD,PROD_HMS,WORK_ORDER,SPLIT_WORK_ORDER,ITEM_CODE,
            LINE_NO,PART_CODE,S_WORK_CODE,PROD_QTY,WORK_CODE,FINISH_FLAG,PROD_USER_ID,UPDATE_USER_ID)
            OUTPUT INSERTED.ID VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            ymd, hms, wo, swo, item, line, part, sw, qty, work, fin, usr, usr)
        nid = cur.fetchone()[0]
        return {"ok": True, "id": int(nid), "mode": "insert"}
    finally:
        nx.close()

@app.post("/api/procreg/delete")
def procreg_delete(payload: dict = Body(...)):
    ids = [int(x) for x in (payload.get("ids", []) or []) if str(x).strip()]
    if not ids:
        return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        ph = ",".join("?" * len(ids))
        cur.execute(f"DELETE FROM nx.proc_result WHERE ID IN ({ph})", *ids)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()


# ================= 150 자재출고(창고간 출고이동) — ★Phase3 단일원장 fold: nx.stock_ledger(STOCK_POINT='MAT', MV) =================
# 파트창고간 이동 = MAT 이동(net 0). 그룹당 2행(−FROM/+TO, MAINT_GROUP_SEQ 링크, tag='MV').
# ★이중차감 경계(결정 I): 이동은 net-0 relocation → 자재소비 아님 → 백플러시(−MAT 소비)와 구조적으로 겹치지 않음.
#   생산소비는 백플러시(Phase2)가 담당하며 이 화면은 소비 경로가 아님. ID = "YMD-GROUP".
@app.get("/api/matissue/list")
def matissue_list(from_ymd: str = Query(""), to_ymd: str = Query(""), mat: str = Query(""),
                  frompart: str = Query(""), topart: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:  # 대표행 = − 출고행(MAINT_QTY<0): FROM=GAGONG_PROC_CODE, TO=TO_GAGONG_PROC_CODE
        w = ["l.STOCK_POINT='MAT'", "l.MAINT_TAG='MV'", "l.MAINT_QTY<0", "l.MAINT_GROUP_SEQ IS NOT NULL"]; p = []
        if from_ymd: w.append("l.MAINT_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("l.MAINT_YMD<=?"); p.append(_d6(to_ymd))
        if mat.strip():  w.append("(l.MAT_CODE LIKE ? OR l.ITEM_CODE LIKE ?)"); p += [f"%{mat.strip()}%"]*2
        if frompart.strip(): w.append("l.GAGONG_PROC_CODE=?"); p.append(frompart.strip())
        if topart.strip():   w.append("l.TO_GAGONG_PROC_CODE=?"); p.append(topart.strip())
        cur.execute(f"""SELECT TOP 3000 l.MAINT_YMD ISSUE_YMD, l.MAINT_GROUP_SEQ,
              ISNULL(l.GAGONG_PROC_CODE,'') frompart, ISNULL(l.TO_GAGONG_PROC_CODE,'') topart,
              ISNULL(l.WORK_CODE,'') work_code, ISNULL(l.MAT_CODE,'') mat_code, ISNULL(im.ITEM_DESC,'') mat_nm,
              ISNULL(l.ITEM_CODE,'') item_code, ABS(l.MAINT_QTY) ISSUE_QTY, ISNULL(l.REMARKS,'') remarks,
              ISNULL(l.INSERT_USER_ID,'') usr, l.INSERT_DATETIME
            FROM nx.stock_ledger l LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM im ON im.ITEM_CODE=l.MAT_CODE
            WHERE {' AND '.join(w)} ORDER BY l.MAINT_YMD DESC, l.MAINT_GROUP_SEQ DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r["ID"] = f'{r["ISSUE_YMD"]}-{r["MAINT_GROUP_SEQ"]}'
            r["ISSUE_QTY"] = float(r["ISSUE_QTY"] or 0)
            r["INSERT_DATETIME"] = str(r["INSERT_DATETIME"] or "")[:19]
        return {"rows": rows, "cnt": len(rows), "sum_qty": sum(r["ISSUE_QTY"] for r in rows)}
    finally:
        nx.close()

@app.post("/api/matissue/save")
def matissue_save(payload: dict = Body(...)):
    p = payload
    ymd = _d6(str(p.get("issue_ymd", "")))
    mat = str(p.get("mat_code", "")).strip()[:20]
    if not ymd or not mat:
        raise HTTPException(400, "출고일자·자재코드는 필수입니다.")
    frompart = str(p.get("from_part_code", "")).strip()[:10]
    topart = str(p.get("part_code", "")).strip()[:10]
    work = str(p.get("work_code", "")).strip()[:10]
    item = str(p.get("item_code", "")).strip()[:20]
    qty = float(p.get("issue_qty") or 0)
    rem = str(p.get("remarks", "")).strip()[:255]
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    mid = p.get("id")
    if qty <= 0:
        raise HTTPException(400, "출고수량은 0보다 커야 합니다.")
    if not frompart or not topart:
        raise HTTPException(400, "FROM파트·TO파트는 필수입니다(창고간 이동).")
    if frompart == topart:
        raise HTTPException(400, "FROM파트와 TO파트가 같습니다.")
    nx = _nx_tx(); cur = nx.cursor()   # ★원자성: MV 이동 2행(±) 그룹 트랜잭션
    try:
        if _closed(cur, ymd):
            raise HTTPException(400, f"마감월({_ym(ymd)}) 편집 불가")
        cur.execute("SELECT ISNULL(MAX(MAINT_GROUP_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_TAG='MV'")
        gseq = cur.fetchone()[0]   # 삭제 전 채번 → 수정 시 신규 그룹번호(기존과 상이)
        if mid:  # 수정 = 기존 그룹(2행) 삭제 후 재생성
            try:
                oy, og = str(mid).split("-"); og = int(og)
                if _closed(cur, oy):
                    raise HTTPException(400, f"마감월({_ym(oy)}) 편집 불가")
                cur.execute("DELETE FROM nx.stock_ledger WHERE STOCK_POINT='MAT' AND MAINT_TAG='MV' AND MAINT_YMD=? AND MAINT_GROUP_SEQ=?", oy, og)
            except (ValueError, AttributeError):
                pass
        # FROM파트 가용재고 이내(음수재고 방지). 현재고 = 원장 SUM(MAT·해당 파트).
        cur.execute("""SELECT ISNULL(SUM(MAINT_QTY),0) FROM nx.stock_ledger
            WHERE STOCK_POINT='MAT' AND MAT_CODE=? AND ISNULL(GAGONG_PROC_CODE,'')=?""", mat, frompart)
        avail = float(cur.fetchone()[0] or 0)
        if qty > avail:
            raise HTTPException(400, f"FROM파트 재고부족 ({mat}@{frompart} 가용 {avail:g} < 이동 {qty:g})")
        for gpc, to_gpc, sq in ((frompart, topart, -qty), (topart, frompart, qty)):  # −FROM, +TO
            cur.execute("SELECT ISNULL(MAX(MAINT_SEQ),0)+1 FROM nx.stock_ledger WHERE MAINT_YMD=?", ymd)
            seq = cur.fetchone()[0]
            cur.execute("""INSERT INTO nx.stock_ledger
                (STOCK_POINT,MAINT_YMD,MAINT_SEQ,MAINT_GROUP_SEQ,MAINT_TAG,GAGONG_PROC_CODE,TO_GAGONG_PROC_CODE,
                 WORK_CODE,MAT_CODE,ITEM_CODE,MAINT_QTY,REMARKS,INSERT_USER_ID,INSERT_DATETIME)
                VALUES('MAT',?,?,?, 'MV', ?,?,?,?,?,?,?,?,GETDATE())""",
                ymd, seq, gseq, gpc, to_gpc, (work or None), mat, (item or None), sq, (rem or None), usr)
        nx.commit()   # ★2행(−FROM/+TO) 원자 커밋
        return {"ok": True, "id": f"{ymd}-{gseq}", "mode": ("update" if mid else "insert")}
    except Exception:
        nx.rollback(); raise   # 부분실패 시 net-0 불변식 보존(전체 롤백)
    finally:
        nx.close()

@app.post("/api/matissue/delete")
def matissue_delete(payload: dict = Body(...)):
    ids = [str(x) for x in (payload.get("ids", []) or []) if str(x).strip()]
    if not ids:
        return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        dl = 0
        for x in ids:
            try:
                y, g = x.split("-"); g = int(g)
            except ValueError:
                continue
            if _closed(cur, y):
                raise HTTPException(400, f"마감월({_ym(y)}) 삭제 불가")
            cur.execute("DELETE FROM nx.stock_ledger WHERE STOCK_POINT='MAT' AND MAINT_TAG='MV' AND MAINT_YMD=? AND MAINT_GROUP_SEQ=?", y, g)
            dl += cur.rowcount
        return {"ok": True, "deleted": dl}
    finally:
        nx.close()


# ================= 조달경로(SUB변형) 그룹 — 접미사 품번을 '동일결과 SUB' 그룹으로 묶어 조달처 배분 =================
# 모델: Assy의 SUB가 여러 조달처(생산처)로 제작되며 레거시는 품번접미사로 복제. 동일결과(자식set)로 그룹핑,
#       생산처(IN_CUST_CODE)=실제 조달처. 다조달처 그룹은 유효기간+배분%(합100%)로 편성 라우팅. nx.procgroup_alloc
import re as _re
_MK = {"1": "자체", "2": "외주(유상사급)", "3": "매입", "": "미지정"}

def _base_of(code):
    return _re.sub(r'\(.*?\)', '', code).split('-')[0].strip()

def _jac(a, b):
    if not a and not b: return 1.0
    if not a or not b: return 0.0
    return len(a & b) / len(a | b)

def _ensure_procgroup_tbl(cur):
    cur.execute("""IF OBJECT_ID('nx.procgroup_alloc') IS NULL CREATE TABLE nx.procgroup_alloc(
        variant_item varchar(120) PRIMARY KEY, base_item varchar(30), group_key varchar(80),
        vendor_code varchar(10), vendor_name varchar(80), is_new bit DEFAULT 0,
        apply_from date, apply_to date, alloc_ratio decimal(9,4), is_active bit DEFAULT 1,
        priority int, upd_dt datetime DEFAULT getdate())""")
    for col, ddl in [("vendor_code", "varchar(10)"), ("vendor_name", "varchar(80)"), ("is_new", "bit")]:
        cur.execute(f"IF COL_LENGTH('nx.procgroup_alloc','{col}') IS NULL ALTER TABLE nx.procgroup_alloc ADD {col} {ddl}")

@app.get("/api/procgroup/vendors")
def procgroup_vendors(q: str = Query("")):
    """조달처(거래처) 검색 — 새 조달처 후보 추가용"""
    q = q.strip()
    cn = _conn(); cur = cn.cursor()
    try:
        like = f"%{q}%"
        cur.execute("""SELECT TOP 40 CUST_CODE, ISNULL(CUST_DESC,'') FROM CM_M_CUST
            WHERE CUST_CODE LIKE ? OR CUST_DESC LIKE ? ORDER BY CUST_DESC""", like, like)
        return {"rows": [{"code": r[0], "nm": r[1]} for r in cur.fetchall()]}
    finally:
        cn.close()

@app.get("/api/procgroup/get")
def procgroup_get(base: str = Query(...), ymd: str = Query("")):
    """base Assy의 변형들을 '동일결과 SUB' 그룹으로 묶어 조달처 후보 반환 + 저장된 배분 병합."""
    base = base.strip()
    if not base:
        raise HTTPException(400, "base 품번 필요")
    ay = _d6(ymd) if ymd else ""
    cn = _conn(); cur = cn.cursor()
    try:
        # 변형 마스터
        # 레거시 개발품목BOM관리(w_cs_master_120)와 동일하게 전 변형 표기(status 필터 제거).
        # 실제 생산단은 아래 nk>0(현재유효 BOM 보유)로 자동 선별 → (CI적용)/예상가 더미 자동제외.
        cur.execute("""SELECT i.ITEM_CODE, ISNULL(i.ITEM_DESC,''), ISNULL(i.IN_CUST_CODE,''),
              ISNULL(cu.CUST_DESC,''), ISNULL(i.MAKE_TYPE,''), ISNULL(i.ITEM_STATUS,'')
            FROM PR_M_ITEM i LEFT JOIN CM_M_CUST cu ON cu.CUST_CODE=i.IN_CUST_CODE
            WHERE i.ITEM_CODE LIKE ?""", base + '%')
        vs = []
        for ic, nm, cc, cnm, mk, st in cur.fetchall():
            if _base_of(ic) != base and ic != base:
                continue
            vs.append({"item": ic, "nm": nm, "cust_code": cc, "cust": (cnm or ("자체" if not cc else cc)),
                       "mk": mk, "mk_label": _MK.get(mk, mk), "status": st})
        if not vs:
            return {"base": base, "groups": [], "msg": "변형 없음(활성품목)"}
        # 각 변형 BOM 자식 set (현재유효 또는 지정일)
        for v in vs:
            if ay:
                cur.execute("""SELECT MAT_CODE, ISNULL(SAGUB_FLAG,'0') FROM CS_M_ITEM_BOM
                    WHERE ITEM_CODE=? AND FROM_APPLY_YMD<=? AND TO_APPLY_YMD>=?""", v["item"], ay, ay)
            else:
                cur.execute("""SELECT MAT_CODE, ISNULL(SAGUB_FLAG,'0') FROM CS_M_ITEM_BOM
                    WHERE ITEM_CODE=? AND TO_APPLY_YMD>='260601'""", v["item"])
            ch = cur.fetchall()
            v["_kset"] = frozenset(x[0] for x in ch)
            v["nk"] = len(v["_kset"]); v["sag"] = sum(1 for x in ch if x[1] == '1')
            v["is_self"] = (v["item"] == base)
        prod = [v for v in vs if v["nk"] > 0]  # 실제 생산단(BOM 보유)
        # 사용여부 신호: as-built BOM 트리(현재 실제 투입경로) + 26년 확정입고
        cur.execute("SELECT ITEM_CODE, MAT_CODE FROM CS_M_ITEM_BOM WHERE TO_APPLY_YMD>='260601'")
        _ch = {}
        for p, m in cur.fetchall(): _ch.setdefault(p, set()).add(m)
        tree = {base}; stack = [base]
        while stack:
            x = stack.pop()
            for k in _ch.get(x, ()):
                if k not in tree: tree.add(k); stack.append(k)
        for v in vs:
            v["in_tree"] = v["item"] in tree
            # ★현행 = 실입고(recv>0). 개발(조달경로 통합검토 subvariant/get)과 동일 기준(동일 쿼리: 태그무관·MAINT_QTY>0·260101~)
            q = cur.execute("""SELECT ISNULL(SUM(MAINT_QTY),0) FROM PU_T_STOCK_MAINT
                WHERE UPPER(LTRIM(RTRIM(MAT_CODE)))=? AND MAINT_YMD>='260101' AND MAINT_QTY>0""", v["item"].upper()).fetchone()
            v["recv"] = float(q[0] or 0)
            v["used"] = v["recv"] > 0   # 현행=실입고만(개발과 일치). in_tree(BOM경로) 과다판정 제거
            v["mk_conflict"] = (v["mk"] == '3' and v["sag"] > 0)  # 매입인데 사급자식有 = 분류오류 의심
        # 클러스터: 자식 Jaccard>=0.75
        used = [False] * len(prod); clusters = []
        for i in range(len(prod)):
            if used[i]: continue
            cl = [i]; used[i] = True
            for j in range(i + 1, len(prod)):
                if not used[j] and _jac(prod[i]["_kset"], prod[j]["_kset"]) >= 0.75:
                    cl.append(j); used[j] = True
            clusters.append(cl)
    finally:
        cn.close()
    # 저장된 배분 병합
    nx = _nx(); ncur = nx.cursor()
    try:
        _ensure_procgroup_tbl(ncur)
        ncur.execute("""SELECT variant_item, group_key, ISNULL(vendor_code,''), ISNULL(vendor_name,''),
            ISNULL(is_new,0), CONVERT(varchar(10),apply_from,23), CONVERT(varchar(10),apply_to,23),
            alloc_ratio, is_active FROM nx.procgroup_alloc WHERE base_item=?""", base)
        saved = {}; synth = {}
        for r in ncur.fetchall():
            rec = {"group_key": r[1], "vendor_code": r[2], "vendor_name": r[3], "is_new": bool(r[4]),
                   "apply_from": r[5], "apply_to": r[6],
                   "alloc_ratio": float(r[7]) if r[7] is not None else None,
                   "is_active": bool(r[8]) if r[8] is not None else True}
            saved[r[0]] = rec
            if rec["is_new"]:
                synth.setdefault(r[1], []).append((r[0], rec))
    finally:
        nx.close()
    groups = []
    for gi, cl in enumerate(clusters, 1):
        members = [prod[k] for k in cl]
        gkey = f"{base}#{min(m['item'] for m in members)}"
        out_members = []
        for m in members:
            s = saved.get(m["item"], {})
            out_members.append({
                "item": m["item"], "nm": m["nm"], "cust": m["cust"], "cust_code": m["cust_code"],
                "mk": m["mk"], "mk_label": m["mk_label"], "sag": m["sag"], "nk": m["nk"],
                "is_self": m["is_self"], "is_new": False,
                "used": m["used"], "in_tree": m["in_tree"], "recv": m["recv"], "mk_conflict": m["mk_conflict"],
                "apply_from": s.get("apply_from"), "apply_to": s.get("apply_to"),
                # 현행 100% 단일운영 시스템에서 이관 → 저장전 기본값 = 현행(사용중)만 활성
                "alloc_ratio": s.get("alloc_ratio"), "is_active": s.get("is_active", m["used"]),
            })
        # 신규 조달처 후보(가상 — 품번복제 없이 거래처만) 병합
        for (vi, rec) in synth.get(gkey, []):
            out_members.append({
                "item": vi, "nm": "＋신규 조달처", "cust": rec["vendor_name"] or rec["vendor_code"],
                "cust_code": rec["vendor_code"], "mk": "2", "mk_label": "외주(유상사급)", "sag": 0,
                "nk": members[0]["nk"], "is_self": False, "is_new": True,
                "used": True, "in_tree": False, "recv": 0.0, "mk_conflict": False,
                "apply_from": rec["apply_from"], "apply_to": rec["apply_to"],
                "alloc_ratio": rec["alloc_ratio"], "is_active": rec["is_active"],
            })
        if not out_members:
            continue   # 빈 그룹 방지
        custs = sorted(set(m["cust"] for m in out_members))
        groups.append({"group_key": gkey, "gi": gi, "nk": members[0]["nk"],
                       "multi_source": len(custs) >= 2, "custs": custs, "members": out_members})
    # 다조달처(배분대상) 우선 정렬
    groups.sort(key=lambda g: (not g["multi_source"], g["gi"]))
    base_nm = next((v["nm"] for v in vs if v["item"] == base), "")
    return {"base": base, "base_nm": base_nm, "n_variants": len(vs), "n_prod": len(prod),
            "n_groups": len(groups), "groups": groups,
            "msg": ("" if groups else "개발(조달경로 통합검토)에서 이 품번의 조달경로를 아직 포함하지 않았습니다. 개발에서 후보를 ➕포함하면 여기 표시됩니다.")}

@app.post("/api/procgroup/save")
def procgroup_save(payload: dict = Body(...)):
    """조달그룹 배분 저장. 검증: 각 그룹의 활성 조달처 배분합=100%(겹치는 기간마다). 위반시 미기록."""
    base = str(payload.get("base", "")).strip()
    rows = payload.get("rows", []) or []
    # 그룹별 활성 배분 수집 → 검증
    bygroup = {}
    norm = []
    for r in rows:
        vi = str(r.get("variant_item", "")).strip()
        if not vi: continue
        gk = str(r.get("group_key", "")).strip()
        vc = str(r.get("vendor_code", "")).strip()
        vn = str(r.get("vendor_name", "")).strip()
        isnew = 1 if r.get("is_new") else 0
        af = _d(r.get("apply_from")) or "2000-01-01"
        at = _d(r.get("apply_to"))
        act = bool(r.get("is_active"))
        ratio = r.get("alloc_ratio"); ratio = float(ratio) if (ratio not in (None, "", "null")) else None
        norm.append((vi, base, gk, vc, vn, isnew, af, at, 1 if act else 0, ratio))
        if act and ratio is not None:   # 배분참여 = 활성+배분% 입력된 것만(미유효/공란 제외)
            bygroup.setdefault(gk, []).append((af, at, ratio))
    errs = []
    for gk, profs in bygroup.items():
        if len(profs) >= 1:  # 배분값 입력된 조달처들: 유효기간 겹치는 구간마다 합=100%
            errs += [f"[{gk}] {e}" for e in _validate_alloc(profs)]
    if errs:
        return {"ok": False, "errors": list(dict.fromkeys(errs))}
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_procgroup_tbl(cur)
        for (vi, bs, gk, vc, vn, isnew, af, at, act, ratio) in norm:
            cur.execute("DELETE FROM nx.procgroup_alloc WHERE variant_item=?", vi)
            cur.execute("""INSERT INTO nx.procgroup_alloc(variant_item,base_item,group_key,vendor_code,vendor_name,is_new,apply_from,apply_to,alloc_ratio,is_active,upd_dt)
                VALUES(?,?,?,?,?,?,?,?,?,?,getdate())""", vi, bs, gk, vc, vn, isnew, af, at, ratio, act)
        return {"ok": True, "count": len(norm)}
    finally:
        nx.close()


# ============ 공통 SUB 통합 검토화면(nx.sub_variant_map) + 사급·현행 정합 ============
@app.get("/api/subvariant/bases")
def subvariant_bases(q: str = Query("")):
    """통합대상 베이스 목록(다변형). q=품번검색."""
    nx = _nx(); cur = nx.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT m.base_item, COUNT(*) nv, COUNT(DISTINCT m.common_sub) ng,
              SUM(CAST(m.is_current AS int)) cur, MAX(ISNULL(i.item_name,'')) nm
            FROM nx.sub_variant_map m LEFT JOIN nx.item i ON i.item_code=m.base_item
            WHERE m.base_item LIKE ? GROUP BY m.base_item ORDER BY COUNT(*) DESC""", like)
        rows = [{"base": r[0], "nv": r[1], "ng": r[2], "cur": r[3], "nm": r[4]} for r in cur.fetchall()]
        return {"rows": rows[:200]}
    finally:
        nx.close()

@app.get("/api/subvariant/get")
def subvariant_get(base: str = Query(...)):
    """베이스의 -S그룹 + 멤버(공급처·현행·구분·사급부품·오분류경고) + 저장된 승인."""
    base = base.strip()
    nx = _nx(); ncur = nx.cursor()
    try:
        ncur.execute("""SELECT struct_group, common_sub, variant_item, vendor_code, ISNULL(vendor_name,''),
              n_child, is_current FROM nx.sub_variant_map WHERE base_item=? ORDER BY struct_group, variant_item""", base)
        recs = ncur.fetchall()
        # 승인상태
        ncur.execute("""IF OBJECT_ID('nx.subvariant_approve') IS NULL CREATE TABLE nx.subvariant_approve(
            common_sub NVARCHAR(35) PRIMARY KEY, approved BIT, note NVARCHAR(200), approver NVARCHAR(30), upd_dt datetime DEFAULT getdate())""")
        ncur.execute("SELECT common_sub, approved, ISNULL(note,'') FROM nx.subvariant_approve")
        appr = {r[0]: {"approved": bool(r[1]), "note": r[2]} for r in ncur.fetchall()}
        # 조달경로 포함(nx 전용, 마이그레이션 안전) — 담당이 후보 변형을 조달경로로 포함
        ncur.execute("""IF OBJECT_ID('nx.sourcing_path','U') IS NULL CREATE TABLE nx.sourcing_path(
            base_item NVARCHAR(60), variant_item NVARCHAR(60), vendor NVARCHAR(60), included BIT DEFAULT 1,
            note NVARCHAR(200), upd_dt datetime DEFAULT getdate(), CONSTRAINT PK_nx_sourcing_path PRIMARY KEY(base_item, variant_item))""")
        ncur.execute("SELECT LTRIM(RTRIM(variant_item)), included FROM nx.sourcing_path WHERE base_item=?", base)
        incl = {str(r[0]).strip(): bool(r[1]) for r in ncur.fetchall()}
        # ★후보별 공정 = 개발 지정 정본(nx.routing=CS_T_ITEM_PROC=품목별 공정관리)만. 협력사견적(coop)은 '그냥 견적(단가)'이라 공정 기준 아님.
        #   proc_code≥90=관리/운반/이윤 제외, work_qty>0=개발이 실제 지정한 공정.
        rtcode = {}
        try:
            ncur.execute("""SELECT r.item_code, r.proc_code FROM nx.routing r
                JOIN nx.sub_variant_map m ON m.variant_item=r.item_code AND m.base_item=?
                WHERE ISNULL(r.proc_code,'')<>'' AND ISNULL(TRY_CONVERT(int,r.proc_code),0)<90
                AND ISNULL(TRY_CONVERT(float,r.work_qty),0)>0 ORDER BY r.item_code, r.sort_seq""", base)
            for r in ncur.fetchall(): rtcode.setdefault(str(r[0]).strip(), []).append(str(r[1]).strip())
        except Exception:
            pass
        # 저장된 조달 프로파일(유효기간·배분%·활성) — 조달 프로파일 화면 편집값
        prof = {}
        try:
            _ensure_procgroup_tbl(ncur)
            ncur.execute("""SELECT LTRIM(RTRIM(variant_item)), CONVERT(varchar(10),apply_from,23),
                CONVERT(varchar(10),apply_to,23), alloc_ratio, is_active FROM nx.procgroup_alloc WHERE base_item=?""", base)
            for r in ncur.fetchall():
                prof[str(r[0]).strip()] = {"apply_from": r[1], "apply_to": r[2],
                    "alloc_ratio": (float(r[3]) if r[3] is not None else None),
                    "is_active": (bool(r[4]) if r[4] is not None else None)}
        except Exception:
            pass
    finally:
        nx.close()
    if not recs:
        return {"base": base, "groups": [], "msg": "통합대상 아님(단일변형)"}
    items = [r[2] for r in recs]
    # live 보강: 구분(MAKE_TYPE)·사급부품
    cn = _conn(); cur = cn.cursor()
    try:
        mk = {}; nm = {}
        for i in range(0, len(items), 900):
            ch = items[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"SELECT ITEM_CODE, ISNULL(MAKE_TYPE,''), ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
            for r in cur.fetchall(): mk[r[0]] = r[1]; nm[r[0]] = r[2]
        sag = {}
        for i in range(0, len(items), 900):
            ch = items[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"""SELECT LTRIM(RTRIM(ITEM_CODE)), LTRIM(RTRIM(MAT_CODE)) FROM CS_M_ITEM_BOM
                WHERE ITEM_CODE IN ({ph}) AND SAGUB_FLAG='1' AND TO_APPLY_YMD>='260601'""", *ch)
            for r in cur.fetchall(): sag.setdefault(r[0], []).append(r[1])
        # 변형별 실입고(2026, MAINT_QTY>0) — 현행 판정을 플래그 대신 실거래로 (레거시 is_current 오표시 정정)
        recv = {}
        try:
            up = [x.upper() for x in items]
            for i in range(0, len(up), 900):
                ch = up[i:i+900]; ph = ",".join("?" * len(ch))
                cur.execute(f"""SELECT UPPER(LTRIM(RTRIM(MAT_CODE))), SUM(MAINT_QTY), MAX(MAINT_YMD)
                    FROM PU_T_STOCK_MAINT WHERE UPPER(LTRIM(RTRIM(MAT_CODE))) IN ({ph})
                    AND MAINT_YMD>='260101' AND MAINT_QTY>0 GROUP BY UPPER(LTRIM(RTRIM(MAT_CODE)))""", *ch)
                for r in cur.fetchall(): recv[r[0]] = {"qty": float(r[1] or 0), "last": str(r[2] or "")}
        except Exception:
            pass
        # 공정코드→이름(CS_M_PROC): nx.routing proc_code 라벨(컷팅/면취/CNC…)
        pname = {}
        try:
            allc = sorted({c for lst in rtcode.values() for c in lst})
            for i in range(0, len(allc), 900):
                ch = allc[i:i+900]; ph = ",".join("?" * len(ch))
                cur.execute(f"SELECT PROC_CODE, PROC_DESC FROM CS_M_PROC WHERE PROC_CODE IN ({ph})", *ch)
                for r in cur.fetchall(): pname[str(r[0]).strip()] = str(r[1]).strip()
        except Exception:
            pass
    finally:
        cn.close()
    # 그룹 구성
    gmap = {}
    for sg, csub, vi, vc, vn, nch, iscur in recs:
        g = gmap.setdefault(csub, {"struct_group": sg, "common_sub": csub, "n_child": nch, "members": [],
                                    "approved": appr.get(csub, {}).get("approved", False), "note": appr.get(csub, {}).get("note", "")})
        mkc = mk.get(vi, ""); slist = sag.get(vi, []); rv = recv.get(vi.upper(), {})
        ops = [pname.get(c, c) for c in rtcode.get(str(vi).strip(), [])]   # ★개발 정본(품목별 공정관리)만. 협력사견적은 공정 기준 아님
        g["members"].append({
            "variant": vi, "nm": nm.get(vi, ""), "vendor": vn or ("자체" if not vc else vc), "vendor_code": vc,
            "mk": mkc, "mk_label": _MK.get(mkc, mkc), "sag_parts": slist, "sag_cnt": len(slist),
            "mk_conflict": (mkc == '3' and len(slist) > 0), "is_current": bool(iscur),
            "recv_qty": rv.get("qty", 0), "recv_last": rv.get("last", ""), "real_current": rv.get("qty", 0) > 0,
            "included": incl.get(str(vi).strip(), False),
            "ops": ops, "proc_sig": "|".join(ops),
            "needs_proc": True,
            # 조달 프로파일 편집값(유효기간·배분%·활성). 저장 전엔 None
            "apply_from": prof.get(str(vi).strip(), {}).get("apply_from"),
            "apply_to": prof.get(str(vi).strip(), {}).get("apply_to"),
            "alloc_ratio": prof.get(str(vi).strip(), {}).get("alloc_ratio"),
            "prof_active": prof.get(str(vi).strip(), {}).get("is_active")})
    groups = list(gmap.values())
    for g in groups:
        g["vendors"] = sorted(set(m["vendor"] for m in g["members"]))
        g["multi"] = len(g["vendors"]) >= 2  # 통합가치=다공급처
    groups.sort(key=lambda g: (not g["multi"], g["struct_group"]))
    return {"base": base, "base_nm": nm.get(base, "") or next((m["nm"] for g in groups for m in g["members"]), ""),
            "n_variants": len(items), "n_groups": len(groups), "groups": groups}

@app.post("/api/subvariant/approve")
def subvariant_approve(payload: dict = Body(...)):
    """담당 승인: -S그룹의 '동일결과' 통합 승인/보류 + 사유."""
    csub = str(payload.get("common_sub", "")).strip()
    if not csub:
        raise HTTPException(400, "common_sub 필요")
    approved = 1 if payload.get("approved") else 0
    note = str(payload.get("note", "")).strip()[:200]
    user = (str(payload.get("user", "")).strip() or "담당")[:30]
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""IF OBJECT_ID('nx.subvariant_approve') IS NULL CREATE TABLE nx.subvariant_approve(
            common_sub NVARCHAR(35) PRIMARY KEY, approved BIT, note NVARCHAR(200), approver NVARCHAR(30), upd_dt datetime DEFAULT getdate())""")
        cur.execute("DELETE FROM nx.subvariant_approve WHERE common_sub=?", csub)
        cur.execute("INSERT INTO nx.subvariant_approve(common_sub,approved,note,approver,upd_dt) VALUES(?,?,?,?,getdate())",
                    csub, approved, note, user)
        return {"ok": True}
    finally:
        nx.close()

@app.post("/api/subvariant/include")
def subvariant_include(payload: dict = Body(...)):
    """조달경로 포함 토글(nx.sourcing_path, nx 전용·마이그레이션 안전) — 후보 변형을 이 품번의 조달경로로 포함/제외."""
    base = str(payload.get("base", "")).strip()
    variant = str(payload.get("variant", "")).strip()
    if not base or not variant:
        raise HTTPException(400, "base·variant 필요")
    included = 1 if payload.get("included", True) else 0
    vendor = str(payload.get("vendor", "")).strip()[:60]
    note = str(payload.get("note", "")).strip()[:200]
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""IF OBJECT_ID('nx.sourcing_path','U') IS NULL CREATE TABLE nx.sourcing_path(
            base_item NVARCHAR(60), variant_item NVARCHAR(60), vendor NVARCHAR(60), included BIT DEFAULT 1,
            note NVARCHAR(200), upd_dt datetime DEFAULT getdate(), CONSTRAINT PK_nx_sourcing_path PRIMARY KEY(base_item, variant_item))""")
        cur.execute("DELETE FROM nx.sourcing_path WHERE base_item=? AND variant_item=?", base, variant)
        cur.execute("""INSERT INTO nx.sourcing_path(base_item,variant_item,vendor,included,note,upd_dt)
                       VALUES(?,?,?,?,?,getdate())""", base, variant, vendor, included, note)
        return {"ok": True, "included": bool(included)}
    finally:
        nx.close()


# ================= 조달경로 통합검토 재설계 (route 기반) — nx.sourcing_route(헤더) + nx.sourcing_route_line(라인) =================
# 상단=우리 기준 BOM(전 공정 우리가 만든다, bom/tree real=0). 하단=조달경로 후보 CRUD(경로1=현행 baseline·경로2..=대안).
# 승인게이트: 저장/편집 시 approve_flag=0 → 개발 승인(approve)해야 조달프로파일 후보로 노출(단일 소스 정합).
_ROUTE_GUBUN = ["자체", "매입", "외주유상", "외주무상"]          # 경로 헤더 구분
_LINE_GUBUN = ["제작", "매입", "사급"]                            # 라인 구분

def _ensure_route_tbl(cur):
    cur.execute("""IF OBJECT_ID('nx.sourcing_route','U') IS NULL CREATE TABLE nx.sourcing_route(
        route_id INT IDENTITY(1,1) PRIMARY KEY, item_code NVARCHAR(60) NOT NULL, route_no INT NOT NULL,
        route_name NVARCHAR(80), vendor_code NVARCHAR(20), gubun NVARCHAR(20), current_flag BIT DEFAULT 0,
        approve_flag BIT DEFAULT 0, apply_from DATE, note NVARCHAR(200),
        ins_user NVARCHAR(30), ins_dt datetime DEFAULT getdate(), upd_user NVARCHAR(30), upd_dt datetime DEFAULT getdate())""")
    cur.execute("""IF OBJECT_ID('nx.sourcing_route_line','U') IS NULL CREATE TABLE nx.sourcing_route_line(
        line_id INT IDENTITY(1,1) PRIMARY KEY, route_id INT NOT NULL, sort_seq INT DEFAULT 0,
        child_item NVARCHAR(60), child_name NVARCHAR(120), qty FLOAT, gubun NVARCHAR(20), vendor_code NVARCHAR(20),
        is_rawmat BIT DEFAULT 0, diam FLOAT, thick FLOAT, len_val FLOAT, material NVARCHAR(40),
        spec NVARCHAR(80), note NVARCHAR(200))""")
    # 반려(승인관리) 컬럼 멱등 추가
    cur.execute("IF COL_LENGTH('nx.sourcing_route','reject_flag') IS NULL ALTER TABLE nx.sourcing_route ADD reject_flag BIT DEFAULT 0")
    cur.execute("IF COL_LENGTH('nx.sourcing_route','reject_reason') IS NULL ALTER TABLE nx.sourcing_route ADD reject_reason NVARCHAR(200)")
    cur.execute("IF COL_LENGTH('nx.sourcing_route','reject_user') IS NULL ALTER TABLE nx.sourcing_route ADD reject_user NVARCHAR(30)")
    cur.execute("IF COL_LENGTH('nx.sourcing_route','reject_dt') IS NULL ALTER TABLE nx.sourcing_route ADD reject_dt datetime")

def _route_baseline_lines(item):
    """현행(baseline) 경로 라인 = 대상(item)의 실사용 BOM 직하위(level1). 구분: 사급/제작/매입, 공급처=IN_CUST, 치수 보강.
    ★조달후보=SUB/조달대상 단위(하단): 대상별 현행 경로1 = 그 대상의 직하위 구성/공급처."""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("""SELECT LTRIM(RTRIM(b.MAT_CODE)) child, CAST(b.USE_QTY AS float) q, ISNULL(b.SAGUB_FLAG,'0') sag,
              ISNULL(m.ITEM_DESC,'') nm, ISNULL(m.MAKE_TYPE,'') mk, ISNULL(m.IN_CUST_CODE,'') cust,
              ISNULL(c.CUST_DESC,'') custnm, ISNULL(m.METAL_GUBUN,'') metal,
              ISNULL(m.ITEM_DIAM,0) diam, ISNULL(m.ITEM_THICK,0) thick, ISNULL(m.ITEM_LENGTH,0) len,
              ISNULL(b.BOM_SEQ,0) sq
            FROM CS_M_ITEM_BOM b
            LEFT JOIN PR_M_ITEM m ON m.ITEM_CODE=b.MAT_CODE
            LEFT JOIN CM_M_CUST c ON c.CUST_CODE=m.IN_CUST_CODE
            WHERE b.ITEM_CODE=? AND b.FROM_APPLY_YMD<='991231' AND b.TO_APPLY_YMD>='260101'
              AND ISNULL(b.CS_CALC_EXCEPT_FLAG,'0')<>'1' ORDER BY b.BOM_SEQ""", item.strip())
        out = []
        for i, r in enumerate(cur.fetchall(), 1):
            gub = "사급" if str(r.sag) == '1' else ("제작" if str(r.mk) == '1' else "매입")
            out.append({"line_id": 0, "sort_seq": i, "child_item": r.child, "child_name": r.nm, "qty": float(r.q or 0),
                        "gubun": gub, "vendor_code": str(r.cust).strip(), "vendor_name": r.custnm,
                        "is_rawmat": 1 if str(r.metal).strip() else 0, "diam": float(r.diam or 0), "thick": float(r.thick or 0),
                        "len_val": float(r.len or 0), "material": str(r.metal).strip(), "spec": "", "note": ""})
        return out
    finally:
        cn.close()

def _custnm_map(cur, codes):
    m = {}
    codes = sorted({str(c).strip() for c in codes if str(c or "").strip()})
    for i in range(0, len(codes), 900):
        ch = codes[i:i+900]; ph = ",".join("?" * len(ch))
        cur.execute(f"SELECT CUST_CODE, ISNULL(CUST_DESC,'') FROM PARTNER_ERP.dbo.CM_M_CUST WHERE CUST_CODE IN ({ph})", *ch)
        for r in cur.fetchall(): m[str(r[0]).strip()] = r[1]
    return m

@app.get("/api/sourcing/routes")
def sourcing_routes(item: str = Query(...), show_unapproved: int = Query(1), for_profile: int = Query(0)):
    """품목의 조달경로 후보 목록. 경로1=현행(baseline, 실사용BOM 파생·읽기전용) 항상 포함. 경로2..=저장된 대안(nx.sourcing_route).
    for_profile=1: approve_flag=1(승인)만 편성가능; show_unapproved=0 이면 미승인 완전제외, =1 이면 미승인도 회색/읽기전용 포함."""
    item = item.strip()
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        cur.execute("""SELECT r.route_id, r.route_no, ISNULL(r.route_name,''), ISNULL(r.vendor_code,''), ISNULL(r.gubun,''),
              r.current_flag, r.approve_flag, CONVERT(varchar(10),r.apply_from,23), ISNULL(r.note,''),
              ISNULL(r.reject_flag,0), ISNULL(r.reject_reason,'')
            FROM nx.sourcing_route r WHERE r.item_code=? ORDER BY r.route_no""", item)
        hdrs = cur.fetchall()
        routes = []
        vcodes = set()
        for h in hdrs:
            rid = int(h[0])
            cur.execute("""SELECT line_id, sort_seq, ISNULL(child_item,''), ISNULL(child_name,''), qty, ISNULL(gubun,''),
                  ISNULL(vendor_code,''), is_rawmat, diam, thick, len_val, ISNULL(material,''), ISNULL(spec,''), ISNULL(note,'')
                FROM nx.sourcing_route_line WHERE route_id=? ORDER BY sort_seq, line_id""", rid)
            lines = []
            for l in cur.fetchall():
                vcodes.add(l[6])
                lines.append({"line_id": int(l[0]), "sort_seq": int(l[1] or 0), "child_item": l[2], "child_name": l[3],
                              "qty": float(l[4] or 0), "gubun": l[5], "vendor_code": str(l[6]).strip(),
                              "is_rawmat": int(l[7] or 0), "diam": float(l[8] or 0), "thick": float(l[9] or 0),
                              "len_val": float(l[10] or 0), "material": l[11], "spec": l[12], "note": l[13]})
            vcodes.add(str(h[3]).strip())
            routes.append({"route_id": rid, "route_no": int(h[1]), "route_name": h[2], "vendor_code": str(h[3]).strip(),
                           "gubun": h[4], "current_flag": bool(h[5]), "approve_flag": bool(h[6]), "apply_from": h[7],
                           "note": h[8], "reject_flag": bool(h[9]), "reject_reason": h[10], "baseline": False, "lines": lines})
        # 현행 baseline 합성(저장된 route_no=1 이 없을 때만) — 읽기전용·자동승인 기준선
        has_saved_current = any(r["current_flag"] or r["route_no"] == 1 for r in routes)
        if not has_saved_current:
            blines = _route_baseline_lines(item)
            for l in blines: vcodes.add(l["vendor_code"])
            routes.insert(0, {"route_id": 0, "route_no": 1, "route_name": "현행(실사용 BOM)", "vendor_code": "",
                              "gubun": "자체", "current_flag": True, "approve_flag": True, "apply_from": None,
                              "note": "레거시 실사용 BOM 파생(기준선·읽기전용). 대안은 [복사]로 생성.",
                              "reject_flag": False, "reject_reason": "", "baseline": True, "lines": blines})
        # 벤더 코드→이름
        vmap = _custnm_map(cur, vcodes)
        for r in routes:
            r["vendor_name"] = vmap.get(r["vendor_code"], r["vendor_code"])
            for l in r["lines"]: l["vendor_name"] = vmap.get(l["vendor_code"], l["vendor_code"])
        cur.execute("SELECT ISNULL(item_name,'') FROM nx.item WHERE item_code=?", item)
        r = cur.fetchone(); nm = r[0] if r else ""
        # 조달프로파일용 필터: 승인분만(선택적으로 미승인 회색포함)
        if for_profile:
            def keep(r):
                if r["approve_flag"]: return True
                return bool(show_unapproved)
            fr = [dict(r, readonly=(not r["approve_flag"])) for r in routes if keep(r)]
            routes = fr
        return {"item": item, "item_name": nm, "gubun_opts": _ROUTE_GUBUN, "line_gubun_opts": _LINE_GUBUN,
                "routes": routes}
    finally:
        nx.close()

def _route_hdr_errors(p):
    errs = []
    if not str(p.get("gubun", "")).strip(): errs.append("구분은 필수입니다")
    if str(p.get("gubun", "")).strip() != "자체" and not str(p.get("vendor_code", "")).strip():
        errs.append("공급처는 필수입니다(자체 제외)")
    if not str(p.get("apply_from", "")).strip(): errs.append("유효일자(적용시작)는 필수입니다")
    return errs

@app.post("/api/sourcing/route/save")
def sourcing_route_save(payload: dict = Body(...)):
    """경로 헤더 추가/수정 → nx.sourcing_route. 필수=구분·공급처(자체제외)·유효일자·현행여부. ★편집 시 approve_flag=0(승인 리셋)."""
    p = payload
    item = str(p.get("item_code", "")).strip()
    if not item: raise HTTPException(400, "item_code 필요")
    errs = _route_hdr_errors(p)
    if errs: return {"ok": False, "errors": errs}
    rid = p.get("route_id")
    gub = str(p.get("gubun", "")).strip()[:20]
    ven = str(p.get("vendor_code", "")).strip()[:20]
    cur_f = 1 if p.get("current_flag") else 0
    apf = _d(p.get("apply_from"))
    note = str(p.get("note", "")).strip()[:200]
    rname = str(p.get("route_name", "")).strip()[:80]
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        if rid and int(rid) > 0:
            cur.execute("""UPDATE nx.sourcing_route SET route_name=?, vendor_code=?, gubun=?, current_flag=?,
                  apply_from=?, note=?, approve_flag=0, upd_user=?, upd_dt=getdate() WHERE route_id=?""",
                  rname, ven, gub, cur_f, apf, note, usr, int(rid))
            if cur.rowcount == 0: raise HTTPException(404, f"대상 없음(route_id={rid})")
            return {"ok": True, "route_id": int(rid), "mode": "update", "approve_reset": True}
        cur.execute("SELECT ISNULL(MAX(route_no),1) FROM nx.sourcing_route WHERE item_code=?", item)
        rno = int(cur.fetchone()[0]) + 1
        cur.execute("""INSERT INTO nx.sourcing_route(item_code,route_no,route_name,vendor_code,gubun,current_flag,
              approve_flag,apply_from,note,ins_user) OUTPUT INSERTED.route_id VALUES(?,?,?,?,?,?,0,?,?,?)""",
              item, rno, (rname or f"대안 {rno}"), ven, gub, cur_f, apf, note, usr)
        nid = int(cur.fetchone()[0])
        return {"ok": True, "route_id": nid, "route_no": rno, "mode": "insert"}
    finally:
        nx.close()

@app.post("/api/sourcing/route/copy")
def sourcing_route_copy(payload: dict = Body(...)):
    """경로 복사 → 새 대안 후보(approve_flag=0). 원본=저장경로(source_route_id>0) / 현행 baseline(source_route_id=0)
    / ★특정 품번(source_item): 그 품번의 현행 BOM(실사용) 직하위를 seed로 복사(대상 품목과 달라도 참조복사 가능).
    copy_children=1: 하위품번을 신규 채번(접미사 suffix, 기본 -S{route_no})으로 복제(nx.item 최소등록). =0: 기존 품번 유지."""
    p = payload
    item = str(p.get("item_code", "")).strip()
    if not item: raise HTTPException(400, "item_code 필요")
    src_rid = int(p.get("source_route_id") or 0)
    src_item = str(p.get("source_item", "")).strip()   # 특정 품번에서 복사(현행 BOM seed)
    copy_children = 1 if p.get("copy_children") else 0
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    nx = _nx_tx(); cur = nx.cursor()   # 헤더+라인 원자적
    try:
        _ensure_route_tbl(cur)
        # 원본 라인 확보
        if src_item:   # ★특정 품번의 현행 BOM을 seed로
            bl = _route_baseline_lines(src_item)
            if not bl: raise HTTPException(404, f"「{src_item}」의 현행 BOM 구성이 없습니다(품번 확인)")
            src_hdr = {"route_name": f"{src_item} 참조복사", "vendor_code": "", "gubun": "자체",
                       "apply_from": None, "note": f"{src_item} 현행 BOM 참조복사"}
            src_lines = [[l["sort_seq"], l["child_item"], l["child_name"], l["qty"], l["gubun"], l["vendor_code"],
                         l["is_rawmat"], l["diam"], l["thick"], l["len_val"], l["material"], l["spec"], l["note"]] for l in bl]
        elif src_rid > 0:
            cur.execute("""SELECT ISNULL(route_name,''), ISNULL(vendor_code,''), ISNULL(gubun,''),
                  CONVERT(varchar(10),apply_from,23), ISNULL(note,'') FROM nx.sourcing_route WHERE route_id=? AND item_code=?""", src_rid, item)
            h = cur.fetchone()
            if not h: raise HTTPException(404, "원본 경로 없음")
            src_hdr = {"route_name": h[0], "vendor_code": h[1], "gubun": h[2], "apply_from": h[3], "note": h[4]}
            cur.execute("""SELECT sort_seq, ISNULL(child_item,''), ISNULL(child_name,''), qty, ISNULL(gubun,''),
                  ISNULL(vendor_code,''), is_rawmat, diam, thick, len_val, ISNULL(material,''), ISNULL(spec,''), ISNULL(note,'')
                FROM nx.sourcing_route_line WHERE route_id=? ORDER BY sort_seq, line_id""", src_rid)
            src_lines = [list(r) for r in cur.fetchall()]
        else:   # 현행 baseline 파생
            src_hdr = {"route_name": "현행 복사", "vendor_code": "", "gubun": "자체", "apply_from": None, "note": "현행(실사용 BOM) 복사"}
            bl = _route_baseline_lines(item)
            src_lines = [[l["sort_seq"], l["child_item"], l["child_name"], l["qty"], l["gubun"], l["vendor_code"],
                         l["is_rawmat"], l["diam"], l["thick"], l["len_val"], l["material"], l["spec"], l["note"]] for l in bl]
        # 새 route_no
        cur.execute("SELECT ISNULL(MAX(route_no),1) FROM nx.sourcing_route WHERE item_code=?", item)
        rno = int(cur.fetchone()[0]) + 1
        suffix = str(p.get("suffix", "") or f"-S{rno}").strip()[:8]
        cur.execute("""INSERT INTO nx.sourcing_route(item_code,route_no,route_name,vendor_code,gubun,current_flag,
              approve_flag,apply_from,note,ins_user) OUTPUT INSERTED.route_id VALUES(?,?,?,?,?,0,0,?,?,?)""",
              item, rno, f"대안 {rno} (복사)", src_hdr["vendor_code"], src_hdr["gubun"],
              _d(src_hdr["apply_from"]), src_hdr["note"], usr)
        nid = int(cur.fetchone()[0])
        new_children = []
        for ln in src_lines:
            child = str(ln[1]).strip()
            if copy_children and child:
                newc = (child + suffix)[:60]
                cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", newc)
                if not cur.fetchone():
                    cur.execute("INSERT INTO nx.item(item_code,item_name,item_type) VALUES(?,?,N'부품')",
                                newc, (str(ln[2]) or child)[:120])
                new_children.append({"old": child, "new": newc})
                child = newc
            cur.execute("""INSERT INTO nx.sourcing_route_line(route_id,sort_seq,child_item,child_name,qty,gubun,
                  vendor_code,is_rawmat,diam,thick,len_val,material,spec,note) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  nid, ln[0], child, ln[2], ln[3], ln[4], ln[5], ln[6], ln[7], ln[8], ln[9], ln[10], ln[11], ln[12])
        nx.commit()
        return {"ok": True, "route_id": nid, "route_no": rno, "lines": len(src_lines),
                "copied_children": new_children, "suffix": suffix}
    except Exception:
        nx.rollback(); raise
    finally:
        nx.close()

@app.post("/api/sourcing/route/delete")
def sourcing_route_delete(payload: dict = Body(...)):
    """경로 삭제(헤더+라인). 현행 baseline(route_id=0)은 삭제 불가. 근거키=route_id."""
    rid = int(payload.get("route_id") or 0)
    if rid <= 0: raise HTTPException(400, "현행(기준선)은 삭제할 수 없습니다 — 대안 경로만 삭제 가능")
    nx = _nx_tx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        cur.execute("SELECT current_flag FROM nx.sourcing_route WHERE route_id=?", rid)
        r = cur.fetchone()
        if not r: raise HTTPException(404, "대상 없음")
        cur.execute("DELETE FROM nx.sourcing_route_line WHERE route_id=?", rid)
        cur.execute("DELETE FROM nx.sourcing_route WHERE route_id=?", rid)
        nx.commit()
        return {"ok": True, "deleted": rid}
    except Exception:
        nx.rollback(); raise
    finally:
        nx.close()

@app.post("/api/sourcing/route/approve")
def sourcing_route_approve(payload: dict = Body(...)):
    """개발 승인 토글(approve_flag). =1 이라야 조달프로파일 후보로 노출. 현행 baseline(route_id=0)은 항상 승인상태."""
    rid = int(payload.get("route_id") or 0)
    if rid <= 0: return {"ok": True, "approve_flag": True}   # baseline 자동승인
    ap = 1 if payload.get("approve") else 0
    usr = (str(payload.get("user", "")).strip() or "개발")[:30]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        cur.execute("UPDATE nx.sourcing_route SET approve_flag=?, upd_user=?, upd_dt=getdate() WHERE route_id=?", ap, usr, rid)
        if cur.rowcount == 0: raise HTTPException(404, "대상 없음")
        return {"ok": True, "approve_flag": bool(ap)}
    finally:
        nx.close()

def _line_errors(p):
    errs = []
    if not str(p.get("child_item", "")).strip(): errs.append("하위품번은 필수입니다")
    if not str(p.get("child_name", "")).strip(): errs.append("품명은 필수입니다")
    try: q = float(p.get("qty") or 0)
    except Exception: q = 0
    if q <= 0: errs.append("소요량(수량)은 0보다 커야 합니다(계산 필수)")
    g = str(p.get("gubun", "")).strip()
    if not g: errs.append("구분은 필수입니다")
    if g == "매입" and not str(p.get("vendor_code", "")).strip(): errs.append("매입 구분은 공급처가 필수입니다")
    if p.get("is_rawmat"):
        for k, lab in (("diam", "외경"), ("thick", "두께"), ("len_val", "길이"), ("material", "재질")):
            v = p.get(k)
            if k == "material":
                if not str(v or "").strip(): errs.append(f"{lab}은(는) 소재계산 대상 시 필수입니다")
            else:
                try: fv = float(v or 0)
                except Exception: fv = 0
                if fv <= 0: errs.append(f"{lab}은(는) 소재계산 대상 시 필수입니다(>0)")
    return errs

@app.post("/api/sourcing/line/save")
def sourcing_line_save(payload: dict = Body(...)):
    """경로 BOM 라인 추가/수정. 필수=하위품번·품명·소요량(>0)·구분; 매입이면 공급처; 소재계산 대상이면 외경/두께/길이/재질.
    ★라인 편집 시 해당 경로 approve_flag=0(승인 리셋). 현행 baseline(route_id=0)은 편집불가."""
    p = payload
    rid = int(p.get("route_id") or 0)
    if rid <= 0: raise HTTPException(409, "현행(기준선)은 편집할 수 없습니다 — [복사]로 대안 경로를 만들어 편집하세요")
    errs = _line_errors(p)
    if errs: return {"ok": False, "errors": errs}
    lid = int(p.get("line_id") or 0)
    vals = (str(p.get("child_item", "")).strip()[:60], str(p.get("child_name", "")).strip()[:120],
            float(p.get("qty") or 0), str(p.get("gubun", "")).strip()[:20], str(p.get("vendor_code", "")).strip()[:20],
            1 if p.get("is_rawmat") else 0, float(p.get("diam") or 0), float(p.get("thick") or 0),
            float(p.get("len_val") or 0), str(p.get("material", "")).strip()[:40],
            str(p.get("spec", "")).strip()[:80], str(p.get("note", "")).strip()[:200])
    usr = (str(p.get("user", "")).strip() or "웹사용자")[:30]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        if lid > 0:
            cur.execute("""UPDATE nx.sourcing_route_line SET child_item=?, child_name=?, qty=?, gubun=?, vendor_code=?,
                  is_rawmat=?, diam=?, thick=?, len_val=?, material=?, spec=?, note=? WHERE line_id=?""", *vals, lid)
            if cur.rowcount == 0: raise HTTPException(404, "대상 라인 없음")
            out_lid = lid
        else:
            cur.execute("SELECT ISNULL(MAX(sort_seq),0)+1 FROM nx.sourcing_route_line WHERE route_id=?", rid)
            seq = int(cur.fetchone()[0])
            cur.execute("""INSERT INTO nx.sourcing_route_line(route_id,sort_seq,child_item,child_name,qty,gubun,vendor_code,
                  is_rawmat,diam,thick,len_val,material,spec,note) OUTPUT INSERTED.line_id VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  rid, seq, *vals)
            out_lid = int(cur.fetchone()[0])
        cur.execute("UPDATE nx.sourcing_route SET approve_flag=0, upd_user=?, upd_dt=getdate() WHERE route_id=?", usr, rid)
        return {"ok": True, "line_id": out_lid, "approve_reset": True}
    finally:
        nx.close()

@app.post("/api/sourcing/line/delete")
def sourcing_line_delete(payload: dict = Body(...)):
    """경로 라인 삭제(nx). 삭제 시 해당 경로 approve_flag=0. 근거키=line_id."""
    lid = int(payload.get("line_id") or 0)
    if lid <= 0: raise HTTPException(400, "line_id 필요")
    usr = (str(payload.get("user", "")).strip() or "웹사용자")[:30]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        cur.execute("SELECT route_id FROM nx.sourcing_route_line WHERE line_id=?", lid)
        r = cur.fetchone()
        if not r: raise HTTPException(404, "대상 없음")
        rid = int(r[0])
        cur.execute("DELETE FROM nx.sourcing_route_line WHERE line_id=?", lid)
        cur.execute("UPDATE nx.sourcing_route SET approve_flag=0, upd_user=?, upd_dt=getdate() WHERE route_id=?", usr, rid)
        return {"ok": True, "deleted": lid}
    finally:
        nx.close()

@app.post("/api/sourcing/child/new")
def sourcing_child_new(payload: dict = Body(...)):
    """신규 하위품번 채번(기존 복사 + 접미사). nx.item 최소등록. 반환=신규코드·품명."""
    base = str(payload.get("base_child", "")).strip()
    suffix = str(payload.get("suffix", "") or "-S1").strip()[:8]
    name = str(payload.get("name", "")).strip()[:120]
    if not base: raise HTTPException(400, "base_child(원본 하위품번) 필요")
    newc = (base + suffix)[:60]
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT ISNULL(item_name,'') FROM nx.item WHERE item_code=?", base)
        r = cur.fetchone()
        bn = name or (r[0] if r else base)
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", newc)
        if cur.fetchone(): return {"ok": True, "code": newc, "name": bn, "existed": True}
        cur.execute("INSERT INTO nx.item(item_code,item_name,item_type) VALUES(?,?,N'부품')", newc, bn)
        return {"ok": True, "code": newc, "name": bn, "existed": False}
    finally:
        nx.close()


# ================= 조달후보 승인관리 (개발) — 미승인(approve_flag=0·반려아님) 목록·상세·개별/일괄 승인·반려 =================
@app.get("/api/sourcing/pending")
def sourcing_pending(item: str = Query(""), gubun: str = Query(""), user: str = Query(""),
                     from_ymd: str = Query(""), to_ymd: str = Query(""), include_rejected: int = Query(0)):
    """미승인 조달후보 목록(approve_flag=0). 기본 반려제외(include_rejected=1이면 반려도 포함). 품목·구분·등록자·기간(등록일) 필터."""
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        w = ["r.approve_flag=0", "ISNULL(r.current_flag,0)=0"]; p = []
        if not include_rejected: w.append("ISNULL(r.reject_flag,0)=0")
        if item.strip(): w.append("r.item_code LIKE ?"); p.append(f"%{item.strip()}%")
        if gubun.strip(): w.append("r.gubun=?"); p.append(gubun.strip())
        if user.strip(): w.append("r.ins_user LIKE ?"); p.append(f"%{user.strip()}%")
        if from_ymd.strip(): w.append("CONVERT(varchar(10),r.ins_dt,23)>=?"); p.append(from_ymd.strip()[:10])
        if to_ymd.strip(): w.append("CONVERT(varchar(10),r.ins_dt,23)<=?"); p.append(to_ymd.strip()[:10])
        cur.execute(f"""SELECT r.route_id, r.item_code, ISNULL(r.route_no,0), ISNULL(r.route_name,''), ISNULL(r.gubun,''),
              ISNULL(r.vendor_code,''), ISNULL(r.ins_user,''), CONVERT(varchar(16),r.ins_dt,120), ISNULL(r.reject_flag,0),
              ISNULL(r.reject_reason,''), (SELECT COUNT(*) FROM nx.sourcing_route_line l WHERE l.route_id=r.route_id) nline
            FROM nx.sourcing_route r WHERE {' AND '.join(w)} ORDER BY r.ins_dt DESC, r.route_id DESC""", *p)
        cols = [c[0] for c in cur.description]; recs = cur.fetchall()
        rows = []; vcodes = set(); icodes = set()
        for rr in recs:
            d = {"route_id": int(rr[0]), "item_code": str(rr[1]).strip(), "route_no": int(rr[2]),
                 "route_name": rr[3], "gubun": rr[4], "vendor_code": str(rr[5]).strip(), "ins_user": rr[6],
                 "ins_dt": rr[7], "reject_flag": bool(rr[8]), "reject_reason": rr[9], "n_line": int(rr[10] or 0)}
            vcodes.add(d["vendor_code"]); icodes.add(d["item_code"]); rows.append(d)
        # 코드→이름(품목=live PR_M_ITEM, 공급처=CM_M_CUST)
        vmap = _custnm_map(cur, vcodes)
        imap = {}
        if icodes:
            cn = _conn(); c2 = cn.cursor()
            try:
                il = list(icodes)
                for i in range(0, len(il), 900):
                    ch = il[i:i+900]; ph = ",".join("?" * len(ch))
                    c2.execute(f"SELECT ITEM_CODE, ISNULL(ITEM_DESC,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
                    for r in c2.fetchall(): imap[str(r[0]).strip()] = r[1]
            finally: cn.close()
        for d in rows:
            d["item_name"] = imap.get(d["item_code"], "")
            d["vendor_name"] = vmap.get(d["vendor_code"], d["vendor_code"])
        return {"rows": rows, "cnt": len(rows), "gubun_opts": _ROUTE_GUBUN}
    finally:
        nx.close()

@app.get("/api/sourcing/route/detail")
def sourcing_route_detail(route_id: int = Query(...)):
    """승인관리 상세: 후보 헤더 + 라인(공급처 코드→이름)."""
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        cur.execute("""SELECT route_id,item_code,ISNULL(route_no,0),ISNULL(route_name,''),ISNULL(gubun,''),ISNULL(vendor_code,''),
              approve_flag,ISNULL(reject_flag,0),ISNULL(reject_reason,''),CONVERT(varchar(10),apply_from,23),ISNULL(note,''),ISNULL(ins_user,'')
            FROM nx.sourcing_route WHERE route_id=?""", route_id)
        h = cur.fetchone()
        if not h: raise HTTPException(404, "대상 없음")
        cur.execute("""SELECT line_id,sort_seq,ISNULL(child_item,''),ISNULL(child_name,''),qty,ISNULL(gubun,''),
              ISNULL(vendor_code,''),is_rawmat,diam,thick,len_val,ISNULL(material,''),ISNULL(spec,'')
            FROM nx.sourcing_route_line WHERE route_id=? ORDER BY sort_seq,line_id""", route_id)
        lines = []; vcodes = {str(h[5]).strip()}
        for l in cur.fetchall():
            vcodes.add(str(l[6]).strip())
            lines.append({"line_id": int(l[0]), "child_item": l[2], "child_name": l[3], "qty": float(l[4] or 0),
                          "gubun": l[5], "vendor_code": str(l[6]).strip(), "is_rawmat": int(l[7] or 0),
                          "diam": float(l[8] or 0), "thick": float(l[9] or 0), "len_val": float(l[10] or 0),
                          "material": l[11], "spec": l[12]})
        vmap = _custnm_map(cur, vcodes)
        for l in lines: l["vendor_name"] = vmap.get(l["vendor_code"], l["vendor_code"])
        hdr = {"route_id": int(h[0]), "item_code": str(h[1]).strip(), "route_no": int(h[2]), "route_name": h[3],
               "gubun": h[4], "vendor_code": str(h[5]).strip(), "vendor_name": vmap.get(str(h[5]).strip(), str(h[5]).strip()),
               "approve_flag": bool(h[6]), "reject_flag": bool(h[7]), "reject_reason": h[8], "apply_from": h[9],
               "note": h[10], "ins_user": h[11]}
        return {"header": hdr, "lines": lines}
    finally:
        nx.close()

@app.post("/api/sourcing/route/approve_bulk")
def sourcing_route_approve_bulk(payload: dict = Body(...)):
    """일괄 승인(개발). route_ids[] approve_flag=1 + reject 해제. 근거키=route_id 목록."""
    ids = [int(x) for x in (payload.get("route_ids", []) or []) if str(x).strip().isdigit()]
    usr = (str(payload.get("user", "")).strip() or "개발")[:30]
    if not ids: return {"ok": True, "approved": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        ph = ",".join("?" * len(ids))
        cur.execute(f"UPDATE nx.sourcing_route SET approve_flag=1, reject_flag=0, reject_reason=NULL, upd_user=?, upd_dt=getdate() WHERE route_id IN ({ph})", usr, *ids)
        return {"ok": True, "approved": cur.rowcount}
    finally:
        nx.close()

@app.post("/api/sourcing/route/reject")
def sourcing_route_reject(payload: dict = Body(...)):
    """반려(개발): approve_flag=0 유지 + reject_flag=1 + 사유 기록. 조달프로파일 미노출. 근거키=route_id."""
    rid = int(payload.get("route_id") or 0)
    reason = str(payload.get("reason", "")).strip()[:200]
    usr = (str(payload.get("user", "")).strip() or "개발")[:30]
    if rid <= 0: raise HTTPException(400, "route_id 필요")
    if not reason: raise HTTPException(400, "반려 사유는 필수입니다")
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_route_tbl(cur)
        cur.execute("UPDATE nx.sourcing_route SET approve_flag=0, reject_flag=1, reject_reason=?, reject_user=?, reject_dt=getdate() WHERE route_id=?",
                    reason, usr, rid)
        if cur.rowcount == 0: raise HTTPException(404, "대상 없음")
        return {"ok": True, "route_id": rid}
    finally:
        nx.close()


# ============ 기준정보: 단가변동내역(전사 라이브 피드) — 품목단가조회에 통합 ============
_COST_TAG = {"1": "매입", "E": "판매(수출)", "S": "판매(내수)"}
@app.get("/api/price/history")
def price_history(from_ymd: str = Query(""), to_ymd: str = Query(""), item: str = Query(""),
                  tag: str = Query(""), changed: str = Query("")):
    """전사 단가변동 피드: PR_M_ITEM_COST 적용일 내림차순 + 직전단가 대비 Δ(LAG). 라이브·읽기전용."""
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("H.apply_ymd>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("H.apply_ymd<=?"); p.append(_d6(to_ymd))
        if item.strip(): w.append("H.item LIKE ?"); p.append(f"%{item.strip()}%")
        if tag.strip():  w.append("H.tag=?"); p.append(tag.strip())
        if changed == "1": w.append("H.prev IS NOT NULL AND H.cost<>H.prev")
        cur.execute(f"""WITH H AS (
            SELECT ITEM_CODE item, COST_TAG tag, ISNULL(CUST_CODE,'') cust, ISNULL(MKT,'') mkt,
                   ISNULL(CURRENCY,'') curr, COST_APPLY_YMD apply_ymd, ITEM_COST cost, MAT_COST mat,
                   PROC_COST procc, OTHER_COST oth, PUR_RATE rate, ISNULL(INSERT_USER_ID,'') usr,
                   INSERT_DATETIME idt, ISNULL(REMARKS,'') remarks,
                   LAG(ITEM_COST) OVER (PARTITION BY ITEM_CODE,COST_TAG,ISNULL(CUST_CODE,''),ISNULL(MKT,'')
                                        ORDER BY COST_APPLY_YMD, INSERT_DATETIME) prev
            FROM PR_M_ITEM_COST)
          SELECT TOP 3000 H.item, ISNULL(i.ITEM_DESC,'') nm, H.tag, H.cust, ISNULL(c.CUST_DESC,'') cust_nm,
                 H.mkt, H.curr, H.apply_ymd, H.cost, H.mat, H.procc, H.oth, H.rate, H.prev, H.usr,
                 H.idt, H.remarks
          FROM H LEFT JOIN PR_M_ITEM i ON i.ITEM_CODE=H.item LEFT JOIN CM_M_CUST c ON c.CUST_CODE=H.cust
          WHERE {' AND '.join(w)} ORDER BY H.apply_ymd DESC, H.idt DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            for k in ("cost", "mat", "procc", "oth", "rate", "prev"):
                r[k] = float(r[k]) if r[k] is not None else None
            r["delta"] = round(r["cost"] - r["prev"], 2) if (r["prev"] is not None and r["cost"] is not None) else None
            r["tag_nm"] = _COST_TAG.get(str(r["tag"]).strip(), str(r["tag"]))
            r["idt"] = str(r["idt"] or "")[:19]
        return {"rows": rows, "cnt": len(rows),
                "changed": sum(1 for r in rows if r["delta"] not in (None, 0))}
    finally:
        cn.close()

# ---------- 품목별 단가조회(라이브) — 레거시 w_pr_master_150: 거래처별·적용월 시계열 ----------
_CURR_NM = {"KRW": "원화", "USD": "달러", "RMB": "위안", "EUR": "유로", "JPY": "엔"}
@app.get("/api/price/search")
def price_search(q: str = Query(""), lgroup: str = Query(""), sgroup: str = Query(""), limit: int = Query(1000)):
    """단가보유 품목 검색(품번/품명 + 대/소분류 필터). 좌측 리스트. 분류 코드→이름."""
    cn = _conn(); cur = cn.cursor()
    try:
        dLG = _kindmap(cur, "PR005"); dSG = _kindmap(cur, "PR006")
        w = ""; p = []
        if q.strip(): w += " AND (i.ITEM_CODE LIKE ? OR i.ITEM_DESC LIKE ?)"; p += [f"%{q.strip()}%"] * 2
        if lgroup.strip(): w += " AND i.ITEM_LGROUP=?"; p.append(lgroup.strip())
        if sgroup.strip(): w += " AND i.ITEM_SGROUP=?"; p.append(sgroup.strip())
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),1000))} i.ITEM_CODE, ISNULL(i.ITEM_DESC,'') nm, ISNULL(i.ITEM_SPEC,'') spec,
              ISNULL(i.ITEM_LGROUP,'') lg, ISNULL(i.ITEM_SGROUP,'') sg, (SELECT COUNT(*) FROM PR_M_ITEM_COST x WHERE x.ITEM_CODE=i.ITEM_CODE) cnt
            FROM PR_M_ITEM i
            WHERE EXISTS(SELECT 1 FROM PR_M_ITEM_COST x WHERE x.ITEM_CODE=i.ITEM_CODE){w}
            ORDER BY i.ITEM_CODE""", *p)
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            d["lg_nm"] = dLG.get(str(d.get("lg", "")).strip(), str(d.get("lg", "")).strip())
            d["sg_nm"] = dSG.get(str(d.get("sg", "")).strip(), str(d.get("sg", "")).strip())
            rows.append(d)
        cur.execute("""SELECT DISTINCT ISNULL(i.ITEM_LGROUP,'') lg, ISNULL(i.ITEM_SGROUP,'') sg FROM PR_M_ITEM i
            WHERE EXISTS(SELECT 1 FROM PR_M_ITEM_COST x WHERE x.ITEM_CODE=i.ITEM_CODE)""")
        lgs, sgs = set(), set()
        for r in cur.fetchall():
            if str(r[0]).strip(): lgs.add(str(r[0]).strip())
            if str(r[1]).strip(): sgs.add(str(r[1]).strip())
        return {"rows": rows, "cnt": len(rows),
                "lgroups": [{"code": s, "nm": dLG.get(s, s)} for s in sorted(lgs)],
                "sgroups": [{"code": s, "nm": dSG.get(s, s)} for s in sorted(sgs)]}
    finally:
        cn.close()

@app.get("/api/price/item")
def price_item(item: str = Query("")):
    """품번별 단가 이력(거래처별·단가구분·적용월 시계열, 코드→이름). 소급조회의 원장."""
    item = item.strip()
    if not item: return {"item": item, "nm": "", "spec": "", "rows": []}
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT ISNULL(ITEM_DESC,''), ISNULL(ITEM_SPEC,'') FROM PR_M_ITEM WHERE ITEM_CODE=?", item)
        r0 = cur.fetchone(); nm, spec = (r0[0], r0[1]) if r0 else ("", "")
        cur.execute("""SELECT h.COST_TAG, ISNULL(h.CUST_CODE,'') cust, ISNULL(c.CUST_DESC,'') cust_nm,
              h.COST_APPLY_YMD, ISNULL(h.CURRENCY,'') curr, ISNULL(h.MAIN_FLAG,'') main_flag, ISNULL(h.MKT,'') mkt,
              h.ITEM_COST, h.MAT_COST, h.PROC_COST, h.OTHER_COST, ISNULL(h.REMARKS,'') remarks
            FROM PR_M_ITEM_COST h LEFT JOIN CM_M_CUST c ON c.CUST_CODE=h.CUST_CODE
            WHERE h.ITEM_CODE=? ORDER BY h.COST_TAG, h.CUST_CODE, h.COST_APPLY_YMD DESC""", item)
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            for k in ("ITEM_COST", "MAT_COST", "PROC_COST", "OTHER_COST"):
                d[k.lower()] = float(d.pop(k)) if d.get(k) is not None else 0.0
            d["tag_nm"] = _COST_TAG.get(str(d["COST_TAG"]).strip(), str(d["COST_TAG"]))
            d["curr_nm"] = _CURR_NM.get(str(d["curr"]).strip(), str(d["curr"]).strip())
            d["main"] = 1 if str(d["main_flag"]).strip() == "1" else 0
            d["apply_ymd"] = str(d.pop("COST_APPLY_YMD") or "")
            d["tag"] = str(d.pop("COST_TAG")).strip()
            rows.append(d)
        return {"item": item, "nm": nm, "spec": spec, "rows": rows, "cnt": len(rows)}
    finally:
        cn.close()


# ---------- 품목조회(라이브 PR_M_ITEM 전 컬럼, 코드→이름) — 레거시 w_pr_master_010 ----------
_ITEM_MAKE = {"": "", "1": "자체생산", "2": "외주가공", "3": "매입", "4": "사급가공", "5": "외주완성"}
_ITEM_WORK = {"": "", "P1": "용접", "P2": "가공", "D1": "직납"}
def _kindmap(cur, kind):
    cur.execute("SELECT DETAIL_CODE, DETAIL_DESC FROM CM_M_MASTER_DETAIL WHERE KIND_CODE=?", kind)
    return {str(r[0]).strip(): str(r[1] or "").strip() for r in cur.fetchall()}

_MAT_SGROUP = ('210', '220', '230', '310', '910', '991', '992', '993')  # 자재(원소재/원자재/부자재/사급/잡자재/소모품)
@app.get("/api/item/list")
def item_list(q: str = Query(""), lgroup: str = Query(""), sgroup: str = Query(""), mat: str = Query(""),
              nature: str = Query(""), limit: int = Query(500)):
    """품목조회(라이브). 대/소분류·품목형태·단위·재질·협력사·작업처·제작유형 코드→이름. mat=1→자재만. nature=성격6그룹(nx.item 조인)."""
    cn = _conn(); cur = cn.cursor()
    try:
        dLG = _kindmap(cur, "PR005"); dSG = _kindmap(cur, "PR006"); dPK = _kindmap(cur, "PR021")
        dUN = _kindmap(cur, "CM002"); dMT = _kindmap(cur, "PR019")
        cur.execute("SELECT CUST_CODE, ISNULL(CUST_DESC,'') FROM CM_M_CUST")
        dCust = {str(r[0]).strip(): r[1] for r in cur.fetchall()}
        w = ["1=1"]; p = []
        if q.strip(): w.append("(i.ITEM_CODE LIKE ? OR i.ITEM_DESC LIKE ?)"); p += [f"%{q.strip()}%"] * 2
        if lgroup.strip(): w.append("i.ITEM_LGROUP=?"); p.append(lgroup.strip())
        if sgroup.strip(): w.append("i.ITEM_SGROUP=?"); p.append(sgroup.strip())
        if mat == "1": w.append(f"i.ITEM_SGROUP IN ({','.join('?'*len(_MAT_SGROUP))})"); p += list(_MAT_SGROUP)
        if nature.strip(): w.append("nx.nature=?"); p.append(nature.strip())
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),3000))} i.ITEM_CODE, ISNULL(i.ITEM_DESC,'') nm, ISNULL(i.ITEM_SPEC,'') spec,
              ISNULL(i.ITEM_LGROUP,'') lg, ISNULL(i.ITEM_SGROUP,'') sg, ISNULL(i.PIPE_KIND,'') pk, ISNULL(i.UNIT,'') un,
              i.ITEM_DIAM, i.ITEM_THICK, i.ITEM_LENGTH, i.ITEM_WEIGHT, ISNULL(i.METAL_GUBUN,'') metal, ISNULL(i.IN_CUST_CODE,'') incust,
              ISNULL(i.WORK_CODE,'') work, ISNULL(i.MAKE_TYPE,'') mk, ISNULL(i.COST_GUBUN,'') cg, ISNULL(i.ITEM_STATUS,'') status,
              ISNULL(i.SAFE_STOCK_MIN,0), ISNULL(i.SAFE_STOCK_MAX,0), ISNULL(i.KITTING_MIN,0),
              ISNULL(i.WELD_POINT_IN,0), ISNULL(i.WELD_POINT_OUT,0), ISNULL(i.TARIFF_RATE,0), ISNULL(i.REMARKS,'') remarks, ISNULL(i.ITEM_COST,0),
              ISNULL(nx.nature,'') nature, nx.active
            FROM PR_M_ITEM i LEFT JOIN PARTNER_ERP_TEST3.nx.item nx ON nx.item_code = i.ITEM_CODE COLLATE DATABASE_DEFAULT
            WHERE {' AND '.join(w)} ORDER BY i.ITEM_CODE""", *p)
        rows = []
        for r in cur.fetchall():
            g = lambda i: str(r[i] or "").strip()
            num = lambda i: (float(r[i]) if r[i] is not None else 0)
            rows.append({
                "item_code": g(0), "nm": g(1), "spec": g(2),
                "lgroup": dLG.get(g(3), g(3)), "sgroup": dSG.get(g(4), g(4)), "pipe_kind": dPK.get(g(5), g(5)),
                "unit": dUN.get(g(6), g(6)), "diam": num(7), "thick": num(8), "length": num(9), "weight": num(10),
                "metal": dMT.get(g(11), g(11)), "in_cust": dCust.get(g(12), g(12)), "work": _ITEM_WORK.get(g(13), g(13)),
                "make_type": _ITEM_MAKE.get(g(14), g(14)), "cost_gubun": g(15), "status": ("사용" if g(16) in ("", "1", "사용") else g(16)),
                "safe_min": num(17), "safe_max": num(18), "kitting_min": num(19),
                "weld_in": num(20), "weld_out": num(21), "tariff": num(22), "remarks": g(23), "item_cost": num(24),
                "nature": g(25), "active": (0 if (r[26] is not None and not r[26]) else 1)})
        # 분류 드롭다운
        return {"rows": rows, "cnt": len(rows),
                "lgroups": [{"code": k, "nm": v} for k, v in sorted(dLG.items())],
                "sgroups": [{"code": k, "nm": v} for k, v in sorted(dSG.items())],
                "natures": [{"code": n, "nm": n} for n in _NATURE_ALL]}
    finally:
        cn.close()


# ===================== 품목마스터 CRUD (Phase②) — nx.item + item_sub/valve/his =====================
# 근거: _schema/ITEM_MASTER_ANALYSIS.md, nx_item_master_ext.sql, sync_item_master.py
# 코어(nx.item 기존19)=BOM/원가 FK, 업무20컬럼 ADD. 서브=검사/사급/RACK 등. BOM무결성 게이트 내장.
_INSP = {"F": "전수검사", "S": "샘플검사", "N": "무검사"}   # INSP_FLAG (라벨 잠정 — 전산확인)
_ITEM_TYPE = ["원소재", "원자재", "부자재", "서브ASSY", "제품", "완제품"]
# nx.item 관리 컬럼. ★품목마스터는 "품목의 본질"만 관리 — 조달성(매입처/작업장/조달구분)·원가(단가구분)·죽은필드는
# 여기서 제외 → 저장 시 미터치(기존값 보존, 타 프로그램 무영향). 조달은 조달프로파일/BOM, 원가는 원가엔진 소관.
# ★3층 원칙: 품목마스터는 "고정 속성"(이게 뭔가)만. 운영·조달·거래·완제품소속은 전부 분리(저장 미터치, DB 데이터·컬럼 보존).
_IM_BIZ = ["item_status", "pipe_kind", "item_pipe_id"]
_IM_CORE = ["item_name", "item_spec", "item_type", "sgroup", "metal_gubun", "diam", "thick",
            "length", "net_weight", "unit", "status"]
# 분리(저장 미터치, 데이터 보존):
#   조달→sourcing_profile: in_cust, work_code, obtain_gubun, make_type, lg_obtain_flag, pur_lead_time, min_pur_qty
#   원가→원가엔진: cost_gubun / 완제품소속→BOM / 거래→주문: sale_cust, dlvy_except_flag
#   운영정책: safe_stock_qty, kitting_min, set_except_day, sub_mat_flag, sub_mat_wh, prod_rate, proc_gubun, prod_tag,
#            pack_kind, pack_qty, prod_worker, insp_worker / 죽은필드: item_group,item_class,pur_gubun,item_radius,item_pipe_type,item_pipe_material,lgroup,use_gubun
_IM_NUM = {"diam", "thick", "length", "net_weight", "item_pipe_id"}       # decimal
_IM_INT = set()                                                            # smallint (분리로 없음)
_SUB_STR = ["insp_flag", "rack_no", "remarks", "prod_step_memo"]
_SUB_INT = ["pack_qty", "pur_lead_time", "min_pur_qty", "safe_stock_qty"]
_VALVE_F = ["item_od", "item_id", "valve_type", "s_w_type", "h_s_type", "n_s_type", "add_item_type",
            "size1", "size1_limit", "size2", "size2_limit", "size3", "size3_limit", "size4", "size4_limit",
            "size5", "size5_limit", "size6", "size6_limit", "size7", "size7_limit", "size8", "size8_limit"]

# ── 품목 성격(nature) 6그룹 — 소분류 + 공정/BOM 신호로 파생 판정 (실측 24,093건 재분류) ──
_NATURE_MAT = {"210": "1.원소재", "220": "1.원소재",
               "230": "2.부자재/소모품", "910": "2.부자재/소모품", "991": "2.부자재/소모품",
               "992": "2.부자재/소모품", "993": "2.부자재/소모품", "310": "3.사급자재"}
_NATURE_ALL = ["1.원소재", "2.부자재/소모품", "3.사급자재", "4.가공품", "5.용접·조립품", "6.구매·부품"]
# 제품군(대분류) 표시 순서 — 규모/공정흐름 순 (사용자 지정)
_PROD_GROUP_ORDER = ["튜브(절삭단품)", "완제품ASSY", "원소재", "설치자재", "부자재", "서포터", "사급부품", "소모품", "기타"]
# ★파생방식: 코드(접두어+소분류)가 진실, 제품군/제품계열은 조회시 파생(저장 안 함).
_SG_PRODGROUP = {"110": ("완제품ASSY", "완제품ASSY"), "120": ("완제품ASSY", "SUB ASSY"), "130": ("튜브(절삭단품)", "가공품"),
                 "210": ("원소재", "원소재"), "220": ("원소재", "원자재"), "230": ("부자재", "부자재"), "310": ("사급부품", "사급"),
                 "910": ("소모품", "소모품"), "991": ("소모품", "생산소모"), "992": ("소모품", "일반소모"), "993": ("소모품", "수불예외")}

def _load_prefix_map(cur):
    try:
        cur.execute("SELECT prefix, prod_group, prod_line FROM nx.item_prefix_map")
        return {str(r[0]).strip().upper(): (r[1], r[2]) for r in cur.fetchall()}
    except Exception:
        return {}

def _derive_class(code, sgroup, pmap):
    """제품군/제품계열 파생: ①문자접두어→사전 ②치수코드(*)→원소재 ③숫자/기타→소분류매핑 ④else 기타."""
    c = str(code or "").strip()
    m = _re.match(r"^([A-Za-z]+)", c)
    p = m.group(1).upper() if m else ""
    if p in pmap: return (str(pmap[p][0] or ""), str(pmap[p][1] or ""))
    if "*" in c: return ("원소재", "소재컷팅(규격)")
    sg = str(sgroup or "").strip()
    if sg in _SG_PRODGROUP: return _SG_PRODGROUP[sg]
    return ("기타", "기타")
# 그룹별 소프트권장(저장은 되나 경고). 하드필수는 전 그룹 공통: 품번·품명·소분류·단위.
_NATURE_SOFT = {"1.원소재": ["metal_gubun", "diam", "thick", "length", "net_weight"],
                "2.부자재/소모품": ["make_type"], "3.사급자재": ["cost_gubun"],
                "4.가공품": ["lgroup", "make_type", "work_code", "diam", "thick", "length"],
                "5.용접·조립품": ["lgroup", "make_type"], "6.구매·부품": []}
_FIELD_LABEL = {"metal_gubun": "재질", "diam": "외경", "thick": "두께", "length": "길이",
                "net_weight": "중량", "make_type": "생산구분", "cost_gubun": "단가구분",
                "lgroup": "대분류", "work_code": "작업장"}

def _item_nature(cur, code, sgroup):
    """성격 6그룹 + active(정리대상) 판정. 소분류 우선 → 용접/조립(BOM부모) → 가공 → 구매·부품."""
    sg = str(sgroup or "").strip()
    if sg in _NATURE_MAT:
        return _NATURE_MAT[sg], 1
    cur.execute("""SELECT
        (SELECT TOP 1 1 FROM PARTNER_ERP.dbo.CS_T_ITEM_WELD WHERE P_ITEM_CODE=?),
        (SELECT TOP 1 1 FROM PARTNER_ERP.dbo.CS_M_ITEM_BOM WHERE ITEM_CODE=?),
        (SELECT TOP 1 1 FROM PARTNER_ERP.dbo.PR_M_ITEM_PROC_GAGONG WHERE ITEM_CODE=?),
        (SELECT TOP 1 1 FROM PARTNER_ERP.dbo.CS_M_ITEM_BOM WHERE MAT_CODE=?)""", code, code, code, code)
    w, bp, g, bc = cur.fetchone()
    if w or bp: return "5.용접·조립품", 1
    if g: return "4.가공품", 1
    return "6.구매·부품", (1 if bc else 0)   # BOM자식=실부품(1), 완전고아=정리대상(0)

@app.get("/api/itemmaster/opts")
def itemmaster_opts():
    """품목마스터 드롭다운(코드→이름). 코드마스터 + 하드코드 도메인."""
    cn = _conn(); cur = cn.cursor()
    try:
        def lst(kind): return [{"code": k, "nm": v} for k, v in sorted(_kindmap(cur, kind).items()) if k]
        return {
            "lgroup": lst("PR005"), "sgroup": lst("PR006"), "item_group": lst("PR001"),
            "item_class": lst("PR008"), "pipe_kind": lst("PR021"), "unit": lst("CM002"), "metal": lst("PR019"),
            "make_type": [{"code": k, "nm": v} for k, v in _ITEM_MAKE.items() if k],
            "work_code": [{"code": k, "nm": v} for k, v in _ITEM_WORK.items() if k],
            "cost_gubun": [{"code": "1", "nm": "내부단가"}, {"code": "2", "nm": "구매단가"}, {"code": "3", "nm": "소재단가"}, {"code": "5", "nm": "기타"}],
            "insp_flag": [{"code": k, "nm": v} for k, v in _INSP.items()],
            "item_type": [{"code": t, "nm": t} for t in _ITEM_TYPE],
            "status": [{"code": "사용", "nm": "사용"}, {"code": "중지", "nm": "중지"}],
            "yn": [{"code": "1", "nm": "예"}, {"code": "0", "nm": "아니오"}],
            "nature": [{"code": n, "nm": n} for n in _NATURE_ALL],
            "nature_soft": _NATURE_SOFT, "field_label": _FIELD_LABEL,
        }
    finally:
        cn.close()

@app.get("/api/itemmaster/list")
def itemmaster_list(q: str = Query(""), lgroup: str = Query(""), sgroup: str = Query(""),
                    status: str = Query(""), nature: str = Query(""), prod_group: str = Query(""), limit: int = Query(500)):
    """품목마스터 목록(nx.item + item_sub). 코드→이름 디코드. nature=성격6그룹, prod_group=제품군(접두어) 필터."""
    nx = _nx(); cur = nx.cursor()
    cn2 = _conn(); c2 = cn2.cursor()
    try:
        dLG = _kindmap(c2, "PR005"); dSG = _kindmap(c2, "PR006"); dGRP = _kindmap(c2, "PR001")
        dCLS = _kindmap(c2, "PR008"); dPK = _kindmap(c2, "PR021"); dUN = _kindmap(c2, "CM002"); dMT = _kindmap(c2, "PR019")
        c2.execute("SELECT CUST_CODE, ISNULL(CUST_DESC,'') FROM CM_M_CUST")
        dCust = {str(r[0]).strip(): r[1] for r in c2.fetchall()}
        w = ["1=1"]; p = []
        if q.strip(): w.append("(i.item_code LIKE ? OR i.item_name LIKE ?)"); p += [f"%{q.strip()}%"] * 2
        if lgroup.strip(): w.append("i.lgroup=?"); p.append(lgroup.strip())
        if sgroup.strip(): w.append("i.sgroup=?"); p.append(sgroup.strip())
        if status.strip(): w.append("i.status=?"); p.append(status.strip())
        if nature.strip(): w.append("i.nature=?"); p.append(nature.strip())
        if prod_group.strip(): w.append("i.prod_group=?"); p.append(prod_group.strip())
        cur.execute(f"""SELECT TOP {max(1,min(int(limit),3000))} i.item_code,i.item_name,i.item_spec,i.item_type,
              i.lgroup,i.sgroup,i.item_group,i.item_class,i.pipe_kind,i.unit,i.metal_gubun,i.in_cust,i.work_code,
              i.make_type,i.cost_gubun,i.item_status,i.status,i.diam,i.thick,i.length,i.net_weight,i.item_pipe_id,
              i.prod_rate,i.sub_mat_flag, s.insp_flag, s.rack_no, i.nature, i.active, i.prod_group, i.prod_line
            FROM nx.item i LEFT JOIN nx.item_sub s ON s.item_code=i.item_code
            WHERE {' AND '.join(w)} ORDER BY i.item_code""", *p)
        rows = []
        for r in cur.fetchall():
            g = lambda i: str(r[i] or "").strip()
            num = lambda i: (float(r[i]) if r[i] is not None else None)
            rows.append({
                "item_code": g(0), "item_name": g(1), "item_spec": g(2), "item_type": g(3),
                "lgroup": dLG.get(g(4), g(4)), "sgroup": dSG.get(g(5), g(5)), "item_group": dGRP.get(g(6), g(6)),
                "item_class": dCLS.get(g(7), g(7)), "pipe_kind": dPK.get(g(8), g(8)), "unit": dUN.get(g(9), g(9)),
                "metal": dMT.get(g(10), g(10)), "in_cust": dCust.get(g(11), g(11)), "work": _ITEM_WORK.get(g(12), g(12)),
                "make_type": _ITEM_MAKE.get(g(13), g(13)), "cost_gubun": g(14), "item_status": g(15),
                "status": g(16) or "사용", "diam": num(17), "thick": num(18), "length": num(19),
                "net_weight": num(20), "pipe_id": num(21), "prod_rate": num(22), "sub_mat_flag": g(23),
                "insp_flag": _INSP.get(g(24), g(24)), "rack_no": g(25),
                "nature": g(26), "active": (1 if (r[27] is None or r[27]) else 0),
                "prod_group": g(28), "prod_line": g(29)})
        cur.execute("SELECT DISTINCT prod_group FROM nx.item WHERE prod_group IS NOT NULL")
        pgs = [str(r[0]).strip() for r in cur.fetchall() if r[0]]
        pgs = sorted(pgs, key=lambda g: _PROD_GROUP_ORDER.index(g) if g in _PROD_GROUP_ORDER else 99)
        return {"rows": rows, "cnt": len(rows),
                "lgroups": [{"code": k, "nm": v} for k, v in sorted(dLG.items()) if k],
                "sgroups": [{"code": k, "nm": v} for k, v in sorted(dSG.items()) if k],
                "natures": [{"code": n, "nm": n} for n in _NATURE_ALL],
                "prod_groups": [{"code": g, "nm": g} for g in pgs]}
    finally:
        nx.close(); cn2.close()

@app.get("/api/itemmaster/get")
def itemmaster_get(item: str = Query(...)):
    """단일 품목 전체(편집 모달용): nx.item 코어+업무 + item_sub + item_valve(설치품)."""
    code = item.strip()
    nx = _nx(); cur = nx.cursor()
    try:
        allc = ["item_code"] + _IM_CORE + [c for c in _IM_BIZ]
        cur.execute(f"SELECT {','.join(allc)},silver_flag,has_gagong,nature,active,prod_group,prod_line FROM nx.item WHERE item_code=?", code)
        r = cur.fetchone()
        if not r: raise HTTPException(404, "품목 없음")
        cols = [d[0] for d in cur.description]
        item_d = {}
        for k, v in zip(cols, r):
            if isinstance(v, __import__("decimal").Decimal): v = float(v)
            item_d[k] = ("" if v is None else v)
        cur.execute(f"SELECT {','.join(_SUB_STR + _SUB_INT)} FROM nx.item_sub WHERE item_code=?", code)
        rs = cur.fetchone()
        sub_d = dict(zip(_SUB_STR + _SUB_INT, ["" if x is None else x for x in rs])) if rs else {}
        cur.execute(f"SELECT {','.join(_VALVE_F)} FROM nx.item_valve WHERE item_code=?", code)
        rv = cur.fetchone()
        valve_d = dict(zip(_VALVE_F, ["" if x is None else x for x in rv])) if rv else {}
        return {"item": item_d, "sub": sub_d, "valve": valve_d, "has_valve": bool(rv)}
    finally:
        nx.close()

@app.post("/api/itemmaster/save")
def itemmaster_save(payload: dict = Body(...)):
    """품목마스터 등록/수정(nx.item + item_sub + item_valve).
    검증: 필수·공백·품번중복·매입처/작업장 배타. 자동: 내경(diam-thick*2)·make_type4→LG사급.
    품번변경(edit+코드변경): nx.bom/서브 연쇄 + item_his 이력 (BOM 무결성)."""
    p = payload
    code = str(p.get("item_code", "") or "").strip()
    name = str(p.get("item_name", "") or "").strip()
    if not code or not name:
        raise HTTPException(400, "품번·품명은 필수입니다.")
    if code != code.strip() or "  " in code or code != "".join(code.split(" ")) and (code[0] == " " or code[-1] == " "):
        raise HTTPException(400, "품번 앞/뒤에 공백은 사용할 수 없습니다.")
    for k in ("sgroup", "unit"):   # 하드필수(전 그룹 공통): 품번·품명·소분류·단위. 대분류는 성격별 소프트권장으로 이동.
        if not str(p.get(k, "") or "").strip():
            raise HTTPException(400, "소분류·단위는 필수입니다.")
    in_cust = str(p.get("in_cust", "") or "").strip()
    work_code = str(p.get("work_code", "") or "").strip()
    if in_cust and work_code:
        raise HTTPException(400, "매입처(업체)와 작업장은 둘 중 하나만 입력 가능합니다.")

    def sval(k, n=None):
        v = p.get(k)
        if v in (None, ""): return None
        v = str(v).strip()
        return v[:n] if n else v
    def dval(k):
        v = p.get(k)
        try: return float(v) if v not in (None, "") else None
        except Exception: return None
    def ival(k):
        v = p.get(k)
        try: return int(float(v)) if v not in (None, "") else None
        except Exception: return None

    # 내경 자동(외경-두께*2)
    diam, thick = dval("diam"), dval("thick")
    if diam is not None and thick is not None and not p.get("item_pipe_id"):
        p["item_pipe_id"] = round(diam - thick * 2, 4)
    # make_type=4 → LG사급 자동
    if str(p.get("make_type", "")).strip() == "4":
        p["lg_obtain_flag"] = "1"

    is_edit = bool(p.get("_edit"))
    orig = str(p.get("_orig_code", "") or "").strip() or code
    user = (str(p.get("user", "") or "").strip() or "웹사용자")[:20]

    nx = _nx(); nx.autocommit = False; cur = nx.cursor()
    try:
        cur.execute("SELECT 1 FROM nx.item WHERE item_code=?", code)
        exists = cur.fetchone() is not None
        if not is_edit and exists:
            raise HTTPException(400, "동일한 품번이 이미 등록되어 있습니다.")

        # ── 품번 변경(rename): copy→repoint(BOM/서브)→delete old + 이력 ──
        if is_edit and orig and orig != code:
            if exists:
                raise HTTPException(400, "변경할 품번이 이미 존재합니다.")
            cur.execute("""INSERT INTO nx.item(item_code,item_name,item_type,unit,status,silver_flag,has_gagong)
                SELECT ?,item_name,item_type,unit,status,silver_flag,has_gagong FROM nx.item WHERE item_code=?""", code, orig)
            cur.execute("UPDATE nx.item_sub   SET item_code=? WHERE item_code=?", code, orig)
            cur.execute("UPDATE nx.item_valve SET item_code=? WHERE item_code=?", code, orig)
            cur.execute("UPDATE nx.bom_header SET item_code=? WHERE item_code=?", code, orig)
            cur.execute("UPDATE nx.bom_line   SET child_item=? WHERE child_item=?", code, orig)
            cur.execute("DELETE FROM nx.item WHERE item_code=?", orig)
            cur.execute("INSERT INTO nx.item_his(old_code,new_code,change_dt,user_id) VALUES(?,?,getdate(),?)", orig, code, user)
            exists = True   # 새 코드 행 생성됨

        # ── 성격(품목유형) + 제품군/제품계열 자동파생(코드→이름) + 소프트권장 경고 ──
        sgroup_v = str(p.get("sgroup", "") or "").strip()
        nature, active = _item_nature(cur, code, sgroup_v)
        prod_group, prod_line = _derive_class(code, sgroup_v, _load_prefix_map(cur))   # 파생: 접두어+소분류
        warnings = []
        for f in _NATURE_SOFT.get(nature, []):
            miss = (dval(f) in (None, 0) or dval(f) == 0.0) if f in _IM_NUM else (not str(p.get(f, "") or "").strip())
            if miss: warnings.append(_FIELD_LABEL.get(f, f))

        # ── nx.item 코어+업무 upsert (NOT NULL 컬럼은 미전송 시 기존값 보존/기본값) ──
        _DEF = {"item_type": "제품", "unit": "EA", "status": "사용"}   # NOT NULL 컬럼 기본값
        allcols = _IM_CORE + _IM_BIZ
        rawvals = [dval(c) if c in _IM_NUM else (ival(c) if c in _IM_INT else sval(c, 200)) for c in allcols]
        if exists:
            setcols = [f"{c}=ISNULL(?,{c})" if c in _DEF else f"{c}=?" for c in allcols] + ["nature=?", "active=?", "prod_group=?", "prod_line=?"]
            cur.execute(f"UPDATE nx.item SET {','.join(setcols)} WHERE item_code=?", *rawvals, nature, active, prod_group, prod_line, code)
        else:
            ivals = [(v if v is not None else _DEF[c]) if c in _DEF else v for c, v in zip(allcols, rawvals)]
            cur.execute(f"""INSERT INTO nx.item(item_code,{','.join(allcols)},silver_flag,has_gagong,nature,active,prod_group,prod_line)
                VALUES(?,{','.join(['?']*len(allcols))},0,0,?,?,?,?)""", code, *ivals, nature, active, prod_group, prod_line)

        # ── item_sub upsert (delete→insert) ──
        cur.execute("DELETE FROM nx.item_sub WHERE item_code=?", code)
        sub_vals = [sval(k, 500) for k in _SUB_STR] + [ival(k) for k in _SUB_INT]
        if any(v not in (None, "") for v in sub_vals):
            cur.execute(f"INSERT INTO nx.item_sub(item_code,{','.join(_SUB_STR + _SUB_INT)}) "
                        f"VALUES(?,{','.join(['?']*len(_SUB_STR + _SUB_INT))})", code, *sub_vals)

        # ── item_valve upsert (설치품, 값 있을 때만) ──
        cur.execute("DELETE FROM nx.item_valve WHERE item_code=?", code)
        vv = [sval(k, 400) for k in _VALVE_F]
        if any(v not in (None, "") for v in vv):
            cur.execute(f"INSERT INTO nx.item_valve(item_code,{','.join(_VALVE_F)}) "
                        f"VALUES(?,{','.join(['?']*len(_VALVE_F))})", code, *vv)

        nx.commit()
        return {"ok": True, "mode": ("update" if is_edit else "insert"), "item_code": code,
                "renamed": (is_edit and orig != code), "nature": nature, "warnings": warnings}
    except Exception:
        nx.rollback(); raise
    finally:
        nx.close()

@app.post("/api/itemmaster/delete")
def itemmaster_delete(payload: dict = Body(...)):
    """삭제 — BOM 무결성 게이트: nx.bom(모/자) + 레거시 PR_M_ITEM_BOM(모/자) 참조 있으면 거부."""
    codes = [str(x).strip() for x in (payload.get("codes", []) or []) if str(x).strip()]
    if not codes: return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        blocked = []
        for code in codes:
            cur.execute("SELECT 1 FROM nx.bom_header WHERE item_code=?", code); a = cur.fetchone()
            cur.execute("SELECT 1 FROM nx.bom_line WHERE child_item=?", code); b = cur.fetchone()
            cur.execute("SELECT TOP 1 1 FROM PARTNER_ERP.dbo.PR_M_ITEM_BOM WHERE ITEM_CODE=? OR MAT_CODE=?", code, code)
            c = cur.fetchone()
            if a or b or c: blocked.append(code)
        if blocked:
            return {"ok": False, "errors": [f"{c} : BOM에 사용중이라 삭제할 수 없습니다." for c in blocked]}
        for code in codes:
            cur.execute("DELETE FROM nx.item_sub WHERE item_code=?", code)
            cur.execute("DELETE FROM nx.item_valve WHERE item_code=?", code)
            cur.execute("DELETE FROM nx.item WHERE item_code=?", code)
        return {"ok": True, "deleted": len(codes)}
    finally:
        nx.close()


# ============ 원소재 마스터 + 5가격 통합 뷰 (nx.raw_material ↔ price_metal/price_item/price_lme_lg) ============
_MATCODE = {"구리": "CU", "고강도관": "고강도", "알루미늄": "AL", "STS": "STS"}
@app.get("/api/rawmat/list")
def rawmat_list(q: str = Query(""), material: str = Query("")):
    """원소재 마스터(외경×두께×재질×조질) + 5가격: 시세·파트너(사급)·매입·매출·LG인증. 규격/재질/품번 조인."""
    nx = _nx(); cur = nx.cursor()
    try:
        # 원소재 마스터
        w = ["1=1"]; p = []
        if material.strip(): w.append("material=?"); p.append(material.strip())
        if q.strip(): w.append("(CAST(outer_diam AS varchar)+'x'+CAST(thickness AS varchar) LIKE ? OR part_no LIKE ? OR src_codes LIKE ?)"); p += [f"%{q.strip()}%"] * 3
        cur.execute(f"""SELECT raw_id, material, outer_diam, thickness, temper, part_no, unit, std_length,
              y2026_kg, src_codes, vendors FROM nx.raw_material WHERE {' AND '.join(w)}
              ORDER BY material, outer_diam, thickness""", *p)
        rms = cur.fetchall()
        # price_metal 규격별 최신 시세/파트너
        cur.execute("""SELECT metal_gubun, diam, thick, std_price, partner_price FROM nx.price_metal pm
              WHERE apply_ym = (SELECT MAX(apply_ym) FROM nx.price_metal x WHERE x.metal_gubun=pm.metal_gubun AND x.diam=pm.diam AND x.thick=pm.thick)""")
        pmet = {}
        for r in cur.fetchall():
            pmet[(str(r[0]).strip(), float(r[1]) if r[1] is not None else None, float(r[2]) if r[2] is not None else None)] = (r[3], r[4])
        # LG 인증가 (재질/월)
        cur.execute("SELECT metal_gubun, lg_recog_price FROM nx.price_lme_lg pl WHERE apply_ym=(SELECT MAX(apply_ym) FROM nx.price_lme_lg x WHERE x.metal_gubun=pl.metal_gubun)")
        plme = {str(r[0]).strip(): r[1] for r in cur.fetchall()}
        # 품번별 최신 매입/사급/매출 (price_item)
        cur.execute("""SELECT item_code, price_type, price FROM nx.price_item pi
              WHERE apply_ymd=(SELECT MAX(apply_ymd) FROM nx.price_item x WHERE x.item_code=pi.item_code AND x.price_type=pi.price_type)""")
        pit = {}
        for r in cur.fetchall():
            pit.setdefault(str(r[0]).strip(), {})[str(r[1]).strip()] = r[2]
        rows = []
        for r in rms:
            mat = str(r[1] or "").strip(); od = float(r[2]) if r[2] is not None else None; th = float(r[3]) if r[3] is not None else None
            sise, partner = pmet.get((_MATCODE.get(mat, mat), od, th), (None, None))
            # 매입/사급/매출 = 원천품번들 중 대표(최대) 값
            codes = [c.strip() for c in str(r[9] or "").split(",") if c.strip()]
            buy = sale = sagub = None
            for c in codes:
                d = pit.get(c, {})
                for k, dst in (("매입", "buy"), ("TAGS", "sagub"), ("TAGE", "sale")):
                    v = d.get(k)
                    if v and float(v) > 0:
                        if dst == "buy": buy = max(buy or 0, float(v))
                        elif dst == "sagub": sagub = max(sagub or 0, float(v))
                        else: sale = max(sale or 0, float(v))
            rows.append({
                "raw_id": r[0], "material": mat, "outer_diam": od, "thickness": th,
                "temper": r[4] or "", "part_no": r[5] or "", "unit": r[6] or "KG", "std_length": r[7],
                "y2026_kg": float(r[8]) if r[8] is not None else 0, "codes_cnt": len(codes), "vendors": r[10] or "",
                "price_sise": float(sise) if sise is not None else None,     # 직거래 시세
                "price_partner": float(partner) if partner is not None else None,  # 파트너(사급)가
                "price_buy": buy, "price_sale": sale, "price_sagub_item": sagub,
                "price_lg_recog": float(plme[_MATCODE.get(mat, mat)]) if _MATCODE.get(mat, mat) in plme else None})
        return {"rows": rows, "cnt": len(rows),
                "materials": [{"code": m, "nm": m} for m in ["구리", "고강도관", "알루미늄", "STS"]]}
    finally:
        nx.close()


@app.get("/api/rawmat/prices")
def rawmat_prices(raw_id: int = Query(...)):
    """선택 원소재의 월별 단가 시계열: LG인증가(price_lme_lg)·LG사급가(rawmat_lg_sagub, 입력값 미입력=0)·현물가/협력사사급가(price_metal). 최신월 상단."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT material, outer_diam, thickness, temper, part_no FROM nx.raw_material WHERE raw_id=?", raw_id)
        r = cur.fetchone()
        if not r:
            return {"raw_id": raw_id, "material": "", "spec": "", "rows": []}
        mat = str(r[0] or "").strip(); od = r[1]; th = r[2]
        mg = _MATCODE.get(mat, mat)
        series = {}
        cur.execute("SELECT apply_ym, std_price, partner_price FROM nx.price_metal WHERE metal_gubun=? AND diam=? AND thick=?", mg, od, th)
        for ym, sp, pp in cur.fetchall():
            d = series.setdefault(str(ym).strip(), {})
            d["sise"] = float(sp) if sp is not None else None
            d["partner"] = float(pp) if pp is not None else None
        cur.execute("SELECT apply_ym, lg_recog_price FROM nx.price_lme_lg WHERE metal_gubun=?", mg)
        for ym, v in cur.fetchall():
            series.setdefault(str(ym).strip(), {})["lg_recog"] = float(v) if v is not None else None
        cur.execute("SELECT apply_ym, price FROM nx.rawmat_lg_sagub WHERE raw_id=?", raw_id)
        for ym, v in cur.fetchall():
            series.setdefault(str(ym).strip(), {})["lg_sagub"] = float(v) if v is not None else 0
        rows = []
        for ym in sorted(series.keys(), reverse=True):
            d = series[ym]
            rows.append({"ym": ym, "lg_recog": d.get("lg_recog"), "lg_sagub": d.get("lg_sagub", 0),
                         "sise": d.get("sise"), "partner": d.get("partner")})
        spec = f"⌀{od}×{th}" + (f" {r[3]}" if r[3] else "")
        return {"raw_id": raw_id, "material": mat, "spec": spec, "part_no": r[4] or "", "rows": rows}
    finally:
        nx.close()


@app.post("/api/rawmat/lg_sagub/save")
def rawmat_lg_sagub_save(payload: dict = Body(...)):
    """원소재별 월별 LG사급가 저장(upsert). rows=[{ym,price}]. 미입력월은 0."""
    raw_id = int(payload.get("raw_id") or 0)
    rows = payload.get("rows", []) or []
    if not raw_id:
        return {"ok": False, "error": "raw_id 필요"}
    nx = _nx(); cur = nx.cursor()
    try:
        n = 0
        for r in rows:
            ym = "".join(ch for ch in str(r.get("ym", "")) if ch.isdigit())[:6]
            if len(ym) != 6:
                continue
            price = float(r.get("price") or 0)
            cur.execute("""MERGE nx.rawmat_lg_sagub AS t USING (SELECT ? raw_id, ? apply_ym) s
                ON t.raw_id=s.raw_id AND t.apply_ym=s.apply_ym
                WHEN MATCHED THEN UPDATE SET price=?, upd_user='web', upd_dt=GETDATE()
                WHEN NOT MATCHED THEN INSERT(raw_id,apply_ym,price,upd_user,upd_dt) VALUES(?,?,?,'web',GETDATE());""",
                raw_id, ym, price, raw_id, ym, price)
            n += 1
        return {"ok": True, "count": n}
    finally:
        nx.close()


# ============ LG전자 LME인정가 (Cost Table 직거래) — 재료비=f(LME,환율)·가공비=국가믹스 · nx.lg_lme_* ============
def _ensure_lglme_tbl(cur):
    cur.execute("""IF OBJECT_ID('nx.lg_lme_header','U') IS NULL CREATE TABLE nx.lg_lme_header(
        apply_ym CHAR(6) PRIMARY KEY, cu_lme FLOAT, brass_lme FLOAT, cable_lme FLOAT,
        fx_now FLOAT, fx_prev FLOAT, premium FLOAT DEFAULT 152, surcharge FLOAT DEFAULT 1.05,
        upd_user NVARCHAR(30), upd_dt datetime DEFAULT getdate())""")
    cur.execute("""IF OBJECT_ID('nx.lg_lme_gagong','U') IS NULL CREATE TABLE nx.lg_lme_gagong(
        apply_ym CHAR(6), gubun NVARCHAR(20), diam FLOAT, thick FLOAT,
        vn_gagong FLOAT, vn_prem FLOAT, vn_mul FLOAT, vn_naeryuk FLOAT,
        cn_gagong FLOAT, cn_prem FLOAT, cn_mul FLOAT, cn_naeryuk FLOAT,
        duty_vn FLOAT DEFAULT 0, duty_cn FLOAT DEFAULT 0.016, mix_cn FLOAT DEFAULT 0.3, mix_vn FLOAT DEFAULT 0.7,
        CONSTRAINT PK_lglme_gagong PRIMARY KEY(apply_ym,gubun,diam,thick))""")
    cur.execute("""IF OBJECT_ID('nx.lg_lme_costtable','U') IS NULL CREATE TABLE nx.lg_lme_costtable(
        id INT IDENTITY(1,1) PRIMARY KEY, apply_ym CHAR(6), gubun NVARCHAR(20), diam FLOAT, thick FLOAT,
        p_no NVARCHAR(30), jaeryo FLOAT, gagong FLOAT, wonjae FLOAT, seq INT)""")

def _lglme_metal_lme(gubun, h):
    """구분→재질 LME: 황동/Cable 키워드면 해당, 그외 Cu(선물). h=header dict."""
    g = str(gubun or "")
    if "황동" in g: return float(h.get("brass_lme") or 0)
    if "Cable" in g or "케이블" in g: return float(h.get("cable_lme") or 0)
    return float(h.get("cu_lme") or 0)

def _lglme_jaeryo(gubun, h):
    """재료비(원/kg): 직관&P/C=(LME+premium)×surcharge×fx/1000, 그외=LME×fx/1000."""
    lme = _lglme_metal_lme(gubun, h); fx = float(h.get("fx_now") or 0)
    g = str(gubun or "")
    if "직관" in g or "P/C" in g:
        return round((lme + float(h.get("premium") or 0)) * float(h.get("surcharge") or 1.0) * fx / 1000, 2)
    return round(lme * fx / 1000, 2)

def _lglme_gagong_won(gr, lme, fx):
    """가공비(원) = (중국Price×mix_cn + 베트남Price×mix_vn)×fx/1000. 국가Price=가공비+프리미엄+물류+관세+내륙(USD)."""
    vn = (gr["vn_gagong"] + gr["vn_prem"] + gr["vn_mul"]
          + (lme + gr["vn_gagong"] + gr["vn_mul"]) * gr["duty_vn"] + gr["vn_naeryuk"])
    cn = (gr["cn_gagong"] + gr["cn_prem"] + gr["cn_mul"]
          + (lme + gr["cn_gagong"] + gr["cn_mul"]) * gr["duty_cn"] + gr["cn_naeryuk"])
    return round((cn * gr["mix_cn"] + vn * gr["mix_vn"]) * fx / 1000, 2)

@app.get("/api/lglme/months")
def lglme_months():
    """LME인정가 적용월 목록(최신순) + 편집가능(header 존재) 여부."""
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("""SELECT c.apply_ym, COUNT(*) nrow, MAX(CASE WHEN h.apply_ym IS NULL THEN 0 ELSE 1 END) editable
            FROM nx.lg_lme_costtable c LEFT JOIN nx.lg_lme_header h ON h.apply_ym=c.apply_ym
            GROUP BY c.apply_ym ORDER BY c.apply_ym DESC""")
        return {"rows": [{"apply_ym": r[0], "n_row": int(r[1]), "editable": bool(r[2])} for r in cur.fetchall()]}
    finally:
        nx.close()

@app.get("/api/lglme/table")
def lglme_table(ym: str = Query(...)):
    """월별 Cost Table: header(입력값) + costtable 행(구분·외경·두께·P/No·재료비·가공비·원재료가). header 없으면 과거값 읽기전용."""
    ym = ym.strip()[:6]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("""SELECT apply_ym,cu_lme,brass_lme,cable_lme,fx_now,fx_prev,premium,surcharge FROM nx.lg_lme_header WHERE apply_ym=?""", ym)
        h = cur.fetchone()
        header = None
        if h:
            header = {"apply_ym": h[0], "cu_lme": h[1], "brass_lme": h[2], "cable_lme": h[3],
                      "fx_now": h[4], "fx_prev": h[5], "premium": h[6], "surcharge": h[7]}
        cur.execute("""SELECT id,gubun,diam,thick,ISNULL(p_no,''),jaeryo,gagong,wonjae,seq
            FROM nx.lg_lme_costtable WHERE apply_ym=? ORDER BY seq,id""", ym)
        rows = [{"id": r[0], "gubun": r[1], "diam": r[2], "thick": r[3], "p_no": r[4],
                 "jaeryo": r[5], "gagong": r[6], "wonjae": r[7], "seq": r[8]} for r in cur.fetchall()]
        return {"ym": ym, "header": header, "editable": bool(header), "rows": rows}
    finally:
        nx.close()

@app.get("/api/lglme/gagong")
def lglme_gagong(ym: str = Query(...)):
    """월별 국가믹스 원천(spec별 베트남·중국 가공비/프리미엄/물류/내륙 + 관세율·믹스)."""
    ym = ym.strip()[:6]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("""SELECT gubun,diam,thick,vn_gagong,vn_prem,vn_mul,vn_naeryuk,cn_gagong,cn_prem,cn_mul,cn_naeryuk,
            duty_vn,duty_cn,mix_cn,mix_vn FROM nx.lg_lme_gagong WHERE apply_ym=? ORDER BY gubun,diam,thick""", ym)
        cols = ["gubun", "diam", "thick", "vn_gagong", "vn_prem", "vn_mul", "vn_naeryuk",
                "cn_gagong", "cn_prem", "cn_mul", "cn_naeryuk", "duty_vn", "duty_cn", "mix_cn", "mix_vn"]
        return {"ym": ym, "rows": [dict(zip(cols, r)) for r in cur.fetchall()]}
    finally:
        nx.close()

@app.post("/api/lglme/header/save")
def lglme_header_save(payload: dict = Body(...)):
    """헤더(재질 LME·환율·premium·할증) upsert. 필수=apply_ym·cu_lme·fx_now."""
    p = payload; ym = str(p.get("apply_ym", "")).strip()[:6]
    if len(ym) != 6: raise HTTPException(400, "apply_ym(YYYYMM) 필요")
    def f(k, d=0.0):
        try: return float(p.get(k))
        except Exception: return d
    if f("cu_lme") <= 0: return {"ok": False, "errors": ["Cu 적용 LME는 필수(>0)"]}
    if f("fx_now") <= 0: return {"ok": False, "errors": ["적용환율은 필수(>0)"]}
    usr = (str(p.get("user", "")).strip() or "web")[:30]
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("""MERGE nx.lg_lme_header AS t USING (SELECT ? apply_ym) s ON t.apply_ym=s.apply_ym
            WHEN MATCHED THEN UPDATE SET cu_lme=?,brass_lme=?,cable_lme=?,fx_now=?,fx_prev=?,premium=?,surcharge=?,upd_user=?,upd_dt=getdate()
            WHEN NOT MATCHED THEN INSERT(apply_ym,cu_lme,brass_lme,cable_lme,fx_now,fx_prev,premium,surcharge,upd_user)
              VALUES(?,?,?,?,?,?,?,?,?);""",
            ym, f("cu_lme"), (f("brass_lme") or f("cu_lme")), (f("cable_lme") or f("cu_lme")), f("fx_now"), f("fx_prev"),
            (f("premium") or 152.0), (f("surcharge") or 1.05), usr,
            ym, f("cu_lme"), (f("brass_lme") or f("cu_lme")), (f("cable_lme") or f("cu_lme")), f("fx_now"), f("fx_prev"),
            (f("premium") or 152.0), (f("surcharge") or 1.05), usr)
        return {"ok": True, "apply_ym": ym}
    finally:
        nx.close()

@app.post("/api/lglme/gagong/save")
def lglme_gagong_save(payload: dict = Body(...)):
    """국가믹스 원천 1행 upsert(spec별). 필수=apply_ym·gubun·diam·thick."""
    p = payload; ym = str(p.get("apply_ym", "")).strip()[:6]
    gub = str(p.get("gubun", "")).strip()[:20]
    try: diam = float(p.get("diam")); thick = float(p.get("thick"))
    except Exception: raise HTTPException(400, "diam·thick 필요")
    if len(ym) != 6 or not gub: raise HTTPException(400, "apply_ym·gubun 필요")
    def f(k, d=0.0):
        try: return float(p.get(k))
        except Exception: return d
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("""MERGE nx.lg_lme_gagong AS t USING (SELECT ? a,? g,? d,? th) s
            ON t.apply_ym=s.a AND t.gubun=s.g AND t.diam=s.d AND t.thick=s.th
            WHEN MATCHED THEN UPDATE SET vn_gagong=?,vn_prem=?,vn_mul=?,vn_naeryuk=?,cn_gagong=?,cn_prem=?,cn_mul=?,cn_naeryuk=?,duty_vn=?,duty_cn=?,mix_cn=?,mix_vn=?
            WHEN NOT MATCHED THEN INSERT(apply_ym,gubun,diam,thick,vn_gagong,vn_prem,vn_mul,vn_naeryuk,cn_gagong,cn_prem,cn_mul,cn_naeryuk,duty_vn,duty_cn,mix_cn,mix_vn)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);""",
            ym, gub, diam, thick,
            f("vn_gagong"), f("vn_prem"), f("vn_mul"), f("vn_naeryuk"), f("cn_gagong"), f("cn_prem"), f("cn_mul"), f("cn_naeryuk"), f("duty_vn"), f("duty_cn", 0.016), f("mix_cn", 0.3), f("mix_vn", 0.7),
            ym, gub, diam, thick,
            f("vn_gagong"), f("vn_prem"), f("vn_mul"), f("vn_naeryuk"), f("cn_gagong"), f("cn_prem"), f("cn_mul"), f("cn_naeryuk"), f("duty_vn"), f("duty_cn", 0.016), f("mix_cn", 0.3), f("mix_vn", 0.7))
        return {"ok": True}
    finally:
        nx.close()

@app.post("/api/lglme/recompute")
def lglme_recompute(payload: dict = Body(...)):
    """월 Cost Table 재계산: 재료비(구분별 산식)+가공비(국가믹스, spec매칭)+원재료가(합) → costtable upsert. header 필수."""
    ym = str(payload.get("ym", "")).strip()[:6]
    if len(ym) != 6: raise HTTPException(400, "ym 필요")
    nx = _nx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("SELECT cu_lme,brass_lme,cable_lme,fx_now,fx_prev,premium,surcharge FROM nx.lg_lme_header WHERE apply_ym=?", ym)
        h = cur.fetchone()
        if not h: return {"ok": False, "errors": ["헤더(LME/환율)가 없습니다 — 먼저 저장하세요"]}
        H = {"cu_lme": h[0], "brass_lme": h[1], "cable_lme": h[2], "fx_now": h[3], "fx_prev": h[4], "premium": h[5], "surcharge": h[6]}
        cur.execute("""SELECT gubun,diam,thick,vn_gagong,vn_prem,vn_mul,vn_naeryuk,cn_gagong,cn_prem,cn_mul,cn_naeryuk,duty_vn,duty_cn,mix_cn,mix_vn
            FROM nx.lg_lme_gagong WHERE apply_ym=?""", ym)
        gk = ["vn_gagong", "vn_prem", "vn_mul", "vn_naeryuk", "cn_gagong", "cn_prem", "cn_mul", "cn_naeryuk", "duty_vn", "duty_cn", "mix_cn", "mix_vn"]
        gmap = {}
        for r in cur.fetchall():
            gmap[(str(r[0]).strip(), float(r[1]), float(r[2]))] = {k: float(r[3 + i] or 0) for i, k in enumerate(gk)}
        cur.execute("SELECT id,gubun,diam,thick,jaeryo,gagong FROM nx.lg_lme_costtable WHERE apply_ym=?", ym)
        recs = cur.fetchall(); n = 0
        for cid, gub, diam, thick, ojae, ogag in recs:
            jae = _lglme_jaeryo(gub, H)
            gr = gmap.get((str(gub).strip(), float(diam or 0), float(thick or 0)))
            lme = _lglme_metal_lme(gub, H)
            gag = _lglme_gagong_won(gr, lme, float(H["fx_now"] or 0)) if gr else float(ogag or 0)
            cur.execute("UPDATE nx.lg_lme_costtable SET jaeryo=?,gagong=?,wonjae=? WHERE id=?", jae, gag, round(jae + gag, 2), cid)
            n += 1
        return {"ok": True, "updated": n}
    finally:
        nx.close()

@app.post("/api/lglme/copy")
def lglme_copy(payload: dict = Body(...)):
    """전월(from_ym) → 신규월(to_ym) 복사: header+gagong+costtable 복제(to_ym 기존 근거범위 제거 후). 이후 LME/환율/국가단가 갱신·재계산."""
    fr = str(payload.get("from_ym", "")).strip()[:6]; to = str(payload.get("to_ym", "")).strip()[:6]
    usr = (str(payload.get("user", "")).strip() or "web")[:30]
    if len(fr) != 6 or len(to) != 6: raise HTTPException(400, "from_ym·to_ym(YYYYMM) 필요")
    if fr == to: raise HTTPException(400, "원본월과 신규월이 같습니다")
    nx = _nx_tx(); cur = nx.cursor()
    try:
        _ensure_lglme_tbl(cur)
        cur.execute("SELECT 1 FROM nx.lg_lme_costtable WHERE apply_ym=?", to)
        if cur.fetchone(): raise HTTPException(409, f"{to} 는 이미 존재합니다(중복 생성 방지)")
        cur.execute("""INSERT INTO nx.lg_lme_header(apply_ym,cu_lme,brass_lme,cable_lme,fx_now,fx_prev,premium,surcharge,upd_user)
            SELECT ?,cu_lme,brass_lme,cable_lme,fx_now,fx_prev,premium,surcharge,? FROM nx.lg_lme_header WHERE apply_ym=?""", to, usr, fr)
        cur.execute("""INSERT INTO nx.lg_lme_gagong(apply_ym,gubun,diam,thick,vn_gagong,vn_prem,vn_mul,vn_naeryuk,cn_gagong,cn_prem,cn_mul,cn_naeryuk,duty_vn,duty_cn,mix_cn,mix_vn)
            SELECT ?,gubun,diam,thick,vn_gagong,vn_prem,vn_mul,vn_naeryuk,cn_gagong,cn_prem,cn_mul,cn_naeryuk,duty_vn,duty_cn,mix_cn,mix_vn FROM nx.lg_lme_gagong WHERE apply_ym=?""", to, fr)
        cur.execute("""INSERT INTO nx.lg_lme_costtable(apply_ym,gubun,diam,thick,p_no,jaeryo,gagong,wonjae,seq)
            SELECT ?,gubun,diam,thick,p_no,jaeryo,gagong,wonjae,seq FROM nx.lg_lme_costtable WHERE apply_ym=?""", to, fr)
        cur.execute("SELECT COUNT(*) FROM nx.lg_lme_costtable WHERE apply_ym=?", to)
        n = cur.fetchone()[0]
        nx.commit()
        return {"ok": True, "to_ym": to, "rows": n}
    except Exception:
        nx.rollback(); raise
    finally:
        nx.close()


# ============ 직거래 LME 월연동 판가 자동정본화 (w_tc_master_165/090 수작업 자동화) ============
# 산식: 판가(월)=base_item_cost + 동소요량×(LME월 − base_LME). 직거래LME만 재계산, 사급정체는 base 고정.
# 라이브 PR_M_ITEM_COST 읽기전용(대사). nx.dtrade_price(마스터)/dtrade_lme_index(월지수)/dtrade_price_ts(계산결과).
@app.get("/api/dtrade/lme")
def dtrade_lme():
    """월별 직거래 LME index(원/kg)."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT apply_ym, lme_index FROM nx.dtrade_lme_index ORDER BY apply_ym DESC")
        return {"rows": [{"apply_ym": r[0], "lme_index": float(r[1] or 0)} for r in cur.fetchall()]}
    finally:
        nx.close()

@app.get("/api/dtrade/summary")
def dtrade_summary():
    """linkage별 대상 건수(직거래LME/사급정체) + 동소요량 src·사급자재 분포."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""SELECT ISNULL(linkage,''), COUNT(*), SUM(CASE WHEN qty_src='LG392' THEN 1 ELSE 0 END),
              SUM(CASE WHEN qty_src='역산' THEN 1 ELSE 0 END), SUM(CAST(ISNULL(sagub_flag,0) AS int))
            FROM nx.dtrade_price GROUP BY linkage""")
        rows = [{"linkage": r[0], "cnt": r[1], "lg392": r[2], "inv": r[3], "sagub": r[4]} for r in cur.fetchall()]
        return {"rows": rows}
    finally:
        nx.close()

@app.post("/api/dtrade/recompute")
def dtrade_recompute(payload: dict = Body(...)):
    """월 판가 일괄 재계산(=w_tc_master_165 일괄등록 자동판). 직거래LME만: 판가=base+동소요량×(LME월−base_LME).
    사급정체는 base 고정(판가 그대로, ts에 그대로 기록). nx.dtrade_price_ts upsert."""
    ym = str(payload.get("ym", "")).strip()[:6]
    if len(ym) != 6: raise HTTPException(400, "ym(YYYYMM) 필요")
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT lme_index FROM nx.dtrade_lme_index WHERE apply_ym=?", ym)
        r = cur.fetchone()
        if not r: return {"ok": False, "errors": [f"{ym} LME index 없음 — LME인정가 탭에서 해당월 확보 필요"]}
        lme = float(r[0] or 0)
        cur.execute("DELETE FROM nx.dtrade_price_ts WHERE apply_ym=?", ym)
        # 직거래LME: 재계산
        cur.execute("""INSERT INTO nx.dtrade_price_ts(item_code,cust_code,cost_tag,apply_ym,lme_index,mat_cost_calc,item_cost,main_flag,remarks)
            SELECT item_code,cust_code,cost_tag,?, ?, ROUND(dong_qty*?,2),
                   ROUND(base_item_cost + dong_qty*(? - base_lme),2), main_flag, N'동가반영(직거래)'
            FROM nx.dtrade_price WHERE linkage=N'직거래LME'""", ym, lme, lme, lme)
        n_dir = cur.rowcount
        # 사급정체: base 고정
        cur.execute("""INSERT INTO nx.dtrade_price_ts(item_code,cust_code,cost_tag,apply_ym,lme_index,mat_cost_calc,item_cost,main_flag,remarks)
            SELECT item_code,cust_code,cost_tag,?, NULL, NULL, base_item_cost, main_flag, N'사급정체(고정)'
            FROM nx.dtrade_price WHERE linkage=N'사급정체'""", ym)
        n_st = cur.rowcount
        return {"ok": True, "ym": ym, "lme_index": lme, "직거래LME": n_dir, "사급정체": n_st}
    finally:
        nx.close()

@app.get("/api/dtrade/list")
def dtrade_list(ym: str = Query(""), linkage: str = Query(""), q: str = Query(""), limit: int = Query(300)):
    """대상 목록 + (ym 지정시)계산 판가. linkage/품번 필터."""
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["1=1"]; p = []
        if linkage.strip(): w.append("d.linkage=?"); p.append(linkage.strip())
        if q.strip(): w.append("(d.item_code LIKE ? OR d.item_desc LIKE ?)"); p += [f"%{q.strip()}%"] * 2
        cur.execute(f"""SELECT TOP {int(limit)} d.item_code,d.cust_code,d.cost_tag,d.linkage,d.dong_qty,d.qty_src,
              d.base_ym,d.base_item_cost,d.base_lme,d.main_flag,ISNULL(d.item_desc,''),d.last_ymd,d.sagub_flag,
              t.item_cost, t.mat_cost_calc
            FROM nx.dtrade_price d LEFT JOIN nx.dtrade_price_ts t ON t.item_code=d.item_code AND t.cust_code=d.cust_code
              AND t.cost_tag=d.cost_tag AND t.apply_ym=?
            WHERE {' AND '.join(w)} ORDER BY d.linkage, d.item_code""", ym.strip(), *p)
        cols = ["item_code", "cust_code", "cost_tag", "linkage", "dong_qty", "qty_src", "base_ym",
                "base_item_cost", "base_lme", "main_flag", "item_desc", "last_ymd", "sagub_flag", "calc_item_cost", "calc_mat"]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            for k in ("dong_qty", "base_item_cost", "base_lme", "calc_item_cost", "calc_mat"):
                r[k] = float(r[k]) if r[k] is not None else None
        cur.execute("SELECT COUNT(*) FROM nx.dtrade_price d WHERE " + ' AND '.join(w), *p)
        return {"rows": rows, "cnt": cur.fetchone()[0], "ym": ym}
    finally:
        nx.close()

@app.get("/api/dtrade/compare")
def dtrade_compare(ym: str = Query(...), batch_ymd: str = Query(...), tol: float = Query(5.0)):
    """★라이브 대사: 우리 계산판가(ym) vs 라이브 PR_M_ITEM_COST 실판가(batch_ymd 배치) diff·일치율(±tol원).
    직거래LME 대상만. 산식 재현 검증."""
    ym = ym.strip()[:6]; batch = batch_ymd.strip()
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("""SELECT t.item_code,t.cust_code,t.cost_tag,t.item_cost
            FROM nx.dtrade_price_ts t JOIN nx.dtrade_price d ON d.item_code=t.item_code AND d.cust_code=t.cust_code AND d.cost_tag=t.cost_tag
            WHERE t.apply_ym=? AND d.linkage=N'직거래LME'""", ym)
        calc = {(r[0], str(r[1]).strip(), r[2]): float(r[3] or 0) for r in cur.fetchall()}
    finally:
        nx.close()
    if not calc:
        return {"ym": ym, "batch_ymd": batch, "cnt": 0, "msg": "재계산 결과 없음 — 먼저 recompute 실행"}
    cn = _conn(); cur = cn.cursor()
    try:
        items = list({k[0] for k in calc})
        live = {}
        for i in range(0, len(items), 900):
            ch = items[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"""SELECT ITEM_CODE,ISNULL(CUST_CODE,''),COST_TAG,ITEM_COST FROM PR_M_ITEM_COST
                WHERE COST_APPLY_YMD=? AND COST_TAG IN('E','S') AND ITEM_CODE IN ({ph})""", batch, *ch)
            for r in cur.fetchall():
                live[(r[0].strip(), str(r[1]).strip(), r[2])] = float(r[3] or 0)
    finally:
        cn.close()
    matched = both = 0; diffs = []; samples = []
    for k, cv in calc.items():
        lv = live.get(k)
        if lv is None: continue
        both += 1; d = cv - lv; diffs.append(abs(d))
        if abs(d) <= tol: matched += 1
        elif len(samples) < 15: samples.append({"item": k[0], "cust": k[1], "tag": k[2], "calc": round(cv, 1), "live": round(lv, 1), "diff": round(d, 1)})
    import statistics as _st
    return {"ym": ym, "batch_ymd": batch, "tol": tol, "calc_cnt": len(calc), "live_matched_keys": both,
            "within_tol": matched, "match_rate": round(matched / both * 100, 1) if both else 0,
            "mean_abs_diff": round(_st.mean(diffs), 1) if diffs else 0,
            "median_abs_diff": round(_st.median(diffs), 1) if diffs else 0, "mismatch_samples": samples}


# ============ 협력사견적: 견적(원소재비/가공비 분리) vs 현재 입고가 ============
#  모델: 원소재비 = total_weight(kg) × sagub_price(원/kg) · 가공비 고정
#        판가 = 원소재비 + 가공비 → 사급가 변경 시 원소재비만 재계산, 가공비 유지
#        현재 입고가 = PR_M_ITEM_COST.ITEM_COST (라이브, 최신 COST_APPLY_YMD)
_PREV_YMD = "251231"   # 종전입고가 기준일: 작년 12월 이하 최신 실입고
def _incost(cur, assys):
    """assy_code 리스트 → 실제 입고가(PU_T_STOCK_MAINT TAG='S' 협력사 납품 실거래).
       {assy_upper: (현재입고가=최근 MAINT_COST, 종전입고가=<=251231 최근, 최근납품일 yymmdd)}."""
    out = {}
    if not assys:
        return out
    for i in range(0, len(assys), 400):
        chunk = [str(a).replace("'", "").strip().upper() for a in assys[i:i + 400] if a]
        if not chunk:
            continue
        inlist = "','".join(chunk)
        cur.execute(f"""
            WITH M AS (
              SELECT UPPER(LTRIM(RTRIM(MAT_CODE))) IC, MAINT_YMD, MAINT_COST,
                ROW_NUMBER() OVER (PARTITION BY UPPER(LTRIM(RTRIM(MAT_CODE))) ORDER BY MAINT_YMD DESC) rn_cur,
                ROW_NUMBER() OVER (PARTITION BY UPPER(LTRIM(RTRIM(MAT_CODE)))
                   ORDER BY (CASE WHEN MAINT_YMD<='{_PREV_YMD}' THEN 0 ELSE 1 END), MAINT_YMD DESC) rn_prev
              FROM PARTNER_ERP.dbo.PU_T_STOCK_MAINT
              WHERE MAINT_TAG='S' AND MAINT_QTY>0 AND MAINT_COST>0
                AND UPPER(LTRIM(RTRIM(MAT_CODE))) IN ('{inlist}')
            )
            SELECT IC,
              MAX(CASE WHEN rn_cur=1 THEN MAINT_COST END) cur_cost,
              MAX(CASE WHEN rn_cur=1 THEN MAINT_YMD END) cur_ymd,
              MAX(CASE WHEN rn_prev=1 AND MAINT_YMD<='{_PREV_YMD}' THEN MAINT_COST END) prev_cost
            FROM M GROUP BY IC""")
        for r in cur.fetchall():
            out[str(r[0]).strip().upper()] = (
                float(r[1] or 0),
                (float(r[3]) if r[3] is not None else None),
                str(r[2] or '').strip())
    return out


@app.get("/api/coopquote/vendors")
def coopquote_vendors():
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT vendor, COUNT(*) n FROM nx.coop_quote GROUP BY vendor ORDER BY vendor")
        return {"rows": [{"vendor": str(r[0]).strip(), "n": int(r[1])} for r in cur.fetchall()]}
    finally:
        nx.close()


@app.get("/api/coopquote/list")
def coopquote_list(vendor: str = Query(""), q: str = Query(""), active_only: int = Query(0)):
    """협력사 견적 목록 + 실제 입고가(라이브 실거래) + 차이. 원소재비/가공비 분리.
       active_only=1 → 최근 4개월 내 실제 납품(입고) 실적 있는 품목만."""
    nx = _nx(); cur = nx.cursor()
    try:
        where = ["1=1"]; args = []
        if vendor.strip():
            where.append("vendor=?"); args.append(vendor.strip())
        if q.strip():
            where.append("(assy_code LIKE ? OR item_name LIKE ?)")
            args += [f"%{q.strip()}%", f"%{q.strip()}%"]
        cur.execute(f"""SELECT quote_id,vendor,assy_code,item_name,spec,total_weight,sagub_price,
            mat_cost,proc_cost,sale_price,quote_price,final_price,lg_price,currency,status,remark,
            CONVERT(varchar(19),ISNULL(upd_dt,reg_dt),120), ISNULL(grade,N'일반CU'), mat_ratio, ISNULL(fixed_mat,0),
            ISNULL(mat_raw,0), ISNULL(mat_weld,0), ISNULL(mat_part,0)
          FROM nx.coop_quote WHERE {' AND '.join(where)}
          ORDER BY vendor, assy_code""", *args)
        rows = [dict(zip(
            ["quote_id","vendor","assy_code","item_name","spec","total_weight","sagub_price",
             "mat_cost","proc_cost","sale_price","quote_price","final_price","lg_price","currency","status","remark","upd_dt","grade","mat_ratio","fixed_mat","mat_raw","mat_weld","mat_part"],
            [r[0], str(r[1] or ""), str(r[2] or ""), str(r[3] or ""), str(r[4] or ""),
             float(r[5] or 0), float(r[6] or 0), float(r[7] or 0), float(r[8] or 0), float(r[9] or 0),
             (float(r[10]) if r[10] is not None else None), (float(r[11]) if r[11] is not None else None),
             (float(r[12]) if r[12] is not None else None), str(r[13] or "KRW"), str(r[14] or ""),
             str(r[15] or ""), r[16], str(r[17] or "일반CU"),
             (float(r[18]) if r[18] is not None else None), float(r[19] or 0),
             float(r[20] or 0), float(r[21] or 0), float(r[22] or 0)])) for r in cur.fetchall()]
        cost = _incost(cur, [r["assy_code"] for r in rows])
        for r in rows:
            c = cost.get(r["assy_code"].strip().upper())
            r["cur_incost"] = (c[0] if (c and c[0]) else None)   # 현재 실입고가(최근 실거래)
            r["prev_incost"] = c[1] if c else None               # 종전 실입고가(작년 12월 이하 최근)
            r["last_in_ymd"] = (c[2] if (c and c[2]) else None)  # 최근 납품일 yymmdd
            # 최종견적가 = 판가(sale_price = 재료비+가공비) 통합
            fin = r["sale_price"]
            r["final_quote"] = fin
            r["diff"] = (fin - r["cur_incost"]) if (r["cur_incost"] is not None) else None
        if active_only:
            cutoff = (datetime.now() - timedelta(days=120)).strftime('%y%m%d')
            rows = [r for r in rows if r.get("last_in_ymd") and r["last_in_ymd"] >= cutoff]
        return {"rows": rows, "count": len(rows), "active_only": bool(active_only)}
    finally:
        nx.close()


@app.get("/api/coopquote/parts")
def coopquote_parts(assy: str = Query(...), vendor: str = Query("")):
    """Assy 하위부품 상세(3분류: 동관/사급부품/용접봉). 재료비·가공비·합계."""
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["assy_code=?"]; a = [assy.strip()]
        if vendor.strip():
            w.append("vendor=?"); a.append(vendor.strip())
        cur.execute(f"""SELECT seq,part_code,part_name,ptype,spec,mat_cost,proc_cost,part_total
            FROM nx.coop_quote_part WHERE {' AND '.join(w)} ORDER BY seq""", *a)
        rows = [dict(zip(["seq","part_code","part_name","ptype","spec","mat_cost","proc_cost","part_total"],
            [r[0], str(r[1] or ""), str(r[2] or ""), str(r[3] or ""), str(r[4] or ""),
             float(r[5] or 0), float(r[6] or 0), float(r[7] or 0)])) for r in cur.fetchall()]
        return {"rows": rows, "count": len(rows)}
    finally:
        nx.close()


@app.post("/api/coopquote/save")
def coopquote_save(payload: dict = Body(...)):
    """신규/수정 견적 저장. 원소재비=total_weight×sagub_price, 판가=원소재비+가공비."""
    qid = int(payload.get("quote_id") or 0)
    vendor = str(payload.get("vendor") or "").strip()
    assy = str(payload.get("assy_code") or "").strip()
    if not vendor or not assy:
        return {"ok": False, "error": "협력사·품번 필수"}
    def fnum(k, d=0.0):
        try: return float(payload.get(k) if payload.get(k) not in (None, "") else d)
        except: return d
    def fopt(k):
        v = payload.get(k)
        try: return float(v) if v not in (None, "") else None
        except: return None
    w = fnum("total_weight"); sg = fnum("sagub_price"); proc = fnum("proc_cost")
    mat = round(w * sg); sale = mat + round(proc)
    item_name = str(payload.get("item_name") or "")[:120]
    spec = str(payload.get("spec") or "")[:80]
    remark = str(payload.get("remark") or "")[:200]
    qp = fopt("quote_price"); fp = fopt("final_price"); lg = fopt("lg_price")
    status = "확정" if fp is not None else str(payload.get("status") or "견적")
    grade = str(payload.get("grade") or "").strip()
    if grade not in ("일반CU", "고강도CU"):
        grade = "일반CU"
    nx = _nx(); cur = nx.cursor()
    try:
        if qid:
            cur.execute("""UPDATE nx.coop_quote SET vendor=?,assy_code=?,item_name=?,spec=?,
                total_weight=?,sagub_price=?,mat_cost=?,proc_cost=?,sale_price=?,
                quote_price=?,final_price=?,lg_price=?,status=?,grade=?,remark=?,upd_user='web',upd_dt=GETDATE()
                WHERE quote_id=?""",
                vendor, assy, item_name, spec, w, sg, mat, round(proc), sale, qp, fp, lg, status, grade, remark, qid)
            return {"ok": True, "quote_id": qid, "mat_cost": mat, "sale_price": sale}
        cur.execute("""INSERT INTO nx.coop_quote(vendor,assy_code,item_name,spec,total_weight,sagub_price,
            mat_cost,proc_cost,sale_price,quote_price,final_price,lg_price,status,grade,remark,src,reg_user)
            OUTPUT INSERTED.quote_id
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'수동입력','web')""",
            vendor, assy, item_name, spec, w, sg, mat, round(proc), sale, qp, fp, lg, status, grade, remark)
        newid = cur.fetchone()[0]
        return {"ok": True, "quote_id": int(newid), "mat_cost": mat, "sale_price": sale}
    finally:
        nx.close()


@app.post("/api/coopquote/recalc")
def coopquote_recalc(payload: dict = Body(...)):
    """사급가 변경 → 원소재비/판가 재계산(가공비 유지). 등급별 2단가 적용:
       일반CU=price_normal, 고강도CU=price_high. scope=all|vendor|ids.
       (구버전 단일 sagub_price도 허용 → 양 등급 동일 적용)"""
    def fnum(k):
        try: return float(payload.get(k))
        except (TypeError, ValueError): return None
    pn = fnum("price_normal"); ph = fnum("price_high"); single = fnum("sagub_price")
    if single is not None and pn is None and ph is None:
        pn = ph = single
    if not pn and not ph:
        return {"ok": False, "error": "사급가 필요"}
    scope = str(payload.get("scope") or "all")
    # 공통 scope where
    scope_sql = ""; scope_args = []
    if scope == "vendor" and payload.get("vendor"):
        scope_sql = " AND vendor=?"; scope_args = [str(payload["vendor"]).strip()]
    elif scope == "ids" and payload.get("ids"):
        ids = [int(x) for x in payload["ids"] if str(x).isdigit()]
        if not ids:
            return {"ok": False, "error": "대상 없음"}
        scope_sql = " AND quote_id IN (" + ",".join(str(i) for i in ids) + ")"
    nx = _nx(); cur = nx.cursor()
    try:
        total = 0; detail = {}
        for grade, price in (("일반CU", pn), ("고강도CU", ph)):
            if not price or price <= 0:
                continue
            cur.execute(f"""UPDATE nx.coop_quote
                SET sagub_price=?, mat_cost=ROUND(total_weight*?,0),
                    sale_price=ROUND(total_weight*{float(price)},0)+proc_cost,
                    upd_user='web', upd_dt=GETDATE()
                WHERE total_weight>0 AND grade=?{scope_sql}""",
                float(price), float(price), grade, *scope_args)
            detail[grade] = int(cur.rowcount); total += int(cur.rowcount)
        return {"ok": True, "count": total, "detail": detail, "price_normal": pn, "price_high": ph}
    finally:
        nx.close()


@app.get("/api/coopquote/bom-form")
def coopquote_bom_form(item: str = Query(..., description="품번(Assy)"), vendor: str = Query("")):
    """견적 입력폼용: 현 BOM(CS_M_ITEM_BOM) 구성 전개 + 부품별 역할/스펙/매입가/공정 프리필.
       구성=현BOM 정본(고정). 직원은 '제작동관'의 협력사 스펙(외경/두께/길이)만 채우면 됨.
       공정=coop_part_proc(견적 유래), 가공비=Σ(임율/표준ST_공정)×횟수×소요량."""
    import math
    vendor = vendor.strip()
    item = item.strip()
    cn = _conn(); cur = cn.cursor()
    try:
        # 1) BOM 재귀전개 (현행)
        cur.execute("""WITH tree AS (
            SELECT ITEM_CODE p, MAT_CODE c, CAST(USE_QTY AS decimal(18,6)) q, ISNULL(SAGUB_FLAG,'0') sag, ISNULL(BOM_SEQ,0) sq, 1 lvl
            FROM CS_M_ITEM_BOM WHERE ITEM_CODE=? AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101'
            UNION ALL
            SELECT b.ITEM_CODE, b.MAT_CODE, CAST(b.USE_QTY AS decimal(18,6)), ISNULL(b.SAGUB_FLAG,'0'), ISNULL(b.BOM_SEQ,0), t.lvl+1
            FROM tree t JOIN CS_M_ITEM_BOM b ON b.ITEM_CODE=t.c AND b.FROM_APPLY_YMD<='991231' AND b.TO_APPLY_YMD>='260101'
            WHERE t.lvl < 8)
            SELECT p,c,q,sag,sq,lvl FROM tree OPTION(MAXRECURSION 50)""", item)
        edges = {}
        for r in cur.fetchall():
            edges.setdefault(r[0], []).append({"child": str(r[1]).strip(), "q": float(r[2] or 0), "sag": str(r[3]), "sq": int(r[4] or 0)})
        for p in edges:
            edges[p].sort(key=lambda x: x["sq"])
        nodes = {item} | {e["child"] for lst in edges.values() for e in lst} | set(edges.keys())
        # 2) LG 마스터 정보
        info = {}
        nl = list(nodes)
        for i in range(0, len(nl), 900):
            chunk = nl[i:i+900]; ph = ",".join("?" * len(chunk))
            cur.execute(f"""SELECT ITEM_CODE, ISNULL(ITEM_DESC,''), ISNULL(ITEM_SPEC,''), ISNULL(METAL_GUBUN,''),
                  ISNULL(ITEM_DIAM,0), ISNULL(ITEM_THICK,0), ISNULL(ITEM_LENGTH,0)
                FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})""", *chunk)
            for r in cur.fetchall():
                info[str(r[0]).strip()] = {"nm": r[1], "spec": r[2], "metal": str(r[3]).strip(),
                    "diam": float(r[4] or 0), "thick": float(r[5] or 0), "length": float(r[6] or 0)}
        # 3) 매입가(PR_M_ITEM_COST 최신, 제이에스2228 제외, 매입TAG='1' 우선)
        pur = {}
        for i in range(0, len(nl), 900):
            chunk = [c.replace("'", "") for c in nl[i:i+900]]; inl = "','".join(chunk)
            cur.execute(f"""WITH C AS (
                  SELECT UPPER(LTRIM(RTRIM(ITEM_CODE))) ic, ITEM_COST, COST_APPLY_YMD,
                    ROW_NUMBER() OVER (PARTITION BY UPPER(LTRIM(RTRIM(ITEM_CODE)))
                      ORDER BY (CASE WHEN COST_TAG='1' THEN 0 ELSE 1 END), COST_APPLY_YMD DESC) rn
                  FROM PR_M_ITEM_COST
                  WHERE UPPER(LTRIM(RTRIM(ITEM_CODE))) IN ('{inl}')
                    AND LTRIM(RTRIM(CUST_CODE))<>'2228' AND ITEM_COST>0)
                SELECT ic, ITEM_COST FROM C WHERE rn=1""")
            for r in cur.fetchall():
                pur[str(r[0]).strip().upper()] = float(r[1] or 0)
    finally:
        cn.close()
    # 4) 협력사 스펙 프리필(coop_raw_spec) + 기존 견적여부
    nx = _nx(); ncur = nx.cursor()
    coop = {}; quoted = False
    try:
        for i in range(0, len(nl), 900):
            chunk = [c.replace("'", "") for c in nl[i:i+900]]; inl = "','".join(chunk)
            ncur.execute(f"""SELECT UPPER(LTRIM(RTRIM(item_code))), diam, thick, length_mm, unit_weight, sagub_price
                FROM nx.coop_raw_spec WHERE UPPER(LTRIM(RTRIM(item_code))) IN ('{inl}')""")
            for r in ncur.fetchall():
                coop[str(r[0]).strip().upper()] = {"diam": float(r[1] or 0), "thick": float(r[2] or 0),
                    "length": float(r[3] or 0), "uw": float(r[4] or 0), "sagub": float(r[5] or 0)}
        ncur.execute("SELECT COUNT(*) FROM nx.coop_quote WHERE assy_code=?", item)
        quoted = ncur.fetchone()[0] > 0
        # 저장 부품 합계(엑셀 AQ) — bottom-up 합산 표시용
        partmap = {}; part_sum = 0.0; sale_stored = 0.0
        try:
            ncur.execute("SELECT UPPER(LTRIM(RTRIM(part_code))), part_total, mat_cost, proc_cost, ISNULL(ptype,'') FROM nx.coop_quote_part WHERE assy_code=?", item)
            for r in ncur.fetchall():
                pt = float(r[1] or 0); part_sum += pt
                partmap[str(r[0]).strip().upper()] = {"total": pt, "mat": float(r[2] or 0), "proc": float(r[3] or 0)}
            ncur.execute("SELECT ISNULL(sale_price,0) FROM nx.coop_quote WHERE assy_code=?", item)
            rr = ncur.fetchone()
            if rr: sale_stored = float(rr[0] or 0)
        except Exception:
            partmap = {}
        # 서브 조립 공정비 = 판가 − Σ하위부품합계 (견적서엔 용접봉줄에 넣었던 실제 조립작업 공정)
        assembly_proc = round(sale_stored - part_sum) if (sale_stored and part_sum) else 0
        # 서브조립 상세(공정ST·가공비·관리/운반/이윤) from coop_assembly
        assembly = None
        try:
            import json as _json
            ncur.execute("SELECT procs, gagong, mgmt, transport, profit, total FROM nx.coop_assembly WHERE assy_code=?", item)
            ar = ncur.fetchone()
            if ar:
                assembly = {"procs": _json.loads(ar[0] or '{}'), "gagong": int(ar[1] or 0), "mgmt": int(ar[2] or 0),
                            "transport": int(ar[3] or 0), "profit": int(ar[4] or 0), "total": int(ar[5] or 0)}
        except Exception:
            assembly = None
        # 공정(부품별 op:cnt) from coop_part_proc — (Assy,부품)별 정확값
        procmap = {}
        try:
            ncur.execute("SELECT UPPER(LTRIM(RTRIM(part_code))), op, cnt FROM nx.coop_part_proc WHERE assy_code=?", item)
            for r in ncur.fetchall():
                procmap.setdefault(str(r[0]).strip().upper(), {})[str(r[1]).strip()] = int(r[2] or 0)
        except Exception:
            procmap = {}
        # 표준ST(공정 divisor): 해당 벤더 우선, 없으면 전체 평균
        rate = {}
        try:
            if vendor:
                ncur.execute("SELECT op, divisor FROM nx.coop_gagong_rate WHERE vendor=?", vendor)
                for r in ncur.fetchall(): rate[str(r[0]).strip()] = float(r[1] or 0)
            ncur.execute("SELECT op, AVG(divisor) FROM nx.coop_gagong_rate GROUP BY op")
            for r in ncur.fetchall(): rate.setdefault(str(r[0]).strip(), float(r[1] or 0))
        except Exception:
            rate = {}
        # 임율
        labor = 6300.0
        try:
            ncur.execute("SELECT TOP 1 CAST(val AS float) FROM nx.coop_config WHERE k IN ('임율','labor_rate','rate')")
            rr = ncur.fetchone()
            if rr and rr[0]: labor = float(rr[0])
        except Exception:
            pass
    finally:
        nx.close()

    def gagong_piece(code):
        """부품 1개당 가공비 = Σ(임율/표준ST_공정)×횟수. 세척 등 제수없는 공정 제외."""
        ops = procmap.get(code.upper()) or {}
        tot = 0.0
        for op, cnt in ops.items():
            div = rate.get(op, 0)
            if div and div > 0: tot += (labor / div) * cnt
        return tot

    def role_of(code, sag, metal, haskids):
        u = code.upper()
        if haskids: return "반제품"
        if u.startswith("RAC") or u.startswith("BCUP"): return "용접봉"
        if sag == "1": return "사급"
        if metal in ("CU", "고강도"): return "제작동관"
        return "매입부품"

    rows = []; need = 0
    seen = set()
    def geom(d, t, L): return round(math.pi*(d-t)*t*L*8.94/1e6, 5) if (d and t and L) else 0.0
    def walk(code, lvl, cumq):
        if code in seen: return
        seen.add(code)
        for e in edges.get(code, []):
            ch = e["child"]; ci = info.get(ch, {}); haskids = ch in edges
            role = role_of(ch, e["sag"], ci.get("metal", ""), haskids)
            cq = cumq * e["q"]
            cs = coop.get(ch.upper()); uw = 0.0; src = ""
            if cs and (cs["uw"] > 0 or (cs["diam"] and cs["thick"] and cs["length"])):
                uw = cs["uw"] if cs["uw"] > 0 else geom(cs["diam"], cs["thick"], cs["length"]); src = "협력사"
            elif role == "제작동관" and ci.get("diam") and ci.get("thick") and ci.get("length"):
                uw = geom(ci["diam"], ci["thick"], ci["length"]); src = "LG참고"
            need_input = (role == "제작동관" and not (cs and (cs["uw"] > 0 or (cs["diam"] and cs["thick"]))))
            pp = pur.get(ch.upper(), None)
            # 용접봉 = 재료비 아님 → 공정(용접) 부자재비. 비용 = 단가 × 소요량
            weld_cost = round(pp * cq) if (role == "용접봉" and pp) else 0
            procs = procmap.get(ch.upper()) if role == "제작동관" else None
            gp_piece = gagong_piece(ch) if role == "제작동관" else 0
            proc_cost = round(gp_piece * cq) if role == "제작동관" else 0
            pt = partmap.get(ch.upper())   # 저장 합계(엑셀 AQ)
            rows.append({
                "part_total": (round(pt["total"]) if pt else None),
                "part_mat": (round(pt["mat"]) if pt else None),
                "level": lvl, "code": ch, "name": ci.get("nm", ""), "role": role,
                "use_qty": e["q"], "cum_qty": round(cq, 5), "sagub": e["sag"] == "1",
                "lg_diam": ci.get("diam", 0), "lg_thick": ci.get("thick", 0), "lg_length": ci.get("length", 0),
                "coop_diam": (cs["diam"] if cs else 0), "coop_thick": (cs["thick"] if cs else 0),
                "coop_length": (cs["length"] if cs else 0), "coop_sagub": (cs.get("sagub", 0) if cs else 0),
                "unit_weight": uw, "weight_src": src, "soyo_weight": round(uw*cq, 5),
                "pur_price": pp, "weld_cost": weld_cost, "is_proc": role == "용접봉",
                "procs": procs, "proc_cost": proc_cost,
                "need_input": need_input, "haskids": haskids})
            walk(ch, lvl+1, cq)
        seen.discard(code)
    walk(item, 1, 1.0)
    need = sum(1 for r in rows if r["need_input"])
    total_soyo = round(sum(r["soyo_weight"] for r in rows if r["role"] == "제작동관"), 4)
    total_weld = round(sum(r["weld_cost"] for r in rows if r["role"] == "용접봉"))
    total_proc = round(sum(r["proc_cost"] for r in rows if r["role"] == "제작동관"))
    # 공정 컬럼 고정(엑셀 전체 공정 순서) — 품번마다 변동하지 않음
    proc_ops = ['컷팅','면취','벤딩','CNC','딤플','벌징','원교정','피어싱','압착','T뽑기','후레아','축확관','실링','망삽입','막음','코킹','세척','용접','부삽입','교/체','수몰검사','포장']
    root = info.get(item, {})
    return {"item": item, "name": root.get("nm", ""), "already_quoted": quoted, "labor_rate": labor,
            "rows": rows, "count": len(rows), "need_input": need, "proc_ops": proc_ops, "rate": rate,
            "total_soyo_weight": total_soyo, "total_weld_cost": total_weld, "total_proc_cost": total_proc,
            "part_sum": round(part_sum), "assembly_proc": assembly_proc, "sale_stored": round(sale_stored),
            "assembly": assembly}


def _coop_soyo(item):
    """현 BOM 재귀전개 × coop_raw_spec(협력사 개당중량) → 제작동관 소요중량 합.
       returns (total_soyo_weight, leaf_rows[{code,cum_qty,unit_weight,role}])."""
    import math
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("""WITH tree AS (
            SELECT ITEM_CODE p, MAT_CODE c, CAST(USE_QTY AS decimal(18,6)) q, ISNULL(SAGUB_FLAG,'0') sag, 1 lvl
            FROM CS_M_ITEM_BOM WHERE ITEM_CODE=? AND FROM_APPLY_YMD<='991231' AND TO_APPLY_YMD>='260101'
            UNION ALL
            SELECT b.ITEM_CODE, b.MAT_CODE, CAST(b.USE_QTY AS decimal(18,6)), ISNULL(b.SAGUB_FLAG,'0'), t.lvl+1
            FROM tree t JOIN CS_M_ITEM_BOM b ON b.ITEM_CODE=t.c AND b.FROM_APPLY_YMD<='991231' AND b.TO_APPLY_YMD>='260101'
            WHERE t.lvl<8)
            SELECT p,c,q,sag FROM tree OPTION(MAXRECURSION 50)""", item)
        edges = {}
        for r in cur.fetchall():
            edges.setdefault(str(r[0]).strip(), []).append((str(r[1]).strip(), float(r[2] or 0), str(r[3])))
        nodes = {item} | {c for lst in edges.values() for (c, q, s) in lst}
        metal = {}
        nl = list(nodes)
        for i in range(0, len(nl), 900):
            ch = nl[i:i+900]; ph = ",".join("?" * len(ch))
            cur.execute(f"SELECT ITEM_CODE, ISNULL(METAL_GUBUN,'') FROM PR_M_ITEM WHERE ITEM_CODE IN ({ph})", *ch)
            for r in cur.fetchall(): metal[str(r[0]).strip()] = str(r[1]).strip()
    finally:
        cn.close()
    nx = _nx(); ncur = nx.cursor(); coop = {}
    try:
        for i in range(0, len(nl), 900):
            ch = [c.replace("'", "") for c in nl[i:i+900]]; inl = "','".join(ch)
            ncur.execute(f"""SELECT UPPER(LTRIM(RTRIM(item_code))), diam, thick, length_mm, unit_weight, sagub_price
                FROM nx.coop_raw_spec WHERE UPPER(LTRIM(RTRIM(item_code))) IN ('{inl}')""")
            for r in ncur.fetchall():
                coop[str(r[0]).strip().upper()] = (float(r[1] or 0), float(r[2] or 0), float(r[3] or 0), float(r[4] or 0), float(r[5] or 0))
    finally:
        nx.close()
    def geom(d, t, L): return round(math.pi*(d-t)*t*L*8.94/1e6, 5) if (d and t and L) else 0.0
    leaves = []; seen = set()
    def walk(code, cq):
        if code in seen: return
        seen.add(code)
        for (ch, q, sag) in edges.get(code, []):
            haskids = ch in edges
            role = ("반제품" if haskids else ("용접봉" if ch.upper().startswith("RAC") or ch.upper().startswith("BCUP")
                    else ("사급" if sag == "1" else ("제작동관" if metal.get(ch, "") in ("CU", "고강도") else "매입부품"))))
            cc = cq * q
            if role == "제작동관":
                cs = coop.get(ch.upper()); uw = 0.0; psag = 0.0
                if cs: uw = cs[3] if cs[3] > 0 else geom(cs[0], cs[1], cs[2]); psag = cs[4] if len(cs) > 4 else 0.0
                leaves.append({"code": ch, "cum_qty": round(cc, 5), "unit_weight": uw, "role": role, "sagub": psag})
            walk(ch, cc)
        seen.discard(code)
    walk(item, 1.0)
    total = round(sum(l["unit_weight"] * l["cum_qty"] for l in leaves), 5)
    return total, leaves


@app.post("/api/coopquote/bom-save")
def coopquote_bom_save(payload: dict = Body(...)):
    """견적 입력폼 저장: 제작동관 협력사스펙 upsert(coop_raw_spec) → 소요중량 재계산 → 견적 생성/수정.
       payload: item, vendor, grade, sagub_price, proc_cost, specs:[{code,diam,thick,length}]."""
    import math
    item = str(payload.get("item") or "").strip()
    vendor = str(payload.get("vendor") or "").strip()
    if not item or not vendor:
        return {"ok": False, "error": "품번·협력사 필수"}
    grade = str(payload.get("grade") or "일반CU").strip()
    if grade not in ("일반CU", "고강도CU"): grade = "일반CU"
    def fnum(v, d=0.0):
        try: return float(v) if v not in (None, "") else d
        except: return d
    sagub = fnum(payload.get("sagub_price"))
    proc = round(fnum(payload.get("proc_cost")))
    base_mat = round(fnum(payload.get("base_mat")))   # 부속품+용접봉(held) 재료비
    specs = payload.get("specs") or []
    nx = _nx(); cur = nx.cursor()
    try:
        # 1) 협력사 스펙 upsert
        upd = 0
        for s in specs:
            code = str(s.get("code") or "").strip()
            d = fnum(s.get("diam")); t = fnum(s.get("thick")); L = fnum(s.get("length"))
            psag = fnum(s.get("sagub"))            # ★관경별 사급가(부품별). 미입력 시 헤더 기본 사급가 사용
            if psag <= 0: psag = sagub
            if not code or not (d and t and L):
                continue
            uw = round(math.pi*(d-t)*t*L*8.94/1e6, 5)
            cur.execute("SELECT 1 FROM nx.coop_raw_spec WHERE UPPER(LTRIM(RTRIM(item_code)))=?", code.upper())
            if cur.fetchone():
                cur.execute("""UPDATE nx.coop_raw_spec SET diam=?,thick=?,length_mm=?,unit_weight=?,sagub_price=?
                    WHERE UPPER(LTRIM(RTRIM(item_code)))=?""", d, t, L, uw, psag, code.upper())
            else:
                cur.execute("""INSERT INTO nx.coop_raw_spec(item_code,vendor,diam,thick,length_mm,unit_weight,sagub_price)
                    VALUES(?,?,?,?,?,?,?)""", code, vendor, d, t, L, uw, psag)
            upd += 1
        # 1b) 공정 ST upsert (제작동관, assy+part별) — 직원 편집 반영
        for pr in (payload.get("procs") or []):
            pc = str(pr.get("code") or "").strip()
            if not pc: continue
            cur.execute("DELETE FROM nx.coop_part_proc WHERE assy_code=? AND part_code=?", item, pc)
            for op, cnt in (pr.get("ops") or {}).items():
                try: c = int(cnt)
                except: c = 0
                if c > 0:
                    cur.execute("INSERT INTO nx.coop_part_proc(assy_code,part_code,op,cnt) VALUES(?,?,?,?)", item[:60], pc[:60], str(op)[:20], c)
        # 1c) 서브조립 공정/관리/운반/이윤 upsert
        asm = payload.get("assembly")
        if asm is not None:
            import json as _json
            aprocs = _json.dumps(asm.get("procs") or {}, ensure_ascii=False)
            gg = round(fnum(asm.get("gagong"))); mg = round(fnum(asm.get("mgmt"))); tr = round(fnum(asm.get("transport"))); pf = round(fnum(asm.get("profit")))
            atot = gg + mg + tr + pf
            cur.execute("SELECT 1 FROM nx.coop_assembly WHERE assy_code=?", item)
            if cur.fetchone():
                cur.execute("UPDATE nx.coop_assembly SET procs=?,gagong=?,mgmt=?,transport=?,profit=?,total=? WHERE assy_code=?", aprocs, gg, mg, tr, pf, atot, item)
            else:
                cur.execute("INSERT INTO nx.coop_assembly(assy_code,procs,gagong,mgmt,transport,profit,total) VALUES(?,?,?,?,?,?,?)", item[:60], aprocs, gg, mg, tr, pf, atot)
        nx.commit()
    finally:
        nx.close()
    # 2) 소요중량 재계산 (upsert 반영 후). 재료비=원소재비+부속품/용접봉(held), 판가=재료비+가공비
    #    ★원소재비 = Σ(리프 소요중량 × 해당 관경 사급가) — 관경별 사급가 반영(미저장 리프는 헤더 sagub)
    total_soyo, leaves = _coop_soyo(item)
    mat_raw = round(sum(l["unit_weight"] * l["cum_qty"] * ((l.get("sagub") or 0) or sagub) for l in leaves))
    mat = mat_raw + base_mat
    sale = mat + proc
    # 3) 견적 upsert
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT quote_id FROM nx.coop_quote WHERE assy_code=? AND vendor=?", item, vendor)
        row = cur.fetchone()
        nm = str(payload.get("item_name") or "")[:120]; spec = str(payload.get("spec") or "")[:80]
        if row:
            cur.execute("""UPDATE nx.coop_quote SET item_name=?,spec=?,total_weight=?,sagub_price=?,mat_cost=?,
                proc_cost=?,sale_price=?,mat_raw=?,grade=?,src='폼입력',upd_user='web',upd_dt=GETDATE()
                WHERE quote_id=?""", nm, spec, total_soyo, sagub, mat, proc, sale, mat_raw, grade, row[0])
            qid = row[0]
        else:
            cur.execute("""INSERT INTO nx.coop_quote(vendor,assy_code,item_name,spec,total_weight,sagub_price,
                mat_cost,proc_cost,sale_price,mat_raw,status,grade,src,reg_user)
                OUTPUT INSERTED.quote_id VALUES(?,?,?,?,?,?,?,?,?,?,'견적',?,'폼입력','web')""",
                vendor, item, nm, spec, total_soyo, sagub, mat, proc, sale, mat_raw, grade)
            qid = cur.fetchone()[0]
        # 작업목록 완료 처리(직원 입력 검증 후 resolved)
        try:
            cur.execute("UPDATE nx.coop_worklist SET resolved=1 WHERE assy_code=?", item)
        except Exception:
            pass
        nx.commit()
        return {"ok": True, "quote_id": int(qid), "spec_updated": upd,
                "total_soyo_weight": total_soyo, "mat_cost": mat, "sale_price": sale}
    finally:
        nx.close()


@app.get("/api/coopquote/worklist")
def coopquote_worklist(wtype: str = Query("")):
    """직원 입력 작업목록: ①데이터문제(견적재료비 불일치) ②신규(견적없는 입고품). resolved=0만, 입고수량 우선."""
    nx = _nx(); cur = nx.cursor()
    try:
        w = ["resolved=0"]; a = []
        if wtype.strip():
            w.append("wtype=?"); a.append(wtype.strip())
        cur.execute(f"""SELECT assy_code,vendor,wtype,reason,in_qty,db_mat,xl_mat
            FROM nx.coop_worklist WHERE {' AND '.join(w)} ORDER BY in_qty DESC, assy_code""", *a)
        rows = [dict(zip(["assy_code","vendor","wtype","reason","in_qty","db_mat","xl_mat"],
            [str(r[0]).strip(), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""), int(r[4] or 0), int(r[5] or 0), str(r[6] or "")])) for r in cur.fetchall()]
        cur.execute("SELECT wtype, COUNT(*) FROM nx.coop_worklist WHERE resolved=0 GROUP BY wtype")
        by = {str(r[0]): int(r[1]) for r in cur.fetchall()}
        cur.execute("SELECT COUNT(*) FROM nx.coop_worklist WHERE resolved=1")
        done = int(cur.fetchone()[0])
        return {"rows": rows, "count": len(rows), "by_type": by, "resolved": done}
    finally:
        nx.close()


# ============ 기준정보: 업체별 재고금액(월재고 스냅샷 → 매입처 집계) ============
@app.get("/api/stockval/list")
def stockval_list(ym: str = Query(""), incust: str = Query("")):
    """업체별(매입처) 재고금액. 월재고 스냅샷 PU_T_MONTH_STOCK_WH를 MAT→PR_M_ITEM.IN_CUST_CODE로 집계.
    incust 지정 시 해당 매입처 자재 명세. 라이브·읽기전용."""
    cn = _conn(); cur = cn.cursor()
    try:
        y = _dig4(ym)
        if not y:
            cur.execute("SELECT MAX(STOCK_YYMM) FROM PU_T_MONTH_STOCK_WH"); y = cur.fetchone()[0]
        cur.execute("SELECT DISTINCT TOP 24 STOCK_YYMM FROM PU_T_MONTH_STOCK_WH ORDER BY STOCK_YYMM DESC")
        months = [r[0] for r in cur.fetchall()]
        if incust.strip():
            cur.execute("""SELECT TOP 5000 W.MAT_CODE mat, ISNULL(M.ITEM_DESC,'') nm, ISNULL(M.ITEM_SPEC,'') spec,
                  ISNULL(M.UNIT,'') unit, SUM(W.STOCK_QTY) qty, MAX(W.STOCK_COST) cost, SUM(W.STOCK_AMT) amt
                FROM PU_T_MONTH_STOCK_WH W JOIN PR_M_ITEM M ON M.ITEM_CODE=W.MAT_CODE
                WHERE W.STOCK_YYMM=? AND ISNULL(M.IN_CUST_CODE,'')=?
                GROUP BY W.MAT_CODE, M.ITEM_DESC, M.ITEM_SPEC, M.UNIT HAVING SUM(W.STOCK_QTY)<>0
                ORDER BY SUM(W.STOCK_AMT) DESC""", y, incust.strip())
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            for r in rows:
                r["qty"] = float(r["qty"] or 0); r["cost"] = float(r["cost"] or 0); r["amt"] = float(r["amt"] or 0)
            return {"mode": "detail", "ym": y, "months": months, "incust": incust.strip(),
                    "rows": rows, "cnt": len(rows), "sum_amt": sum(r["amt"] for r in rows)}
        cur.execute("""SELECT ISNULL(M.IN_CUST_CODE,'') incust, MAX(ISNULL(C.CUST_DESC,'')) nm,
              COUNT(DISTINCT W.MAT_CODE) items, SUM(W.STOCK_QTY) qty, SUM(W.STOCK_AMT) amt
            FROM PU_T_MONTH_STOCK_WH W JOIN PR_M_ITEM M ON M.ITEM_CODE=W.MAT_CODE
            LEFT JOIN CM_M_CUST C ON C.CUST_CODE=M.IN_CUST_CODE
            WHERE W.STOCK_YYMM=? GROUP BY M.IN_CUST_CODE HAVING SUM(W.STOCK_AMT)<>0
            ORDER BY SUM(W.STOCK_AMT) DESC""", y)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r["items"] = int(r["items"] or 0); r["qty"] = float(r["qty"] or 0); r["amt"] = float(r["amt"] or 0)
        return {"mode": "summary", "ym": y, "months": months, "rows": rows, "cnt": len(rows),
                "sum_amt": sum(r["amt"] for r in rows)}
    finally:
        cn.close()


# ============ 기준정보: 기준MASTER관리(생산 요청) — 부서·라인·조립공정·단품공정 (라이브 조회) ============
_BASEMASTER = {
    "dept":  {"t": "HR_M_DEPT", "title": "부서MASTER", "src": "HR_M_DEPT", "order": "SORT_KEY",
              "cols": [("DEPT_CODE", "부서코드"), ("DEPT_DESC", "부서명"), ("SORT_KEY", "정렬"),
                       ("ENTERPRISE_DEPT", "전사부서"), ("WH_CODE", "창고"), ("USE_FLAG", "사용")]},
    "line":  {"t": "PR_M_LINE_NO", "title": "LINE-NO MASTER", "src": "PR_M_LINE_NO", "order": "LINE_NO",
              "cols": [("LINE_NO", "라인번호"), ("APPLY_YMD", "적용일"), ("MAINT_DAY", "리드(일)"),
                       ("MAINT_HHMM", "시각"), ("LINK_CUST_CODE", "연계거래처"), ("CUST_MAINT_DAY", "거래처리드")]},
    "assem": {"t": "CS_M_ASSEM_PROC", "title": "조립공정MASTER", "src": "CS_M_ASSEM_PROC", "order": "SORT_SEQ",
              "cols": [("ASSEM_PROC_CODE", "공정코드"), ("ASSEM_PROC_DESC", "공정명"), ("STD_ST", "표준ST"),
                       ("SORT_SEQ", "정렬"), ("USE_FLAG", "사용")]},
    "proc":  {"t": "CS_M_PROC", "title": "단품공정MASTER", "src": "CS_M_PROC", "order": "SORT_SEQ",
              "cols": [("PROC_CODE", "공정코드"), ("PROC_DESC", "공정명"), ("ITEM_LGROUP", "대분류"),
                       ("SORT_SEQ", "정렬"), ("PROD_UPH", "표준UPH"), ("USE_FLAG", "사용")]},
}
def _basemaster_partner(q):
    """거래처MASTER(라이브 CM_M_CUST). 레거시 w_cm_master_055. 코드→이름: CUST_TYPE=PR011 거래처구분, 역할=IN/OUT/OUTSIDE 플래그."""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT DETAIL_CODE, DETAIL_DESC FROM CM_M_MASTER_DETAIL WHERE KIND_CODE='PR011'")
        dec = {str(r[0]).strip(): str(r[1]).strip() for r in cur.fetchall()}
        w = ""; p = []
        if q.strip():
            w = " AND (CUST_CODE LIKE ? OR CUST_DESC LIKE ? OR ISNULL(OWNER_NAME,'') LIKE ?)"
            p = [f"%{q.strip()}%"] * 3
        cur.execute(f"""SELECT CUST_CODE, ISNULL(CUST_DESC,''), ISNULL(OWNER_NAME,''), ISNULL(BUSINESS_NO,''),
              ISNULL(CUST_TYPE,''), ISNULL(IN_FLAG,'0'), ISNULL(OUT_FLAG,'0'), ISNULL(OUTSIDE_FLAG,'0'),
              ISNULL(BUSI_TYPE,''), ISNULL(BUSI_KIND,''), ISNULL(CHARGE_USER_ID,''), ISNULL(PHONE_NO,''),
              ISNULL(FAX_NO,''), ISNULL(ADDRESS,''), ISNULL(DLVY_DAY,0), ISNULL(DLVY_DAY2,0),
              ISNULL(SET_IN_FLAG,'0'), ISNULL(SAGUB_OUT_FLAG,'0'), ISNULL(HEAT_LABEL_FLAG,'0'),
              ISNULL(USE_FLAG,'0'), ISNULL(REMARKS,'')
            FROM CM_M_CUST WHERE CUST_CODE>''{w} ORDER BY CUST_CODE""", *p)
        yn = lambda v: 'Y' if str(v).strip() == '1' else ''
        rows = []
        for r in cur.fetchall():
            roles = []
            if str(r[5]).strip() == '1': roles.append('매입')
            if str(r[6]).strip() == '1': roles.append('매출')
            if str(r[7]).strip() == '1': roles.append('외주')
            rows.append([
                str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip(), str(r[3]).strip(),
                dec.get(str(r[4]).strip(), str(r[4]).strip()), '·'.join(roles),
                str(r[8]).strip(), str(r[9]).strip(), str(r[10]).strip(), str(r[11]).strip(),
                str(r[12]).strip(), str(r[13]).strip(), str(int(r[14] or 0)), str(int(r[15] or 0)),
                yn(r[16]), yn(r[17]), yn(r[18]), '사용' if str(r[19]).strip() == '1' else '중지', str(r[20]).strip(),
            ])
        headers = ['거래처코드', '상호', '대표자', '사업자번호', '거래처구분', '역할', '업태', '종목', '담당자',
                   '전화', '팩스', '주소', '납기일', '납기일2', '세트입고', '사급출고', '열처리라벨', '사용', '비고']
        return {"kind": "partner", "title": "거래처MASTER", "table": "CM_M_CUST",
                "headers": headers, "rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

@app.get("/api/basemaster/list")
def basemaster_list(kind: str = Query("dept"), q: str = Query("")):
    """기준MASTER 라이브 조회(읽기전용). kind=partner/dept/line/assem/proc."""
    if kind == "partner":
        return _basemaster_partner(q)
    m = _BASEMASTER.get(kind)
    if not m:
        raise HTTPException(400, "알 수 없는 마스터 종류")
    cn = _conn(); cur = cn.cursor()
    try:
        sel = ", ".join(f"ISNULL(CAST([{col}] AS NVARCHAR(120)),'') c{i}" for i, (col, _) in enumerate(m["cols"]))
        w = ""
        p = []
        if q.strip():
            code_col, nm_col = m["cols"][0][0], m["cols"][1][0]
            w = f" WHERE [{code_col}] LIKE ? OR [{nm_col}] LIKE ?"
            p = [f"%{q.strip()}%"] * 2
        cur.execute(f"SELECT {sel} FROM {m['t']}{w} ORDER BY [{m['order']}]", *p)
        rows = [list(r) for r in cur.fetchall()]
        return {"kind": kind, "title": m["title"], "table": m["src"],
                "headers": [h for _, h in m["cols"]], "rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

# ---- 달력 마스터(근무/라인별/파트별) — 엔티티+기간 필터, 요일/근무 파생 ----
# 소스근거(w_pr_plan_020): work_stats in('1','2','5','6')=근무일, '4'=제외(비근무). WEEKLY 1=일~7=토.
_CAL = {
    "cal_work": {"t": "HR_M_CALENDAR", "title": "근무달력MASTER", "ent": "WORK_TEAM", "entlbl": "근무팀", "date": "CALENDAR_YYMD", "d8": True},
    "cal_line": {"t": "PR_M_LINE_CALENDAR", "title": "라인별 달력관리", "ent": "LINE_NO", "entlbl": "라인", "date": "CALENDAR_YMD", "d8": False},
    "cal_part": {"t": "PR_M_PART_CALENDAR", "title": "파트별 달력관리", "ent": "PART_CODE", "entlbl": "파트", "date": "CALENDAR_YMD", "d8": False},
}
_WEEKDAY = ["", "일", "월", "화", "수", "목", "금", "토"]
def _wstats(v):
    s = str(v).strip()
    return "근무" if s in ("1", "2", "5", "6") else ("휴무" if s == "4" else "기타")
@app.get("/api/basemaster/cal")
def basemaster_cal(kind: str = Query("cal_line"), ent: str = Query(""),
                   from_ymd: str = Query(""), to_ymd: str = Query("")):
    """달력 마스터 라이브 조회(읽기전용). 엔티티(팀/라인/파트)+기간 필터."""
    m = _CAL.get(kind)
    if not m:
        raise HTTPException(400, "알 수 없는 달력 종류")
    cn = _conn(); cur = cn.cursor()
    try:
        dcol, ecol = m["date"], m["ent"]
        cur.execute(f"SELECT DISTINCT [{ecol}] FROM {m['t']} WHERE [{ecol}] IS NOT NULL ORDER BY [{ecol}]")
        ents = [str(r[0]).strip() for r in cur.fetchall() if str(r[0]).strip()]
        def cvt(s):
            d = "".join(ch for ch in str(s or "") if ch.isdigit())[-8:]
            if m["d8"]:
                return d if len(d) == 8 else (("20" + d[-6:]) if len(d) >= 6 else d)
            return d[-6:] if len(d) >= 6 else d
        w = ["1=1"]; p = []
        if from_ymd: w.append(f"[{dcol}]>=?"); p.append(cvt(from_ymd))
        if to_ymd:   w.append(f"[{dcol}]<=?"); p.append(cvt(to_ymd))
        if ent.strip(): w.append(f"[{ecol}]=?"); p.append(ent.strip())
        cur.execute(f"""SELECT TOP 3000 [{ecol}] ent, [{dcol}] ymd, ISNULL(WEEKLY,0) wk,
              ISNULL(WORK_STATS,'') ws, ISNULL(REMARKS,'') remarks
            FROM {m['t']} WHERE {' AND '.join(w)} ORDER BY [{dcol}] DESC""", *p)
        rows = []
        for r in cur.fetchall():
            ymd = str(r[1] or ""); ymd6 = ymd[2:] if len(ymd) == 8 else ymd
            try: wk = int(r[2] or 0)
            except Exception: wk = 0
            rows.append({"ent": str(r[0] or "").strip(), "ymd": ymd6,
                         "weekday": _WEEKDAY[wk] if 1 <= wk <= 7 else "",
                         "ws": str(r[3] or "").strip(), "ws_nm": _wstats(r[3]),
                         "remarks": str(r[4] or "").strip()})
        work = sum(1 for r in rows if r["ws_nm"] == "근무")
        return {"kind": kind, "title": m["title"], "table": m["t"], "entlbl": m["entlbl"],
                "ents": ents, "rows": rows, "cnt": len(rows), "work_days": work}
    finally:
        cn.close()


# ============ 기준정보: 거래처MASTER CRUD (nx.cust, 위하고정합) — 레거시 w_cm_master_055 ============
_BIZTAG = {"0": "개인", "1": "사업자", "2": "관공서", "3": "국외업체"}
def _valid_bizno(s):
    """사업자등록번호 체크섬(f_check_saupjano 대응). 빈값은 통과(선택)."""
    d = [c for c in str(s or "") if c.isdigit()]
    if not d: return True
    if len(d) != 10: return False
    d = [int(x) for x in d]; key = [1, 3, 7, 1, 3, 7, 1, 3, 5]
    tot = sum(d[i] * key[i] for i in range(9)) + (d[8] * 5) // 10
    return (10 - (tot % 10)) % 10 == d[9]

@app.get("/api/cust/opts")
def cust_opts():
    """거래처 드롭다운 소스: 거래처구분(PR011)·사업자구분·역할·결제조건·은행(CM701)."""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT DETAIL_CODE, DETAIL_DESC FROM CM_M_MASTER_DETAIL WHERE KIND_CODE='PR011' ORDER BY DETAIL_CODE")
        cust_type = [{"code": str(r[0]).strip(), "nm": str(r[1]).strip()} for r in cur.fetchall()]
        cur.execute("SELECT DETAIL_CODE, DETAIL_DESC FROM CM_M_MASTER_DETAIL WHERE KIND_CODE='CM701' ORDER BY DETAIL_CODE")
        banks = [{"code": str(r[0]).strip(), "nm": str(r[1]).strip()} for r in cur.fetchall()]
        return {"cust_type": cust_type,
                "biztag": [{"code": k, "nm": v} for k, v in _BIZTAG.items()],
                "yn": [{"code": "1", "nm": "예"}, {"code": "0", "nm": "아니오"}],
                "ue_date": [{"code": "0", "nm": "당월결제"}, {"code": "1", "nm": "1개월후"}, {"code": "2", "nm": "2개월후"}, {"code": "3", "nm": "3개월후"}, {"code": "4", "nm": "4개월후"}],
                "ue_week": [{"code": "10", "nm": "10일"}, {"code": "25", "nm": "25일"}, {"code": "31", "nm": "31일"}],
                "banks": banks}
    finally:
        cn.close()

@app.get("/api/cust/newcode")
def cust_newcode():
    """신규 거래처코드 = 숫자코드 최댓값+1 (레거시 wf_last_cust_code)."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("SELECT MAX(CAST(cust_code AS INT)) FROM nx.cust WHERE cust_code NOT LIKE '%[^0-9]%'")
        mx = cur.fetchone()[0] or 0
        return {"code": str(int(mx) + 1).zfill(6)}
    finally:
        nx.close()

@app.get("/api/cust/list")
def cust_list(q: str = Query(""), use: str = Query(""), ctype: str = Query("")):
    """거래처MASTER 목록(nx.cust). 코드→이름 디코드."""
    nx = _nx(); cur = nx.cursor()
    try:
        cur2 = _conn().cursor()
        cur2.execute("SELECT DETAIL_CODE, DETAIL_DESC FROM CM_M_MASTER_DETAIL WHERE KIND_CODE='PR011'")
        dec = {str(r[0]).strip(): str(r[1]).strip() for r in cur2.fetchall()}
        w = ["1=1"]; p = []
        if q.strip(): w.append("(cust_code LIKE ? OR cust_name LIKE ? OR owner_name LIKE ?)"); p += [f"%{q.strip()}%"] * 3
        if use in ("0", "1"): w.append("use_flag=?"); p.append(int(use))
        if ctype.strip(): w.append("cust_type=?"); p.append(ctype.strip())
        cur.execute(f"""SELECT cust_code,cust_name,biz_no,owner_name,biz_type,biz_item,cust_type,
            in_flag,out_flag,outside_flag,business_tag,tel,fax,address1,charge_name,charge_tel,charge_email,
            homepage,dlvy_day,dlvy_day2,ue_date,ue_week,use_flag,remarks,resident_no,bank_flag,
            recv_address,sagub_out_flag,set_in_flag,heat_label_flag,print_name,corp_no,charge_user_id,
            charge_rank,charge_hp,post_no,address2,credit_limit,collateral_amt,gc_gubun
            FROM nx.cust WHERE {' AND '.join(w)} ORDER BY cust_code""", *p)
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            roles = []
            if d["in_flag"]: roles.append("매입")
            if d["out_flag"]: roles.append("매출")
            if d["outside_flag"]: roles.append("외주")
            d["roles"] = "·".join(roles)
            d["cust_type_nm"] = dec.get(str(d["cust_type"] or "").strip(), str(d["cust_type"] or ""))
            d["biztag_nm"] = _BIZTAG.get(str(d["business_tag"] or "").strip(), "")
            rows.append(d)
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()

@app.post("/api/cust/save")
def cust_save(payload: dict = Body(...)):
    """거래처 등록/수정(nx.cust). 검증: 사업자번호 체크섬·거래처구분 필수·역할 최소1·코드중복."""
    p = payload
    code = str(p.get("cust_code", "")).strip()[:10]
    name = str(p.get("cust_name", "")).strip()[:50]
    if not code or not name:
        raise HTTPException(400, "거래처코드·거래처명은 필수입니다.")
    if not str(p.get("cust_type", "")).strip():
        raise HTTPException(400, "거래처구분을 선택해야 합니다.")
    if not (p.get("in_flag") or p.get("out_flag") or p.get("outside_flag")):
        raise HTTPException(400, "역할(매입/매출/외주) 최소 하나를 선택해야 합니다.")
    if not _valid_bizno(p.get("biz_no")):
        raise HTTPException(400, "사업자등록번호가 올바르지 않습니다.")
    def s(k, n): return str(p.get(k, "") or "").strip()[:n]
    def bit(k): return 1 if p.get(k) in (1, "1", True, "true") else 0
    def num(k):
        try: return int(float(p.get(k) or 0))
        except Exception: return 0
    is_new = not p.get("_edit")
    nx = _nx(); cur = nx.cursor()
    try:
        if is_new:
            cur.execute("SELECT 1 FROM nx.cust WHERE cust_code=?", code)
            if cur.fetchone():
                raise HTTPException(400, "동일한 거래처코드가 이미 등록되어 있습니다.")
        vals = (name, s("biz_no", 12), s("resident_no", 13), s("owner_name", 30), s("biz_type", 50), s("biz_item", 100),
                s("post_no", 6), s("address1", 100), s("address2", 100), s("tel", 50), s("fax", 20), s("print_name", 50),
                s("trade_start", 8), s("trade_end", 8), bit("use_flag"), s("dept_name", 30), s("charge_name", 30), s("charge_rank", 20),
                s("charge_tel", 20), s("charge_hp", 20), s("charge_email", 40), s("homepage", 50), num("credit_limit"), num("collateral_amt"),
                s("cust_type", 2), bit("in_flag"), bit("out_flag"), bit("outside_flag"), bit("bank_flag"), s("business_tag", 1),
                s("charge_user_id", 20), s("corp_no", 13), s("recv_post_no", 6), s("recv_address", 100), s("recv_address_dtl", 100),
                bit("sagub_out_flag"), bit("set_in_flag"), bit("heat_label_flag"), bit("prod_check_flag"),
                num("dlvy_day"), num("dlvy_day2"), s("ue_date", 2), s("ue_week", 2), s("ue_day", 2), s("gc_gubun", 10),
                s("bank_code", 10), s("bank_bookno", 20), s("bank_person_name", 30), s("cms_no", 20), s("remarks", 255),
                (s("user", 40) or "웹사용자"))
        setcols = ("cust_name=?,biz_no=?,resident_no=?,owner_name=?,biz_type=?,biz_item=?,post_no=?,address1=?,address2=?,"
                   "tel=?,fax=?,print_name=?,trade_start=?,trade_end=?,use_flag=?,dept_name=?,charge_name=?,charge_rank=?,"
                   "charge_tel=?,charge_hp=?,charge_email=?,homepage=?,credit_limit=?,collateral_amt=?,cust_type=?,in_flag=?,"
                   "out_flag=?,outside_flag=?,bank_flag=?,business_tag=?,charge_user_id=?,corp_no=?,recv_post_no=?,recv_address=?,"
                   "recv_address_dtl=?,sagub_out_flag=?,set_in_flag=?,heat_label_flag=?,prod_check_flag=?,dlvy_day=?,dlvy_day2=?,"
                   "ue_date=?,ue_week=?,ue_day=?,gc_gubun=?,bank_code=?,bank_bookno=?,bank_person_name=?,cms_no=?,remarks=?,upd_user=?")
        if is_new:
            cur.execute(
                "INSERT INTO nx.cust(cust_code,cust_name,biz_no,resident_no,owner_name,biz_type,biz_item,post_no,address1,address2,"
                "tel,fax,print_name,trade_start,trade_end,use_flag,dept_name,charge_name,charge_rank,charge_tel,charge_hp,charge_email,"
                "homepage,credit_limit,collateral_amt,cust_type,in_flag,out_flag,outside_flag,bank_flag,business_tag,charge_user_id,"
                "corp_no,recv_post_no,recv_address,recv_address_dtl,sagub_out_flag,set_in_flag,heat_label_flag,prod_check_flag,"
                "dlvy_day,dlvy_day2,ue_date,ue_week,ue_day,gc_gubun,bank_code,bank_bookno,bank_person_name,cms_no,remarks,upd_user,upd_dt) "
                "VALUES(" + ",".join(["?"] * 52) + ",getdate())", code, *vals)
            return {"ok": True, "mode": "insert", "cust_code": code}
        cur.execute(f"UPDATE nx.cust SET {setcols},upd_dt=getdate() WHERE cust_code=?", *vals, code)
        return {"ok": True, "mode": "update", "cust_code": code}
    finally:
        nx.close()

@app.post("/api/cust/delete")
def cust_delete(payload: dict = Body(...)):
    codes = [str(x).strip() for x in (payload.get("codes", []) or []) if str(x).strip()]
    if not codes: return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute(f"DELETE FROM nx.cust WHERE cust_code IN ({','.join('?'*len(codes))})", *codes)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()

# ---------- 부서MASTER CRUD (nx.dept, 레거시 w_hr_master_010) ----------
_DEPT_F = ["dept_desc", "sort_key", "dept_desch", "dept_from_ymd", "dept_to_ymd", "fin_dept_code",
           "fin_from_ymd", "fin_to_ymd", "enterprise_dept", "wh_code", "use_flag", "remarks"]
@app.get("/api/dept/list")
def dept_list(q: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        w = ""; p = []
        if q.strip(): w = " WHERE dept_code LIKE ? OR dept_desc LIKE ?"; p = [f"%{q.strip()}%"] * 2
        cur.execute(f"SELECT dept_code,dept_desc,sort_key,dept_desch,dept_from_ymd,dept_to_ymd,fin_dept_code,fin_from_ymd,fin_to_ymd,enterprise_dept,wh_code,use_flag,remarks FROM nx.dept{w} ORDER BY sort_key,dept_code", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()

@app.post("/api/dept/save")
def dept_save(payload: dict = Body(...)):
    p = payload
    code = str(p.get("dept_code", "")).strip()[:4]
    if not code or not str(p.get("dept_desc", "")).strip():
        raise HTTPException(400, "부서코드·부서명은 필수입니다.")
    def s(k, n): return str(p.get(k, "") or "").strip()[:n]
    def num(k):
        try: return int(float(p.get(k) or 0))
        except Exception: return 0
    vals = (s("dept_desc", 30), num("sort_key"), s("dept_desch", 30), s("dept_from_ymd", 8), s("dept_to_ymd", 8),
            s("fin_dept_code", 4), s("fin_from_ymd", 8), s("fin_to_ymd", 8), s("enterprise_dept", 2), s("wh_code", 2),
            1 if p.get("use_flag") in (1, "1", True) else 0, s("remarks", 100), s("user", 40) or "웹사용자")
    nx = _nx(); cur = nx.cursor()
    try:
        if not p.get("_edit"):
            cur.execute("SELECT 1 FROM nx.dept WHERE dept_code=?", code)
            if cur.fetchone(): raise HTTPException(400, "이미 등록된 부서코드입니다.")
            cur.execute("INSERT INTO nx.dept(dept_code,dept_desc,sort_key,dept_desch,dept_from_ymd,dept_to_ymd,fin_dept_code,fin_from_ymd,fin_to_ymd,enterprise_dept,wh_code,use_flag,remarks,upd_user,upd_dt) VALUES(?," + ",".join(["?"] * 13) + ",getdate())", code, *vals)
            return {"ok": True, "mode": "insert", "dept_code": code}
        cur.execute("UPDATE nx.dept SET dept_desc=?,sort_key=?,dept_desch=?,dept_from_ymd=?,dept_to_ymd=?,fin_dept_code=?,fin_from_ymd=?,fin_to_ymd=?,enterprise_dept=?,wh_code=?,use_flag=?,remarks=?,upd_user=?,upd_dt=getdate() WHERE dept_code=?", *vals, code)
        return {"ok": True, "mode": "update", "dept_code": code}
    finally:
        nx.close()

@app.post("/api/dept/delete")
def dept_delete(payload: dict = Body(...)):
    codes = [str(x).strip() for x in (payload.get("codes", []) or []) if str(x).strip()]
    if not codes: return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute(f"DELETE FROM nx.dept WHERE dept_code IN ({','.join('?'*len(codes))})", *codes)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()

# ---------- LINE-NO MASTER CRUD (nx.line_no, 레거시 w_pr_master_190) ----------
def _valid_hhmm(s):
    d = str(s or "").strip()
    if not d: return True
    return len(d) == 4 and d.isdigit() and int(d[:2]) < 24 and int(d[2:]) < 60

@app.get("/api/line/list")
def line_list(q: str = Query("")):
    nx = _nx(); cur = nx.cursor()
    try:
        w = ""; p = []
        if q.strip(): w = " WHERE l.line_no LIKE ?"; p = [f"%{q.strip()}%"]
        cur.execute(f"""SELECT l.line_no,l.apply_ymd,l.maint_day,l.maint_hhmm,l.link_cust_code,
              ISNULL(c.cust_name,'') link_cust_name, l.cust_maint_day
            FROM nx.line_no l LEFT JOIN nx.cust c ON c.cust_code=l.link_cust_code{w} ORDER BY l.line_no""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        nx.close()

# ==================================================================================
#  생산정보등록 (기준정보) — w_pr_master_090 우측 3패널 재구현
#  조회 = 라이브 PARTNER_ERP ∪ nx(nx우선).  편집/저장 = PARTNER_ERP_TEST3.nx 만.
#  원천/nx: PR_M_ITEM_ASSY_RT→nx.prodinfo_assy, PR_M_WORK_SINGLE→nx.prodinfo_single,
#          PR_M_ITEM_PROC_GAGONG→nx.prodinfo_proc, PR_M_ITEM_ST→nx.prodinfo_item_st
#  * _nx() 커넥션으로 크로스DB 조회(라이브는 PARTNER_ERP.dbo. 로 정규화, nx는 nx.).
# ==================================================================================
_JP_METHOD = {"J": "전표처리", "G": "가간판", "L": "라벨"}
_PROC_GUBUN_ASSY = {"1": "용접", "2": "검사", "3": "조립", "21": "검사1", "31": "조립1"}
# 단품 매트릭스 외경 컬럼(실재 9열, 6.35~28.00). ∅4.76/5.00은 원천 부재 → 프론트 표시만/공란.
_OD_COLS = [("st_635", "6.35"), ("st_794", "7.94"), ("st_952", "9.52"), ("st_127", "12.70"),
            ("st_1588", "15.88"), ("st_1905", "19.05"), ("st_22", "22.20"), ("st_254", "25.40"), ("st_28", "28.00")]

# 양산준비 문서 14종 고정목록(순서=화면 표시순). 코드=저장/경로 키, 이름=표시.
YANGSAN_DOCS = [
    ("SHIP_DWG1",    "(출하검사확인)도면1"),
    ("SHIP_DWG2",    "(출하검사확인)도면2"),
    ("QMAP",         "Q-map"),
    ("QC_FLOW",      "QC공정도"),
    ("HSMS",         "HSMS입력"),
    ("INSP_GUIDE",   "검사지도서"),
    ("XRF",          "XRF"),
    ("WORK_STD",     "작업표준서"),
    ("PROC_PHOTO1",  "(공정전표)사진1"),
    ("PROC_PHOTO2",  "(공정전표)사진2"),
    ("TEST_PLAN",    "시험기획서"),
    ("DEV_REPORT",   "개발완료보고서"),
    ("CHANGE_POINT", "변경점"),
    ("INSP_CERT",    "검사성적서"),
]
_YANGSAN_DOC_NM = dict(YANGSAN_DOCS)

@app.get("/api/prodinfo/search")
def prodinfo_search(q: str = Query("")):
    """품번 검색(라이브 PR_M_ITEM)."""
    cn = _nx(); cur = cn.cursor()
    try:
        like = f"%{q.strip()}%"
        cur.execute("""SELECT TOP 60 ITEM_CODE, ISNULL(ITEM_DESC,''), ISNULL(ITEM_SPEC,''),
              ISNULL(ITEM_DIAM,0), ISNULL(ITEM_THICK,0), ISNULL(ITEM_LENGTH,0), ISNULL(PROD_RATE,100)
            FROM PARTNER_ERP.dbo.PR_M_ITEM
            WHERE ITEM_CODE LIKE ? OR ITEM_DESC LIKE ? ORDER BY ITEM_CODE""", like, like)
        rows = [{"item": r[0], "name": r[1], "spec": r[2], "diam": float(r[3] or 0),
                 "thick": float(r[4] or 0), "length": float(r[5] or 0), "prod_rate": float(r[6] or 0)}
                for r in cur.fetchall()]
        return {"rows": rows, "cnt": len(rows)}
    finally:
        cn.close()

def _pi_proc_rows(cur, item, use_nx):
    """생산공정순서 행 조회(use_nx=True→nx.prodinfo_proc, False→레거시). 마스터 조인으로 표시명·회수율 포함."""
    src = ("nx.prodinfo_proc a" if use_nx
           else "PARTNER_ERP.dbo.PR_M_ITEM_PROC_GAGONG a")
    C = (lambda c: c.lower()) if use_nx else (lambda c: c)  # nx는 소문자 컬럼
    cur.execute(f"""
        SELECT a.{C('PROC_SEQ')}, ISNULL(a.{C('WORK_CODE')},'') , ISNULL(a.{C('GAGONG_PROC_CODE')},''),
               ISNULL(a.{C('S_WORK_CODE')},0), ISNULL(a.{C('MACH_CODE')},''), ISNULL(a.{C('WORK_QTY')},0),
               ISNULL(a.{C('STD_SIZE')},''), ISNULL(a.{C('GAGONG_PROC_SEQ')},1), ISNULL(a.{C('READY_ST')},0),
               ISNULL(a.{C('MACH_CT')},0), ISNULL(a.{C('INWON')},0), ISNULL(a.{C('HUMAN_ST')},0),
               ISNULL(a.{C('TOT_ST')},0), ISNULL(a.{C('JP_PROC_METHOD')},'J'), ISNULL(a.{C('LT_HR')},0),
               ISNULL(w.WORK_DESC,''), ISNULL(g.GAGONG_PROC_DESC,''), ISNULL(s.WORK_DESC,''), ISNULL(m.MACH_DESC,''),
               ISNULL(g.PROD_RATE,0), ISNULL(w.PROD_RATE,0)
        FROM {src}
        LEFT JOIN PARTNER_ERP.dbo.PR_M_WORK        w ON w.WORK_CODE        = a.{C('WORK_CODE')}
        LEFT JOIN PARTNER_ERP.dbo.PR_M_PROC_GAGONG g ON g.GAGONG_PROC_CODE = a.{C('GAGONG_PROC_CODE')}
        LEFT JOIN PARTNER_ERP.dbo.PR_M_WORK_SINGLE s ON s.S_WORK_CODE      = a.{C('S_WORK_CODE')}
        LEFT JOIN PARTNER_ERP.dbo.QA_M_MACHINE     m ON m.MACH_CODE        = a.{C('MACH_CODE')}
        WHERE a.{C('ITEM_CODE')}=? ORDER BY a.{C('PROC_SEQ')}""", item)
    out = []
    for r in cur.fetchall():
        out.append({"proc_seq": int(r[0]), "work_code": str(r[1]).strip(), "gagong_proc_code": str(r[2]).strip(),
                    "s_work_code": int(r[3] or 0), "mach_code": str(r[4]).strip(), "work_qty": float(r[5] or 0),
                    "std_size": str(r[6]), "gagong_proc_seq": int(r[7] or 1), "ready_st": float(r[8] or 0),
                    "mach_ct": float(r[9] or 0), "inwon": int(r[10] or 0), "human_st": float(r[11] or 0),
                    "tot_st": float(r[12] or 0), "jp_proc_method": str(r[13]).strip() or "J", "lt_hr": float(r[14] or 0),
                    "work_desc": str(r[15]).strip(), "part_desc": str(r[16]).strip(), "s_work_desc": str(r[17]).strip(),
                    "mach_desc": str(r[18]).strip(), "part_rate": float(r[19] or 0), "work_rate": float(r[20] or 0)})
    return out

@app.get("/api/prodinfo/get")
def prodinfo_get(item: str = Query(...), assyall: int = Query(0)):
    """품번의 3패널 + 하단 탭 데이터 로드(nx우선 병합)."""
    item = item.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""SELECT ISNULL(ITEM_DESC,''), ISNULL(ITEM_SPEC,''), ISNULL(ITEM_DIAM,0),
              ISNULL(ITEM_THICK,0), ISNULL(ITEM_LENGTH,0), ISNULL(PROD_RATE,100), ISNULL(JIG_CODE,''), ISNULL(JIG_KEEP_AREA,'')
            FROM PARTNER_ERP.dbo.PR_M_ITEM WHERE ITEM_CODE=?""", item)
        pi = cur.fetchone()
        if not pi:
            raise HTTPException(404, f"품번 {item} 없음")
        head = {"item": item, "name": pi[0], "spec": pi[1], "diam": float(pi[2] or 0), "thick": float(pi[3] or 0),
                "length": float(pi[4] or 0), "prod_rate": float(pi[5] or 0), "jig_code": str(pi[6]).strip(),
                "jig_area": str(pi[7]).strip()}

        # ── 패널① 조립(공정수): 마스터 전량 LEFT JOIN 품목 work_qty(nx우선). 기본=사용공정(qty>0) ──
        flt = "" if assyall else "WHERE ISNULL(nx.work_qty, rt.WORK_QTY) > 0"
        cur.execute(f"""
            SELECT a.A_WORK_CODE, ISNULL(a.WORK_DESC,''), ISNULL(a.WORK_ST,0), ISNULL(a.PROC_GUBUN,''),
                   ISNULL(a.WELDING_GUBUN,0), ISNULL(a.SORT_SEQ,0),
                   ISNULL(nx.work_qty, rt.WORK_QTY) AS work_qty,
                   CASE WHEN nx.item_code IS NOT NULL THEN 1 ELSE 0 END nx_flag
            FROM PARTNER_ERP.dbo.PR_M_WORK_ASSY a
            LEFT JOIN PARTNER_ERP.dbo.PR_M_ITEM_ASSY_RT rt ON rt.A_WORK_CODE=a.A_WORK_CODE AND rt.ITEM_CODE=?
            LEFT JOIN nx.prodinfo_assy nx ON nx.a_work_code=a.A_WORK_CODE AND nx.item_code=?
            {flt}
            ORDER BY a.SORT_SEQ, a.A_WORK_CODE""", item, item)
        assy = []
        for r in cur.fetchall():
            pg = str(r[3]).strip()
            assy.append({"a_work_code": int(r[0]), "work_desc": str(r[1]).strip(), "work_st": float(r[2] or 0),
                         "proc_gubun": pg, "proc_gubun_nm": _PROC_GUBUN_ASSY.get(pg, pg), "welding_gubun": int(r[4] or 0),
                         "sort_seq": int(r[5] or 0), "work_qty": (None if r[6] is None else float(r[6])),
                         "nx_flag": int(r[7])})
        cur.execute("SELECT COUNT(*) FROM PARTNER_ERP.dbo.PR_M_WORK_ASSY")
        assy_master_cnt = cur.fetchone()[0]

        # ── 패널② 단품(공정수) = 외경별 표준ST 매트릭스(전사 마스터, nx우선) ──
        stsel = ", ".join([f"ISNULL(n.{k}, s.{k.upper()})" for k, _ in _OD_COLS])
        cur.execute(f"""
            SELECT s.S_WORK_CODE, ISNULL(n.work_desc, s.WORK_DESC), ISNULL(s.WORK_CODE,''),
                   ISNULL(s.GAGONG_PROC_CODE,''), ISNULL(s.HOUR_PAY,0), ISNULL(s.CUTTING_PROC_FLAG,''),
                   ISNULL(s.SUB_WELD_FLAG,''), ISNULL(s.SORT_SEQ,0),
                   CASE WHEN n.s_work_code IS NOT NULL THEN 1 ELSE 0 END, {stsel}
            FROM PARTNER_ERP.dbo.PR_M_WORK_SINGLE s
            LEFT JOIN nx.prodinfo_single n ON n.s_work_code = s.S_WORK_CODE
            ORDER BY s.WORK_CODE, s.SORT_SEQ, s.S_WORK_CODE""")
        single = []
        for r in cur.fetchall():
            d = {"s_work_code": int(r[0]), "work_desc": str(r[1] or "").strip(), "work_code": str(r[2]).strip(),
                 "gagong_proc_code": str(r[3]).strip(), "hour_pay": int(r[4] or 0), "cutting_flag": str(r[5]).strip(),
                 "sub_weld_flag": str(r[6]).strip(), "sort_seq": int(r[7] or 0), "nx_flag": int(r[8])}
            for i, (k, _) in enumerate(_OD_COLS):
                v = r[9 + i]
                d[k] = (None if v is None else round(float(v), 3))
            single.append(d)

        # ── 패널③ 생산공정순서(nx우선 by item) ──
        cur.execute("SELECT COUNT(*) FROM nx.prodinfo_proc WHERE item_code=?", item)
        use_nx = cur.fetchone()[0] > 0
        proc = _pi_proc_rows(cur, item, use_nx)
        proc_src = "nx" if use_nx else "legacy"

        # ── 하단 탭: LOB(item_st, nx우선) ──
        cur.execute("""SELECT prod_gubun, ISNULL(member_qty,0), ISNULL(capa_qty,0), 'nx'
              FROM nx.prodinfo_item_st WHERE item_code=?
            UNION ALL
            SELECT ISNULL(PROD_GUBUN,''), ISNULL(MEMBER_QTY,0), ISNULL(CAPA_QTY,0), 'legacy'
              FROM PARTNER_ERP.dbo.PR_M_ITEM_ST
              WHERE ITEM_CODE=? AND ISNULL(PROD_GUBUN,'') NOT IN (SELECT prod_gubun FROM nx.prodinfo_item_st WHERE item_code=?)
            ORDER BY 1""", item, item, item)
        item_st = [{"prod_gubun": str(r[0]).strip(), "member_qty": int(r[1] or 0), "capa_qty": int(r[2] or 0),
                    "src": r[3]} for r in cur.fetchall()]

        # ── 하단 탭: 양산준비/지그(PR_M_ITEM_SUB 실측 후보 컬럼, 읽기전용 [재구성]) ──
        cur.execute("""SELECT ISNULL(PROD_STEP_MEMO,''), ISNULL(PROD_STEP_MEMO2,''), ISNULL(PROD_WORKER,''),
              ISNULL(INSP_WORKER,''), ISNULL(MAIN_MACH_CODE,''), ISNULL(ZIG_QTY,0), ISNULL(INSP_COUNT,0), ISNULL(ERR_RATE,0)
            FROM PARTNER_ERP.dbo.PR_M_ITEM_SUB WHERE ITEM_CODE=?""", item)
        sub = cur.fetchone()
        mach_nm = ""
        if sub and str(sub[4]).strip():
            cur.execute("SELECT ISNULL(MACH_DESC,'') FROM PARTNER_ERP.dbo.QA_M_MACHINE WHERE MACH_CODE=?", str(sub[4]).strip())
            mm = cur.fetchone(); mach_nm = str(mm[0]).strip() if mm else ""
        subd = ({"prod_step_memo": str(sub[0]), "prod_step_memo2": str(sub[1]), "prod_worker": str(sub[2]).strip(),
                 "insp_worker": str(sub[3]).strip(), "main_mach_code": str(sub[4]).strip(), "main_mach_nm": mach_nm,
                 "zig_qty": int(sub[5] or 0), "insp_count": int(sub[6] or 0), "err_rate": float(sub[7] or 0)}
                if sub else {})

        # ── 하단 탭 ⑦ 지그정보(nx 다행) + 레거시 단건 참조(읽기) ──
        cur.execute("""SELECT seq, ISNULL(jig_gubun,''), ISNULL(jig_qty,0), ISNULL(rack_loc,''), ISNULL(make_ymd,'')
            FROM nx.prodinfo_jig WHERE item_code=? ORDER BY seq""", item)
        jig = [{"seq": int(r[0]), "jig_gubun": str(r[1]), "jig_qty": int(r[2] or 0),
                "rack_loc": str(r[3]), "make_ymd": str(r[4]), "src": "nx"} for r in cur.fetchall()]
        jig_legacy = {"jig_code": head["jig_code"], "jig_area": head["jig_area"],
                      "zig_qty": (int(subd.get("zig_qty", 0)) if subd else 0)}

        # ── 하단 탭 ⑧ 수율(공정수) nx ──
        cur.execute("""SELECT seq, ISNULL(yield_proc,''), ISNULL(proc_qty,0), ISNULL(std_st,0), ISNULL(st,0)
            FROM nx.prodinfo_yield WHERE item_code=? ORDER BY seq""", item)
        yield_rows = [{"seq": int(r[0]), "yield_proc": str(r[1]), "proc_qty": float(r[2] or 0),
                       "std_st": float(r[3] or 0), "st": float(r[4] or 0)} for r in cur.fetchall()]

        # ── 하단 탭 ⑥ 양산준비 파일첨부(14종 × 첨부목록, nx) ──
        cur.execute("""SELECT doc_type, yid, orig_filename, byte_size, upd_user, upd_at
            FROM nx.prodinfo_yangsan WHERE item_code=? AND del_flag=0 ORDER BY doc_type, upd_at DESC""", item)
        yfiles = {}
        for r in cur.fetchall():
            yfiles.setdefault(str(r[0]), []).append({
                "yid": int(r[1]), "filename": r[2], "size": int(r[3] or 0), "user": r[4] or "",
                "dt": (r[5].isoformat() if hasattr(r[5], "isoformat") else str(r[5] or "")).replace("T", " ")[:19]})
        yangsan = [{"doc_type": c, "doc_nm": nm, "files": yfiles.get(c, [])} for c, nm in YANGSAN_DOCS]

        return {"head": head, "assy": assy, "assy_master_cnt": assy_master_cnt, "single": single,
                "proc": proc, "proc_src": proc_src, "item_st": item_st, "sub": subd,
                "jig": jig, "jig_legacy": jig_legacy, "yield": yield_rows, "yangsan": yangsan,
                "od_cols": [od for _, od in _OD_COLS]}
    finally:
        cn.close()

@app.get("/api/prodinfo/opts")
def prodinfo_opts(work_code: str = Query("")):
    """드롭다운 옵션 + 캐스케이드. work_code 지정 시 파트/가공공정/설비를 해당 작업처로 필터."""
    wc = work_code.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT WORK_CODE, ISNULL(WORK_DESC,'') FROM PARTNER_ERP.dbo.PR_M_WORK ORDER BY WORK_CODE")
        works = [{"code": str(r[0]).strip(), "name": str(r[1]).strip()} for r in cur.fetchall()]
        # 파트(PR_M_PROC_GAGONG) — work_code 포함(프론트 캐스케이드용)
        pw = "WHERE WORK_CODE=?" if wc else ""
        cur.execute(f"""SELECT GAGONG_PROC_CODE, ISNULL(GAGONG_PROC_DESC,''), ISNULL(WORK_CODE,'') FROM PARTNER_ERP.dbo.PR_M_PROC_GAGONG
            {pw} ORDER BY SORT_KEY, GAGONG_PROC_CODE""", *( [wc] if wc else [] ))
        parts = [{"code": str(r[0]).strip(), "name": str(r[1]).strip(), "work_code": str(r[2]).strip()} for r in cur.fetchall()]
        # 가공공정(PR_M_WORK_SINGLE, nx우선 명칭)
        cur.execute(f"""SELECT s.S_WORK_CODE, ISNULL(n.work_desc, s.WORK_DESC), ISNULL(s.WORK_CODE,'')
            FROM PARTNER_ERP.dbo.PR_M_WORK_SINGLE s LEFT JOIN nx.prodinfo_single n ON n.s_work_code=s.S_WORK_CODE
            {('WHERE s.WORK_CODE=?' if wc else '')} ORDER BY s.SORT_SEQ, s.S_WORK_CODE""", *( [wc] if wc else [] ))
        singles = [{"code": int(r[0]), "name": (str(r[1] or "").strip() or str(r[0])), "work_code": str(r[2]).strip()} for r in cur.fetchall()]
        # 설비(QA_M_MACHINE) — 작업처 지정 시 해당 작업처 + 미지정 설비 포함
        if wc:
            cur.execute("""SELECT TOP 400 MACH_CODE, ISNULL(MACH_DESC,''), ISNULL(WORK_CODE,'') FROM PARTNER_ERP.dbo.QA_M_MACHINE
                WHERE ISNULL(USE_FLAG,'1')='1' AND (ISNULL(WORK_CODE,'')='' OR WORK_CODE=?) ORDER BY MACH_DESC""", wc)
        else:
            cur.execute("""SELECT TOP 400 MACH_CODE, ISNULL(MACH_DESC,''), ISNULL(WORK_CODE,'') FROM PARTNER_ERP.dbo.QA_M_MACHINE
                WHERE ISNULL(USE_FLAG,'1')='1' ORDER BY MACH_DESC""")
        machs = [{"code": str(r[0]).strip(), "name": str(r[1]).strip(), "work_code": str(r[2]).strip()} for r in cur.fetchall()]
        return {"works": works, "parts": parts, "singles": singles, "machs": machs, "jp_methods": _JP_METHOD}
    finally:
        cn.close()

@app.post("/api/prodinfo/proc/save")
def prodinfo_proc_save(payload: dict = Body(...)):
    """생산공정순서 저장(nx.prodinfo_proc replace-all by item). 편집=nx만."""
    item = str(payload.get("item", "")).strip()
    if not item: return {"ok": False, "detail": "품번 필수"}
    rows = payload.get("rows", []) or []
    user = (payload.get("user") or "웹")[:30]
    cn = _nx(); cur = cn.cursor()
    try:
        seqs = set()
        for r in rows:
            s = int(r.get("proc_seq") or 0)
            if s <= 0: return {"ok": False, "detail": "공정SEQ는 1 이상 필수"}
            if s in seqs: return {"ok": False, "detail": f"공정SEQ 중복: {s}"}
            seqs.add(s)
        cur.execute("DELETE FROM nx.prodinfo_proc WHERE item_code=?", item)
        def n(r, k, d=0):
            v = r.get(k)
            try: return float(v) if v not in (None, "") else d
            except: return d
        for r in rows:
            cur.execute("""INSERT INTO nx.prodinfo_proc
                (item_code, proc_seq, work_code, gagong_proc_code, s_work_code, mach_code, work_qty, std_size,
                 mix_gagong, gagong_proc_flag, gagong_proc_seq, ready_st, mach_ct, inwon, human_st, tot_st,
                 jp_proc_method, lt_hr, key_id, upd_user, upd_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,getdate())""",
                item, int(r.get("proc_seq")), (r.get("work_code") or "")[:4], (r.get("gagong_proc_code") or "")[:10],
                int(n(r, "s_work_code")), (r.get("mach_code") or "")[:10], n(r, "work_qty"), (r.get("std_size") or "")[:100],
                int(n(r, "mix_gagong")), (r.get("gagong_proc_flag") or "")[:1], int(n(r, "gagong_proc_seq", 1)),
                n(r, "ready_st"), n(r, "mach_ct"), int(n(r, "inwon")), n(r, "human_st"), n(r, "tot_st"),
                (r.get("jp_proc_method") or "J")[:1], n(r, "lt_hr"), int(n(r, "key_id")), user)
        cn.commit()
        return {"ok": True, "saved": len(rows)}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()

@app.post("/api/prodinfo/assy/save")
def prodinfo_assy_save(payload: dict = Body(...)):
    """조립(공정수) 저장(nx.prodinfo_assy replace-all by item, work_qty>0만 보존)."""
    item = str(payload.get("item", "")).strip()
    if not item: return {"ok": False, "detail": "품번 필수"}
    rows = payload.get("rows", []) or []
    user = (payload.get("user") or "웹")[:30]
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("DELETE FROM nx.prodinfo_assy WHERE item_code=?", item)
        saved = 0
        for r in rows:
            try: q = int(float(r.get("work_qty")))
            except: q = 0
            aw = int(r.get("a_work_code") or 0)
            if aw and q > 0:
                cur.execute("""INSERT INTO nx.prodinfo_assy(item_code,a_work_code,work_qty,upd_user,upd_at)
                    VALUES(?,?,?,?,getdate())""", item, aw, q, user)
                saved += 1
        cn.commit()
        return {"ok": True, "saved": saved}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()

@app.post("/api/prodinfo/single/save")
def prodinfo_single_save(payload: dict = Body(...)):
    """단품 외경별 표준ST 마스터 저장(nx.prodinfo_single upsert by s_work_code). 전사 공유 마스터."""
    rows = payload.get("rows", []) or []
    user = (payload.get("user") or "웹")[:30]
    cn = _nx(); cur = cn.cursor()
    try:
        saved = 0
        for r in rows:
            sw = int(r.get("s_work_code") or 0)
            if not sw: continue
            def st(k):
                v = r.get(k)
                try: return float(v) if v not in (None, "") else None
                except: return None
            stvals = [st(k) for k, _ in _OD_COLS]
            cur.execute("SELECT 1 FROM nx.prodinfo_single WHERE s_work_code=?", sw)
            setclause = ", ".join([f"{k}=?" for k, _ in _OD_COLS])
            if cur.fetchone():
                cur.execute(f"""UPDATE nx.prodinfo_single SET work_desc=?, work_code=?, gagong_proc_code=?,
                    hour_pay=?, cutting_proc_flag=?, sub_weld_flag=?, sort_seq=?, {setclause},
                    upd_user=?, upd_at=getdate() WHERE s_work_code=?""",
                    (r.get("work_desc") or None), (r.get("work_code") or None), (r.get("gagong_proc_code") or None),
                    int(float(r.get("hour_pay") or 0)), (r.get("cutting_flag") or None), (r.get("sub_weld_flag") or None),
                    int(float(r.get("sort_seq") or 0)), *stvals, user, sw)
            else:
                cols = ", ".join([k for k, _ in _OD_COLS])
                ph = ", ".join(["?"] * len(_OD_COLS))
                cur.execute(f"""INSERT INTO nx.prodinfo_single
                    (s_work_code, work_desc, work_code, gagong_proc_code, hour_pay, cutting_proc_flag, sub_weld_flag, sort_seq, {cols}, upd_user, upd_at)
                    VALUES (?,?,?,?,?,?,?,?,{ph},?,getdate())""",
                    sw, (r.get("work_desc") or None), (r.get("work_code") or None), (r.get("gagong_proc_code") or None),
                    int(float(r.get("hour_pay") or 0)), (r.get("cutting_flag") or None), (r.get("sub_weld_flag") or None),
                    int(float(r.get("sort_seq") or 0)), *stvals, user)
            saved += 1
        cn.commit()
        return {"ok": True, "saved": saved}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()

@app.post("/api/prodinfo/itemst/save")
def prodinfo_itemst_save(payload: dict = Body(...)):
    """LOB분석(생산구분별 인원/CAPA) 저장(nx.prodinfo_item_st replace-all by item)."""
    item = str(payload.get("item", "")).strip()
    if not item: return {"ok": False, "detail": "품번 필수"}
    rows = payload.get("rows", []) or []
    user = (payload.get("user") or "웹")[:30]
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("DELETE FROM nx.prodinfo_item_st WHERE item_code=?", item)
        saved = 0
        for r in rows:
            pg = (r.get("prod_gubun") or "").strip()[:2]
            if not pg: continue
            cur.execute("""INSERT INTO nx.prodinfo_item_st(item_code,prod_gubun,member_qty,capa_qty,upd_user,upd_at)
                VALUES(?,?,?,?,?,getdate())""", item, pg, int(float(r.get("member_qty") or 0)),
                int(float(r.get("capa_qty") or 0)), user)
            saved += 1
        cn.commit()
        return {"ok": True, "saved": saved}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()

@app.post("/api/prodinfo/jig/save")
def prodinfo_jig_save(payload: dict = Body(...)):
    """지그정보 저장(nx.prodinfo_jig replace-all by item). 편집=nx만."""
    item = str(payload.get("item", "")).strip()
    if not item: return {"ok": False, "detail": "품번 필수"}
    rows = payload.get("rows", []) or []
    user = (payload.get("user") or "웹")[:30]
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("DELETE FROM nx.prodinfo_jig WHERE item_code=?", item)
        saved = 0
        for r in rows:
            gb = (r.get("jig_gubun") or "").strip()
            rk = (r.get("rack_loc") or "").strip()
            my = (r.get("make_ymd") or "").strip().replace("-", "")[:8]
            try: q = int(float(r.get("jig_qty") or 0))
            except: q = 0
            if not (gb or rk or my or q):
                continue
            cur.execute("""INSERT INTO nx.prodinfo_jig(item_code,seq,jig_gubun,jig_qty,rack_loc,make_ymd,upd_user,upd_at)
                VALUES(?,?,?,?,?,?,?,getdate())""", item, saved + 1, (gb[:40] or None), q,
                (rk[:40] or None), (my or None), user)
            saved += 1
        cn.commit()
        return {"ok": True, "saved": saved}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()


@app.post("/api/prodinfo/yield/save")
def prodinfo_yield_save(payload: dict = Body(...)):
    """수율(공정수) 저장(nx.prodinfo_yield replace-all by item). 편집=nx만."""
    item = str(payload.get("item", "")).strip()
    if not item: return {"ok": False, "detail": "품번 필수"}
    rows = payload.get("rows", []) or []
    user = (payload.get("user") or "웹")[:30]
    def fnum(v):
        try: return float(v) if v not in (None, "") else 0.0
        except: return 0.0
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("DELETE FROM nx.prodinfo_yield WHERE item_code=?", item)
        saved = 0
        for r in rows:
            nm = (r.get("yield_proc") or "").strip()
            pq, ss, s = fnum(r.get("proc_qty")), fnum(r.get("std_st")), fnum(r.get("st"))
            if not nm and not (pq or ss or s):
                continue
            cur.execute("""INSERT INTO nx.prodinfo_yield(item_code,seq,yield_proc,proc_qty,std_st,st,upd_user,upd_at)
                VALUES(?,?,?,?,?,?,?,getdate())""", item, saved + 1, (nm[:60] or None), pq, ss, s, user)
            saved += 1
        cn.commit()
        return {"ok": True, "saved": saved}
    except Exception as e:
        return {"ok": False, "detail": str(e)[:200]}
    finally:
        cn.close()


@app.get("/api/prodinfo/yangsan/list")
def prodinfo_yangsan_list(item: str = Query(...)):
    """양산준비 14종 문서 × 첨부파일 목록(nx.prodinfo_yangsan)."""
    it = item.strip()
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""SELECT doc_type, yid, orig_filename, byte_size, upd_user, upd_at
            FROM nx.prodinfo_yangsan WHERE item_code=? AND del_flag=0 ORDER BY doc_type, upd_at DESC""", it)
        m = {}
        for r in cur.fetchall():
            m.setdefault(str(r[0]), []).append({"yid": int(r[1]), "filename": r[2], "size": int(r[3] or 0),
                "user": r[4] or "", "dt": (r[5].isoformat() if hasattr(r[5], "isoformat") else str(r[5] or "")).replace("T", " ")[:19]})
        return {"rows": [{"doc_type": c, "doc_nm": nm, "files": m.get(c, [])} for c, nm in YANGSAN_DOCS]}
    finally:
        cn.close()


@app.post("/api/prodinfo/yangsan/upload")
async def prodinfo_yangsan_upload(file: UploadFile = File(...), item: str = Form(...),
                                  doc_type: str = Form(...), user: str = Form("웹")):
    """양산준비 문서 업로드(DOC_STORAGE_PATH\\YANGSAN\\item\\doc_type 저장 + nx.prodinfo_yangsan 메타)."""
    dt = doc_type.strip()
    if dt not in _YANGSAN_DOC_NM: raise HTTPException(400, f"문서구분 오류: {dt}")
    it = item.strip()
    if not it: raise HTTPException(400, "품번 필수")
    raw = await file.read()
    if not raw: raise HTTPException(400, "빈 파일입니다.")
    fname = file.filename or "file"
    ext = ((fname.rsplit(".", 1)[-1] if "." in fname else "") or "").lower()[:10]
    sha = _hashlib.sha256(raw).hexdigest()
    sub = _os.path.join("YANGSAN", it, dt)
    d = _os.path.join(DOC_STORAGE_PATH, sub)
    try:
        _os.makedirs(d, exist_ok=True)
    except Exception as e:
        raise HTTPException(500, f"저장경로 생성 실패({DOC_STORAGE_PATH}): {e}")
    safe = f"{sha[:12]}_{fname}"
    with open(_os.path.join(d, safe), "wb") as fp: fp.write(raw)
    rel = _os.path.join(sub, safe)
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("""INSERT INTO nx.prodinfo_yangsan(item_code,doc_type,orig_filename,storage_uri,ext,byte_size,sha256,del_flag,upd_user,upd_at)
            OUTPUT INSERTED.yid VALUES(?,?,?,?,?,?,?,0,?,getdate())""",
            it, dt, fname, rel, ext, len(raw), sha, (user or "웹")[:30])
        yid = cur.fetchone()[0]
        cn.commit()
        return {"ok": True, "yid": int(yid), "size": len(raw), "path": rel}
    finally:
        cn.close()


@app.get("/api/prodinfo/yangsan/download")
def prodinfo_yangsan_download(yid: int = Query(...), disp: str = Query("attach")):
    """양산준비 첨부 다운로드/열기(disp=inline이면 브라우저 뷰)."""
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT orig_filename, storage_uri FROM nx.prodinfo_yangsan WHERE yid=? AND del_flag=0", int(yid))
        r = cur.fetchone()
        if not r: raise HTTPException(404, "첨부 없음")
        path = _os.path.join(DOC_STORAGE_PATH, r[1]); fname = r[0] or "file"
        if not _os.path.exists(path): raise HTTPException(404, f"파일 없음: {r[1]}")
        with open(path, "rb") as fp: data = fp.read()
    finally:
        cn.close()
    mime = _mimetypes.guess_type(fname)[0] or "application/octet-stream"
    cd = "inline" if str(disp).lower() == "inline" else "attachment"
    return Response(content=data, media_type=mime,
                    headers={"Content-Disposition": f"{cd}; filename*=UTF-8''{_urlquote(fname)}"})


@app.post("/api/prodinfo/yangsan/delete")
def prodinfo_yangsan_delete(payload: dict = Body(...)):
    """양산준비 첨부 삭제(soft delete + 파일 제거)."""
    yid = payload.get("yid")
    if not yid: return {"ok": False, "detail": "yid 필요"}
    cn = _nx(); cur = cn.cursor()
    try:
        cur.execute("SELECT storage_uri FROM nx.prodinfo_yangsan WHERE yid=? AND del_flag=0", int(yid))
        r = cur.fetchone()
        if not r: return {"ok": False, "detail": "첨부 없음"}
        cur.execute("UPDATE nx.prodinfo_yangsan SET del_flag=1, upd_at=getdate() WHERE yid=?", int(yid))
        cn.commit()
        try:
            fp = _os.path.join(DOC_STORAGE_PATH, r[0])
            if _os.path.exists(fp): _os.remove(fp)
        except Exception: pass
        return {"ok": True}
    finally:
        cn.close()


@app.post("/api/line/save")
def line_save(payload: dict = Body(...)):
    p = payload
    code = str(p.get("line_no", "")).strip()[:20]
    if not code:
        raise HTTPException(400, "라인번호는 필수입니다.")
    if not _valid_hhmm(p.get("maint_hhmm")):
        raise HTTPException(400, "변경시각은 HHMM(4자리, 시<24·분<60) 형식이어야 합니다.")
    link = str(p.get("link_cust_code", "") or "").strip()[:10]
    def num(k):
        try: return int(float(p.get(k) or 0))
        except Exception: return 0
    nx = _nx(); cur = nx.cursor()
    try:
        if link:
            cur.execute("SELECT 1 FROM nx.cust WHERE cust_code=?", link)
            if not cur.fetchone(): raise HTTPException(400, "등록되어 있지 않은 거래처입니다.")
        vals = (str(p.get("apply_ymd", "") or "").strip()[:6], num("maint_day"), str(p.get("maint_hhmm", "") or "").strip()[:4],
                link, num("cust_maint_day"), str(p.get("user", "") or "웹사용자")[:40])
        if not p.get("_edit"):
            cur.execute("SELECT 1 FROM nx.line_no WHERE line_no=?", code)
            if cur.fetchone(): raise HTTPException(400, "이미 등록된 라인번호입니다.")
            cur.execute("INSERT INTO nx.line_no(line_no,apply_ymd,maint_day,maint_hhmm,link_cust_code,cust_maint_day,upd_user,upd_dt) VALUES(?,?,?,?,?,?,?,getdate())", code, *vals)
            return {"ok": True, "mode": "insert", "line_no": code}
        cur.execute("UPDATE nx.line_no SET apply_ymd=?,maint_day=?,maint_hhmm=?,link_cust_code=?,cust_maint_day=?,upd_user=?,upd_dt=getdate() WHERE line_no=?", *vals, code)
        return {"ok": True, "mode": "update", "line_no": code}
    finally:
        nx.close()

@app.post("/api/line/delete")
def line_delete(payload: dict = Body(...)):
    codes = [str(x).strip() for x in (payload.get("codes", []) or []) if str(x).strip()]
    if not codes: return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute(f"DELETE FROM nx.line_no WHERE line_no IN ({','.join('?'*len(codes))})", *codes)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()

# ---------- 라인별달력 (LG 라인스케줄 엑셀 업로드 → nx.line_calendar) : 생산계획 가동캘린더 ----------
_WD_KO = ['월', '화', '수', '목', '금', '토', '일']
def _ensure_linecal():
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute("IF OBJECT_ID('nx.line_calendar') IS NULL CREATE TABLE nx.line_calendar (line_no NVARCHAR(20), cal_ymd DATE, work_code NVARCHAR(20), note NVARCHAR(50), upd_dt DATETIME DEFAULT GETDATE(), CONSTRAINT pk_lcal PRIMARY KEY(line_no,cal_ymd))")
        cur.execute("IF OBJECT_ID('nx.line_cal_event') IS NULL CREATE TABLE nx.line_cal_event (cal_ymd DATE, event NVARCHAR(50), CONSTRAINT pk_lcev PRIMARY KEY(cal_ymd,event))")
        cur.execute("IF OBJECT_ID('nx.line_cal_meta') IS NULL CREATE TABLE nx.line_cal_meta (line_no NVARCHAR(20) PRIMARY KEY, sort_ord INT, gubun NVARCHAR(20), model_no NVARCHAR(20), jindo NVARCHAR(30), upd_dt DATETIME DEFAULT GETDATE())")
    finally:
        nx.close()

@app.post("/api/linecal/upload")
async def linecal_upload(file: UploadFile = File(...), anchor_ymd: str = Form(...)):
    """LG 라인스케줄 엑셀 업로드. anchor_ymd(YYYY-MM-DD, 기준일)로 날짜 앵커링 → '잔업' 시트 파싱 → 덮어쓰기."""
    import sys as _sys, os as _os, datetime as _dt
    _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), '..', '..', '_schema'))
    from linecal_parser import parse_line_schedule
    d = "".join(ch for ch in str(anchor_ymd) if ch.isdigit())
    if len(d) != 8: raise HTTPException(400, "기준일(YYYY-MM-DD)을 입력하세요.")
    anchor = _dt.date(int(d[:4]), int(d[4:6]), int(d[6:8]))
    data = await file.read()
    try:
        res = parse_line_schedule(data, anchor)
    except Exception as e:
        raise HTTPException(400, f"엑셀 파싱 실패: {e}")
    _ensure_linecal()
    nx = _nx(); cur = nx.cursor()
    try:
        df, dt = res["date_from"].isoformat(), res["date_to"].isoformat()
        cur.execute("DELETE FROM nx.line_calendar WHERE cal_ymd BETWEEN ? AND ?", df, dt)
        cur.execute("DELETE FROM nx.line_cal_event WHERE cal_ymd BETWEEN ? AND ?", df, dt)
        for x in res["recs"]:
            cur.execute("INSERT INTO nx.line_calendar(line_no,cal_ymd,work_code) VALUES(?,?,?)", x["line_no"], x["ymd"].isoformat(), x["code"])
        for x in res["events"]:
            cur.execute("INSERT INTO nx.line_cal_event(cal_ymd,event) VALUES(?,?)", x["ymd"].isoformat(), x["event"])
        for x in res["meta"]:
            cur.execute("DELETE FROM nx.line_cal_meta WHERE line_no=?", x["line_no"])
            cur.execute("INSERT INTO nx.line_cal_meta(line_no,sort_ord,gubun,model_no,jindo) VALUES(?,?,?,?,?)", x["line_no"], x["sort_ord"], x["gubun"], x["model_no"], x["jindo"])
        return {"ok": True, "recs": len(res["recs"]), "events": len(res["events"]), "lines": len(res["meta"]),
                "date_from": df, "date_to": dt, "anchor": anchor.isoformat()}
    finally:
        nx.close()

@app.get("/api/linecal/matrix")
def linecal_matrix(from_ymd: str = Query(""), weeks: int = Query(4)):
    """라인별달력 매트릭스(라인×날짜). from_ymd 없으면 데이터 최신 앵커 주 월요일부터."""
    import datetime as _dt
    _ensure_linecal()
    nx = _nx(); cur = nx.cursor()
    try:
        dd = "".join(ch for ch in from_ymd if ch.isdigit())
        if len(dd) >= 8:
            start = _dt.date(int(dd[:4]), int(dd[4:6]), int(dd[6:8]))
        else:
            cur.execute("SELECT MIN(cal_ymd) FROM nx.line_calendar WHERE cal_ymd >= DATEADD(day,-7,CAST(GETDATE() AS DATE))")
            base = cur.fetchone()[0]
            base = base if base else _dt.date.today()
            if not isinstance(base, _dt.date): base = _dt.date.fromisoformat(str(base)[:10])
            start = base - _dt.timedelta(days=base.weekday())
        weeks = max(1, min(int(weeks or 4), 8))
        dates = [start + _dt.timedelta(days=i) for i in range(weeks * 7)]
        end = dates[-1]
        cur.execute("SELECT CONVERT(varchar(10),cal_ymd,120),event FROM nx.line_cal_event WHERE cal_ymd BETWEEN ? AND ?", start.isoformat(), end.isoformat())
        ev = {}
        for r in cur.fetchall(): ev.setdefault(r[0], []).append(r[1])
        cur.execute("SELECT line_no,ISNULL(sort_ord,0),ISNULL(gubun,''),ISNULL(model_no,''),ISNULL(jindo,'') FROM nx.line_cal_meta ORDER BY sort_ord,line_no")
        metas = [{"line_no": r[0], "gubun": r[2], "model_no": r[3], "jindo": r[4]} for r in cur.fetchall()]
        cur.execute("SELECT line_no,CONVERT(varchar(10),cal_ymd,120),work_code FROM nx.line_calendar WHERE cal_ymd BETWEEN ? AND ?", start.isoformat(), end.isoformat())
        cells = {}
        for r in cur.fetchall(): cells.setdefault(r[0], {})[r[1]] = r[2]
        for m in metas: m["cells"] = cells.get(m["line_no"], {})
        return {"dates": [{"ymd": d.isoformat(), "dow": _WD_KO[d.weekday()], "day": d.day, "mon": d.month,
                           "events": ev.get(d.isoformat(), [])} for d in dates],
                "lines": metas, "from": start.isoformat(), "to": end.isoformat(), "weeks": weeks}
    finally:
        nx.close()


# ---------- 근무달력(nx.work_calendar)·파트별달력(nx.part_calendar) 매트릭스 + 편집 ----------
_PR004 = {'1': '출근(잔업2시간)', '2': '출근(정상근무)', '3': '일요일', '4': '휴무',
          '5': '출근(잔업3시간)', '6': '출근(잔업4시간)', '7': '출근(4시간근무)'}
# 공장운영 달력 파트 순서(사용자 지정, 레거시 PR_M_PROC_GAGONG 참고) (코드, 표시명, 들여쓰기)
_PART_ORDER = [
    ("P0005", "제품물류파트", 0), ("P0006", "부품물류파트", 0),
    ("P0003", "이지링크", 0), ("P0002", "가공", 0),
    ("S5", "01라인(용접)", 0), ("S5-2", "01라인(조립)", 0), ("S1", "02라인", 0),
    ("S6", "03라인", 0), ("S4", "04라인", 0), ("S11", "05라인", 0), ("RAC", "06라인", 0),
    ("S8", "서포터(08)", 0), ("S10", "자동은납(10)", 0),
]
def _wcal_partnames():
    """PART_CODE→이름(PR_M_PROC_GAGONG), 이름 속 'PART'→'파트'."""
    cn = _conn(); cur = cn.cursor()
    try:
        cur.execute("SELECT GAGONG_PROC_CODE, GAGONG_PROC_DESC FROM PR_M_PROC_GAGONG")
        return {str(r[0]).strip(): str(r[1] or '').strip().replace('PART', '파트') for r in cur.fetchall()}
    finally:
        cn.close()

@app.get("/api/wcal/matrix")
def wcal_matrix(kind: str = Query("part"), from_ymd: str = Query(""), weeks: int = Query(4)):
    """근무/파트 달력 매트릭스(엔티티×날짜, WORK_STATS 디코드). 파트=공통행(근무달력) + 파트행."""
    import datetime as _dt
    nx = _nx(); cur = nx.cursor()
    try:
        dd = "".join(ch for ch in from_ymd if ch.isdigit())
        if len(dd) >= 8:
            start = _dt.date(int(dd[:4]), int(dd[4:6]), int(dd[6:8]))
        else:
            t = _dt.date.today(); start = t - _dt.timedelta(days=t.weekday())
        weeks = max(1, min(int(weeks or 4), 8))
        dates = [start + _dt.timedelta(days=i) for i in range(weeks * 7)]
        s, e = start.isoformat(), dates[-1].isoformat()
        # 공통(근무달력 team=A)
        cur.execute("SELECT CONVERT(varchar(10),cal_ymd,120),work_stats FROM nx.work_calendar WHERE cal_ymd BETWEEN ? AND ?", s, e)
        common = {r[0]: r[1] for r in cur.fetchall()}
        rows = [{"ent": "공통", "name": "공통", "indent": 0, "cells": common, "common": True}]
        if kind == "part":
            cur.execute("SELECT part_code,CONVERT(varchar(10),cal_ymd,120),work_stats FROM nx.part_calendar WHERE cal_ymd BETWEEN ? AND ?", s, e)
            pc = {}
            for r in cur.fetchall(): pc.setdefault(r[0], {})[r[1]] = r[2]
            for code, nm, indent in _PART_ORDER:   # 사용자 지정 순서·이름(코드 아님)
                rows.append({"ent": code, "name": nm, "indent": indent, "cells": pc.get(code, {}), "common": False})
        return {"kind": kind, "from": s, "to": e, "weeks": weeks,
                "dates": [{"ymd": d.isoformat(), "dow": _WD_KO[d.weekday()], "day": d.day, "mon": d.month} for d in dates],
                "rows": rows, "decode": _PR004}
    finally:
        nx.close()

@app.post("/api/wcal/save")
def wcal_save(payload: dict = Body(...)):
    """셀 편집 저장(upsert). kind=work→nx.work_calendar(team=A), part→nx.part_calendar. cells=[{ent,ymd,ws}]."""
    kind = payload.get("kind", "part")
    cells = payload.get("cells", []) or []
    tbl = "nx.work_calendar" if kind == "work" else "nx.part_calendar"
    keycol = "team" if kind == "work" else "part_code"
    nx = _nx(); cur = nx.cursor()
    try:
        n = 0
        for x in cells:
            ent = "A" if kind == "work" else str(x.get("ent", "")).strip()
            if kind == "work": ent = "A"
            elif str(x.get("ent", "")).strip() == "공통": ent = None
            ymd = str(x.get("ymd", "")).strip()[:10]
            ws = str(x.get("ws", "")).strip()[:2]
            if not ymd: continue
            # 공통행 편집 → work_calendar
            t2, k2, e2 = ("nx.work_calendar", "team", "A") if (kind == "work" or x.get("ent") == "공통") else (tbl, keycol, ent)
            if e2 is None: e2 = "A"; t2 = "nx.work_calendar"; k2 = "team"
            cur.execute(f"DELETE FROM {t2} WHERE {k2}=? AND cal_ymd=?", e2, ymd)
            if ws:
                cur.execute(f"INSERT INTO {t2}({k2},cal_ymd,work_stats,upd_user) VALUES(?,?,?,?)", e2, ymd, ws, str(payload.get("user", "") or "웹사용자")[:40])
            n += 1
        return {"ok": True, "saved": n}
    finally:
        nx.close()

# ============ 생산: 생산계획현황(라이브 SA_T_PLAN_DTL) — 제번×일자 피벗 ============
@app.get("/api/prodplan/status")
def prodplan_status(from_ymd: str = Query(""), to_ymd: str = Query(""), line: str = Query(""),
                    wo: str = Query(""), model: str = Query(""), cr: str = Query("")):
    """라이브 생산계획현황. SA_T_PLAN_DTL(현행 LG 생산계획)을 제번(WO)×일자로 피벗(읽기전용)."""
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("PLAN_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("PLAN_YMD<=?"); p.append(_d6(to_ymd))
        if line.strip():  w.append("LINE_NO=?"); p.append(line.strip())
        if wo.strip():    w.append("WORK_ORDER LIKE ?"); p.append(f"%{wo.strip()}%")
        if model.strip(): w.append("MODEL_NO LIKE ?"); p.append(f"%{model.strip()}%")
        if cr in ("C", "R"): w.append("CR_FLAG=?"); p.append(cr)
        cur.execute(f"""SELECT PLAN_YMD, WORK_ORDER, MODEL_NO, LINE_NO, ISNULL(PLAN_QTY,0) PLAN_QTY,
              ISNULL(LOT_QTY,0) LOT_QTY, ISNULL(TOOLS_DESC,'') TOOLS_DESC, ISNULL(CR_FLAG,'') CR_FLAG,
              ISNULL(OUTPUT_HM,'') OUTPUT_HM
            FROM SA_T_PLAN_DTL WHERE {' AND '.join(w)}""", *p)
        cols = [d[0] for d in cur.description]
        raw = [dict(zip(cols, row)) for row in cur.fetchall()]
        dates = sorted({r["PLAN_YMD"] for r in raw})
        wos = {}
        for r in raw:
            k = r["WORK_ORDER"]
            g = wos.get(k)
            if not g:
                g = {"wo": k, "model": r["MODEL_NO"], "line": r["LINE_NO"], "tool": r["TOOLS_DESC"],
                     "cr": r["CR_FLAG"], "hm": r["OUTPUT_HM"], "total": 0, "days": {}}
                wos[k] = g
            q = float(r["PLAN_QTY"] or 0)
            g["days"][r["PLAN_YMD"]] = g["days"].get(r["PLAN_YMD"], 0) + q
            g["total"] += q
        rows = sorted(wos.values(), key=lambda x: (x["line"] or "", x["wo"]))
        return {"dates": dates, "rows": rows, "wo_count": len(rows),
                "sum_qty": sum(float(r["PLAN_QTY"] or 0) for r in raw), "src": "SA_T_PLAN_DTL(라이브)"}
    finally:
        cn.close()


# ============ 일반업무: 공수등록(근무/지원) — HR_M_WORK_INFO(라이브)↔nx.hr_work_info ============
_HRCHK = {"1": "연차", "2": "반차", "3": "조퇴"}  # 소스 dw_pr_worktime_001_t2(4~6은 빈 라벨=정상)
def _hrchk(v):
    return _HRCHK.get(str(v or "").strip(), "정상")
@app.get("/api/gongsu/list")
def gongsu_list(from_ymd: str = Query(""), to_ymd: str = Query(""), dept: str = Query(""),
                gubun: str = Query(""), user: str = Query(""), src: str = Query("legacy")):
    """공수 조회. src=legacy(HR_M_WORK_INFO)/nx(nx.hr_work_info). gubun=근무/지원."""
    if src == "nx":
        nx = _nx(); cur = nx.cursor()
        try:
            w = ["1=1"]; p = []
            if from_ymd: w.append("work_ymd>=?"); p.append(_d6(from_ymd))
            if to_ymd:   w.append("work_ymd<=?"); p.append(_d6(to_ymd))
            if dept.strip(): w.append("dept_code=?"); p.append(dept.strip())
            if gubun.strip(): w.append("gubun=?"); p.append(gubun.strip())
            if user.strip(): w.append("user_id LIKE ?"); p.append(f"%{user.strip()}%")
            cur.execute(f"""SELECT id ID, gubun, work_ymd, dept_code, ISNULL(user_id,'') user_id, ISNULL(line,'') line,
                  ISNULL(start_time,'') start_time, ISNULL(end_time,'') end_time, ISNULL(work_hr,0) work_hr,
                  ISNULL(support_line,'') support_line, ISNULL(support_hr,0) support_hr, ISNULL(hr_check,'0') hr_check,
                  ISNULL(remarks,'') remarks FROM nx.hr_work_info WHERE {' AND '.join(w)}
                ORDER BY work_ymd DESC, id DESC""", *p)
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            for r in rows:
                r["src"] = "nx"; r["work_hr"] = float(r["work_hr"] or 0); r["support_hr"] = float(r["support_hr"] or 0)
                r["dept_nm"] = ""; r["hr_check_nm"] = _hrchk(r["hr_check"])
            return {"rows": rows, "cnt": len(rows), "sum_hr": sum(r["work_hr"] + r["support_hr"] for r in rows)}
        finally:
            nx.close()
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("A.WORK_YMD>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("A.WORK_YMD<=?"); p.append(_d6(to_ymd))
        if dept.strip(): w.append("A.DEPT_CODE=?"); p.append(dept.strip())
        if user.strip(): w.append("A.USER_ID LIKE ?"); p.append(f"%{user.strip()}%")
        if gubun == "지원": w.append("ISNULL(A.SUPPORT_SHEET_NO,0)>0")
        elif gubun == "근무": w.append("ISNULL(A.SUPPORT_SHEET_NO,0)=0")
        cur.execute(f"""SELECT TOP 3000
              CASE WHEN ISNULL(A.SUPPORT_SHEET_NO,0)>0 THEN '지원' ELSE '근무' END gubun,
              A.WORK_YMD work_ymd, ISNULL(A.DEPT_CODE,'') dept_code,
              COALESCE(NULLIF(D.DEPT_DESC,''), G.GAGONG_PROC_DESC, A.DEPT_CODE) dept_nm,
              ISNULL(A.USER_ID,'') user_id, ISNULL(A.CUST_CODE,'') line, ISNULL(A.START_TIME,'') start_time,
              ISNULL(A.END_TIME,'') end_time, ISNULL(A.WORK_HR,0) work_hr, ISNULL(A.SUPPORT_LINE,'') support_line,
              ISNULL(A.SUPPORT_HR,0) support_hr, ISNULL(A.HR_CHECK_POINT,'0') hr_check, ISNULL(A.REMARKS,'') remarks
            FROM HR_M_WORK_INFO A LEFT JOIN HR_M_DEPT D ON D.DEPT_CODE=A.DEPT_CODE
              LEFT JOIN PR_M_PROC_GAGONG G ON G.GAGONG_PROC_CODE COLLATE DATABASE_DEFAULT=A.DEPT_CODE COLLATE DATABASE_DEFAULT
            WHERE {' AND '.join(w)} ORDER BY A.WORK_YMD DESC, A.DEPT_CODE, A.MAINT_SEQ""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            r["src"] = "legacy"; r["ID"] = None
            r["work_hr"] = float(r["work_hr"] or 0); r["support_hr"] = float(r["support_hr"] or 0)
            r["remarks"] = str(r["remarks"]).strip(); r["hr_check_nm"] = _hrchk(r["hr_check"])
        return {"rows": rows, "cnt": len(rows), "sum_hr": sum(r["work_hr"] + r["support_hr"] for r in rows)}
    finally:
        cn.close()

@app.post("/api/gongsu/save")
def gongsu_save(payload: dict = Body(...)):
    p = payload
    ymd = _d6(str(p.get("work_ymd", "")))
    user = str(p.get("user_id", "")).strip()[:40]
    if not ymd or not user:
        raise HTTPException(400, "근무일·작업자는 필수입니다.")
    def s(k, n): return str(p.get(k, "")).strip()[:n]
    def f(k):
        try: return float(p.get(k) or 0)
        except Exception: return 0.0
    vals = ((s("gubun", 4) or "근무"), ymd, s("dept_code", 20), user, s("line", 20), s("start_time", 6),
            s("end_time", 6), f("work_hr"), s("support_line", 20), s("support_start", 6), s("support_end", 6),
            f("support_hr"), (s("hr_check", 4) or "0"), s("remarks", 200), (s("uuser", 40) or "웹사용자"))
    mid = p.get("id")
    nx = _nx(); cur = nx.cursor()
    try:
        if mid:
            cur.execute("""UPDATE nx.hr_work_info SET gubun=?,work_ymd=?,dept_code=?,user_id=?,line=?,start_time=?,
                end_time=?,work_hr=?,support_line=?,support_start=?,support_end=?,support_hr=?,hr_check=?,
                remarks=?,upd_user=?,upd_dt=getdate() WHERE id=?""", *vals, int(mid))
            return {"ok": True, "id": int(mid), "mode": "update"}
        cur.execute("""INSERT INTO nx.hr_work_info(gubun,work_ymd,dept_code,user_id,line,start_time,end_time,
            work_hr,support_line,support_start,support_end,support_hr,hr_check,remarks,upd_user)
            OUTPUT INSERTED.id VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", *vals)
        return {"ok": True, "id": int(cur.fetchone()[0]), "mode": "insert"}
    finally:
        nx.close()

@app.post("/api/gongsu/delete")
def gongsu_delete(payload: dict = Body(...)):
    ids = [int(x) for x in (payload.get("ids", []) or []) if str(x).strip()]
    if not ids: return {"ok": True, "deleted": 0}
    nx = _nx(); cur = nx.cursor()
    try:
        cur.execute(f"DELETE FROM nx.hr_work_info WHERE id IN ({','.join('?'*len(ids))})", *ids)
        return {"ok": True, "deleted": cur.rowcount}
    finally:
        nx.close()

# ============ 일반업무: 일일체크리스트(부서간 일일 이슈/체크, DAY_CHECK_LIST 라이브 조회) ============
@app.get("/api/daycheck/list")
def daycheck_list(from_ymd: str = Query(""), to_ymd: str = Query(""), dept: str = Query("")):
    """일일체크리스트 조회(읽기전용). ※원본 DAY_CHECK_LIST는 현재 과거이력(≈2016)만 보유."""
    cn = _conn(); cur = cn.cursor()
    try:
        w = ["1=1"]; p = []
        if from_ymd: w.append("check_ymd>=?"); p.append(_d6(from_ymd))
        if to_ymd:   w.append("check_ymd<=?"); p.append(_d6(to_ymd))
        if dept.strip(): w.append("check_dept LIKE ?"); p.append(f"%{dept.strip()}%")
        cur.execute(f"""SELECT TOP 2000 check_ymd ymd, check_seq seq, ISNULL(check_dept,'') dept,
              ISNULL(request_member,'') req, ISNULL(issue_item,'') item, ISNULL(issue_note,'') note,
              ISNULL(contents,'') contents, ISNULL(result_check,'') result, ISNULL(result_member,'') rmember,
              ISNULL(imp_check,'0') imp
            FROM DAY_CHECK_LIST WHERE {' AND '.join(w)} ORDER BY check_ymd DESC, check_seq DESC""", *p)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            for k in ("dept", "req", "item", "note", "contents", "rmember"): r[k] = str(r[k]).strip()
            r["imp"] = 1 if str(r["imp"]).strip() == "1" else 0
        cur.execute("SELECT MAX(check_ymd) FROM DAY_CHECK_LIST")
        return {"rows": rows, "cnt": len(rows), "max_ymd": cur.fetchone()[0],
                "note": "원본(DAY_CHECK_LIST)은 과거이력(≈2016)만 보유 — 현행 일일점검은 설비/안전 체크리스트 별도"}
    finally:
        cn.close()


# ===== 프론트엔드 정적 서빙 (내부망 단일 포트 운영) =====
# ★반드시 모든 API 라우트 정의 이후(파일 최하단)에 위치. /api/* · /live/* 가 먼저 매칭되고 나머지는 정적파일로.
# 프론트도 8010에서 서빙 → 브라우저 API_BASE=location.origin 자동일치(직원 PC 어디서 열어도 동작).
import os as _os
from fastapi.staticfiles import StaticFiles as _StaticFiles
_FRONT_DIR = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))   # backend의 상위 = PNC_ERP_Web
app.mount("/", _StaticFiles(directory=_FRONT_DIR, html=True), name="front")
