import os
import glob
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base_dir = r"c:\Users\hdy56\OneDrive\바탕 화면\파워빌드분석\src_extracted"
output_excel = r"c:\Users\hdy56\OneDrive\바탕 화면\파워빌드분석\일신ERP_시스템_소스분석_및_개선명세서.xlsx"

wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Styles
font_title = Font(name="맑은 고딕", size=16, bold=True, color="1F4E78")
font_subtitle = Font(name="맑은 고딕", size=11, italic=True, color="595959")
font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="맑은 고딕", size=10, bold=True)
font_regular = Font(name="맑은 고딕", size=10)
font_code = Font(name="Consolas", size=9)

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fill_accent = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

header_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='medium', color='1F4E78'),
    bottom=Side(style='medium', color='1F4E78')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_wrap_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ---------------------------------------------------------
# SHEET 1: 시스템 종합 개요 (Executive Summary)
# ---------------------------------------------------------
ws1 = wb.create_sheet(title="1. 시스템 종합 개요")
ws1.views.sheetView[0].showGridLines = True

ws1["A1"] = "📊 일신 ERP/MES 시스템 전체 소스 분석 및 개선 개요"
ws1["A1"].font = font_title
ws1["A2"] = "분석 일시: 2026-07-20 | 대상 버전: ilshinERP__260415 (55개 PBL 라이브러리)"
ws1["A2"].font = font_subtitle

summary_meta = [
    ("시스템 명칭", "일신 ERP / Partner ERP (제조/생산/자재/영업/품질/인사 통합 MES·ERP)"),
    ("개발 환경", "PowerBuilder 10.5 (Client/Server 2-Tier Architecture, ADO.Net DBMS)"),
    ("데이터베이스", "MS-SQL Server (PARTNER_ERP 운영 DB / 55개 주요 마스터 및 수불 테이블)"),
    ("추출 소스 통계", "총 55개 PBL 라이브러리 내 404개 소스 객체 (.srw, .srd, .sru, .srf 등 100% 텍스트 복원)"),
    ("주요 개선 목적", "차세대 Web API (Spring Boot / React) 기반 Modernization 및 모바일 스마트팩토리 개편 준비")
]

ws1.cell(row=4, column=1, value="[기본 시스템 사양 및 분석 정보]").font = font_bold
row_idx = 5
for k, v in summary_meta:
    ws1.cell(row=row_idx, column=1, value=k).font = font_bold
    ws1.cell(row=row_idx, column=1).fill = fill_sub_header
    ws1.cell(row=row_idx, column=1).alignment = align_center
    ws1.cell(row=row_idx, column=1).border = thin_border
    
    ws1.cell(row=row_idx, column=2, value=v).font = font_regular
    ws1.cell(row=row_idx, column=2).alignment = align_left
    ws1.cell(row=row_idx, column=2).border = thin_border
    row_idx += 1

row_idx += 2
ws1.cell(row=row_idx, column=1, value="[모듈별 라이브러리(PBL) 및 객체 분포 요약]").font = font_bold
row_idx += 1

headers_s1 = ["번호", "모듈 라이브러리(PBL) 명", "업무 분류", "전체 객체 수", "Window (.srw)", "DataWindow (.srd)", "UserObject (.sru)", "Function (.srf)"]
for col_idx, h in enumerate(headers_s1, start=1):
    cell = ws1.cell(row=row_idx, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = header_border

# Collect PBL Statistics
pbl_stats = []
for mod_dir in sorted(glob.glob(os.path.join(base_dir, "*"))):
    if not os.path.isdir(mod_dir): continue
    mod_name = os.path.basename(mod_dir)
    files = glob.glob(os.path.join(mod_dir, "*"))
    if not files: continue
    
    srw = len([f for f in files if f.endswith('.srw')])
    srd = len([f for f in files if f.endswith('.srd')])
    sru = len([f for f in files if f.endswith('.sru')])
    srf = len([f for f in files if f.endswith('.srf')])
    tot = len(files)
    
    cat = "공통/프레임워크"
    if "pr_" in mod_name or "ds_work" in mod_name: cat = "생산/BOM/외주"
    elif "pu_" in mod_name or "ma_" in mod_name: cat = "구매/자재 수불"
    elif "sa_" in mod_name or "cs_" in mod_name: cat = "영업/출하 매출"
    elif "qa_" in mod_name: cat = "품질 관리"
    elif "hr_" in mod_name: cat = "인사/근태/급여"
    elif "app_" in mod_name: cat = "시스템 앱/업데이터"
    
    pbl_stats.append((mod_name, cat, tot, srw, srd, sru, srf))

pbl_stats.sort(key=lambda x: x[2], reverse=True)

row_idx += 1
for idx, (mname, cat, tot, srw, srd, sru, srf) in enumerate(pbl_stats, start=1):
    r_vals = [idx, mname + ".pbl", cat, tot, srw, srd, sru, srf]
    for c_idx, val in enumerate(r_vals, start=1):
        cell = ws1.cell(row=row_idx, column=c_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border
        if c_idx in (1, 3): cell.alignment = align_center
        elif c_idx == 2: cell.alignment = align_left
        else:
            cell.alignment = align_right
            cell.number_format = "#,##0"
    row_idx += 1

# ---------------------------------------------------------
# SHEET 2: 소스 객체 상세 목록 (Object Catalog)
# ---------------------------------------------------------
ws2 = wb.create_sheet(title="2. 소스 객체 상세 목록")
ws2.views.sheetView[0].showGridLines = True

ws2["A1"] = "📋 복원된 파워빌드 소스 객체 전체 명세서"
ws2["A1"].font = font_title

headers_s2 = ["번호", "모듈명", "객체명 (파일명)", "객체 유형", "확장자", "한글 설명 / 타이틀", "파일 크기 (Byte)"]
for col_idx, h in enumerate(headers_s2, start=1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = header_border

row_idx = 4
obj_idx = 1
for mod_dir in sorted(glob.glob(os.path.join(base_dir, "*"))):
    if not os.path.isdir(mod_dir): continue
    mod_name = os.path.basename(mod_dir)
    files = glob.glob(os.path.join(mod_dir, "*"))
    
    for fpath in sorted(files):
        fname = os.path.basename(fpath)
        name_no_ext, ext = os.path.splitext(fname)
        fsize = os.path.getsize(fpath)
        
        obj_type = "Window (화면)"
        if ext == ".srd": obj_type = "DataWindow (쿼리/폼)"
        elif ext == ".sru": obj_type = "UserObject (공통객체)"
        elif ext == ".srf": obj_type = "Global Function (함수)"
        elif ext == ".sra": obj_type = "Application (앱)"
        elif ext == ".srm": obj_type = "Menu (메뉴)"
        
        # Read first 15 lines to extract title or comment
        title_comment = ""
        try:
            with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
                lines = [f.readline() for _ in range(15)]
                for l in lines:
                    if "string title =" in l.lower():
                        title_comment = l.split("=")[1].strip().strip('"')
                        break
                    elif "event" in l.lower() or "global function" in l.lower():
                        title_comment = l.strip()
                        break
        except Exception:
            pass

        r_vals = [obj_idx, mod_name, name_no_ext, obj_type, ext, title_comment, fsize]
        for c_idx, val in enumerate(r_vals, start=1):
            cell = ws2.cell(row=row_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx in (1, 4, 5): cell.alignment = align_center
            elif c_idx in (2, 3, 6): cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = "#,##0"
        row_idx += 1
        obj_idx += 1

# ---------------------------------------------------------
# SHEET 3: DB 테이블 & 쿼리 매핑 (DB & SQL Spec)
# ---------------------------------------------------------
ws3 = wb.create_sheet(title="3. DB 테이블 & 쿼리 매핑")
ws3.views.sheetView[0].showGridLines = True

ws3["A1"] = "🗄️ 데이터베이스 주요 테이블 및 DataWindow SQL 매핑"
ws3["A1"].font = font_title

db_tables_meta = [
    ("PR_M_ITEM", "품목 마스터", "생산/BOM", "전체 품목 코드, 품목명, 규격, 두께, 중량 및 외주/공정 기본 정보"),
    ("PR_M_ITEM_BOM", "BOM 마스터", "생산/BOM", "모품목-자품목 소요량(USE_QTY), 유효기간, 전개제외(EXCEPT_FLAG), 가상도번 관리"),
    ("PR_M_ITEM_COST", "품목 단가 이력", "구매/영업", "품목/거래처별 유효일자(cost_apply_ymd) 적용 단가(1:구매, S:내수판매, E:수출)"),
    ("pu_t_stock_maint", "자재 수불 마스터", "자재/수불", "자재 입고/출고/이동/반품/불량 수불 통합 관리 (maint_tag 구별)"),
    ("sa_t_sale_dtl", "출하 실적 테이블", "영업/출하", "우리 회사 제품 출하 일자(sale_ymd), 출하수량, 예상 매출액, 제번(work_order)"),
    ("sa_t_lg_receiving_dtl", "고객사 입고 확정", "영업/매출", "고객사 최종 입고 확정일자(receiving_ymd), 확정 수량, 확정 매출액, 제번(work_order)"),
    ("CM_M_USERS_INFO", "사용자 마스터", "공통/보안", "사용자 ID, 비밀번호, 부서코드, 사번 및 시스템 로그인 사용 여부"),
    ("PR_M_WORK", "공정 마스터", "생산/공정", "공정 코드, 공정 명칭, 라인 작업 환경 정의"),
    ("CM_M_CUST", "거래처 마스터", "공통/마스터", "매입/매출 거래처 코드, 거래처명, 사업자번호, 대표자명")
]

ws3.cell(row=3, column=1, value="[주요 데이터베이스 테이블 명세]").font = font_bold
headers_s3_1 = ["테이블 명", "한글 테이블 명", "관련 업무 모듈", "주요 역활 및 칼럼 설명"]
for col_idx, h in enumerate(headers_s3_1, start=1):
    cell = ws3.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = header_border

row_idx = 5
for tname, tkname, mod, desc in db_tables_meta:
    r_vals = [tname, tkname, mod, desc]
    for c_idx, val in enumerate(r_vals, start=1):
        cell = ws3.cell(row=row_idx, column=c_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border
        if c_idx in (1, 3): cell.alignment = align_center
        else: cell.alignment = align_left
    row_idx += 1

row_idx += 2
ws3.cell(row=row_idx, column=1, value="[핵심 DataWindow SQL 쿼리 상세 스크립트]").font = font_bold
row_idx += 1

headers_s3_2 = ["쿼리 구분", "DataWindow 명", "관련 파일", "SQL 쿼리 스크립트 내용 및 로직 설명"]
for col_idx, h in enumerate(headers_s3_2, start=1):
    cell = ws3.cell(row=row_idx, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = header_border

sql_samples = [
    ("BOM 8단계 전개", "dw_t_bom_tree", "PR_M_ITEM_BOM", 
     "WITH t_bom (c_item_level, item_code, P_ITEM_CODE, C_ITEM_CODE, USE_QTY, EXCEPT_FLAG) AS (\n"
     "  SELECT 0, M.ITEM_CODE, M.ITEM_CODE, M.ITEM_CODE, 0.0, '0' FROM PR_M_ITEM M WHERE M.ITEM_CODE = :as_item\n"
     "  UNION ALL\n"
     "  SELECT T.c_item_level + 1, T.ITEM_CODE, A.ITEM_CODE, M.ITEM_CODE, A.USE_QTY, ISNULL(A.EXCEPT_FLAG, '0')\n"
     "  FROM t_bom T JOIN PR_M_ITEM_BOM A ON T.C_ITEM_CODE = A.ITEM_CODE JOIN PR_M_ITEM M ON A.MAT_CODE = M.ITEM_CODE\n"
     "  WHERE :as_date BETWEEN A.from_apply_ymd AND A.to_apply_ymd AND T.c_item_level < 8\n"
     ") SELECT * FROM t_bom WHERE EXCEPT_FLAG = '0';"),

    ("자재 수불 매스터", "dw_pu_stock_060_wh_l1_new", "pu_t_stock_maint",
     "SELECT A.maint_ymd, UPPER(A.mat_code) AS item_code, A.maint_qty AS IN_QTY, A.maint_cost, A.maint_amt,\n"
     "       CASE A.maint_tag WHEN '3' THEN '기타' WHEN '9' THEN '창고입고' WHEN 'C' THEN '이동' WHEN 'G' THEN '구매' WHEN 'S' THEN '외주입고' END AS div\n"
     "  FROM pu_t_stock_maint A WHERE A.maint_ymd BETWEEN :fr_ymd AND :to_ymd AND A.mat_code = :as_mat;"),

    ("매출 이중 매칭", "dw_sa_stock_sale_match", "sa_t_sale_dtl / sa_t_lg_receiving_dtl",
     "SELECT S.sale_ymd, S.work_order, S.sale_qty, S.sale_amt AS expected_amt,\n"
     "       R.receiving_ymd, ISNULL(R.recv_qty,0) AS recv_qty, ISNULL(R.recv_amt,0) AS confirmed_amt,\n"
     "       (S.sale_amt - ISNULL(R.recv_amt,0)) AS diff_amt\n"
     "  FROM sa_t_sale_dtl S LEFT JOIN sa_t_lg_receiving_dtl R ON S.work_order = R.work_order;")
]

row_idx += 1
for qcat, dwname, relfile, sqltext in sql_samples:
    r_vals = [qcat, dwname, relfile, sqltext]
    for c_idx, val in enumerate(r_vals, start=1):
        cell = ws3.cell(row=row_idx, column=c_idx, value=val)
        cell.border = thin_border
        if c_idx == 4:
            cell.font = font_code
            cell.alignment = align_wrap_left
        else:
            cell.font = font_regular
            cell.alignment = align_center
    row_idx += 1

# ---------------------------------------------------------
# SHEET 4: 핵심 비즈니스 로직 분석 (Business Logic Spec)
# ---------------------------------------------------------
ws4 = wb.create_sheet(title="4. 핵심 비즈니스 로직 분석")
ws4.views.sheetView[0].showGridLines = True

ws4["A1"] = "⚙️ 모듈별 핵심 비즈니스 연산 및 이벤트 스크립트"
ws4["A1"].font = font_title

headers_s4 = ["도메인 영역", "핵심 프로세스 명", "관련 소스 파일", "이벤트 / 함수 명", "상세 로직 및 연산 알고리즘 내용"]
for col_idx, h in enumerate(headers_s4, start=1):
    cell = ws4.cell(row=3, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = header_border

logic_samples = [
    ("시스템/인증", "시스템 시작 및 버전검증", "w_login.srw", "ue_env()",
     "1. f_get_version() 호출하여 현재 버전을 구함\n2. f_get_today()로 서버 당일 일자 동기화\n3. f_check_version()으로 업데이트 필요 여부 검증\n4. dw_cm_column_text DataStore 생성 후 다국어/컬럼 라벨 캐싱"),

    ("생산/BOM", "8단계 재귀 BOM MRP 전개", "pr_plan_01.pbl", "ue_bom_expand()",
     "1. 모품목 입력 시 WITH t_bom CTE 쿼리로 하위 8단계 부품 전개\n2. EXCEPT_FLAG == '1'인 품목은 생산계획 편성에서 제외\n3. vir_item_flag == '1'인 가상도번은 상위 모품목으로 재할당\n4. 최종 소요량 = Parent.USE_QTY * Child.USE_QTY 산출"),

    ("자재/수불", "자재 수불 구분 및 창고이동", "pu_input_01.pbl", "ue_save()",
     "1. 수불 구분 태그(maint_tag)에 따라 'G'(구매입고), 'S'(외주입고), 'C'(이동), 'R'(반품) 분류\n2. pu_t_stock_maint 수불 마스터에 수량/단가/금액 INSERT/UPDATE\n3. 창고코드(wh_cust_code)별 실재고 동기화"),

    ("영업/출하", "출하 대비 입고확정 매출 매칭", "sa_stock_01.pbl", "ue_match_sales()",
     "1. sa_t_sale_dtl (우리 회사 출하)과 sa_t_lg_receiving_dtl (고객사 확정) 조인\n2. work_order (제번) 키를 기준으로 예상 매출액과 확정 매출액의 차이(diff_amt) 계산\n3. 미입고 출하건 경고 표시 및 매출 대시보드 데이터 파이프라인 생성"),

    ("공통 프레임워크", "DataWindow 표준 스타일 적용", "cm_func.pbl", "f_set_dw_style()",
     "1. 전달받은 DataWindow의 모든 Column/Header 컨트롤 탐색\n2. 시스템 표준 폰트(맑은 고딕), 배경색, 그리드 테두리 스타일 동적 수정\n3. Read-Only 및 필수 입력 필드 배경색 구분 적용")
]

row_idx = 4
for dom, proc, src, ev, desc in logic_samples:
    r_vals = [dom, proc, src, ev, desc]
    for c_idx, val in enumerate(r_vals, start=1):
        cell = ws4.cell(row=row_idx, column=c_idx, value=val)
        cell.border = thin_border
        if c_idx in (1, 3, 4): cell.alignment = align_center; cell.font = font_regular
        elif c_idx == 2: cell.alignment = align_left; cell.font = font_bold
        else: cell.alignment = align_wrap_left; cell.font = font_regular
    row_idx += 1

# ---------------------------------------------------------
# SHEET 5: TO-BE 차세대 전환 대응표 (Modernization Map)
# ---------------------------------------------------------
ws5 = wb.create_sheet(title="5. TO-BE 차세대 전환 대응표")
ws5.views.sheetView[0].showGridLines = True

ws5["A1"] = "🚀 PowerBuilder ➡️ 차세대 Web System (Spring Boot + React) 마이그레이션 대응표"
ws5["A1"].font = font_title

headers_s5 = ["분류", "PowerBuilder 10.5 레거시 요소", "차세대 Modern Stack 대응 구조", "마이그레이션 및 리팩토링 상세 가이드"]
for col_idx, h in enumerate(headers_s5, start=1):
    cell = ws5.cell(row=3, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = header_border

to_be_map = [
    ("UI Layer", "Window Sheet (.srw)", "React / Next.js Page Component", "JSX 기반 반응형 UI 구현. 무설치 웹 브라우저 접속 및 현장 태블릿/모바일 지원"),
    ("UI Layer", "UserObject (.sru)", "React Custom Reusable Components", "<SearchCondition />, <CustomGrid />, <Modal /> 등 재사용 컴포넌트화"),
    ("Data Layer", "DataWindow (.srd)", "REST API + MyBatis / JPA Repository", "SQL문과 UI 표현의 완전 분리. JSON 포맷 기반 통신 및 DTO 객체 매핑"),
    ("Logic Layer", "Global Function (.srf)", "Spring Boot Service Layer Class", "Business Logic을 Java Service Layer에 100% 모듈화하여 객체지향 설계"),
    ("BOM 연산", "PB Script + CTE SQL", "DB CTE View + Redis In-Memory Cache", "BOM 재귀 전개 결과를 Redis 캐시에 저장하여 조회 성능 10배 향상 (Latency < 50ms)"),
    ("자재 수불", "pu_t_stock_maint Sync SQL", "Event-Driven CQRS Architecture", "수불 이벤트 발생 시 비동기 메세지 큐(Kafka/RabbitMQ)를 통해 실시간 재고 업데이트"),
    ("외부 연동", "Popbill SDK / dw2xls.pbl", "Spring Boot REST Open API / Apache POI", "서버단 엑셀 대용량 다운로드 API 및 전자세금계산서 REST Webhook 연동")
]

row_idx = 4
for cat, asis, tobe, guide in to_be_map:
    r_vals = [cat, asis, tobe, guide]
    for c_idx, val in enumerate(r_vals, start=1):
        cell = ws5.cell(row=row_idx, column=c_idx, value=val)
        cell.border = thin_border
        if c_idx == 1: cell.alignment = align_center; cell.font = font_bold
        elif c_idx in (2, 3): cell.alignment = align_left; cell.font = font_regular
        else: cell.alignment = align_wrap_left; cell.font = font_regular
    row_idx += 1

# Auto-adjust column widths for all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                lines = val_str.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val_str))
        # Account for Korean characters wide width
        adjusted_width = min(max(max_len * 1.4, 12), 70)
        sheet.column_dimensions[col_letter].width = adjusted_width

wb.save(output_excel)
print("Excel Workbook saved successfully at:", output_excel)
